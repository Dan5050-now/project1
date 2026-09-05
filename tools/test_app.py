"""Drive app/PRAP.html in a real browser and check it against the reference.

A screenshot proves the page rendered. These checks prove it is RIGHT: the
calculation is compared cell-by-cell with tools/verify_source_workbook.py, and the
export is re-read by openpyxl - an independent .xlsx implementation - and put back
through the same verifier.

    python tools/test_app.py
"""

import json
import pathlib
import subprocess
import math
import sys
from collections import defaultdict
from datetime import date

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

ROOT = pathlib.Path(__file__).resolve().parents[1]
APP = (ROOT / "app" / "PRAP.html").as_uri()
# Both dummy sizes are checked. The small one is not a lighter version of the same
# test - it has a different period mix, a different overlap profile and a single-person
# role pool, so it exercises paths the large set happens not to reach.
FIXTURES = [ROOT / "templates" / "PRAP_SourceData_Dummy_v1.16.xlsx",
            ROOT / "templates" / "PRAP_SourceData_Dummy_10x10_v1.8.xlsx"]
CHROME = "/opt/pw-browsers/chromium-1194/chrome-linux/chrome"

sys.path.insert(0, str(ROOT / "tools"))
_ref = (ROOT / "tools" / "verify_source_workbook.py").read_text().split("def main")[0]
exec(_ref)                                    # rows, d, months_between, coverage, CLINICAL_TYPES


# REQ-CAL-20: every FTE figure is a whole number of hundredths, and a project-month's
# hundredths are handed out by largest remainder so the parts sum to the month exactly.
# Written out here rather than imported, because the point of this file is to be an
# INDEPENDENT check - importing the thing under test would only prove it agrees with
# itself. The tie-break is part of the rule: two programs breaking ties differently
# would differ by 0.01 and both believe themselves right.
CENTS = 100


def to_cents(v):
    """Half away from zero, matching JavaScript's Math.round for positive values.
    Python's round() is half-to-even and would disagree on exact halves."""
    v = (v or 0.0) * CENTS
    return int(math.floor(v + 0.5)) if v >= 0 else -int(math.floor(-v + 0.5))


def largest_remainder(items, total_cents):
    """items: [(weight, tie_break_key)] -> cents each, summing to total_cents."""
    n = len(items)
    if not n:
        return []
    if total_cents <= 0:
        return [0] * n
    tot = sum(w for w, _ in items if w > 0)
    if not tot > 0:
        return [0] * n
    exact = [(w if w > 0 else 0.0) / tot * total_cents for w, _ in items]
    out = [int(math.floor(v)) for v in exact]
    left = total_cents - sum(out)
    order = sorted(range(n), key=lambda i: (-(exact[i] - out[i]), items[i][1]))
    i = 0
    while left > 0:
        out[order[i % n]] += 1
        i += 1
        left -= 1
    return out


def reference_person_months(path):
    wb = load_workbook(path)
    P = {r["project_id"]: r for r in rows(wb["Project"])}
    PER = defaultdict(list)
    for r in rows(wb["ProjectPeriod"]):
        PER[r["project_id"]].append(r)
    # Schema 6 keys the table on the work scope too, with an EMPTY scope meaning every
    # scope - see lookup() in verify_source_workbook.py, which this reuses.
    RF = {(r["project_type"], r["clinical_phase"], scope_of(r), r["period_name"],
           r["role_name"]): r["role_factor"] for r in rows(wb["RoleFactor"])}
    # REQ-CAL-19: the month's demand in FTE, keyed exactly as the factors are.
    PWS = {(r["project_type"], r["clinical_phase"], scope_of(r), r["period_name"]):
           r["standard_fte"] for r in rows(wb["PeriodFTEStandard"])}
    # REQ-CAL-16, worked out here independently: which absent roles land on this one.
    ABSORB = defaultdict(list)
    for r in rows(wb["RoleFactor"]):
        if r.get("absorbed_by"):
            ABSORB[(r["project_type"], r["clinical_phase"], scope_of(r),
                    r["period_name"], r["absorbed_by"])].append(r["role_name"])
    PPW = defaultdict(list)
    for r in rows(wb["PersonPeriodWeight"]):
        PPW[r["assignment_id"]].append(r)

    def seg(pid, y, m):
        for s in PER.get(pid, []):
            if d(s["period_start"]) <= date(y, m, 1) <= d(s["period_end"]):
                return s
        return None

    def weight(a, y, m):
        for w in PPW.get(a["assignment_id"], []):
            if d(w["period_start"]) <= date(y, m, 1) <= d(w["period_end"]):
                return w["weight_override"]
        return a["person_weight"]

    # Who holds which role on which project, month by month. The role factor is what
    # the ROLE costs the project, not what each person holding it costs, so it is
    # divided between them - worked out here independently of the application, which
    # is the whole point of a reference.
    ASG = list(rows(wb["Assignment"]))
    sharers = defaultdict(set)
    for a in ASG:
        pr = P[a["project_id"]]
        s, e = assignment_window(PER, pr, a)
        for y, m in months_between(s, e):
            if coverage(y, m, s, e) > 0:
                sharers[(a["project_id"], a["role_name"], y, m)].add(a["person_id"])

    def eff(pr, ph, pn, role, y, m):
        """One role's effective factor, absorption included - worked out here rather
        than borrowed, which is the whole point of a reference implementation."""
        rf = lookup(RF, (pr["project_type"], ph), scope_of(pr), (pn, role))
        rf = 1.0 if rf is None else rf
        for absent in (lookup(ABSORB, (pr["project_type"], ph), scope_of(pr),
                              (pn, role)) or []):
            if sharers[(pr["project_id"], absent, y, m)]:
                continue
            extra = lookup(RF, (pr["project_type"], ph), scope_of(pr), (pn, absent))
            rf += 0.0 if extra is None else extra
        return rf

    lines = []
    for a in ASG:
        pr = P[a["project_id"]]
        s, e = assignment_window(PER, pr, a)
        ph = pr["clinical_phase"] if pr["project_type"] in CLINICAL_TYPES else None
        for y, m in months_between(s, e):
            cov = coverage(y, m, s, e)
            if cov <= 0:
                continue
            sg = seg(a["project_id"], y, m)
            pn = sg["period_name"] if sg else None
            rf = eff(pr, ph, pn, a["role_name"], y, m)
            share = len(sharers[(a["project_id"], a["role_name"], y, m)]) or 1
            # REQ-CAL-19. A project-month IS its standard x its own period weight,
            # scaled by the month it actually runs; the people on it DIVIDE that in
            # proportion to their claims. Worked out below, once the month is whole.
            std = lookup(PWS, (pr["project_type"], pr["clinical_phase"]),
                         scope_of(pr), (pn,)) if pn is not None else None
            std = 1.0 if std is None else float(std)
            lines.append({"aid": a["assignment_id"], "pid": a["project_id"],
                          "sid": a["person_id"], "y": y, "m": m, "fte": 0.0,
                          "claim": (rf / share) * weight(a, y, m) * cov, "cov": cov,
                          "std": std, "pw": (sg["weight"] if sg else 1.0)})

    # REQ-CAL-19: the demand, divided between the claims on it.
    grp0 = defaultdict(list)
    for L in lines:
        grp0[(L["pid"], L["y"], L["m"])].append(L)
    for group in grp0.values():
        claims = sum(L["claim"] for L in group)
        ran = max(L["cov"] for L in group)
        demand_cents = to_cents(group[0]["std"] * group[0]["pw"] * ran)
        cents = largest_remainder(
            [((L["claim"] if claims > 0 else 0.0), L.get("aid") or "") for L in group],
            demand_cents)
        for L, c in zip(group, cents):
            L["fte"] = c / CENTS

    # REQ-CAL-18. A manual assignment takes the figure it was given - and 0.00 where it
    # was given none, which is the only reading of "the user owns every month" that does
    # not quietly fall back to the arithmetic they rejected. A manual PROJECT keeps its
    # people in proportion: scale the lines of that project-month to the stated total.
    EST = {}
    if "MonthlyEstimate" in wb.sheetnames:
        for r in rows(wb["MonthlyEstimate"]):
            if r.get("scope") and r.get("ref_id") and r.get("month"):
                EST[(str(r["scope"]), str(r["ref_id"]), str(r["month"]))] = r.get("fte")
    man_p = {pid for pid, pr in P.items()
             if str(pr.get("estimation_type") or "").strip().lower() == "manual"}
    man_a = {a["assignment_id"] for a in ASG
             if str(a.get("estimation_type") or "").strip().lower() == "manual"}
    for L in lines:
        if L["aid"] in man_a:
            v = EST.get(("assignment", L["aid"], f"{L['y']}-{L['m']:02d}"))
            L["fte"] = 0.0 if v is None else to_cents(float(v)) / CENTS
    grp = defaultdict(list)
    for L in lines:
        if L["pid"] in man_p:
            grp[(L["pid"], L["y"], L["m"])].append(L)
    for (pid, y, m), g in grp.items():
        want = EST.get(("project", pid, f"{y}-{m:02d}"))
        if want is None:
            for L in g:
                L["fte"] = 0.0
            continue
        have = sum(L["fte"] for L in g)
        if abs(have) < 1e-9:            # nobody to carry it - left alone, and reported
            continue
        # The stated month shared to the hundredth, the same rule as above (REQ-CAL-20).
        cents = largest_remainder(
            [(L["fte"], L.get("aid") or "") for L in g], to_cents(float(want)))
        for L, c in zip(g, cents):
            L["fte"] = c / CENTS

    out = defaultdict(float)
    for L in lines:
        out[(L["sid"], L["y"] * 12 + L["m"] - 1)] += L["fte"]
    return out


def main(DUMMY):
    failures, notes = [], []
    with sync_playwright() as pw:
        b = pw.chromium.launch(executable_path=CHROME, downloads_path="/tmp")
        ctx = b.new_context(accept_downloads=True)
        pg = ctx.new_page()
        errors = []
        pg.on("pageerror", lambda e: errors.append(str(e)))
        pg.on("dialog", lambda dlg: dlg.accept())
        pg.goto(APP)
        pg.wait_for_timeout(200)
        pg.set_input_files("#picker", str(DUMMY))
        pg.wait_for_timeout(4000)

        # Informational findings are explanations, not problems: V-14 and V-21 both fire
        # on these fixtures to say WHY an inspection after the DB lock is legitimate.
        # Anything above information on a file the verifier calls clean is a failure.
        findings = pg.evaluate("S.model.findings.filter(f => f.sev !== 'information')"
                               ".map(f => f.sev + ' ' + f.rule + ': ' + f.msg)")
        info = pg.evaluate("S.model.findings.filter(f => f.sev === 'information').length")
        if findings:
            failures.append(f"app reports {len(findings)} findings above information on a file the "
                            f"verifier calls clean; first: {findings[0][:110]}")
        else:
            notes.append(f"no findings above information ({info} informational)")

        app_pm = pg.evaluate("() => { const o = {}; "
                             "for (const [k, v] of S.calc.persMonth) o[k] = v; return o; }")
        ref = reference_person_months(DUMMY)
        got = {(k.split("|")[0], int(k.split("|")[1])): v for k, v in app_pm.items()}
        if set(got) != set(ref):
            failures.append(f"person-month coverage differs: app {len(got)}, reference {len(ref)}")
        bad = [k for k in set(got) & set(ref) if abs(got[k] - ref[k]) > 1e-6]
        if bad:
            k = bad[0]
            failures.append(f"{len(bad)} calculated values differ; e.g. {k}: "
                            f"app {got[k]:.6f} vs reference {ref[k]:.6f}")
        else:
            notes.append(f"calculation matches the reference on all {len(ref)} person-months")

        with pg.expect_download() as dl:
            pg.click("#exportBtn")            # opens the export menu
            pg.click("#exportBtn2")           # …source data .xlsx
        exported = "/tmp/prap_export_test.xlsx"
        dl.value.save_as(exported)

        pg2 = ctx.new_page()
        pg2.goto(APP)
        pg2.wait_for_timeout(200)
        pg2.set_input_files("#picker", exported)
        pg2.wait_for_timeout(4000)
        n = pg2.evaluate("S.model.findings.filter(f => f.sev !== 'information').length")
        if n:
            failures.append(f"the exported file re-imports with {n} findings above information")
        else:
            notes.append("export re-imports with no findings above information")

        if errors:
            failures.append(f"page errors: {errors[:2]}")
        b.close()

    # The export must also satisfy a reader that is not ours.
    r = subprocess.run([sys.executable, str(ROOT / "tools" / "verify_source_workbook.py"), exported],
                       capture_output=True, text=True)
    if "ERRORS: none" not in r.stdout:
        failures.append("the exported file does not pass verify_source_workbook.py")
    else:
        notes.append("export passes the Python verifier via openpyxl")

    print(f"app/PRAP.html  x  {DUMMY.name}")
    for n_ in notes:
        print("  ok  ", n_)
    for f in failures:
        print("  FAIL", f)
    return failures


if __name__ == "__main__":
    only = [pathlib.Path(a) for a in sys.argv[1:]]
    all_failures = []
    for fixture in (only or FIXTURES):
        all_failures += main(fixture)
        print()
    print("FAILURES: none" if not all_failures else f"FAILURES ({len(all_failures)})")
    sys.exit(1 if all_failures else 0)
