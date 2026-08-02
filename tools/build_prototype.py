"""Generate the Step 3 UI prototype - a static, self-contained HTML mock-up.

DESIGN ONLY. There is no parsing, no calculation and no import/export behind this
page: the figures are a snapshot computed here, in Python, and baked into the markup.
The only JavaScript is tab switching, row expansion and the hover tooltip. The point is
to review layout, components and
information hierarchy before any application code is written (plan sheet 08, task 3.1).

    python tools/build_prototype.py

Output: app/PRAP_Prototype_v0.7.html
"""

import calendar
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DUMMY = ROOT / "templates" / "PRAP_SourceData_Dummy_v1.8.xlsx"
OUT = ROOT / "app" / "PRAP_Prototype_v0.7.html"

APP_VERSION = "prototype v0.7"
SCHEMA_EXPECTED = 5
WIN_FROM, WIN_TO = (2027, 1), (2027, 12)      # the 12 months the mock-up shows

# ---- design tokens (validated: see dataviz palette reference) --------------
CAT_LIGHT = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100", "#e87ba4", "#008300", "#4a3aa7"]
CAT_DARK = ["#3987e5", "#d95926", "#199e70", "#c98500", "#d55181", "#008300", "#9085e9"]
SEQ = ["#cde2fb", "#b7d3f6", "#9ec5f4", "#86b6ef", "#6da7ec", "#5598e7",
       "#3987e5", "#2a78d6", "#256abf", "#1c5cab", "#184f95"]


def rows(ws):
    hdr = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not all(v is None for v in r):
            yield dict(zip(hdr, r))


def d(v):
    return v.date() if hasattr(v, "date") else v


def months(a, b):
    y, m = a
    while (y, m) <= b:
        yield y, m
        m += 1
        if m == 13:
            y, m = y + 1, 1


def coverage(y, m, s, e):
    dm = calendar.monthrange(y, m)[1]
    lo, hi = max(date(y, m, 1), s), min(date(y, m, dm), e)
    return 0.0 if hi < lo else ((hi - lo).days + 1) / dm


def esc(t):
    return (str(t).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


# ---------------------------------------------------------------- snapshot
def snapshot():
    wb = load_workbook(DUMMY, data_only=False)
    P = {r["project_id"]: r for r in rows(wb["Project"])}
    PER = defaultdict(list)
    for r in rows(wb["ProjectPeriod"]):
        PER[r["project_id"]].append(r)
    MS = defaultdict(list)
    for r in rows(wb["Milestone"]):
        MS[r["project_id"]].append(r)
    RF_ROWS = list(rows(wb["RoleFactor"]))
    # R-10: keyed on project type + clinical phase + period + role.
    RF = {(r["project_type"], r["clinical_phase"], r["period_name"], r["role_name"]):
          r["role_factor"] for r in RF_ROWS}
    PWS_ROWS = list(rows(wb["PeriodWeightStandard"]))
    LIST_ROWS = list(rows(wb["Lists"]))
    CT = {"NewDrug CT", "Biosimilar CT"}
    TYPE_RANK = {"NewDrug CT": 0, "Biosimilar CT": 1, "Others": 2}
    PSN = {r["person_id"]: r for r in rows(wb["Person"])}
    ASG = list(rows(wb["Assignment"]))
    PPW = defaultdict(list)
    for r in rows(wb["PersonPeriodWeight"]):
        PPW[r["assignment_id"]].append(r)
    CFG = {r["parameter"]: r["value"] for r in rows(wb["Config"])}

    grid = list(months(WIN_FROM, WIN_TO))
    gridset = set(grid)

    def pweight(pid, y, m):
        for s in PER.get(pid, []):
            if d(s["period_start"]) <= date(y, m, 1) <= d(s["period_end"]):
                return s["weight"]
        return 1.00

    def period_of(pid, y, m):
        for s in PER.get(pid, []):
            if d(s["period_start"]) <= date(y, m, 1) <= d(s["period_end"]):
                return s["period_name"]
        return None

    def rfactor(pr, a, y, m):
        ph = pr["clinical_phase"] if pr["project_type"] in CT else None
        return RF.get((pr["project_type"], ph, period_of(a["project_id"], y, m),
                       a["role_name"]), 1.00)

    def wweight(a, y, m):
        for w in PPW.get(a["assignment_id"], []):
            if d(w["period_start"]) <= date(y, m, 1) <= d(w["period_end"]):
                return w["weight_override"]
        return a["person_weight"]

    proj_m = defaultdict(float)
    pers_m = defaultdict(float)
    cell_m = defaultdict(float)        # (project, person, role, y, m) -> FTE
    # The window aggregates above cover the 12 months on screen. The Gantt spans whole
    # projects, most of which start before the window and end after it, so its tooltip
    # needs a load figure that exists outside the window too (O-10).
    proj_all = defaultdict(float)      # (project, y, m) -> FTE, full project span
    who = defaultdict(set)             # (project, y, m) -> {(person_id, role)}
    for a in ASG:
        pr = P.get(a["project_id"])
        if not pr or a["person_id"] not in PSN:
            continue
        s, e = d(a["assign_start_date"]), d(a["assign_end_date"]) or d(pr["end_date"])
        for (y, m) in months((s.year, s.month), (e.year, e.month)):
            cov = coverage(y, m, s, e)
            if cov <= 0:
                continue
            v = (pweight(a["project_id"], y, m) * rfactor(pr, a, y, m)
                 * wweight(a, y, m) * cov)
            proj_all[(a["project_id"], y, m)] += v
            if (y, m) in gridset:
                proj_m[(a["project_id"], y, m)] += v
                pers_m[(a["person_id"], y, m)] += v
                cell_m[(a["project_id"], a["person_id"], a["role_name"], y, m)] += v
                who[(a["project_id"], y, m)].add((a["person_id"], a["role_name"]))

    tot = defaultdict(float)
    for (pid, y, m), v in proj_m.items():
        tot[pid] += v
    top = [pid for pid, _ in sorted(tot.items(), key=lambda kv: -kv[1])[:10]]

    return dict(P=P, PER=PER, MS=MS, PSN=PSN, ASG=ASG, PPW=PPW, CFG=CFG, CT=CT,
                TYPE_RANK=TYPE_RANK, grid=grid, proj_m=proj_m, pers_m=pers_m,
                cell_m=cell_m, top=top, tot=tot, proj_all=proj_all, who=who,
                RF_ROWS=RF_ROWS, PWS_ROWS=PWS_ROWS, LIST_ROWS=LIST_ROWS)


S = snapshot()
GRID = S["grid"]
MLAB = [f"{calendar.month_abbr[m]}<br><span class='yr'>{y}</span>" for y, m in GRID]
OVER = float(S["CFG"]["over_allocation_fte"])
UNDER = float(S["CFG"]["under_allocation_fte"])
HOURS = float(S["CFG"]["fte_hours_per_month"])



CT = None  # set after snapshot


def is_ct(pid):
    return S["P"][pid]["project_type"] in S["CT"]


# REQ-DSH-10: where a period name occurs more than once in a project, each occurrence
# must be distinguishable on screen. Since R-11 named the two conduct stretches apart,
# no name repeats and this never fires - the requirement is now satisfied by the data
# model rather than by a display rule. Kept as the guard for that: V-18 rejects a
# repeated name on import, but a hand-built model could still reach the renderer.
_PLABEL = {}
for _pid, _segs in S["PER"].items():
    _segs = sorted(_segs, key=lambda r: r["period_seq"])
    _n = defaultdict(int)
    for _s in _segs:
        _n[_s["period_name"]] += 1
    _seen = defaultdict(int)
    for _s in _segs:
        _nm = _s["period_name"]
        _seen[_nm] += 1
        _PLABEL[(_pid, _s["period_seq"])] = (f"{_nm} ({_seen[_nm]})" if _n[_nm] > 1 else _nm)


def period_label(pid, seg):
    return _PLABEL[(pid, seg["period_seq"])]


def prank(pid):
    """Requested order: NewDrug CT, then Biosimilar CT, then Others; earlier first."""
    pr = S["P"][pid]
    return (S["TYPE_RANK"].get(pr["project_type"], 9), d(pr["start_date"]), pid)


def phase_pill(pid):
    pr = S["P"][pid]
    if not is_ct(pid):
        return ''
    n = (pr["clinical_phase"] or "").replace("Phase ", "")
    return f'<span class="ph ph{n}">{esc(pr["clinical_phase"])}</span>'


def type_pill(pid):
    t = S["P"][pid]["project_type"]
    k = {"NewDrug CT": "nd", "Biosimilar CT": "bs"}.get(t, "ot")
    return f'<span class="ty {k}">{esc(t)}</span>'


def _mix(hex_, f):
    """Lighten (f>1) or darken (f<1) a hex colour, staying on its own hue."""
    r, g, b = (int(hex_[i:i + 2], 16) for i in (1, 3, 5))
    if f >= 1:
        t = min(1.0, f - 1.0)
        r, g, b = (int(c + (255 - c) * t) for c in (r, g, b))
    else:
        r, g, b = (int(c * f) for c in (r, g, b))
    return "#%02x%02x%02x" % (max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b)))


# The validated set caps at 8 hues. Per-project colour was asked for explicitly, so
# the palette is EXTENDED systematically: the seven validated hues, each stepped in
# lightness. Beyond the first seven, hue alone no longer separates reliably - the
# tooltip, the legend order and the table below carry identity.
BASE7 = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100", "#e87ba4", "#008300", "#4a3aa7"]
STEPS = [1.0, 0.70, 1.30, 0.85, 1.15, 0.55, 1.45, 0.62]


def proj_colour(i):
    return _mix(BASE7[i % 7], STEPS[(i // 7) % len(STEPS)])


# P-01 / S-01: every row in every editable section carries an insert control, and a
# new row lands directly below the row whose button was pressed - so a row is added
# where the user is looking, not appended to the bottom of a 62-row table.
#
# It leads the row rather than trailing it. The project table is 23 columns wide and
# scrolls; a trailing control would sit off-screen and need a horizontal scroll every
# time. Leading, it reads as a row-action gutter and is always in view.
INS_TD = ('<td class="ins"><button class="btn tiny" '
          'title="Insert a new row directly below this one">+ row</button></td>')


def att(t):
    """Escape for an HTML attribute - esc() plus the quote that would close it."""
    return esc(t).replace('"', "&quot;")


# O-10: the Gantt now colours bands by PERIOD NAME rather than by weight, which
# overturns D-06. Your example mapping is followed: grey before start-up, red at
# start-up, green through conduct, orange at close-out, dark grey after.
#
# One deliberate departure. Red beside green is the pair that red-green colour
# blindness collapses, and D-04 - which you accepted - makes "never colour alone"
# an accessibility floor for this UI. So the red is shifted a little toward orange,
# every band wide enough carries its name as text, and the tooltip names the period
# outright. Hue speeds recognition here; it is never the only carrier.
PERIOD_HUE = {
    "Before-Start-up":        "#adaca6",   # grey, light
    "Start-up":               "#d9472f",   # red, shifted toward orange
    "Conduct (interim)":      "#3fc795",   # green, light - the first stretch
    "Conduct (final)":        "#159068",   # green, deep  - the stretch that ends the trial
    "Close-out (interim)":    "#f2b53d",   # orange, light - the interim stop
    "Close-out (final)":      "#d97e0a",   # orange, deep - the final one
    "After Close-out (final)": "#6f6f68",  # grey, dark - still legible on a dark surface
    "Planning":               "#adaca6",   # 'Others' set, same semantics
    "Develop":                "#1baf7a",
    "Close":                  "#eda100",
}


def band_fill(name, w, wmax):
    """Hue carries the period; lightness within that hue carries the weight.

    The lightness range is deliberately narrow. Widen it and the weight steps start
    competing with the hue for the reader's attention, which is exactly the confusion
    that moving colour onto the period name was meant to remove.
    """
    base = PERIOD_HUE.get(name, "#adaca6")
    f = 1.14 - 0.24 * (min(w, wmax) / wmax if wmax else 0)
    return _mix(base, f)


def seq_step(v, vmax):
    if v <= 0 or vmax <= 0:
        return None
    i = min(len(SEQ) - 1, int(round((v / vmax) * (len(SEQ) - 1))))
    return i


# ---------------------------------------------------------------- charts
def chart_stacked():
    """Monthly demand, one band per project.

    Bands are ordered by total resource with the LARGEST AT THE BOTTOM, so the
    heaviest projects sit on the baseline where they are easiest to read. Every
    'Others' project is grey; clinical trials take the extended colour set.
    """
    order = sorted(S["P"], key=lambda pid: (-S["tot"].get(pid, 0.0), pid))
    active = [pid for pid in order if S["tot"].get(pid, 0.0) > 0.004]
    colour, ci = {}, 0
    for pid in active:
        if S["P"][pid]["project_type"] == "Others":
            colour[pid] = "var(--other)"
        else:
            colour[pid] = proj_colour(ci)
            ci += 1

    W, H = 1180, 300
    pad_l, pad_b, pad_t = 56, 46, 26
    bw = (W - pad_l - 14) / len(GRID)
    vmax = max(sum(S["proj_m"].get((pid, y, m), 0.0) for pid in active)
               for (y, m) in GRID) or 1
    scale = (H - pad_b - pad_t) / (vmax * 1.08)

    out = [f'<svg viewBox="0 0 {W} {H}" class="chart" style="min-width:{W}px" role="img" '
           f'aria-label="Monthly resource demand, one stacked band per project">']
    out.append(f'<text class="ax" x="{pad_l - 44}" y="{pad_t - 10}">FTE</text>')
    for k in range(5):
        v = vmax * 1.08 * k / 4
        yy = H - pad_b - v * scale
        out.append(f'<line class="grid" x1="{pad_l}" y1="{yy:.1f}" x2="{W - 8}" y2="{yy:.1f}"/>')
        out.append(f'<text class="ax" x="{pad_l - 8}" y="{yy + 3:.1f}" text-anchor="end">{v:.0f}</text>')
    for i, (y, m) in enumerate(GRID):
        x = pad_l + i * bw + 3
        base = H - pad_b
        for pid in active:                       # largest first => lowest in the stack
            v = S["proj_m"].get((pid, y, m), 0.0)
            if v <= 0.004:
                continue
            h = v * scale
            # O-06: the legend is gone, so the tooltip is now the only thing that
            # names a band. It therefore has to carry everything the legend did and
            # the detail the legend never could - who is actually on the project
            # that month, which is the question a stack of 62 bands invites.
            crew = sorted(S["who"].get((pid, y, m), ()))
            crew_html = "<br>".join(
                f'&#183; {esc(S["PSN"][ps]["person_name"])} '
                f'<span class="tr">{esc(role)}</span>' for ps, role in crew[:8])
            if len(crew) > 8:
                crew_html += f"<br>&#183; and {len(crew) - 8} more"
            tip = (f'<b>{esc(S["P"][pid]["project_name"])}</b> '
                   f'<span class="tr">{esc(S["P"][pid]["project_type"])}</span><br>'
                   f'{calendar.month_abbr[m]} {y} &#183; <b>{v:.2f} FTE</b> '
                   f'({v * HOURS:.0f} h)<br>'
                   f'<span class="th">{len(crew)} '
                   f'{"person" if len(crew) == 1 else "people"} this month</span><br>'
                   f'{crew_html or "&#183; nobody assigned"}')
            out.append(f'<rect class="band" x="{x:.1f}" y="{base - h:.1f}" width="{bw - 8:.1f}" '
                       f'height="{max(0.6, h):.1f}" fill="{colour[pid]}" '
                       f'data-tip="{att(tip)}"></rect>')
            base -= h
        out.append(f'<text class="ax" x="{x + (bw - 8) / 2:.1f}" y="{H - pad_b + 16}" '
                   f'text-anchor="middle">{calendar.month_abbr[m]}</text>')
    out.append(f'<text class="ax yr" x="{pad_l}" y="{H - pad_b + 30}">{GRID[0][0]}</text>')
    out.append(f'<line class="base" x1="{pad_l}" y1="{H - pad_b}" x2="{W - 8}" y2="{H - pad_b}"/>')
    out.append("</svg>")

    # O-06: no legend. With 62 bands it was a list nobody could scan against the
    # chart anyway; identity now comes from hovering the band itself.
    return "".join(out)


def chart_people():
    """One bar per person: mean load over the window, with both thresholds marked."""
    people = sorted(S["PSN"])
    vals = []
    for p in people:
        vs = [S["pers_m"].get((p, y, m), 0.0) for (y, m) in GRID]
        vals.append(sum(vs) / len(vs))
    W, H = 980, 240
    pad_l, pad_b, pad_t = 52, 52, 26
    bw = (W - pad_l - 12) / len(people)
    vmax = max(max(vals), OVER) * 1.15
    scale = (H - pad_b - pad_t) / vmax

    out = [f'<svg viewBox="0 0 {W} {H}" class="chart" role="img" '
           f'aria-label="Mean monthly load per person over the window, against both thresholds">']
    for k in range(4):
        v = vmax * k / 3
        yy = H - pad_b - v * scale
        out.append(f'<line class="grid" x1="{pad_l}" y1="{yy:.1f}" x2="{W - 6}" y2="{yy:.1f}"/>')
        out.append(f'<text class="ax" x="{pad_l - 8}" y="{yy + 3:.1f}" text-anchor="end">{v:.1f}</text>')
    for p, v in zip(people, vals):
        i = people.index(p)
        x = pad_l + i * bw + 3
        h = v * scale
        breach = "over" if v > OVER else ("under" if v < UNDER else "")
        cls = f"bar {breach}" if breach else "bar"
        out.append(f'<rect class="{cls}" x="{x:.1f}" y="{H - pad_b - h:.1f}" '
                   f'width="{bw - 8:.1f}" height="{max(0, h):.1f}" rx="2">'
                   f'<title>{esc(S["PSN"][p]["person_name"])} ({p}) — mean {v:.2f} FTE'
                   f'{" · over ceiling" if breach == "over" else (" · under floor" if breach == "under" else "")}'
                   f'</title></rect>')
        if breach:
            out.append(f'<text class="flag {breach}" x="{x + (bw - 8) / 2:.1f}" '
                       f'y="{H - pad_b - h - 6:.1f}" text-anchor="middle">'
                       f'{"▲" if breach == "over" else "▼"} {v:.2f}</text>')
        out.append(f'<text class="ax rot" x="{x + (bw - 8) / 2:.1f}" y="{H - pad_b + 14}" '
                   f'text-anchor="middle">{p[-3:]}</text>')
    for v, lab, cls in ((OVER, f"over-allocation ceiling {OVER:.2f}", "th-over"),
                        (UNDER, f"under-allocation floor {UNDER:.2f}", "th-under")):
        yy = H - pad_b - v * scale
        out.append(f'<line class="halo" x1="{pad_l}" y1="{yy:.1f}" x2="{W - 6}" y2="{yy:.1f}"/>')
        out.append(f'<line class="{cls}" x1="{pad_l}" y1="{yy:.1f}" x2="{W - 6}" y2="{yy:.1f}"/>')
        out.append(f'<rect class="thbg" x="{pad_l + 2}" y="{yy - 14:.1f}" '
                   f'width="{7.0 * len(lab):.0f}" height="12" rx="2"/>')
        out.append(f'<text class="thlab {cls}" x="{pad_l + 6}" y="{yy - 4:.1f}">{lab}</text>')
    out.append(f'<line class="base" x1="{pad_l}" y1="{H - pad_b}" x2="{W - 6}" y2="{H - pad_b}"/>')
    out.append(f'<text class="ax" x="{pad_l - 44}" y="{pad_t - 2}">FTE</text>')
    out.append(f'<text class="ax" x="{pad_l}" y="{H - 8}">person (PSN-…)</text>')
    out.append("</svg>")
    return "".join(out)


def gantt_months(pid, s, e):
    """Mean monthly FTE this project draws over one period's months (O-10)."""
    vals = [S["proj_all"].get((pid, y, m), 0.0)
            for (y, m) in months((s.year, s.month), (e.year, e.month))]
    return (sum(vals) / len(vals)) if vals else 0.0


def chart_gantt():
    """Timeline per project: bands coloured by period name, milestones as markers.

    Colour is by period name, not by weight (O-10, overturning D-06). Weight survives
    as a lightness step within each hue, and exactly as before in the tooltip - so
    nothing that drove the simulation stopped being visible, it just stopped owning
    the hue.
    """
    pids = S["top"] + [p for p in list(S["P"])[:60] if p not in S["top"]][:5]
    pids = pids[:12]
    lo = min(d(S["P"][p]["start_date"]) for p in pids)
    hi = max(d(S["P"][p]["end_date"]) for p in pids)
    span = (hi - lo).days or 1
    W, rowh, pad_l, pad_t = 1180, 38, 210, 30
    H = pad_t + rowh * len(pids) + 30
    inner = W - pad_l - 14
    wmax = max((s["weight"] or 0) for p in pids for s in S["PER"][p]) or 1

    def x_of(day):
        return pad_l + (day - lo).days / span * inner

    out = [f'<svg viewBox="0 0 {W} {H}" class="chart" style="min-width:{W}px" role="img" '
           f'aria-label="Project timeline: one row per project, bands coloured by period, '
           f'milestones marked">']
    yr = lo.year
    while yr <= hi.year:
        x = x_of(date(yr, 1, 1))
        if x >= pad_l:
            out.append(f'<line class="grid" x1="{x:.1f}" y1="{pad_t - 8}" x2="{x:.1f}" y2="{H - 22}"/>')
            out.append(f'<text class="ax" x="{x:.1f}" y="{pad_t - 12}" text-anchor="middle">{yr}</text>')
        yr += 1

    for i, p in enumerate(pids):
        pr = S["P"][p]
        y = pad_t + i * rowh
        ps, pe = d(pr["start_date"]), d(pr["end_date"])
        nmon = (pe.year - ps.year) * 12 + (pe.month - ps.month) + 1
        # O-10.1: the duration sits under the project name, so the row says how long
        # the project runs without the reader measuring the bar against the axis.
        out.append(f'<text class="rowlab" x="{pad_l - 12}" y="{y + 14}" text-anchor="end">'
                   f'{esc(pr["project_name"][:26])}</text>')
        out.append(f'<text class="rowsub" x="{pad_l - 12}" y="{y + 27}" text-anchor="end">'
                   f'{ps} &#8594; {pe} &middot; {nmon} months</text>')

        for s in sorted(S["PER"][p], key=lambda r: r["period_seq"]):
            bs, be = d(s["period_start"]), d(s["period_end"])
            x0, x1 = x_of(bs), x_of(be)
            w = max(1.5, x1 - x0 - 2)
            lab = period_label(p, s)
            fte = gantt_months(p, bs, be)
            tip = (f'<b>{esc(pr["project_name"])}</b><br>{esc(lab)}<br>'
                   f'{bs} to {be}<br>period weight {s["weight"]:.2f}<br>'
                   f'<b>{fte:.2f} FTE per month</b> on average across this period')
            out.append(f'<rect class="band" x="{x0:.1f}" y="{y + 13}" width="{w:.1f}" '
                       f'height="{rowh - 20}" fill="{band_fill(s["period_name"], s["weight"], wmax)}" '
                       f'rx="2" data-tip="{att(tip)}"><title>{esc(lab)} &#183; '
                       f'{s["weight"]:.2f} &#183; {fte:.2f} FTE/month</title></rect>')
            # Colour never carries the meaning on its own (D-04): where the band is
            # wide enough to hold it, the period names itself.
            # Only label a band that can actually hold the label. Text that overruns
            # its band reads as belonging to the next one along.
            if w > len(lab) * 5.7 + 12:
                out.append(f'<text class="bandlab" x="{x0 + w / 2:.1f}" y="{y + rowh / 2 + 5:.1f}" '
                           f'text-anchor="middle">{esc(lab)}</text>')

        # O-10.3: 'Inspection' is a milestone like any other, so it takes the same
        # marker. One shape for one kind of thing.
        for ms in S["MS"][p]:
            x = x_of(d(ms["milestone_date"]))
            tip = (f'<b>{esc(pr["project_name"])}</b><br>{esc(ms["milestone_name"])}<br>'
                   f'{d(ms["milestone_date"])}')
            cy = y + 1
            out.append(f'<polygon class="ms" points="{x - 5:.1f},{cy:.1f} {x + 5:.1f},{cy:.1f} '
                       f'{x:.1f},{cy + 9:.1f}" data-tip="{att(tip)}">'
                       f'<title>{esc(ms["milestone_name"])} &#183; {d(ms["milestone_date"])}</title>'
                       f'</polygon>')
    out.append("</svg>")

    seen, leg = [], ['<ul class="legend">']
    for nm in ["Before-Start-up", "Start-up", "Conduct (interim)", "Close-out (interim)",
               "Conduct (final)", "Close-out (final)", "After Close-out (final)"]:
        if nm in seen:
            continue
        seen.append(nm)
        leg.append(f'<li><span class="sw" style="background:{PERIOD_HUE[nm]}"></span>{esc(nm)}</li>')
    leg.append('<li><span class="sw tri"></span>milestone</li>')
    leg.append('<li class="hint">darker band = higher period weight</li>')
    leg.append("</ul>")
    return "".join(out) + "".join(leg)


# ---------------------------------------------------------------- tables
def _cells(get, vmax=None, flag=False):
    tds, tot = [], 0.0
    for (y, m) in GRID:
        v = get(y, m)
        tot += v
        if flag and v > OVER:
            tds.append(f'<td class="c over">&#9650; {v:.2f}</td>')
        elif flag and 0 < v < UNDER:
            tds.append(f'<td class="c under">&#9660; {v:.2f}</td>')
        elif v > 0.004:
            i = seq_step(v, vmax) if vmax else None
            tds.append(f'<td class="c c{i}">{v:.2f}</td>' if i is not None
                       else f'<td class="c">{v:.2f}</td>')
        else:
            tds.append('<td class="c z">&middot;</td>')
    return tds, tot


def table_projects():
    """Sorted NewDrug CT, Biosimilar CT, Others; earlier project first.

    Clicking a project name expands it to the people and roles on it.
    """
    # The requested order puts all 34 NewDrug CT projects first, so a flat top-14
    # would never reach the other two types and the ordering could not be seen
    # working. The mock-up therefore samples the head of each type, in order, with
    # a marker row where rows are skipped. The real table lists all 62.
    ordered = sorted([p for p in S["P"] if S["tot"].get(p, 0.0) > 0.004], key=prank)
    listed, breaks, per_type = [], {}, {"NewDrug CT": 6, "Biosimilar CT": 4, "Others": 3}
    seen = {k: 0 for k in per_type}
    for pid in ordered:
        t = S["P"][pid]["project_type"]
        if seen[t] < per_type[t]:
            listed.append(pid)
        seen[t] += 1
    for t, n in per_type.items():
        skipped = seen[t] - n
        if skipped > 0:
            last = [q for q in listed if S["P"][q]["project_type"] == t][-1]
            breaks[last] = (t, skipped)
    vmax = max([S["proj_m"].get((p, y, m), 0.0) for p in listed for (y, m) in GRID] or [1])
    head = "".join(f"<th>{lab}</th>" for lab in MLAB)
    body = []
    for pid in listed:
        tds, tot = _cells(lambda y, m, p=pid: S["proj_m"].get((p, y, m), 0.0), vmax)
        body.append(
            f'<tr class="parent" data-k="p-{pid}" tabindex="0" role="button" '
            f'aria-expanded="false"><th class="rh"><span class="exp">&#9656;</span>'
            f'<span class="nm">{esc(S["P"][pid]["project_name"])}</span> '
            f'{type_pill(pid)}{phase_pill(pid)}'
            f'<span class="sub">{pid} &middot; starts {d(S["P"][pid]["start_date"])}</span></th>'
            f'{"".join(tds)}<td class="tot">{tot:.1f}</td></tr>')
        det = sorted([a for a in S["ASG"] if a["project_id"] == pid],
                     key=lambda a: (a["role_name"], a["person_id"]))
        for a in det:
            k = (pid, a["person_id"], a["role_name"])
            if not any(S["cell_m"].get(k + (y, m), 0.0) > 0.004 for (y, m) in GRID):
                continue
            dtds, dtot = _cells(lambda y, m, kk=k: S["cell_m"].get(kk + (y, m), 0.0))
            body.append(
                f'<tr class="child c-p-{pid}" hidden><th class="rh sub2">'
                f'&#8627; {esc(S["PSN"][a["person_id"]]["person_name"])}'
                f'<span class="role">{esc(a["role_name"])}</span></th>'
                f'{"".join(dtds)}<td class="tot">{dtot:.1f}</td></tr>')
        if pid in breaks:
            t, n = breaks[pid]
            body.append(f'<tr class="skip"><th class="rh">&hellip; {n} more {esc(t)} '
                        f'project{"s" if n != 1 else ""} in this position</th>'
                        f'<td class="c" colspan="{len(GRID) + 1}"></td></tr>')
    rest = [p for p in S["P"] if p not in listed]
    ocells, _ = _cells(lambda y, m: sum(S["proj_m"].get((p, y, m), 0.0) for p in rest))
    ocells = [c.replace('class="c c', 'class="c agg c').replace('class="c"', 'class="c agg"')
              for c in ocells]
    otot = sum(S["tot"].get(p, 0.0) for p in rest)
    body.append(f'<tr class="other"><th class="rh">Other <span class="sub">'
                f'{len(rest)} projects</span></th>{"".join(ocells)}'
                f'<td class="tot">{otot:.1f}</td></tr>')
    tcells, gtot = _cells(lambda y, m: sum(S["proj_m"].get((p, y, m), 0.0) for p in S["P"]))
    body.append(f'<tr class="grand"><th class="rh">All projects</th>'
                f'{"".join(t.replace(chr(34) + "c z" + chr(34), chr(34) + "z" + chr(34)) for t in tcells)}'
                f'<td class="tot">{gtot:.1f}</td></tr>')
    return (f'<table class="grid-t"><thead><tr><th class="rh">Project</th>{head}'
            f'<th class="tot">Total</th></tr></thead><tbody>{"".join(body)}</tbody></table>')


def table_people():
    """Clicking a person expands to their projects and roles.

    The detail rows carry the same order as the project table: NewDrug CT,
    Biosimilar CT, Others, then earlier project first.
    """
    head = "".join(f"<th>{lab}</th>" for lab in MLAB)
    body = []
    for sid in sorted(S["PSN"]):
        tds, tot = _cells(lambda y, m, p=sid: S["pers_m"].get((p, y, m), 0.0), flag=True)
        cap = S["PSN"][sid].get("capacity_fte") or 1.0
        pt = ' <span class="pt">part-time</span>' if cap < 1 else ""
        body.append(
            f'<tr class="parent" data-k="s-{sid}" tabindex="0" role="button" '
            f'aria-expanded="false"><th class="rh"><span class="exp">&#9656;</span>'
            f'<span class="nm">{esc(S["PSN"][sid]["person_name"])}</span>{pt}'
            f'<span class="sub">{sid} &middot; {cap:.2f} FTE &middot; '
            f'{esc(S["PSN"][sid]["department"])}</span></th>'
            f'{"".join(tds)}<td class="tot">{tot / len(GRID):.2f}</td></tr>')
        det = sorted([a for a in S["ASG"] if a["person_id"] == sid],
                     key=lambda a: prank(a["project_id"]))
        for a in det:
            k = (a["project_id"], sid, a["role_name"])
            if not any(S["cell_m"].get(k + (y, m), 0.0) > 0.004 for (y, m) in GRID):
                continue
            dtds, dtot = _cells(lambda y, m, kk=k: S["cell_m"].get(kk + (y, m), 0.0))
            body.append(
                f'<tr class="child c-s-{sid}" hidden><th class="rh sub2">'
                f'&#8627; {esc(S["P"][a["project_id"]]["project_name"])} '
                f'{type_pill(a["project_id"])}{phase_pill(a["project_id"])}'
                f'<span class="role">{esc(a["role_name"])}</span></th>'
                f'{"".join(dtds)}<td class="tot">{dtot / len(GRID):.2f}</td></tr>')
    return (f'<table class="grid-t"><thead><tr><th class="rh">Person</th>{head}'
            f'<th class="tot">Mean</th></tr></thead><tbody>{"".join(body)}</tbody></table>')


# ---------------------------------------------------------------- tiles
def tiles():
    over = sum(1 for (p, y, m), v in S["pers_m"].items() if v > OVER)
    under = sum(1 for (p, y, m), v in S["pers_m"].items() if 0 < v < UNDER)
    total = sum(S["proj_m"].values())
    active = len({p for (p, y, m) in S["proj_m"]})
    ppl = len({p for (p, y, m) in S["pers_m"]})
    items = [("Projects active in window", f"{active}", "of 62 in the file", ""),
             ("People assigned", f"{ppl}", "of 20 in the file", ""),
             ("Total demand", f"{total:,.0f}", f"FTE-months · {total * HOURS:,.0f} hours", ""),
             ("Over-allocated", f"{over}", "person-months above the ceiling", "over"),
             ("Under-allocated", f"{under}", "person-months below the floor", "under")]
    out = ['<div class="tiles">']
    for lab, val, sub, cls in items:
        icon = "▲ " if cls == "over" else ("▼ " if cls == "under" else "")
        out.append(f'<div class="tile {cls}"><div class="tl">{lab}</div>'
                   f'<div class="tv">{icon}{val}</div><div class="ts">{sub}</div></div>')
    out.append("</div>")
    return "".join(out)


# ---------------------------------------------------------------- source tabs
def project_tab():
    # Prefer an example that actually exercises REQ-DSH-10: a caption promising numbered
    # period names against a project that has none leaves the reviewer nothing to check.
    # The pick stays inside the nine rows listed above, so the sub-tables still describe
    # a project the reader can see.
    listed = S["top"][:9]
    pid = next((q for q in listed
                if len({s["period_name"] for s in S["PER"][q]}) < len(S["PER"][q])),
               listed[0])
    pr = S["P"][pid]
    cols = ["project_id", "project_name", "project_type", "project_category", "clinical_phase",
            "outsourcing_type", "EDC_setup", "EDC_system", "planned_member_count",
            "start_date", "end_date", "status"]
    head = '<th class="ins">insert</th>' + "".join(f"<th>{c}</th>" for c in cols)
    body = []
    for qid in S["top"][:9]:
        q = S["P"][qid]
        sel = ' class="sel"' if qid == pid else ""
        tds = "".join(f'<td contenteditable="false">{esc(d(q[c]) if "date" in c else (q[c] or "—"))}</td>'
                      for c in cols)
        body.append(f"<tr{sel}>{INS_TD}{tds}</tr>")
    tbl = (f'<div class="scrollx"><table class="data-t"><thead><tr>{head}</tr></thead>'
           f'<tbody>{"".join(body)}</tbody></table></div>')

    mrows = "".join(
        f'<tr>{INS_TD}<td>{esc(m["milestone_name"])}</td><td>{d(m["milestone_date"])}</td>'
        f'<td>{m["milestone_seq"]}</td><td class="muted">{esc(m.get("note_1") or "—")}</td></tr>'
        for m in sorted(S["MS"][pid], key=lambda r: d(r["milestone_date"])))
    prows = "".join(
        f'<tr>{INS_TD}<td>{p["period_seq"]}</td><td>{esc(period_label(pid, p))}</td>'
        f'<td>{d(p["period_start"])}</td><td>{d(p["period_end"])}</td>'
        f'<td>{p["weight"]:.2f}</td><td class="muted">{esc(p.get("note_1") or "—")}</td></tr>'
        for p in sorted(S["PER"][pid], key=lambda r: r["period_seq"]))
    return tbl, mrows, prows, pr, pid


def person_tab():
    pid = "PSN-001"
    per = S["PSN"][pid]
    cols = ["person_id", "person_name", "department", "primary_role", "capacity_fte"]
    head = '<th class="ins">insert</th>' + "".join(f"<th>{c}</th>" for c in cols)
    body = []
    for qid, q in list(S["PSN"].items())[:9]:
        sel = ' class="sel"' if qid == pid else ""
        body.append(f"<tr{sel}>" + INS_TD
                    + "".join(f"<td>{esc(q[c] if q[c] is not None else '—')}</td>" for c in cols)
                    + "</tr>")
    tbl = (f'<div class="scrollx"><table class="data-t"><thead><tr>{head}</tr></thead>'
           f'<tbody>{"".join(body)}</tbody></table></div>')
    arows = "".join(
        f'<tr>{INS_TD}<td>{esc(a["assignment_id"])}</td>'
        f'<td>{esc(S["P"][a["project_id"]]["project_name"])}</td>'
        f'<td>{esc(a["role_name"])}</td><td>{d(a["assign_start_date"])}</td>'
        f'<td>{d(a["assign_end_date"])}</td><td>{a["person_weight"]:.2f}</td></tr>'
        for a in S["ASG"] if a["person_id"] == pid)
    orows = "".join(
        f'<tr>{INS_TD}<td>{esc(k)}</td><td>{d(w["period_start"])}</td><td>{d(w["period_end"])}</td>'
        f'<td>{w["weight_override"]:.2f}</td><td class="muted">{esc(w["reason"])}</td></tr>'
        for k, ws_ in S["PPW"].items() for w in ws_)

    vals = [S["pers_m"].get((pid, y, m), 0.0) for (y, m) in GRID]
    W, H, padl = 640, 116, 30
    vmax = max(max(vals), OVER) * 1.15
    bw = (W - padl) / len(vals)
    base, top = H - 18, 12
    strip = [f'<svg viewBox="0 0 {W} {H}" class="strip" role="img" '
             f'aria-label="Monthly load for the selected person against both thresholds">']
    for i, v in enumerate(vals):
        h = (v / vmax) * (base - top)
        breach = "over" if v > OVER else ("under" if 0 < v < UNDER else "")
        strip.append(f'<rect class="bar {breach}" x="{padl + i * bw + 2:.1f}" y="{base - h:.1f}" '
                     f'width="{bw - 4:.1f}" height="{max(0, h):.1f}" rx="2">'
                     f'<title>{GRID[i][0]}-{GRID[i][1]:02d}: {v:.2f} FTE</title></rect>')
        strip.append(f'<text class="ax tiny" x="{padl + i * bw + bw / 2:.1f}" y="{H - 5}" '
                     f'text-anchor="middle">{calendar.month_abbr[GRID[i][1]][0]}</text>')
    # thresholds drawn last, so they read over the bars rather than behind them
    for v, cls, lab in ((OVER, "th-over", f"{OVER:.2f}"), (UNDER, "th-under", f"{UNDER:.2f}")):
        y = base - (v / vmax) * (base - top)
        strip.append(f'<line class="halo" x1="{padl}" y1="{y:.1f}" x2="{W - 4}" y2="{y:.1f}"/>')
        strip.append(f'<line class="{cls}" x1="{padl}" y1="{y:.1f}" x2="{W - 4}" y2="{y:.1f}"/>')
        strip.append(f'<text class="thlab {cls}" x="{padl - 5}" y="{y + 3:.1f}" '
                     f'text-anchor="end">{lab}</text>')
    strip.append(f'<line class="base" x1="{padl}" y1="{base}" x2="{W - 4}" y2="{base}"/>')
    strip.append("</svg>")
    return tbl, arows, orows, per, pid, "".join(strip)


def _wmatrix(rows_, row_key, row_label, col_key, val_key, cols=None):
    """A weight table as a matrix: one row per key, one column per period."""
    if cols is None:
        cols = []
        for r in rows_:
            if r[col_key] not in cols:
                cols.append(r[col_key])
    grid, keys = {}, []
    for r in rows_:
        k = row_key(r)
        if k not in keys:
            keys.append(k)
        grid[(k, r[col_key])] = r[val_key]
    vmax = max((r[val_key] for r in rows_), default=1) or 1
    head = "".join(f"<th>{esc(c)}</th>" for c in cols)
    body = []
    for k in keys:
        tds = []
        for c in cols:
            v = grid.get((k, c))
            tds.append('<td class="c z">&middot;</td>' if v is None
                       else f'<td class="c c{seq_step(v, vmax)}">{v:.2f}</td>')
        nm, sub = row_label(k)
        body.append(f'<tr><th class="rh"><span class="nm">{esc(nm)}</span>'
                    f'<span class="sub">{esc(sub)}</span></th>{"".join(tds)}</tr>')
    return (f'<table class="grid-t"><thead><tr><th class="rh">&nbsp;</th>{head}</tr></thead>'
            f'<tbody>{"".join(body)}</tbody></table>')


def general_tab():
    """G-07: the fourth tab - the standards the whole simulation is multiplied by.

    PeriodWeightStandard and RoleFactor were reachable only by opening the workbook,
    yet every figure on the Overall tab is their product. Config sits here too, so
    the thresholds that colour the tables are visible next to what they mean (O-04).
    """
    # Both weight tables are shown as matrices. Flat, RoleFactor is now 249 rows -
    # a list that long is unreadable, and the shape of a role across the periods is
    # the whole point of keying it that way (R-10).
    pws = S["PWS_ROWS"]
    pws_tbl = _wmatrix(pws, lambda r: (r["project_type"], r["clinical_phase"]),
                       lambda k: (k[1], k[0]), "period_name", "weight")

    ct = [r for r in S["RF_ROWS"] if r["project_type"] in S["CT"]]
    ot = [r for r in S["RF_ROWS"] if r["project_type"] not in S["CT"]]
    rf_ct = _wmatrix(ct, lambda r: (r["project_type"], r["clinical_phase"], r["role_name"]),
                     lambda k: (k[2], f"{k[0]} · {k[1]}"), "period_name", "role_factor")
    rf_ot = _wmatrix(ot, lambda r: (r["project_type"], r["role_name"]),
                     lambda k: (k[1], k[0]), "period_name", "role_factor")

    cfg_rows = "".join(
        f'<tr>{INS_TD}<td>{esc(r["parameter"])}</td>'
        f'<td class="num" contenteditable="false">{esc(r["value"])}</td>'
        f'<td class="muted">{esc(r.get("note") or "—")}</td></tr>'
        for r in rows(load_workbook(DUMMY)["Config"]))

    lst = defaultdict(list)
    for r in S["LIST_ROWS"]:
        lst[r["list_name"]].append(str(r["value"]))
    # Read-only: the lists are what the other sheets are checked against, and a row
    # inserted here would be a value with nothing referring to it. No insert control.
    list_rows = "".join(
        f'<tr><td>{esc(k)}</td><td class="vals">{esc(", ".join(v))}</td>'
        f'<td class="num">{len(v)}</td></tr>' for k, v in lst.items())

    return pws_tbl, rf_ct, rf_ot, cfg_rows, list_rows


PTBL, MROWS, PROWS, PROJ, PPID = project_tab()
STBL, AROWS, OROWS, PERS, SPID, STRIP = person_tab()
PWSTBL, RFCT, RFOT, CFGROWS, LISTROWS = general_tab()

# ---------------------------------------------------------------- page
CSS = """
:root{color-scheme:light dark;
 --page:#f9f9f7;--surface:#fcfcfb;--ink:#0b0b0b;--ink2:#52514e;--muted:#898781;
 --grid:#e1e0d9;--base:#c3c2b7;--ring:rgba(11,11,11,.10);
 --s1:#2a78d6;--s2:#eb6834;--s3:#1baf7a;--s4:#eda100;--s5:#e87ba4;--s6:#008300;--s7:#4a3aa7;
 --other:#b8b7b0;
 --good:#0ca30c;--warn:#fab219;--serious:#ec835a;--crit:#d03b3b;
 --overbg:#fbe9e9;--underbg:#fdf3dd;--underink:#8a6100;--accent:#2a78d6;}
@media (prefers-color-scheme:dark){:root:where(:not([data-theme=light])){
 --page:#0d0d0d;--surface:#1a1a19;--ink:#fff;--ink2:#c3c2b7;--muted:#898781;
 --grid:#2c2c2a;--base:#383835;--ring:rgba(255,255,255,.10);
 --s1:#3987e5;--s2:#d95926;--s3:#199e70;--s4:#c98500;--s5:#d55181;--s6:#008300;--s7:#9085e9;
 --other:#5a5a55;--overbg:#3a1e1e;--underbg:#33290f;--underink:#fab219;--accent:#3987e5;}
 :root:where(:not([data-theme=light])) .ty.nd{background:#16283f;color:#8fbdf0}
 :root:where(:not([data-theme=light])) .ty.bs{background:#12302a;color:#6fcfae}
 :root:where(:not([data-theme=light])) .ph1{background:#3a2712;color:#f0b46a}
 :root:where(:not([data-theme=light])) .ph2{background:#12302a;color:#6fcfae}
 :root:where(:not([data-theme=light])) .ph3{background:#16283f;color:#8fbdf0}
 :root:where(:not([data-theme=light])) .ph4{background:#241f3d;color:#b3a6f2}}
:root[data-theme=dark]{--page:#0d0d0d;--surface:#1a1a19;--ink:#fff;--ink2:#c3c2b7;
 --grid:#2c2c2a;--base:#383835;--ring:rgba(255,255,255,.10);
 --s1:#3987e5;--s2:#d95926;--s3:#199e70;--s4:#c98500;--s5:#d55181;--s6:#008300;--s7:#9085e9;
 --other:#5a5a55;--overbg:#3a1e1e;--underbg:#33290f;--underink:#fab219;--accent:#3987e5;}
:root[data-theme=dark] .ty.nd{background:#16283f;color:#8fbdf0}
:root[data-theme=dark] .ty.bs{background:#12302a;color:#6fcfae}
:root[data-theme=dark] .ph1{background:#3a2712;color:#f0b46a}
:root[data-theme=dark] .ph2{background:#12302a;color:#6fcfae}
:root[data-theme=dark] .ph3{background:#16283f;color:#8fbdf0}
:root[data-theme=dark] .ph4{background:#241f3d;color:#b3a6f2}
*{box-sizing:border-box}
body{margin:0;background:var(--page);color:var(--ink);
 font:14px/1.5 system-ui,-apple-system,"Segoe UI",sans-serif;}
.wrap{max-width:1120px;margin:0 auto;padding:0 16px 56px}
.proto{background:var(--warn);color:#0b0b0b;text-align:center;padding:6px 12px;
 font-size:12px;font-weight:600;letter-spacing:.02em}
header{display:flex;flex-wrap:wrap;gap:12px;align-items:baseline;
 padding:16px 0 10px;border-bottom:1px solid var(--ring)}
h1{font-size:17px;margin:0;font-weight:650}
.vers{color:var(--muted);font-size:12px}
.file{margin-left:auto;color:var(--ink2);font-size:12px}
.btn{border:1px solid var(--ring);background:var(--surface);color:var(--ink);
 border-radius:6px;padding:5px 11px;font-size:12.5px;cursor:default}
.btn.primary{background:var(--accent);color:#fff;border-color:transparent}
.banner{display:flex;gap:10px;align-items:center;margin:12px 0;padding:9px 12px;
 border-radius:8px;background:var(--underbg);border:1px solid var(--ring);font-size:13px}
.banner .lk{margin-left:auto;color:var(--accent);text-decoration:underline}
.dirty{margin:10px 0;font-size:13px;color:var(--ink2);line-height:1.55}
.dirty .dot-w{margin-right:6px;vertical-align:0}
.dirty .ok{color:var(--good);font-weight:600}
.dot-w{width:8px;height:8px;border-radius:50%;background:var(--warn);display:inline-block}
.file .tz{color:var(--muted)}
nav{display:flex;gap:2px;margin:14px 0 0;border-bottom:1px solid var(--ring)}
nav button{border:0;background:none;color:var(--ink2);padding:9px 15px;font-size:13.5px;
 cursor:pointer;border-bottom:2px solid transparent;font-family:inherit}
nav button[aria-selected=true]{color:var(--ink);border-bottom-color:var(--accent);font-weight:600}
section[hidden]{display:none}
.panel{background:var(--surface);border:1px solid var(--ring);border-radius:10px;
 padding:14px 16px;margin:16px 0}
.panel h2{font-size:13px;margin:0 0 3px;font-weight:650;letter-spacing:.01em}
.panel p.cap{margin:0 0 12px;color:var(--muted);font-size:12px}
.controls{display:flex;flex-wrap:wrap;gap:10px;align-items:flex-end}
.ctl label{display:block;font-size:11px;color:var(--muted);margin-bottom:3px}
.ctl select,.ctl input{font:inherit;font-size:12.5px;padding:5px 8px;border-radius:6px;
 border:1px solid var(--ring);background:var(--surface);color:var(--ink)}
.tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(168px,1fr));gap:10px;margin:16px 0}
.tile{background:var(--surface);border:1px solid var(--ring);border-radius:10px;padding:12px 14px}
.tile.over{background:var(--overbg)}.tile.under{background:var(--underbg)}
.tl{font-size:11.5px;color:var(--ink2)}
.tv{font-size:25px;font-weight:600;line-height:1.25;font-variant-numeric:tabular-nums}
.tile.over .tv{color:var(--crit)}.tile.under .tv{color:var(--underink)}
.ts{font-size:11px;color:var(--muted)}
.chart{width:100%;height:auto;display:block}
.strip{width:100%;max-width:640px;height:auto}
.grid{stroke:var(--grid);stroke-width:1}
.base{stroke:var(--base);stroke-width:1}
.ax{fill:var(--muted);font-size:10px}
.ax.tiny{font-size:8px}.ax.yr{font-size:9px}
.rowlab{fill:var(--ink);font-size:11.5px;font-weight:600}
.rowsub{fill:var(--muted);font-size:9.5px;font-variant-numeric:tabular-nums}
.bandlab{fill:#fff;font-size:9.5px;font-weight:600;pointer-events:none;
 paint-order:stroke;stroke:rgba(0,0,0,.30);stroke-width:2.4px}
.band{cursor:pointer}
.band:hover{stroke:var(--ink);stroke-width:1.4}
.bar{fill:var(--s1)}.bar.over{fill:var(--crit)}.bar.under{fill:var(--warn)}
.flag{font-size:9.5px;font-weight:600}
.flag.over{fill:var(--crit)}.flag.under{fill:var(--underink)}
.halo{stroke:var(--surface);stroke-width:4;opacity:.9}
.th-over{stroke:var(--crit);stroke-width:1.5;stroke-dasharray:5 3}
.th-under{stroke:var(--warn);stroke-width:1.5;stroke-dasharray:5 3}
.thlab{font-size:9.5px;font-weight:600}
.thbg{fill:var(--surface);opacity:.85}.thlab.th-over{fill:var(--crit);stroke:none}
.thlab.th-under{fill:var(--underink);stroke:none}
.ms{fill:var(--ink);stroke:var(--surface);stroke-width:1.2;cursor:pointer}
.ms:hover{fill:var(--accent)}
.legend{list-style:none;display:flex;flex-wrap:wrap;gap:6px 16px;margin:8px 0 0;padding:0;
 font-size:11.5px;color:var(--ink2)}
.legend .sw{width:11px;height:11px;border-radius:2px;display:inline-block;
 margin-right:6px;vertical-align:-1px}
.legend .sw.ramp{background:linear-gradient(90deg,#86b6ef,#184f95);width:34px}
.legend .sw.dot{border-radius:50%;background:var(--ink2);width:9px;height:9px}
.legend .sw.tri{width:0;height:0;border-radius:0;border-left:6px solid transparent;
 border-right:6px solid transparent;border-top:9px solid var(--ink);background:none}
.legend .hint{color:var(--muted);font-style:italic}
/* O-06 / O-10: a real hover pop-up, not the native SVG title. The native tooltip
   cannot show a list of people, and waits half a second before appearing. */
.tip{position:fixed;z-index:50;max-width:320px;pointer-events:none;
 background:var(--surface);color:var(--ink);border:1px solid var(--base);
 border-radius:8px;padding:8px 10px;font-size:11.5px;line-height:1.5;
 box-shadow:0 6px 22px rgba(0,0,0,.18)}
.tip b{font-weight:650}
.tip .tr{color:var(--muted);font-size:10.5px}
.tip .th{color:var(--ink2);font-weight:600;font-size:10.5px;
 display:inline-block;margin-top:2px}
.btn.tiny{font-size:10.5px;padding:1px 7px;border-radius:6px;line-height:1.6}
td.ins,th.ins{width:1%;white-space:nowrap;text-align:center}
td.num{text-align:right;font-variant-numeric:tabular-nums}
td.vals{white-space:normal;color:var(--ink2)}
.scrollx{overflow-x:auto;-webkit-overflow-scrolling:touch}
/* A table taller or wider than its panel must scroll INSIDE the panel. Without the
   cap a long sub-table simply grew the page and pushed everything below it down. */
.scrollx.tall{max-height:340px;overflow:auto}
.scrollx.tall thead th{position:sticky;top:0;z-index:2}
.scrollx.tall .grid-t thead th,.scrollx.tall .data-t thead th{background:var(--page)}
table{border-collapse:collapse;width:100%;font-size:12px}
.grid-t th,.grid-t td{border:1px solid var(--grid);padding:4px 7px;text-align:right;
 font-variant-numeric:tabular-nums;white-space:nowrap}
.grid-t thead th{background:var(--page);color:var(--ink2);font-weight:600;font-size:10.5px;
 text-align:center;line-height:1.25}
.grid-t th.rh{text-align:left;font-weight:500;min-width:200px;position:sticky;left:0;
 background:var(--surface);z-index:1}
.grid-t thead th.rh{background:var(--page)}
.rh .sub{display:block;color:var(--muted);font-size:10px}
.rh .exp{color:var(--muted);margin-right:3px}
.pt{font-size:10px;color:var(--ink2);border:1px solid var(--ring);border-radius:9px;padding:0 5px}
td.c{min-width:52px}
td.z{color:var(--muted)}
td.over{background:var(--overbg);color:var(--crit);font-weight:600}
td.under{background:var(--underbg);color:var(--underink);font-weight:600}
tr.other td,tr.other th{color:var(--ink2)}
td.agg{background:var(--page);color:var(--ink2)}
tr.grand th,tr.grand td{border-top:2px solid var(--base);font-weight:650;background:var(--page)}
.tot{font-weight:600}
.data-t th,.data-t td{border:1px solid var(--grid);padding:5px 8px;text-align:left;white-space:nowrap}
.data-t thead th{background:var(--page);color:var(--ink2);font-weight:600;font-size:10.5px}
.data-t tr.sel td{background:color-mix(in srgb,var(--accent) 11%,transparent)}
.muted{color:var(--muted)}
.two{display:grid;grid-template-columns:1fr 1fr;gap:16px}
/* A grid track defaults to min-content:auto, so a wide table stretches the track
   instead of scrolling inside it. This is what keeps the panel on the page. */
.two>*{min-width:0}
@media(max-width:820px){.two{grid-template-columns:1fr}}
.note{font-size:11.5px;color:var(--muted);margin-top:9px}
.filterbar{background:var(--surface);border-color:var(--accent)}
.scope{font-weight:400;font-size:11px;color:var(--muted);
 border:1px solid var(--ring);border-radius:9px;padding:1px 7px;margin-left:6px}
.btn.reset{border-color:var(--accent);color:var(--accent)}
.scrollx>.chart,.scrollx>.strip{display:block}
.legendbox{max-height:118px;overflow-y:auto;margin-top:8px;padding-right:6px;
 border-top:1px solid var(--ring)}
.legend.proj{display:grid;grid-template-columns:repeat(auto-fill,minmax(210px,1fr));
 gap:2px 14px;margin:8px 0 2px}
.legend.proj li{display:flex;align-items:center;gap:0}
.legend .lv{margin-left:auto;color:var(--muted);font-variant-numeric:tabular-nums;
 font-size:10.5px;padding-left:8px}
.ty{font-size:9.5px;border-radius:9px;padding:1px 6px;margin-left:6px;
 border:1px solid var(--ring);white-space:nowrap}
.ty.nd{background:#e7f0fb;color:#184f95}.ty.bs{background:#eaf7f1;color:#0d6b4a}
.ty.ot{background:var(--page);color:var(--ink2)}
.ph{font-size:9.5px;border-radius:9px;padding:1px 6px;margin-left:4px;font-weight:600;
 white-space:nowrap}
.ph1{color:#8a4b00;background:#fdf0e3}.ph2{color:#0d6b4a;background:#eaf7f1}
.ph3{color:#184f95;background:#e7f0fb}.ph4{color:#5b3ea8;background:#efeafb}
tr.parent{cursor:pointer}
tr.parent:hover th.rh{background:color-mix(in srgb,var(--accent) 8%,var(--surface))}
tr.parent .nm{font-weight:600}
tr.parent[aria-expanded=true] .exp{transform:rotate(90deg);display:inline-block}
tr.child th.rh.sub2{font-weight:400;padding-left:20px;color:var(--ink2);font-size:11.5px}
tr.child .role{color:var(--muted);font-size:10.5px;margin-left:6px}
tr.child td{background:color-mix(in srgb,var(--accent) 4%,var(--surface))}
tr.skip th,tr.skip td{background:var(--page);color:var(--muted);font-style:italic;
 font-size:11px;font-weight:400}
:root{--h0:#cde2fb;--h1:#b7d3f6;--h2:#9ec5f4;--h3:#86b6ef;--h4:#6da7ec;--h5:#5598e7;
 --h6:#3987e5;--h7:#2a78d6;--h8:#256abf;--h9:#1c5cab;--h10:#184f95;--hink:#0b0b0b;--hink2:#fff;}

@media(prefers-color-scheme:dark){:root:where(:not([data-theme=light])){
 --h0:#16233a;--h1:#182b48;--h2:#1a3355;--h3:#1c3c64;--h4:#1e4573;--h5:#215083;
 --h6:#245c98;--h7:#2a6cb4;--h8:#3180d4;--h9:#4a92e2;--h10:#6da7ec;--hink:#fff;--hink2:#fff;}}
:root[data-theme=dark]{
 --h0:#16233a;--h1:#182b48;--h2:#1a3355;--h3:#1c3c64;--h4:#1e4573;--h5:#215083;
 --h6:#245c98;--h7:#2a6cb4;--h8:#3180d4;--h9:#4a92e2;--h10:#6da7ec;--hink:#fff;--hink2:#0b0b0b;}

.c0{background:var(--h0)}.c1{background:var(--h1)}.c2{background:var(--h2)}
.c3{background:var(--h3)}.c4{background:var(--h4)}.c5{background:var(--h5)}
.c0,.c1,.c2,.c3,.c4,.c5{color:var(--hink)}
.c6{background:var(--h6)}.c7{background:var(--h7)}.c8{background:var(--h8)}
.c9{background:var(--h9)}.c10{background:var(--h10)}
.c6,.c7,.c8{color:#fff}.c9,.c10{color:var(--hink2)}
"""

JS = """
document.querySelectorAll('tr.parent').forEach(function(tr){
  function toggle(){
    var open = tr.getAttribute('aria-expanded') === 'true';
    tr.setAttribute('aria-expanded', open ? 'false' : 'true');
    document.querySelectorAll('tr.c-' + tr.dataset.k).forEach(function(c){ c.hidden = open; });
  }
  tr.addEventListener('click', toggle);
  tr.addEventListener('keydown', function(e){
    if (e.key === 'Enter' || e.key === ' ') { e.preventDefault(); toggle(); }
  });
});
// O-06 / O-10: hover pop-up. Follows the cursor and flips before it runs off the
// right or bottom edge, so a band at the far end of a scrolled chart is still readable.
(function(){
  var tip = document.getElementById('tip');
  function place(e){
    var pad = 14, w = tip.offsetWidth, h = tip.offsetHeight;
    var x = e.clientX + pad, y = e.clientY + pad;
    if (x + w > window.innerWidth - 8)  { x = e.clientX - w - pad; }
    if (y + h > window.innerHeight - 8) { y = e.clientY - h - pad; }
    tip.style.left = Math.max(8, x) + 'px';
    tip.style.top  = Math.max(8, y) + 'px';
  }
  document.addEventListener('mouseover', function(e){
    var el = e.target.closest('[data-tip]');
    if (!el) return;
    tip.innerHTML = el.dataset.tip;
    tip.hidden = false;
    place(e);
  });
  document.addEventListener('mousemove', function(e){
    if (!tip.hidden) place(e);
  });
  document.addEventListener('mouseout', function(e){
    if (e.target.closest('[data-tip]')) tip.hidden = true;
  });
})();
document.querySelectorAll('nav button').forEach(function(b){
  b.addEventListener('click',function(){
    document.querySelectorAll('nav button').forEach(function(x){x.setAttribute('aria-selected','false');});
    document.querySelectorAll('section.tab').forEach(function(s){s.hidden=true;});
    b.setAttribute('aria-selected','true');
    document.getElementById(b.dataset.tab).hidden=false;
    window.scrollTo({top:0});
  });
});
"""

html = f"""<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>PRAP — UI prototype v0.7</title>
<style>{CSS}</style>
<div class="proto">PROTOTYPE — layout and components only. Figures are a fixed snapshot;
nothing on this page loads, calculates or exports.</div>
<div class="wrap">
<header>
  <h1>Project Resource Assignment Program</h1>
  <span class="vers">{APP_VERSION} · expects source schema v{SCHEMA_EXPECTED}</span>
  <span class="file">PRAP_SourceData_Dummy_v1.8.xlsx · loaded 2026-08-01 09:14
    <span class="tz">(GMT+9, KST)</span></span>
  <button class="btn">Load workbook</button>
  <button class="btn primary">Export</button>
</header>

<div class="banner"><strong>Last import: 0 errors, 1 information notice.</strong>
  PRJ-002 — an Inspection on or before the final DB lock is treated as a marker (V-21).
  <a class="lk" href="#">Open full report</a></div>

<div class="dirty"><span class="dot-w"></span> <strong>3 unsaved edits</strong> ·
  <span class="ok">all 3 pass validation</span> — export writes every validated edit back
  to the workbook. An edit that fails a rule is rejected at entry and never reaches
  the model, so nothing invalid can be exported.</div>

<div class="panel filterbar">
  <h2>Horizon and filters <span class="scope">applies to every tab</span></h2>
  <p class="cap">One setting drives the whole page — the charts, both Overall tables and
    both source-data tabs. Opens on 24 months.</p>
  <div class="controls">
    <div class="ctl"><label>From</label><input value="2027-01" size="8"></div>
    <div class="ctl"><label>To</label><input value="2027-12" size="8"></div>
    <div class="ctl"><label>&nbsp;</label><button class="btn">Expand to all projects</button></div>
    <div class="ctl"><label>Project type</label><select><option>All</option>
      <option>NewDrug CT</option><option>Biosimilar CT</option><option>Others</option></select></div>
    <div class="ctl"><label>Clinical phase</label><select><option>All</option>
      <option>Phase 1</option><option>Phase 2</option><option>Phase 3</option>
      <option>Phase 4</option></select></div>
    <div class="ctl"><label>Project</label><select><option>All (62)</option></select></div>
    <div class="ctl"><label>Person</label><select><option>All (20)</option></select></div>
    <div class="ctl"><label>Role</label><select><option>All (13)</option></select></div>
    <div class="ctl"><label>Department</label><select><option>All (5)</option></select></div>
    <div class="ctl"><label>&nbsp;</label><button class="btn reset">Reset filters</button></div>
  </div>
</div>

<nav role="tablist">
  <button role="tab" aria-selected="true" data-tab="t-overall">Overall</button>
  <button role="tab" aria-selected="false" data-tab="t-proj">Source data (project)</button>
  <button role="tab" aria-selected="false" data-tab="t-pers">Source data (person)</button>
  <button role="tab" aria-selected="false" data-tab="t-gen">General assumptions</button>
</nav>

<section class="tab" id="t-overall">

  <div class="panel">
    <h2>Project timeline</h2>
    <p class="cap">One row per project, with its start, end and length under the name.
      Bands are coloured by period — grey before start-up, red through start-up, green
      through conduct, orange at close-out, dark grey after — and shaded darker as the
      period weight rises. Trials with an interim DB lock show two conduct stretches,
      named <em>Conduct (interim)</em> and <em>Conduct (final)</em>; trials without one show
      a single <em>Conduct (final)</em>. <strong>Hover any band</strong>
      for its dates, weight and the FTE per month the project draws across it; hover a
      marker for the milestone.</p>
    <div class="scrollx">{chart_gantt()}</div>
  </div>

  {tiles()}

  <div class="panel">
    <h2>Monthly demand by project</h2>
    <p class="cap">One band per project, ordered by total resource with the largest on
      the baseline. Every “Others” project is grey; trials take the extended colour set.
      <strong>Hover any band</strong> for the project, its FTE that month and who is on it —
      with 62 projects that is where identity lives now, not in a legend nobody could scan
      against the chart.</p>
    <div class="scrollx">{chart_stacked()}</div>
  </div>

  <div class="panel">
    <h2>Resource by project</h2>
    <p class="cap">Sorted NewDrug CT, then Biosimilar CT, then Others; earlier projects
      first. <strong>Click a project name</strong> to expand it to the people and roles on
      it, and again to collapse.</p>
    <div class="scrollx">{table_projects()}</div>
  </div>

  <div class="panel">
    <h2>Mean load per person</h2>
    <p class="cap">Averaged over the window, against the two absolute thresholds. Bars that
      breach are labelled with their value, so the flag never rests on colour alone.</p>
    <div class="scrollx">{chart_people()}</div>
  </div>

  <div class="panel">
    <h2>Resource by person</h2>
    <p class="cap">Summed across every project. ▲ above the ceiling, ▼ below the floor.
      <strong>Click a person name</strong> to expand it to their projects and roles, in the
      same order as the project table.</p>
    <div class="scrollx">{table_people()}</div>
    <p class="note">Under-allocation is reported as a run of three or more consecutive
      months, not per month — the run is what matters, a single quiet month is not.</p>
  </div>

</section>

<section class="tab" id="t-proj" hidden>
  <div class="panel">
    <h2>Projects</h2>
    <p class="cap">All 23 columns, sortable and filterable. Every field is editable;
      changing an identifier cascades to the rows that reference it.
      <strong>+ row</strong> inserts a new row directly below the one you press it on,
      in this and every other section.</p>
    {PTBL}
    <p class="note">Showing 9 of 62 rows in this mock-up, with the selected row tinted. The real table lists all 62, sorted and filtered.</p>
  </div>
  <div class="two">
    <div class="panel">
      <h2>Milestones — {esc(PROJ['project_name'])}</h2>
      <p class="cap">Only CTA submission and the DB locks set period boundaries.</p>
      <div class="scrollx tall"><table class="data-t"><thead><tr><th class="ins">insert</th>
        <th>milestone</th><th>date</th><th>seq</th><th>note_1</th></tr></thead>
        <tbody>{MROWS}</tbody></table></div>
    </div>
    <div class="panel">
      <h2>Periods — {esc(PROJ['project_name'])}</h2>
      <p class="cap">Derived from the milestones above. Period names are unique within a
        project, so <code>project_id + period_name</code> identifies a row.
        <button class="btn">Recompute periods</button></p>
      <div class="scrollx tall"><table class="data-t"><thead><tr><th class="ins">insert</th>
        <th>seq</th><th>period</th><th>start</th><th>end</th><th>weight</th><th>note_1</th>
        </tr></thead><tbody>{PROWS}</tbody></table></div>
    </div>
  </div>
</section>

<section class="tab" id="t-pers" hidden>
  <div class="panel">
    <h2>People</h2>
    <p class="cap">Editable, same rules as the project table — including <strong>+ row</strong>
      on every row, which inserts directly below it.</p>
    {STBL}
    <p class="note">Showing 9 of 20 rows in this mock-up.</p>
  </div>
  <div class="panel">
    <h2>Utilisation — {esc(PERS['person_name'])} ({SPID})</h2>
    <p class="cap">Monthly load across the window. Dashed lines are the ceiling and floor \u2014
      the same two figures for everyone, since both thresholds are absolute.</p>
    <div class="scrollx">{STRIP}</div>
  </div>
  <div class="two">
    <div class="panel">
      <h2>Assignments — {SPID}</h2>
      <div class="scrollx tall"><table class="data-t"><thead><tr><th class="ins">insert</th>
        <th>id</th><th>project</th><th>role</th><th>start</th><th>end</th><th>weight</th>
        </tr></thead><tbody>{AROWS}</tbody></table></div>
    </div>
    <div class="panel">
      <h2>Weight overrides</h2>
      <p class="cap">Replaces person_weight for the window it covers.</p>
      <div class="scrollx tall"><table class="data-t"><thead><tr><th class="ins">insert</th>
        <th>assignment</th><th>start</th><th>end</th><th>override</th><th>reason</th>
        </tr></thead><tbody>{OROWS}</tbody></table></div>
    </div>
  </div>
</section>

<section class="tab" id="t-gen" hidden>
  <div class="panel">
    <h2>Standard period weights <span class="scope">PeriodWeightStandard</span></h2>
    <p class="cap">The weight every clinical trial period is multiplied by, selected by the
      project's type and clinical phase. Shown as a matrix rather than 48 rows, because it is
      a standard and standards are read across, not down. The shading follows the same ramp as
      the Overall tables, so a heavier weight reads as a stronger cell. <strong>Others</strong>
      projects do not appear here — their weights are entered by hand on each project.</p>
    <div class="scrollx">{PWSTBL}</div>
  </div>

  <div class="panel">
    <h2>Role factors — clinical trials <span class="scope">RoleFactor</span></h2>
    <p class="cap">What one person in this role costs the project per month, before their own
      weight and the period weight are applied. Keyed on project type, clinical phase, period
      and role — so a role's burden can move across the life of a project rather than being one
      number for the whole run. Read a row across: the database programmer peaks at start-up,
      the data associator through conduct, the analyst at lock.</p>
    <div class="scrollx tall">{RFCT}</div>
    <p class="note">All 240 clinical-trial rows, folded into 40 matrix rows of six periods
      each — the flat sheet is 2 types × 4 phases × 6 periods × 5 roles, plus 9 rows for
      Others. That is a lot to maintain by hand, and it now varies over the same three
      dimensions as the period weight above; see the note on double-counting in the
      specification, sheet 05.</p>
  </div>

  <div class="panel">
    <h2>Role factors — Others <span class="scope">RoleFactor</span></h2>
    <p class="cap">The same table for non-trial projects, which carry no clinical phase and
      run on the three-period set.</p>
    <div class="scrollx">{RFOT}</div>
  </div>

  <div class="two">
    <div class="panel">
      <h2>Configuration <span class="scope">Config</span></h2>
      <p class="cap">The thresholds and settings the whole page reads. Both allocation
        thresholds are absolute — they are not scaled by anyone's capacity.</p>
      <div class="scrollx"><table class="data-t"><thead><tr><th class="ins">insert</th>
        <th>parameter</th><th>value</th><th>note</th></tr></thead>
        <tbody>{CFGROWS}</tbody></table></div>
      <p class="note">The display unit lives here rather than in the filter bar: it is a
        setting, not a filter, and it changes how every figure is written rather than
        which figures are shown.</p>
      <div class="controls">
        <div class="ctl"><label>Display unit</label><select><option>FTE</option>
          <option>Hours</option></select></div>
      </div>
    </div>
    <div class="panel">
      <h2>Value lists <span class="scope">Lists</span></h2>
      <p class="cap">What each list-typed column will accept. A value outside its list is
        kept and reported (V-11), never silently dropped. Read-only — a value here with
        nothing referring to it would be noise, so there is no insert control.</p>
      <div class="scrollx tall"><table class="data-t"><thead><tr><th>list_name</th>
        <th>values</th><th>n</th></tr></thead><tbody>{LISTROWS}</tbody></table></div>
    </div>
  </div>
</section>
</div>
<div id="tip" class="tip" hidden></div>
<script>{JS}</script>
"""

OUT.parent.mkdir(exist_ok=True)
OUT.write_text(html, encoding="utf-8")
print(f"Written: {OUT}  ({len(html):,} bytes)")
