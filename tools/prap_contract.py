"""Assemble the PRAP contract - one machine-readable description of everything an
outside program (or an AI agent) needs in order to read, write and reason about a
PRAP source workbook without opening the application.

Nothing here is retyped. Every fact is read from the artifact that already owns it:

    the sheets, columns, value lists, Config defaults, date columns
        -> tools/build_source_workbook.py, which BUILDS the template
    the validation rules V-01..V-24
        -> docs/PRAP_Development_Plan_v2.x.xlsx sheet 04_Data_Model, the baselined table
    the application version, expected schema version, derived columns,
        column help text, period sets, milestone list
        -> app/PRAP.html, the running code

So the contract cannot drift from the thing it describes: if a column is added to the
template, or a rule reworded in the plan, the contract changes on the next build, and
tools/check_consistency.py fails if the three disagree.

    python tools/build_ai_reference.py      # writes docs/prap_contract.json
"""

import ast
import json
import re
from pathlib import Path

from openpyxl import load_workbook

import build_source_workbook as B

ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "app" / "PRAP.html"

CONTRACT_VERSION = "1.0"

# The current issue of each controlled document. check_consistency.py verifies the
# files exist and that the versions agree with the application's provenance strip.
PLAN = "PRAP_Development_Plan_v2.46.xlsx"
SPEC = "PRAP_Programming_Specification_v1.20.xlsx"
UIL = "PRAP_UI_Component_List_v1.0.xlsx"
TEMPLATE = f"PRAP_SourceData_Template_v{B.TEMPLATE_VERSION}.xlsx"
DUMMY = f"PRAP_SourceData_Dummy_v{B.DUMMY_VERSION}.xlsx"
DUMMY_SMALL = f"PRAP_SourceData_Dummy_10x10_v{B.DUMMY_SMALL_VERSION}.xlsx"


# ---------------------------------------------------------------- from the app
def app_text():
    return APP.read_text(encoding="utf-8")


def app_const(name, src=None):
    """Read `const NAME = <literal>;` out of the application source.

    The application is the running truth for these, and a copy typed here would be a
    second truth. Only literals are read - anything computed stays in the app.
    """
    src = src or app_text()
    m = re.search(rf'^const {re.escape(name)} = (.+?);\s*$', src, re.M)
    if not m:
        raise SystemExit(f"prap_contract: {name} not found in app/PRAP.html")
    raw = m.group(1).strip()
    if raw.startswith('"') or raw.startswith("'"):
        return raw[1:-1]
    if re.fullmatch(r'-?\d+', raw):
        return int(raw)
    raise SystemExit(f"prap_contract: {name} is not a simple literal: {raw!r}")


def app_js_object(name, src=None):
    """Read a one-level `const NAME = { ... };` or `[ ... ];` block as Python.

    JavaScript object literals with bare keys are not JSON, so the keys are quoted
    before parsing. Only used for the flat, string-valued tables (DERIVED_COLS,
    CLINICAL_PERIODS, OTHER_PERIODS, KEY_COL).
    """
    src = src or app_text()
    i = src.index(f"const {name} = ")
    j = i + len(f"const {name} = ")
    open_c = src[j]
    close_c = {"{": "}", "[": "]"}[open_c]
    depth, k, instr, quote = 0, j, False, ""
    while k < len(src):
        c = src[k]
        if instr:
            if c == "\\":
                k += 2
                continue
            if c == quote:
                instr = False
        elif c in "\"'":
            instr, quote = True, c
        elif c == open_c:
            depth += 1
        elif c == close_c:
            depth -= 1
            if depth == 0:
                break
        k += 1
    body = src[j:k + 1]
    body = re.sub(r'//[^\n]*', '', body)
    body = re.sub(r'([{,]\s*)([A-Za-z_][\w]*)\s*:', r'\1"\2":', body)
    return ast.literal_eval(body)


# --------------------------------------------------------------- from the plan
def rules_from_plan(plan_path):
    """The V-rule table, read from sheet 04_Data_Model of the baselined plan."""
    ws = load_workbook(plan_path)["04_Data_Model"]
    out = {}
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if isinstance(a, str) and re.fullmatch(r'V-\d\d', a.strip()):
            out[a.strip()] = {"statement": (row[1] or "").strip(),
                              "effect": (row[2] or "").strip()}
    if not out:
        raise SystemExit(f"prap_contract: no V-rules found in {plan_path.name}")
    return dict(sorted(out.items()))


def rules_reported_by_app(src=None):
    """Which rules the application actually raises, and at what severity.

    Read from the calls themselves rather than from a list, so a rule that is
    documented but never reported shows up as the gap it is.
    """
    src = src or app_text()
    found = {}
    for sev, rule in re.findall(r'add\("(\w+)","(V-\d\d)"', src):
        found.setdefault(rule, set()).add(sev)
    for sev, rule in re.findall(r'sev:"(\w+)", rule:"(V-\d\d)"', src):
        found.setdefault(rule, set()).add(sev)
    order = {"fatal": 0, "error": 1, "warning": 2, "information": 3}
    return {k: sorted(v, key=lambda s: order.get(s, 9)) for k, v in sorted(found.items())}


# ------------------------------------------------------------------- assembly
KIND_TYPE = {"key": "identifier", "calc": "derived", "fill": "decimal", "": "text"}

SHEET_ROLE = {
    "Project": {"role": "master", "key": ["project_id"], "parent": None},
    "Milestone": {"role": "child", "key": ["project_id", "milestone_name", "milestone_date"],
                  "parent": "Project"},
    "ProjectPeriod": {"role": "child", "key": ["project_id", "period_name"], "parent": "Project"},
    # Schema 6 put the work scope in both keys. A row whose work_scope_type is EMPTY
    # applies to every scope, and a project falls back to it when there is no row for
    # its own - so an outside program reading these tables must do the same, or it will
    # report a missing weight where the application finds one.
    "PeriodFTEStandard": {"role": "reference",
                             "key": ["project_type", "clinical_phase", "work_scope_type",
                                     "period_name"],
                             "fallback": "work_scope_type empty means every scope",
                             "parent": None},
    "RoleFactor": {"role": "reference",
                   "key": ["project_type", "clinical_phase", "work_scope_type",
                           "period_name", "role_name"],
                   "fallback": "work_scope_type empty means every scope",
                   "parent": None},
    "Person": {"role": "master", "key": ["person_id"], "parent": None},
    "Assignment": {"role": "child", "key": ["assignment_id"], "parent": "Person"},
    "PersonPeriodWeight": {"role": "child", "key": ["assignment_id", "period_start"],
                           "parent": "Assignment"},
    # Schema 9. The one sheet whose parent depends on a value IN the row: `scope` says
    # whether ref_id names a project or an assignment, so it is keyed on all three of
    # scope, ref_id and month and has no single parent sheet. A reading program must
    # branch on scope rather than assume either table.
    "MonthlyEstimate": {"role": "child", "key": ["scope", "ref_id", "month"],
                        "parent": "Project|Assignment"},
    "Lists": {"role": "vocabulary", "key": ["list_name", "value"], "parent": None},
    "Config": {"role": "settings", "key": ["parameter"], "parent": None},
}

FOREIGN_KEYS = [
    {"from": "Milestone.project_id", "to": "Project.project_id", "rule": "V-01"},
    {"from": "ProjectPeriod.project_id", "to": "Project.project_id", "rule": "V-01"},
    {"from": "Assignment.project_id", "to": "Project.project_id", "rule": "V-01"},
    {"from": "Assignment.person_id", "to": "Person.person_id", "rule": "V-02"},
    {"from": "PersonPeriodWeight.assignment_id", "to": "Assignment.assignment_id",
     "rule": "V-24"},
]


def build(plan_path):
    src = app_text()
    derived = app_js_object("DERIVED_COLS", src)
    key_col = app_js_object("KEY_COL", src)
    plan_rules = rules_from_plan(plan_path)
    reported = rules_reported_by_app(src)

    sheets = {}
    for name, cols in B.SHEETS.items():
        dates = set(B.DATE_COLS.get(name, []))
        drops = B.DROPDOWNS.get(name, {})
        entries = []
        for col, note, kind in cols:
            t = "date" if col in dates else KIND_TYPE[kind]
            if kind == "calc":
                t = "derived"
            e = {"name": col, "type": t, "editable": kind != "calc", "note": note}
            if col in drops:
                e["value_list"] = drops[col]
            entries.append(e)
        meta = SHEET_ROLE[name]
        sheets[name] = {
            "role": meta["role"],
            "parent_sheet": meta["parent"],
            "key": meta["key"],
            "identifier_column": key_col.get(name),
            "derived_columns": derived.get(name, []),
            "date_columns": sorted(dates),
            "columns": entries,
        }

    rules = {}
    for rid, r in plan_rules.items():
        rules[rid] = {
            "statement": r["statement"],
            "effect": r["effect"],
            "severity_reported_by_app": reported.get(rid, []),
            "enforced_by_application": rid in reported or rid == "V-17",
        }
    # V-00 is the application's own gate: it is not a data rule from the plan but an
    # agent that reads findings will meet it, so it belongs in the catalogue.
    rules["V-00"] = {
        "statement": "The workbook contains all ten required sheets, and every sheet has "
                     "a header row whose names match this contract.",
        "effect": "Fatal - nothing is loaded, and the findings report names the missing sheet.",
        "severity_reported_by_app": ["fatal", "error"],
        "enforced_by_application": True,
    }

    return {
        "$comment": "Machine-readable contract for the Project Resource Assignment Program "
                    "(PRAP). Generated by tools/build_ai_reference.py - do not edit by hand.",
        "contract_version": CONTRACT_VERSION,
        "generated_by": "tools/build_ai_reference.py",
        "application": {
            "file": "app/PRAP.html",
            "version": app_const("APP_VERSION", src),
            "kind": "single-file offline HTML application, no network access, no install",
            "reads": ["Excel .xlsx source workbook", "PRAP JSON interchange file (.prap.json)"],
            "writes": ["Excel .xlsx source workbook", "PRAP JSON interchange file (.prap.json)"],
            "ways_in": BLANK_START,
        },
        "schema_version": app_const("SCHEMA_EXPECTED", src),
        "documents": {
            "development_plan": f"docs/{PLAN}",
            "programming_specification": f"docs/{SPEC}",
            "ui_component_list": f"docs/{UIL}",
            "source_data_template": f"templates/{TEMPLATE}",
            "worked_example_large": f"templates/{DUMMY}",
            "worked_example_small": f"templates/{DUMMY_SMALL}",
            "agent_guide": "docs/PRAP_AI_Agent_Guide.md",
        },
        "sheets": sheets,
        "sheet_order": list(B.SHEETS.keys()),
        "foreign_keys": FOREIGN_KEYS,
        "value_lists": {name: list(vals) for name, vals in B.LISTS},
        "config": {
            p: {"default": v, "meaning": note} for p, v, note in B.CONFIG
        },
        "period_sets": {
            "clinical": app_js_object("CLINICAL_PERIODS", src),
            "others": app_js_object("OTHER_PERIODS", src),
        },
        "project_types": {
            "clinical": ["NewDrug CT", "Biosimilar CT"],
            "non_clinical": ["Others"],
        },
        "calculation": CALCULATION,
        "period_derivation": PERIOD_DERIVATION,
        "validation_rules": dict(sorted(rules.items())),
        "date_handling": DATE_HANDLING,
        "editing_contract": EDITING_CONTRACT,
        "interchange_format": INTERCHANGE,
        "agent_recipes": RECIPES,
        "verification": VERIFICATION,
    }


# ------------------------------------------------------ narrative, in one place
BLANK_START = {
    "load_a_file": "Drop a .xlsx workbook or a .prap.json file on the page. This is the "
                   "path for anything you produce.",
    "start_blank": "'Start blank' begins a plan with nothing in it but the standard value "
                   "lists and settings, and opens every tab for typing. It is how someone "
                   "plans from scratch without a workbook.",
    "what_blank_contains": "The Lists and Config sheets of the delivered template, plus a "
                           "complete PeriodFTEStandard and RoleFactor grid built from "
                           "those lists - every (type, phase, period) and (type, phase, "
                           "period, role) combination, so nothing silently falls back to "
                           "1.00 and V-23 never fires.",
    "placeholders": "Every seeded weight and role factor is 1.00 and says so in its note. "
                    "They are NOT a company standard. If a user hands you a plan started "
                    "this way, check whether those figures are still 1.00 before reading "
                    "the FTE totals as real - until they are set, the load reduces to "
                    "person_weight x month_coverage.",
    "both_end_the_same": "A plan started blank exports to the same .xlsx and the same "
                         ".prap.json as any other, so nothing downstream needs to know "
                         "which way it began.",
}

CALCULATION = {
    "statement": "monthly_load_fte = period_weight x role_factor x person_weight x month_coverage",
    "unit": "FTE. 1.00 FTE is fte_hours_per_month hours (Config, default 160 = 8 h/day x "
            "5 days/week x 4 weeks). The dashboard can show hours instead; the stored "
            "numbers are always FTE.",
    "evaluated_for": "every (assignment, calendar month) pair the assignment covers",
    "terms": {
        "period_weight": "ProjectPeriod.weight of the period containing the FIRST DAY of "
                         "the month, for the assignment's project. A month in no period "
                         "uses 1.00 and is reported under V-12.",
        "role_factor": "RoleFactor.role_factor for (project_type, clinical_phase, "
                       "period_name, role_name). clinical_phase is null for 'Others' "
                       "projects. A missing factor is an error under V-23 - the "
                       "calculation would otherwise silently use 1.00.",
        "person_weight": "Assignment.person_weight, UNLESS a PersonPeriodWeight window "
                         "contains the first day of the month, in which case "
                         "weight_override REPLACES it. It does not multiply it.",
        "month_coverage": "the fraction of that calendar month's days the assignment "
                          "actually spans, inclusive of both end dates: "
                          "(min(month_end, assign_end) - max(month_start, assign_start) "
                          "+ 1 days) / days_in_month.",
    },
    "assignment_window": "assign_start_date .. assign_end_date. An empty "
                         "assign_end_date means the project's end_date.",
    "aggregation": [
        "per project per month  - sum over that project's assignments",
        "per person per month   - sum over that person's assignments (this is what the "
        "over/under-allocation thresholds are compared against)",
        "per project per person per role per month - the finest cell the dashboard shows",
    ],
    "thresholds": "over_allocation_fte and under_allocation_fte are ABSOLUTE FTE figures. "
                  "They are NOT scaled by a person's capacity_fte (decision S2-01). "
                  "under_allocation_min_months consecutive months below the floor make a "
                  "run; a single low month is not flagged.",
    "rows_excluded": "an assignment whose project or person does not exist, or which has "
                     "no start date, contributes nothing and is reported instead.",
}

PERIOD_DERIVATION = {
    "when": "Only for a clinical-trial project (NewDrug CT / Biosimilar CT) that has NO "
            "rows on ProjectPeriod at all. Any period the file supplies is used as it "
            "stands - derivation never overrides typed data (REQ-CAL-13).",
    "requires": ["CTA submission", "final DB lock (or interim DB lock in its place)"],
    "if_missing": "V-16 error - the project has no periods and none can be computed.",
    "steps": [
        "Start-up starts the day after 'Protocol (v1)', or one month before "
        "'CTA submission' when there is no protocol date; never before the project start.",
        "Start-up ends at 'First SIV' (or 'FPI'), or four months after it began.",
        "'After Close-out (final)' opens only on an Inspection date AFTER the final DB "
        "lock. Earlier inspections are markers inside the existing periods (V-21).",
        "'Close-out (final)' starts three months before the final DB lock and runs to the "
        "day before period 7, or to the later of the DB lock and the project end.",
        "With an interim DB lock earlier than the final one, the conduct stretch splits: "
        "'Conduct (interim)' up to three months before the interim lock, "
        "'Close-out (interim)' to the interim lock, then 'Conduct (final)'.",
        "Without an interim lock there is one 'Conduct (final)' stretch.",
        "A period squeezed to nothing is omitted, and the remainder is renumbered from 1 "
        "(REQ-CAL-12, decision C-11).",
    ],
    "note": "'Others' projects are never derived. Enter their three periods "
            "(Planning / Develop / Close) and their weights directly.",
}

DATE_HANDLING = {
    "accepted": ["a real Excel date cell", "the text yyyy-mm-dd"],
    "rejected": "any ambiguous format such as 03/04/2026. It is reported, never guessed at.",
    "excel_serial": "Serial day numbers are read against 1899-12-30, which reproduces "
                    "Excel's 1900 leap-year bug. Write 45658 for 2025-01-01.",
    "inclusive": "Every start/end pair in PRAP is INCLUSIVE of both endpoints.",
    "timezone": "All dates are handled as UTC calendar dates. There are no times.",
    "json": "In the JSON interchange format always write dates as the string yyyy-mm-dd.",
}

EDITING_CONTRACT = {
    "provisional_edits": "Every change made in the application is provisional. The file on "
                         "disk is never touched. 'Save' commits changes into the in-memory "
                         "model; 'Leave without change' reverts to the state at load.",
    "export_is_the_only_write": "The application writes nothing until Export. Export "
                                "produces a NEW file named <source>_<yyyy-mm-dd>, so the "
                                "source workbook is never overwritten.",
    "export_blocked_when": [
        "there are unsaved pending changes - press Save or Leave without change first",
        "a newly inserted row has no identifier, or has only its parent key filled in - "
        "such a row would be silently dropped on re-import",
    ],
    "identifier_edits_cascade": "Editing project_id, person_id or assignment_id rewrites "
                                "every row that references it (REQ-IMP-10).",
    "deletes_never_cascade": "Deleting a row is refused while anything still references it, "
                             "and the refusal names what does (V-17).",
    "derived_columns_are_not_edited": "total_period_months, Milestone.project_name and "
                                      "Assignment.person_name are recomputed on every load. "
                                      "A value in the file that disagrees is reported (V-13) "
                                      "and then overwritten. In the template these cells are "
                                      "locked and carry a comment.",
    "drafts": "A row inserted in the application is a draft until Save. Drafts are exempt "
              "from validation, because a half-typed row is incomplete rather than wrong.",
}

INTERCHANGE = {
    "purpose": "A plain-text form of the source workbook, so a program or an AI agent that "
               "cannot write .xlsx can still produce a file the application loads, and can "
               "read one without a spreadsheet library.",
    "media_type": "application/json",
    "extension": ".prap.json",
    "shape": {
        "prap_format": "prap-source-data  (constant - identifies the file)",
        "format_version": 1,
        "schema_version": "must equal the contract's schema_version",
        "sheets": "an object: sheet name -> array of row objects, keyed by column name. "
                  "Column order does not matter; unknown columns are rejected so a typo "
                  "cannot be mistaken for data.",
    },
    "example": {
        "prap_format": "prap-source-data",
        "format_version": 1,
        "schema_version": 5,
        "sheets": {
            "Project": [{"project_id": "PRJ-001", "project_name": "ONV-101 Phase 1",
                         "project_type": "NewDrug CT", "clinical_phase": "Phase 1",
                         "start_date": "2026-01-01", "end_date": "2028-06-30",
                         "status": "Active"}],
            "Assignment": [{"assignment_id": "ASG-001", "person_id": "PSN-001",
                            "project_id": "PRJ-001", "role_name": "Project oversight",
                            "assign_start_date": "2026-01-01", "person_weight": 0.2}],
        },
    },
    "round_trip": "xlsx -> json -> xlsx reproduces every cell. Proved on both worked "
                  "examples by tools/test_interop.py.",
    "converter": "python tools/prap_io.py to-json <file.xlsx> -o <file.prap.json> and "
                 "python tools/prap_io.py to-xlsx <file.prap.json> -o <file.xlsx>",
}

RECIPES = [
    {
        "task": "Draft a resource assignment for a project from its milestones",
        "when": "The user gives you a project - type, phase, dates, milestones - and asks "
                "who should work on it and how much.",
        "steps": [
            "Read the current workbook: python tools/prap_io.py to-json <file.xlsx> -o draft.prap.json",
            "Add the Project row. For a clinical trial, project_type, clinical_phase, "
            "project_category, start_date and end_date are all needed, or the project "
            "cannot be weighted (V-19, V-04).",
            "Add the Milestone rows. 'CTA submission' and a DB lock are what the period "
            "derivation hangs on (V-16); the other eight names are optional markers.",
            "Leave ProjectPeriod EMPTY for a clinical trial and let the application derive "
            "the seven periods. For an 'Others' project you must write Planning / Develop / "
            "Close yourself, with no gap and no overlap (V-06, V-12).",
            "Choose people from the Person sheet. For each, add ONE Assignment row per "
            "(person, project, role). role_name must exist in RoleFactor for that "
            "project_type (V-03); use nextKey-style ids - the smallest unused ASG-nnn.",
            "Set person_weight to the fraction of that person that goes to this project. "
            "Where the fraction changes for a stretch of months, add a PersonPeriodWeight "
            "window instead of a second assignment - the override REPLACES the weight for "
            "those months, and windows on one assignment must not overlap (V-06, V-24).",
            "Check the draft: python tools/prap_io.py validate draft.prap.json",
            "Look at the load it produces: python tools/prap_io.py calculate draft.prap.json "
            "--by person. Keep every person-month between under_allocation_fte and "
            "over_allocation_fte; those are absolute FTE, not shares of capacity.",
            "Convert back and hand the workbook over: python tools/prap_io.py to-xlsx "
            "draft.prap.json -o <file>_draft.xlsx. Say which rows you added and why.",
        ],
        "do_not": [
            "do not invent role names, period names, systems or statuses - every one of "
            "them comes from the Lists sheet, and a value outside its list is reported (V-11)",
            "do not write the derived columns; they are recomputed and your value is discarded",
            "do not overwrite the user's source workbook - always deliver a new file",
        ],
    },
    {
        "task": "Level an over-allocated person",
        "when": "The dashboard flags a person above over_allocation_fte for some months.",
        "steps": [
            "python tools/prap_io.py calculate <file> --by person --flags shows every month "
            "over the ceiling and which projects make it up.",
            "Reduce person_weight on the least critical assignment, or add a "
            "PersonPeriodWeight window covering only the offending months.",
            "Re-run calculate. Moving load off one person usually pushes it onto another, "
            "so check the whole person list, not only the one you changed.",
        ],
        "do_not": ["do not shorten an assignment to hide load - that changes the plan, not "
                   "the demand. Say what you changed."],
    },
    {
        "task": "Answer a question about the data without changing it",
        "when": "The user asks who is on what, when a project peaks, where the gaps are.",
        "steps": [
            "python tools/prap_io.py to-json <file.xlsx> -o /tmp/read.prap.json, then read "
            "the JSON directly. It is the whole workbook as plain text.",
            "python tools/prap_io.py calculate <file> --by project|person [--from YYYY-MM "
            "--to YYYY-MM] gives the same monthly figures the dashboard draws.",
            "Quote figures as FTE to two decimals, and say which months you looked at.",
        ],
        "do_not": ["do not re-derive the formula yourself; use calculate, so your answer and "
                   "the dashboard cannot disagree"],
    },
    {
        "task": "Continue a plan somebody started inside the application",
        "when": "The user hands you a workbook they built by typing into PRAP rather than "
                "one exported from a system of record.",
        "steps": [
            "Check the reference tables FIRST: python tools/prap_io.py to-json <file.xlsx> "
            "-o plan.prap.json, then look at PeriodFTEStandard and RoleFactor. A plan "
            "started blank seeds every one of them at 1.00 with a note saying so.",
            "If they are still 1.00, say so before quoting any figure. The load formula "
            "then reduces to person_weight x month_coverage, which is a real number but "
            "not a resourced estimate.",
            "Offer to fill them in from whatever the user can tell you about their own "
            "standards, one (type, phase, period) at a time. Do not invent them.",
            "Everything else is an ordinary workbook - the same schema, the same rules.",
        ],
        "do_not": ["do not treat a seeded 1.00 as a measured value, and do not quietly "
                   "replace it with a number of your own"],
    },
    {
        "task": "Prepare a workbook for someone to open in the application",
        "when": "You have produced or edited data and the user wants to look at it.",
        "steps": [
            "Validate first. A file with errors still loads, but the user meets a findings "
            "report before they see any numbers.",
            "Deliver the .xlsx. Tell the user to open app/PRAP.html in a browser and drop "
            "the file on it - nothing is uploaded, and the application needs no network.",
            "Say what you changed, in rows: 'added 4 Assignment rows, 1 PersonPeriodWeight "
            "window' beats 'updated the plan'.",
        ],
        "do_not": ["do not send a .prap.json to a person expecting a spreadsheet - convert "
                   "it to .xlsx first, unless they asked for the JSON"],
    },
]

VERIFICATION = {
    "validate a workbook (no browser)": "python tools/verify_source_workbook.py <file.xlsx>",
    "validate a draft in either format": "python tools/prap_io.py validate <file>",
    "documents against artifacts": "python tools/check_consistency.py",
    "calculation matches the reference implementation":
        "python tools/test_app.py  (drives the real application in a browser)",
    "JSON round-trip and cross-tool agreement": "python tools/test_interop.py",
    "row insert/delete/identity": "python tools/test_rows.py",
    "type-ahead value lists": "python tools/test_valuelist.py",
    "charts": "python tools/test_charts.py",
    "text contrast": "python tools/test_contrast.py",
}
