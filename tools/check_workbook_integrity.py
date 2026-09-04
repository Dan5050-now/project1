#!/usr/bin/env python3
"""Fail the build on workbook defects that make Excel show a repair prompt.

Written after TEA-SPEC-001 v1.1.0 shipped with an empty <dataValidations count="0"/>
element on the Interpretations sheet, which is schema-invalid and made Excel offer to
repair the file on open. openpyxl writes such a workbook without complaint and reads it
back without complaint, so nothing in the normal build catches it — only Excel does.

Checks every controlled workbook for:
  - XML parts that do not parse
  - empty <dataValidations> containers, and dataValidation entries with no sqref
  - overlapping merged ranges
  - cell values beyond Excel's 32,767-character limit
  - row heights beyond 409 points, column widths beyond 255 characters
  - control characters that are illegal in XML

Usage:  python3 tools/check_workbook_integrity.py [workbook.xlsx ...]
        with no arguments, checks every .xlsx under docs/
"""
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent

MAX_CELL_CHARS = 32767
MAX_ROW_HEIGHT = 409
MAX_COL_WIDTH = 255
ILLEGAL_XML = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def check(path):
    problems = []
    z = zipfile.ZipFile(path)

    for name in z.namelist():
        if not name.endswith((".xml", ".rels")):
            continue
        raw = z.read(name)
        try:
            ET.fromstring(raw)
        except ET.ParseError as e:
            problems.append(f"{name}: XML does not parse — {e}")
            continue
        if not name.startswith("xl/worksheets/sheet"):
            continue
        text = raw.decode("utf-8", "replace")
        if re.search(r'<dataValidations[^>]*count="0"', text):
            problems.append(
                f"{name}: empty <dataValidations count=\"0\"/> — Excel will offer to repair "
                f"the file. A DataValidation was attached to the sheet but no range was added.")
        for m in re.finditer(r"<dataValidation\b[^>]*>", text):
            sq = re.search(r'sqref="([^"]*)"', m.group(0))
            if not sq or not sq.group(1).strip():
                problems.append(f"{name}: <dataValidation> with no sqref")

    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries

    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        boxes = [(str(r),) + range_boundaries(str(r)) for r in ws.merged_cells.ranges]
        for i in range(len(boxes)):
            for j in range(i + 1, len(boxes)):
                _, a1, b1, a2, b2 = boxes[i]
                _, c1, d1, c2, d2 = boxes[j]
                if a1 <= c2 and c1 <= a2 and b1 <= d2 and d1 <= b2:
                    problems.append(
                        f"{sheet}: merged ranges overlap — {boxes[i][0]} and {boxes[j][0]}")

        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                if len(v) > MAX_CELL_CHARS:
                    problems.append(
                        f"{sheet}!{cell.coordinate}: {len(v)} chars exceeds the "
                        f"{MAX_CELL_CHARS} limit")
                if ILLEGAL_XML.search(v):
                    problems.append(
                        f"{sheet}!{cell.coordinate}: illegal XML control character")

        for idx, rd in ws.row_dimensions.items():
            if rd.height and rd.height > MAX_ROW_HEIGHT:
                problems.append(f"{sheet}: row {idx} height {rd.height} exceeds {MAX_ROW_HEIGHT}")
        for key, cd in ws.column_dimensions.items():
            if cd.width and cd.width > MAX_COL_WIDTH:
                problems.append(f"{sheet}: column {key} width {cd.width} exceeds {MAX_COL_WIDTH}")

    return problems


def main(argv):
    paths = [Path(a) for a in argv[1:]] or sorted((REPO / "docs").rglob("*.xlsx"))
    if not paths:
        print("no workbooks found")
        return 1

    failed = False
    for p in paths:
        if not p.exists():
            print(f"FAIL {p}: not found")
            failed = True
            continue
        problems = check(p)
        if problems:
            failed = True
            print(f"FAIL {p.name} — {len(problems)} problem(s):")
            for q in problems:
                print(f"  {q}")
        else:
            print(f"ok   {p.name}")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
