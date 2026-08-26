"""Generate the Project Resource Assignment Program (PRAP) development plan workbook.

The workbook is a generated artifact: edit this script, re-run it, and commit both
so the plan stays diffable under version control.

    python tools/build_dev_plan.py

Output: docs/PRAP_Development_Plan_v0.2.xlsx

v0.1 is rebuildable from commit 45412d1.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DOC_VERSION = "2.29"
DOC_STATUS = ("Baseline v2.0 + Step 4 progress. Application v1.25 - Gate 4 refinements rounds 1-25, "
              "plus SCHEMA 6 (the work scope, the biosimilar split), the shared-role division "
              "and the delivered default assumptions.")
DOC_DATE = "2026-07-31"
OUT = Path(__file__).resolve().parents[1] / "docs" / f"PRAP_Development_Plan_v{DOC_VERSION}.xlsx"

FONT = "Arial"
NAVY = "1F3864"
BLUE_HDR = "2F5597"
BAND = "F2F5FB"
YELLOW = "FFFF00"
GREEN = "C6E0B4"
ORANGE = "FCE4D6"
GREY = "808080"

TITLE_F = Font(name=FONT, size=16, bold=True, color=NAVY)
H1_F = Font(name=FONT, size=12, bold=True, color=NAVY)
HDR_F = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_F = Font(name=FONT, size=10)
BOLD_F = Font(name=FONT, size=10, bold=True)
NOTE_F = Font(name=FONT, size=9, italic=True, color=GREY)
MONO_F = Font(name="Consolas", size=10)

HDR_FILL = PatternFill("solid", fgColor=BLUE_HDR)
BAND_FILL = PatternFill("solid", fgColor=BAND)
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
NEW_FILL = PatternFill("solid", fgColor=GREEN)
CHG_FILL = PatternFill("solid", fgColor=ORANGE)

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(vertical="top", wrap_text=True)
WRAP_C = Alignment(vertical="top", wrap_text=True, horizontal="center")

# Rows whose first cell starts with these markers get a highlight fill.
MARK_NEW = "[NEW]"
MARK_CHG = "[CHANGED]"


def sheet(wb, name, title, subtitle=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_F
    row = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = NOTE_F
        row = 3
    ws.freeze_panes = f"A{row + 2}"
    return ws, row + 1


def table(ws, start_row, headers, rows, widths, wrap_cols=(), mark_col=None):
    """Write a banded header+body table. Returns the row after the table.

    mark_col: 1-based column whose value may carry a [NEW]/[CHANGED] marker,
    which is stripped and turned into a row highlight.
    """
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = HDR_F
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = WRAP_C
    ws.row_dimensions[start_row].height = 28

    for r, data in enumerate(rows, start=start_row + 1):
        data = list(data)
        fill = None
        if mark_col is not None and isinstance(data[mark_col - 1], str):
            v = data[mark_col - 1]
            if v.startswith(MARK_NEW):
                fill = NEW_FILL
                data[mark_col - 1] = v[len(MARK_NEW):].strip()
            elif v.startswith(MARK_CHG):
                fill = CHG_FILL
                data[mark_col - 1] = v[len(MARK_CHG):].strip()
        for i, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = BODY_F
            c.border = BOX
            c.alignment = WRAP if i in wrap_cols else Alignment(vertical="top")
            if fill is not None:
                c.fill = fill
            elif (r - start_row) % 2 == 0:
                c.fill = BAND_FILL

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return start_row + len(rows) + 2


def section(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = H1_F
    return row + 1


def note(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = NOTE_F
    return row + 1


def lines(ws, row, texts, mono=False):
    for t in texts:
        c = ws.cell(row=row, column=1, value=t)
        c.font = MONO_F if (mono and t.startswith(" ")) else BODY_F
        row += 1
    return row


def legend(ws, row):
    ws.cell(row=row, column=1, value="Legend").font = BOLD_F
    row += 1
    for fill, txt in (
        (NEW_FILL, "Green = new in v0.2, added from your review answers."),
        (CHG_FILL, "Orange = changed in v0.2 as a consequence of your review."),
        (INPUT_FILL, "Yellow = still needs your input."),
    ):
        c = ws.cell(row=row, column=1, value=txt)
        c.font = BODY_F
        c.fill = fill
        row += 1
    return row + 1


wb = Workbook()
wb.remove(wb.active)

# ---- 00 Cover -------------------------------------------------------------
ws = wb.create_sheet("00_Cover")
ws.sheet_view.showGridLines = False
ws["A1"] = "Project Resource Assignment Program (PRAP)"
ws["A1"].font = Font(name=FONT, size=20, bold=True, color=NAVY)
ws["A2"] = "Application Development Plan"
ws["A2"].font = Font(name=FONT, size=14, color=NAVY)

cover = [
    ("Document ID", "PRAP-PLAN-001"),
    ("Document type", "Development plan (Step 1 deliverable) - change against the v1.3 baseline"),
    ("Version", f"v{DOC_VERSION}"),
    ("Status", DOC_STATUS),
    ("Issue date", DOC_DATE),
    ("Author", "Claude Code"),
    ("Reviewer", "Requester - four review rounds: v0.11, v0.2, v0.3, v0.4 reviewed"),
    ("Baseline", "v1.3, approved by Dan 2026-08-01"),
    ("Repository", "Dan5050-now/project1"),
    ("Branch", "claude/project-resource-assignment-app-1vjdzh"),
    ("Supersedes", "v1.4 (draft) and, once approved, v1.3. v1.3 remains the baseline until then"),
]
r = 4
for k, v in cover:
    ws.cell(row=r, column=1, value=k).font = BOLD_F
    c = ws.cell(row=r, column=2, value=v)
    c.font = BODY_F
    c.alignment = WRAP
    r += 1
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 84

r += 1
r = section(ws, r, "Status of this issue")
r = lines(ws, r, [
    "v1.0 was approved by Dan on 2026-08-01, closing Gate 1. It remains the approved baseline.",
    "",
    "This issue, v1.1, carries a change request raised after that approval: two additions marked on sheet 11",
    "and a set of edits made directly to the derivation table on sheet 05. Because v1.0 is a baseline, these",
    "cannot simply be absorbed - they are presented as a numbered change against it, and v1.1 needs its own",
    "approval before it supersedes v1.0. The approval block on sheet 12 now holds both.",
    "",
    "WHAT CHANGES",
    "  1. 'Inspection' becomes a standard milestone, and unlike the others it may occur SEVERAL TIMES in one",
    "     project. A milestone name is therefore no longer unique within a project.",
    "  2. A seventh period, 'After Close-out (final)', spans the inspection activity. The clinical-trial",
    "     period set grows from five names to six.",
    "  3. Three boundaries now prefer a milestone where one is recorded: Before-Start-up ends at",
    "     'Protocol (v1)', Start-up begins the day after it, and Start-up ends at 'First SIV'.",
    "  4. 'FPI' returns as a permitted milestone - as the fallback for 'First SIV'. It had been dropped from",
    "     the standard list at Q-15, so this reverses that.",
    "",
    "The revised derivation was tested against seven timelines and is contiguous in all of them. One case",
    "needed a rule that was not in the mark-up: an inspection dated BEFORE the final DB lock. Taken",
    "literally it makes the seventh period start before the sixth. Inspections on or before the final DB",
    "lock are therefore treated as markers inside the existing periods, and only later inspection activity",
    "opens period 7. Raised as R-03 and CONFIRMED by the reviewer at the v1.1 review.",
    "",
    "v1.3 was approved by Dan on 2026-08-01 and remains the baseline. This issue carries five changes against",
    "it - R-05 from the Step 3 review, R-06 to R-08 from the specification v0.3 review, and R-09 to R-13",
    "from the component-list reviews. All were approved by Dan on 2026-08-02, and this issue is now THE",
    "BASELINE, superseding v1.3. Steps 1 to 3 are closed; Step 4, code generation, is authorised.",
    "",
    "v1.3 carries change R-04: every sheet of the source workbook now holds at least one free-text note",
    "column. Four sheets had none - Milestone, ProjectPeriod, PeriodWeightStandard and Lists - and each gains",
    "note_1. The other six already had one: Project and Person carry note_1..note_5, Assignment",
    "note_1..note_3, RoleFactor role_note, PersonPeriodWeight reason, and Config note.",
    "",
    "Because columns change, the source schema version steps from 1 to 2. Nothing else moves: notes are",
    "carried through import and export unchanged and are never read by the calculation, so re-running the",
    "dummy dataset gives identical figures to v1.2.",
    "",
    "WHAT THIS PLAN NOW FIXES",
    "  65 requirements (sheet 03) as the contract for Steps 2-5",
    "  21 validation rules (sheet 04) applied on import and on every on-screen edit",
    "  11 engineering decisions, C-01 to C-11 (sheet 05), all confirmed",
    "  a source schema of 10 sheets at version 2 (sheet 04)",
    "  the five-step build with a review gate at the end of each (sheet 08)",
    "",
    "Nothing in this document is open. Six review rounds produced 28 questions and 4 change requests; all 32",
    "are answered and applied, and the review log on sheet 12 carries no open item.",
])
r += 1
r = section(ws, r, "The plan in one page")
r = lines(ws, r, [
    "  Output        One HTML file, opened by double-click on Windows, offline, no install. One Excel",
    "                workbook alongside it holds the data and remains the archive of record.",
    "",
    "  Calculation   load = project period weight x role factor x person weight x month coverage,",
    "                in FTE where 1.00 FTE = 160 hours per month.",
    "                Over-allocated above 1.50 FTE in a month; under-allocated below 0.80 FTE for three",
    "                or more consecutive months, reported as a run.",
    "",
    "  Periods       Clinical Trial: five period names derived from milestone dates and month offsets,",
    "                with the conduct phase split into 'Conduct (interim)' and 'Conduct (final)' where an",
    "                interim DB lock divides it. Weights come",
    "                from the trial's clinical phase.",
    "                Others: three periods, dates and weights both entered by hand.",
    "",
    "  Dashboard     Overall (monthly simulation per project and per person, tables and graphs),",
    "                Source data (project), Source data (person). Every field editable, with identifier",
    "                edits cascading to referencing rows.",
    "",
    "  Build         Five steps, each ending at a review gate. This document closes Step 1.",
])
r += 1
r = section(ws, r, "What approving this document means")
r = lines(ws, r, [
    "Gate 1 approval on sheet 12 baselines the 63 requirements on sheet 03 as the contract for Steps 2-5,",
    "and confirms the six engineering decisions C-06 to C-11 on sheet 05 that were proposed but never",
    "explicitly answered across the review rounds. Those six are listed there with their rationale; if any",
    "is wrong, say so at approval rather than after.",
])

r += 1
r = section(ws, r, "How to read this workbook")
guide = [
    ("00_Cover", "Document control, what changed, reading guide."),
    ("01_Version_History", "Change log for this plan."),
    ("02_Scope", "Background, objectives, in scope / out of scope."),
    ("03_Requirements", "Numbered requirement register - the contract for Steps 2-5."),
    ("04_Data_Model", "Every sheet and column of the single source workbook, with rules."),
    ("05_Resource_Logic", "The calculation that turns assignments and weights into monthly FTE."),
    ("06_Dashboard_Design", "The three dashboard tabs: tables, graphs, filters, editing."),
    ("07_Architecture", "Single-file HTML + Excel-on-disk design and the file-access decision."),
    ("08_WBS_Schedule", "The five steps broken into tasks, deliverables and review gates."),
    ("09_Version_Control", "How plan, specification and output files are versioned together."),
    ("10_Risks", "Risks and assumptions with mitigations."),
    ("11_Open_Questions", "All 28 questions and their answers. Nothing open."),
    ("12_Review_Log", "Every change made across the four review rounds, and why. Gate 1 approval block."),
]
r = table(ws, r, ["Sheet", "Contents"], guide, [26, 84], wrap_cols=(2,))
r = legend(ws, r)

# ---- 01 Version history ---------------------------------------------------
ws, r = sheet(wb, "01_Version_History", "Version history",
              "One row per issued version. Numbering rules are on sheet 09_Version_Control.")
rows = [
    ["0.1", "2026-07-25", "Claude Code", "-",
     "First draft issued for review. Built from the requirements supplied by the requester, incorporating two "
     "confirmed decisions: browser file-picker + SheetJS for Excel access (option a), and the five-step plan.",
     "Superseded"],
    ["0.11", "2026-07-31", "Requester", "Requester",
     "Reviewer mark-up returned as PRAP_Development_Plan_v0.11_reviewed.xlsx. Answered all 12 open questions; "
     "added one Import/Export requirement; added clinical phase, four system-responsibility columns, three "
     "system-name columns, and denormalised name columns to the data model.",
     "Review input"],
    ["0.2", "2026-07-31", "Claude Code", "Requester",
     "Reviewer answers propagated across all sheets: single source workbook; calculation reduced to "
     "project weight x role factor x person weight x coverage; FTE capacity with 1.50 over-allocation and "
     "0.80/3-month under-allocation thresholds; in-application editing added; real role, period and milestone "
     "names adopted; performance target reduced to the stated data volume. Eight follow-up questions raised.",
     "Superseded"],
    ["0.21", DOC_DATE, "Requester", "Requester",
     "Second reviewer mark-up returned as PRAP_Development_Plan_v0.2_reviewed.xlsx, answering Q-13 to Q-20.",
     "Review input"],
    ["0.3", DOC_DATE, "Claude Code", "Requester",
     "Q-13 and Q-17 closed, clearing both blockers. Milestone list replaced (First SIV in, FPI out, DB lock "
     "cut-off and DB lock split interim/final). Period sets made type-specific: Clinical Trial keeps its four, "
     "'Others' gets Planning / Develop / Close. Period boundaries now computed from milestone dates plus month "
     "offsets. outsourcing_type fixed to three values; system_prepared_by dropped. All fields made editable with "
     "cascading referential integrity. Three coverage gaps in the supplied period mapping identified and raised "
     "as Q-21 to Q-23.",
     "Superseded"],
    ["0.31", DOC_DATE, "Requester", "Requester",
     "Third reviewer mark-up returned as PRAP_Development_Plan_v0.3_reviewed.xlsx, answering Q-21 to Q-26.",
     "Review input"],
    ["0.4", DOC_DATE, "Claude Code", "Requester",
     "Period model settled. Close-out starts 3 months before DB lock; a trial with an interim lock carries "
     "two close-out periods with Conduct resuming between them, so 'Conduct' may occur twice and "
     "ProjectPeriod gains period_seq. Clinical-trial period set grows to five names. PeriodWeightStandard "
     "re-keyed on clinical_phase. 'Protocol (v1)' added as the eighth milestone. Derivation verified "
     "contiguous against five timelines including four degenerate ones. Two questions remain.",
     "Superseded"],
    ["0.41", DOC_DATE, "Requester", "Requester",
     "Fourth reviewer mark-up returned as PRAP_Development_Plan_v0.4_reviewed.xlsx, answering Q-27 and Q-28.",
     "Review input"],
    ["1.0", DOC_DATE, "Claude Code", "Pending",
     "FINAL Step 1 issue. Both remaining answers confirmed the v0.4 design. A single close-out keeps the name "
     "'Close-out (final)'. 'Others' projects need no distinguishing and take manual weights as well as manual "
     "dates, so PeriodWeightStandard becomes a clinical-trial reference table only. All 28 review questions "
     "closed; no open items. Issued for Gate 1 approval.",
     "APPROVED 2026-08-01 by Dan - baseline"],
    ["1.01", DOC_DATE, "Requester", "Requester",
     "Post-approval change request returned as PRAP_Development_Plan_v1.0_reviewed.xlsx: add 'Inspection' as a "
     "standard milestone, add a seventh period covering inspection activity, and prefer 'Protocol (v1)' and "
     "'First SIV' as period boundaries where recorded.",
     "Change request"],
    ["1.1", DOC_DATE, "Claude Code", "Pending",
     "Change against the v1.0 baseline. 'Inspection' added as a standard milestone that may repeat within a "
     "project; 'FPI' restored as the fallback for 'First SIV'; seventh period 'After Close-out (final)' added, "
     "taking the clinical period set to six names; Before-Start-up, Start-up start and Start-up end now prefer "
     "a recorded milestone. Derivation re-verified contiguous against seven timelines. Needs its own approval.",
     "Superseded by v1.2"],
    ["1.11", DOC_DATE, "Requester", "Requester",
     "Reviewed v1.1 and confirmed R-03 - inspections dated on or before the final DB lock stay markers and do "
     "not open the seventh period. No other change requested.",
     "Review input"],
    ["1.2", DOC_DATE, "Claude Code", "Pending",
     "R-03 confirmed and closed. No content change beyond recording it: the derivation, requirements and "
     "validation rules are as issued at v1.1. Every question and change request raised across five review "
     "rounds is now settled.",
     "APPROVED 2026-08-01 by Dan - superseded by v1.3 once approved"],
    ["1.3", DOC_DATE, "Claude Code", "Pending",
     "Change against the v1.2 baseline (R-04): a free-text note column added to every source sheet that "
     "lacked one - Milestone, ProjectPeriod, PeriodWeightStandard and Lists each gain note_1. Source schema "
     "version steps from 1 to 2. No calculation, validation or UI behaviour changes. Issued FINAL: this is "
     "the last planned issue of the development plan.",
     "APPROVED 2026-08-01 by Dan - FINAL"],
    ["1.4", DOC_DATE, "Claude Code", "Pending",
     "Change against the v1.3 baseline (R-05), from the Step 3 prototype review: project_type splits into "
     "'NewDrug CT' and 'Biosimilar CT'. REQ-PRJ-01 and REQ-PRJ-02 reworded; RoleFactor and "
     "PeriodWeightStandard keyed on the type; source schema version steps 2 to 3.",
     "Superseded by v1.5"],
    ["1.5", DOC_DATE, "Claude Code", "Pending",
     "Changes R-06, R-07 and R-08 against the v1.3 baseline, from the specification v0.3 review. "
     "REQ-NFR-03 re-baselined to 100 projects and 1,000 people and raised to Must; both allocation "
     "thresholds confirmed ABSOLUTE with the under-allocation floor moved 0.80 to 0.60; repeated period "
     "names must be distinguishable on screen. REQ-DSH-09, REQ-DSH-10 and V-22 added.",
     "Superseded by v1.6"],
    [f"{MARK_NEW}2.29", "2026-08-25", "Claude Code", "Pending",
     "R-15, REQ-CAL-15: BOTH ASSIGNMENT DATES ARE OPTIONAL, and a blank one means the "
     "project's own. Most people are on a project for the whole of it, and requiring two "
     "dates that simply repeat the project's asks somebody to copy the same pair onto "
     "every row and then keep them in step when the project moves. The end date already "
     "worked this way; the start did not, and a blank one was the worst of the three "
     "possible behaviours - the assignment contributed NOTHING, the person looked "
     "unassigned, and no finding said why. The window is now computed by one function "
     "shared by the calculation and by the role-sharing pre-pass, so the months a person "
     "is counted IN cannot differ from the months they are counted AMONG; an undated "
     "person is a sharer like any other. Roughly a third of the dummy fixture now leaves "
     "both dates blank, which is what a real file looks like and puts the inherited "
     "window under the four-way comparison rather than in one unit test.",
     "Issued for review"],
    ["2.28", "2026-08-22", "Claude Code", "Pending",
     "TWO CHANGES, BOTH REQUESTED, AND THE FIRST CHANGES EVERY FIGURE A SHARED ROLE EVER "
     "PRODUCED. (1) R-13, REQ-CAL-14: where several people hold the same role on the same "
     "project in the same month, the role factor is DIVIDED BETWEEN THEM. The factor states "
     "what the ROLE costs the project in that period, not what each holder costs, so until "
     "now a trial staffed by two data managers cost twice a trial staffed by one - the same "
     "work, priced by how many people were named against it. Counted per month, so a sharer "
     "leaving restores the others without an edit; by distinct people, so one person on two "
     "rows is one person; and each person's own weight still applies to their share. Kept "
     "as the setting split_shared_role_fte, default 1, so a figure produced before this "
     "rule can still be reproduced and compared. (2) R-14: THE DELIVERED DEFAULTS. "
     "PeriodWeightStandard and RoleFactor arrived empty in the template and were filled "
     "with a placeholder 1.00 by a blank start - and 1.00 everywhere is not a starting "
     "point, it is the absence of one: at 1.00 the period weight and the role factor cancel "
     "out of the arithmetic entirely, so the application produced figures that looked like "
     "an answer and were not. Both now arrive filled in - 84 period weights and 429 role "
     "factors - defined once and read by the template, the blank start and the dummy "
     "datasets alike, so the workbook and the application hold the same assumptions by "
     "construction. check_consistency.py fails if they drift, and fails too if the figures "
     "ever revert to one flat value. A defect was found while testing: prap_io read Config "
     "with `or default`, so a setting deliberately set to 0 silently came back on.",
     "Issued for review"],
    ["2.27", "2026-08-20", "Claude Code", "Pending",
     "SCHEMA 6 - THE WORK SCOPE, AND THE BIOSIMILAR SPLIT. Requested 2026-08-20, and the "
     "first change to the data model since the plan was baselined. (1) work_scope_type joins "
     "PeriodWeightStandard and RoleFactor beside clinical_phase, and the keys become "
     "(project_type, clinical_phase, work_scope_type, period_name) and that plus role_name. "
     "How much of the work is kept in-house changes how much of it lands on this team, and "
     "until now the model had nowhere to say so. (2) A ROW WITH AN EMPTY work_scope_type "
     "APPLIES TO EVERY SCOPE, and a project falls back to it when there is no row for its own "
     "scope. That is a judgment made here rather than requested: without it RoleFactor is "
     "1,269 hand-entered rows instead of 429, two thirds of them repeating their neighbour, "
     "and every schema 5 file would stop calculating on the day it was opened. With it, a "
     "schema 5 file's rows ARE the every-scope rows and keep working. (3) Project gains "
     "work_scope_type, because the key selects a row FOR A PROJECT and nothing could choose "
     "between the three scope rows without it - implied by the request rather than stated in "
     "it. outsourcing_type stays, descriptive, and V-25 reports the two contradicting each "
     "other. (4) 'Biosimilar CT' becomes 'Biosimilar CT (Healthy)' and 'Biosimilar CT "
     "(Patient)'. A project type is now recognised as a clinical trial by the START of its "
     "name, so the next subdivision is a value-list change rather than a code change. V-26 "
     "reports the retired value with the two it became, because only the reader knows which "
     "of them a given trial was. Template v1.8, dummies v1.10 and v1.2.",
     "Issued for review"],
    ["2.26", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 25, application v1.24. (1) EVERY FILTER NOW TAKES SEVERAL VALUES AT ONCE. "
     "Each is a drop-down of tick boxes built on <details>, which gives open, close, keyboard "
     "operation and focus for nothing; a plain <select multiple> would have cost seven tall "
     "boxes across the bar and asked the reader to know that ctrl-click means 'and also'. "
     "Nothing ticked means All, which is what the closed control says, so the bar reads as it "
     "did with single selects; one ticked reads as that value, more than one as a count with "
     "the full list on hover. Values within one filter are an OR - ticking a second project "
     "type WIDENS the view - while the filters remain an AND with each other, which is the "
     "only reading that makes 'NewDrug CT and Others, in-house' mean anything. Each has its "
     "own Clear; Reset still empties them all; and a value that has gone from the file is "
     "dropped from the selection, since the page must not stay filtered by something the "
     "reader can no longer see. (2) AUTO DERIVATION in the Periods section builds a project's "
     "periods from its milestones by the rule on sheet 05, and (3) BLANK LIST in the "
     "Milestones section lays out the ten standard milestone names with their dates empty, so "
     "only the dates have to be typed. Both produce ORDINARY ROWS - provisional until Save, "
     "editable and deletable afterwards, subject to every rule - because the point is to save "
     "the typing, not to take the decision away. Auto derivation reads the milestones from the "
     "raw sheet so a set just typed and not yet saved still counts, which is the ordinary case "
     "when the two buttons are used one after the other; it asks before replacing an existing "
     "set; and it refuses what it cannot do, naming the reason: an 'Others' project (the rule "
     "hangs on CTA submission and the DB locks) or a trial missing either of those (V-16). "
     "Blank list does not repeat a name already listed, since a second 'CTA submission' is the "
     "duplicate V-20 exists to catch. A LATENT CRASH was found while testing them and fixed: "
     "clicking a row that had not been saved yet made its identifier the selection, and the "
     "detail panels below need a RECORD - dates, a name, a calculation - which a draft has "
     "none of. projDetail threw, taking the whole re-render with it, so the click appeared to "
     "do nothing at all. Both detail panels now fall back to the scratch panels built for "
     "exactly that state. tools/test_generate.py checks the derivation against the rule term "
     "by term on a project built to exercise every branch. No requirement, rule, calculation "
     "or schema version changed.",
     "Issued for review"],
    ["2.25", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 24, application v1.23. Three changes to the filter bar and to what the "
     "unsaved-change counter can tell you. (1) OUTSOURCING TYPE joins the filter conditions, "
     "driving the same machinery as the other six - the charts, both Overall tables and the "
     "source-data tabs - and Reset puts it back with the rest. (2) THE HORIZON NOW FOLLOWS THE "
     "FILTERS. Narrowing to one project type left the window the whole portfolio needed, so a "
     "two-year span went mostly empty and the reader was looking at a chart of nothing with no "
     "way to tell whether that was the answer or the view. Changing a filter now pulls From and "
     "To in to the months the surviving rows actually reach. Two things are deliberately left "
     "alone: a combination that matches NOTHING keeps the window where it was, because jumping "
     "to an arbitrary span would hide the reason the screen is empty; and typing in From or To "
     "is the user moving the window themselves, so the re-fit runs only when a filter DROPDOWN "
     "moves. (3) A SHOW DETAILS button beside the counter opens every change waiting to be "
     "saved, as a table: the time it happened, which tab, which section, which row, which item, "
     "what it was before and what it is now. Newest first, because the question it answers is "
     "almost always 'what did I just do' - the edit you are unsure about is the one you made a "
     "moment ago - and the first row is tinted for the same reason. The dialog closes, goes "
     "full screen for a log too wide or too long for the ordinary box, and scrolls with the "
     "same visible bar and edge shading as every other scroll region. The section column shows "
     "the name the panel goes by on screen rather than the sheet name, because the point of the "
     "log is being able to walk back to the thing you changed; the sheet is still what is "
     "recorded underneath. Every pending entry now carries the time it was made, which nothing "
     "did before. tools/test_filters.py covers all three. No requirement, rule, calculation or "
     "schema version changed.",
     "Superseded by v2.26"],
    ["2.24", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 23, application v1.22. Four values that were typed by hand while the rows "
     "underneath already said what they should be. (1) A project's WINDOW is now recalculated "
     "on Save: start_date becomes the earliest of its milestone dates and end_date the latest, "
     "and total_period_months follows without anyone touching it. (2) planned_member_count "
     "becomes the number of DISTINCT people assigned to the project - people, not assignment "
     "rows, since one person may hold several. Both are done at SAVE rather than at load, "
     "deliberately: doing them at load would rewrite a workbook produced elsewhere the moment "
     "somebody opened it, and the delivered examples set a project start before its first "
     "milestone quite legitimately. At Save it is a consequence of an edit just made, and the "
     "banner names every value it changed. Two guards keep the change from destroying "
     "information: a project with NO milestone dates keeps the window that was typed, because "
     "there is nothing to derive it from; and a project with NO assignments keeps the team "
     "size that was typed, because 'nobody is assigned yet' is not the statement 'this needs "
     "nobody'. ONE CONSEQUENCE IS WORTH STATING: a project now ends at its last milestone, so "
     "work that legitimately continues past the final DB lock - close-out, archiving, an "
     "inspection - has to be carried by a milestone if it is to be inside the window, and an "
     "assignment running past it is reported as V-07. The period derivation already allowed "
     "for that headroom by running Close-out (final) to the later of the DB lock and the "
     "project end; deriving the end from the milestones removes it. (3) A new row in Periods "
     "is given the next PERIOD_SEQ for its project. The request named a 'period_id': the sheet "
     "has none, it is keyed on (project_id, period_name) per change R-11, and period_seq is "
     "the field that carries the order - so that is the number allocated, one past the highest "
     "in that project, exactly as milestone_seq is. (4) A weight-override window now shows the "
     "PROJECT and ROLE of its assignment even when that assignment has not been saved. It read "
     "them from the validated model, which excludes anything still being entered, so the two "
     "lookup columns went blank until Save - and they exist precisely so that, while typing, "
     "the user can confirm the window is attached to the right piece of work. They are read "
     "from the raw sheets now, and the same two columns were added to the scratch panel that "
     "a plan started blank uses. tools/test_derive.py covers all four; seven of its ten checks "
     "fail against the previous build. No requirement, rule, calculation or schema version "
     "changed.",
     "Superseded by v2.25"],
    ["2.23", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 22, application v1.21. Three changes to what a new row arrives with, and "
     "to whether the reader can tell there is more to see. (1) IDENTIFIERS are now "
     "allocated on insert for project_id, person_id and milestone_seq as well as "
     "assignment_id, and the rule is ONE PAST THE HIGHEST already in the sheet rather than "
     "the smallest unused. The old rule was about tidiness - a file whose ids ran 1..46 "
     "and then jumped to 901 for a few hand-placed rows would not start climbing from 902; "
     "the new one is about being able to tell at a glance which rows were added today, and "
     "a gap left by a deleted row now stays a gap, which is how an identifier sequence "
     "behaves everywhere else. milestone_seq counts within its own PROJECT, not across the "
     "file, because that is the list it orders. (2) WEIGHTS start at 1.00: "
     "ProjectPeriod.weight, Person.capacity_fte, Assignment.person_weight and "
     "PersonPeriodWeight.weight_override. 1.00 is the neutral multiplier, and the "
     "alternative is not 'no value' but ZERO - an empty weight reads as 0.00 in the "
     "calculation, so a row left alone contributed nothing at all and nothing on screen "
     "said so. A row that contributes too much is noticed; a row that contributes nothing "
     "is not. This required a second concept alongside 'blank': a row carrying only the "
     "values the APPLICATION put there is still an empty row, so Save does not promote it "
     "and Export refuses to write it - otherwise every insert would immediately become a "
     "half-record complaining about what it is missing. (3) SCROLL REGIONS now say which "
     "way there is more. Every section was already bounded on both axes, but "
     "scrollbar-width:thin bought a browser OVERLAY bar that occupies no layout space and "
     "fades when idle - measured, offsetWidth === clientWidth - so a table with eleven "
     "columns off to the right looked exactly like a table with none. The bar is now a "
     "real 12px with a visible thumb; and because overlay scrollbars are the default on "
     "some platforms whatever the CSS asks for, each region ALSO carries a soft shade on "
     "any edge that has content beyond it, drawn over the box so table rows cannot paint "
     "on top of it, and removed the moment that edge is reached so it always means 'more "
     "this way'. tools/test_newrow.py covers all three. Two existing tests had to change, "
     "and both were encoding old behaviour rather than finding new faults: test_rows was "
     "promoting an override row with no window and no weight in it, which isSkeleton now "
     "declines to do, and test_nokey could no longer reach a keyless row by skipping the "
     "identifier - the application supplies one - so it clears it explicitly instead. No "
     "requirement, rule, calculation or schema version changed.",
     "Superseded by v2.24"],
    ["2.22", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 21, application v1.20. A ROW WITH NO IDENTIFIER IS NO LONGER TREATED AS "
     "A RECORD, which fixes three reported faults that were all one fault. A project saved "
     "with every field filled in EXCEPT project_id was indexed as M.projects[null] - and "
     "JavaScript turns a null key into the STRING 'null'. The application then believed in "
     "a project called 'null'. It became the selection; and the row filter, which keeps a "
     "row when the set of visible project_ids contains it, compared the set {'null'} "
     "against the row's actual null and missed - so the Projects table the user had just "
     "typed into reported 'No rows. Use + row to add one.' The Milestones and Periods "
     "sections beneath went looking for a parent named 'null', found none, and reported "
     "the same; a row added there would have been stamped 'null' and lost. A person saved "
     "without a person_id did exactly the same, which is why the Assignments and Weight "
     "overrides sections beneath them offered nothing to fill in. One cause, three "
     "symptoms, on both tabs. The rule now is explicit: a row becomes a RECORD when it "
     "carries its sheet's identifier and not before. Until then it is not indexed, it is "
     "reported as an error naming the column to fill in, and Save is refused rather than "
     "silently creating the phantom - which matches the export guard, that has always "
     "refused to write such a row. Crucially the row STAYS ON SCREEN, because it is the "
     "row the user is in the middle of repairing: both the project and people tables keep "
     "any row that has no identifier alongside the real ones, and a child table now shows "
     "a row carrying no parent key wherever the user is, since a row that cannot be seen "
     "cannot be repaired or deleted, only silently dropped. Supplying the identifier "
     "recovers completely. The same rule went into tools/prap_io.py, so the command line "
     "and the page still report the same findings. tools/test_nokey.py covers all of it "
     "for a project, for a person, and for a workbook that already carries such a row; "
     "four of its eleven checks fail against the previous build, reproducing the reported "
     "symptoms exactly - '0 rows on screen', projects=['null'], and 'No rows. Use + row to "
     "add one.' No requirement, calculation or schema version changed.",
     "Superseded by v2.23"],
    ["2.21", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 20, application v1.19. SCROLL POSITION now survives a re-render. "
     "Committing a cell re-renders the panel it lives in, and a freshly built element "
     "starts at the top left - so filling one cell in a table twenty-two columns wide sent "
     "every scroll box back to the first row and the first column, and the next cell the "
     "user meant to fill was off screen again. On a dashboard that is invisible; during "
     "data entry it costs a re-scroll for every single cell, which is the difference "
     "between a usable table and an unusable one. It showed up as soon as a plan could be "
     "typed in from scratch (round 18) because that is the first time anyone fills in "
     "cells one after another. Every scroll offset on the page is now captured and put "
     "back around each re-render, along with the page's own scroll position, which a "
     "re-render can also move because the edit banner changes height between 'no changes' "
     "and 'n changes not yet saved'. The key has to survive the DOM being rebuilt: a data "
     "table names itself by its sheet, and everything else - the charts - is keyed by "
     "where it sits among the other keyless boxes in its own pane. The same wrapper covers "
     "the three other places that redraw part of the page: selecting a parent row, "
     "selecting an assignment, and the matrix/rows toggle on General assumptions. Changing "
     "TAB still takes the page to the top, because that is the user moving somewhere else "
     "rather than the page moving underneath them. tools/test_scroll.py drives the real "
     "browser and checks all four commit paths, the page scroll, a row 900px down a "
     "289-row table, and that the cell just filled is still on screen afterwards; five of "
     "its eleven checks fail against the previous build. The test itself needed care: "
     "Playwright's own click scrolls its target into view first, and Chromium's "
     "scrollIntoView scrolls EVERY scrollable ancestor, so a naive click on an off-screen "
     "cell moves the box before the application has run a line - the first draft of the "
     "test was measuring the driver rather than the page. No requirement, rule, "
     "calculation or schema version changed.",
     "Superseded by v2.22"],
    ["2.20", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 19, application v1.18. The four CHILD sections are now on screen from the "
     "moment a blank plan is started, and can be filled in before their parent is saved. "
     "Round 18 made a plan enterable but only in one order: the project tab showed the "
     "Projects table alone until a project had been saved, and the person tab the People "
     "table alone. Both were literally true - the detail panels need a parent RECORD, with "
     "dates for the timeline and a name for every heading, and a row still being typed is "
     "not even parsed into the model - but the consequence was that Milestones, Periods, "
     "Assignments and Weight overrides were invisible to anyone building their first plan. "
     "A section that appears only after you have done something else reads as a section "
     "that does not exist. So the tables are drawn from the start, in the same two-column "
     "arrangement they have when a workbook is loaded, and they behave in three states. "
     "With no parent at all they are LOCKED and carry the reason where the '+ row' button "
     "would be, because a child row whose parent does not exist has nothing to attach to "
     "and would be dropped when the file is read back - saying so beats offering a button "
     "that creates a row nobody can rescue. The moment the parent row carries an "
     "IDENTIFIER - before it is saved - they unlock and scope themselves to it, so a plan "
     "can be entered the way a person actually thinks about one: the project and its "
     "milestones together, the person and their assignments together, each committed in a "
     "single Save. Two fixes were needed for that to work. The draft row now becomes the "
     "tab's selection, since child rows inherit their parent key from the selection and "
     "without it they were created parentless. And the fallback that finds the assignment "
     "for a new override row now reads the raw sheet rather than the validated array: the "
     "validated array excludes anything still being entered, so on a plan being typed from "
     "scratch the only assignment on screen was not found. tools/test_blank.py now enters "
     "the milestones BEFORE saving the project and the assignment BEFORE saving the "
     "person, checks each child inherited the right foreign key, and adds a weight "
     "override window to prove the fourth section works and REPLACES the person weight for "
     "its three months and no others. No requirement, rule, calculation or schema version "
     "changed.",
     "Superseded by v2.21"],
    ["2.19", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 18, application v1.17. A PLAN CAN NOW BE STARTED IN THE APPLICATION "
     "ITSELF. The landing screen offers two ways in, given equal weight: load a source "
     "workbook, or start blank and type. 'Start blank' opens every tab and every section "
     "an upload opens - it is the same code path, so it is not a lesser mode with its own "
     "rules - and the plan it produces exports to the same .xlsx as any other. A blank "
     "start is not an EMPTY workbook: with no Lists sheet there would be no vocabulary, so "
     "every typed value would be reported as unrecognised (V-11) and no field could offer "
     "a choice, and with no Config there would be no thresholds. It therefore begins from "
     "the reference content of the delivered template, embedded in the HTML by "
     "tools/build_app_seed.py and held to the template by check_consistency.py. The "
     "PeriodWeightStandard and RoleFactor grids are NOT embedded: the application builds "
     "them from the value lists it was just seeded with - 56 standard weights and 289 role "
     "factors, every combination a project can reach - so a company that adds a role to "
     "Lists gets it without anybody regenerating anything. Every seeded figure is a "
     "placeholder 1.00 that says so on its own row, because an invented weight that looks "
     "like a company standard is worse than an obvious placeholder: only one of the two "
     "gets questioned. Four things had to be fixed before the path worked at all, each "
     "found by walking it rather than by reading it. (1) A DEADLOCK at the first save: a "
     "clinical trial saved before its milestones exist raises V-16 by definition, and its "
     "milestones cannot be entered until it is saved, because the milestone table hangs "
     "off a SELECTED project. The save guard now treats V-12 and V-16 as INCOMPLETENESS "
     "rather than error - the same line the specification already draws for drafts - names "
     "what is still missing in the banner, and refuses everything else exactly as before. "
     "(2) A row still being typed is not yet parsed into the model, so on a plan with no "
     "saved projects the Projects table said 'nothing matches the filters' and hid the "
     "very row the user was filling in. (3) The first assignment was the one row the "
     "application refused to name, having named every one after it: nextKey had nothing to "
     "copy a house style from, so the sheets that allocate their own keys now carry a "
     "pattern to start from. (4) FIVE PROJECT COLUMNS HAD NO HOME ANYWHERE IN THE "
     "APPLICATION. The project and people tables showed a curated subset, which was "
     "survivable while every plan arrived as a workbook filled in elsewhere; it is not "
     "survivable now that a plan can be built here, because DataReviewSystem_setup could "
     "never be entered and V-10 would warn about it forever. Both tables now carry every "
     "column the schema lets a user type into. Also: a person with no assignment yet is "
     "listed on the person tab, which is the state everybody is in for the minute after "
     "they are created; and the default horizon no longer anchors to year 0 when nothing "
     "has been calculated. tools/test_blank.py never touches a fixture - it clicks Start "
     "blank and builds a plan the way a person would, then checks every monthly figure "
     "against the formula worked by hand. No requirement, rule, calculation or schema "
     "version changed.",
     "Superseded by v2.20"],
    ["2.18", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 17, application v1.16. REFERENCE MATERIAL FOR ANOTHER AI SYSTEM, and an "
     "interoperability audit of everything delivered so far. The audit found one thing, and it was "
     "structural: of 95 files in the repository, 72 were .xlsx and 2 were Markdown. Every fact this "
     "project has established - the schema, the value lists, the formula, the twenty-four validation "
     "rules - existed only inside a ZIP of XML or inside the application's own JavaScript. A person "
     "with Excel could read all of it; a program, a script or a language model could read none of "
     "it. Four deliverables close that. (1) docs/prap_contract.json, a machine-readable contract "
     "carrying the schema, every column with its type and meaning, the value lists, the Config "
     "parameters, the formula term by term, the period derivation, all twenty-five rules with the "
     "severity the application actually reports, the interchange format and the task recipes. It is "
     "GENERATED - from the template builder, from this plan's own rule table, and from the "
     "application source - so it cannot drift from what it describes. (2) docs/PRAP_AI_Agent_Guide.md, "
     "the same contract written for a reader in the order an agent needs it, plus the same document "
     "as a workbook for human review. (3) A JSON INTERCHANGE FORMAT: the whole workbook as row "
     "objects keyed by column name, dates as yyyy-mm-dd. The application loads it and exports it, "
     "and tools/prap_io.py converts either way, so an agent that cannot write a ZIP of XML can still "
     "produce a file PRAP opens. (4) tools/prap_io.py also VALIDATES and CALCULATES from the command "
     "line, implementing the same rules and the same formula as the page - so an agent can check its "
     "own draft without a browser. tools/test_interop.py holds the two implementations to each "
     "other: same findings at the same severities, and every person-month equal to 1e-6, on both "
     "worked examples. The audit also found a rule the documents promised and the code did not keep: "
     "V-14 has been in the data model since v1.0 and the application never reported it. It does "
     "now - a milestone outside its project's window, or a boundary milestone out of order - with "
     "the one exception that is legitimate rather than merely noisy, an Inspection after the final "
     "DB lock, reported as information because the derivation deliberately extends the timeline to "
     "reach it. No requirement, calculation or schema version changed.",
     "Superseded by v2.19"],
    ["2.17", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 16, application v1.15. A visual design pass over the whole interface, taking "
     "Apple's Human Interface Guidelines as the reference and applying them to a dense data tool. "
     "The stylesheet is rebuilt around TOKENS - one palette, one type scale, one radius scale, one "
     "motion curve - so that no value is chosen by eye at the point of use, which is what keeps a "
     "hundred small decisions consistent. Four changes carry most of the difference. (1) MATERIAL: "
     "white cards on a grey field, separated by soft shadow instead of by a 1px outline round every "
     "panel; dark mode inverts to lifted grey cards on near-black. (2) VIBRANCY: the two sticky bars "
     "are translucent and blurred, so content scrolling beneath stays legible as context - at an "
     "opacity that keeps their own labels readable over a chart, because blur alone is not enough "
     "when what is underneath is high-contrast. (3) The tab strip becomes a SEGMENTED CONTROL, and "
     "buttons take Apple's filled / tinted / plain hierarchy; the row actions that appear twice on "
     "every table row are furniture at rest and take colour only on hover, where blue means 'adds' "
     "and red means 'removes'. (4) COLOUR MEANS SOMETHING: blue is interactive, red is over the "
     "ceiling, amber is under the floor, everything else is a grey. Chart series colours are exempt, "
     "because there the colour IS the data, and they are unchanged. Contrast is now measured rather "
     "than judged: tools/test_contrast.py composites every text colour onto the surface it actually "
     "sits on and applies WCAG AA at the rendered size. The previous design had five colours below "
     "AA; this one has none. Two defects were found and fixed in passing: the drop zone still named "
     "template v1.6, and #editbar set display:flex from an id selector, which outranks the `hidden` "
     "attribute - so an empty edit bar showed before any workbook was loaded. No requirement, rule, "
     "calculation or schema changed, and every existing selector is retained.",
     "Superseded by v2.18"],
    ["2.16", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 15, application v1.14, template v1.7, dummy v1.9 and 10x10 v1.1. DERIVED "
     "columns in the source workbook are now distinguished from entry columns by more than a "
     "colour: their cells are LOCKED, and the column heading carries a note saying what the column "
     "is, that the lock has no password, and what a hand-typed value would actually do. A green "
     "fill is a convention the reader must have been told about, and the telling is on a README "
     "sheet they may never open; a lock is the file itself refusing the edit at the moment it is "
     "attempted. Every other cell on those sheets is explicitly UNLOCKED, and protection is set so "
     "that inserting, deleting and sorting rows all still work - only adding or removing COLUMNS is "
     "blocked, because the column set is the schema. There is no password: it is a guard rail, not "
     "security, and anyone who genuinely needs to paste a column can turn it off in two clicks. The "
     "application's EXPORT now writes the same locks and the same sheet protection, so the guard "
     "rail survives a round trip instead of disappearing the first time a file goes through PRAP. "
     "check_consistency.py now verifies all of it against the plan's own data model - which "
     "surfaced a documentation gap the change had to close: Milestone.project_name and "
     "Assignment.person_name were typed 'Text' in the data model although the application "
     "recomputes both on import and reports any disagreement as V-13. They are typed 'Derived' now, "
     "which is what they have always been. No requirement, rule, calculation or schema version "
     "changed.",
     "Superseded by v2.17"],
    ["2.15", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 14, application v1.13. (1) A LINE CHART of monthly resource now stands as the "
     "FIRST panel of every data tab - one coloured line per project on Overall and on the project "
     "tab, one per person on the person tab, each keeping the colour it already carries elsewhere. "
     "The stacked charts say what a month's total is MADE OF; they cannot say whether one series is "
     "rising or falling, because in a stack every band's baseline moves with the bands beneath it. "
     "Lines share one baseline, which is the whole reason to draw the same numbers a second way. "
     "Each line carries its total, mean and peak month, and is capped at the twelve largest by total "
     "because sixty lines on one axis is a texture, not a reading. On the person chart the "
     "over-allocation ceiling is marked, since there it applies to a single series' own value. "
     "(2) A design pass over every section. Summary tiles move to the TOP of the Overall tab: they "
     "are the headline figures and sat below a chart. Every panel now carries ONE header shape - "
     "title on the left, what the panel covers on the right - where scope had been stated three "
     "different ways: inside the caption, in a note under the table, or not at all, so the reader "
     "had to hunt for it in a different place in each section. Captions are bounded to a readable "
     "measure. Two long-standing chart faults were fixed on the way: a transparent stroke is not a "
     "PAINTED one, so the line hit strips needed pointer-events stated explicitly, and the visible "
     "stroke sits on top of them and had to be taken out of hit testing or it swallowed the pointer "
     "and offered no tip of its own. No requirement, rule, calculation or schema changed.",
     "Superseded by v2.16"],
    ["2.14", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 13, application v1.12. PersonPeriodWeight is treated as what the schema says "
     "it is - a CHILD of Assignment - so on the person tab it now follows the assignment SELECTED "
     "above it, exactly as Milestones and Periods follow the project selected above them. Clicking "
     "a row in Assignments selects it and the overrides table redraws to that assignment's windows "
     "and no others; a new override row is seeded with the selected assignment rather than the "
     "person's first, so it is filed under the project on screen. Scoping the table to the PERSON "
     "instead - which is what round 12 left it doing - listed every window that person carries "
     "across every project at once, a set of rows with nothing to do with each other. Both headings "
     "now name the person ('Assignments - Kim S. (PSN-001)'), and the overrides panel restates the "
     "assignment it belongs to above the table: identifier, project, role, dates and weight. One "
     "structural change was needed to make this safe: the overrides panel is rendered into its own "
     "element, because clicking a cell both selects the row and puts the caret in it, and redrawing "
     "the whole person panel would rebuild the Assignments table under the caret so the edit could "
     "never be typed. The project and person tables escape this by living outside the panel they "
     "drive; this table drives a panel beside it. No requirement, rule, calculation or schema "
     "changed.",
     "Issued for review"],
    ["2.13", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 12, application v1.11. Three changes, none touching a requirement, rule, "
     "calculation or schema. (1) The Overall tab's 'Mean load per person' is replaced by "
     "'MONTHLY DEMAND BY PERSON': the same months as 'Monthly demand by project' directly above it, "
     "cut the other way, one stacked band per person with a colour per person. The two charts total "
     "the same figure every month, because they are the same person-months summed along different "
     "axes - and that is now an executable check, not a claim. The mean-and-thresholds reading the "
     "old chart gave is not lost: a segment is OUTLINED where that person's own month crosses the "
     "ceiling or the floor, so the flag never rests on fill colour, which here already means 'which "
     "person' (D-04). The 20-person cap from D-14 carries over to the stack and its legend "
     "(REQ-DSH-09). (2) 'Utilisation' on the project tab is stacked by PERSON on the same terms, "
     "with each segment's pop-up naming the person, their role on this project, their FTE and share "
     "of the month, and the project's total. Its relative reference lines are unchanged, and a month "
     "whose TOTAL crosses one is outlined behind the stack rather than coloured into any one "
     "person's segment. (3) DEFECT: a row drafted on one person's tab appeared on every other "
     "person's tab. The child-table filters carried a bare '|| r.__new', added at round 6 so a new "
     "row would be visible before its key was filled in; being unconditional it admitted every draft "
     "everywhere. On 'Weight overrides', whose project and role columns are looked up FROM the "
     "assignment, such a row then described another person's work - the incorrect mapping reported. "
     "A draft is now admitted only while its parent key is empty, applied through one helper to all "
     "four child tables. The overrides table is also scoped correctly as a GRANDCHILD of the person: "
     "by the identifiers of the assignments shown above it, not by the person's own.",
     "Superseded by v2.14"],
    ["2.12", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 refinements round 11, application v1.10. Two changes, neither touching a requirement, "
     "rule, calculation or schema. (1) A PROJECT TIMELINE run-chart now stands before the Utilisation "
     "panel on the project tab - the Overall tab's chart drawn for the selected project alone, with "
     "one band per period shaded by period weight, every milestone marked, the two DB locks in red, "
     "and each band labelled with its average FTE per month. The row-label gutter the Overall chart "
     "needs is dropped, since the panel heading already names the project. (2) The UTILISATION bars on "
     "the person tab are now STACKED BY PROJECT: one coloured segment per project per month. The total "
     "says whether someone is over the ceiling; only the split says because of what. Each segment's "
     "pop-up answers both halves at once - the project (name, type, the milestones that project "
     "passes THAT month, this person's FTE on it and its share of the month) and the person (name, "
     "identifier, total FTE that month, how many projects it is spread across, capacity, and which "
     "threshold if any it crosses). Over- and under-allocation is now drawn as an outline BEHIND the "
     "stack, because it is a property of the month's total and not of any one project in it. "
     "Supporting this, project colours are now allocated from the sorted list of ALL project ids and "
     "held for the session, so one project carries ONE colour everywhere - the Overall stacked chart "
     "picked its colours from the filtered draw order, which meant the same colour meant different "
     "projects on different tabs and after every filter change.",
     "Superseded by v2.13"],
    ["2.11", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 defect fix round 10, application v1.9. Reviewer reported that dragging the scroll bar of "
     "a type-ahead value list closed the list, making a list taller than its box impossible to read to "
     "the bottom of. Two independent causes, both now fixed. (1) A capturing scroll listener closed "
     "the list on ANY scroll anywhere in the document - including the list's OWN scrolling. It existed "
     "because the list is a fixed box anchored to its cell and a page scroll would leave it stranded; "
     "closing was the crude answer. It now ignores its own scrolling and RE-ANCHORS on any other, "
     "closing only once the cell itself has left the screen. (2) Pressing on the list's scroll bar took "
     "focus off the cell, and the focus-out handler closed the list underneath the hand scrolling it. "
     "The press is now held for its whole duration and the caret is restored on release. The list "
     "closes on exactly the three things asked for: choosing a value, Escape, or a click outside. "
     "Because Escape leaves the caret in the cell and so fires no focus event, a click in the field "
     "now re-opens the list. Matching is also improved: tokens match in ANY ORDER, so 'phase 1 onv' "
     "finds 'ONV-101 Phase 1'; every matched fragment is marked, not just the first; results rank a "
     "typed prefix above a mere containment; and a query matching nothing now shows the whole "
     "vocabulary with an explanation instead of vanishing and stranding the reader. That last change "
     "exposed a pre-existing trap - render() highlighted the first row, so Enter silently swapped in a "
     "value the user never chose, which had escaped notice only because a query matching nothing used "
     "to hide the list and put Enter out of its reach. Nothing is highlighted now until an arrow key "
     "highlights it. No requirement, rule, calculation or schema changed.",
     "Issued for review"],
    ["2.10", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 defect fix round 9, application v1.8. Reviewer reported that choosing an assignment_id "
     "on a new Weight-overrides row raised a warning offering to change every other override of the "
     "assignment being moved away from, so a second window could not be added. Cause: the edit path "
     "cascaded an identifier change whenever the edited column was the sheet's KEY_COL, but on a "
     "CHILD sheet that column is a FOREIGN key - PersonPeriodWeight.assignment_id points at an "
     "assignment, it does not define one. Re-pointing one row therefore rewrote every row sharing "
     "the old value, which on the overrides table is the sibling windows of that same assignment. "
     "Accepting the prompt destroyed them; declining abandoned the edit, so the assignment could "
     "not be changed at all and the row stayed on whichever assignment it was seeded with. The "
     "cascade is now confined to the sheet that OWNS the identifier, using the OWNER map that "
     "deleteRow has used for the same distinction since round 4. In place of the warning, picking "
     "an assignment that already carries windows now states how many the row joins and names V-06 "
     "and V-24 - an additional window is the normal case, and only an overlapping one is refused, "
     "at Save. No requirement, rule, calculation or schema changed.",
     "Superseded by v2.11"],
    ["2.9", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 refinements round 8, application v1.7. Three changes, none touching a requirement, "
     "rule, calculation or schema. (1) EVERY scroll region now scrolls in BOTH directions and is "
     "bounded on both axes. A panel bounded on one axis only can still hide content on the other, "
     "and the reader has no way of telling that it has - there is no scrollbar to suggest anything "
     "is missing. The findings dialog and the provenance strip were the two regions with no bound "
     "at all; the three chart panels were bounded sideways only. (2) On the Assignments table the "
     "project is now identified by NAME: project_name is the field typed into and project_id is "
     "derived from it, with a type-ahead over the project names. Editing the identifier directly "
     "still works and the name follows instead - the two are one field seen from two ends. A name "
     "matching no project is refused, naming near matches; an ambiguous one is refused too, since "
     "only the identifier is guaranteed unique. The name is never STORED on the assignment row: "
     "the sheet has no such column, so a copy could not survive an export and re-import, and it "
     "would be a second version of a fact the Project sheet already owns. (3) '+ row' on the "
     "Assignments table allocates the next free assignment_id - the smallest unused number in the "
     "pattern the file already uses, not the largest plus one, so a file whose ids jump to 901 for "
     "a few hand-placed rows does not start allocating 905.",
     "Superseded by v2.10"],
    ["2.8", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 round 7, application unchanged at v1.6. A second dummy dataset at the reviewer's "
     "requested size - 10 projects and 10 people - issued as PRAP_SourceData_Dummy_10x10_v1.0.xlsx "
     "on schema 5, alongside the 62-project set rather than replacing it. The generator now runs "
     "from size PROFILES, so one code path produces both and neither can drift from the schema "
     "the other follows; the 62-project file rebuilds with every sheet byte-identical, confirmed "
     "by unpacking both .xlsx containers. The small set is not a lighter test: it holds both "
     "clinical types, all four phases, trials with and without an interim DB lock, inspections "
     "that open the seventh period, hand-entered 'Others' periods, two part-timers, multi-window "
     "weight overrides, and both allocation thresholds crossed. Its load distribution matches the "
     "large set's (median 0.80 against 0.87 FTE). tools/test_app.py now runs over BOTH fixtures. "
     "No requirement, rule, calculation or schema changed.",
     "Superseded by v2.9"],
    ["2.7", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 defect fix round 6, application v1.6. Reviewer reported that inserting and deleting rows "
     "did not work on Milestones, Periods, Assignments and Weight overrides - the four CHILD tables, "
     "which are the four rendered filtered to a selected parent. Three defects, all in the shared "
     "insert/rebuild path, none touching a requirement, rule, calculation or schema. (D-1) A new row "
     "was created with every column empty, so the parent filter that decides which rows the table shows "
     "excluded it: the row existed but was invisible. A new child row is now seeded with the parent it "
     "was created under, and each filter admits a row still being drafted. (D-2) Every edit and delete "
     "re-parses the data, and the parser discards blank rows - correct for a file being read, wrong for "
     "a row the user has not filled in yet, which was silently destroyed by the next action. Rows the "
     "parser would discard are now held out of the parse and put back in place. (D-3) The parser numbers "
     "rows by position, so deleting a row renumbered every row below it while pending edits, edit "
     "highlights and the rendered cells still named the old numbers - an edit could end up attached to a "
     "different record, and a drafted row could be duplicated. Row identity is now preserved across a "
     "rebuild. A drafted row is exempt from validation while it is being written and is validated when "
     "Save is pressed; a draft that would break a rule is refused by name rather than kept.",
     "Superseded by v2.8"],
    ["2.6", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 refinements round 5, application v1.5. Five changes, none touching a requirement, rule, "
     "calculation or schema. (1) and (2) The Assignments and Weight-overrides tables carry the project "
     "name, and the overrides table also the role - looked up from the master row and marked as such, "
     "since an identifier alone does not tell a reader which project they are looking at. (3) Every data "
     "cell shows its value and its column's meaning on hover or focus. (4) The Periods table and the "
     "General-assumptions tables gain row actions. The two weight sheets and the value lists keep their "
     "matrix view for reading and gain a rows view for editing, because a matrix row is six or seven "
     "workbook rows and cannot carry a per-row action. (5) Type-ahead on every column with a vocabulary: "
     "the Lists sheet where the column is backed by one, the distinct values already present where it is "
     "not, and the type-dependent set for period_name and role_name.",
     "Superseded by v2.7"],
    ["2.5", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 refinements round 4, application v1.4. (1) Information pop-ups: every column heading on the "
     "source-data tabs explains what its column is for, and the summary tiles, month headings and project "
     "type pills explain what they count. They open on HOVER and PIN on CLICK, which is also what makes "
     "them reachable on a touch screen and from the keyboard, where there is no hover at all. (2) Every "
     "editable table gains a Delete control beside Insert. V-17 governs it: a row still referenced cannot "
     "be deleted and the refusal names what refers to it, and a delete is NEVER cascaded - that asymmetry "
     "with the cascading identifier edit is deliberate, because a cascaded rename is reversible and a "
     "cascaded delete is not. Deletion is provisional like any other change. No requirement, rule, "
     "calculation or schema changed.",
     "Issued for review"],
    ["2.4", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 refinements round 3, application v1.3. Three changes. (1) Column headings stay visible in "
     "EVERY scroll region, and the corner cell of the two Overall grids is pinned in both directions - a "
     "month heading that scrolls away turns the table into unlabelled numbers. (2) The two utilisation "
     "charts carry the year on the x axis, shown where the year changes with a divider at each boundary. "
     "(3) EDITING IS NOW PROVISIONAL: a change is applied on screen so its effect is visible, but is held "
     "as pending until the user presses Save or Leave without change. Save commits it to the working data; "
     "Leave without change restores the snapshot taken before the first pending edit. Export is HELD while "
     "anything is pending, so nothing unconfirmed can reach a file. This is a behavioural change to "
     "REQ-IMP-07 and REQ-IMP-08 and is recorded for the specification to take up at the next issue.",
     "Issued for review"],
    ["2.3", DOC_DATE, "Claude Code", "Pending",
     "Gate 4 refinements round 2, application v1.2. The application now shows, at the top of the page, "
     "which controlled documents the build implements - development plan, programming specification, UI "
     "component list and source-data template - collapsed to one line with the full table on demand, and "
     "a row comparing the loaded workbook's schema version against the one the application expects. This "
     "is REQ-VC-01 made visible in the product rather than only in the repository. The list is data in one "
     "place and is verified against the repository by tools/check_consistency.py on every run, which was "
     "proved to fire by naming a superseded plan version. No requirement, rule, calculation or schema "
     "changed.",
     "Issued for review"],
    ["2.2", DOC_DATE, "Claude Code", "Pending",
     "First refinements from the Gate 4 review (WBS 4.9), applied to application v1.1. Three UI changes, "
     "no change to any requirement, rule, calculation or schema. (1) The three named Overall sections - "
     "monthly demand, resource by project, resource by person - are up to twice their previous size. "
     "(2) The project timeline scrolls in both directions inside its panel, which also let the 20-project "
     "display cap be removed: every project in the filter is now drawn. (3) Chart text is larger and "
     "higher contrast, and each threshold or reference label moved out of the plot area into a right-hand "
     "margin - drawn over the bars they were unreadable at any size, and enlarging them made the "
     "collision worse rather than better. Breaching bars on the person chart now state their own value.",
     "Issued for review"],
    ["2.1", DOC_DATE, "Claude Code", "Pending",
     "Step 4 progress against the v2.0 baseline. The application is built: app/PRAP.html, one offline "
     "file with its own .xlsx reader and writer. WBS tasks 4.1 to 4.8 move to Complete; 4.9 - your "
     "review of the output against real data - is the Gate 4 review and is what remains. No requirement, "
     "rule or schema changed: this issue records progress, not a change of intent. The calculation "
     "engine was checked cell-by-cell against tools/verify_source_workbook.py and matches exactly across "
     "all 1,225 person-months of the dummy dataset.",
     "Issued for review"],
    ["2.0", "2026-08-02", "Claude Code", "Dan",
     "APPROVED BASELINE. Content identical to v1.10; the version number changes because the numbering "
     "rule on sheet 09 reserves x.0 for an approved baseline and v1.10 was a change awaiting signature. "
     "Approved together with programming specification v1.0 and UI component list v1.0. Changes R-05 to "
     "R-13 are now part of the baseline rather than pending against v1.3, and Step 4 is authorised.",
     "APPROVED 2026-08-02 by Dan - BASELINE"],
    ["1.10", DOC_DATE, "Claude Code", "Pending",
     "Change R-13, from the component-list v0.7 review. All 13 changed components and all 7 changed "
     "decisions were confirmed OK; nothing was marked Rework. Two new requests were raised and both are "
     "applied: the interim and final DB lock milestones are emphasised on the timeline, and the project "
     "source-data tab gains a utilisation graph. The second needed a new requirement, REQ-DSH-12, because "
     "a project has no absolute ceiling or floor - its reference lines are relative to the portfolio "
     "average and to the project's own lifetime average. No schema change.",
     "Issued for approval"],
    ["1.9", DOC_DATE, "Claude Code", "Pending",
     "Change R-12. A reviewer question about the PersonPeriodWeight key - why period_start is needed "
     "when assignment_id looks unique - was answered by confirming the key is correct: one assignment "
     "may carry several non-overlapping override windows, so assignment_id alone would cap it at one. "
     "Checking the question found two rules that were specified but never implemented. V-06's "
     "assignment-window half now runs, and V-24 is added for referential integrity and key uniqueness "
     "on PersonPeriodWeight. The dummy fixture gains an assignment with two windows, which is what "
     "would have caught both. No schema change.",
     "Issued for approval"],
    ["1.8", DOC_DATE, "Claude Code", "Pending",
     "Change R-11 against the v1.3 baseline. The conduct phase is split by NAME rather than by sequence: "
     "'Conduct (interim)' where an interim DB lock exists and the stretch runs before it, 'Conduct (final)' "
     "after it or where there is no interim lock. Period names are therefore unique within a project and "
     "ProjectPeriod is keyed on (project_id, period_name); period_seq carries order only. The clinical "
     "period set grows to seven names, V-18 becomes a uniqueness check, REQ-CAL-11 and REQ-PRJ-12 are "
     "reworded, and REQ-DSH-10 is now satisfied structurally rather than by display numbering. Schema "
     "version steps 4 to 5. Weights are unchanged, so the dummy dataset produces identical figures.",
     "Superseded by v1.9"],
    ["1.7", DOC_DATE, "Claude Code", "Pending",
     "Change R-10 against the v1.3 baseline, from the component-list v0.4 review. The role factor is now "
     "keyed on project type, clinical phase, period and role rather than type and role alone, so a role's "
     "burden can vary across the life of a project. RoleFactor gains clinical_phase and period_name and "
     "grows from 13 rows to 249; REQ-CAL-02 reworded; V-23 added for a factor missing on a period an "
     "assignment actually spans; source schema version steps 3 to 4. A consequence is recorded on sheet 04: "
     "RoleFactor and PeriodWeightStandard now vary over the same three dimensions and multiply together, "
     "so they must not both be edited for the same reason.",
     "Superseded by v1.8"],
    ["1.6", DOC_DATE, "Claude Code", "Pending",
     "Change R-09 against the v1.3 baseline, from the component-list v0.3 review. Nine components changed; "
     "two needed requirements the plan did not have. REQ-DSH-11 added - the standing assumptions get their "
     "own tab. REQ-IMP-11 added - a row can be inserted, positioned below the row acted on. REQ-DSH-05 "
     "reworded for the clinical-phase filter and to stop implying the display unit is a filter; REQ-IMP-05 "
     "reworded for the time zone. Sheet 06 now describes four tabs. Design decision D-06 is superseded by "
     "the reviewer's own O-10, which asks for the opposite. Version history rows for v1.4 and v1.5 added "
     "above - they were issued without one, which REQ-VC-04 requires.",
     "Issued for approval"],
]
r = table(ws, r, ["Version", "Date", "Author", "Reviewer", "Summary of change", "Status"],
          rows, [10, 12, 15, 14, 92, 16], wrap_cols=(5,), mark_col=1)
r = note(ws, r, "The reviewer's file was named v0.11. Under the numbering rule on sheet 09 that reads as a draft "
                "increment of v0.1, so it is logged above as review input and this incorporating draft is v0.2.")

# ---- 02 Scope -------------------------------------------------------------
ws, r = sheet(wb, "02_Scope", "Scope and objectives")

r = section(ws, r, "Background - the problem being solved")
r = lines(ws, r, [
    "Multiple projects run at the same time, and their timelines change frequently. Deciding who to assign, in which",
    "role, to which project is therefore difficult to do by hand. The resource burden a project places on a person is",
    "not constant: it varies by project, by period within that project, and by the role the person holds.",
    "",
    "Today there is no single view that answers 'who is over-committed next quarter, and on which project?' -",
    "nor the quieter question, 'who has been under-used for three months running?'.",
])
r += 1

r = section(ws, r, "Objectives")
objectives = [
    ["OBJ-1", "Hold project and person data in one controlled place: a single Excel workbook kept outside the application.",
     "The workbook loads without error and round-trips through export.", "Q-09"],
    ["OBJ-2", "Simulate monthly resource demand per project and per person, respecting period weights and role factors.",
     "Monthly FTE reconciles to the worked example on sheet 05.", "Q-01"],
    ["OBJ-3", "Make both over- and under-allocation visible before they bite.",
     "Person-months above 1.50 FTE are flagged; runs of three or more months below 0.80 FTE are flagged.", "Q-08"],
    ["OBJ-4", "Let the user re-run the simulation immediately after a timeline change.",
     "Editing dates in the workbook - or in the application - and re-running updates the dashboard with no code change.", "Q-01"],
    ["OBJ-5", "Run entirely on a local Windows PC with no install, no server and no network.",
     "The HTML file opens by double-click in Edge and Chrome and works offline.", "Q-05"],
]
r = table(ws, r, ["ID", "Objective", "How success is measured", "Basis"],
          objectives, [10, 60, 60, 10], wrap_cols=(2, 3))

r = section(ws, r, "In scope")
in_scope = [
    ["Single-file HTML application runnable from local disk on Windows.", "v0.1"],
    [f"{MARK_CHG}Import of all project and person data from ONE Excel workbook via a browser file picker or drag-and-drop.", "Q-09"],
    [f"{MARK_NEW}Editing imported data inside the application, with the edits carried into the export.", "REQ-IMP-07"],
    ["Export of the current working data back to .xlsx, so the workbook remains the archive of record.", "v0.1"],
    [f"{MARK_CHG}Monthly resource simulation in FTE per project and per person, default 24-month horizon, expandable to the latest project end date.", "Q-08, Q-11"],
    ["Three dashboard tabs: Overall, Source data (project), Source data (person).", "v0.1"],
    ["Tables plus graphs on the Overall tab.", "v0.1"],
    [f"{MARK_NEW}Over-allocation and under-allocation detection against the agreed FTE thresholds.", "Q-08"],
    ["Version control of plan, specification and application across the development lifecycle.", "v0.1"],
]
r = table(ws, r, ["In scope for v1.0", "Basis"], in_scope, [124, 12], wrap_cols=(1,), mark_col=1)

r = section(ws, r, "Out of scope")
out_scope = [
    ["Multi-user concurrent editing, or any server / database component.", "Requirement is a local single-file HTML tool.", "Confirmed Q-12"],
    ["Authentication, user accounts, permissions.", "No server; the workbook carries whatever access control the file share provides.", "Confirmed Q-12"],
    ["Automatic writing to the workbook on disk without user action.", "Browsers cannot do this from a file:// page - see 07_Architecture. Export is always a deliberate act.", "Confirmed Q-12"],
    ["Cost, budget or salary calculation.", "Not raised. The data model leaves room to add it.", "Confirmed Q-12"],
    ["Integration with an HR, CTMS or timesheet system.", "Not raised.", "Confirmed Q-12"],
    ["Resource optimisation / automatic assignment suggestions.", "v1.0 reports and simulates; it does not decide.", "Confirmed Q-12"],
    ["Working-day and holiday calendars.", "Q-02 confirmed calendar-day pro-rating. A working-day calendar is a later option.", "Confirmed Q-02"],
]
r = table(ws, r, ["Excluded from v1.0", "Reason", "Status"], out_scope, [66, 60, 16], wrap_cols=(1, 2))
r = note(ws, r, "Q-12 answer was 'Fine' - the v0.1 exclusions stand. Note that in-application editing has moved INTO scope "
                "via REQ-IMP-07; only automatic writing to disk without user action remains excluded.")

r = section(ws, r, "Users")
users = [
    ["Resource planner / manager", "Primary", "Loads the workbook, reviews monthly simulation, spots over- and under-allocation, tries timeline scenarios."],
    ["Project lead", "Secondary", "Checks the resource profile of their own project and who is assigned in which role."],
    ["Data maintainer", "Secondary", "Keeps the source workbook accurate; may be the same person as the planner."],
]
r = table(ws, r, ["User", "Type", "What they do with the application"], users, [28, 12, 96], wrap_cols=(3,))

# ---- 03 Requirements ------------------------------------------------------
ws, r = sheet(wb, "03_Requirements", "Requirement register",
              "The contract for Steps 2-5. Every specification section and every code module cites the REQ-IDs it satisfies.")

reqs = [
    ["REQ-OUT-01", "Output", "The application is a single HTML file that runs on a local Windows PC by double-click, with no install and no network access.", "Must", "Requester", "4"],
    ["REQ-OUT-02", "Output", "All program logic (HTML, CSS, JavaScript, libraries) is embedded in that one HTML file so it can be copied between PCs as a single artifact.", "Must", "Requester", "4"],
    [f"{MARK_CHG}REQ-OUT-03", "Output", "Source data lives in ONE Excel workbook held separately from the HTML file, and that workbook is the archive of record.", "Must", "Q-09", "4"],
    ["REQ-OUT-04", "Output", "The application reads the source workbook and uses it to drive every table and graph.", "Must", "Requester", "4"],
    ["REQ-OUT-05", "Output", "Plans and specifications are delivered as Excel workbooks.", "Must", "Requester", "1,2"],

    [f"{MARK_CHG}REQ-PRJ-01", "Project data", "Each project records a project type, restricted to 'NewDrug CT', 'Biosimilar CT' or 'Others'. The first two are clinical trials and every clinical trial is exactly one of them.", "Must", "R-05", "2"],
    [f"{MARK_CHG}REQ-PRJ-02", "Project data", "Each project records a project category. For either clinical trial type the category is the product name; the field is optional for 'Others'.", "Must", "R-05", "2"],
    ["REQ-PRJ-03", "Project data", "Each project records a project name, unique within the source workbook.", "Must", "Requester", "2"],
    [f"{MARK_CHG}REQ-PRJ-04", "Project data", "Each project records its conditions: outsourcing type ('Full outsourcing' / 'Partial outsourcing' / 'Full In-house') and the number of project members.", "Must", "Q-14", "2"],
    [f"{MARK_CHG}REQ-PRJ-06", "Project data", "Each project carries a resource weight per period. For a clinical trial the weight is seeded from a standard keyed on clinical phase; for 'Others' it is entered by hand.", "Must", "Q-26, Q-28", "2"],
    ["REQ-PRJ-05", "Project data", "Each project records a timeline: start date, major milestone dates, and total period.", "Must", "Requester", "2"],
    ["REQ-PRJ-07", "Project data", "The project record accepts further project-related information without a schema change (free/extension columns).", "Should", "Requester", "2"],
    ["REQ-PRJ-08", "Project data", "Total period is derived from start and end dates rather than typed by hand, so it cannot contradict the timeline.", "Should", "Derived", "2"],
    [f"{MARK_CHG}REQ-PRJ-12", "Project data", "The set of periods a project carries depends on its type: either clinical trial type uses Before-Start-up / Start-up / Conduct (interim) / Close-out (interim) / Conduct (final) / Close-out (final) / After Close-out (final); 'Others' uses Planning / Develop / Close. No name occurs twice in one project.", "Must", "Q-18, Q-23, R-02, R-11", "2"],
    [f"{MARK_NEW}REQ-PRJ-13", "Project data", "A milestone name may occur more than once in one project. 'Inspection' in particular may record several events; the others are expected once.", "Must", "R-01", "2"],
    [f"{MARK_CHG}REQ-PRJ-09", "Project data", "A clinical trial of either type records its clinical phase (phase 1 / 2 / 3 / 4). The phase determines the project's period weights, so it drives the simulation rather than merely describing the project.", "Must", "Q-26", "2"],
    [f"{MARK_NEW}REQ-PRJ-10", "Project data", "A clinical trial of either type records who performs each of EDC set-up, data-review-system set-up, RBQM set-up and DM conduct ('by CRO' / 'by SB').", "Must", "Reviewer v0.11", "2"],
    [f"{MARK_NEW}REQ-PRJ-11", "Project data", "A clinical trial of either type records the EDC system, data review system and RBQM system in use, from data-driven value lists.", "Must", "Reviewer v0.11", "2"],

    ["REQ-PSN-01", "Person data", "Data is managed by person, by project assigned, and by project role assigned.", "Must", "Requester", "2"],
    ["REQ-PSN-02", "Person data", "Each assignment records the project(s) the person is assigned to.", "Must", "Requester", "2"],
    [f"{MARK_CHG}REQ-PSN-03", "Person data", "Each assignment records the role(s) the person holds on that project, drawn from the role list valid for that project's type.", "Must", "Q-03", "2"],
    ["REQ-PSN-04", "Person data", "Each assignment records the start date the person joins and the end date they leave that study.", "Must", "Requester", "2"],
    [f"{MARK_CHG}REQ-PSN-05", "Person data", "Each assignment carries a person weight saying how much that person works on that project, with optional period-specific overrides.", "Must", "Q-01", "2"],
    ["REQ-PSN-06", "Person data", "The person record accepts further project-related information without a schema change.", "Should", "Requester", "2"],
    ["REQ-PSN-07", "Person data", "One person may hold assignments on several projects simultaneously, and may hold more than one role on the same project.", "Must", "Derived", "2"],

    [f"{MARK_CHG}REQ-CAL-01", "Calculation", "Resource is simulated on a monthly grid, default horizon 24 months, expandable to the latest project end date.", "Must", "Q-11", "4"],
    [f"{MARK_CHG}REQ-CAL-02", "Calculation", "Monthly load for an assignment = project period weight x (role factor / people sharing that role) x person weight x fraction of the month covered. There is no separate base allocation. The role factor is selected by project type, clinical phase, WORK SCOPE, the period the month falls in, and the role - so a role's burden can vary across the life of a project and with how much of the work is kept in-house.", "Must", "Q-01, R-10, R-12, R-13", "2,4"],
    [f"{MARK_NEW}REQ-CAL-15", "Calculation", "BOTH ASSIGNMENT DATES ARE OPTIONAL. A blank assign_start_date means the project's own start date, and a blank assign_end_date the project's own end date - so an assignment with neither runs for the whole project, which is what most assignments are. Filling them in is for a PARTIAL involvement. The end date already behaved this way; the start did not, and a blank one made the assignment contribute nothing at all with no finding to say why.", "Must", "R-15", "4"],
    [f"{MARK_NEW}REQ-CAL-14", "Calculation", "Where SEVERAL PEOPLE hold the same role on the same project in the same month, the role factor is DIVIDED BETWEEN THEM. The role factor states what the ROLE costs the project in that period, not what each person holding it costs, so a project staffed by two data managers must not cost twice a project staffed by one. The count is per month, so a sharer leaving restores the others' full share without any edit; it counts distinct PEOPLE, so one person recorded on two rows is one person; and each person's own person_weight then applies to their share. Kept as the setting split_shared_role_fte so a figure produced before this rule can still be reproduced.", "Must", "R-13", "4"],
    ["REQ-CAL-03", "Calculation", "Project monthly load is the sum of its assignments; person monthly load is the sum across all their projects.", "Must", "Requester", "4"],
    [f"{MARK_CHG}REQ-CAL-04", "Calculation", "A person-month whose total exceeds the over-allocation threshold (default 1.50 FTE) is flagged as over-allocated. The threshold is absolute, not scaled by capacity (S2-01).", "Must", "Q-08, S2-01", "4"],
    [f"{MARK_CHG}REQ-CAL-05", "Calculation", "A partial first or last month is pro-rated by calendar days worked, not counted as a whole month.", "Must", "Q-02", "4"],
    ["REQ-CAL-06", "Calculation", "All weights, factors and thresholds are data, held in the source workbook, never hardcoded in the program.", "Must", "Derived", "4"],
    [f"{MARK_CHG}REQ-CAL-07", "Calculation", "A person whose monthly total stays below the under-allocation threshold (default 0.60 FTE) for three or more consecutive months is flagged as under-allocated, with the run's start and length reported. The threshold is absolute, not scaled by capacity.", "Must", "Q-08, S2-01, S2-05", "4"],
    [f"{MARK_NEW}REQ-CAL-08", "Calculation", "Load is expressed in FTE, where 1.00 FTE = 160 hours per month (8 h/day x 5 days/week x 20 days/month). Hours are shown alongside FTE where useful.", "Must", "Q-08", "4"],
    [f"{MARK_CHG}REQ-CAL-09", "Calculation", "For a clinical trial of either type, period boundaries are computed from milestone dates plus fixed month offsets (sheet 05), so a timeline change re-shapes the periods without re-typing them. For 'Others', periods are entered directly - those projects have no milestone mapping.", "Must", "Q-16, Q-22, Q-23, Q-25", "2,4"],
    [f"{MARK_NEW}REQ-CAL-10", "Calculation", "Every month of a project's timeline falls in exactly one period. A month left uncovered by the derivation is reported and carries weight 1.00 rather than being dropped from the simulation.", "Must", "Derived from Q-16", "2,4"],
    [f"{MARK_CHG}REQ-CAL-11", "Calculation", "Where an interim DB lock splits the conduct phase, the two stretches are separate rows carrying DIFFERENT names: 'Conduct (interim)' before the interim lock, 'Conduct (final)' after it. A project with no interim lock carries a single 'Conduct (final)'. Period names are therefore unique within a project.", "Must", "Q-23, R-11", "2,4"],
    [f"{MARK_NEW}REQ-CAL-13", "Calculation", "Where a period boundary has both a milestone and a month-offset definition, a recorded milestone date wins. The offset is the fallback for a project that does not carry that milestone.", "Must", "R-02", "2,4"],
    [f"{MARK_NEW}REQ-CAL-12", "Calculation", "Where a timeline squeezes a derived period to zero or negative length, that period is omitted and the omission reported. Boundaries are applied in order so the periods always stay contiguous.", "Must", "Derived from Q-22, Q-23", "2,4"],

    ["REQ-DSH-01", "Dashboard", "Tab 'Overall' shows monthly resource simulation per project and per person as tables.", "Must", "Requester", "3,4"],
    ["REQ-DSH-02", "Dashboard", "Tab 'Overall' shows appropriate graphs of the same simulation.", "Must", "Requester", "3,4"],
    ["REQ-DSH-03", "Dashboard", "Tab 'Source data (project)' shows project information as a table.", "Must", "Requester", "3,4"],
    ["REQ-DSH-04", "Dashboard", "Tab 'Source data (person)' shows person information as a table.", "Must", "Requester", "3,4"],
    [f"{MARK_CHG}REQ-DSH-05", "Dashboard", "One filter set drives every tab: date horizon, project type, clinical phase, project, person, role and department, with a single action to clear them all. The display unit is a setting, not a filter - it changes how figures are written, not which are shown.", "Should", "Derived / O-03", "3,4"],
    ["REQ-DSH-06", "Dashboard", "Any table on screen can be exported to Excel.", "Should", "Derived", "4"],
    [f"{MARK_NEW}REQ-DSH-07", "Dashboard", "The horizon control offers 24 months by default and a one-click expansion to cover the latest project end date across all projects.", "Must", "Q-11", "3,4"],
    [f"{MARK_NEW}REQ-DSH-08", "Dashboard", "Over-allocated and under-allocated person-months are distinguishable at a glance, and both are counted in the summary tiles.", "Must", "Q-08", "3,4"],
    [f"{MARK_NEW}REQ-DSH-09", "Dashboard", "At the target volume a table taller than the viewport renders only the visible rows, and the per-person chart shows an aggregate or a ranked subset rather than one bar per person. A thousand bars is not a chart.", "Must", "S2-06", "3,4"],
    [f"{MARK_CHG}REQ-DSH-10", "Dashboard", "Every period of a project is distinguishable from every other on screen. Since R-11 this is satisfied by the data model rather than by a display rule - no name repeats, so a name alone identifies a period. The display numbering that satisfied it before is retained only as a guard against a repeat reaching the renderer.", "Must", "S2-04, R-11", "3,4"],
    [f"{MARK_NEW}REQ-DSH-12", "Dashboard", "The project source-data tab shows the selected project's monthly resource over the horizon, against RELATIVE reference lines - a multiple and a fraction of the average an active project draws across the portfolio, and the project's own average over its full life. A project has no absolute ceiling or floor, so the references are context rather than pass or fail.", "Should", "Reviewer v0.7", "3,4"],
    [f"{MARK_NEW}REQ-DSH-11", "Dashboard", "The standing assumptions the simulation multiplies by - standard period weights, role factors, configuration and value lists - are presented on their own tab, so a reader can see what every figure was derived from without opening the workbook.", "Must", "G-07", "3,4"],

    [f"{MARK_CHG}REQ-IMP-01", "Import/Export", "The user loads the source workbook through a file picker or drag-and-drop, chosen because a local HTML page cannot open a file from disk unaided.", "Must", "Decision D-01", "4"],
    ["REQ-IMP-02", "Import/Export", "Loading validates the workbook and reports every problem found - missing sheet, missing column, bad date, unknown project reference - without stopping at the first one.", "Must", "Derived", "4"],
    [f"{MARK_CHG}REQ-IMP-03", "Import/Export", "A blank source workbook template with correct sheet names, headers, value lists and one example row is delivered with the application.", "Must", "Derived", "4"],
    ["REQ-IMP-04", "Import/Export", "The user can export current data back to .xlsx, preserving the template layout so the export can be re-imported.", "Must", "Derived", "4"],
    [f"{MARK_CHG}REQ-IMP-05", "Import/Export", "The application records which file was loaded and when, and shows it on screen with its time zone - a bare timestamp is ambiguous to anyone reading it from another country.", "Should", "Derived / G-02", "4"],
    ["REQ-IMP-06", "Import/Export", "Loaded data may be cached in the browser so re-opening the page does not force a re-import; the cache never replaces the workbook as the record.", "Could", "Derived", "4"],
    [f"{MARK_NEW}REQ-IMP-07", "Import/Export", "After import, the application lets the user update the data on screen, and those updates are carried into the file produced on export.", "Must", "Reviewer v0.11", "4"],
    [f"{MARK_NEW}REQ-IMP-08", "Import/Export", "Unsaved edits are visible as such, and the user is warned before any action that would discard them (closing the page, loading another file).", "Must", "Derived from REQ-IMP-07", "4"],
    [f"{MARK_CHG}REQ-IMP-09", "Import/Export", "Every field is editable, including identifiers. An on-screen edit is re-validated against the same rules as an imported value, so editing cannot introduce data the import would have rejected.", "Must", "Q-20", "4"],
    [f"{MARK_NEW}REQ-IMP-10", "Import/Export", "Editing an identifier that other sheets reference cascades to every referencing row, after showing how many rows will change. Deleting a referenced row is refused, naming what still points at it.", "Must", "Q-20", "4"],
    [f"{MARK_NEW}REQ-IMP-11", "Import/Export", "A new row can be inserted into any editable table, positioned immediately below the row the user acts on rather than appended at the end, and validated on entry like any other edit.", "Must", "P-01, S-01", "4"],

    ["REQ-VC-01", "Version control", "Development plan, programming specification and output files are version-controlled together and their versions are cross-referenced.", "Must", "Requester", "1-5"],
    ["REQ-VC-02", "Version control", "The application displays its own version, and the version of the source data schema it expects.", "Must", "Requester", "4"],
    ["REQ-VC-03", "Version control", "Loading a source file whose schema version is newer or older than the application expects produces a clear warning.", "Should", "Derived", "4"],
    ["REQ-VC-04", "Version control", "Every document re-issue adds a version-history row stating what changed and why.", "Must", "Requester", "1-5"],

    ["REQ-NFR-01", "Non-functional", "The application is built so later requirements can be added without restructuring: parsing, calculation and presentation are separated.", "Must", "Requester", "4"],
    ["REQ-NFR-02", "Non-functional", "Works in Microsoft Edge and Google Chrome on Windows 10/11, offline.", "Must", "Q-05", "4"],
    [f"{MARK_CHG}REQ-NFR-03", "Non-functional", "Handles 100 projects and 1,000 people, with the assignments that implies (order 8,000) over a 60-month horizon. Tables of that height are virtualised and the per-person chart aggregates rather than drawing a bar each - see REQ-DSH-09.", "Must", "S2-06", "4"],
    ["REQ-NFR-04", "Non-functional", "No data leaves the PC: no network calls, no external CDN, no telemetry.", "Must", "Derived", "4"],
    ["REQ-NFR-05", "Non-functional", "Dates are handled unambiguously (ISO yyyy-mm-dd internally) regardless of Windows regional settings.", "Must", "Derived", "4"],
]

r_start = r
r = table(ws, r, ["REQ-ID", "Category", "Requirement", "Priority", "Source", "Step"],
          reqs, [15, 15, 92, 10, 18, 8], wrap_cols=(3,), mark_col=1)
last = r_start + len(reqs)

dv = DataValidation(type="list", formula1='"Must,Should,Could,Won\'t"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"D{r_start + 1}:D{last}")

r = section(ws, r, "Requirement count by priority")
counts = [
    ["Must", f'=COUNTIF(D{r_start + 1}:D{last},"Must")'],
    ["Should", f'=COUNTIF(D{r_start + 1}:D{last},"Should")'],
    ["Could", f'=COUNTIF(D{r_start + 1}:D{last},"Could")'],
    ["Total", f"=COUNTA(A{r_start + 1}:A{last})"],
]
r = table(ws, r, ["Priority", "Count"], counts, [15, 12])
r = note(ws, r, "47 requirements at v0.1, 58 at v0.2, 61 at v0.3, 63 at v0.4 and v1.0. v1.1 adds 2 and rewords 3, "
                "as a change against the approved baseline.")
r = note(ws, r, "'Source' = Requester (in the original request), a Q-number (answered at v0.11 review), "
                "'Reviewer v0.11' (added by the reviewer directly), Derived (engineering consequence), "
                "or a Decision ID from 07_Architecture.")
r = legend(ws, r)

# ---- 04 Data model --------------------------------------------------------
ws, r = sheet(wb, "04_Data_Model", "Data model - source Excel structure",
              "ONE workbook, PRAP_SourceData.xlsx, with the sheets below (changed at Q-09). "
              "This is the schema Step 2 will fix and the code will parse.")

r = section(ws, r, "Sheets in PRAP_SourceData.xlsx")
sheets_tbl = [
    ["Project", "One row per project.", "Master"],
    ["Milestone", "One row per project milestone.", "Child of Project"],
    [f"{MARK_NEW}ProjectPeriod", "The four standard periods per project, with their start/end and weight.", "Child of Project"],
    [f"{MARK_NEW}PeriodWeightStandard", "Default weight per project category x period. Seeds ProjectPeriod.", "Reference"],
    [f"{MARK_CHG}RoleFactor", "Role burden factor, per project type, clinical phase and period.", "Reference"],
    ["Person", "One row per person.", "Master"],
    ["Assignment", "One row per person + project + role.", "Link"],
    ["PersonPeriodWeight", "Optional time-varying override of the person weight.", "Child of Assignment"],
    [f"{MARK_NEW}Lists", "Value lists for every data-driven dropdown, so lists change without code.", "Reference"],
    ["Config", "Parameters: schema version, thresholds, horizon, FTE hours.", "Reference"],
]
r = table(ws, r, ["Sheet", "Purpose", "Kind"], sheets_tbl, [26, 84, 20], wrap_cols=(2,), mark_col=1)

r = section(ws, r, "Sheet: Project")
proj = [
    ["project_id", "Text", "Yes", "Unique key, e.g. PRJ-001. Referenced by every other sheet.", "REQ-PRJ-03"],
    ["project_name", "Text", "Yes", "Unique display name.", "REQ-PRJ-03"],
    [f"{MARK_CHG}project_type", "List", "Yes", "'NewDrug CT', 'Biosimilar CT (Healthy)', 'Biosimilar CT (Patient)' or 'Others'. Everything but 'Others' is a clinical trial. SCHEMA 6 split the single 'Biosimilar CT' in two: a healthy-volunteer study and a patient study are not the same workload, and one value could not say which. A project type is recognised as a trial by the START of its name, so a further subdivision is a change to the value list rather than to the code.", "REQ-PRJ-01, R-12"],
    ["project_category", "Text", "Conditional", "Product name. Required for either clinical trial type.", "REQ-PRJ-02"],
    [f"{MARK_NEW}clinical_phase", "List", "Conditional", "'Phase 1' / 'Phase 2' / 'Phase 3' / 'Phase 4'. Required for either clinical trial type.", "REQ-PRJ-09"],
    [f"{MARK_NEW}work_scope_type", "List", "Yes", "How much of the work is done in-house: 'fully in-housed' / 'fully outsourced' / 'Partially outsourced (in-house for EDC)', extensible on the Lists sheet. SCHEMA 6. Part of the key into PeriodWeightStandard and RoleFactor - a trial run entirely in-house costs this team more than the same trial handed to a CRO, and the weights are where that belongs.", "R-12"],
    [f"{MARK_CHG}outsourcing_type", "List", "Yes", "'Full outsourcing' / 'Partial outsourcing' / 'Full In-house'. Three values, fixed at Q-14. DESCRIPTIVE ONLY since schema 6 - work_scope_type is what the weights are keyed on. The two are checked against each other (V-25) but not merged: merging them would rewrite a column in every existing file.", "REQ-PRJ-04, R-12"],
    [f"{MARK_NEW}EDC_setup", "List", "Conditional", "Who sets up the EDC system. 'by CRO' / 'by SB'. Required for either clinical trial type.", "REQ-PRJ-10"],
    [f"{MARK_NEW}DataReviewSystem_setup", "List", "Conditional", "Who sets up the data review system. 'by CRO' / 'by SB'.", "REQ-PRJ-10"],
    [f"{MARK_NEW}RBQM_setup", "List", "Conditional", "Who sets up the RBQM system. 'by CRO' / 'by SB'.", "REQ-PRJ-10"],
    [f"{MARK_NEW}DM_conduct", "List", "Conditional", "Who reviews the clinical trial data. 'by CRO' / 'by SB'.", "REQ-PRJ-10"],
    [f"{MARK_NEW}EDC_system", "List", "Conditional", "EDC system name, e.g. 'Veeva EDC' / 'Rave' / 'eSOURCE'. Data-driven list.", "REQ-PRJ-11"],
    [f"{MARK_NEW}DataReviewSystem", "List", "Conditional", "e.g. 'Veeva DQS' / 'Medidata CDS' / 'No system (manual)'. Data-driven list.", "REQ-PRJ-11"],
    [f"{MARK_NEW}RBQM_system", "List", "Conditional", "e.g. 'CluePoints' / 'Medidata CDS' / 'No system (manual)'. Data-driven list.", "REQ-PRJ-11"],
    ["planned_member_count", "Integer", "No", "Planned number of project members; compared against actual assignments.", "REQ-PRJ-04"],
    ["start_date", "Date", "Yes", "Project start.", "REQ-PRJ-05"],
    ["end_date", "Date", "Yes", "Planned project end.", "REQ-PRJ-05"],
    ["total_period_months", "Derived", "-", "Calculated from start_date and end_date; not entered by hand.", "REQ-PRJ-08"],
    ["status", "List", "No", "Planned / Active / On hold / Completed. Drives default dashboard filtering.", "REQ-PRJ-07"],
    ["note_1 .. note_5", "Text", "No", "Free extension columns.", "REQ-PRJ-07"],
]
r = table(ws, r, ["Column", "Type", "Required", "Definition / rule", "REQ-ID"],
          proj, [26, 11, 12, 88, 14], wrap_cols=(4,), mark_col=1)
r = note(ws, r, "system_prepared_by is dropped at v0.3 (Q-19): the four *_setup columns replace it with more detail.")

r = section(ws, r, "Sheet: Milestone")
mile = [
    ["project_id", "Text", "Yes", "Foreign key to Project.", "REQ-PRJ-05"],
    ["project_name", "Derived", "No", "DERIVED - looked up from Project on import and refreshed from it, so a mismatch cannot corrupt the link. Locked in the template; a value typed here is discarded and the disagreement reported as V-13.", "REQ-PRJ-05"],
    [f"{MARK_CHG}milestone_name", "List", "Yes", "Standard list (10) at v1.1: 'Protocol (v1)', 'CTA submission', 'FPI', 'First SIV', 'LPI', 'interim DB lock cut-off', 'interim DB lock', 'final DB lock cut-off', 'final DB lock', 'Inspection'. FPI returns as the fallback for First SIV; Inspection is new and MAY REPEAT within a project. Held in Lists, not fixed in code.", "REQ-PRJ-05, REQ-PRJ-13"],
    ["milestone_date", "Date", "Yes", "Planned date.", "REQ-PRJ-05"],
    ["milestone_seq", "Integer", "No", "Display order on the timeline.", "REQ-PRJ-05"],
    [f"{MARK_NEW}note_1", "Text", "No", "Free extension column.", "REQ-PRJ-07"],
]
r = table(ws, r, ["Column", "Type", "Required", "Definition / rule", "REQ-ID"],
          mile, [26, 11, 12, 88, 14], wrap_cols=(4,), mark_col=1)
r = note(ws, r, "Boundary milestones at v1.1: Protocol (v1), First SIV (or FPI), CTA submission, the DB locks, and Inspection. Only LPI and the two cut-off dates remain pure markers. milestone_seq orders repeated Inspection rows; uniqueness is on (project_id, milestone_name, milestone_date), not on name alone.")

r = section(ws, r, "Sheet: ProjectPeriod  [was ProjectPeriodWeight]")
pp = [
    ["project_id", "Text", "Yes", "Foreign key to Project.", "REQ-PRJ-06"],
    [f"{MARK_CHG}period_name", "List", "Yes", "From the period set for the project's TYPE - see the two sets below. A period name from the wrong set is rejected. UNIQUE within a project since R-11, so (project_id, period_name) is the key.", "REQ-PRJ-12, REQ-CAL-11, R-11"],
    [f"{MARK_CHG}period_seq", "Integer", "Yes", "Orders the periods along the timeline. Unique within a project. No longer part of the key - it carries order, not identity (R-11).", "REQ-CAL-11"],
    [f"{MARK_CHG}period_start", "Date", "Yes", "Inclusive. For Clinical Trial, computed from milestones (sheet 05) and then editable. For Others, entered directly.", "REQ-CAL-09"],
    [f"{MARK_CHG}period_end", "Date", "Yes", "Inclusive. Periods for one project must not overlap and must leave no gap (REQ-CAL-10).", "REQ-CAL-09"],
    [f"{MARK_CHG}weight", "Decimal", "Yes", "Effort multiplier for this period. For a clinical trial, seeded from PeriodWeightStandard by clinical phase and then editable. For 'Others', entered by hand (Q-28). Values are data (Q-17), not fixed in this plan.", "REQ-PRJ-06"],
    [f"{MARK_NEW}note_1", "Text", "No", "Free extension column, e.g. why a derived date was overridden.", "REQ-PRJ-07"],
]
r = table(ws, r, ["Column", "Type", "Required", "Definition / rule", "REQ-ID"],
          pp, [26, 11, 12, 88, 14], wrap_cols=(4,), mark_col=1)

r = note(ws, r, "Period sets are type-specific (Q-18). The derivation itself is on sheet 05.")
sets = [
    [f"{MARK_CHG}NewDrug CT / Biosimilar CT", "Before-Start-up, Start-up, Conduct (interim), Close-out (interim), Conduct (final), Close-out (final), After Close-out (final)", "Seven names, in timeline order. Both trial types share one period set and one derivation; they differ in their WEIGHTS, not their shape. A trial with no interim DB lock simply omits 'Conduct (interim)' and 'Close-out (interim)'."],
    ["Others", "Planning, Develop, Close", "Entered directly - dates confirmed manual at Q-25, weights at Q-28. 'Others' projects are hand-entered throughout."],
]
r = table(ws, r, ["project_type", "Period set", "How boundaries are set"], sets, [18, 50, 66], wrap_cols=(2, 3), mark_col=1)

r = section(ws, r, "Sheet: PeriodWeightStandard")
pws = [
    [f"{MARK_CHG}project_type", "List", "Yes", "A clinical trial type. Part of the key: a biosimilar trial at a given phase is not the same workload as a new-drug trial at that phase. 'Others' take manual weights (Q-28).", "R-05"],
    [f"{MARK_CHG}clinical_phase", "List", "Yes", "The phase the standard applies to.", "Q-26"],
    [f"{MARK_NEW}work_scope_type", "List", "No", "The work scope this standard applies to. SCHEMA 6. EMPTY means the row applies to EVERY scope - see the note below, which is the important part of this change.", "R-12"],
    ["period_name", "List", "Yes", "A period from that type's set.", "Q-04, Q-18"],
    [f"{MARK_CHG}weight", "Decimal", "Yes", "Default multiplier. You fill these in the source workbook (Q-17); the plan fixes only where they live.", "Q-01"],
    [f"{MARK_NEW}note_1", "Text", "No", "Free extension column, e.g. the basis for the weight.", "Q-01"],
]
r = table(ws, r, ["Column", "Type", "Required", "Definition / rule", "Basis"],
          pws, [26, 11, 12, 88, 14], wrap_cols=(4,), mark_col=1)
r = note(ws, r, "SCHEMA 6 - THE EMPTY SCOPE. The key is (project_type, clinical_phase, work_scope_type, "
                "period_name) and it is unique. A row whose work_scope_type is EMPTY applies to every scope; a "
                "project looks for its own scope first and falls back to that row. Without the fallback this "
                "sheet would need 3 types x 4 phases x 7 periods x 3 scopes = 252 rows, of which two thirds would "
                "repeat their neighbour, and every one of them hand-entered. With it, 84 rows carry the baseline "
                "and a scope-specific row is added only where the scope really changes the number. It also means "
                "a schema 5 file keeps working: its rows have no scope, so they are the every-scope rows.")
r += 1
r = note(ws, r, "Re-keyed at Q-26: a clinical trial's weights come from its CLINICAL PHASE, not its product category. "
                "So a phase 1 trial and a phase 3 trial of the same product carry different period weights, which is "
                "the intended behaviour. Q-28 then confirmed that 'Others' projects are not distinguished at all and "
                "take their weights by hand - which makes this sheet a clinical-trial table, and leaves 'Others' "
                "projects entirely hand-entered: dates and weights alike.")

r = section(ws, r, "Sheet: RoleFactor   [key extended at R-10]")
rf = [
    [f"{MARK_CHG}project_type", "List", "Yes", "'NewDrug CT', 'Biosimilar CT (Healthy)', 'Biosimilar CT (Patient)' or 'Others'. Roles are keyed by type, so a factor can differ between the trial types.", "Q-03, R-05, R-12"],
    [f"{MARK_NEW}clinical_phase", "List", "Trials only", "The phase this factor applies to. EMPTY on 'Others' rows, which carry no phase.", "R-10"],
    [f"{MARK_NEW}work_scope_type", "List", "No", "The work scope this factor applies to. SCHEMA 6. EMPTY means EVERY scope, exactly as on PeriodWeightStandard - and it matters more here, because this is the larger sheet.", "R-12"],
    [f"{MARK_CHG}period_name", "Text", "Yes", "The period this factor applies to. One of the seven clinical periods, or one of the three 'Others' periods, matching the project_type.", "R-10, R-11"],
    ["role_name", "Text", "Yes", "Clinical Trial: 'Project oversight', 'Lead data manager', 'Clinical Data Associator', 'Clinical Database Programmer', 'Data Analyst'. Others: 'Project lead', 'Main staff', 'Other staff'.", "Q-03"],
    [f"{MARK_CHG}role_factor", "Decimal", "Yes", "Relative burden of THIS role in THIS period, the same for everyone holding the role. A role's share of the work is not flat across a project: the database programmer is heaviest while the database is built, the data associator while data arrives, the analyst at lock.", "Q-01, R-10"],
    ["role_note", "Text", "No", "Basis for the factor.", "REQ-CAL-06"],
]
r = table(ws, r, ["Column", "Type", "Required", "Definition / rule", "Basis"],
          rf, [26, 11, 12, 88, 14], wrap_cols=(4,), mark_col=1)
r = note(ws, r, "The key is (project_type, clinical_phase, work_scope_type, period_name, role_name) since "
                "schema 6, which makes this comfortably the largest sheet in the workbook: 3 types x 4 phases x 7 "
                "periods x 5 roles = 420 baseline rows, plus 9 for 'Others'. Filling all three scopes as well "
                "would make it 1,269. THAT is why the empty scope exists: fill the baseline once, and add a "
                "scope-specific row only where the scope really changes the number. The cost of the change still "
                "falls on whoever maintains the file, and the fallback is what keeps it payable.")
r += 1
r = note(ws, r, "NOTE A CONSEQUENCE. PeriodWeightStandard and RoleFactor now both vary over (project_type, "
                "clinical_phase, period_name), and the calculation multiplies them together. The pair is "
                "therefore mathematically collapsible into one table keyed on all four dimensions. They are "
                "kept separate deliberately - one says how busy the PROJECT is in a period, the other how much "
                "of that falls on a ROLE - but the separation only holds if they are maintained that way. "
                "Raising a whole project's Conduct load belongs in PeriodWeightStandard, once; doing it in "
                "RoleFactor means five edits and double-counts the next time the period weight moves.")

r = section(ws, r, "Sheet: Person")
pers = [
    ["person_id", "Text", "Yes", "Unique key, e.g. PSN-001.", "REQ-PSN-01"],
    ["person_name", "Text", "Yes", "Display name.", "REQ-PSN-01"],
    ["department", "Text", "No", "Grouping for the dashboard.", "REQ-PSN-06"],
    ["primary_role", "Text", "No", "Usual role; an assignment can override it.", "REQ-PSN-03"],
    [f"{MARK_CHG}capacity_fte", "Decimal", "No", "Available capacity in FTE, default 1.00 (= 160 h/month). Lower it for a part-timer.", "REQ-CAL-08"],
    ["employment_start / employment_end", "Date", "No", "Bounds availability; blank end = open.", "REQ-PSN-04"],
    ["note_1 .. note_5", "Text", "No", "Free extension columns.", "REQ-PSN-06"],
]
r = table(ws, r, ["Column", "Type", "Required", "Definition / rule", "REQ-ID"],
          pers, [26, 11, 12, 88, 14], wrap_cols=(4,), mark_col=1)

r = section(ws, r, "Sheet: Assignment")
asg = [
    ["assignment_id", "Text", "Yes", "Unique key. One row per person + project + role.", "REQ-PSN-07"],
    ["person_id", "Text", "Yes", "Foreign key to Person.", "REQ-PSN-01"],
    ["person_name", "Derived", "No", "DERIVED - looked up from Person on import and refreshed from it. Locked in the template; a value typed here is discarded and the disagreement reported as V-13.", "REQ-PSN-01"],
    ["project_id", "Text", "Yes", "Foreign key to Project.", "REQ-PSN-02"],
    [f"{MARK_CHG}role_name", "Text", "Yes", "Foreign key to RoleFactor, matched on (project's type, role_name); the phase and period are supplied by the project and the month being calculated. Several rows = several roles on one project.", "REQ-PSN-03"],
    [f"{MARK_CHG}assign_start_date", "Date", "No", "Date the person joins the study. BLANK = the project's own start date (REQ-CAL-15). Optional since v2.29: leave both dates blank for somebody on the project throughout, and fill them in only for a partial involvement.", "REQ-PSN-04, REQ-CAL-15"],
    [f"{MARK_CHG}assign_end_date", "Date", "No", "Date the person leaves the study. BLANK = the project's own end date. This one has always behaved that way; v2.29 makes the start match it.", "REQ-PSN-04, REQ-CAL-15"],
    [f"{MARK_CHG}person_weight", "Decimal", "Yes", "RENAMED from base_allocation. How much this person works on this project, e.g. 0.40. Q-01 folded the two former fields into this one.", "REQ-CAL-02"],
    ["note_1 .. note_3", "Text", "No", "Free extension columns.", "REQ-PSN-06"],
]
r = table(ws, r, ["Column", "Type", "Required", "Definition / rule", "REQ-ID"],
          asg, [26, 11, 12, 88, 14], wrap_cols=(4,), mark_col=1)

r = section(ws, r, "Sheet: PersonPeriodWeight")
ppw = [
    ["assignment_id", "Text", "Yes", "Foreign key to Assignment.", "REQ-PSN-05"],
    ["period_start", "Date", "Yes", "Inclusive. Part of the key: one assignment may carry SEVERAL non-overlapping windows, so assignment_id alone does not identify a row.", "REQ-PSN-05"],
    ["period_end", "Date", "Yes", "Inclusive. Periods within one assignment must not overlap.", "REQ-PSN-05"],
    [f"{MARK_CHG}weight_override", "Decimal", "Yes", "REPLACES person_weight for months inside this window - it no longer multiplies it. Where no row covers a month, person_weight applies unchanged.", "REQ-PSN-05"],
    ["reason", "Text", "No", "Why the weight differs, e.g. 'part-time', 'covering start-up peak'.", "REQ-PSN-05"],
]
r = table(ws, r, ["Column", "Type", "Required", "Definition / rule", "REQ-ID"],
          ppw, [26, 11, 12, 88, 14], wrap_cols=(4,), mark_col=1)
r = note(ws, r, "Changed at Q-01. The answer named exactly three factors, so a fourth multiplier would have been an "
                "invention. This sheet now overrides rather than scales - the same effect, one fewer factor, and no "
                "hidden compounding. It is optional: leave it empty and every assignment simply uses person_weight.")

r = section(ws, r, "Sheet: Config")
cfg = [
    [f"{MARK_CHG}schema_version", "Version of this workbook structure; checked on import. Steps to 3 at v1.4, 4 at v1.7, 5 at v1.8, 6 at v2.27 (R-12).", "6", "REQ-VC-02"],
    [f"{MARK_NEW}split_shared_role_fte", "1 = where several people hold the same role on one project in a month, the role factor is divided between them (REQ-CAL-14). 0 = each carries the whole factor, which is how every version before v2.28 behaved. A setting rather than a constant because it changes every figure a shared role ever produced, and somebody comparing this month's report with last year's has to be able to see where the difference came from.", "1", "REQ-CAL-14"],
    [f"{MARK_NEW}fte_hours_per_month", "Hours equal to 1.00 FTE.", "160", "REQ-CAL-08"],
    [f"{MARK_NEW}over_allocation_fte", "Person-month total above this is over-allocated.", "1.50", "REQ-CAL-04"],
    [f"{MARK_CHG}under_allocation_fte", "Person-month total below this counts toward an under-allocated run. Absolute, not scaled by capacity.", "0.60", "REQ-CAL-07"],
    [f"{MARK_NEW}under_allocation_min_months", "Consecutive months below the threshold before flagging.", "3", "REQ-CAL-07"],
    [f"{MARK_CHG}default_horizon_months", "Months shown on opening.", "24", "REQ-CAL-01"],
    [f"{MARK_CHG}capacity_unit", "Display unit: 'FTE' or 'percent'.", "FTE", "REQ-CAL-08"],
]
r = table(ws, r, ["parameter", "Meaning", "Default value", "REQ-ID"],
          cfg, [30, 74, 16, 14], wrap_cols=(2,), mark_col=1)

r = section(ws, r, "Sheet: Lists   [new]")
r = lines(ws, r, [
    "Two columns - list_name, value - holding every data-driven dropdown: outsourcing_type, clinical_phase,",
    "milestone_name, period_name, EDC_system, DataReviewSystem, RBQM_system, setup_party, project_status.",
    "Adding a permitted value is a data edit, never a code change (REQ-CAL-06). A third column, note_1,",
    "carries free text about a value - when it was added, or what it means.",
])
r += 1

r = section(ws, r, "Referential rules checked on import")
rules = [
    ["V-01", "Every Assignment.project_id exists in Project.", "Error - row rejected, reported with its row number."],
    ["V-02", "Every Assignment.person_id exists in Person.", "Error - row rejected."],
    [f"{MARK_CHG}V-03", "Every Assignment.role_name appears in RoleFactor for that project's type.", "Error - row rejected. Was a warning in v0.1; roles are now type-specific, so a mismatch is a real error."],
    [f"{MARK_NEW}V-24", "Every PersonPeriodWeight.assignment_id exists in Assignment, and (assignment_id, period_start) is unique.", "Error - an override on an assignment that does not exist is silently ignored, so the weight the user typed never applies and nothing says so."],
    [f"{MARK_NEW}V-25", "A project's work_scope_type does not contradict its outsourcing_type. The two unambiguous ends are checked: 'Full outsourcing' against 'fully outsourced', and 'Full In-house' against 'fully in-housed'.", "Warning - two columns on the same axis and only one of them drives the weights. A project marked 'Full In-house' and 'fully outsourced' is a slip, and the calculation would follow the field nobody was looking at. 'Partial outsourcing' is compatible with more than one scope and is not checked: a warning nobody can act on is a warning everybody learns to ignore."],
    [f"{MARK_NEW}V-26", "No project carries a project_type that schema 6 retired - at present 'Biosimilar CT', which became 'Biosimilar CT (Healthy)' and 'Biosimilar CT (Patient)'.", "Error - reported as itself rather than as a generic unknown value, because the remedy is a choice the file cannot make on the user's behalf. Only they know whether the trial ran in healthy volunteers or in patients, and guessing would put a wrong weight on real work."],
    [f"{MARK_NEW}V-23", "For every (project_type, clinical_phase, period_name, role_name) an assignment can actually reach, a RoleFactor row exists.", "Error - a factor missing for ONE period of a project silently drops that stretch to 1.00, which is a wrong number rather than an obvious gap. Checked against the periods each assignment spans, not against the whole cross-product."],
    ["V-04", "project_category is present for either clinical trial type.", "Warning - shown as blank in the dashboard."],
    ["V-05", "end_date is on or after start_date, for projects, assignments and all weight periods.", "Error - row rejected."],
    [f"{MARK_CHG}V-06", "Periods within one project, and override windows within one assignment, do not overlap.", "Error - overlapping pair reported. The assignment half was specified from v1.0 but not implemented until R-12; an overlap there makes the applied weight depend on row order."],
    ["V-07", "Assignment dates fall inside the project's own start and end dates.", "Warning - kept, but listed for review."],
    ["V-08", "project_id, person_id and assignment_id are unique in their sheet.", "Error - duplicate rejected."],
    ["V-09", "schema_version in Config matches the version the application expects.", "Warning - proceeds, banner shown."],
    [f"{MARK_NEW}V-10", "Clinical-trial projects carry clinical_phase and the four *_setup values.", "Warning - the project still simulates; the gap is listed."],
    [f"{MARK_NEW}V-11", "Every list-typed value appears in the Lists sheet for its list.", "Warning - value kept, reported as unrecognised."],
    [f"{MARK_CHG}V-12", "A project's periods leave no gap and no overlap across its timeline. The full set need not be present - a period may be legitimately omitted (REQ-CAL-12).", "Warning - months with no period use weight 1.00 and are listed (REQ-CAL-10)."],
    [f"{MARK_NEW}V-13", "Denormalised project_name / person_name match their master row.", "Warning - the master value wins and the copy is refreshed."],
    [f"{MARK_CHG}V-14", "A milestone date falls inside its project's start..end window, and the boundary milestones appear in chronological order. Repeated 'Inspection' rows are exempt from the uniqueness part of this check.", "Error for period-defining milestones - the derivation cannot run. Warning for markers."],
    [f"{MARK_NEW}V-15", "A period_name belongs to the period set of its project's type.", "Error - a 'Planning' period on a clinical trial is a category mistake, not a typo."],
    [f"{MARK_NEW}V-16", "A clinical trial carries the milestones the derivation needs: CTA submission, and at least one DB lock.", "Error - without them no period boundary can be computed."],
    [f"{MARK_CHG}V-18", "Within a project, period_name is unique, and period_seq is unique.", "Error - since R-11 the key is (project_id, period_name), so a repeated name is a duplicate key rather than a legitimate second stretch. period_seq stays unique because it carries the order."],
    [f"{MARK_CHG}V-19", "A clinical trial carries a clinical_phase, and PeriodWeightStandard has rows for that phase. Not applied to 'Others' projects, whose weights are entered directly.", "Error - since Q-26 the phase selects the weights, so a missing phase leaves the project unweighted."],
    [f"{MARK_NEW}V-20", "A milestone other than 'Inspection' appears at most once per project.", "Warning - repeats are allowed but usually a data-entry slip outside Inspection."],
    [f"{MARK_NEW}V-21", "'Inspection' dates on or before the final DB lock are treated as markers, not as the start of period 7.", "Information - listed so the reader can see why an early inspection did not open a period."],
    [f"{MARK_NEW}V-22", "No person carries a capacity_fte below the under-allocation floor.", "Warning - such a person can never clear the floor however fully they are booked, so the flag would be permanent and meaningless."],
    [f"{MARK_NEW}V-17", "On editing an identifier, every referencing row is updated; on deleting a row, nothing may still reference it.", "Cascade after confirmation / deletion refused (REQ-IMP-10)."],
]
r = table(ws, r, ["ID", "Rule", "On failure"], rules, [9, 84, 60], wrap_cols=(2, 3), mark_col=1)
r = note(ws, r, "Import never stops at the first problem: it collects every finding and presents one report (REQ-IMP-02). "
                "Under REQ-IMP-09 the same rules run again on any value edited on screen.")
r = legend(ws, r)

# ---- 05 Resource logic ----------------------------------------------------
ws, r = sheet(wb, "05_Resource_Logic", "Resource calculation logic",
              "Rewritten at v0.2 from the Q-01, Q-02 and Q-08 answers. Step 2 will fix this as pseudocode "
              "plus the worked example as an acceptance test.")

r = section(ws, r, "Monthly load for one assignment, in one month")
r = lines(ws, r, [
    "    load(assignment, month) = project_period_weight(project, month)",
    "                            x role_factor(project_type, clinical_phase, period, role)",
    "                            x person_weight(assignment, month)",
    "                            x coverage(assignment, month)",
    "",
    "The result is FTE, where 1.00 FTE = 160 hours per month (8 h/day x 5 days/week x 20 days/month).",
    "",
    "project_period_weight  is set by the project's category for the milestone period the month falls in.",
    "                       It is identical for everyone assigned to that project in that period.",
    "role_factor            is the standard burden of the role IN THE PERIOD THE MONTH FALLS IN,",
    "                       identical for everyone holding that role on that kind of project (R-10).",
    "                       Roles are drawn from the list valid for the project's type. Where no row",
    "                       matches, the factor falls back to 1.00 and V-23 reports it.",
    "person_weight          is how much this person works on this project, given per assignment;",
    "                       a PersonPeriodWeight row replaces it for the months it covers.",
    "coverage               = calendar days worked in the month / calendar days in the month.",
    "                       1.00 for a full month, a fraction in the joining and leaving months.",
], mono=True)
r += 1
r = note(ws, r, "Changed from v0.1: base_allocation is gone. Q-01 named three factors, and base_allocation was a fourth "
                "that would have silently scaled all of them. person_weight now carries that meaning alone.")
r += 1

r = section(ws, r, "Aggregation")
agg = [
    ["Per project, per month", "Sum of load over every assignment on that project.", "Overall tab, project table + stacked graph", "REQ-CAL-03"],
    ["Per person, per month", "Sum of load over every assignment that person holds, across all projects.", "Overall tab, person table + graph", "REQ-CAL-03"],
    ["Per person, per project, per month", "The individual load value.", "Drill-down when a person row is expanded", "REQ-CAL-03"],
    [f"{MARK_CHG}Over-allocation flag", "Person-month total > over_allocation_fte (1.50).", "Red cell + count in summary tiles", "REQ-CAL-04"],
    [f"{MARK_NEW}Under-allocation run", "Person-month total < under_allocation_fte (0.80) for >= under_allocation_min_months (3) consecutive months. Reported as a run with its start month and length.", "Amber cells across the run + count in summary tiles", "REQ-CAL-07"],
    [f"{MARK_NEW}Hours view", "load x fte_hours_per_month.", "Alternative display unit", "REQ-CAL-08"],
    ["Head-count per project, per month", "Count of assignments with load > 0 that month.", "Compared against planned_member_count", "REQ-PRJ-04"],
]
r = table(ws, r, ["Output", "How it is computed", "Where it appears", "REQ-ID"],
          agg, [30, 66, 44, 14], wrap_cols=(2, 3), mark_col=1)

r = note(ws, r, "Under-allocation needs a rolling window, not a per-month test: a single quiet month is normal, three "
                "in a row is a signal. A run is reported once, at its start, rather than as three separate flags.")
r += 1

r = section(ws, r, "Period derivation for a clinical trial - either type   [v1.1]")
r = lines(ws, r, [
    "Boundaries come from milestone dates plus month offsets, so moving a milestone re-shapes the periods.",
    "Where a boundary has both a milestone and an offset definition, the RECORDED MILESTONE WINS and the",
    "offset is the fallback (REQ-CAL-13). Boundary milestones: Protocol (v1), CTA submission, First SIV",
    "(or FPI), the DB locks, and Inspection. LPI and the two cut-off dates are markers only.",
])
r += 1
deriv = [
    [f"{MARK_CHG}1", "Before-Start-up", "project start_date", "'Protocol (v1)' where recorded, otherwise the day before Start-up begins", "R-02"],
    [f"{MARK_CHG}2", "Start-up", "day after 'Protocol (v1)' where recorded, otherwise one month before 'CTA submission'", "'First SIV' where recorded (or 'FPI'), otherwise start + 4 months - 1 day", "R-02"],
    [f"{MARK_CHG}3", "Conduct (interim)", "day after Start-up ends", "day before Close-out (interim) begins", "Q-22, R-11"],
    ["4", "Close-out (interim)", "3 months before 'interim DB lock'", "'interim DB lock'", "Q-22, Q-23"],
    [f"{MARK_CHG}5", "Conduct (final)", "day after 'interim DB lock'", "day before Close-out (final) begins", "Q-23, R-11"],
    [f"{MARK_CHG}6", "Close-out (final)", "3 months before 'final DB lock'", "day before period 7 begins; where there is no period 7, 'final DB lock' or project end_date if later", "R-02"],
    [f"{MARK_NEW}7", "After Close-out (final)", "earliest 'Inspection' date that falls after the final DB lock", "latest 'Inspection' date, or project end_date if later", "R-02"],
]
r = table(ws, r, ["Seq", "Period", "Starts", "Ends", "Basis"],
          deriv, [6, 24, 42, 50, 10], wrap_cols=(3, 4), mark_col=1)
r = note(ws, r, "Rows 3 and 4 exist ONLY where the project has an interim DB lock. Where there is none, rows 3 "
                "and 4 are omitted and row 5 - 'Conduct (final)' - runs from the day after Start-up ends. That is "
                "what makes the names unique: 'Conduct (interim)' exists only when there is an interim lock to be "
                "interim to, and every other conduct stretch is 'Conduct (final)' (R-11).")
r += 1
r = note(ws, r, "Row 7 exists only where an "
                "'Inspection' milestone falls after the final DB lock - 'Inspection' may be recorded several times, "
                "and the period spans from the first such event to the last.")
r += 1
r = note(ws, r, "Note the interaction at rows 1 and 2: where 'Protocol (v1)' is recorded, Before-Start-up ends on it "
                "and Start-up begins the day after, so the two rules meet exactly and no gap can open between them.")
r += 1

r = section(ws, r, "Degenerate timelines   [REQ-CAL-12]")
r = lines(ws, r, [
    "Real timelines do not always leave room for every period. Boundaries are applied in sequence order and a",
    "later one wins, so the periods stay contiguous; any period squeezed to zero or negative length is omitted",
    "and reported rather than written with a nonsensical range.",
])
r += 1
degen = [
    ["CTA submission is within a month of the project start", "Before-Start-up would be empty or negative.", "Before-Start-up omitted; Start-up begins at the project start date."],
    ["Interim and final DB lock less than 3 months apart", "'Conduct (final)' would be negative, and Close-out (final) would start before the interim lock.", "'Conduct (final)' omitted; Close-out (final) begins the day after the interim DB lock. The project then carries 'Conduct (interim)' but no 'Conduct (final)' - legitimate, since the names describe position, not a required pair."],
    ["No interim DB lock recorded", "Rows 3 and 4 have no anchor.", "Both omitted; a single 'Conduct (final)' runs from Start-up into Close-out (final)."],
    ["Trial too short for a conduct phase", "Start-up's fixed 4 months would overrun the Close-out start.", "The conduct period is omitted; Start-up is clipped where Close-out begins."],
    ["No CTA submission, or no DB lock at all", "Nothing to derive from.", "Error, not a silent default - V-16. The project's periods must then be entered by hand."],
    ["'Inspection' dated on or before the final DB lock", "Period 7 would start before period 6, overlapping the conduct phase.", "Those inspections are treated as markers inside the existing periods; only later inspection activity opens period 7. Reported by V-21. See R-03."],
    ["'Protocol (v1)' earlier than the project start date", "Before-Start-up would be empty or negative.", "Before-Start-up omitted; Start-up begins at the project start date."],
    ["'First SIV' earlier than Start-up begins", "Start-up would end before it starts.", "The offset fallback is used instead - start + 4 months - 1 day - and the inconsistency is reported."],
]
r = table(ws, r, ["Timeline", "What would go wrong", "What happens instead"],
          degen, [40, 52, 54], wrap_cols=(1, 2, 3))
r = note(ws, r, "All eight cases were run through the derivation, together with a full timeline carrying protocol, "
                "SIV, interim lock and three inspections. Every one produces a contiguous set of periods with no gap "
                "and no overlap. The check is worth repeating as an automated test at Step 4.")
r += 1

r = section(ws, r, "Worked example - the acceptance test for the calculation engine")
ex = [
    ["Assignment", "PSN-001 on PRJ-001 (Clinical Trial), role = Lead data manager", ""],
    ["person_weight", "0.40", "How much this person works on this project"],
    ["Assignment window", "2026-03-10 to 2026-12-31", "Joins part-way through March"],
    ["Project periods", "Start-up 2026-01-01..2026-06-30 weight 1.50; Conduct (final) from 2026-07-01 weight 1.00", "Seeded from the project's CLINICAL PHASE (Q-26)"],
    ["role_factor", "1.20 for Lead data manager", "Illustrative. Real values are entered in the source workbook (Q-17), not fixed here"],
    ["Coverage, March 2026", "22 / 31 = 0.7097", "10 Mar to 31 Mar inclusive = 22 days"],
    ["Load, March 2026", "1.50 x 1.20 x 0.40 x 0.7097 = 0.511 FTE", "81.8 hours"],
    ["Load, April 2026", "1.50 x 1.20 x 0.40 x 1.0000 = 0.720 FTE", "115.2 hours. Full month inside Start-up"],
    ["Load, July 2026", "1.00 x 1.20 x 0.40 x 1.0000 = 0.480 FTE", "76.8 hours. The Conduct (final) weight applies"],
    ["Over-allocated?", "No - 0.720 is below the 1.50 threshold", "This person would need other assignments to breach it"],
    ["Under-allocated?", "Only if their total across ALL projects stays below 0.80 for 3 months running", "Judged on the person total, never on one assignment"],
]
r = table(ws, r, ["Element", "Value", "Note"], ex, [26, 66, 50], wrap_cols=(2, 3))
r = note(ws, r, "Please confirm these numbers at Step 2. The role_factor of 1.20 is a placeholder: the arithmetic is "
                "the thing being agreed, not that particular value.")
r += 1

r = section(ws, r, "Design decisions inside the calculation")
dec = [
    ["C-01", "Weights multiply rather than add.", "Confirmed at Q-01: project weight x role factor x person weight.", "CONFIRMED"],
    ["C-02", "Partial months are pro-rated by calendar days.", "Confirmed at Q-02: work is expected in proportion to calendar days from the joined date to the end date.", "CONFIRMED"],
    [f"{MARK_CHG}C-03", "A missing PersonPeriodWeight row means person_weight applies unchanged.", "Follows from the override semantics. No row is the normal case, not an exception.", "CONFIRMED"],
    [f"{MARK_CHG}C-04", "Capacity is FTE. 1.00 FTE = 160 h/month.", "Confirmed at Q-08. Note the flag is NOT at 1.00 - over-allocation begins above 1.50.", "CONFIRMED"],
    [f"{MARK_CHG}C-05", "The unit shown on screen is FTE, with hours available.", "Q-08 was expressed in FTE and hours, so FTE is now the default rather than percent.", "CONFIRMED"],
    [f"{MARK_NEW}C-06", "Over- and under-allocation are judged on the person's total across all projects.", "The question 'is this person over-committed' is only meaningful across their whole workload.", "Confirm"],
    [f"{MARK_NEW}C-07", "Period weights are seeded from the category standard, then editable per project.", "Q-01 ties weight to category; real projects still deviate. Seeding keeps them consistent without freezing them.", "Confirm"],
    [f"{MARK_NEW}C-08", "A month is attributed to the period containing its first day.", "Avoids splitting a month across two weights. Simpler to explain and to check by hand.", "Confirm"],
    [f"{MARK_NEW}C-09", "Editing an identifier cascades to every row referencing it; deleting a referenced row is refused.", "Q-20 asked for every field to be editable with rules preventing inconsistency. Cascading is what makes an editable ID safe - blocking the edit instead would make the field read-only in all but name.", "Confirm"],
    [f"{MARK_NEW}C-10", "Period boundaries are recomputed whenever a milestone date changes, unless the period dates were edited by hand.", "Keeps derived dates true to the timeline - the frequent-change problem this tool exists for - without silently discarding a deliberate manual override.", "Confirm"],
    [f"{MARK_NEW}C-11", "Derived boundaries are applied in sequence order, a later one overriding an earlier; a period squeezed to nothing is omitted.", "The only rule that keeps periods contiguous on every timeline, including the awkward ones. Verified against five cases on this sheet.", "Confirm"],
]
r = table(ws, r, ["ID", "Decision", "Rationale", "Status"], dec, [9, 46, 76, 14], wrap_cols=(2, 3), mark_col=1)
r = legend(ws, r)

# ---- 06 Dashboard ---------------------------------------------------------
ws, r = sheet(wb, "06_Dashboard_Design", "Dashboard design",
              "Four tabs: the three specified, plus 'General assumptions' added at the component-list "
              "review (R-09). Detailed layout is fixed at the Step 3 UI review.")

r = section(ws, r, "Tab 1 - Overall")
overall = [
    [f"{MARK_CHG}Control bar", "Horizon (default 24 months, one click to expand to the latest project end date), project type, clinical phase, project, person, role, department, and one action to reset them all. Plus 'Load workbook', 'Export', and the loaded-file name and time with its time zone. The FTE/hours unit is NOT here - it is a setting, and it lives on the assumptions tab.", "REQ-DSH-05, REQ-DSH-07, REQ-IMP-05"],
    ["Table A - resource by project", "Rows = project, columns = months, cells = monthly FTE. Row and column totals. Expanding a project reveals its assigned people.", "REQ-DSH-01"],
    [f"{MARK_CHG}Table B - resource by person", "Rows = person, columns = months, cells = FTE summed across projects. Over-allocated cells red, under-allocated runs amber. Expanding a person reveals their projects.", "REQ-DSH-01, REQ-DSH-08"],
    ["Graph 1 - stacked area or bar, by project", "Total monthly demand, one band per project. Shows how demand builds and where the peaks land.", "REQ-DSH-02"],
    [f"{MARK_CHG}Graph 2 - grouped bar, by person", "Monthly FTE per person with reference lines at 1.50 and 0.80, so both thresholds are visible.", "REQ-DSH-02, REQ-DSH-08"],
    ["Graph 3 - project timeline (Gantt-style)", "One bar per project across the horizon, with milestone markers and period-weight shading.", "REQ-DSH-02, REQ-PRJ-05"],
    [f"{MARK_CHG}Summary tiles", "Active projects, people assigned, total FTE in the horizon, over-allocated person-months, under-allocated runs.", "REQ-DSH-08"],
]
r = table(ws, r, ["Component", "Content", "REQ-ID"], overall, [34, 96, 22], wrap_cols=(2,), mark_col=1)

r = section(ws, r, "Tab 2 - Source data (project)")
p2 = [
    [f"{MARK_CHG}Project table", "Every Project column from sheet 04, sortable and filterable, with derived total period. Editable under REQ-IMP-07.", "REQ-DSH-03"],
    ["Milestone sub-table", "Milestones of the selected project, in sequence.", "REQ-PRJ-05"],
    [f"{MARK_CHG}Period sub-table", "The selected project's four periods, their derived dates and their weights.", "REQ-PRJ-06"],
    ["Export", "Export the visible table to .xlsx.", "REQ-DSH-06"],
]
r = table(ws, r, ["Component", "Content", "REQ-ID"], p2, [34, 96, 22], wrap_cols=(2,), mark_col=1)

r = section(ws, r, "Tab 3 - Source data (person)")
p3 = [
    [f"{MARK_CHG}Person table", "Every Person column from sheet 04, sortable and filterable. Editable under REQ-IMP-07.", "REQ-DSH-04"],
    [f"{MARK_CHG}Assignment sub-table", "The selected person's assignments: project, role, start, end, person weight.", "REQ-PSN-02, REQ-PSN-03"],
    ["Person period weight sub-table", "Override periods for the selected assignment.", "REQ-PSN-05"],
    ["Export", "Export the visible table to .xlsx.", "REQ-DSH-06"],
]
r = table(ws, r, ["Component", "Content", "REQ-ID"], p3, [34, 96, 22], wrap_cols=(2,), mark_col=1)

r = section(ws, r, "Editing rules   [new at v0.2, from REQ-IMP-07]")
edit = [
    [f"{MARK_CHG}The v0.1 rule 'nothing on screen is editable' is withdrawn. Imported data can be edited in the application.", "REQ-IMP-07"],
    ["An edited cell is marked as changed, and a running count of unsaved edits is always visible.", "REQ-IMP-08"],
    ["Closing the page or loading another workbook while edits are unsaved raises a warning first.", "REQ-IMP-08"],
    [f"{MARK_CHG}Every field is editable, identifiers included (Q-20). Consistency is protected by rule, not by locking fields.", "REQ-IMP-09"],
    ["An edited value is validated exactly as an imported one would be; a rejected edit is refused at the point of entry, not at export.", "REQ-IMP-09"],
    [f"{MARK_NEW}Changing an identifier updates every row that references it, after showing how many will change. Deleting a row that is still referenced is refused, naming what points at it.", "REQ-IMP-10, V-17"],
    [f"{MARK_NEW}Editing a milestone date re-computes that project's derived periods, unless its period dates were set by hand.", "C-10"],
    ["Export writes the edited data in the template layout, so the exported file can be re-imported unchanged.", "REQ-IMP-04, REQ-IMP-07"],
    ["The application never writes to the workbook on disk by itself - export is always a deliberate act.", "Out of scope, sheet 02"],
]
r = table(ws, r, ["Rule", "REQ-ID"], edit, [124, 22], wrap_cols=(1,), mark_col=1)

r = section(ws, r, "Cross-cutting UI rules")
ui = [
    ["A validation banner appears whenever the last import produced findings, and opens the full report on click."],
    ["Empty state: with nothing loaded, every tab shows the same clear 'Load your source workbook' panel and a link to download the blank template."],
    ["The application version and expected schema version are shown in the footer at all times (REQ-VC-02)."],
]
r = table(ws, r, ["Rule"], ui, [146], wrap_cols=(1,))
r = legend(ws, r)

# ---- 07 Architecture ------------------------------------------------------
ws, r = sheet(wb, "07_Architecture", "Architecture and technical approach")

r = section(ws, r, "Decision D-01 - how the HTML file reads the Excel data  [CONFIRMED: option (a)]")
r = lines(ws, r, [
    "A page opened from local disk (file://) cannot read an .xlsx from a fixed path on its own - the browser's",
    "security model blocks it. Q-05 confirmed local .html files do open on the work PC, so the approach is viable.",
])
r += 1
opts = [
    ["(a)", "User selects the workbook through a file picker or drag-and-drop; an embedded SheetJS-style parser reads it in memory.",
     "True .xlsx support; no install; no server; user controls which file is loaded.",
     "The file must be selected each session (mitigated by optional caching, REQ-IMP-06).", "CONFIRMED - adopted"],
    ["(b)", "Data kept as .csv next to the HTML, still opened through a picker.",
     "Simplest possible parsing.", "Loses formatting, multiple sheets and validation - and Q-09 asks for one multi-sheet workbook, which .csv cannot express.", "Rejected"],
    ["(c)", "Excel is only an export target; browser storage is the working record.",
     "No load step at all.", "Contradicts the workbook being the archive of record; clearing browser data would lose everything.", "Rejected"],
]
r = table(ws, r, ["Option", "Description", "Advantages", "Disadvantages", "Outcome"],
          opts, [9, 50, 40, 48, 18], wrap_cols=(2, 3, 4))
r = note(ws, r, "Q-09 strengthens this decision: one workbook with ten sheets is exactly what .xlsx does well and .csv cannot do at all.")

r = section(ws, r, "Technology")
tech = [
    ["Container", "One .html file, opened by double-click.", "REQ-OUT-01"],
    ["Excel parsing/writing", "SheetJS (xlsx) community build, embedded inline - no CDN, so it works offline.", "REQ-OUT-02, REQ-NFR-04"],
    ["Charts", "Inline SVG drawn by the application's own code, or a small embedded chart library. No external requests either way.", "REQ-DSH-02, REQ-NFR-04"],
    ["Framework", "None. Plain modern JavaScript.", "REQ-OUT-02"],
    ["Styling", "Embedded CSS. Readable on a laptop screen and when printed.", "REQ-OUT-01"],
    ["Persistence", "Optional browser cache of the last load, marked as a convenience, never as the record.", "REQ-IMP-06"],
]
r = table(ws, r, ["Concern", "Choice and reason", "REQ-ID"], tech, [24, 108, 20], wrap_cols=(2,))

r = section(ws, r, "Internal structure - built for later change (REQ-NFR-01)")
layers = [
    ["1. IO layer", "Reads the chosen workbook into raw rows; writes .xlsx on export.", "Swapping the file-access method touches this layer only."],
    ["2. Validation layer", "Applies rules V-01..V-21, produces one findings report.", "New rules are added here without touching parsing or maths."],
    ["3. Model layer", "Typed in-memory objects: Project, Milestone, ProjectPeriod, RoleFactor, Person, Assignment, PersonPeriodWeight, Config, Lists.", "New fields are added here and flow outward."],
    [f"{MARK_NEW}4. Edit buffer", "Holds on-screen changes, tracks which rows are dirty, re-runs validation per edit, and feeds export.", "Added for REQ-IMP-07. Keeping edits out of the model layer means the calculation never has to care whether a value came from the file or the screen."],
    [f"{MARK_CHG}5. Calculation layer", "The monthly engine above. Pure functions - no DOM, no file access.", "Directly unit-testable, and the piece most likely to be refined once you see real output."],
    [f"{MARK_CHG}6. Presentation layer", "Tabs, tables, graphs, filters, editing controls, export buttons.", "A new tab or graph is added here alone."],
]
r = table(ws, r, ["Layer", "Responsibility", "Why the boundary is there"],
          layers, [22, 66, 64], wrap_cols=(2, 3), mark_col=1)
r = note(ws, r, "The calculation layer never touches the screen and never touches files. That is what makes both future "
                "requirement changes and testing cheap.")

r = section(ws, r, "Deployment")
dep = [
    [f"{MARK_CHG}The user keeps one folder containing the .html file and ONE source workbook. Copying the folder moves the whole tool."],
    ["Updating the application means replacing the .html file. Source data is untouched, because it lives outside it (REQ-OUT-03)."],
    ["The application checks the source schema version on load and warns on a mismatch (REQ-VC-03)."],
]
r = table(ws, r, ["Point"], dep, [146], wrap_cols=(1,), mark_col=1)
r = legend(ws, r)

# ---- 08 WBS ---------------------------------------------------------------
ws, r = sheet(wb, "08_WBS_Schedule", "Work breakdown - the five steps",
              "Each step ends at a review gate. Work on the next step starts only after that gate is passed.")

wbs = [
    ["1", "1.1", "Draft the development plan.", "PRAP_Development_Plan_v0.1.xlsx", "Complete"],
    ["1", "1.2", "Requester reviews; answers the open questions.", "v0.11_reviewed mark-up", "Complete"],
    ["1", "1.3", "Propagate the answers across all sheets and re-issue.", "PRAP_Development_Plan_v0.2.xlsx", "Complete"],
    ["1", "1.4", "Answer the follow-up questions Q-13..Q-20.", "v0.2_reviewed mark-up", "Complete"],
    ["1", "1.5", "Incorporate round 2; work the period mapping through for coverage gaps.", "PRAP_Development_Plan_v0.3.xlsx", "Complete"],
    ["1", "1.6", "Settle the period-boundary questions Q-21..Q-26.", "v0.3_reviewed mark-up", "Complete"],
    ["1", "1.7", "Incorporate round 3; verify the derivation against degenerate timelines.", "PRAP_Development_Plan_v0.4.xlsx", "Complete - issued for review"],
    ["1", "1.8", "Answer Q-27 and Q-28.", "v0.4_reviewed mark-up", "Complete"],
    ["1", "1.9", "Final plan issue.", "PRAP_Development_Plan_v1.0.xlsx", "Complete"],
    ["1", "1.10", "Apply the post-approval change request (Inspection milestone, seventh period).", "PRAP_Development_Plan_v1.1.xlsx", "Complete"],
    ["1", "1.11", "Confirm R-03.", "v1.1_reviewed mark-up", "Complete"],
    ["1", "1.12", "Record R-03 as closed and re-issue for approval.", "PRAP_Development_Plan_v1.2.xlsx", "Complete - issued for review"],
    ["1", "1.13", "Approve plan v1.2 as a change against the v1.0 baseline.", "Approval on 12_Review_Log", "Complete"],
    ["1", "G1b", "GATE 1 RE-CONFIRMED - plan v1.2 approved by Dan, 2026-08-01. Step 1 closed.", "Approval recorded on 12_Review_Log", "Complete"],
    ["1", "1.14", "Apply change R-04 (note column on every source sheet).", "PRAP_Development_Plan_v1.3.xlsx", "Complete - issued for review"],
    ["1", "1.15", "Approve plan v1.3 as final.", "Approval on 12_Review_Log", "Complete"],
    ["1", "G1c", "Plan v1.3 approved by Dan, 2026-08-01.", "Approval recorded on 12_Review_Log", "Complete"],
    ["1", "1.16", "Apply change R-05 (project_type split).", "PRAP_Development_Plan_v1.4.xlsx", "Complete - issued for review"],
    ["1", "1.17", "Apply the specification-review answers (R-06, R-07, R-08).", "PRAP_Development_Plan_v1.5.xlsx", "Complete - issued for review"],
    ["1", "1.18", "Apply the component-list review outcome (R-09).", "PRAP_Development_Plan_v1.6.xlsx", "Complete - issued for review"],
    ["1", "1.19", "Apply change R-10 (role factor keyed on phase and period).", "PRAP_Development_Plan_v1.7.xlsx", "Complete - issued for review"],
    ["1", "1.20", "Apply change R-11 (conduct stretches named apart; ProjectPeriod natural key).", "PRAP_Development_Plan_v1.8.xlsx", "Complete - issued for review"],
    ["1", "1.21", "Close the validation gaps found while answering the PersonPeriodWeight key question (R-12).", "PRAP_Development_Plan_v1.9.xlsx", "Complete - issued for review"],
    ["1", "1.22", "Apply the component-list v0.7 review outcome (R-13).", "PRAP_Development_Plan_v1.10.xlsx", "Complete - issued for review"],
    ["1", "1.23", "Approve the plan as a baseline.", "Approval on 12_Review_Log", "Complete - APPROVED 2026-08-02"],
    ["1", "G1d", "GATE 1 CLOSED - plan v2.0 approved by Dan, 2026-08-02, superseding v1.3.", "Approval recorded on 12_Review_Log", "Complete"],
    ["1", "G1", "GATE 1 - development plan approved by Dan, 2026-08-01. Decisions C-06..C-11 confirmed.", "Approval recorded on 12_Review_Log", "Complete"],

    ["2", "2.0", "Generate the source workbook template and a dummy data file for review.", "PRAP_SourceData_Template + _Dummy v1.1", "Complete - issued for review"],
    ["2", "2.1", "Fix the source workbook schema: exact sheet names, column headers, types, value lists.", "Specification sheet 'Data schema'", "In progress"],
    ["2", "2.2", "Specify the calculation engine as pseudocode plus the worked example as an acceptance test.", "Specification sheet 'Calculation'", "Not started"],
    ["2", "2.3", "Specify period derivation per project type, and the over/under-allocation detection.", "Specification sheet 'Calculation'", "Not started"],
    ["2", "2.4", "Specify import validation rules V-01..V-21 and the findings report.", "Specification sheet 'Validation'", "Not started"],
    ["2", "2.5", "Specify each dashboard tab: every table, column, graph, filter and interaction.", "Specification sheet 'UI spec'", "Not started"],
    ["2", "2.6", "Specify on-screen editing, the dirty-state model and export round-tripping.", "Specification sheet 'Editing & IO'", "Not started"],
    ["2", "2.7", "Specify the version/compatibility check.", "Specification sheet 'IO & versioning'", "Not started"],
    ["2", "2.8", "Build the requirement-to-specification traceability matrix.", "Specification sheet 'Traceability'", "Not started"],
    ["2", "G2", "GATE 2 - programming specification approved.", "PRAP_Programming_Specification_v1.0.xlsx", "Not started"],

    ["3", "3.1", "High-level UI design: tab structure, page regions, navigation. No code behind it.", "Prototype HTML (UI only)", "Not started"],
    ["3", "3.2", "Requester reviews the high-level component design.", "Review comments", "Not started"],
    ["3", "3.3", "Final UI design with clear per-component requirements.", "Prototype HTML v1.0 + component list", "Not started"],
    ["3", "G3", "GATE 3 - final design and component list approved. Code generation starts only for approved components.", "Approved component list", "Not started"],

    ["4", "4.1", "IO layer: workbook load, sheet/column mapping, export.", "app/PRAP.html", "Complete"],
    ["4", "4.2", "Validation layer: rules V-00..V-24 and the findings report.", "app/PRAP.html", "Complete"],
    ["4", "4.3", "Model layer, Lists and Config handling.", "app/PRAP.html", "Complete"],
    ["4", "4.4", "Edit buffer: on-screen editing, dirty state, per-edit validation, cascading identifier edits, row insertion.", "app/PRAP.html", "Complete"],
    ["4", "4.5", "Calculation engine, verified against the reference implementation.", "app/PRAP.html + test evidence", "Complete - EXACT MATCH on all 1,225 person-months of the dummy dataset"],
    ["4", "4.6", "Overall tab: tables, graphs, filters, over/under-allocation flagging.", "app/PRAP.html", "Complete"],
    ["4", "4.7", "Source data (project), Source data (person) and General assumptions tabs.", "app/PRAP.html", "Complete"],
    ["4", "4.8", "Blank source workbook template with value lists and example rows.", "PRAP_SourceData_Template_v1.6.xlsx", "Complete"],
    ["4", "4.9", "Requester reviews output against real data; refinements folded in.", "Updated code", "In progress - rounds 1-25 applied (app v1.24); GATE 4 open"],
    ["4", "G4", "GATE 4 - application functionally complete.", "PRAP_Application_v0.9.html", "Not started"],

    ["5", "5.1", "Full pass over the traceability matrix: every Must requirement demonstrated.", "Completed traceability matrix", "Not started"],
    ["5", "5.2", "Test on a clean Windows PC in Edge and Chrome, offline.", "Test evidence", "Not started"],
    ["5", "5.3", "Write the user guide: folder layout, how to load, how to maintain source data.", "User guide", "Not started"],
    ["5", "5.4", "Final version alignment across plan, specification and application.", "Version alignment table (sheet 09)", "Not started"],
    ["5", "G5", "GATE 5 - release.", "PRAP_Application_v1.0.html", "Not started"],
]
r_start = r
r = table(ws, r, ["Step", "Task", "Activity", "Deliverable", "Status"],
          wbs, [7, 8, 86, 42, 24], wrap_cols=(3, 4))
last = r_start + len(wbs)

dv2 = DataValidation(type="list",
                     formula1='"Not started,In progress,Pending you,Complete - issued for review,Complete,Blocked"',
                     allow_blank=True)
ws.add_data_validation(dv2)
dv2.add(f"E{r_start + 1}:E{last}")

r = section(ws, r, "Progress")
prog = [
    ["Tasks total", f"=COUNTA(B{r_start + 1}:B{last})"],
    ["Complete", f'=COUNTIF(E{r_start + 1}:E{last},"Complete")+COUNTIF(E{r_start + 1}:E{last},"Complete - issued for review")'],
    ["Awaiting the requester", f'=COUNTIF(E{r_start + 1}:E{last},"Pending you")'],
    ["Not started", f'=COUNTIF(E{r_start + 1}:E{last},"Not started")'],
]
r = table(ws, r, ["Measure", "Count"], prog, [26, 12])
r = note(ws, r, "No duration estimates: the pace is set by review turnaround at each gate, not by build effort.")

# ---- 09 Version control ---------------------------------------------------
ws, r = sheet(wb, "09_Version_Control", "Version control across documents (REQ-VC-01)")

r = section(ws, r, "What is version-controlled")
vc = [
    ["Development plan", "PRAP_Development_Plan_v<ver>.xlsx", "docs/", "Sheet 01_Version_History"],
    ["Programming specification", "PRAP_Programming_Specification_v<ver>.xlsx", "docs/", "Its own version-history sheet"],
    ["Application", "PRAP_Application_v<ver>.html", "app/", "Version constant shown in the footer"],
    [f"{MARK_CHG}Source data template", "PRAP_SourceData_v<ver>.xlsx  (one workbook, was two)", "templates/", "schema_version in the Config sheet"],
    ["Output / evidence files", "PRAP_<content>_v<ver>_<yyyymmdd>.xlsx", "output/", "Filename plus a header block"],
    ["Generator scripts", "tools/*.py", "tools/", "Git history"],
]
r = table(ws, r, ["Artifact", "Naming convention", "Folder", "Where its version is recorded"],
          vc, [28, 56, 16, 44], wrap_cols=(2, 4), mark_col=1)

r = section(ws, r, "Version numbering")
num = [
    ["0.x", "Draft, under review. Content may change substantially.", "0.1 -> 0.2 on each re-issue during review."],
    ["x.0", "Approved baseline. Changing it requires a new approval.", "0.9 -> 1.0 at the gate."],
    ["x.y", "Approved change against a baseline: added requirement, bug fix, layout change.", "1.0 -> 1.1"],
    ["Schema version", "Independent counter for the source workbook structure. Steps only when a sheet or column changes.", "schema 1 -> 2"],
]
r = table(ws, r, ["Pattern", "Meaning", "Example"], num, [16, 86, 34], wrap_cols=(2,))
r = note(ws, r, "A reviewer mark-up carries the reviewed version's number with a suffix (v0.11_reviewed against v0.1). "
                "The incorporating issue takes the next draft number - here v0.2.")

r = section(ws, r, "Version alignment - which versions belong together")
align = [
    ["v0.1", "-", "-", "-", "2026-07-25", "Plan issued for review"],
    ["v0.2", "-", "-", "1 (draft)", "2026-07-31", "Reviewer answers incorporated; single-workbook schema defined"],
    ["v0.3", "-", "-", "1 (draft)", "2026-07-31", "Round 2 incorporated; type-specific period sets; milestone list replaced"],
    ["v0.4", "-", "-", "1 (draft)", "2026-07-31", "Round 3 incorporated; period model settled and verified"],
    ["v1.0", "-", "-", "1", "2026-07-31", "Step 1 baseline. APPROVED 2026-08-01 by Dan (Gate 1)."],
    ["v1.1", "-", "-", "1", "2026-08-01", "Change against baseline: Inspection milestone, seventh period. Superseded."],
    ["v1.2", "-", "-", "1", "2026-08-01", "APPROVED 2026-08-01 by Dan. Baseline until v1.3 is approved."],
    ["v1.3", "v0.2", "-", "2", "2026-08-01", "R-04: note column on every source sheet. APPROVED 2026-08-01 by Dan."],
    ["v1.4", "v0.3", "prototype v0.2", "3", "2026-08-01", "R-05: project_type split. Superseded."],
    ["v1.5", "v0.4", "prototype v0.3", "3", "2026-08-01", "R-05..R-08. Superseded."],
    ["v1.6", "v0.5", "prototype v0.4", "3", "2026-08-01", "R-05..R-09. Superseded."],
    ["v1.7", "v0.6", "prototype v0.5", "4", "2026-08-02", "R-05..R-10. Superseded."],
    ["v1.8", "v0.7", "prototype v0.6", "5", "2026-08-02", "R-05..R-11. Superseded."],
    ["v1.9", "v0.8", "prototype v0.7", "5", "2026-08-02", "R-05..R-12. Superseded."],
    ["v1.10", "v0.9", "prototype v0.8", "5", "2026-08-02", "R-05..R-13. Superseded by v2.0."],
    ["v2.0", "v1.0", "prototype v0.8", "5", "2026-08-02", "APPROVED BASELINE. Steps 1-3 closed; Step 4 authorised."],
    ["v2.1", "v1.0", "application v1.0", "5", "2026-08-02", "Step 4 built: app/PRAP.html. Tasks 4.1-4.8 complete."],
    ["v2.2", "v1.0", "application v1.1", "5", "2026-08-02", "Gate 4 refinements round 1: three UI changes."],
    ["v2.3", "v1.0", "application v1.2", "5", "2026-08-02", "Gate 4 round 2: the application carries its own document provenance."],
    ["v2.4", "v1.0", "application v1.3", "5", "2026-08-02", "Gate 4 round 3: sticky headers, year axis, provisional editing."],
    ["v2.5", "v1.0", "application v1.4", "5", "2026-08-02", "Gate 4 round 4: information pop-ups and row deletion."],
    ["v2.6", "v1.0", "application v1.5", "5", DOC_DATE, "Gate 4 round 5: lookup columns, cell values, row actions everywhere, type-ahead."],
    ["v2.7", "v1.0", "application v1.6", "5", DOC_DATE, "Gate 4 round 6: insert/delete fixed on the four child tables."],
    ["v2.8", "v1.0", "application v1.6", "5", DOC_DATE, "Gate 4 round 7: second dummy dataset, 10 projects x 10 people."],
    ["v2.9", "v1.0", "application v1.7", "5", DOC_DATE, "Gate 4 round 8: two-way scrolling everywhere, project by name, allocated assignment_id."],
    ["v2.10", "v1.0", "application v1.8", "5", DOC_DATE, "Gate 4 round 9: a foreign key no longer cascades; second override windows can be added."],
    ["v2.11", "v1.0", "application v1.9", "5", DOC_DATE, "Gate 4 round 10: the value list survives being scrolled, and matches on tokens."],
    ["v2.12", "v1.0", "application v1.10", "5", DOC_DATE, "Gate 4 round 11: a per-project timeline, and utilisation stacked by project."],
    ["v2.13", "v1.0", "application v1.11", "5", DOC_DATE, "Gate 4 round 12: demand by person, project utilisation by person, child tables scoped."],
    ["v2.14", "v1.0", "application v1.12", "5", DOC_DATE, "Gate 4 round 13: weight overrides follow the selected assignment."],
    ["v2.15", "v1.0", "application v1.13", "5", DOC_DATE, "Gate 4 round 14: monthly trend line charts, and one header shape for every panel."],
    ["v2.16", "v1.0", "application v1.14", "5", DOC_DATE, "Gate 4 round 15: derived columns locked in the workbook, template v1.7."],
    ["v2.17", "v1.0", "application v1.15", "5", DOC_DATE, "Gate 4 round 16: visual design pass; contrast measured, not judged."],
    ["v2.18", "v1.0", "application v1.16", "5", DOC_DATE, "Gate 4 round 17: AI-agent reference, JSON interchange, interoperability audit."],
    ["v2.19", "v1.0", "application v1.17", "5", DOC_DATE, "Gate 4 round 18: a plan can be started in the application, with no workbook."],
    ["v2.20", "v1.0", "application v1.18", "5", DOC_DATE, "Gate 4 round 19: the child sections are enterable from the start, before their parent is saved."],
    ["v2.21", "v1.0", "application v1.19", "5", DOC_DATE, "Gate 4 round 20: scroll position survives a re-render."],
    ["v2.22", "v1.0", "application v1.20", "5", DOC_DATE, "Gate 4 round 21: a row with no identifier is not a record."],
    ["v2.23", "v1.0", "application v1.21", "5", DOC_DATE, "Gate 4 round 22: allocated ids, neutral weights, and scroll regions that say there is more."],
    ["v2.24", "v1.0", "application v1.22", "5", DOC_DATE, "Gate 4 round 23: the project window and team size derived from the rows beneath."],
    ["v2.25", "v1.0", "application v1.23", "5", DOC_DATE, "Gate 4 round 24: outsourcing filter, a horizon that follows the filters, and a change log."],
    ["v2.26", "v1.0", "application v1.24", "5", DOC_DATE, "Gate 4 round 25: multi-value filters, and the two generators on the project tab. 4.9 continues."],
    ["", "", "", "", "", ""],
]
r = table(ws, r, ["Plan version", "Specification version", "Application version", "Schema version", "Date", "Note"],
          align, [15, 22, 20, 16, 14, 52], wrap_cols=(6,))
r = note(ws, r, "One row per release. Answers 'which specification produced this .html, and which source template does it expect?'.")

r = section(ws, r, "Repository rules")
git = [
    ["All work is committed to branch claude/project-resource-assignment-app-1vjdzh."],
    ["Documents are generated from scripts in tools/ so their content is reviewable as a text diff, not only as a binary."],
    ["One commit per meaningful change, with a message naming the step and the affected REQ-IDs."],
    ["A released version is tagged, so any release can be reproduced exactly."],
    [f"{MARK_NEW}Reviewer mark-ups are committed as received, unedited, so the review trail is auditable."],
]
r = table(ws, r, ["Rule"], git, [146], wrap_cols=(1,), mark_col=1)

# ---- 10 Risks -------------------------------------------------------------
ws, r = sheet(wb, "10_Risks", "Risks and assumptions")

r = section(ws, r, "Risks")
risks = [
    [f"{MARK_CHG}R-01", "The resource formula does not match how burden is actually judged, so output looks wrong.", "Medium", "High",
     "Reduced from High at Q-01, which fixed the three factors. A worked example (sheet 05) is agreed at Step 2 before code is written, and every factor stays in data."],
    ["R-02", "The source workbook is maintained inconsistently - typos, overlapping periods - and the simulation silently misleads.", "High", "Medium",
     "Fourteen validation rules on import, ID-based references instead of name matching, and a findings report the user cannot miss."],
    [f"{MARK_CHG}R-03", "Corporate policy blocks local .html files or restricts the browser.", "-", "-",
     "CLOSED at Q-05: local .html files open on the work PC. Re-test on the target PC at Step 5.2 all the same."],
    ["R-04", "Requirements grow after you see real output.", "High", "Medium",
     "Expected and planned for: six separated layers (sheet 07) plus extension columns already in the schema."],
    [f"{MARK_CHG}R-05", "Data volume grows beyond what a browser handles comfortably.", "Low", "Low",
     "Downgraded at Q-06: under 20 projects and 30 people is comfortably inside browser limits, with the headroom target in REQ-NFR-03."],
    ["R-06", "Excel date handling differs between Windows regional settings and misreads dates.", "Medium", "High",
     "Dates normalised to ISO on import; an unparseable date is a hard error, never a guess (REQ-NFR-05)."],
    ["R-07", "The file must be re-selected each session and this becomes irritating.", "Medium", "Low",
     "Optional browser caching (REQ-IMP-06) plus a one-click reload of the last file."],
    [f"{MARK_NEW}R-08", "On-screen edits are lost because the user closes the page without exporting.", "High", "High",
     "The direct cost of REQ-IMP-07. Mitigated by REQ-IMP-08: a visible unsaved-edit count and a warning before any action that discards them. The browser cache (REQ-IMP-06) becomes more valuable here as a second net."],
    [f"{MARK_CHG}R-09", "Milestone-derived periods break for a project with missing or out-of-order milestones.", "Medium", "High",
     "Raised from Medium impact: at v0.3 the derivation depends on specific milestones, so a missing CTA submission or DB lock now stops it outright. V-14 and V-16 make those errors rather than warnings; V-12 catches any remaining gap and falls back to weight 1.00 with a report."],
    [f"{MARK_CHG}R-10", "The weight and role-factor values are unknown to the plan, so output cannot be checked against reality until they are entered.", "Medium", "Medium",
     "Downgraded at Q-17: the values are data you enter in the source workbook, not a plan input, which is what REQ-CAL-06 always required. The engine is validated against the worked example on sheet 05; realism of the numbers is validated at WBS 4.9 against your real data."],
    [f"{MARK_CHG}R-11", "The period mapping leaves parts of a project with no period, so months silently carry a default weight.", "Low", "Medium",
     "Closed in substance at Q-21..Q-23: the mapping is now contiguous, verified against five timelines including four degenerate ones. REQ-CAL-10 and V-12 keep any residual gap visible rather than silent."],
    [f"{MARK_NEW}R-12", "Fully editable identifiers let a user break the links between sheets.", "Medium", "High",
     "The direct cost of the Q-20 answer. Mitigated by V-17 and REQ-IMP-10: identifier edits cascade after confirmation, and a referenced row cannot be deleted. Without that, 'every field editable' and 'no inconsistency' could not both hold."],
]
r = table(ws, r, ["ID", "Risk", "Likelihood", "Impact", "Mitigation"],
          risks, [9, 58, 12, 10, 64], wrap_cols=(2, 5), mark_col=1)

r = section(ws, r, "Assumptions")
assum = [
    [f"{MARK_CHG}A-01", "Data volume is under 20 projects and 30 people.", "CONFIRMED Q-06"],
    [f"{MARK_CHG}A-02", "One person's capacity is 1.00 FTE = 160 hours per month; part-timers carry a lower capacity_fte.", "CONFIRMED Q-08"],
    ["A-03", "Monthly granularity is sufficient; weekly or daily simulation is not required in v1.0.", "Standing"],
    ["A-04", "The source workbook is maintained by hand, not generated by another system.", "Standing"],
    [f"{MARK_CHG}A-05", "Windows 10 or 11 with Edge or Chrome, and local .html files open.", "CONFIRMED Q-05"],
    ["A-06", "Only one person maintains the workbook at a time - no concurrent editing to reconcile.", "Standing"],
    [f"{MARK_CHG}A-07", "Milestone dates now DRIVE period boundaries (REQ-CAL-09); they are no longer merely informational.", "Changed by Q-01/Q-04"],
    [f"{MARK_CHG}A-08", "Period sets are type-specific: seven names for either clinical trial type, three for 'Others'. No name repeats within a project.", "CONFIRMED Q-18, Q-23; revised at R-11"],
    [f"{MARK_CHG}A-09", "A clinical trial's period weights are selected by its clinical phase.", "CONFIRMED Q-26"],
    [f"{MARK_CHG}A-10", "'Others' projects have no milestone mapping, so their three periods are entered by hand.", "CONFIRMED Q-25"],
    [f"{MARK_NEW}A-11", "No example source file is coming; the structure in this plan is taken as matching the existing one.", "CONFIRMED Q-13"],
]
r = table(ws, r, ["ID", "Assumption", "Status"], assum, [9, 110, 24], wrap_cols=(2,), mark_col=1)
r = legend(ws, r)

# ---- 11 Open questions ----------------------------------------------------
ws, r = sheet(wb, "11_Open_Questions", "Questions and answers",
              "All 28 questions raised across four review rounds, with their answers and how each was applied. "
              "Nothing is open.")

r = section(ws, r, "Answered at the v0.4 review  [round 4 - final]")
round4 = [
    ["Q-27", "Naming", "What is a close-out called on a trial with no interim lock?", "With no interim DB lock, the final DB lock falls into 'Close-out (final)'.", "Confirms v0.4 as built - one name, no special case in the weight lookup."],
    ["Q-28", "Calculation", "What selects period weights for 'Others' projects?", "No need to distinguish each 'Others' project. They are given manual input for the weight of periods.", "Applied - 'Others' weights are entered by hand, so PeriodWeightStandard becomes a clinical-trial table only. Together with Q-25 this makes 'Others' projects hand-entered throughout: dates and weights alike."],
]
r = table(ws, r, ["ID", "Topic", "Question (short)", "Your answer", "How it was applied in v1.0"],
          round4, [8, 15, 34, 62, 52], wrap_cols=(3, 4, 5))

r = section(ws, r, "Answered at the v0.3 review  [round 3]")
round3 = [
    ["Q-21", "Calculation", "Is 'Before-Start-up' the stretch before Start-up?", "Yes - project start date to 1 month before CTA submission.", "Applied - confirms the v0.3 reading. Sequence 1 on sheet 05."],
    ["Q-22", "Calculation", "Conduct ends 3 months prior, Close-out starts 2 - which?", "Close-out starts from 3 months prior, and runs until DB lock.", "Applied - Close-out is now 3 months long and Conduct ends the day before it. This is the opposite of the v0.3 default, which had shortened Close-out instead; corrected."],
    ["Q-23", "Calculation", "Which DB lock does Close-out end at?", "With an interim lock, split into 'Close-out (interim)' and 'Close-out (final)'. The stretches before and after 'Close-out (interim)' are both Conduct.", "Applied, and the largest change in v0.4: 'Conduct' can now occur twice in one project. ProjectPeriod gains period_seq; REQ-CAL-11 added; the clinical-trial period set grows to five names."],
    ["Q-24", "Data model", "Which is the eighth standard milestone?", "'Protocol (v1)'.", "Applied - milestone list now holds eight."],
    ["Q-25", "Data model", "How do 'Others' projects get their period dates?", "'Others' projects don't have milestone mapping.", "Applied - confirms the v0.3 assumption; their periods are entered by hand. Assumption A-10 confirmed."],
    ["Q-26", "Calculation", "What selects a project's weight?", "Clinical phase, where project_type is 'Clinical Trial'.", "Applied - PeriodWeightStandard re-keyed from project_category to clinical_phase. clinical_phase now drives the simulation, so V-19 makes a missing phase an error."],
]
r = table(ws, r, ["ID", "Topic", "Question (short)", "Your answer", "How it was applied in v0.4"],
          round3, [8, 15, 34, 62, 52], wrap_cols=(3, 4, 5))

r = section(ws, r, "Answered at the v0.2 review  [round 2]")
round2 = [
    ["Q-13", "Blocker", "The '13_Templete_example' sheet is missing.", "Structure already provided is similar to the existing one. No example file coming. Close this.", "CLOSED. WBS task 1.5 removed; assumption A-11 records that the schema in this plan is taken as matching yours."],
    ["Q-14", "Data model", "outsourcing_type listed a value twice.", "'Full outsourcing' / 'Partial outsourcing' / 'Full In-house'.", "Applied - three values, REQ-PRJ-04 and the Lists sheet."],
    ["Q-15", "Data model", "Which milestone set is standard?", "'CTA submission', 'First SIV', 'LPI', 'interim DB lock cut-off', 'interim DB lock', 'final DB lock cut-off', 'final DB lock'.", "Applied - list replaced; FPI dropped. The answer says eight but names seven, so the count is queried at Q-24."],
    ["Q-16", "Calculation", "Confirm the milestone-to-period mapping.", "Start-up = 1 month before CTA submission, 4 months long. Conduct = after Start-up to 3 months before interim DB lock (final if no interim). Close-out = 2 months before that DB lock, to DB lock.", "Applied with three corrections - the mapping as written leaves gaps. See sheet 05 and Q-21, Q-22, Q-23."],
    ["Q-17", "Calculation", "The weight and role-factor values.", "Actual numbers to be filled into the source file (excel).", "CLOSED. The plan fixes where the values live, not what they are - which is what REQ-CAL-06 always required. Risk R-10 downgraded. Part (c) of the question is handled by design and re-asked lightly at Q-26."],
    ["Q-18", "Calculation", "Do 'Others' projects share the same periods?", "No - 'Others' uses 'Planning' / 'Develop' / 'Close'.", "Applied - REQ-PRJ-12, type-specific period sets, V-15. Raises Q-25: those periods have no milestones to derive from."],
    ["Q-19", "Data model", "Can system_prepared_by be dropped?", "Yes - replaced by the EDC_setup .. DM_conduct fields.", "Applied - column removed from the Project sheet."],
    ["Q-20", "Editing", "Which fields should be editable?", "Every field. Validation rules prevent inconsistency between sheets instead.", "Applied - REQ-IMP-09 widened, REQ-IMP-10 and V-17 added, decision C-09, risk R-12."],
]
r = table(ws, r, ["ID", "Topic", "Question (short)", "Your answer", "How it was applied in v0.3"],
          round2, [8, 15, 34, 62, 52], wrap_cols=(3, 4, 5))

r = section(ws, r, "Answered at the v0.11 review  [round 1]")
answered = [
    ["Q-01", "Calculation", "Do weights multiply?", "Yes: project weight x role factor x person weight. Project weight comes from the project's category and is the same for everyone on that project in that milestone period. Role factor is the standard weight of the role. Person weight says how much the person works on the project.", "Applied - sheet 05. base_allocation removed; PersonPeriodWeight became an override."],
    ["Q-02", "Calculation", "Pro-rate partial months by calendar days?", "Yes - expected work follows the calendar days worked between the joined date and the end date.", "Applied - C-02 confirmed; REQ-CAL-05 raised to Must."],
    ["Q-03", "Roles", "Which roles, and do they differ in burden?", "Clinical Trial: Project oversight, Lead data manager, Clinical Data Associator, Clinical Database Programmer, Data Analyst. Others: Project lead, Main staff, Other staff.", "Applied - RoleFactor gains a project_type column; V-03 raised to an error."],
    ["Q-04", "Project periods", "Standard period names and weights?", "Before-Start-up, Start-up, Conduct, Close-out.", "Applied - sheet 04 ProjectPeriod. Weight VALUES still needed - Q-17."],
    ["Q-05", "Environment", "Can you open a local .html file?", "Yes.", "Applied - risk R-03 closed."],
    ["Q-06", "Data volume", "How many projects and people?", "Under 20 projects and 30 people.", "Applied - REQ-NFR-03 target reduced; R-05 downgraded."],
    ["Q-07", "Milestones", "Which milestones?", "CTA submission, FPI, LPI, DB lock cut-off, DB lock.", "Applied - Milestone list. Conflicts with the v0.11 mark-up in sheet 04 - see Q-15."],
    ["Q-08", "Capacity", "Capacity and thresholds?", "1 FTE = 160 h/month (8 h/day, 5 days/week, 20 days/month). Over-allocation: total above 1.5 FTE in a month. Under-allocation: below 0.8 FTE for 3 months or longer.", "Applied - REQ-CAL-04, REQ-CAL-07, REQ-CAL-08; Config parameters; sheet 06 flagging."],
    ["Q-09", "Source files", "One workbook or two?", "One single workbook with all sheets.", "Applied - sheet 04 restructured; REQ-OUT-03 reworded; architecture and deployment updated."],
    ["Q-10", "Existing data", "Existing spreadsheet structure?", "Refer to '13_Templete_example'.", "NOT APPLIED - that sheet is not in the returned file. See Q-13."],
    ["Q-11", "Horizon", "Planning horizon?", "24 months default, expandable to the latest project end date.", "Applied - REQ-CAL-01, REQ-DSH-07, Config default_horizon_months."],
    ["Q-12", "Scope check", "Anything wrongly out of scope?", "Fine.", "Applied - exclusions confirmed on sheet 02."],
]
r = table(ws, r, ["ID", "Topic", "Question (short)", "Your answer", "How it was applied in v0.2"],
          answered, [8, 15, 34, 62, 52], wrap_cols=(3, 4, 5))

r = section(ws, r, "Change requests against the approved v1.0 baseline")
r = note(ws, r, "Raised after Gate 1, so these are handled as a numbered change rather than absorbed silently. "
                "All are applied and confirmed; the change needs its approval signature on sheet 12. "
                "Note that this register and the risk register on sheet 10 both number their entries R-nn. "
                "They are separate ID spaces - change R-07 is the threshold answer, risk R-07 is the "
                "file-reselection nuisance. Renumbering either now would invalidate the approval signatures "
                "and cross-references already given against these IDs.")
chg = [
    ["R-01", "Data model", "Add 'Inspection' as a standard milestone.", "Applied. Milestone list grows to ten. Unlike the others, 'Inspection' MAY REPEAT within one project, so REQ-PRJ-13 and V-20 were added and V-14's uniqueness check now exempts it.", "Applied"],
    ["R-02", "Calculation", "Sheet 05 derivation edits: Before-Start-up ends at 'Protocol (v1)'; Start-up begins the day after it and ends at 'First SIV' (or 'FPI'); a seventh period 'After Close-out (final)' spans the Inspection dates.", "Applied. Clinical period set grows to six names; REQ-PRJ-12 and REQ-CAL-09 reworded; REQ-CAL-13 added for the milestone-beats-offset rule; 'FPI' restored to the milestone list as the First SIV fallback.", "Applied"],
    ["R-03", "Calculation", "Not requested - found while applying R-02.", "Period 7 was defined as earliest to latest 'Inspection'. Where an inspection is dated on or before the final DB lock, that makes period 7 start before period 6 and overlap Conduct. Only inspections AFTER the final DB lock open period 7; earlier ones stay markers, reported by V-21. CONFIRMED by the reviewer at the v1.1 review.", "CONFIRMED"],
    ["R-13", "UI", "Two requests at the component-list v0.7 review: highlight the interim and final DB lock milestones in red, and add a project utilisation graph to the 'Source data (project)' tab.", "Both applied. The DB locks are the milestones the whole period derivation hangs on - move a lock and every period after it moves - so red plus a larger marker is warranted; size carries it as well as hue, because D-04 makes 'never colour alone' a floor. The cut-off milestones stay ordinary: the cut-off is preparation, the lock is the event. The utilisation graph needed a new requirement, REQ-DSH-12, because the person tab's equivalent rests on absolute thresholds a project does not have - which the reviewer identified. The references are therefore relative: 2x and 0.5x the portfolio average for an active project-month, plus the project's own lifetime average.", "Applied"],
    ["R-12", "Validation", "Raised as a question about the PersonPeriodWeight key - why is period_start needed when assignment_id looks unique?", "The key is correct and unchanged: (assignment_id, period_start). assignment_id is unique in Assignment, but PersonPeriodWeight is a CHILD of it, and REQ-PSN-05, V-06 and the data model all allow one assignment to carry several non-overlapping override windows. Keying on assignment_id alone would cap it at one. But checking the question exposed two real gaps, both now closed: V-06's assignment-window half had never been implemented, so an overlapping pair passed silently and the applied weight depended on row order; and nothing checked that a PersonPeriodWeight row referred to a real assignment, so an orphan override was accepted and ignored without a word. V-24 added for the second. The dummy fixture now carries an assignment with TWO windows - it only ever had one each, which is why neither gap surfaced.", "Applied"],
    ["R-11", "Data model", "Name the two conduct stretches apart - 'Conduct (interim)' before an interim DB lock, 'Conduct (final)' after it or where there is no interim lock - and key ProjectPeriod on (project_id, period_name).", "Applied, and it simplifies more than it costs. The clinical period set grows from six names to seven, but no name now repeats in a project, so ProjectPeriod has a natural key again and period_seq goes back to carrying order rather than identity. V-18 becomes a plain uniqueness check. REQ-CAL-11 and REQ-PRJ-12 reworded. REQ-DSH-10, which existed to number repeated names on screen, is now satisfied by the data model instead - the display numbering is kept only as a guard. This is the alternative offered against decision D-15 at the component-list review, so D-15 is superseded. Schema 4 -> 5: the columns are unchanged but the period_name value set is not, so a v4 file's 'Conduct' rows would fail V-15. PeriodWeightStandard grows 48 -> 56 rows and RoleFactor 249 -> 289; both new Conduct entries carry the same weights the single 'Conduct' did, so the change renames without reweighting - the dummy dataset returns byte-identical load figures.", "Applied"],
    ["R-10", "Data model", "The role factor must be given by role AND project type AND clinical phase AND period, not by role and type alone.", "Applied. RoleFactor gains clinical_phase and period_name; its key becomes all four columns and the sheet grows from 13 rows to 249. REQ-CAL-02 reworded, V-23 added, schema 3 -> 4. This is a real gain in expressiveness - a role's burden genuinely is not flat across a project, and the dummy data now shows the database programmer peaking at start-up and the analyst at lock. It carries two costs, both stated rather than hidden: 249 rows is a large hand-maintained table, and RoleFactor now varies over the same three dimensions as PeriodWeightStandard, so the two multiply and can double-count if both are edited for the same reason. See the note on sheet 04.", "Applied"],
    ["R-09", "UI", "Nine components changed at the component-list v0.3 review: a fourth tab for the standing assumptions, insert-row on every editable table, clinical-phase filter, the unit toggle demoted from filter to setting, the demand-chart legend replaced by a hover pop-up, four timeline changes, a time zone on the load stamp, and the edit counter stating its validation standing.", "Applied. Seven were satisfiable by design alone. Two were not: nothing in the plan required the assumptions to be reachable in the application, and nothing said a row could be created at all - so REQ-DSH-11 and REQ-IMP-11 were added. REQ-DSH-05 and REQ-IMP-05 reworded. One conflict surfaced: O-10 asks the timeline to be coloured by period name, which contradicts accepted decision D-06 (shade by weight). O-10 is the more specific instruction, so D-06 is superseded and weight becomes a lightness step within each period hue.", "Applied"],
    ["R-08", "UI", "The two 'Conduct' stretches of a project must be distinguishable on screen (S2-04).", "Applied as REQ-DSH-10. A display rule, not a data change: period_name stays 'Conduct' in the workbook and the screen shows it with its sequence, 'Conduct (1)' and 'Conduct (2)'. Changing the stored name would have broken the weight lookup and V-15.", "Applied"],
    ["R-07", "Calculation", "Thresholds stay ABSOLUTE, not relative to capacity_fte (S2-01), and the under-allocation floor moves to 0.60 FTE (S2-05).", "Applied. Taken together these resolve the defect S2-01 was raised about: at a 0.60 floor every capacity in the data can clear it - 1.00 needs 60% utilisation, 0.80 needs 75%, 0.60 needs 100%. The risk only returns for a capacity BELOW the floor, so V-22 warns on exactly that. REQ-CAL-04 and REQ-CAL-07 reworded; Config default 0.80 -> 0.60.", "Applied"],
    ["R-06", "Non-functional", "Re-baseline REQ-NFR-03 to 100 projects and 1,000 people (S2-06).", "Applied, and it is the most consequential answer in this round. At that volume the person table is 1,000 rows x 60 months = 60,000 cells, and the per-person chart would be 1,000 bars at roughly 1.2px each. Neither survives a naive build, so virtualisation and aggregation become requirements (REQ-DSH-09) rather than optimisations. Priority raised from Should to Must.", "Applied"],
    ["R-05", "Data model", "Split project_type: 'Clinical Trial' becomes 'NewDrug CT' and 'Biosimilar CT'; every clinical trial is one or the other.", "Applied. REQ-PRJ-01 and REQ-PRJ-02 reworded. Both new types share the clinical period set and the same derivation - they differ in weights, not shape. RoleFactor and PeriodWeightStandard are now keyed on the type, so the split can carry real differences rather than being only a label. Schema 2 -> 3. All outputs regenerated.", "Applied"],
    ["R-04", "Data model", "Add at least one free-text note column to every sheet of the source workbook.", "Applied. Four sheets had none and each gains note_1: Milestone, ProjectPeriod, PeriodWeightStandard, Lists. Six already had one. Source schema version steps 1 -> 2. Note columns are carried through import and export unchanged and never read by the calculation, so the dummy dataset produces identical figures.", "Applied"],
]
r_start = r
r = table(ws, r, ["ID", "Topic", "Change requested", "How it was applied", "Status"],
          chg, [8, 14, 62, 66, 12], wrap_cols=(3, 4), mark_col=1)
for rr in range(r_start + 1, r_start + 1 + len(chg)):
    ws.cell(row=rr, column=5).fill = NEW_FILL
r = note(ws, r, "All four are settled. Nothing on this sheet is open, and the plan is final.")
r += 1
r = note(ws, r, "Six engineering decisions - C-06 to C-11 on sheet 05 - were proposed during the review rounds and "
                "never explicitly answered. Gate 1 approval of v1.0 confirmed them.")

# ---- 12 Review log --------------------------------------------------------
ws, r = sheet(wb, "12_Review_Log", "Review log and approval",
              "Every change made in v0.2, and why. Reviewer input was given as direct mark-up in "
              "PRAP_Development_Plan_v0.11_reviewed.xlsx rather than as cell comments.")

log = [
    ["1", "11_Open_Questions", "All 12 questions answered.", "Accepted in full.", "Propagated into sheets 02-10; see rows below.", "Closed"],
    ["2", "03_Requirements!A43", "New requirement added: 'After importing the source excel file, HTML to provide updating data and the updated data to be absorbed in the exported file when it extracts from the HTML.'", "Accepted. Numbered REQ-IMP-07.", "Added REQ-IMP-07, plus REQ-IMP-08 (unsaved-edit warning) and REQ-IMP-09 (re-validate edits), which it implies. Withdrew the v0.1 rule that nothing is editable (sheet 06). Added edit-buffer layer (sheet 07) and risk R-08.", "Closed"],
    ["3", "04_Data_Model - Project", "Added clinical_phase; EDC_setup, DataReviewSystem_setup, RBQM_setup, DM_conduct; EDC_system, DataReviewSystem, RBQM_system. Made outsourcing_type required.", "Accepted.", "All carried into the Project sheet with REQ-IDs REQ-PRJ-09/10/11. Value lists moved to the new Lists sheet. V-10 and V-11 added.", "Closed"],
    ["4", "04_Data_Model - Project", "outsourcing_type list reads 'Full outsourcing' / 'Partial outsourcing' / 'Full In-house' / 'Partial outsourcing'.", "Query - 'Partial outsourcing' appears twice.", "Flagged as Q-14. Resolved at round 2 - three values, see item 17.", "Closed"],
    ["5", "04_Data_Model - Milestone", "Milestone examples changed to FPI, LPI, LPO, LPLV, interim DB Lock, final DB Lock, CSR.", "Query - conflicts with the Q-07 answer, which names five different milestones.", "Q-07's list adopted as standard; raised as Q-15. Superseded at round 2 by a replacement list - see item 18.", "Closed"],
    ["6", "04_Data_Model", "Added project_name to Milestone and person_name to Assignment.", "Accepted with a safeguard.", "Kept as display conveniences, explicitly non-authoritative. V-13 checks them against the master row and refreshes rather than trusting them, so a stale copy cannot break a link.", "Closed"],
    ["7", "Q-09 answer", "One single workbook instead of two.", "Accepted.", "Sheet 04 restructured into one workbook of ten sheets; REQ-OUT-03, REQ-IMP-01/03 reworded; architecture option (b) rejection strengthened; deployment updated.", "Closed"],
    ["8", "Q-01 answer", "Three factors: project weight x role factor x person weight.", "Accepted, with one consequence.", "base_allocation removed - it was a fourth factor the answer did not name. Renamed to person_weight. PersonPeriodWeight changed from a multiplier to an override so no hidden fourth factor returns. Sheet 05 rewritten.", "Closed"],
    ["9", "Q-08 answer", "1 FTE = 160 h/month; over-allocation above 1.5 FTE; under-allocation below 0.8 FTE for 3+ months.", "Accepted.", "REQ-CAL-04 rewritten, REQ-CAL-07 and REQ-CAL-08 added. Under-allocation needs a rolling 3-month window, so it is specified as a run with a start and length, not a per-month flag. Config parameters added; sheet 06 flags both.", "Closed"],
    ["10", "Q-04 + Q-07 answers", "Standard periods and milestones named.", "Accepted, with a derivation.", "Periods are now derived from milestone dates (REQ-CAL-09) so a timeline change re-shapes them automatically. Mapping raised as Q-16 and answered at round 2 with a different mapping - see item 19. Assumption A-07 changed: milestones now drive weights.", "Closed"],
    ["11", "Q-06 answer", "Under 20 projects and 30 people.", "Accepted.", "REQ-NFR-03 target cut from 100 projects / 200 people / 1,000 assignments to a realistic figure with headroom; R-05 downgraded to Low/Low.", "Closed"],
    ["12", "Q-10 answer", "'Refer to 13_Templete_example'.", "Cannot action - the sheet is absent from the returned workbook.", "Verified against the file itself: it holds 13 sheets, 00 to 12, none hidden. Raised as Q-13; closed at round 2 - see item 16.", "Closed"],
    ["13", "02_Scope!A1", "Sheet title cell was blank in the returned file.", "Treated as accidental.", "Title 'Scope and objectives' restored.", "Closed"],
    ["14", "Q-03 answer", "Role lists differ between project types.", "Accepted.", "RoleFactor gains a project_type column; roles are matched on (project type, role name). V-03 raised from warning to error, since a role from the wrong type is now a real mistake rather than a typo.", "Closed"],
    ["15", "-", "Role factors and period weight values.", "Not supplied at round 1.", "Resolved at round 2 - see item 20.", "Closed"],

    ["16", "Q-13 answer (round 2)", "'No example file provided from me. Close this.'", "Accepted.", "Blocker cleared. WBS task 1.5 removed; assumption A-11 records that the schema here is taken as matching the existing one.", "Closed"],
    ["17", "Q-14 answer", "outsourcing_type has three values.", "Accepted.", "REQ-PRJ-04 and the Project sheet corrected.", "Closed"],
    ["18", "Q-15 answer", "Milestone list replaced - First SIV in, FPI out, DB lock cut-off and DB lock split interim/final.", "Accepted, with one query.", "Milestone list replaced throughout. The count query was answered at Q-24 - 'Protocol (v1)' is the eighth. See item 27.", "Closed"],
    ["19", "Q-16 answer", "Milestone-to-period mapping given for Start-up, Conduct and Close-out.", "Accepted in substance; three gaps found.", "Worked the mapping through a test project: it leaves 59 days before Start-up, 30 days between Conduct and Close-out, and 183 days after the interim DB lock with no period at all. A no-gap reading was adopted and the three points raised as Q-21..Q-23, all answered at round 3 - see items 24..26. REQ-CAL-10 and risk R-11 added so any residual gap is reported rather than silent.", "Closed"],
    ["20", "Q-17 answer", "'Actual numbers to be filled into the source file (excel).'", "Accepted - this closes the question rather than deferring it.", "The plan fixes where the values live, not what they are, which is what REQ-CAL-06 required all along. R-10 downgraded from High/High to Medium/Medium. Engine correctness is proven against the worked example; realism against your data at WBS 4.9.", "Closed"],
    ["21", "Q-18 answer", "'Others' projects use Planning / Develop / Close.", "Accepted.", "REQ-PRJ-12 added; ProjectPeriod and PeriodWeightStandard keyed by project_type; V-15 rejects a period from the wrong set. Q-25 confirmed at round 3 that 'Others' periods are entered by hand.", "Closed"],
    ["22", "Q-19 answer", "system_prepared_by can be dropped.", "Accepted.", "Column removed from the Project sheet.", "Closed"],
    ["23", "Q-20 answer", "Every field editable; validation rules prevent inconsistency.", "Accepted, with the mechanism made explicit.", "REQ-IMP-09 widened to all fields; REQ-IMP-10 and V-17 added for cascading identifier edits and refused deletes; decision C-09; risk R-12. Cascading is what makes an editable identifier safe - blocking the edit would have made it read-only in all but name.", "Closed"],

    ["24", "Q-21 answer (round 3)", "Before-Start-up runs from project start to one month before CTA submission.", "Accepted - confirms the v0.3 reading.", "Sequence 1 of the derivation on sheet 05, unchanged.", "Closed"],
    ["25", "Q-22 answer", "Close-out starts 3 months prior to DB lock and runs until DB lock.", "Accepted - and it reverses the v0.3 default.", "v0.3 had closed the one-month gap by shortening Close-out to two months. Your answer closes it the other way: Close-out is three months and Conduct ends the day before it. Corrected on sheet 05.", "Closed"],
    ["26", "Q-23 answer", "With an interim DB lock, split into Close-out (interim) and Close-out (final); the stretches either side of Close-out (interim) are both Conduct.", "Accepted. The structural change of v0.4.", "A period name is no longer unique within a project - Conduct now occurs twice. ProjectPeriod gains period_seq; REQ-CAL-11 added; V-18 enforces sequence uniqueness; the clinical-trial period set grows from four names to five. Derivation re-verified contiguous.", "Closed"],
    ["27", "Q-24 answer", "The eighth standard milestone is 'Protocol (v1)'.", "Accepted.", "Milestone list now holds eight. Only CTA submission and the DB locks are boundaries; the other six are markers.", "Closed"],
    ["28", "Q-25 answer", "'Others' projects have no milestone mapping.", "Accepted.", "Confirms the v0.3 assumption. A-10 moves from 'to confirm' to confirmed; their three periods are entered by hand.", "Closed"],
    ["29", "Q-26 answer", "A clinical trial's weight relies on clinical_phase.", "Accepted, with one consequence worth noting.", "PeriodWeightStandard re-keyed from project_category to clinical_phase. clinical_phase stops being descriptive and starts driving the simulation, so V-19 makes a missing phase an error rather than the warning it was under V-10. Raises Q-28: 'Others' projects have no phase.", "Closed"],
    ["30", "Sheet 05 - derivation", "Not reviewer input - a check made while applying the answers.", "Five timelines tested.", "The revised mapping was run against the normal case plus four degenerate ones: interim and final locks a month apart, no interim lock, CTA submission before project start, and a trial too short to have a Conduct phase. All five come out contiguous. Rules recorded as REQ-CAL-12 and decision C-11, and the omission cases documented on sheet 05.", "Closed"],
    ["31", "-", "Naming and 'Others' weighting.", "Two small points raised.", "Raised as Q-27 and Q-28; both answered at round 4 - see items 32 and 33.", "Closed"],

    ["32", "Q-27 answer (round 4)", "With no interim DB lock, the final DB lock falls into 'Close-out (final)'.", "Accepted - confirms v0.4 as built.", "No change needed. One period name covers both cases, so the weight lookup needs no special branch.", "Closed"],
    ["33", "Q-28 answer", "'Others' projects are not distinguished, and take manual input for period weights.", "Accepted, and it simplifies the model.", "PeriodWeightStandard becomes a clinical-trial reference table only; its '*' rows are gone. Combined with Q-25, 'Others' projects are now hand-entered throughout - both period dates and period weights. V-19 scoped to clinical trials. REQ-PRJ-06 reworded.", "Closed"],
    ["34", "Sheet 05 - decisions C-06..C-11", "Not reviewer input - six decisions proposed during the review rounds and never explicitly answered.", "Resolved at approval rather than by a fifth round.", "Gate 1 approval confirms C-06 (over/under-allocation judged on the person total), C-07 (period weights seeded then editable), C-08 (a month belongs to the period containing its first day), C-09 (identifier edits cascade), C-10 (boundaries recomputed on milestone change unless hand-edited) and C-11 (ordered clipping). Each is stated with its rationale on sheet 05.", "Closed"],

    ["35", "Gate 1", "Plan v1.0 approved by Dan, 2026-08-01, with direction to proceed to Step 2 starting from the workbook template and a dummy data file.", "Recorded.", "v1.0 is the approved baseline. Step 2 opened; template and dummy workbook issued for review as WBS task 2.0.", "Closed"],
    ["36", "11_Open_Questions (post-approval)", "Add 'Inspection' as a standard milestone.", "Accepted as change R-01.", "Milestone list grows to ten. 'Inspection' may repeat within a project, which no other milestone does - REQ-PRJ-13 and V-20 added, V-14 relaxed accordingly. Uniqueness is now on (project, name, date) rather than on name.", "Closed"],
    ["37", "05_Resource_Logic (post-approval)", "Derivation edits: Before-Start-up ends at 'Protocol (v1)'; Start-up starts the day after it and ends at 'First SIV' or 'FPI'; new seventh period 'After Close-out (final)' spanning the Inspection dates.", "Accepted as change R-02.", "Derivation table rewritten. Clinical period set 5 -> 6 names. REQ-CAL-13 added: where a boundary has both a milestone and an offset definition, the recorded milestone wins. 'FPI' restored to the milestone list - it had been dropped at Q-15, so this reverses that.", "Closed"],
    ["38", "05_Resource_Logic", "Not reviewer input - found while applying R-02.", "One case in the mark-up does not resolve.", "Period 7 is defined as earliest to latest 'Inspection'. An inspection dated on or before the final DB lock therefore makes period 7 start before period 6 and overlap Conduct. v1.1 counts only inspections after the final DB lock, treating earlier ones as markers (V-21). Verified contiguous across seven timelines. Raised as R-03 - see item 40.", "Closed"],
    ["39", "Version control", "v1.0 is an approved baseline.", "Changes cannot be absorbed silently.", "Per the rules on sheet 09, a change against a baseline takes the next point release and needs its own approval. The approval block on sheet 12 holds v1.0 (signed) and this change (awaiting signature).", "Closed"],
    ["40", "11_Open_Questions (v1.1 review)", "R-03 confirmed: an inspection on or before the final DB lock stays a marker and does not open the seventh period.", "Accepted - no content change needed.", "R-03 closed. The derivation, requirements and validation rules stand exactly as issued at v1.1; v1.2 records the confirmation and re-issues for signature. V-21 continues to report the case so it is visible rather than silent.", "Closed"],
    ["41", "-", "Status of the document as a whole.", "Nothing open.", "Across five review rounds: 28 questions and 3 change requests, all answered and applied. 65 requirements, 21 validation rules, 11 engineering decisions.", "Closed"],
    ["42", "Gate 1", "Plan v1.2 approved by Dan, 2026-08-01, with direction to continue Step 2.", "Recorded.", "v1.2 became the baseline and Step 1 closed. Programming specification v0.1 issued.", "Closed"],
    ["43", "Step 2 request", "Add at least one note column to every sheet of the source workbook.", "Accepted as change R-04.", "Four sheets gain note_1. Source schema version steps 1 -> 2, so the data model on sheet 04 and the specification's parse contract both change. Template and dummy workbooks regenerated at v1.2 and re-verified: identical calculation results, confirming the note columns are inert.", "Closed"],
    ["45", "Specification v0.3 review", "Six open points answered (S2-01 to S2-06).", "Accepted; three change the plan.", "S2-02 and S2-03 confirmed the draft with no change. S2-01 + S2-05 became R-07, S2-04 became R-08, S2-06 became R-06. The two calculation answers interact usefully: keeping thresholds absolute would have left the part-timer defect in place, but moving the floor to 0.60 removes it for every capacity in the data.", "Closed"],
    ["46", "Specification 03_Data_Schema!B26", "ProjectPeriod key changed from project_id + period_seq to project_id + period_name + period_start.", "Accepted.", "A natural key rather than a surrogate, and it disambiguates the two Conduct stretches just as well since their start dates differ. period_seq stays as the ordering column, and V-18 now checks both: the natural key for uniqueness, the sequence for deterministic ordering.", "Closed"],
    ["44", "Approval", "Plan v1.3 approved FINAL by Dan, 2026-08-01.", "Recorded.", "Step 1 is complete. The plan is closed at 65 requirements, 21 validation rules, 11 decisions and a version-2 source schema. Work proceeds under PRAP_Programming_Specification_v0.2.xlsx, whose sheet 10 carries the one decision still outstanding (S2-01) - a specification matter, not a plan matter.", "Closed"],
]
r_start = r
r = table(ws, r, ["No.", "Sheet / source", "Reviewer input", "Response", "Action taken in v0.2", "Status"],
          log, [6, 22, 56, 40, 76, 12], wrap_cols=(3, 4, 5))

dv3 = DataValidation(type="list", formula1='"Open,Accepted,Rejected,Deferred,Closed"', allow_blank=True)
ws.add_data_validation(dv3)
dv3.add(f"F{r_start + 1}:F{r_start + len(log)}")

r = section(ws, r, "Disposition summary")
disp = [
    ["Closed in v0.2", f'=COUNTIF(F{r_start + 1}:F{r_start + len(log)},"Closed")'],
    ["Open - awaiting your answer", f'=COUNTIF(F{r_start + 1}:F{r_start + len(log)},"Open")'],
    ["Total items", f"=COUNTA(A{r_start + 1}:A{r_start + len(log)})"],
]
r = table(ws, r, ["Measure", "Count"], disp, [30, 12])

r = section(ws, r, "Approval - Gate 1")
appr = [["PRAP Development Plan v1.0", "Dan", "2026-08-01",
         "APPROVED. Finalise plan v1.0 and proceed to Step 2, beginning with generation of the workbook "
         "template and a dummy data file for review."],
        ["PRAP Development Plan v1.2", "Dan", "2026-08-01",
         "APPROVED. Finalise plan v1.2 and continue Step 2. This issue supersedes v1.0 as the baseline; "
         "R-01, R-02 and R-03 are part of it."],
        ["PRAP Development Plan v1.3", "Dan", "2026-08-01",
         "APPROVED - FINAL. Change R-04 accepted; this issue supersedes v1.2 and closes the development "
         "plan. Step 1 is complete."],
        ["PRAP Development Plan v2.0", "Dan", "2026-08-02",
         "APPROVED - BASELINE. Development plan v2.0, programming specification v1.0 and UI component "
         "list v1.0 all approved together; Steps 1 to 3 are closed and Step 4, code generation, is "
         "authorised. This issue supersedes v1.3 as the baseline. Changes R-05 (project_type split), R-06 (volume re-baselined to "
         "100 projects / 1,000 people), R-07 (absolute thresholds, floor 0.60), R-08 (repeated period "
         "names distinguishable on screen), R-09 (component-list review: assumptions tab and insert-row "
         "become REQ-DSH-11 and REQ-IMP-11), R-10 (role factor keyed on type, phase, period and role) "
         "R-11 (conduct stretches named apart; ProjectPeriod keyed on project_id + period_name; "
         "schema 4 -> 5) and R-12 (two unimplemented validation rules on PersonPeriodWeight closed; "
         "V-24 added) and R-13 (component-list review: DB lock milestones emphasised, and a project "
         "utilisation graph added as REQ-DSH-12)."]]
r_start2 = r
r = table(ws, r, ["Document", "Approver", "Date", "Decision"], appr, [30, 16, 14, 88], wrap_cols=(4,))
for cc in (1, 2, 3, 4):
    ws.cell(row=r_start2 + 1, column=cc).fill = NEW_FILL
for cc in (1, 2, 3, 4):
    ws.cell(row=r_start2 + 2, column=cc).fill = NEW_FILL
for cc in (1, 2, 3, 4):
    ws.cell(row=r_start2 + 3, column=cc).fill = NEW_FILL
for cc in (2, 3):
    ws.cell(row=r_start2 + 4, column=cc).fill = INPUT_FILL
r = note(ws, r, "STEP 1 COMPLETE. v1.3 is the final development plan: the 65 requirements on sheet 03 are the "
                "contract for Steps 2-5, decisions C-01 to C-11 on sheet 05 are confirmed, and the source schema "
                "is version 2. The plan is closed; further change would need a new approval and a version "
                "increment, but none is expected. Work now proceeds under the programming specification.")

wb.save(OUT)
print(f"Written: {OUT}")
