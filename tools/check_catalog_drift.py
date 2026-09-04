#!/usr/bin/env python3
"""Fail the build when the rule catalog and the specification workbook disagree.

The Step 1 review chose to maintain the machine-readable catalog separately from
the specification workbook. That is workable only if divergence is detected, so
this check compares the contractual fields of every rule in:

  docs/spec/rule-catalog.yaml                              (engine input)
  docs/spec/TEA-SPEC-001_programming-specification.xlsx    (reviewed document)

and, when rule modules exist, that every catalog id has an implementation and
vice versa. Exit code 1 on any mismatch.

Usage:  python3 tools/check_catalog_drift.py [--rules-dir src/rules]
"""
import argparse
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
CATALOG = REPO / "docs" / "spec" / "rule-catalog.yaml"
WORKBOOK = REPO / "docs" / "spec" / "TEA-SPEC-001_programming-specification.xlsx"

FIELDS = ["family", "title", "guideline", "severity", "mode", "status",
          "confidence_base_rate"]


def load_catalog(path):
    """Minimal reader for the flat structure build_rule_catalog.py emits."""
    rules, cur = {}, None
    for line in path.read_text().splitlines():
        m = re.match(r"^  - id: (TE-[A-Z]{2}-\d{3})$", line)
        if m:
            cur = {"id": m.group(1)}
            rules[cur["id"]] = cur
            continue
        if cur is None:
            continue
        m = re.match(r"^    (\w+): (.*)$", line)
        if m and m.group(1) in FIELDS:
            val = m.group(2).strip()
            if val.startswith('"') and val.endswith('"'):
                val = val[1:-1].replace('\\"', '"').replace("\\n", "\n").replace("\\\\", "\\")
            cur[m.group(1)] = val
    return rules


def load_workbook_rules(path):
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True, data_only=True)["Rules"]
    header, rules = None, {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "ID":
            header = [str(c).strip() if c else "" for c in row]
            continue
        if not header or not row or not row[0]:
            continue
        if not re.fullmatch(r"TE-[A-Z]{2}-\d{3}", str(row[0])):
            continue
        d = dict(zip(header, row))
        rules[d["ID"]] = {
            "id": d["ID"],
            "family": d.get("Family"),
            "title": d.get("Title"),
            "guideline": d.get("Guideline"),
            "severity": d.get("Severity"),
            "mode": d.get("Mode"),
            "status": d.get("Status"),
            "confidence_base_rate": d.get("Conf. base"),
        }
    return rules


def norm(field, v):
    if v is None:
        return ""
    if field == "confidence_base_rate":
        return f"{float(v):.2f}"
    return re.sub(r"\s+", " ", str(v)).strip()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rules-dir", default=None,
                    help="Directory of implemented rule modules (checked once Step 5 begins)")
    args = ap.parse_args()

    problems = []

    if not CATALOG.exists():
        print(f"FAIL: catalog not found at {CATALOG}")
        return 1
    if not WORKBOOK.exists():
        print(f"FAIL: specification workbook not found at {WORKBOOK}")
        return 1

    cat = load_catalog(CATALOG)
    doc = load_workbook_rules(WORKBOOK)

    only_cat = sorted(set(cat) - set(doc))
    only_doc = sorted(set(doc) - set(cat))
    for rid in only_cat:
        problems.append(f"{rid}: in catalog but not in the specification workbook")
    for rid in only_doc:
        problems.append(f"{rid}: in the specification workbook but not in the catalog")

    for rid in sorted(set(cat) & set(doc)):
        for f in FIELDS:
            a, b = norm(f, cat[rid].get(f)), norm(f, doc[rid].get(f))
            if a != b:
                problems.append(f"{rid}.{f}: catalog={a!r} workbook={b!r}")

    if args.rules_dir:
        d = REPO / args.rules_dir
        if d.exists():
            impl = {p.stem.upper().replace("_", "-") for p in d.glob("TE_*.py")}
            for rid in sorted(set(cat) - impl):
                problems.append(f"{rid}: catalogued but not implemented")
            for rid in sorted(impl - set(cat)):
                problems.append(f"{rid}: implemented but not catalogued")

    if problems:
        print(f"CATALOG DRIFT — {len(problems)} problem(s):")
        for p in problems:
            print(f"  {p}")
        return 1

    print(f"catalog and specification workbook agree on all {len(cat)} rules")
    return 0


if __name__ == "__main__":
    sys.exit(main())
