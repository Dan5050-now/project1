"""Prove the pieces another AI system will use actually agree with the application.

The reference material tells an agent it can work in JSON, check its draft with
tools/prap_io.py, and trust that a file which passes there will load cleanly in
app/PRAP.html. That is a promise about two independent implementations - one in
JavaScript inside the page, one in Python on the command line - and a promise nobody
checks is a promise that quietly stops being true.

What is checked, on BOTH worked examples:

  1. xlsx -> json -> xlsx -> json reproduces every cell
  2. the browser loads the .prap.json and builds the SAME model as from the .xlsx
     (row counts per sheet, and every calculated person-month)
  3. the browser's findings and prap_io's findings are the same rules at the same
     severities - so 'validate' predicts what the user will see
  4. every calculated person-month agrees between browser and command line, exactly
  5. the application's own JSON export re-loads into an identical model
  6. a JSON file with a mistyped column name is REFUSED, not silently ignored,
     by both implementations
  7. the generated contract describes the real schema: its column lists match the
     template, the application's SHEET_HEADERS and prap_io's headers

    python tools/test_interop.py
"""

import json
import pathlib
import subprocess
import sys
import tempfile

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

ROOT = pathlib.Path(__file__).resolve().parents[1]
APP = (ROOT / "app" / "PRAP.html").as_uri()
CHROME = "/opt/pw-browsers/chromium-1194/chrome-linux/chrome"
FIXTURES = [ROOT / "templates" / "PRAP_SourceData_Dummy_v1.14.xlsx",
            ROOT / "templates" / "PRAP_SourceData_Dummy_10x10_v1.6.xlsx"]

sys.path.insert(0, str(ROOT / "tools"))
import prap_io                                                       # noqa: E402

TMP = pathlib.Path(tempfile.mkdtemp(prefix="prap_interop_"))
fails = []


def check(ok, label, detail=""):
    print(f"  {'ok  ' if ok else 'FAIL'} {label}{'   ' + detail if detail else ''}")
    if not ok:
        fails.append(label)


def sheets_of(pg):
    return pg.evaluate("() => { const o = {}; for (const s of REQUIRED_SHEETS) "
                       "o[s] = S.model.raw[s].length; return o; }")


def person_months(pg):
    return pg.evaluate("() => { const o = {}; for (const [k, v] of S.calc.persMonth) "
                       "o[k] = v; return o; }")


def findings(pg):
    return pg.evaluate("S.model.findings.map(f => f.sev + '|' + f.rule)")


def load(pg, path):
    pg.goto(APP)
    pg.wait_for_timeout(200)
    pg.set_input_files("#picker", str(path))
    pg.wait_for_timeout(4500)


def run(fixture, ctx):
    name = fixture.name
    print(f"app/PRAP.html  x  tools/prap_io.py  x  {name}")

    # ---- 1. round trip, entirely outside the browser ----------------------
    j1 = TMP / (fixture.stem + ".prap.json")
    x2 = TMP / (fixture.stem + "_rt.xlsx")
    j2 = TMP / (fixture.stem + "_rt.prap.json")
    prap_io.write_json(prap_io.read_xlsx(fixture), j1)
    prap_io.write_xlsx(prap_io.read_json(j1), x2)
    prap_io.write_json(prap_io.read_xlsx(x2), j2)
    a = json.loads(j1.read_text())["sheets"]
    b = json.loads(j2.read_text())["sheets"]
    diff = [s for s in a if a[s] != b[s]]
    check(not diff, "xlsx -> json -> xlsx -> json reproduces every cell",
          f"{sum(len(v) for v in a.values())} rows"
          + (f"; differs on {', '.join(diff)}" if diff else ""))

    # ---- 2/3/4. the browser, from each form ------------------------------
    pg = ctx.new_page()
    errors = []
    pg.on("pageerror", lambda e: errors.append(str(e)))

    load(pg, fixture)
    x_rows, x_pm, x_find = sheets_of(pg), person_months(pg), findings(pg)
    load(pg, j1)
    j_rows, j_pm, j_find = sheets_of(pg), person_months(pg), findings(pg)

    check(x_rows == j_rows, "the browser builds the same row counts from .xlsx and .prap.json",
          f"{sum(x_rows.values())} rows across {len(x_rows)} sheets"
          + ("" if x_rows == j_rows else f"  xlsx={x_rows}  json={j_rows}"))

    bad = [k for k in set(x_pm) | set(j_pm) if abs(x_pm.get(k, 0) - j_pm.get(k, 0)) > 1e-9]
    check(not bad, "every calculated person-month is identical from either form",
          f"{len(x_pm)} person-months" + (f"; {len(bad)} differ" if bad else ""))

    check(sorted(x_find) == sorted(j_find),
          "the same findings are raised from either form", f"{len(x_find)} findings")

    # prap_io must predict what the browser reports: same rules, same severities.
    M = prap_io.Model(prap_io.read_xlsx(fixture))
    cli = sorted(f"{f['sev']}|{f['rule']}" for f in M.findings)
    check(cli == sorted(x_find),
          "prap_io validate reports the same rules at the same severities as the app",
          f"cli {len(cli)} vs app {len(x_find)}"
          + ("" if cli == sorted(x_find)
             else f"\n        cli only: {sorted(set(cli) - set(x_find))[:4]}"
                  f"\n        app only: {sorted(set(x_find) - set(cli))[:4]}"))

    C = prap_io.calculate(M)
    ref = {f"{sid}|{k}": v for (sid, k), v in C["pers_month"].items()}
    missing = set(ref) ^ set(x_pm)
    bad = [k for k in set(ref) & set(x_pm) if abs(ref[k] - x_pm[k]) > 1e-6]
    check(not missing and not bad,
          "prap_io calculate matches the app on every person-month",
          f"{len(ref)} person-months"
          + (f"; {len(missing)} only on one side" if missing else "")
          + (f"; {len(bad)} differ, e.g. {sorted(bad)[0]}" if bad else ""))

    # ---- 5. the app's own JSON export --------------------------------------
    load(pg, fixture)
    with pg.expect_download() as dl:
        pg.click("#exportBtn")
        pg.click("#exportJsonBtn")
    out = TMP / (fixture.stem + "_exported.prap.json")
    dl.value.save_as(out)
    doc = json.loads(out.read_text())
    shaped = (doc.get("prap_format") == "prap-source-data"
              and doc.get("format_version") == 1
              and set(doc.get("sheets", {})) == set(prap_io.SHEET_ORDER))
    load(pg, out)
    e_rows, e_pm = sheets_of(pg), person_months(pg)
    bad = [k for k in set(x_pm) | set(e_pm) if abs(x_pm.get(k, 0) - e_pm.get(k, 0)) > 1e-9]
    check(shaped and e_rows == x_rows and not bad,
          "the app's own Export JSON re-loads into an identical model",
          f"{out.stat().st_size:,} bytes, {sum(e_rows.values())} rows"
          + ("" if shaped else "; wrong envelope")
          + (f"; {len(bad)} person-months differ" if bad else ""))

    # ---- 6. a mistyped column is refused by both --------------------------
    doc["sheets"]["Assignment"][0]["persn_weight"] = 0.5      # one letter short
    typo = TMP / (fixture.stem + "_typo.prap.json")
    typo.write_text(json.dumps(doc))
    try:
        prap_io.read_json(typo)
        py_refused, why = False, "prap_io accepted it"
    except prap_io.Problem as e:
        py_refused, why = "persn_weight" in str(e), str(e)[:70]
    pg.goto(APP)
    pg.wait_for_timeout(200)
    pg.set_input_files("#picker", str(typo))
    pg.wait_for_timeout(2500)
    banner = pg.inner_text("#banner")
    js_refused = "persn_weight" in banner and pg.eval_on_selector("#tabs", "e => e.hidden")
    check(py_refused and js_refused,
          "a mistyped column name is refused by both, and named",
          f"python: {why}" if not py_refused else f"both name it; app says “{banner.strip()[:70]}…”")

    check(not errors, "no uncaught errors in the page", "; ".join(errors[:2]))
    pg.close()


def contract_checks():
    print("docs/prap_contract.json  x  the schema it describes")
    contract = json.loads((ROOT / "docs" / "prap_contract.json").read_text())

    tpl = ROOT / "templates" / pathlib.Path(
        contract["documents"]["source_data_template"]).name
    wb = load_workbook(tpl)
    tpl_cols = {s: [c.value for c in wb[s][1] if c.value]
                for s in wb.sheetnames if s != "00_README"}
    c_cols = {s: [c["name"] for c in v["columns"]] for s, v in contract["sheets"].items()}
    bad = [s for s in c_cols if tpl_cols.get(s) != c_cols[s]]
    check(not bad, "the contract's columns are the template's columns",
          f"{sum(len(v) for v in c_cols.values())} columns across {len(c_cols)} sheets"
          + (f"; differ on {bad}" if bad else ""))

    app_hdr = json.loads(subprocess.run(
        [sys.executable, "-c",
         "import re,json,sys;s=open(sys.argv[1],encoding='utf-8').read();"
         "i=s.index('const SHEET_HEADERS = ');j=s.index('};',i);"
         "b=s[i+len('const SHEET_HEADERS = '):j+1];"
         "b=re.sub(r'([{,]\\s*)([A-Za-z_]\\w*)\\s*:',r'\\1\"\\2\":',b);"
         "print(json.dumps(json.loads(b.replace(',\\n}','\\n}').replace(',}','}'))))",
         str(ROOT / "app" / "PRAP.html")],
        capture_output=True, text=True, check=True).stdout)
    bad = [s for s in c_cols if app_hdr.get(s) != c_cols[s]]
    check(not bad, "the application's SHEET_HEADERS are the contract's columns",
          f"{len(app_hdr)} sheets" + (f"; differ on {bad}" if bad else ""))

    bad = [s for s in c_cols if prap_io.HEADERS.get(s) != c_cols[s]]
    check(not bad, "prap_io's headers are the contract's columns",
          f"{len(prap_io.HEADERS)} sheets" + (f"; differ on {bad}" if bad else ""))

    # A rule the plan documents but nothing enforces is a promise to an agent that the
    # application does not keep - unless the plan says the rule is RETIRED. A retired id
    # stays in the register deliberately: one that simply vanished would leave an agent
    # that had seen it with no way to learn what became of it.
    retired = [r for r, v in contract["validation_rules"].items()
               if not v["enforced_by_application"]
               and "RETIRED" in (v.get("statement") or "").upper()]
    unenforced = [r for r, v in contract["validation_rules"].items()
                  if not v["enforced_by_application"] and r not in retired]
    check(not unenforced,
          "every documented validation rule is enforced somewhere, or says it is retired",
          f"{len(contract['validation_rules'])} rules"
          + (f", {len(retired)} retired ({', '.join(retired)})" if retired else "")
          + (f"; not enforced: {', '.join(unenforced)}" if unenforced else ""))

    manifest = json.loads((ROOT / "docs" / "PRAP_Manifest.json").read_text())
    missing = [r["path"] for r in manifest["current"] if not (ROOT / r["path"]).exists()]
    check(not missing, "every file the manifest points at exists",
          f"{len(manifest['current'])} entries"
          + (f"; missing {', '.join(missing)}" if missing else ""))
    print()


if __name__ == "__main__":
    contract_checks()
    with sync_playwright() as pw:
        b = pw.chromium.launch(executable_path=CHROME, downloads_path=str(TMP))
        ctx = b.new_context(accept_downloads=True)
        for f in FIXTURES:
            run(f, ctx)
            print()
        b.close()
    print("FAILURES: " + (", ".join(fails) if fails else "none"))
    sys.exit(1 if fails else 0)
