"""Write the reference material another AI system reads before touching PRAP.

Four files, all generated, none hand-maintained:

    docs/prap_contract.json          the machine-readable contract - schema, value lists,
                                     Config, the formula, the period derivation, every
                                     validation rule, the interchange format, the recipes
    docs/PRAP_AI_Agent_Guide.md      the same contract written out for a reader, in the
                                     order an agent actually needs it
    docs/PRAP_AI_Agent_Guide_v1.0.xlsx   the guide as a workbook, so it sits with the
                                     other controlled documents and a person can review it
    docs/PRAP_Manifest.json          which file is current, and its sha256 - so an agent
                                     never has to guess which of 70-odd versioned
                                     workbooks is the one to read

Everything is derived from tools/prap_contract.py, which in turn reads the template
builder, the plan and the application. Nothing is retyped, so nothing can drift.

    python tools/build_ai_reference.py
"""

import hashlib
import json
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook                                        # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter                         # noqa: E402

import prap_contract as PC                                           # noqa: E402

ROOT = PC.ROOT
DOCS = ROOT / "docs"
GUIDE_VERSION = "1.0"
PLAN_PATH = DOCS / PC.PLAN
# The guide is issued against whichever plan is current at build time; before the new
# plan exists the previous issue still carries the rule table.
if not PLAN_PATH.exists():
    cands = sorted(DOCS.glob("PRAP_Development_Plan_v*.xlsx"),
                   key=lambda p: [int(x) for x in p.stem.split("_v")[1].split(".")])
    PLAN_PATH = cands[-1]


LABEL = {
    "development_plan": "Development plan",
    "programming_specification": "Programming specification",
    "ui_component_list": "UI component list",
    "source_data_template": "Source data template",
    "worked_example_large": "Worked example (62 projects, 20 people)",
    "worked_example_small": "Worked example (10 projects, 10 people)",
    "agent_guide": "This guide",
}


# ================================================================== Markdown
def md_table(headers, rows):
    out = ["| " + " | ".join(headers) + " |",
           "|" + "|".join("---" for _ in headers) + "|"]
    for r in rows:
        cells = [str(c).replace("|", "\\|").replace("\n", " ") for c in r]
        out.append("| " + " | ".join(cells) + " |")
    return "\n".join(out)


def guide_markdown(C):
    A, out = C["application"], []
    w = out.append

    w(f"# PRAP — reference for an AI agent")
    w("")
    w("**Project Resource Assignment Program.** A single-file offline HTML application "
      "that simulates monthly resource demand across simultaneously running projects — "
      "per project and per person. The data lives in an Excel workbook next to it; the "
      "workbook is the archive of record, and the application never writes to it.")
    w("")
    w("This document is written for a language model or an agent, not for a person. "
      "It tells you what you may change, what you must not, how to check your own work "
      "before handing it over, and what to say when you do. Its machine-readable half is "
      "`docs/prap_contract.json` — the same facts as JSON, generated from the same "
      "sources. Read whichever suits you; they cannot disagree.")
    w("")
    w(md_table(["", ""], [
        ["Application", f"`{A['file']}` v{A['version']}"],
        ["Source schema version", C["schema_version"]],
        ["Contract version", C["contract_version"]],
        ["Guide version", GUIDE_VERSION],
        ["Generated", date.today().isoformat()],
    ]))
    w("")
    w("---")
    w("")

    # ---- 0. the shortest useful thing ------------------------------------
    w("## 0. If you read nothing else")
    w("")
    w("```")
    w("# read a workbook as plain text")
    w("python tools/prap_io.py to-json  data.xlsx  -o data.prap.json")
    w("")
    w("# edit data.prap.json  (row objects keyed by column name; dates yyyy-mm-dd)")
    w("")
    w("# check it before anyone sees it")
    w("python tools/prap_io.py validate  data.prap.json")
    w("python tools/prap_io.py calculate data.prap.json --by person --flags")
    w("")
    w("# hand back a workbook")
    w("python tools/prap_io.py to-xlsx   data.prap.json -o data_draft.xlsx")
    w("```")
    w("")
    w("Seven rules that account for most mistakes:")
    w("")
    for line in [
        "**Never overwrite the user's workbook.** Always deliver a new file.",
        "**Never invent a value from a controlled list** — role names, period names, "
        "statuses, systems and phases all come from the `Lists` sheet (§4).",
        "**Never write a derived column** (§3.3). It is recomputed on load and your "
        "value is discarded.",
        "**One assignment per (person, project, role).** A changing workload is a "
        "`PersonPeriodWeight` window, not a second assignment (§5.3).",
        "**A weight override replaces the person weight** for the months it covers. It "
        "does not multiply it.",
        "**The allocation thresholds are absolute FTE**, not shares of a person's "
        "capacity (§5.4).",
        "**Validate before you deliver**, and say what you changed in rows.",
    ]:
        w(f"- {line}")
    w("")

    # ---- 1. what the thing is -------------------------------------------
    w("## 1. The system in one page")
    w("")
    w("```")
    w("  Excel workbook (.xlsx)  ──load──▶  app/PRAP.html  ──export──▶  new .xlsx")
    w("  or .prap.json (text)                in a browser                or .prap.json")
    w("                                     offline, no server")
    w("```")
    w("")
    w("- The application is **one HTML file**. Opening it in a browser is the whole "
      "installation. It makes no network requests, so a file dropped on it never "
      "leaves the machine.")
    w("- **There are two ways in**, and the second one changes what you may assume about "
      "a file you are handed:")
    for k, v in C["application"]["ways_in"].items():
        w(f"  - *{k.replace('_', ' ')}* — {v}")
    w("- **The workbook is the record.** Edits inside the application are provisional "
      "until Save, and are written nowhere until Export — which produces a *new* file "
      "stamped with the date (§6).")
    w("- **You cannot drive the application.** It has no API and no command line. What "
      "you can do is produce and check the file it reads, which is what "
      "`tools/prap_io.py` is for.")
    w("")
    w("### 1.1 Which file to read")
    w("")
    w("The repository keeps every issue of every document, so pick from "
      "`docs/PRAP_Manifest.json` rather than by sorting filenames. The current set:")
    w("")
    w(md_table(["What", "Path"], [[LABEL.get(k, k.replace("_", " ")), f"`{v}`"]
                                  for k, v in C["documents"].items()]))
    w("")

    # ---- 2. vocabulary ---------------------------------------------------
    w("## 2. The eight words you need")
    w("")
    w(md_table(["Term", "Meaning"], [
        ["Project", "A study or piece of work with a start, an end and a type. "
                    "`NewDrug CT` and `Biosimilar CT` are clinical trials; `Others` is "
                    "everything else, and the two kinds have different period sets and "
                    "different role lists."],
        ["Milestone", "A dated event in a project. Ten standard names. Two of them — "
                      "`CTA submission` and a DB lock — are what the period derivation "
                      "hangs on."],
        ["Period", "A named stretch of a project carrying an effort `weight`. A clinical "
                   "trial has up to seven; an `Others` project has three."],
        ["Person", "Someone who can be assigned, with a `capacity_fte`."],
        ["Assignment", "One person on one project in one role, between two dates, at a "
                       "`person_weight`."],
        ["Weight override", "A `PersonPeriodWeight` window that REPLACES `person_weight` "
                            "for the months it covers."],
        ["Role factor", "What one person in a role costs a project per month, before "
                        "their own weight. Keyed on type, phase, period and role."],
        ["FTE", f"The unit of everything. 1.00 FTE = "
                f"{C['config']['fte_hours_per_month']['default']} hours per month."],
    ]))
    w("")

    # ---- 3. the data ------------------------------------------------------
    w("## 3. The data model")
    w("")
    w("Ten sheets, all required, in this order. A missing sheet is fatal (V-00).")
    w("")
    w(md_table(["Sheet", "Role", "Parent", "Key", "Columns"],
               [[f"`{n}`", s["role"], f"`{s['parent_sheet']}`" if s["parent_sheet"] else "—",
                 ", ".join(f"`{k}`" for k in s["key"]), len(s["columns"])]
                for n, s in C["sheets"].items()]))
    w("")
    w("```")
    w("Project ──┬── Milestone")
    w("          ├── ProjectPeriod")
    w("          └── Assignment ── PersonPeriodWeight")
    w("Person  ──────┘")
    w("")
    w("PeriodFTEStandard, RoleFactor   reference tables, keyed on type/phase/period")
    w("Lists, Config                      vocabulary and settings")
    w("```")
    w("")
    w("### 3.1 Columns")
    w("")
    for name, s in C["sheets"].items():
        w(f"#### `{name}`")
        w("")
        rows = []
        for c in s["columns"]:
            t = c["type"]
            if c.get("value_list"):
                t += f" · list `{c['value_list']}`"
            if not c["editable"]:
                t += " · **do not write**"
            rows.append([f"`{c['name']}`", t, c["note"] or ""])
        w(md_table(["Column", "Type", "Meaning"], rows))
        w("")

    w("### 3.2 Foreign keys")
    w("")
    w(md_table(["From", "Must exist in", "Rule if it does not"],
               [[f"`{f['from']}`", f"`{f['to']}`", f["rule"]] for f in C["foreign_keys"]]))
    w("")
    w("Editing an identifier inside the application rewrites every row that references "
      "it. Deleting a row is refused while anything still references it, and the "
      "refusal names what does — a delete is never cascaded (V-17).")
    w("")

    w("### 3.3 Derived columns — read them, never write them")
    w("")
    derived = [(n, c) for n, s in C["sheets"].items() for c in s["derived_columns"]]
    w(md_table(["Sheet", "Column", "Recomputed from"], [
        ["`Project`", "`total_period_months`", "`end_date` − `start_date`, in whole months, inclusive"],
        ["`Milestone`", "`project_name`", "the `Project` row it points at"],
        ["`Assignment`", "`person_name`", "the `Person` row it points at"],
    ]))
    w("")
    w(f"({len(derived)} columns.) The application recomputes all three on every load, "
      "reports any disagreement as V-13, and then uses the master value. In the "
      "template these cells are locked and carry a comment saying so. Leave them out of "
      "a JSON file entirely — that is the clearest signal that you did not intend to "
      "set them.")
    w("")

    w("### 3.4 Dates")
    w("")
    for k, v in C["date_handling"].items():
        label = k.replace("_", " ")
        w(f"- **{label}** — {v if not isinstance(v, list) else ', '.join(v)}")
    w("")

    # ---- 4. vocabularies -------------------------------------------------
    w("## 4. Value lists")
    w("")
    w("These live on the `Lists` sheet of the workbook you are given — read them from "
      "there, because a company may have added values. The delivered set is:")
    w("")
    w(md_table(["List", "Values"],
               [[f"`{k}`", ", ".join(f"`{x}`" for x in v)] for k, v in C["value_lists"].items()]))
    w("")
    w("A value outside its list is kept and reported (V-11), never dropped — so a "
      "mistake of yours reaches the user as a warning rather than as missing data. That "
      "is not a licence to invent values.")
    w("")
    w("Note which list applies where: `role_clinical` for `NewDrug CT` and "
      "`Biosimilar CT`, `role_others` for `Others`; likewise `period_name_clinical` "
      "against `period_name_others`. Using a clinical role on an `Others` project is "
      "V-03, an error.")
    w("")

    w("## 5. How the numbers are produced")
    w("")
    w("### 5.1 The formula")
    w("")
    w("```")
    w(C["calculation"]["statement"])
    w("```")
    w("")
    w(f"Evaluated for {C['calculation']['evaluated_for']}, in FTE.")
    w("")
    w(md_table(["Term", "Where it comes from"],
               [[f"`{k}`", v] for k, v in C["calculation"]["terms"].items()]))
    w("")
    w(f"**Assignment window** — {C['calculation']['assignment_window']}")
    w("")
    w("**Aggregation**")
    w("")
    for a in C["calculation"]["aggregation"]:
        w(f"- {a}")
    w("")
    w("### 5.2 Worked example")
    w("")
    w("A person on `Phase 2` `NewDrug CT`, role `Lead data manager`, "
      "`person_weight` 0.40, assigned 2026-03-10 to 2026-12-31, in a month whose period "
      "carries weight 1.20 and whose role factor is 0.90:")
    w("")
    w("```")
    w("March 2026    coverage = (31 − 10 + 1) / 31 = 0.7097")
    w("              load     = 1.20 × 0.90 × 0.40 × 0.7097 = 0.3066 FTE")
    w("April 2026    coverage = 1.0")
    w("              load     = 1.20 × 0.90 × 0.40 × 1.0    = 0.4320 FTE")
    w("```")
    w("")
    w("Both endpoints are inclusive, which is why March counts 22 days and not 21.")
    w("")
    w("### 5.3 Periods")
    w("")
    w(md_table(["Project type", "Period set"],
               [["`NewDrug CT`, `Biosimilar CT`",
                 " → ".join(f"`{p}`" for p in C["period_sets"]["clinical"])],
                ["`Others`", " → ".join(f"`{p}`" for p in C["period_sets"]["others"])]]))
    w("")
    w(f"**Derivation.** {C['period_derivation']['when']}")
    w("")
    w(f"Requires: {', '.join(C['period_derivation']['requires'])}. "
      f"{C['period_derivation']['if_missing']}")
    w("")
    for i, s in enumerate(C["period_derivation"]["steps"], 1):
        w(f"{i}. {s}")
    w("")
    w(C["period_derivation"]["note"])
    w("")
    w("Periods within a project must leave **no gap and no overlap** (V-06, V-12). A "
      "month in no period is calculated at weight 1.00 and reported — which is a wrong "
      "number rather than a visible blank, so treat a V-12 warning as something to fix.")
    w("")
    w("### 5.4 Thresholds")
    w("")
    w(C["calculation"]["thresholds"])
    w("")
    w(md_table(["Parameter", "Default", "Controls"],
               [[f"`{k}`", v["default"], v["meaning"]] for k, v in C["config"].items()]))
    w("")
    w("Read the actual values from the `Config` sheet of the workbook in front of you; "
      "the table above is what the delivered template ships with.")
    w("")

    # ---- 6. the editing contract ----------------------------------------
    w("## 6. What the application does with a file")
    w("")
    for k, v in C["editing_contract"].items():
        label = k.replace("_", " ")
        if isinstance(v, list):
            w(f"- **{label}**")
            for x in v:
                w(f"  - {x}")
        else:
            w(f"- **{label}** — {v}")
    w("")
    w("The practical consequence for you: **a file you produce is never the file the "
      "user is looking at.** Say which file you wrote, and tell them to load it.")
    w("")

    # ---- 7. rules --------------------------------------------------------
    w("## 7. Validation rules")
    w("")
    w("The application collects every finding and shows them as one report; it never "
      "stops at the first. `python tools/prap_io.py validate <file>` produces the same "
      "list without a browser, and `--json` gives it to you as data.")
    w("")
    w("Severities: **fatal** nothing loads · **error** the figures would be wrong · "
      "**warning** the figures stand, the data is doubtful · **information** an "
      "explanation, not a problem.")
    w("")
    rows = []
    for rid, r in C["validation_rules"].items():
        sev = "/".join(r["severity_reported_by_app"]) or "—"
        rows.append([f"**{rid}**", sev, r["statement"]])
    w(md_table(["Rule", "Severity", "What it requires"], rows))
    w("")
    w("**Aim for zero errors and zero warnings you cannot explain.** A file that loads "
      "with errors still shows numbers, and those numbers are wrong in ways the user "
      "will not see.")
    w("")

    # ---- 8. interchange ---------------------------------------------------
    w("## 8. The JSON interchange format")
    w("")
    w(C["interchange_format"]["purpose"])
    w("")
    w(md_table(["Field", "Meaning"],
               [[f"`{k}`", v] for k, v in C["interchange_format"]["shape"].items()]))
    w("")
    w("```json")
    w(json.dumps(C["interchange_format"]["example"], indent=2))
    w("```")
    w("")
    w(f"- {C['interchange_format']['round_trip']}")
    w("- Omit a column rather than writing `null` — an absent key and an empty cell mean "
      "the same thing, and omission reads as deliberate.")
    w("- Every one of the ten sheets must be present, even where it has no rows.")
    w("- The application loads a `.prap.json` directly (drop it on the page) and its "
      "**Export JSON** button writes one, so text and workbook are interchangeable in "
      "both directions.")
    w("")

    # ---- 9. recipes ------------------------------------------------------
    w("## 9. Task recipes")
    w("")
    for i, r in enumerate(C["agent_recipes"], 1):
        w(f"### 9.{i} {r['task']}")
        w("")
        w(f"*{r['when']}*")
        w("")
        for j, s in enumerate(r["steps"], 1):
            w(f"{j}. {s}")
        w("")
        w("**Do not:**")
        for d in r["do_not"]:
            w(f"- {d}")
        w("")

    # ---- 10. checking ----------------------------------------------------
    w("## 10. Checking your work")
    w("")
    w(md_table(["To check", "Run"], [[k, f"`{v}`"] for k, v in C["verification"].items()]))
    w("")
    w("`prap_io.py` implements the same rules and the same formula as the application, "
      "and `tools/test_interop.py` proves the two agree — finding for finding and figure "
      "for figure — on both worked examples. If they ever stop agreeing, that test "
      "fails. So a file that passes `validate` is a file the application will load "
      "cleanly.")
    w("")

    # ---- 11. how to answer ------------------------------------------------
    w("## 11. How to report back")
    w("")
    w("- **Quantify in rows.** \"Added 4 `Assignment` rows and 1 `PersonPeriodWeight` "
      "window\" beats \"updated the plan\".")
    w("- **Quote FTE to two decimals**, and name the months you looked at.")
    w("- **Say what you assumed.** A `person_weight` you chose is a judgement, not a "
      "fact from the file. Mark it as yours.")
    w("- **Report the findings you left behind.** If `validate` still shows two "
      "warnings, say which and why they are acceptable.")
    w("- **Do not claim the dashboard shows something you did not check.** Run "
      "`calculate` and quote it.")
    w("- **Hand over a path.** Name the file you wrote and tell the user to open "
      "`app/PRAP.html` and drop it on the page.")
    w("")
    w("---")
    w("")
    w(f"Generated by `tools/build_ai_reference.py` on {date.today().isoformat()} from "
      f"`{A['file']}` v{A['version']}, `{PLAN_PATH.name}` and "
      f"`tools/build_source_workbook.py`. Do not edit by hand — rebuild it.")
    w("")
    return "\n".join(out)


# ================================================================== Workbook
FONT = "Arial"
NAVY = "1F3864"
HDR = PatternFill("solid", fgColor="2F5597")
BAND = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sheet(wb, name, title, intro, headers, rows, widths):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws["A2"] = intro
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    ws.row_dimensions[2].height = 30
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HDR
        c.border = BOX
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for r, row in enumerate(rows, start=5):
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = BOX
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if r % 2 == 1:
                c.fill = BAND
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.freeze_panes = "A5"
    return ws


def guide_workbook(C, path):
    wb = Workbook()
    wb.remove(wb.active)
    A = C["application"]

    ws = wb.create_sheet("00_Read_me_first")
    ws["A1"] = "PRAP — reference for an AI agent"
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
    lines = [
        "",
        "This workbook is the human-reviewable copy of docs/PRAP_AI_Agent_Guide.md.",
        "An AI agent should read the Markdown, or docs/prap_contract.json, which carry",
        "the same facts generated from the same sources.",
        "",
        f"Application            app/PRAP.html v{A['version']}",
        f"Source schema version  {C['schema_version']}",
        f"Contract version       {C['contract_version']}",
        f"Guide version          {GUIDE_VERSION}",
        f"Generated              {date.today().isoformat()}",
        "",
        "Seven rules that account for most mistakes:",
        "  1. Never overwrite the user's workbook — always deliver a new file.",
        "  2. Never invent a value from a controlled list; read the Lists sheet.",
        "  3. Never write a derived column; it is recomputed and your value discarded.",
        "  4. One assignment per (person, project, role); a changing workload is a",
        "     PersonPeriodWeight window, not a second assignment.",
        "  5. A weight override REPLACES the person weight — it does not multiply it.",
        "  6. The allocation thresholds are absolute FTE, not shares of capacity.",
        "  7. Validate before delivering, and say what changed, in rows.",
    ]
    for i, t in enumerate(lines, start=2):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(name="Consolas" if t.startswith(" ") or "  " in t else FONT, size=10)
    ws.column_dimensions["A"].width = 100

    sheet(wb, "01_Sheets", "The ten sheets",
          "All ten must be present. A missing sheet is fatal (V-00).",
          ["Sheet", "Role", "Parent", "Key", "Derived columns", "Columns"],
          [[n, s["role"], s["parent_sheet"] or "-", ", ".join(s["key"]),
            ", ".join(s["derived_columns"]) or "-", len(s["columns"])]
           for n, s in C["sheets"].items()],
          [22, 12, 14, 34, 24, 9])

    rows = []
    for n, s in C["sheets"].items():
        for c in s["columns"]:
            rows.append([n, c["name"], c["type"], "no" if not c["editable"] else "yes",
                         c.get("value_list", ""), c["note"]])
    sheet(wb, "02_Columns", "Every column",
          "'Editable' no means the column is derived: it is recomputed on load, any "
          "disagreement is reported as V-13, and the master value wins.",
          ["Sheet", "Column", "Type", "Editable", "Value list", "Meaning"],
          rows, [20, 24, 12, 10, 20, 70])

    sheet(wb, "03_Value_lists", "Value lists",
          "Read the actual values from the Lists sheet of the workbook you are given — "
          "a company may have added to them. A value outside its list is kept and "
          "reported (V-11), never dropped.",
          ["List", "Values"],
          [[k, ", ".join(v)] for k, v in C["value_lists"].items()], [26, 100])

    calc = C["calculation"]
    sheet(wb, "04_Calculation", "How the numbers are produced",
          calc["statement"] + "   —   " + calc["unit"],
          ["Term", "Where it comes from"],
          [[k, v] for k, v in calc["terms"].items()]
          + [["assignment window", calc["assignment_window"]],
             ["thresholds", calc["thresholds"]],
             ["rows excluded", calc["rows_excluded"]]],
          [24, 100])

    pd_ = C["period_derivation"]
    sheet(wb, "05_Periods", "Periods and their derivation", pd_["when"],
          ["#", "Step"],
          [["clinical set", " -> ".join(C["period_sets"]["clinical"])],
           ["others set", " -> ".join(C["period_sets"]["others"])],
           ["requires", ", ".join(pd_["requires"]) + " — " + pd_["if_missing"]]]
          + [[str(i), s] for i, s in enumerate(pd_["steps"], 1)]
          + [["note", pd_["note"]]],
          [14, 110])

    sheet(wb, "06_Validation_rules", "Validation rules",
          "The application collects every finding into one report. "
          "python tools/prap_io.py validate <file> produces the same list.",
          ["Rule", "Severity", "What it requires", "Effect"],
          [[rid, "/".join(r["severity_reported_by_app"]) or "-", r["statement"], r["effect"]]
           for rid, r in C["validation_rules"].items()],
          [9, 16, 74, 60])

    sheet(wb, "07_Config", "Config parameters",
          "Defaults as delivered. Read the real values from the Config sheet of the "
          "workbook in front of you.",
          ["Parameter", "Default", "Controls"],
          [[k, v["default"], v["meaning"]] for k, v in C["config"].items()], [30, 12, 90])

    ic = C["interchange_format"]
    sheet(wb, "08_Interchange", "The JSON interchange format", ic["purpose"],
          ["Field", "Meaning"],
          [[k, str(v)] for k, v in ic["shape"].items()]
          + [["round trip", ic["round_trip"]], ["converter", ic["converter"]],
             ["example", json.dumps(ic["example"], indent=1)]],
          [22, 100])

    rows = []
    for r in C["agent_recipes"]:
        rows.append([r["task"], "when", r["when"]])
        for i, s in enumerate(r["steps"], 1):
            rows.append(["", str(i), s])
        for d in r["do_not"]:
            rows.append(["", "do NOT", d])
    sheet(wb, "09_Recipes", "Task recipes",
          "What to actually do when the user asks for one of these.",
          ["Task", "#", "Step"], rows, [42, 8, 92])

    sheet(wb, "10_Verification", "Checking your work",
          "prap_io.py implements the same rules and formula as the application, and "
          "tools/test_interop.py proves the two agree on both worked examples.",
          ["To check", "Run"], [[k, v] for k, v in C["verification"].items()], [46, 74])

    sheet(wb, "10b_Ways_in", "How a plan gets into the application",
          "The second one changes what you may assume about a file you are handed.",
          ["", "What"],
          [[k.replace("_", " "), v] for k, v in C["application"]["ways_in"].items()],
          [26, 96])

    sheet(wb, "11_Documents", "Which file is current",
          "The repository keeps every issue. Pick from docs/PRAP_Manifest.json rather "
          "than by sorting filenames.",
          ["What", "Path"],
          [[LABEL.get(k, k.replace("_", " ")), v] for k, v in C["documents"].items()],
          [30, 70])

    wb.save(path)


# ================================================================== Manifest
def sha256(p):
    return hashlib.sha256(p.read_bytes()).hexdigest()


def manifest(C):
    items = []
    for what, rel in C["documents"].items():
        p = ROOT / rel
        items.append({
            "what": what,
            "path": rel,
            "exists": p.exists(),
            "bytes": p.stat().st_size if p.exists() else None,
            "sha256": sha256(p) if p.exists() else None,
        })
    for what, rel in [("application", "app/PRAP.html"),
                      ("machine_readable_contract", "docs/prap_contract.json"),
                      ("interchange_tool", "tools/prap_io.py"),
                      ("agent_guide_workbook",
                       f"docs/PRAP_AI_Agent_Guide_v{GUIDE_VERSION}.xlsx")]:
        p = ROOT / rel
        items.append({"what": what, "path": rel, "exists": p.exists(),
                      "bytes": p.stat().st_size if p.exists() else None,
                      "sha256": sha256(p) if p.exists() else None})
    return {
        "$comment": "The current issue of every PRAP artifact. The repository keeps "
                    "superseded versions alongside, so resolve a document through this "
                    "file rather than by sorting filenames. Regenerated by "
                    "tools/build_ai_reference.py.",
        "generated": date.today().isoformat(),
        "application_version": C["application"]["version"],
        "schema_version": C["schema_version"],
        "contract_version": C["contract_version"],
        "guide_version": GUIDE_VERSION,
        "start_here": ["docs/PRAP_AI_Agent_Guide.md", "docs/prap_contract.json"],
        "current": items,
    }


def main():
    C = PC.build(PLAN_PATH)
    DOCS.mkdir(exist_ok=True)

    contract = DOCS / "prap_contract.json"
    contract.write_text(json.dumps(C, indent=1, ensure_ascii=False) + "\n", encoding="utf-8")

    md = DOCS / "PRAP_AI_Agent_Guide.md"
    md.write_text(guide_markdown(C), encoding="utf-8")

    xl = DOCS / f"PRAP_AI_Agent_Guide_v{GUIDE_VERSION}.xlsx"
    guide_workbook(C, xl)

    mf = DOCS / "PRAP_Manifest.json"
    mf.write_text(json.dumps(manifest(C), indent=1, ensure_ascii=False) + "\n",
                  encoding="utf-8")

    print(f"built against {PLAN_PATH.name}, app v{C['application']['version']}, "
          f"schema v{C['schema_version']}")
    for p in (contract, md, xl, mf):
        print(f"  {p.relative_to(ROOT)}  {p.stat().st_size:,} bytes")
    missing = [r["path"] for r in manifest(C)["current"] if not r["exists"]]
    if missing:
        print("  NOTE - referenced but not present yet: " + ", ".join(missing))


if __name__ == "__main__":
    main()
