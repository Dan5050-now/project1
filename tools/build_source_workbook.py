"""Generate the PRAP source-data workbook: a blank template and a populated dummy file.

Both follow the schema baselined in PRAP_Development_Plan_v1.3.xlsx, sheet 04_Data_Model.

    python tools/build_source_workbook.py

Outputs into templates/:
    PRAP_SourceData_Template_v1.2.xlsx   headers, value lists, one example row per sheet
    PRAP_SourceData_Dummy_v1.3.xlsx      populated dataset that exercises every rule

The two carry different version numbers because the template's structure is unchanged
since v1.2; only the dummy dataset was regenerated.

The dummy files are generated deterministically (seeded), so every sheet rebuilds
byte-for-byte. The .xlsx as a whole does not: openpyxl stamps the build time into
docProps/core.xml, so the container differs while the data does not.

Two sizes are produced, from one generator driven by PROFILES:

    Dummy_v1.10        50 clinical trials + 12 'Others', 20 people - the review set
    Dummy_10x10_v1.2    8 clinical trials +  2 'Others', 10 people - small enough to
                        read every row and check the arithmetic by hand

Both are built to exercise the rules rather than merely fill cells: interim DB locks
that open a 'Conduct (interim)' stretch, inspections that open the final period,
hand-entered 'Others' periods, part-time capacities, multi-window weight overrides, and
both allocation thresholds crossed.
"""

import random
from datetime import date, timedelta
from pathlib import Path

from dateutil.relativedelta import relativedelta as rd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SCHEMA_VERSION = 6
TEMPLATE_VERSION = "1.8"
DUMMY_VERSION = "1.10"
DUMMY_SMALL_VERSION = "1.2"
OUTDIR = Path(__file__).resolve().parents[1] / "templates"

FONT = "Arial"
NAVY = "1F3864"
HDR_FILL = PatternFill("solid", fgColor="2F5597")
KEY_FILL = PatternFill("solid", fgColor="DDEBF7")   # identifier columns
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")  # derived - do not type here
FILL_ME = PatternFill("solid", fgColor="FFFF00")    # you must supply these
EG_FILL = PatternFill("solid", fgColor="F2F2F2")    # example row

HDR_F = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_F = Font(name=FONT, size=10)
BOLD_F = Font(name=FONT, size=10, bold=True)
TITLE_F = Font(name=FONT, size=14, bold=True, color=NAVY)
NOTE_F = Font(name=FONT, size=9, italic=True, color="808080")
EG_F = Font(name=FONT, size=10, italic=True, color="7F7F7F")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DATE_FMT = "yyyy-mm-dd"

# --------------------------------------------------------------------------
# Value lists (Lists sheet, long format per the approved data model)
# --------------------------------------------------------------------------
LISTS = [
    ("project_type", ["NewDrug CT", "Biosimilar CT (Healthy)", "Biosimilar CT (Patient)",
                      "Others"]),
    ("clinical_phase", ["Phase 1", "Phase 2", "Phase 3", "Phase 4"]),
    # Schema 6. How much of the work is done in-house decides how much of it lands on
    # this team, so it belongs in the weights rather than only in the description of a
    # project. Add a row inside the block to extend the list - the application reads
    # the list from the file, so a category of your own needs no code change.
    ("work_scope_type", ["fully in-housed", "fully outsourced",
                         "Partially outsourced (in-house for EDC)"]),
    ("outsourcing_type", ["Full outsourcing", "Partial outsourcing", "Full In-house"]),
    ("setup_party", ["by CRO", "by SB"]),
    ("EDC_system", ["Veeva EDC", "Rave", "eSOURCE"]),
    ("DataReviewSystem", ["Veeva DQS", "Medidata CDS", "No system (manual)"]),
    ("RBQM_system", ["CluePoints", "Medidata CDS", "No system (manual)"]),
    ("project_status", ["Planned", "Active", "On hold", "Completed"]),
    ("milestone_name", ["Protocol (v1)", "CTA submission", "FPI", "First SIV", "LPI",
                        "interim DB lock cut-off", "interim DB lock",
                        "final DB lock cut-off", "final DB lock", "Inspection"]),
    ("period_name_clinical", ["Before-Start-up", "Start-up", "Conduct (interim)",
                              "Close-out (interim)", "Conduct (final)",
                              "Close-out (final)", "After Close-out (final)"]),
    ("period_name_others", ["Planning", "Develop", "Close"]),
    ("role_clinical", ["Project oversight", "Lead data manager", "Clinical Data Associator",
                       "Clinical Database Programmer", "Data Analyst"]),
    ("role_others", ["Project lead", "Main staff", "Other staff"]),
]

CONFIG = [
    ("schema_version", SCHEMA_VERSION, "Structure version of this workbook. The application warns on a mismatch."),
    ("fte_hours_per_month", 160, "Hours equal to 1.00 FTE: 8 h/day x 5 days/week x 20 days/month."),
    ("over_allocation_fte", 1.50, "A person-month total above this is flagged as over-allocated. Absolute, not scaled by capacity (S2-01)."),
    ("under_allocation_fte", 0.60, "A person-month total below this counts toward an under-allocated run. Absolute, not scaled by capacity (S2-01)."),
    ("under_allocation_min_months", 3, "Consecutive months below the threshold before a run is flagged."),
    ("default_horizon_months", 24, "Months shown when the dashboard opens."),
    ("capacity_unit", "FTE", "Display unit: 'FTE' or 'percent'."),
]

# --------------------------------------------------------------------------
# Sheet definitions: (column, note, kind)
#   kind: 'key' identifier | 'calc' derived | 'fill' you must supply | '' normal
# --------------------------------------------------------------------------
SHEETS = {
    "Project": [
        ("project_id", "Unique key, e.g. PRJ-001.", "key"),
        ("project_name", "Unique display name.", "key"),
        ("project_type", "'NewDrug CT', 'Biosimilar CT (Healthy)', 'Biosimilar CT (Patient)' or 'Others'. Everything but 'Others' is a clinical trial.", ""),
        ("project_category", "Product name. Required for either clinical trial type.", ""),
        ("clinical_phase", "Required for any clinical trial type - with the type and the work scope it selects the period weights.", ""),
        ("work_scope_type", "How much of the work is done in-house. With the type and phase it selects the standard weights and role factors.", ""),
        ("outsourcing_type", "Full / Partial outsourcing, or Full In-house. Descriptive - work_scope_type is what the weights are keyed on.", ""),
        ("EDC_setup", "Who sets up EDC. Clinical trial types only.", ""),
        ("DataReviewSystem_setup", "Who sets up the data review system.", ""),
        ("RBQM_setup", "Who sets up RBQM.", ""),
        ("DM_conduct", "Who reviews the data.", ""),
        ("EDC_system", "EDC system in use.", ""),
        ("DataReviewSystem", "Data review system in use.", ""),
        ("RBQM_system", "RBQM system in use.", ""),
        ("planned_member_count", "Planned team size; compared against actual assignments.", ""),
        ("start_date", "Project start.", ""),
        ("end_date", "Planned project end.", ""),
        ("total_period_months", "DERIVED - formula, do not type.", "calc"),
        ("status", "Planned / Active / On hold / Completed.", ""),
        ("note_1", "Free text.", ""), ("note_2", "", ""), ("note_3", "", ""),
        ("note_4", "", ""), ("note_5", "", ""),
    ],
    "Milestone": [
        ("project_id", "Foreign key to Project.", "key"),
        ("project_name", "DERIVED - looked up from Project, do not type.", "calc"),
        ("milestone_name", "From the standard list of ten. 'Inspection' may appear on several rows.", ""),
        ("milestone_date", "Planned date.", ""),
        ("milestone_seq", "Display order on the timeline.", ""),
        ("note_1", "Free text. e.g. why a date moved, or which inspection body.", ""),
    ],
    "ProjectPeriod": [
        ("project_id", "Foreign key to Project.", "key"),
        ("period_name", "From the set for this project's type. UNIQUE within a project, so (project_id, period_name) is the key (R-11).", "key"),
        ("period_seq", "Orders periods along the timeline. Unique within a project.", ""),
        ("period_start", "Inclusive.", ""),
        ("period_end", "Inclusive. Periods must not overlap or leave a gap.", ""),
        ("weight", "Effort multiplier. Clinical trial types: seeded from PeriodWeightStandard. Others: type it.", ""),
        ("note_1", "Free text. e.g. why a derived date was overridden by hand.", ""),
    ],
    "PeriodWeightStandard": [
        ("project_type", "A clinical trial type. 'Others' projects take manual weights instead.", "key"),
        ("clinical_phase", "The phase this standard applies to.", "key"),
        ("work_scope_type", "The work scope this standard applies to. LEAVE EMPTY for a row that applies to EVERY scope - fill only the scopes that really differ.", "key"),
        ("period_name", "One of the seven clinical periods. Unique within a project (R-11).", "key"),
        ("weight", "YOU SUPPLY. Default multiplier for this phase and period.", "fill"),
        ("note_1", "Free text. e.g. the basis for this weight.", ""),
    ],
    "RoleFactor": [
        ("project_type", "Which type's role list this row belongs to.", "key"),
        ("clinical_phase", "The phase this factor applies to. Leave EMPTY for 'Others'.", "key"),
        ("work_scope_type", "The work scope this factor applies to. LEAVE EMPTY for a row that applies to EVERY scope - fill only the scopes that really differ.", "key"),
        ("period_name", "The period this factor applies to.", "key"),
        ("role_name", "The role.", "key"),
        ("role_factor", "YOU SUPPLY. Relative burden of this role in this period.", "fill"),
        ("role_note", "Basis for the factor.", ""),
    ],
    "Person": [
        ("person_id", "Unique key, e.g. PSN-001.", "key"),
        ("person_name", "Display name.", ""),
        ("department", "Grouping for the dashboard.", ""),
        ("primary_role", "Usual role; an assignment can override it.", ""),
        ("capacity_fte", "Available capacity. 1.00 = full time. Lower for a part-timer.", ""),
        ("employment_start", "Blank = open.", ""),
        ("employment_end", "Blank = open.", ""),
        ("note_1", "Free text.", ""), ("note_2", "", ""), ("note_3", "", ""),
        ("note_4", "", ""), ("note_5", "", ""),
    ],
    "Assignment": [
        ("assignment_id", "Unique key. One row per person + project + role.", "key"),
        ("person_id", "Foreign key to Person.", "key"),
        ("person_name", "DERIVED - looked up from Person, do not type.", "calc"),
        ("project_id", "Foreign key to Project.", "key"),
        ("role_name", "Must exist in RoleFactor for this project's type.", ""),
        ("assign_start_date", "Date the person joins the study.", ""),
        ("assign_end_date", "Date the person leaves. Blank = runs to project end.", ""),
        ("person_weight", "How much this person works on this project, e.g. 0.40.", ""),
        ("note_1", "Free text.", ""), ("note_2", "", ""), ("note_3", "", ""),
    ],
    "PersonPeriodWeight": [
        ("assignment_id", "Foreign key to Assignment.", "key"),
        ("period_start", "Inclusive.", ""),
        ("period_end", "Inclusive. Windows within one assignment must not overlap.", ""),
        ("weight_override", "REPLACES person_weight for these months - it does not multiply it.", ""),
        ("reason", "Why the weight differs.", ""),
    ],
    "Lists": [
        ("list_name", "Which list this value belongs to.", "key"),
        ("value", "A permitted value. Add a row inside the block to extend a list.", ""),
        ("note_1", "Free text. e.g. when a value was added, or what it means.", ""),
    ],
    "Config": [
        ("parameter", "Setting name.", "key"),
        ("value", "Setting value.", ""),
        ("note", "What it controls.", ""),
    ],
}

DATE_COLS = {
    "Project": ["start_date", "end_date"],
    "Milestone": ["milestone_date"],
    "ProjectPeriod": ["period_start", "period_end"],
    "Person": ["employment_start", "employment_end"],
    "Assignment": ["assign_start_date", "assign_end_date"],
    "PersonPeriodWeight": ["period_start", "period_end"],
}

# Dropdown bindings: sheet -> {column: list_name}
DROPDOWNS = {
    "Project": {"project_type": "project_type", "clinical_phase": "clinical_phase",
                "work_scope_type": "work_scope_type",
                "outsourcing_type": "outsourcing_type", "EDC_setup": "setup_party",
                "DataReviewSystem_setup": "setup_party", "RBQM_setup": "setup_party",
                "DM_conduct": "setup_party", "EDC_system": "EDC_system",
                "DataReviewSystem": "DataReviewSystem", "RBQM_system": "RBQM_system",
                "status": "project_status"},
    "Milestone": {"milestone_name": "milestone_name"},
    "PeriodWeightStandard": {"project_type": "project_type", "clinical_phase": "clinical_phase",
                             "work_scope_type": "work_scope_type",
                             "period_name": "period_name_clinical"},
    "RoleFactor": {"project_type": "project_type", "clinical_phase": "clinical_phase",
                   "work_scope_type": "work_scope_type"},
}


# --------------------------------------------------------------------------
# Period derivation - the algorithm baselined on plan sheet 05
# --------------------------------------------------------------------------
def derive_periods(start, end, milestones, inspections):
    """Return [(period_name, seq, start, end)] for a clinical trial - plan v1.1 sheet 05.

    A recorded milestone beats the month-offset fallback (REQ-CAL-13). Boundaries apply
    in sequence order, a later one winning; a period squeezed to nothing is omitted
    (REQ-CAL-12, decision C-11).
    """
    protocol = milestones.get("Protocol (v1)")
    cta = milestones.get("CTA submission")
    siv = milestones.get("First SIV") or milestones.get("FPI")
    idbl = milestones.get("interim DB lock")
    fdbl = milestones.get("final DB lock") or idbl
    if not cta or not fdbl:
        return []

    su_s = max((protocol + timedelta(days=1)) if protocol else (cta - rd(months=1)), start)
    su_e = siv if siv and siv >= su_s else (su_s + rd(months=4) - timedelta(days=1))

    # Period 7 opens only on inspection activity AFTER the final DB lock (V-21, R-03).
    later = [x for x in inspections if x > fdbl]
    p7_s = min(later) if later else None
    p7_e = max(max(later), end) if later else None

    cof_s = fdbl - rd(months=3)
    cof_e = (p7_s - timedelta(days=1)) if p7_s else max(fdbl, end)

    segs = []
    if su_s > start:
        segs.append(("Before-Start-up", start, su_s - timedelta(days=1)))
    segs.append(("Start-up", su_s, su_e))
    # R-11: the two Conduct stretches carry different names, so period_name is unique
    # within a project. 'Conduct (interim)' only exists where there IS an interim DB
    # lock and the stretch runs before it; everything else is 'Conduct (final)',
    # including the single stretch of a project that never has an interim lock.
    if idbl and idbl < fdbl:
        coi_s = idbl - rd(months=3)
        segs.append(("Conduct (interim)", su_e + timedelta(days=1), coi_s - timedelta(days=1)))
        segs.append(("Close-out (interim)", coi_s, idbl))
        cof_s = max(cof_s, idbl + timedelta(days=1))
        segs.append(("Conduct (final)", idbl + timedelta(days=1), cof_s - timedelta(days=1)))
    else:
        segs.append(("Conduct (final)", su_e + timedelta(days=1), cof_s - timedelta(days=1)))
    segs.append(("Close-out (final)", cof_s, cof_e))
    if p7_s:
        segs.append(("After Close-out (final)", p7_s, p7_e))

    rows = [(n, s, e) for n, s, e in segs if e >= s]
    return [(n, i + 1, s, e) for i, (n, s, e) in enumerate(rows)]


# --------------------------------------------------------------------------
# Dummy dataset
# --------------------------------------------------------------------------
def _extras_large(span):
    """The hand-placed rows of the 62-project set, kept as the literals they shipped as."""
    A = [
        ("ASG-901", "PSN-020", "PRJ-051", "Other staff", date(2026, 1, 1), date(2027, 6, 30), 0.18),
        ("ASG-902", "PSN-001", "PRJ-002", "Project oversight", date(2026, 4, 1), date(2027, 3, 31), 0.35),
        ("ASG-903", "PSN-001", "PRJ-006", "Project oversight", date(2026, 4, 1), date(2027, 3, 31), 0.35),
    ]
    # The key is (assignment_id, period_start) precisely because ONE assignment may
    # carry SEVERAL non-overlapping windows. ASG-902 exercises that: a spell of leave,
    # back to normal in between, then a peak. A fixture that only ever showed one window
    # per assignment left the multi-window path untested, which is how the missing
    # overlap and referential rules went unnoticed until R-12.
    ppw = [
        ("ASG-902", date(2026, 7, 1), date(2026, 9, 30), 0.20, "Part-time - parental leave"),
        ("ASG-902", date(2027, 4, 1), date(2027, 6, 30), 0.75, "Covering start-up peak"),
        ("ASG-903", date(2026, 10, 1), date(2026, 12, 31), 0.45, "Covering interim analysis peak"),
    ]
    return A, ppw


def _extras_small(span):
    """The same three shapes on the 10-project set, with the dates read off the projects.

    Hard-coded dates would be a hostage to the generated calendar: a project that starts
    a month later than the literal assumes turns the fixture into a V-07 warning about an
    assignment running past its project. Deriving them keeps the file clean whatever the
    seed produces.
    """
    def window(pid, frm, to):
        """A stretch inside a project's own dates, given as fractions of its span."""
        s, e = span[pid]
        total = (e - s).days
        a = (s + timedelta(days=int(total * frm))).replace(day=1)
        b = (s + timedelta(days=int(total * to))).replace(day=1) - timedelta(days=1)
        return max(a, s), min(b, e)

    # Both 'Others' projects, so the part-timer has a continuous but thin workload:
    # one short assignment on its own reads as a gap in the data rather than as the
    # under-allocation it is meant to show.
    o1_s, o1_e = window("PRJ-009", 0.05, 0.95)
    o2_s, o2_e = window("PRJ-010", 0.05, 0.95)
    p1_s, p1_e = window("PRJ-002", 0.10, 0.60)
    p2_s, p2_e = window("PRJ-006", 0.10, 0.60)
    A = [
        ("ASG-901", "PSN-010", "PRJ-009", "Other staff", o1_s, o1_e, 0.18),
        ("ASG-902", "PSN-001", "PRJ-002", "Project oversight", p1_s, p1_e, 0.35),
        ("ASG-903", "PSN-001", "PRJ-006", "Project oversight", p2_s, p2_e, 0.35),
        ("ASG-904", "PSN-010", "PRJ-010", "Other staff", o2_s, o2_e, 0.22),
    ]

    def third(s, e, k):
        """The k-th of three equal, non-overlapping windows inside s..e, month-aligned."""
        n = (e.year - s.year) * 12 + e.month - s.month + 1
        a = (s + rd(months=(n * k) // 3)).replace(day=1)
        b = (a + rd(months=max(1, n // 3))).replace(day=1) - timedelta(days=1)
        return max(a, s), min(b, e)

    w1s, w1e = third(p1_s, p1_e, 0)
    w2s, w2e = third(p1_s, p1_e, 2)
    w3s, w3e = third(p2_s, p2_e, 1)
    ppw = [
        ("ASG-902", w1s, w1e, 0.20, "Part-time - parental leave"),
        ("ASG-902", w2s, w2e, 0.75, "Covering start-up peak"),
        ("ASG-903", w3s, w3e, 0.45, "Covering interim analysis peak"),
    ]
    return A, ppw


PROFILES = {
    # The set shipped for Gate 4. Sized to the reviewer's original request rather than
    # to REQ-NFR-03's headroom figure - see the README note.
    "dummy": dict(
        version=DUMMY_VERSION, seed=20260801, n_ct=50, n_other=12,
        ct_base=date(2025, 1, 1), ot_base=date(2025, 6, 1),
        ct_spread=41, ot_spread=36, weight=(0.10, 0.17), lead_bonus=0.04,
        names=["Kim S.", "Park J.", "Lee H.", "Choi M.", "Jung Y.", "Han B.", "Oh K.",
               "Seo W.", "Yoon T.", "Nam R.", "Ahn D.", "Baek C.", "Shin E.", "Koo J.",
               "Ryu P.", "Moon A.", "Jang L.", "Hwang N.", "Cho V.", "Song G."],
        depts=(["Clinical Operations"] * 3 + ["Data Management"] * 4
               + ["Data Management"] * 5 + ["Programming"] * 4
               + ["Biostatistics"] * 3 + ["Business Systems"]),
        primary=(["Project oversight"] * 3 + ["Lead data manager"] * 4
                 + ["Clinical Data Associator"] * 5 + ["Clinical Database Programmer"] * 4
                 + ["Data Analyst"] * 3 + ["Other staff"]),
        capacity={20: 0.60, 4: 0.80},
        pools={
            "Project oversight":            ["PSN-001", "PSN-002", "PSN-003"],
            "Lead data manager":            ["PSN-004", "PSN-005", "PSN-006", "PSN-007"],
            "Clinical Data Associator":     ["PSN-008", "PSN-009", "PSN-010", "PSN-011", "PSN-012"],
            "Clinical Database Programmer": ["PSN-013", "PSN-014", "PSN-015", "PSN-016"],
            "Data Analyst":                 ["PSN-017", "PSN-018", "PSN-019"],
            "Project lead":                 ["PSN-001", "PSN-004"],
            "Main staff":                   ["PSN-008", "PSN-013", "PSN-017"],
            "Other staff":                  ["PSN-011", "PSN-016", "PSN-020"],
        },
        light="PSN-020", extras=_extras_large,
    ),
    # Ten projects and ten people: small enough to read every row on screen and check
    # the arithmetic by hand, which the 62-project set is not. Everything the larger set
    # exercises is still here - both clinical types, all four phases, trials with and
    # without an interim lock, inspections, hand-entered 'Others' periods, a part-timer,
    # multi-window overrides, and both allocation thresholds crossed.
    #
    # Two deliberate outliers: PSN-009 covers every trial as Data Analyst and goes over
    # the ceiling, PSN-010 carries one small assignment and sits under the floor. With
    # ten people and five clinical roles somebody has to double up, so the doubling is
    # placed where it demonstrates something rather than left to fall out of the sums.
    # The starts are packed into a narrower span than the large set's, because ten
    # projects spread over three and a half years would rarely overlap, and overlap is
    # the thing this application exists to show.
    "dummy_small": dict(
        version=DUMMY_SMALL_VERSION, seed=20260803, n_ct=8, n_other=2,
        # The short internal projects start later than the trials, so they are live
        # in the middle of the span rather than finished before it - a demo row that
        # reads 0.00 in the default window teaches nothing.
        ct_base=date(2025, 1, 1), ot_base=date(2026, 6, 1),
        ct_spread=12, ot_spread=10, weight=(0.22, 0.34), lead_bonus=0.05,
        # Someone whose pool is half the size carries twice as many trials, and would
        # not in practice hold the same share of each. Without this the analyst is over
        # the ceiling for the whole span, which demonstrates nothing.
        role_scale={"Data Analyst": 0.55},
        names=["Kim S.", "Park J.", "Lee H.", "Choi M.", "Jung Y.",
               "Han B.", "Oh K.", "Seo W.", "Yoon T.", "Nam R."],
        depts=(["Clinical Operations"] * 2 + ["Data Management"] * 4
               + ["Programming"] * 2 + ["Biostatistics"] + ["Business Systems"]),
        primary=(["Project oversight"] * 2 + ["Lead data manager"] * 2
                 + ["Clinical Data Associator"] * 2 + ["Clinical Database Programmer"] * 2
                 + ["Data Analyst"] + ["Other staff"]),
        capacity={10: 0.60, 4: 0.80},
        pools={
            "Project oversight":            ["PSN-001", "PSN-002"],
            "Lead data manager":            ["PSN-003", "PSN-004"],
            "Clinical Data Associator":     ["PSN-005", "PSN-006"],
            "Clinical Database Programmer": ["PSN-007", "PSN-008"],
            "Data Analyst":                 ["PSN-009"],
            "Project lead":                 ["PSN-001", "PSN-003"],
            "Main staff":                   ["PSN-005", "PSN-007"],
            "Other staff":                  ["PSN-006", "PSN-010"],
        },
        light="PSN-010", extras=_extras_small,
    ),
}


def dummy_data(prof):
    """Deterministic dataset built to a size PROFILE.

    Seeded so each file rebuilds byte-for-byte. The two profiles differ only in how
    many projects and people they hold and how the work is shared out; the shapes they
    exercise - interim locks, inspections, hand-entered 'Others' periods, part-time
    capacities, multi-window overrides, both allocation thresholds - are the same in
    both, because a small dataset that exercises nothing is not a smaller test, it is
    a weaker one.
    """
    rng = random.Random(prof["seed"])

    PRODUCTS = ["Onvelaris", "Cardexa", "Neurexa", "Renvia", "Hepatiq",
                "Immunex", "Osteva", "Pulmora", "Dermaline", "Glycora"]
    PHASES = ["Phase 1", "Phase 2", "Phase 3", "Phase 4"]
    CLINICAL_PERIODS = [v for k, v in LISTS if k == "period_name_clinical"][0]
    OTHER_PERIODS = [v for k, v in LISTS if k == "period_name_others"][0]
    OUTSOURCING = ["Full outsourcing", "Partial outsourcing", "Full In-house"]
    # Schema 6. Kept in step with OUTSOURCING at the same index so the fixture never
    # trips V-25, which reports a project whose two scope fields contradict each other.
    WORK_SCOPE = [v for k, v in LISTS if k == "work_scope_type"][0]
    SCOPE_FOR = {"Full outsourcing": "fully outsourced",
                 "Partial outsourcing": "Partially outsourced (in-house for EDC)",
                 "Full In-house": "fully in-housed"}
    PARTY = ["by CRO", "by SB"]
    EDC = ["Veeva EDC", "Rave", "eSOURCE"]
    DRS = ["Veeva DQS", "Medidata CDS", "No system (manual)"]
    RBQM = ["CluePoints", "Medidata CDS", "No system (manual)"]

    CT_TYPES = ["NewDrug CT", "Biosimilar CT (Healthy)", "Biosimilar CT (Patient)"]
    CT_ROLES = ["Project oversight", "Lead data manager", "Clinical Data Associator",
                "Clinical Database Programmer", "Data Analyst"]
    OT_ROLES = ["Project lead", "Main staff", "Other staff"]

    phase_profile = {
        # The two Conduct entries carry the SAME weight as the single 'Conduct' did
        # before R-11. Splitting the name should not silently reweight anything; if the
        # two stretches really do differ in burden, that is a data edit, not a default.
        "Phase 1": {"Before-Start-up": 0.60, "Start-up": 1.30, "Conduct (interim)": 1.00,
                    "Close-out (interim)": 1.20, "Conduct (final)": 1.00,
                    "Close-out (final)": 1.40, "After Close-out (final)": 0.84},
        "Phase 2": {"Before-Start-up": 0.70, "Start-up": 1.40, "Conduct (interim)": 1.10,
                    "Close-out (interim)": 1.30, "Conduct (final)": 1.10,
                    "Close-out (final)": 1.50, "After Close-out (final)": 0.90},
        "Phase 3": {"Before-Start-up": 0.80, "Start-up": 1.60, "Conduct (interim)": 1.20,
                    "Close-out (interim)": 1.40, "Conduct (final)": 1.20,
                    "Close-out (final)": 1.70, "After Close-out (final)": 1.02},
        "Phase 4": {"Before-Start-up": 0.50, "Start-up": 1.10, "Conduct (interim)": 0.90,
                    "Close-out (interim)": 1.00, "Conduct (final)": 0.90,
                    "Close-out (final)": 1.20, "After Close-out (final)": 0.72},
    }

    N_CT, N_OT = prof["n_ct"], prof["n_other"]
    OT_NAMES = ['CDISC library migration', 'eTMF rollout', 'EDC vendor evaluation',
                'SDTM automation', 'Risk dashboard build', 'Archive remediation',
                'CDR pilot', 'Metadata repository', 'eConsent rollout', 'Query analytics',
                'Training refresh', 'SOP revision']

    # ---- projects -------------------------------------------------------
    projects, ms, inspections = [], {}, {}

    for n in range(1, N_CT + 1):                             # clinical trials
        pid = f"PRJ-{n:03d}"
        phase = PHASES[(n - 1) % 4]
        product = PRODUCTS[(n - 1) % len(PRODUCTS)]
        start = prof["ct_base"] + rd(months=rng.randint(0, prof["ct_spread"]))
        months = rng.choice([18, 24, 30, 36, 42])
        end = start + rd(months=months) - timedelta(days=1)
        has_interim = (n % 5) in (0, 1, 2)                   # 60% carry an interim lock
        projects.append((
            # Type and scope move together on purpose, and the direction matters: it
            # gives the filter suite a pair that CAN co-occur (NewDrug CT with
            # 'Full In-house') and a pair that cannot (either biosimilar with it), so
            # "a combination that matches nothing" is a real case rather than a hope.
            pid, f"{product[:3].upper()}-{100 + n} {phase}", CT_TYPES[n % 3], product, phase,
            SCOPE_FOR[OUTSOURCING[2 - n % 3]],
            OUTSOURCING[2 - n % 3], PARTY[n % 2], PARTY[(n + 1) % 2], PARTY[n % 2],
            PARTY[(n + 1) % 2],
            EDC[n % 3], DRS[n % 3], RBQM[n % 3],
            rng.randint(3, 8), start, end,
            ["Planned", "Active", "Active", "Active", "On hold", "Completed"][n % 6],
        ))

        m = {
            "Protocol (v1)": start + rd(months=1),
            "CTA submission": start + rd(months=3),
            "First SIV": start + rd(months=7),
            "LPI": start + rd(months=int(months * 0.60)),
        }
        if has_interim:
            m["interim DB lock cut-off"] = start + rd(months=int(months * 0.62))
            m["interim DB lock"] = start + rd(months=int(months * 0.68))
        m["final DB lock cut-off"] = start + rd(months=months - 4)
        m["final DB lock"] = start + rd(months=months - 2)
        ms[pid] = m

        if n % 4 == 0:                                       # a quarter carry inspections
            base = m["final DB lock"]
            inspections[pid] = [base + rd(months=k) for k in (2, 5)][: 1 + (n % 2)]

    for n in range(N_CT + 1, N_CT + N_OT + 1):               # 'Others' projects
        pid = f"PRJ-{n:03d}"
        start = prof["ot_base"] + rd(months=rng.randint(0, prof["ot_spread"]))
        months = rng.choice([6, 9, 12, 18])
        projects.append((
            pid, f"{OT_NAMES[n - N_CT - 1]}",
            "Others", "", "", SCOPE_FOR[OUTSOURCING[n % 3]], OUTSOURCING[n % 3],
            "", "", "", "", "", "", "",
            rng.randint(2, 5), start, start + rd(months=months) - timedelta(days=1),
            ["Planned", "Active", "Active", "Completed"][n % 4],
        ))

    # ---- people ---------------------------------------------------------
    people = []
    for i, (nm, dept, role) in enumerate(zip(prof["names"], prof["depts"],
                                             prof["primary"]), start=1):
        cap = prof["capacity"].get(i, 1.00)                  # the rest are full-time
        people.append((f"PSN-{i:03d}", nm, dept, role, cap))

    # Pools sized so the trials spread evenly within each role. Internal ('Others')
    # projects draw on the same group - one team runs both.
    by_role = prof["pools"]
    LIGHT = prof["light"]

    # ---- assignments ----------------------------------------------------
    # Round-robin within each role pool, so no one carries three times their share.
    A = []
    cursor = {role: 0 for role in set(list(by_role) + CT_ROLES + OT_ROLES)}
    for p in projects:
        pid, ptype, pstart, pend = p[0], p[2], p[15], p[16]
        roles = OT_ROLES if ptype == "Others" else CT_ROLES
        for role in roles:
            pool = by_role.get(role) or by_role["Main staff"]
            person = pool[cursor[role] % len(pool)]
            cursor[role] += 1
            if person == LIGHT:                     # kept deliberately light
                person = pool[cursor[role] % len(pool)]
                cursor[role] += 1
            a_start = pstart + rd(months=rng.randint(0, 2))
            a_end = pend - rd(months=rng.randint(0, 2))
            if a_end <= a_start:
                a_end = pend
            # the more projects run at once, the smaller each person's share of any one
            w = round(rng.uniform(*prof["weight"]), 2)
            if role in ("Lead data manager", "Project lead"):
                w = round(w + prof["lead_bonus"], 2)
            w = round(w * prof.get("role_scale", {}).get(role, 1.00), 2)
            A.append((f"ASG-{len(A) + 1:03d}", person, pid, role, a_start, a_end, w))

    # The light person carries one small assignment, to demonstrate an under-allocation
    # run on a part-timer. PSN-001 is deliberately pushed over the ceiling for a year.
    # The dates are taken from the projects themselves so the fixture cannot drift into
    # a V-07 warning when the sizes change.
    span = {p[0]: (p[15], p[16]) for p in projects}
    extra, ppw = prof["extras"](span)
    A += extra

    # Keyed on type as well as phase: a biosimilar trial of a given phase is not the
    # same workload as a new-drug trial of that phase, and the split exists to say so.
    #
    # Schema 6 adds the WORK SCOPE to the key. The table is filled the way the
    # application expects it to be filled in real use: a base row per type, phase and
    # period with the scope column EMPTY - meaning "whatever the scope, unless a more
    # specific row says otherwise" - and specific rows only where the scope really does
    # change the number. Filling all three scopes for every combination would be 252
    # rows of which two thirds would repeat their neighbour.
    #
    # The illustrative shape: work kept in-house costs this team more, work handed to a
    # CRO costs it less, and the partial case is left to fall back on the base row so
    # the fallback is exercised by the fixture rather than only by a unit test.
    SCOPE_FACTOR = {None: 1.00, "fully in-housed": 1.15, "fully outsourced": 0.80}
    TYPE_FACTOR = {"NewDrug CT": 1.00, "Biosimilar CT (Healthy)": 0.85,
                   "Biosimilar CT (Patient)": 0.92}
    pws = []
    for ct in CT_TYPES:
        for scope, sf in SCOPE_FACTOR.items():
            note = ("Illustrative - replace with your figure" if scope is None else
                    f"Illustrative - {scope}")
            for ph, prof in phase_profile.items():
                for pn, w in prof.items():
                    pws.append((ct, ph, scope, pn,
                                round(w * TYPE_FACTOR[ct] * sf, 2), note))

    # A role's share of the work is not flat across a trial: the database programmer
    # is heaviest while the database is being built, the data associator while data is
    # coming in, the analyst at lock. That shape is what makes the factor worth keying
    # on the period as well as the role (R-10).
    ct_role_base = [("Project oversight", 0.80), ("Lead data manager", 1.20),
                    ("Clinical Data Associator", 1.00),
                    ("Clinical Database Programmer", 1.10), ("Data Analyst", 0.90)]
    # Indexed by CLINICAL_PERIODS: BSU, Start-up, Conduct (i), Close-out (i),
    # Conduct (f), Close-out (f), After Close-out. The two Conduct values match, for
    # the same reason the period weights do - R-11 renames, it does not reweight.
    role_shape = {
        "Project oversight":            [0.90, 1.00, 1.00, 1.00, 1.00, 1.00, 0.80],
        "Lead data manager":            [0.80, 1.20, 1.00, 1.10, 1.00, 1.20, 0.70],
        "Clinical Data Associator":     [0.50, 0.80, 1.30, 1.10, 1.30, 0.90, 0.50],
        "Clinical Database Programmer": [0.70, 1.50, 0.80, 1.00, 0.80, 1.10, 0.40],
        "Data Analyst":                 [0.40, 0.60, 0.90, 1.30, 0.90, 1.50, 0.90],
    }
    # Deliberately mild. Phase already drives PeriodWeightStandard, so a strong phase
    # term here would count the same effect twice.
    phase_role_mod = {"Phase 1": 0.95, "Phase 2": 1.00, "Phase 3": 1.05, "Phase 4": 0.90}
    type_role_mod = {"NewDrug CT": 1.00, "Biosimilar CT (Healthy)": 0.95,
                     "Biosimilar CT (Patient)": 0.98}

    # The same arrangement as PeriodWeightStandard above: base rows with an empty
    # scope, and scope-specific rows only for 'fully outsourced' - where the roles
    # shift as well as shrink, because oversight barely changes while the associator
    # and the programmer are largely doing somebody else's work.
    ROLE_SCOPE_MOD = {
        None: {r: 1.00 for r, _ in ct_role_base},
        "fully outsourced": {"Project oversight": 1.05, "Lead data manager": 0.95,
                             "Clinical Data Associator": 0.60,
                             "Clinical Database Programmer": 0.55, "Data Analyst": 0.85},
    }
    roles_tbl = []
    for ct in CT_TYPES:
        for scope, mods in ROLE_SCOPE_MOD.items():
            note = ("Illustrative - replace with your figure" if scope is None
                    else "Illustrative - fully outsourced")
            for ph in PHASES:
                for i, pn in enumerate(CLINICAL_PERIODS):
                    for rn, base in ct_role_base:
                        f = (base * role_shape[rn][i] * phase_role_mod[ph]
                             * type_role_mod[ct] * mods[rn])
                        roles_tbl.append((ct, ph, scope, pn, rn, round(f, 2), note))
    ot_shape = {"Project lead": [1.10, 1.00, 0.90], "Main staff": [0.70, 1.20, 0.90],
                "Other staff": [0.60, 1.10, 1.00]}
    ot_base = {"Project lead": 1.00, "Main staff": 0.90, "Other staff": 0.70}
    for i, pn in enumerate(OTHER_PERIODS):
        for rn in OT_ROLES:
            roles_tbl.append(("Others", None, None, pn, rn,
                              round(ot_base[rn] * ot_shape[rn][i], 2), "Illustrative"))

    # ---- periods --------------------------------------------------------
    periods = []
    for p in projects:
        pid, ptype, pstart, pend = p[0], p[2], p[15], p[16]
        if ptype != "Others":
            phase, scope = p[4], p[5]
            fac = TYPE_FACTOR[ptype] * SCOPE_FACTOR.get(scope, 1.00)
            for name, seq, s_, e_ in derive_periods(pstart, pend, ms[pid], inspections.get(pid, [])):
                periods.append((pid, name, seq, s_, e_,
                                round(phase_profile[phase][name] * fac, 2)))
        else:
            span = (pend - pstart).days
            b1 = pstart + timedelta(days=int(span * 0.25))
            b2 = pstart + timedelta(days=int(span * 0.80))
            periods.append((pid, "Planning", 1, pstart, b1, 0.80))
            periods.append((pid, "Develop", 2, b1 + timedelta(days=1), b2, 1.20))
            periods.append((pid, "Close", 3, b2 + timedelta(days=1), pend, 0.90))

    return projects, ms, periods, pws, roles_tbl, people, A, ppw, inspections


def describe(prof, projects, ms, periods, people, A, ppw, inspections):
    """The 'what is in this file' block, computed from the rows rather than typed.

    The counts in a description like this are exactly the kind that go stale the first
    time the data is regenerated, and nobody notices because the prose still reads well.
    """
    ct = [p for p in projects if p[2] != "Others"]
    ot = [p for p in projects if p[2] == "Others"]
    lo = min(p[15] for p in projects)
    hi = max(p[16] for p in projects)
    span = (hi.year - lo.year) * 12 + hi.month - lo.month + 1
    interim = sum(1 for p in ct if "interim DB lock" in ms[p[0]])
    part = [(s, cap) for s, _, _, _, cap in people if cap < 1.00]
    counted, roles = {}, {}
    for a in A:
        counted[a[1]] = counted.get(a[1], 0) + 1
        roles.setdefault(a[1], set()).add(a[3])
    busiest = max(counted, key=lambda k: counted[k])
    # The deliberate overload is the person the hand-placed rows (ASG-9xx) pile onto -
    # a fact about how the file was built, unlike 'who has the most rows', which is not
    # the same thing as who is over the ceiling and must not be stated as if it were.
    hand = [a[1] for a in A if a[0].startswith("ASG-9") and a[1] != prof["light"]]
    overloaded = max(set(hand), key=hand.count)
    n_light = counted.get(prof["light"], 0)
    windows = {}
    for w in ppw:
        windows[w[0]] = windows.get(w[0], 0) + 1
    phases = sorted({p[4] for p in ct})
    types = sorted({p[2] for p in ct})
    return [
        "",
        "WHAT IS IN THIS FILE",
        f"   {len(projects)} projects   {len(ct)} clinical trials "
        f"(PRJ-001..{len(ct):03d}) + {len(ot)} 'Others' "
        f"(PRJ-{len(ct)+1:03d}..{len(projects):03d})",
        f"   {len(people)} people     PSN-001..{len(people):03d}, including "
        + (", ".join(f"{s} at {c:.2f}" for s, c in part) or "no part-timers")
        + " capacity",
        f"   {len(A)} assignments, {sum(len(v) for v in ms.values()) + sum(len(v) for v in inspections.values())} "
        f"milestones, {len(periods)} periods, spanning {span} months "
        f"({lo:%b %Y} to {hi:%b %Y})",
        f"   Types: {', '.join(types)}.  Phases: {', '.join(phases)}.",
        "",
        "WHAT IT DEMONSTRATES",
        f"   - {interim} of the {len(ct)} trials carry an interim DB lock, so each has TWO 'Conduct'",
        f"     stretches and two close-outs. The other {len(ct) - interim} run Conduct once.",
        f"   - {len(inspections)} trials record an 'Inspection' after the final DB lock and so carry the",
        "     seventh period, 'After Close-out (final)'.",
        f"   - The {len(ot)} 'Others' projects have hand-entered periods and no milestones.",
        f"   - {overloaded} carries two extra oversight assignments on top of a full round-robin",
        "     share, and so crosses the 1.50 FTE ceiling at the peaks.",
        f"   - {busiest} is the busiest by row count, at {counted[busiest]} assignments - every trial, as "
        f"the only",
        f"     {sorted(roles[busiest])[0]}. Their share of each is set lower to match.",
        f"   - {prof['light']} is a part-timer carrying {n_light} light assignment"
        f"{'' if n_light == 1 else 's'}, so an under-allocation",
        "     run is raised. With the floor at 0.60 they CAN clear it, but only at full utilisation -",
        "     V-22 warns if anyone is recorded below the floor and so could never clear it.",
        f"   - {len(windows)} assignments carry PersonPeriodWeight overrides, one of them with "
        f"{max(windows.values())} separate",
        "     windows, which is why that sheet is keyed on assignment_id AND period_start.",
        "   - The months at either end of the span are naturally light: with few projects",
        "     overlapping there, most under-allocation runs sit in the ramp-up and wind-down.",
        "",
        "   The weights and role factors here are ILLUSTRATIVE, so the numbers can be exercised.",
        "   Replace them with your own before drawing any conclusion from the output.",
    ]


# --------------------------------------------------------------------------
# Writing
# --------------------------------------------------------------------------
def col_index(sheet, name):
    return [c for c, _, _ in SHEETS[sheet]].index(name) + 1


def write_sheet(wb, name, rows, example=None, list_ranges=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    cols = SHEETS[name]

    for i, (col, note, kind) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = HDR_F
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        if note:
            width = max(12, min(30, len(col) + 6))
        else:
            width = 12
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    start = 2
    if example is not None:
        for i, v in enumerate(example, start=1):
            c = ws.cell(row=start, column=i, value=v)
            c.font = EG_F
            c.fill = EG_FILL
            c.border = BOX
        start += 1

    for r, row in enumerate(rows, start=start):
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BODY_F
            c.border = BOX

    # column tinting by kind
    last = max(start + len(rows) - 1, start)
    for i, (col, note, kind) in enumerate(cols, start=1):
        fill = {"key": KEY_FILL, "calc": CALC_FILL, "fill": FILL_ME}.get(kind)
        if not fill:
            continue
        for r in range(2, last + 1):
            if ws.cell(row=r, column=i).value is not None or kind == "fill":
                ws.cell(row=r, column=i).fill = fill

    # date formatting
    for col in DATE_COLS.get(name, []):
        i = col_index(name, col)
        for r in range(2, last + 1):
            ws.cell(row=r, column=i).number_format = DATE_FMT

    # dropdowns
    if list_ranges:
        for col, list_name in DROPDOWNS.get(name, {}).items():
            rng = list_ranges.get(list_name)
            if not rng:
                continue
            dv = DataValidation(type="list", formula1=rng, allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv)
            letter = get_column_letter(col_index(name, col))
            dv.add(f"{letter}2:{letter}{max(last, 400)}")

    # ---- derived columns: locked, and said so ------------------------------
    #
    # A green fill is a convention the reader has to have been told about, and the
    # telling is on a sheet they may never open. Locking the cells makes the file itself
    # refuse the edit, at the moment it is attempted, with Excel's own message.
    #
    # No password. This is a GUARD RAIL, not security: anyone who genuinely needs to
    # paste a column can turn protection off in two clicks, and should be able to. What
    # it stops is the accidental type-over, which is the failure that actually happens -
    # a hand-typed total that then disagrees with the dates it was supposed to come from,
    # silently, because nothing recalculates it until the file is opened by something
    # that does.
    derived = [i for i, (_, _, k) in enumerate(cols, start=1) if k == "calc"]
    if derived:
        spare = max(last, 400)
        for i, (col, note, kind) in enumerate(cols, start=1):
            if kind == "calc":
                continue
            for r in range(2, spare + 1):
                ws.cell(row=r, column=i).protection = Protection(locked=False)
        for i in derived:
            col = cols[i - 1][0]
            note = (f"{col} is DERIVED.\n\n"
                    f"{cols[i - 1][1]}\n\n"
                    "The cell is locked so it cannot be typed over by accident. Nothing is "
                    "hidden: Review > Unprotect Sheet turns it off, there is no password. "
                    "But a value typed here is not recalculated, and the application "
                    "recomputes this column on import anyway - so a hand-typed one is "
                    "silently discarded, and any disagreement is reported as V-13.")
            c = ws.cell(row=1, column=i)
            c.comment = Comment(note, "PRAP")
            c.comment.width, c.comment.height = 340, 170
        # Everything a normal edit needs stays allowed; only the schema itself is fixed.
        ws.protection = SheetProtection(
            sheet=True, objects=False, scenarios=False,
            selectLockedCells=False, selectUnlockedCells=False,
            formatCells=False, formatColumns=False, formatRows=False,
            insertRows=False, deleteRows=False, sort=False, autoFilter=False,
            insertColumns=True, deleteColumns=True, insertHyperlinks=True)
    return ws


def add_readme(wb, kind, facts=None):
    ws = wb.create_sheet("00_README", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"PRAP source data - {'dummy' if kind in PROFILES else kind}"
    ws["A1"].font = TITLE_F
    ws.column_dimensions["A"].width = 120

    body = [
        "",
        f"Schema version {SCHEMA_VERSION}. Matches PRAP_Development_Plan_v1.0.xlsx, sheet 04_Data_Model.",
        "",
        "COLOUR KEY",
        "   Blue    identifier or key column - these link the sheets together.",
        "   Green   DERIVED. Do not type in these. They are LOCKED, so the sheet will refuse the",
        "           edit, and the column heading carries a note saying why. See below.",
        "   Yellow  you must supply this value. The application cannot work it out.",
        "   Grey italic row   an example, in the template only. Delete it before use.",
        "",
        "DERIVED COLUMNS ARE LOCKED",
        "   Three columns are computed, not entered:",
        "      Project.total_period_months   from start_date and end_date",
        "      Milestone.project_name        looked up from Project",
        "      Assignment.person_name        looked up from Person",
        "   Their cells are locked and every other cell on the sheet is not, so typing into one",
        "   is refused where it happens rather than going wrong later. There is NO PASSWORD:",
        "   Review > Unprotect Sheet turns it off if you genuinely need to paste a column.",
        "   Adding, deleting and sorting rows all still work with protection on; only adding or",
        "   removing COLUMNS is blocked, because the column set is the schema.",
        "   What the lock is protecting you from: a value typed into one of these is not",
        "   recalculated by anything, and the application recomputes the column on import in any",
        "   case - so a hand-typed one is discarded, and the disagreement is reported as V-13.",
        "",
        "SHEETS",
        "   Project               one row per project.",
        "   Milestone             one row per milestone. Eight standard names.",
        "   ProjectPeriod         the periods each project passes through, with their weights.",
        "   PeriodWeightStandard  default weights per project type, clinical phase, WORK SCOPE and",
        "                         period. Clinical trials only.",
        "   RoleFactor            the relative burden of each role, per project type, clinical phase,",
        "                         WORK SCOPE and period. Leave clinical_phase empty on the 'Others' rows.",
        "   Person                one row per person.",
        "   Assignment            one row per person + project + role.",
        "   PersonPeriodWeight    optional windows where a person's weight differs.",
        "   Lists                 permitted values for every dropdown.",
        "   Config                thresholds and settings.",
        "",
        "FREE-TEXT NOTES",
        "   Every sheet carries at least one free-text column, so anything worth recording has a home:",
        "      Project               note_1 .. note_5",
        "      Milestone             note_1",
        "      ProjectPeriod         note_1",
        "      PeriodWeightStandard  note_1",
        "      RoleFactor            role_note",
        "      Person                note_1 .. note_5",
        "      Assignment            note_1 .. note_3",
        "      PersonPeriodWeight    reason",
        "      Lists                 note_1",
        "      Config                note",
        "   Nothing in a note column affects the calculation. They are carried through import and export",
        "   unchanged, so they survive a round trip.",
        "",
        "HOW THE CALCULATION USES THIS",
        "   monthly load = project period weight  x  role factor  x  person weight  x  month coverage",
        "   The result is FTE, where 1.00 FTE = 160 hours per month.",
        "   Over-allocated:  a person's monthly total above 1.50 FTE.",
        "   Under-allocated: below 0.80 FTE for three or more consecutive months.",
        "",
        "PERIODS",
        "   Clinical trial periods are derived from the milestone dates:",
        "      Before-Start-up      project start, to the day before Start-up",
        "      Start-up             one month before CTA submission, lasting four months",
        "      Conduct              from the end of Start-up to the day before the next close-out",
        "      Close-out (interim)  three months before the interim DB lock, to that lock",
        "      Conduct  (again)     from the interim DB lock to the day before Close-out (final)",
        "      Close-out (final)    three months before the final DB lock, to that lock",
        "      After Close-out (final)  the inspection activity that follows the final DB lock",
        "",
        "   Where 'Protocol (v1)' is recorded, Before-Start-up ends on it and Start-up begins the day after.",
        "   Where 'First SIV' (or 'FPI') is recorded, Start-up ends on it instead of running a fixed four",
        "   months. A recorded milestone always beats the month-offset fallback.",
        "",
        "   'Conduct' therefore appears TWICE where a trial has an interim DB lock. period_seq keeps",
        "   the two apart. A period squeezed to nothing by a tight timeline is simply omitted.",
        "",
        "   'Inspection' is the one milestone that may be recorded SEVERAL TIMES for a project. Only",
        "   inspections dated AFTER the final DB lock open the last period; an earlier one stays a marker.",
        "",
        "   'Others' projects have no milestones. Their period dates and weights are both typed in.",
        "",
        "ADDING A PERMITTED VALUE",
        "   Insert a row inside that list's block on the Lists sheet, so the block stays contiguous -",
        "   the dropdowns read a range, not a scattered set of rows. work_scope_type is meant to be",
        "   extended this way: the three values supplied are a starting point, not a closed set.",
        "",
        "WORK SCOPE, AND THE EMPTY ROW THAT COVERS EVERY SCOPE",
        "   PeriodWeightStandard and RoleFactor are keyed on work_scope_type as well as on the project",
        "   type, the phase and the period. A row with work_scope_type EMPTY applies to EVERY scope.",
        "",
        "   So fill the empty-scope rows first - they are your baseline - and add a scope-specific row",
        "   only where that scope really does change the number. A project looks for its own scope",
        "   first and falls back to the empty row, so nothing has to be entered three times to say",
        "   the same thing three times.",
    ]
    if kind == "template":
        body += [
            "",
            "BEFORE YOU USE THIS FILE",
            "   1. Delete the grey example row from every sheet.",
            "   2. Fill in the yellow columns: PeriodWeightStandard.weight and RoleFactor.role_factor.",
            "      Nothing simulates correctly until those are set.",
            "   3. Enter your projects, people and assignments.",
        ]
    elif kind != "dummy":
        # Every count here is read off the rows that were actually generated, so the
        # description cannot drift from the file it describes.
        body += facts
    else:
        body += [
            "",
            "WHAT IS IN THIS FILE",
            "   62 projects   50 clinical trials (PRJ-001..050) + 12 'Others' (PRJ-051..062)",
            "   20 people     PSN-001..020, including two part-timers (PSN-004 at 0.80,",
            "                 PSN-020 at 0.60 capacity)",
            "   289 assignments, 372 milestones, 308 periods, spanning 73 months",
            "   Trials are spread evenly across Phase 1-4 and staggered from 2025 to 2031.",
            "",
            "WHAT IT DEMONSTRATES",
            "   - 30 trials carry an interim DB lock, so each has TWO 'Conduct' stretches and two",
            "     close-out periods. The other 20 run Conduct once.",
            "   - 12 trials record an 'Inspection' after the final DB lock and so carry the seventh",
            "     period, 'After Close-out (final)'.",
            "   - The 12 'Others' projects have hand-entered periods and no milestones.",
            "   - PSN-001 is pushed above the 1.50 FTE ceiling for a stretch in 2026-27.",
            "   - PSN-020 is a part-timer (0.60 FTE) carrying one light assignment, so an under-allocation run",
            "     is raised. With the floor now at 0.60 (S2-05) they CAN clear it, but only at full",
            "     utilisation - V-22 warns if anyone is recorded below the floor and so could never clear it.",
            "   - Two assignments carry PersonPeriodWeight overrides.",
            "",
            "   Typical load is around 0.93 FTE per person-month (median). Months at the very start",
            "   and end of the 73-month span are naturally light, since few projects overlap there.",
            "",
            "   The weights and role factors here are ILLUSTRATIVE, so the numbers can be exercised.",
            "   Replace them with your own before drawing any conclusion from the output.",
        ]

    for i, line in enumerate(body, start=2):
        c = ws.cell(row=i, column=1, value=line)
        if line.isupper() and line.strip():
            c.font = BOLD_F
        elif line.startswith("   ") or line.startswith("      "):
            c.font = Font(name="Consolas", size=9)
        else:
            c.font = BODY_F
    return ws


def build(kind):
    wb = Workbook()
    wb.remove(wb.active)

    # Lists first, so ranges are known for the dropdowns
    list_rows, list_ranges = [], {}
    r = 2
    for name, values in LISTS:
        for v in values:
            list_rows.append((name, v, None))
        list_ranges[name] = f"Lists!$B${r}:$B${r + len(values) - 1}"
        r += len(values)

    if kind in PROFILES:
        projects, ms, periods, pws, roles, people, A, ppw, inspections = dummy_data(PROFILES[kind])
        P = {p[0]: p[2] for p in projects}
        proj_rows = [list(p[:17]) + [None] + [p[17]] + [None] * 5 for p in projects]
        mile_rows = []
        for pid, mm in ms.items():
            events = sorted(mm.items(), key=lambda kv: kv[1])
            events += [("Inspection", x) for x in inspections.get(pid, [])]
            for seq, (nm, dt) in enumerate(sorted(events, key=lambda kv: kv[1]), start=1):
                mile_rows.append([pid, None, nm, dt, seq,
                                  "Regulatory inspection" if nm == "Inspection" else None])
        period_rows = [list(x) + [("Entered by hand - no milestone mapping"
                                   if P[x[0]] == "Others" else "Derived from milestones")]
                       for x in periods]
        pws_rows = [list(x) for x in pws]
        role_rows = [list(x) for x in roles]
        person_rows = [list(p) + [None, None] + [None] * 5 for p in people]
        asg_rows = [[a[0], a[1], None, a[2], a[3], a[4], a[5], a[6], None, None, None] for a in A]
        ppw_rows = [list(x) for x in ppw]
        examples = {k: None for k in SHEETS}
        facts = describe(PROFILES[kind], projects, ms, periods, people, A, ppw, inspections)
    else:
        facts = None
        proj_rows = mile_rows = period_rows = pws_rows = role_rows = []
        person_rows = asg_rows = ppw_rows = []
        # one example row per sheet (REQ-IMP-03)
        examples = {
            "Project": ["PRJ-001", "ONV-101 First-in-human", "NewDrug CT", "Onvelaris", "Phase 1",
                        "Partially outsourced (in-house for EDC)",
                        "Partial outsourcing", "by SB", "by SB", "by CRO", "by SB", "Veeva EDC",
                        "Veeva DQS", "CluePoints", 5, date(2025, 10, 1), date(2027, 6, 30), None,
                        "Active", "example row - delete before use", None, None, None, None],
            "Milestone": ["PRJ-001", None, "CTA submission", date(2026, 1, 15), 2, "example row - delete before use"],
            "ProjectPeriod": ["PRJ-001", "Start-up", 2, date(2025, 12, 15), date(2026, 4, 14), 1.30, "example row - delete before use"],
            "PeriodWeightStandard": ["NewDrug CT", "Phase 1", None, "Start-up", None,
                                     "example row - delete before use. Leave work_scope_type "
                                     "empty for a row that applies to every scope"],
            "RoleFactor": ["NewDrug CT", "Phase 1", None, "Start-up", "Lead data manager", None,
                           "example row - delete before use"],
            "Person": ["PSN-001", "Kim S.", "Data Management", "Lead data manager", 1.00,
                       None, None, "example row - delete before use", None, None, None, None],
            "Assignment": ["ASG-001", "PSN-001", None, "PRJ-001", "Lead data manager",
                           date(2025, 10, 1), date(2027, 6, 30), 0.45, "example row - delete before use", None, None],
            "PersonPeriodWeight": ["ASG-001", date(2026, 7, 1), date(2026, 9, 30), 0.20, "Part-time - parental leave"],
            "Lists": None,
            "Config": None,
        }

    add_readme(wb, kind, facts)
    write_sheet(wb, "Project", proj_rows, examples["Project"], list_ranges)
    write_sheet(wb, "Milestone", mile_rows, examples["Milestone"], list_ranges)
    write_sheet(wb, "ProjectPeriod", period_rows, examples["ProjectPeriod"], list_ranges)
    write_sheet(wb, "PeriodWeightStandard", pws_rows, examples["PeriodWeightStandard"], list_ranges)
    write_sheet(wb, "RoleFactor", role_rows, examples["RoleFactor"], list_ranges)
    write_sheet(wb, "Person", person_rows, examples["Person"], list_ranges)
    write_sheet(wb, "Assignment", asg_rows, examples["Assignment"], list_ranges)
    write_sheet(wb, "PersonPeriodWeight", ppw_rows, examples["PersonPeriodWeight"], list_ranges)
    write_sheet(wb, "Lists", list_rows, None, None)
    write_sheet(wb, "Config", [list(c) for c in CONFIG], None, None)

    # derived formulas
    ws = wb["Project"]
    ci = col_index("Project", "total_period_months")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value:
            ws.cell(row=row, column=ci,
                    value=f"=IF(OR($O{row}=\"\",$P{row}=\"\"),\"\","
                          f"(YEAR($P{row})-YEAR($O{row}))*12+MONTH($P{row})-MONTH($O{row})+1)")
    ws = wb["Milestone"]
    ci = col_index("Milestone", "project_name")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value:
            ws.cell(row=row, column=ci,
                    value=f'=IFERROR(INDEX(Project!$B:$B,MATCH($A{row},Project!$A:$A,0)),"")')
    ws = wb["Assignment"]
    ci = col_index("Assignment", "person_name")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value:
            ws.cell(row=row, column=ci,
                    value=f'=IFERROR(INDEX(Person!$B:$B,MATCH($B{row},Person!$A:$A,0)),"")')

    OUTDIR.mkdir(exist_ok=True)
    if kind == "template":
        suffix, ver = "Template", TEMPLATE_VERSION
    else:
        p = PROFILES[kind]
        # the size is in the file name, because two files both called 'Dummy' with
        # nothing to tell them apart is how the wrong one gets loaded
        suffix = "Dummy" if kind == "dummy" else f"Dummy_{p['n_ct'] + p['n_other']}x{len(p['names'])}"
        ver = p["version"]
    out = OUTDIR / f"PRAP_SourceData_{suffix}_v{ver}.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    for k in ("template", "dummy", "dummy_small"):
        print("Written:", build(k))
