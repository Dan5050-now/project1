"""The bridge between a PRAP source workbook and plain text, for programs and AI agents.

The application reads .xlsx, which is a ZIP of XML. A language model cannot write one,
and most agent runtimes cannot read one without a spreadsheet library. That made every
artifact in this project opaque to anything except a human with Excel. This tool closes
that gap in four commands:

    python tools/prap_io.py to-json  <file.xlsx>       -o <file.prap.json>
    python tools/prap_io.py to-xlsx  <file.prap.json>  -o <file.xlsx>
    python tools/prap_io.py validate <file>            [--json]
    python tools/prap_io.py calculate <file> --by person|project|cell
                                             [--from YYYY-MM --to YYYY-MM] [--flags] [--json]

`validate` and `calculate` implement the same rules and the same formula as
app/PRAP.html, so an agent can check its own draft without a browser.
tools/test_interop.py proves the two agree - finding for finding and figure for figure -
on both worked examples. If they ever stop agreeing, that test fails.

The JSON form is described in docs/prap_contract.json under "interchange_format" and in
docs/PRAP_AI_Agent_Guide.md. It is the whole workbook as row objects keyed by column
name, with dates as yyyy-mm-dd strings.
"""

import argparse
import calendar
import json
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import build_source_workbook as B                                      # noqa: E402

FORMAT_NAME = "prap-source-data"
FORMAT_VERSION = 1
SHEET_ORDER = list(B.SHEETS.keys())
HEADERS = {s: [c for c, _, _ in cols] for s, cols in B.SHEETS.items()}
DATE_COLS = {s: set(B.DATE_COLS.get(s, [])) for s in SHEET_ORDER}
NUM_COLS = {
    "Project": {"planned_member_count", "total_period_months"},
    "Milestone": {"milestone_seq"},
    "ProjectPeriod": {"period_seq", "weight"},
    "PeriodWeightStandard": {"standard_fte"},
    "RoleFactor": {"role_factor"},
    "Person": {"capacity_fte"},
    "Assignment": {"person_weight"},
    "PersonPeriodWeight": {"weight_override"},
    "MonthlyEstimate": {"fte"},
    "Lists": set(), "Config": set(),
}
class _ClinicalTypes:
    """Which project types are clinical trials.

    Schema 6 split 'Biosimilar CT' into '(Healthy)' and '(Patient)', and a fixed set of
    names would have to be edited by a programmer the next time a type is subdivided -
    for what is a change to a value list. So the question is asked of the NAME: anything
    beginning 'NewDrug CT' or 'Biosimilar CT' is a clinical trial and takes the seven
    clinical periods. `in` rather than a bare function so every call site reads as it
    did, and so this stays a mirror of the JavaScript rather than a variation on it.
    """

    PREFIXES = ("NewDrug CT", "Biosimilar CT")

    def __contains__(self, t):
        return str(t or "").startswith(self.PREFIXES)


CLINICAL_TYPES = _ClinicalTypes()

# Schema 6 retired this value. Named so a file carrying it gets a sentence saying what
# to do, rather than the generic "not a known value" - which is true, unhelpful, and
# identical to the message for a typo.
RETIRED_TYPES = {"Biosimilar CT": "Biosimilar CT (Healthy) or Biosimilar CT (Patient)"}

# Columns the schema RENAMED, and what they are now. A rename is the one schema change
# that loses data in silence - the old column is read into a key nothing looks at, the
# new one comes back empty, and every rule still passes.
RENAMED_COLS = {"Project": {"outsourcing_type": "outsourcing_scope_det"},   # schema 6 -> 7
                # schema 9 -> 10: it always held a monthly FTE, not a multiplier.
                "PeriodWeightStandard": {"weight": "standard_fte"}}

ANY_SCOPE = ""


def scope_of(row):
    """A row's work scope, with an empty one meaning 'every scope' (schema 6)."""
    return str(row.get("work_scope_type") or "") if row else ANY_SCOPE


def std_weight(M, proj, period_name):
    """The standard MONTHLY FTE: this project's own scope first, then the any-scope row.
    One function, so the calculation and V-19 cannot disagree. The phase is nulled for a
    non-clinical type exactly as std_factor does it - see stdWeight() in 05_model.js."""
    ph = proj.get("clinical_phase") if proj.get("project_type") in CLINICAL_TYPES else None
    k = (proj.get("project_type"), ph)
    v = M.pws.get((*k, scope_of(proj), period_name))
    return M.pws.get((*k, ANY_SCOPE, period_name)) if v is None else v


def std_factor(M, proj, period_name, role_name):
    """The role factor, the same way."""
    ph = proj.get("clinical_phase") if proj.get("project_type") in CLINICAL_TYPES else None
    k = (proj.get("project_type"), ph)
    v = M.rf.get((*k, scope_of(proj), period_name, role_name))
    return M.rf.get((*k, ANY_SCOPE, period_name, role_name)) if v is None else v


def absorbed_into(M, proj, period_name, role_name):
    """Which absent roles land on this one, if nobody holds them (REQ-CAL-16).
    The same two-step as std_factor, so a mapping written once on the baseline row
    covers every scope exactly as a factor written once does."""
    ph = proj.get("clinical_phase") if proj.get("project_type") in CLINICAL_TYPES else None
    k = (proj.get("project_type"), ph)
    v = M.rf_absorb.get((*k, scope_of(proj), period_name, role_name))
    if v is None:
        v = M.rf_absorb.get((*k, ANY_SCOPE, period_name, role_name))
    return v or []
CLINICAL_PERIODS = ["Before-Start-up", "Start-up", "Conduct (interim)", "Close-out (interim)",
                    "Conduct (final)", "Close-out (final)", "After Close-out (final)"]
OTHER_PERIODS = ["Planning", "Develop", "Close"]

# V-14's ordering half. Stated as PAIRS, not as one chain, because the ten milestone
# names are not totally ordered in real trials: an INTERIM database lock is precisely the
# one taken while recruitment continues, so it may fall either side of LPI, and a chain
# would report every such trial. Only the pairs below are ordered by definition.
MILESTONE_ORDER = [
    ("Protocol (v1)", "CTA submission"),
    ("CTA submission", "First SIV"),
    ("First SIV", "FPI"),
    ("FPI", "LPI"),
    ("interim DB lock cut-off", "interim DB lock"),
    ("interim DB lock", "final DB lock"),
    ("final DB lock cut-off", "final DB lock"),
    ("CTA submission", "final DB lock"),
]
# The derivation hangs on these three; out of order between them is an error, because the
# periods it computes would be wrong rather than merely surprising.
DERIVATION_MILESTONES = {"CTA submission", "interim DB lock", "final DB lock"}


class Problem(Exception):
    """A file we refuse to read, as opposed to a finding about data we did read."""


def _key(row, col):
    """Whether a row carries its identifier. Blank means it is not a record yet."""
    v = row.get(col)
    return not (v is None or str(v).strip() == "")


def no_key(M, sheet, row, col, owns):
    M.add("error", "V-08", sheet, row["__row"],
          f"A row on {sheet} has no {col}. Nothing can reference it, its {owns} have "
          f"nothing to attach to, and it would be lost when the file is read back. "
          f"Fill in {col} or delete the row.")


# =============================================================== 1. reading
def _as_date(v, where):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):                     # Excel serial, 1900 leap-year bug
        return date(1899, 12, 30) + timedelta(days=int(round(v)))
    s = str(v).strip()
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise Problem(f"{where}: '{s}' is not a date. Write yyyy-mm-dd - an ambiguous "
                      f"format such as 03/04/2026 is never guessed at.")


def _as_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except ValueError:
        return None


def _coerce(sheet, rows):
    out = []
    for i, r in enumerate(rows, start=2):
        rec = {}
        for col, v in r.items():
            if isinstance(v, str):
                v = v.strip() or None
            if col in DATE_COLS[sheet]:
                v = _as_date(v, f"{sheet} row {i}, {col}")
            elif col in NUM_COLS[sheet]:
                v = _as_num(v)
            rec[col] = v
        rec["__row"] = i
        out.append(rec)
    return out


def read_xlsx(path):
    """Workbook -> {sheet: [row dict]}. Blank rows are skipped, as the application does."""
    wb = load_workbook(path, data_only=True)
    # MonthlyEstimate arrived at schema 9; a file written before it simply carries no
    # manual figures, and refusing it would gain nothing. Every other sheet IS the plan.
    missing = [s for s in SHEET_ORDER
               if s not in wb.sheetnames and s != "MonthlyEstimate"]
    if missing:
        raise Problem(f"{Path(path).name}: sheet(s) not found: {', '.join(missing)}. "
                      f"Compare the file against templates/PRAP_SourceData_Template_"
                      f"v{B.TEMPLATE_VERSION}.xlsx.")
    sheets = {}
    for s in SHEET_ORDER:
        if s not in wb.sheetnames:
            sheets[s] = []
            continue
        ws = wb[s]
        renamed = RENAMED_COLS.get(s, {})
        hdr = [renamed.get(c.value, c.value) for c in ws[1]]
        rows = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or v == "" for v in raw):
                continue
            rows.append({h: v for h, v in zip(hdr, raw) if h})
        sheets[s] = _coerce(s, rows)
    return sheets


def read_json(path):
    doc = json.loads(Path(path).read_text(encoding="utf-8"))
    if doc.get("prap_format") != FORMAT_NAME:
        raise Problem(f"{Path(path).name}: not a PRAP interchange file — expected "
                      f'"prap_format": "{FORMAT_NAME}".')
    if doc.get("format_version") != FORMAT_VERSION:
        raise Problem(f"{Path(path).name}: format_version {doc.get('format_version')!r}; "
                      f"this build writes and reads {FORMAT_VERSION}.")
    src = doc.get("sheets") or {}
    # MonthlyEstimate arrived at schema 9, and a file written before it carries no manual
    # figures - a complete plan, not a broken one. The same tolerance read_xlsx extends,
    # for the same reason. Every other sheet has always been there and is still required.
    missing = [s for s in SHEET_ORDER if s not in src and s != "MonthlyEstimate"]
    if missing:
        raise Problem(f"{Path(path).name}: sheets missing: {', '.join(missing)}. All "
                      f"{len(SHEET_ORDER)} must be present, even if empty.")
    sheets = {}
    for s in SHEET_ORDER:
        rows = src.get(s) or []
        renamed = RENAMED_COLS.get(s, {})
        for r in rows:
            for was, now in renamed.items():
                if was in r:
                    r.setdefault(now, r[was])
                    del r[was]
        for i, r in enumerate(rows, start=1):
            # A column name that is not in the schema is refused rather than ignored: a
            # typo silently dropped is a value the user believes they supplied.
            bad = [k for k in r if k not in HEADERS[s]]
            if bad:
                raise Problem(f"{Path(path).name}: {s} row {i} has unknown column(s) "
                              f"{', '.join(sorted(bad))}. Valid: {', '.join(HEADERS[s])}.")
        sheets[s] = _coerce(s, [dict(r) for r in rows])
    return sheets


def read_any(path):
    p = Path(path)
    if p.suffix.lower() == ".json":
        return read_json(p)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        return read_xlsx(p)
    raise Problem(f"{p.name}: expected .xlsx or .prap.json.")


# =============================================================== 2. writing
def _plain(v):
    if isinstance(v, datetime):
        v = v.date()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def to_json_doc(sheets):
    out = {}
    for s in SHEET_ORDER:
        rows = []
        for r in sheets[s]:
            rec = {}
            for col in HEADERS[s]:
                v = _plain(r.get(col))
                if v is not None and v != "":
                    rec[col] = v
            rows.append(rec)
        out[s] = rows
    return {
        "prap_format": FORMAT_NAME,
        "format_version": FORMAT_VERSION,
        "schema_version": B.SCHEMA_VERSION,
        "$comment": "PRAP source data as plain text. Load into app/PRAP.html directly, or "
                    "convert with tools/prap_io.py. Dates are yyyy-mm-dd. Column meanings "
                    "and rules: docs/prap_contract.json.",
        "sheets": out,
    }


def write_json(sheets, path):
    Path(path).write_text(json.dumps(to_json_doc(sheets), indent=1, ensure_ascii=False) + "\n",
                          encoding="utf-8")


def write_xlsx(sheets, path):
    """A minimal, unstyled workbook with the schema's sheets and column order.

    Deliberately not the styled template: this is a data carrier, and the application
    reads only the values. Dates are written as real dates so Excel shows them as dates.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for s in SHEET_ORDER:
        ws = wb.create_sheet(s)
        ws.append(HEADERS[s])
        for r in sheets[s]:
            ws.append([r.get(c) for c in HEADERS[s]])
        for col in DATE_COLS[s]:
            i = HEADERS[s].index(col) + 1
            for cell in ws[get_column_letter(i)][1:]:
                cell.number_format = "yyyy-mm-dd"
        ws.freeze_panes = "A2"
    wb.save(path)


# =============================================================== 3. the model
def month_key(y, m):
    return y * 12 + m


def key_label(k):
    return f"{calendar.month_abbr[k % 12 + 1]} {k // 12}"


def key_ym(k):
    return k // 12, k % 12 + 1


def parse_ym(s):
    y, m = s.split("-")
    return month_key(int(y), int(m) - 1)


def months_between(a, b):
    y, m = a.year, a.month
    while (y, m) <= (b.year, b.month):
        yield y, m
        m += 1
        if m == 13:
            y, m = y + 1, 1


def coverage(y, m, s, e):
    days = calendar.monthrange(y, m)[1]
    lo, hi = max(date(y, m, 1), s), min(date(y, m, days), e)
    return 0.0 if hi < lo else ((hi - lo).days + 1) / days


def add_months(d, n):
    y, m = d.year, d.month + n
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    return date(y, m, min(d.day, calendar.monthrange(y, m)[1]))


def derive_periods(proj, ms):
    """The clinical period derivation, identical to derivePeriods() in the application."""
    get = lambda n: (ms.get(n) or [None])[0]                            # noqa: E731
    protocol, cta = get("Protocol (v1)"), get("CTA submission")
    siv = get("First SIV") or get("FPI")
    idbl = get("interim DB lock")
    fdbl = get("final DB lock") or idbl
    if not cta or not fdbl:
        return None
    # The project's own dates are OPTIONAL here, and a blank one is not an error - see
    # the long note in derivePeriods() in src/core/04_derive.js. Both are floors and
    # nothing more; absent, the milestones alone describe the run.
    start, end = proj.get("start_date"), proj.get("end_date")
    su_s0 = (protocol + timedelta(days=1)) if protocol else add_months(cta, -1)
    su_s = max(su_s0, start) if start else su_s0
    su_e = siv if (siv and siv >= su_s) else add_months(su_s, 4) - timedelta(days=1)
    later = [d for d in (ms.get("Inspection") or []) if d > fdbl]
    p7_s = min(later) if later else None
    p7_e = (max(max(later), end) if end else max(later)) if later else None
    cof_s = add_months(fdbl, -3)
    cof_e = (p7_s - timedelta(days=1)) if p7_s else (max(fdbl, end) if end else fdbl)
    segs = []
    if start and su_s > start:
        segs.append(["Before-Start-up", start, su_s - timedelta(days=1)])
    segs.append(["Start-up", su_s, su_e])
    if idbl and idbl < fdbl:
        coi_s = add_months(idbl, -3)
        segs.append(["Conduct (interim)", su_e + timedelta(days=1), coi_s - timedelta(days=1)])
        segs.append(["Close-out (interim)", coi_s, idbl])
        cof_s = max(cof_s, idbl + timedelta(days=1))
        segs.append(["Conduct (final)", idbl + timedelta(days=1), cof_s - timedelta(days=1)])
    else:
        segs.append(["Conduct (final)", su_e + timedelta(days=1), cof_s - timedelta(days=1)])
    segs.append(["Close-out (final)", cof_s, cof_e])
    if p7_s:
        segs.append(["After Close-out (final)", p7_s, p7_e])
    kept = [s for s in segs if s[2] >= s[1]]
    return [{"period_name": n, "period_seq": i + 1, "period_start": s, "period_end": e,
             "__derived": True} for i, (n, s, e) in enumerate(kept)]


class Model:
    def __init__(self, sheets):
        self.raw = sheets
        self.findings = []
        self.lists = defaultdict(list)
        for r in sheets["Lists"]:
            if r.get("list_name"):
                self.lists[r["list_name"]].append(r.get("value"))
        self.config = {r["parameter"]: r.get("value")
                       for r in sheets["Config"] if r.get("parameter")}
        def cfg(k, d):
            """A Config value, or the default if it is absent or unreadable.

            Written out rather than left as a one-liner because the one-liner ended
            in `or d`, which treats a legitimate ZERO as a missing value - so a
            setting deliberately turned off silently came back on. It cost nothing
            while every setting was a positive threshold, and became a real defect
            the moment one of them was a switch.
            """
            v = _as_num(self.config.get(k))
            return d if v is None else v
        self.OVER = cfg("over_allocation_fte", 1.50)
        self.UNDER = cfg("under_allocation_fte", 0.60)
        self.MINM = int(cfg("under_allocation_min_months", 3))
        self.HOURS = cfg("fte_hours_per_month", 160)
        # Whether the role factor is divided between the people sharing a role. A
        # setting rather than a constant because it changes every figure a shared role
        # ever produced, and somebody comparing this month's report with last year's
        # needs to be able to turn it off and see where the difference came from.
        self.SPLIT = cfg("split_shared_role_fte", 1) != 0
        self.ABSORB = cfg("absorb_unstaffed_role_factor", 1) != 0

        # V-30: a setting the file does not carry at all. Every one of these is read
        # through cfg(), which falls back to the figure below - so a missing row breaks
        # nothing, and that is the danger: the value in force becomes the program's
        # rather than the plan's, and the plan no longer records what it was.
        defaults = [("over_allocation_fte", 1.50, "the over-allocation threshold"),
                    ("under_allocation_fte", 0.60, "the under-allocation floor"),
                    ("under_allocation_min_months", 3,
                     "months below the floor before a run is reported"),
                    ("fte_hours_per_month", 160, "hours equal to 1.00 FTE"),
                    ("default_horizon_months", 24, "months shown when the dashboard opens"),
                    ("capacity_unit", "FTE", "the display unit"),
                    ("split_shared_role_fte", 1,
                     "whether a shared role's factor is divided (REQ-CAL-14)"),
                    ("absorb_unstaffed_role_factor", 1,
                     "whether an unstaffed role's factor is absorbed (REQ-CAL-16)")]
        absent = [d for d in defaults
                  if self.config.get(d[0]) is None or self.config.get(d[0]) == ""]
        if absent:
            self.add("information", "V-30", "Config", "",
                     "Config has no row for "
                     + ", ".join(k for k, _, _ in absent)
                     + ". The application is using its built-in default"
                     + ("" if len(absent) == 1 else "s") + ": "
                     + "; ".join(f"{k} = {v} ({what})" for k, v, what in absent)
                     + ". Nothing is wrong with the file - but the value in force is the "
                       "program's, not the plan's, and the plan no longer records what it "
                       "was. Add the row back to state it explicitly.")

        sv = _as_num(self.config.get("schema_version"))
        if sv is None:
            self.add("warning", "V-09", "Config", "", "No schema_version in Config.")
        elif int(sv) != B.SCHEMA_VERSION:
            self.add("warning", "V-09", "Config", "",
                     f"This file is schema version {int(sv)}; this build expects "
                     f"{B.SCHEMA_VERSION}.")

        self.projects, self.people = {}, {}
        for p in sheets["Project"]:
            # A row with no identifier is not a record: nothing can reference it, its
            # children have nothing to attach to, and it cannot survive a round trip.
            # Indexing it anyway is what let the application believe in a project called
            # "null" and then hide every real row behind it.
            if not _key(p, "project_id"):
                no_key(self, "Project", p, "project_id", "milestones and periods")
                continue
            if p.get("project_id") in self.projects:
                self.add("error", "V-08", "Project", p["__row"],
                         f"project_id {p['project_id']} appears more than once.")
            self.projects[p.get("project_id")] = p
        for p in sheets["Person"]:
            if not _key(p, "person_id"):
                no_key(self, "Person", p, "person_id", "assignments")
                continue
            if p.get("person_id") in self.people:
                self.add("error", "V-08", "Person", p["__row"],
                         f"person_id {p['person_id']} appears more than once.")
            self.people[p.get("person_id")] = p

        self.milestones = defaultdict(lambda: defaultdict(list))
        for m in sheets["Milestone"]:
            if m.get("project_id") and m.get("milestone_name") and m.get("milestone_date"):
                self.milestones[m["project_id"]][m["milestone_name"]].append(m["milestone_date"])
        for pid in self.milestones:
            for n in self.milestones[pid]:
                self.milestones[pid][n].sort()

        # Schema 6: both standards tables are keyed on the work scope as well, and a
        # row with an EMPTY scope applies to every scope.
        self.pws = {(r.get("project_type"), r.get("clinical_phase"), scope_of(r),
                     r.get("period_name")): _as_num(r.get("standard_fte"))
                    for r in sheets["PeriodWeightStandard"]}
        self.rf, self.rf_roles = {}, defaultdict(set)
        # Indexed by the ABSORBING role, because that is the question the calculation
        # asks: standing on the lead data manager, which absent roles land on me?
        self.rf_absorb = defaultdict(list)
        for r in sheets["RoleFactor"]:
            self.rf[(r.get("project_type"), r.get("clinical_phase"), scope_of(r),
                     r.get("period_name"), r.get("role_name"))] = _as_num(r.get("role_factor"))
            self.rf_roles[r.get("project_type")].add(r.get("role_name"))
            if r.get("absorbed_by"):
                self.rf_absorb[(r.get("project_type"), r.get("clinical_phase"),
                                scope_of(r), r.get("period_name"),
                                r["absorbed_by"])].append(r.get("role_name"))

        self.periods = defaultdict(list)
        for r in sheets["ProjectPeriod"]:
            self.periods[r.get("project_id")].append(r)
        for pid, proj in self.projects.items():
            if pid not in self.periods and proj.get("project_type") in CLINICAL_TYPES:
                got = derive_periods(proj, self.milestones.get(pid, {}))
                if got:
                    for d in got:
                        d["project_id"] = pid
                        w = std_weight(self, proj, d["period_name"])
                        d["weight"] = 1.00 if w is None else w
                    self.periods[pid] = got
                else:
                    self.add("error", "V-16", "Project", proj.get("__row", ""),
                             f"Project {pid} has no periods and cannot derive them - it is "
                             f"missing CTA submission or a DB lock.")
        for pid in self.periods:
            self.periods[pid].sort(key=lambda s: (s.get("period_seq") or 0))

        # Manual monthly figures (REQ-CAL-18). A workbook written before schema 9 has no
        # MonthlyEstimate sheet at all, which simply means it carries none.
        self.manual, self.manual_at = {}, {}
        for r in sheets.get("MonthlyEstimate", []) or []:
            if not r.get("scope") or not r.get("ref_id") or not r.get("month"):
                continue
            k = f"{r['scope']}|{r['ref_id']}|{r['month']}"
            self.manual[k] = _as_num(r.get("fte"))
            self.manual_at[k] = r.get("edited_at")

        self.ppw = defaultdict(list)
        for w in sheets["PersonPeriodWeight"]:
            self.ppw[w.get("assignment_id")].append(w)
        self.assignments = []
        validate(self)
        recompute_derived(self)
        # V-23 is a consequence of the arithmetic, so the model is not fully described
        # until the arithmetic has run. Cheap, and it keeps `prap_io validate` telling
        # an agent exactly what the application will show.
        calculate(self)

    def is_manual(self, scope, ref):
        if scope == "project":
            row = self.projects.get(ref)
        else:
            row = next((a for a in self.raw.get("Assignment", [])
                        if a.get("assignment_id") == ref), None)
        return str((row or {}).get("estimation_type") or "").strip().lower() == "manual"

    def add(self, sev, rule, sheet, row, msg):
        self.findings.append({"sev": sev, "rule": rule, "sheet": sheet,
                              "row": row, "msg": msg})


def recompute_derived(M):
    for m in M.raw["Milestone"]:
        master = (M.projects.get(m.get("project_id")) or {}).get("project_name")
        if m.get("project_name") and master and m["project_name"] != master:
            M.add("warning", "V-13", "Milestone", m["__row"],
                  f"Milestone row {m['__row']} records project_name "
                  f"'{m['project_name']}' but {m['project_id']} is '{master}'.")
        m["project_name"] = master
    for a in M.raw["Assignment"]:
        a["person_name"] = (M.people.get(a.get("person_id")) or {}).get("person_name")
    for p in M.raw["Project"]:
        if p.get("start_date") and p.get("end_date"):
            p["total_period_months"] = ((p["end_date"].year - p["start_date"].year) * 12
                                        + p["end_date"].month - p["start_date"].month + 1)


# --------------------------------------------- what an error is allowed to do
# Severity says how wrong something is; the CLASS says what an application may do about
# it. The register lives in the development plan and the same three answers are used by
# the browser (RULE_CLASS in core/05_model.js), so a program written against this file
# can tell a user which of their errors will stop them.
#
#   must         wrong with the row itself - refused
#   conditional  the row is sound, something it depends on is missing - the application
#                asks, and the user may keep it
#   incomplete   the row is still being built - reported, never questioned
RULE_CLASS = {"V-03": "conditional", "V-23": "conditional",
              "V-12": "incomplete", "V-16": "incomplete"}


def rule_class(rule):
    return RULE_CLASS.get(rule, "must")


def refuses(f):
    return f["sev"] in ("error", "fatal") and rule_class(f["rule"]) == "must"


def validate(M):
    for pid, p in M.projects.items():
        is_ct = p.get("project_type") in CLINICAL_TYPES
        if p.get("start_date") and p.get("end_date") and p["end_date"] < p["start_date"]:
            M.add("error", "V-05", "Project", p["__row"],
                  f"Project {pid}: end_date {p['end_date']} is before start_date "
                  f"{p['start_date']}.")
        if is_ct and not p.get("project_category"):
            M.add("warning", "V-04", "Project", p["__row"],
                  f"Project {pid} is a clinical trial with no product category.")
        if is_ct and not p.get("clinical_phase"):
            M.add("error", "V-19", "Project", p["__row"],
                  f"Project {pid} is a clinical trial with no clinical_phase.")
        if is_ct:
            for c in ("EDC_setup", "DataReviewSystem_setup", "RBQM_setup"):
                if p.get(c) is None:
                    M.add("warning", "V-10", "Project", p["__row"],
                          f"Project {pid} has no {c} recorded.")
        for col, lst in (("project_type", "project_type"),
                         ("work_scope_type", "work_scope_type"),
                         ("status", "project_status"), ("clinical_phase", "clinical_phase")):
            v = p.get(col)
            if v and M.lists.get(lst) and v not in M.lists[lst]:
                M.add("warning", "V-11", "Project", p["__row"],
                      f"Project {pid}: {col} '{v}' is not a known value.")
        # V-26: the value schema 6 retired. Only the user knows whether a biosimilar
        # trial ran in healthy volunteers or in patients, so it is reported rather than
        # guessed at - a wrong guess puts a wrong weight on real work.
        if p.get("project_type") in RETIRED_TYPES:
            M.add("error", "V-26", "Project", p["__row"],
                  f"Project {pid}: project_type '{p['project_type']}' was split in schema 6. "
                  f"Change it to {RETIRED_TYPES[p['project_type']]}.")

    for pid, mm in M.milestones.items():
        for nm, dates in mm.items():
            if M.lists.get("milestone_name") and nm not in M.lists["milestone_name"]:
                M.add("warning", "V-11", "Milestone", "",
                      f"Project {pid}: milestone '{nm}' is not in the standard list.")
            if nm != "Inspection" and len(dates) > 1:
                M.add("warning", "V-20", "Milestone", "",
                      f"Project {pid} records '{nm}' {len(dates)} times.")
        lock = (mm.get("final DB lock") or mm.get("interim DB lock") or [None])[0]
        if lock:
            early = [d for d in mm.get("Inspection", []) if d <= lock]
            if early:
                M.add("information", "V-21", "Milestone", "",
                      f"Project {pid}: {len(early)} Inspection date(s) on or before the "
                      f"final DB lock are treated as markers.")
        # V-14: a milestone outside its project's own window, and boundary milestones
        # out of order. Documented in the plan from v1.0 and reported from app v1.16.
        proj = M.projects.get(pid)
        if proj and proj.get("start_date") and proj.get("end_date"):
            for nm, dates in mm.items():
                for dt in dates:
                    if proj["start_date"] <= dt <= proj["end_date"]:
                        continue
                    # An Inspection after the final DB lock is the one milestone that is
                    # MEANT to sit past the project end: it opens 'After Close-out (final)'
                    # and the derivation extends the timeline to reach it (V-21). Saying
                    # 'outside the window' about that would be wrong, not merely noisy.
                    if nm == "Inspection" and lock and dt > lock:
                        M.add("information", "V-14", "Milestone", "",
                              f"Project {pid}: Inspection on {dt} falls after the project "
                              f"end {proj['end_date']}; the timeline is extended to cover "
                              f"it and 'After Close-out (final)' runs to that date.")
                        continue
                    M.add("warning", "V-14", "Milestone", "",
                          f"Project {pid}: '{nm}' on {dt} falls outside the project window "
                          f"{proj['start_date']}..{proj['end_date']}.")
        for n1, n2 in MILESTONE_ORDER:
            if not (mm.get(n1) and mm.get(n2)):
                continue
            d1, d2 = mm[n1][0], mm[n2][0]
            if d2 >= d1:
                continue
            sev = ("error" if {n1, n2} & DERIVATION_MILESTONES else "warning")
            M.add(sev, "V-14", "Milestone", "",
                  f"Project {pid}: '{n2}' ({d2}) is before '{n1}' ({d1}); the period "
                  f"derivation reads these in order.")

    for pid, proj in M.projects.items():
        segs = M.periods.get(pid)
        if not segs:
            if not any(f["rule"] == "V-16" and pid in f["msg"] for f in M.findings):
                M.add("error", "V-12", "ProjectPeriod", "", f"Project {pid} has no periods.")
            continue
        allowed = CLINICAL_PERIODS if proj.get("project_type") in CLINICAL_TYPES else OTHER_PERIODS
        names = [s.get("period_name") for s in segs]
        for n in set(names):
            if names.count(n) > 1:
                M.add("error", "V-18", "ProjectPeriod", "",
                      f"Project {pid}: period_name '{n}' appears {names.count(n)} times.")
        seqs = [s.get("period_seq") for s in segs]
        if len(set(seqs)) != len(seqs):
            M.add("error", "V-18", "ProjectPeriod", "",
                  f"Project {pid}: period_seq is duplicated.")
        prev_end = None
        for s in segs:
            if s.get("period_name") not in allowed:
                M.add("error", "V-15", "ProjectPeriod", "",
                      f"Project {pid} is type '{proj.get('project_type')}' but has a period "
                      f"named '{s.get('period_name')}'.")
            if s.get("period_end") and s.get("period_start") and s["period_end"] < s["period_start"]:
                M.add("error", "V-05", "ProjectPeriod", "",
                      f"Project {pid} period {s.get('period_seq')}: end before start.")
            if prev_end and s.get("period_start"):
                gap = (s["period_start"] - prev_end).days
                if gap > 1:
                    M.add("warning", "V-12", "ProjectPeriod", "",
                          f"Project {pid}: {gap - 1} day(s) before period "
                          f"{s.get('period_seq')} belong to no period.")
                elif gap < 1:
                    M.add("error", "V-06", "ProjectPeriod", "",
                          f"Project {pid}: periods overlap at {s.get('period_name')}.")
            prev_end = s.get("period_end")
            if (proj.get("project_type") in CLINICAL_TYPES and proj.get("clinical_phase")
                    and std_weight(M, proj, s.get("period_name")) is None):
                M.add("error", "V-19", "PeriodWeightStandard", "",
                      f"Project {pid}: no standard weight for {proj['project_type']} / "
                      f"{proj['clinical_phase']} / "
                      f"{proj.get('work_scope_type') or 'any scope'} / "
                      f"{s.get('period_name')}.")

    seen = set()
    for a in M.raw["Assignment"]:
        aid = a.get("assignment_id")
        if not _key(a, "assignment_id"):
            no_key(M, "Assignment", a, "assignment_id", "weight overrides")
            continue
        if aid in seen:
            M.add("error", "V-08", "Assignment", a["__row"],
                  f"assignment_id {aid} appears more than once.")
        seen.add(aid)
        proj = M.projects.get(a.get("project_id"))
        if not proj:
            M.add("error", "V-01", "Assignment", a["__row"],
                  f"Assignment {aid} refers to project {a.get('project_id')}, which does "
                  f"not exist.")
            continue
        if a.get("person_id") not in M.people:
            M.add("error", "V-02", "Assignment", a["__row"],
                  f"Assignment {aid} refers to person {a.get('person_id')}, which does "
                  f"not exist.")
        # The coarse half of the question V-23 asks precisely: does RoleFactor carry this
        # role for this project's TYPE at all. Reported, never a reason to refuse a row
        # in the application (R-19).
        roles = M.rf_roles.get(proj.get("project_type"))
        if not roles or a.get("role_name") not in roles:
            M.add("error", "V-03", "Assignment", a["__row"],
                  f"Assignment {aid}: role '{a.get('role_name')}' has no RoleFactor row "
                  f"for a project of type '{proj.get('project_type')}', so it would be "
                  f"calculated at factor 1.00.")
        if (a.get("assign_end_date") and a.get("assign_start_date")
                and a["assign_end_date"] < a["assign_start_date"]):
            M.add("error", "V-05", "Assignment", a["__row"], f"Assignment {aid}: end before start.")
        if proj.get("end_date") and a.get("assign_end_date") and a["assign_end_date"] > proj["end_date"]:
            M.add("warning", "V-07", "Assignment", a["__row"],
                  f"Assignment {aid} runs to {a['assign_end_date']}, after project "
                  f"{a.get('project_id')} ends on {proj['end_date']}.")
        per = M.people.get(a.get("person_id"))
        if per and a.get("person_name") and a["person_name"] != per.get("person_name"):
            M.add("warning", "V-13", "Assignment", a["__row"],
                  f"Assignment {aid} records person_name '{a['person_name']}' but "
                  f"{a.get('person_id')} is '{per.get('person_name')}'.")
        M.assignments.append(a)

    for aid, wins in M.ppw.items():
        if aid not in seen:
            M.add("error", "V-24", "PersonPeriodWeight", "",
                  f"{aid}: PersonPeriodWeight refers to an assignment that does not exist.")
            continue
        wins.sort(key=lambda w: (w.get("period_start") or date.min))
        starts = [w.get("period_start") for w in wins]
        if len(set(starts)) != len(starts):
            M.add("error", "V-24", "PersonPeriodWeight", "",
                  f"{aid}: two override windows share a period_start.")
        for a1, a2 in zip(wins, wins[1:]):
            if (a1.get("period_end") and a2.get("period_start")
                    and a2["period_start"] <= a1["period_end"]):
                M.add("error", "V-06", "PersonPeriodWeight", "",
                      f"{aid}: override windows overlap - {a1['period_start']}.."
                      f"{a1['period_end']} and {a2['period_start']}..{a2['period_end']}.")
        for w in wins:
            if w.get("period_end") and w.get("period_start") and w["period_end"] < w["period_start"]:
                M.add("error", "V-05", "PersonPeriodWeight", "",
                      f"{aid}: override window ends before it starts.")

    for sid, p in M.people.items():
        cap = _as_num(p.get("capacity_fte"))
        if cap is not None and cap < M.UNDER:
            M.add("warning", "V-22", "Person", p["__row"],
                  f"{sid}: capacity {cap:.2f} FTE is below the under-allocation floor of "
                  f"{M.UNDER:.2f}.")

    # V-23 is NOT raised here. It is raised by calculate() - see report_gaps - keyed on
    # the composition the lookup actually used and only for the person-months that
    # actually made it, so it reports what the arithmetic had to guess at rather than
    # what the sheets might one day need. It is also therefore a finding that cannot
    # refuse an edit in the application, which is the point (R-19).


# =============================================================== 4. calculation
def project_window(M, proj):
    """How long the project runs, for the purpose of working out a number.

    THE PERIODS ARE THE PROJECT (REQ-CAL-17). Milestones are reference dates: the
    derivation reads them to lay the periods out, and several of them are markers
    that sit INSIDE the run rather than bounding it. The periods are the run itself,
    and the only thing any weight in this calculation is attached to.

    A project with no periods keeps its own typed dates - there is nothing to take a
    window from, and V-12 already says so.
    """
    lo = hi = None
    for s in (M.periods.get(proj.get("project_id")) or []):
        ps, pe = s.get("period_start"), s.get("period_end")
        if ps is not None and (lo is None or ps < lo):
            lo = ps
        if pe is not None and (hi is None or pe > hi):
            hi = pe
    return (lo or proj.get("start_date"), hi or proj.get("end_date"))


def assignment_window(M, proj, a):
    """The months an assignment covers.

    Both dates are optional and a blank one means the project's own (REQ-CAL-15):
    most people are on a project for the whole of it, and only a partial involvement
    is worth writing down. One function for the sharing pre-pass and the calculation
    alike, so the months a person is counted IN cannot differ from the months they
    are counted AMONG.
    """
    ps, pe = project_window(M, proj)
    return (a.get("assign_start_date") or ps,
            a.get("assign_end_date") or pe)


def calculate(M):
    proj_month = defaultdict(float)
    pers_month = defaultdict(float)
    cell = defaultdict(float)
    pers_proj = defaultdict(lambda: defaultdict(float))
    lo, hi = None, None

    def period_at(pid, y, m):
        first = date(y, m, 1)
        for s in M.periods.get(pid, []):
            if s.get("period_start") and s.get("period_end") \
                    and s["period_start"] <= first <= s["period_end"]:
                return s
        return None

    def person_weight(a, y, m):
        first = date(y, m, 1)
        for w in M.ppw.get(a.get("assignment_id"), []):
            if w.get("period_start") and w.get("period_end") \
                    and w["period_start"] <= first <= w["period_end"]:
                return _as_num(w.get("weight_override")) or 0.0
        return _as_num(a.get("person_weight")) or 0.0

    # Who shares a role, and when. The role factor is what the ROLE costs the project
    # in a period, not what each person holding it costs - so two data managers on one
    # trial share it rather than being charged for two. Counted PER MONTH, because that
    # is the only count that conserves the total when one of them leaves mid-project,
    # and by distinct PEOPLE, so two rows for one person do not halve their own load.
    # Built ALWAYS and read two ways: how many people hold a role in a month is the
    # divisor, and WHETHER anybody holds it decides absorption (REQ-CAL-16).
    sharers = defaultdict(set)
    for a in M.assignments:
        proj = M.projects.get(a.get("project_id"))
        if not proj or a.get("person_id") not in M.people:
            continue
        s, e = assignment_window(M, proj, a)
        if not s or not e:
            continue
        for y, m in months_between(s, e):
            if coverage(y, m, s, e) > 0:
                sharers[(a["project_id"], a.get("role_name"),
                         month_key(y, m - 1))].add(a["person_id"])

    # Which roles are staffed on a project in a month, so the demand can be divided
    # between them (REQ-CAL-19). Built from the same pass as the divisor and the
    # absorption test: three answers about one month, one picture of that month.
    roles_on = defaultdict(set)
    for (pid_, role_, k_) in sharers:
        roles_on[(pid_, k_)].add(role_)

    def effective_factor(proj, period_name, role_name, k):
        """This role's factor plus the factor of any role that names it as cover and
        that nobody is holding this month. One hop: if the absorbing role is itself
        unstaffed the work is not passed further along - V-29 reports that instead."""
        rf = std_factor(M, proj, period_name, role_name)
        rf = 1.0 if rf is None else rf
        if not M.ABSORB:
            return rf
        for absent in absorbed_into(M, proj, period_name, role_name):
            if sharers.get((proj["project_id"], absent, k)):
                continue
            rf += std_factor(M, proj, period_name, absent) or 0.0
        return rf

    # V-23, recorded as it happens: the composition the lookup used, and only where the
    # lookup fed a person-month. A month in no period at all is V-12's finding, not this
    # one - there is no period name for a factor to be missing FOR.
    _std_cache = {}

    def std_monthly(proj, period_name):
        """The month's demand in FTE, before this project's own adjustment.

        PeriodWeightStandard holds it - see stdMonthly() in core/06_calculate.js for why
        it went unused until schema 10. Missing, it falls back to 1.00 and V-19 reports
        it, which is deliberately the OLD behaviour: the project month becomes its own
        period weight, so an incomplete standards sheet degrades to figures its author
        will recognise rather than to zero.
        """
        if period_name is None:
            return 1.0
        key = (proj.get("project_id"), period_name)
        if key not in _std_cache:
            v = std_weight(M, proj, period_name)
            _std_cache[key] = 1.0 if v is None else float(v)
        return _std_cache[key]

    gaps = {}
    lines = []

    for a in M.assignments:
        proj = M.projects.get(a.get("project_id"))
        if not proj or a.get("person_id") not in M.people:
            continue
        s, e = assignment_window(M, proj, a)
        if not s or not e:
            continue
        for y, m in months_between(s, e):
            cov = coverage(y, m, s, e)
            if cov <= 0:
                continue
            seg = period_at(a["project_id"], y, m)
            pw = (_as_num(seg.get("weight")) if seg else None)
            pw = 1.0 if pw is None else pw
            k = month_key(y, m - 1)
            if seg and std_factor(M, proj, seg.get("period_name"), a.get("role_name")) is None:
                key = (proj.get("project_type"), proj.get("clinical_phase") or "",
                       scope_of(proj) or "", seg.get("period_name"), a.get("role_name"))
                g = gaps.setdefault(key, {"proj": proj, "period_name": seg.get("period_name"),
                                          "role": a.get("role_name"), "projects": set(),
                                          "months": 0})
                g["projects"].add(a["project_id"])
                g["months"] += 1
            pn = seg.get("period_name") if seg else None
            rf = effective_factor(proj, pn, a.get("role_name"), k)
            share = (len(sharers.get((a["project_id"], a.get("role_name"), k), ())) or 1) \
                if M.SPLIT else 1
            # REQ-CAL-19. The STANDARD is the month's demand in FTE; the project's own
            # period weight adjusts it for this study; the role factors then divide that
            # demand between the roles ACTUALLY STAFFED, so the shares add to one.
            denom = sum(effective_factor(proj, pn, r_, k)
                        for r_ in roles_on.get((a["project_id"], k), ()))
            std_f = std_monthly(proj, pn)
            frac = (rf / share) / denom if denom > 0 else 0.0
            v = std_f * pw * frac * person_weight(a, y, m) * cov
            lo = k if lo is None else min(lo, k)
            hi = k if hi is None else max(hi, k)
            lines.append({"month": k, "project_id": a["project_id"],
                          "person_id": a["person_id"],
                          "assignment_id": a.get("assignment_id"),
                          "role_name": a.get("role_name"), "fte": v})

    apply_manual(M, lines)
    # Every map built from the lines, so a manual figure moves the totals and the two
    # can never disagree - the same reason the browser engine does it this way.
    for L in lines:
        proj_month[(L["project_id"], L["month"])] += L["fte"]
        pers_month[(L["person_id"], L["month"])] += L["fte"]
        pers_proj[(L["person_id"], L["month"])][L["project_id"]] += L["fte"]
        cell[(L["project_id"], L["person_id"], L["role_name"], L["month"])] += L["fte"]

    report_gaps(M, gaps)
    return {"proj_month": proj_month, "pers_month": pers_month, "pers_proj": pers_proj,
            "cell": cell, "gaps": gaps, "lines": lines, "lo": lo or 0, "hi": hi or 0}


def apply_manual(M, lines):
    """REQ-CAL-18, exactly as core/06_calculate.js does it.

    Assignment first - the figure REPLACES that person's contribution - then project,
    which sets the whole month and SCALES the people on it so they still add up to it.
    A project figure with nobody assigned is not applied; the browser reports V-32.
    """
    if not getattr(M, "manual", None):
        return

    def iso(k):
        return f"{k // 12}-{k % 12 + 1:02d}"

    for L in lines:
        if not M.is_manual("assignment", L["assignment_id"]):
            continue
        v = M.manual.get(f"assignment|{L['assignment_id']}|{iso(L['month'])}")
        L["fte"] = 0.0 if v is None else v

    groups = defaultdict(list)
    for L in lines:
        if M.is_manual("project", L["project_id"]):
            groups[(L["project_id"], L["month"])].append(L)
    for (pid, k), group in groups.items():
        want = M.manual.get(f"project|{pid}|{iso(k)}")
        if want is None:
            for L in group:
                L["fte"] = 0.0
            continue
        have = sum(L["fte"] for L in group)
        if abs(have) < 1e-9:
            continue                       # nobody to share it out to - V-32
        scale = want / have
        for L in group:
            L["fte"] *= scale


def report_gaps(M, gaps):
    """V-23, written onto the model once the calculation knows what it needed.

    Rewritten rather than appended, so calculating the same model twice does not
    report it twice. Still an ERROR: a role calculated at 1.00 is a different answer,
    not an approximate one. What changed is only where it comes from, and therefore
    what it can do - a finding that exists only after the arithmetic cannot refuse the
    edit that led to it.
    """
    M.findings[:] = [f for f in M.findings if f["rule"] != "V-23"]
    for g in sorted(gaps.values(), key=lambda g: g["months"]):
        proj = g["proj"]
        ph = proj.get("clinical_phase") if proj.get("project_type") in CLINICAL_TYPES else None
        pl = sorted(g["projects"])
        M.add("error", "V-23", "RoleFactor", "",
              f"No role factor for {proj.get('project_type')} / {ph or '-'} / "
              f"{proj.get('work_scope_type') or 'any scope'} / {g['period_name']} / "
              f"{g['role']} - {g['months']} person-month(s) on {len(pl)} project(s) "
              f"({', '.join(pl[:3])}{', ...' if len(pl) > 3 else ''}) were calculated at "
              f"factor 1.00 instead.")


def flags(M, C):
    """Over-allocated months and under-allocated runs, exactly as the dashboard marks them."""
    out = []
    people = sorted({sid for sid, _ in C["pers_month"]})
    for sid in people:
        run = []
        for k in range(C["lo"], C["hi"] + 1):
            v = C["pers_month"].get((sid, k), 0.0)
            if v > M.OVER:
                out.append({"kind": "over", "person_id": sid, "month": key_label(k),
                            "fte": round(v, 4), "threshold": M.OVER})
            if v < M.UNDER:
                run.append(k)
            else:
                if len(run) >= M.MINM:
                    out.append({"kind": "under_run", "person_id": sid,
                                "from": key_label(run[0]), "to": key_label(run[-1]),
                                "months": len(run), "threshold": M.UNDER})
                run = []
        if len(run) >= M.MINM:
            out.append({"kind": "under_run", "person_id": sid, "from": key_label(run[0]),
                        "to": key_label(run[-1]), "months": len(run), "threshold": M.UNDER})
    return out


# =============================================================== 5. the CLI
def _load(path):
    sheets = read_any(path)
    return sheets, Model(sheets)


def cmd_to_json(args):
    sheets = read_any(args.file)
    out = args.out or str(Path(args.file).with_suffix("")) + ".prap.json"
    write_json(sheets, out)
    print(f"wrote {out}  ({sum(len(v) for v in sheets.values())} rows across "
          f"{len(sheets)} sheets)")


def cmd_to_xlsx(args):
    sheets = read_any(args.file)
    out = args.out or str(Path(args.file).name).replace(".prap.json", "") + ".xlsx"
    write_xlsx(sheets, out)
    print(f"wrote {out}  ({sum(len(v) for v in sheets.values())} rows across "
          f"{len(sheets)} sheets)")


SEV_ORDER = {"fatal": 0, "error": 1, "warning": 2, "information": 3}


def cmd_validate(args):
    _, M = _load(args.file)
    F = sorted(M.findings, key=lambda f: (SEV_ORDER.get(f["sev"], 9), f["rule"]))
    if args.json:
        print(json.dumps({"file": str(args.file), "findings": F,
                          "counts": {s: sum(1 for f in F if f["sev"] == s)
                                     for s in ("fatal", "error", "warning", "information")}},
                         indent=1, ensure_ascii=False))
    else:
        counts = defaultdict(int)
        for f in F:
            counts[f["sev"]] += 1
        for f in F:
            row = f"row {f['row']}" if f["row"] else ""
            # The class beside the severity, because "will this stop me" is the question
            # a reader brings to this list. Only errors carry one - a warning never
            # refused anything, so labelling it would suggest a choice that never existed.
            cls = f"[{rule_class(f['rule'])}]" if f["sev"] in ("error", "fatal") else ""
            print(f"  {f['sev']:<11} {cls:<14}{f['rule']}  {f['sheet']:<20} "
                  f"{row:<9} {f['msg']}")
        print()
        by_sev = ", ".join(f"{counts[s]} {s}"
                           for s in ("fatal", "error", "warning", "information") if counts[s])
        print(f"{len(F)} finding(s): {by_sev}" if F else "0 findings — the file is clean.")
        must = sum(1 for f in F if refuses(f))
        if must:
            print(f"{must} of them are [must]: the application refuses those rows until "
                  f"they are corrected. The rest it reports, and asks about at Save.")
    # A non-zero exit means "this file has something wrong with it", which is what a
    # script wants to branch on - so it still counts every error, not only the must ones.
    return 1 if any(f["sev"] in ("fatal", "error") for f in F) else 0


def cmd_calculate(args):
    _, M = _load(args.file)
    C = calculate(M)
    lo = parse_ym(args.frm) if args.frm else C["lo"]
    hi = parse_ym(args.to) if args.to else C["hi"]
    months = list(range(lo, hi + 1))
    hours = args.unit == "hours"
    conv = (lambda v: v * M.HOURS) if hours else (lambda v: v)

    if args.by == "cell":
        rowsrc = {}
        for (pid, sid, role, k), v in C["cell"].items():
            rowsrc.setdefault((pid, sid, role), {})[k] = v
        label = ("project_id", "person_id", "role_name")
    else:
        src = C["proj_month"] if args.by == "project" else C["pers_month"]
        rowsrc = {}
        for (i, k), v in src.items():
            rowsrc.setdefault((i,), {})[k] = v
        label = ("project_id",) if args.by == "project" else ("person_id",)

    table = []
    for keys in sorted(rowsrc):
        rec = dict(zip(label, keys))
        rec["months"] = {key_label(k): round(conv(rowsrc[keys].get(k, 0.0)), 4)
                         for k in months if rowsrc[keys].get(k)}
        rec["total"] = round(sum(conv(rowsrc[keys].get(k, 0.0)) for k in months), 4)
        if rec["months"]:
            table.append(rec)

    payload = {"file": str(args.file), "by": args.by, "unit": "hours" if hours else "FTE",
               "from": key_label(lo), "to": key_label(hi), "rows": table}
    if args.flags:
        payload["flags"] = flags(M, C)
        payload["thresholds"] = {"over_allocation_fte": M.OVER,
                                 "under_allocation_fte": M.UNDER,
                                 "under_allocation_min_months": M.MINM}
    if args.json:
        print(json.dumps(payload, indent=1, ensure_ascii=False))
        return 0

    head = "  ".join(label)
    print(f"{args.by} load, {payload['unit']}, {payload['from']} .. {payload['to']}")
    print()
    for rec in table:
        who = "  ".join(str(rec[c]) for c in label)
        print(f"  {who:<52} total {rec['total']:>9.2f}")
        for mk, v in rec["months"].items():
            print(f"      {mk:<10} {v:>8.2f}")
    if args.flags:
        print()
        for f in payload["flags"]:
            if f["kind"] == "over":
                print(f"  OVER   {f['person_id']}  {f['month']}  {f['fte']:.2f} FTE "
                      f"(ceiling {f['threshold']:.2f})")
            else:
                print(f"  UNDER  {f['person_id']}  {f['from']} .. {f['to']} "
                      f"({f['months']} months below {f['threshold']:.2f})")
    return 0


def main(argv=None):
    ap = argparse.ArgumentParser(
        prog="prap_io", description=__doc__.split("\n")[0],
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Column meanings, value lists and every validation rule: "
               "docs/prap_contract.json and docs/PRAP_AI_Agent_Guide.md")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("to-json", help="workbook -> PRAP JSON interchange file")
    p.add_argument("file")
    p.add_argument("-o", "--out")
    p.set_defaults(fn=cmd_to_json)

    p = sub.add_parser("to-xlsx", help="PRAP JSON interchange file -> workbook")
    p.add_argument("file")
    p.add_argument("-o", "--out")
    p.set_defaults(fn=cmd_to_xlsx)

    p = sub.add_parser("validate", help="report every validation finding, as the app does")
    p.add_argument("file")
    p.add_argument("--json", action="store_true")
    p.set_defaults(fn=cmd_validate)

    p = sub.add_parser("calculate", help="monthly load, by person, project or cell")
    p.add_argument("file")
    p.add_argument("--by", choices=("person", "project", "cell"), default="person")
    p.add_argument("--from", dest="frm", metavar="YYYY-MM")
    p.add_argument("--to", metavar="YYYY-MM")
    p.add_argument("--unit", choices=("FTE", "hours"), default="FTE")
    p.add_argument("--flags", action="store_true",
                   help="also list over-allocated months and under-allocated runs")
    p.add_argument("--json", action="store_true")
    p.set_defaults(fn=cmd_calculate)

    args = ap.parse_args(argv)
    try:
        return args.fn(args) or 0
    except Problem as e:
        print(f"error: {e}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    sys.exit(main())
