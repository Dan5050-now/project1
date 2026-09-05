"""Generate the PM_APP UI component list - the Step N3 deliverable for review.

Two documents come out of Step N3: the prototype you can click
(tools/build_newapp_prototype.py) and this, which is what you mark up.

Every component carries a verdict column. The question is not "is this pretty" but
"is this the behaviour you want when it happens to you at 16:40 on a Friday".

Sheet 03 is the one the plan specifically requires: NR-PAR-04 says the desktop
interface matches the web application's except where a desktop convention requires
otherwise, AND that every such difference is listed. A divergence nobody wrote down is
a divergence nobody agreed to.

    python tools/build_newapp_components.py

Output: docs/PRAP_NewApp_Component_List_v0.1.xlsx
"""

import importlib.util
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DOC_VERSION = "1.0"
DOC_STATUS = "APPROVED - 2026-08-13. Gate N3 closed; this governs Step N4."
DOC_DATE = "2026-08-13"
SPEC = "PRAP_NewApp_Specification_v1.2.xlsx"
PLAN = "PRAP_NewApp_Development_Plan_v1.4.xlsx"
PROTO = "app/PM_APP_Prototype_v0.3.html"
ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / f"PRAP_NewApp_Component_List_v{DOC_VERSION}.xlsx"

FONT = "Arial"
NAVY = "1F3864"
TITLE_F = Font(name=FONT, size=16, bold=True, color=NAVY)
H1_F = Font(name=FONT, size=12, bold=True, color=NAVY)
HDR_F = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_F = Font(name=FONT, size=10)
BOLD_F = Font(name=FONT, size=10, bold=True)
NOTE_F = Font(name=FONT, size=9, italic=True, color="808080")
MONO_F = Font(name="Consolas", size=9)

HDR_FILL = PatternFill("solid", fgColor="2F5597")
BAND_FILL = PatternFill("solid", fgColor="F2F5FB")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
NEW_FILL = PatternFill("solid", fgColor="C6E0B4")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
WRAP_C = Alignment(vertical="top", wrap_text=True, horizontal="center")


def sheet(wb, name, title, subtitle=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_F
    row = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = NOTE_F
        row = 3
    ws.freeze_panes = f"A{row + 2}"
    return ws, row + 1


def table(ws, start_row, headers, rows, widths, wrap_cols=()):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = HDR_F
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = WRAP_C
    ws.row_dimensions[start_row].height = 28
    for r, data in enumerate(rows, start=start_row + 1):
        for i, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = BODY_F
            c.border = BOX
            c.alignment = WRAP if i in wrap_cols else Alignment(vertical="top")
            if (r - start_row) % 2 == 0:
                c.fill = BAND_FILL
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return start_row + len(rows) + 2


def section(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = H1_F
    return row + 1


def note(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = NOTE_F
    return row + 1


def lines(ws, row, texts, mono=False):
    for t in texts:
        c = ws.cell(row=row, column=1, value=t)
        c.font = MONO_F if mono else BODY_F
        row += 1
    return row


def verdict_column(ws, first_row, n, col):
    dv = DataValidation(type="list", formula1='"OK,Change,Discuss"', allow_blank=True)
    ws.add_data_validation(dv)
    letter = get_column_letter(col)
    dv.add(f"{letter}{first_row}:{letter}{first_row + n - 1}")
    for rr in range(first_row, first_row + n):
        ws.cell(row=rr, column=col).fill = INPUT_FILL


# The prototype is the source of the screen list, so the two cannot describe
# different sets of screens.
spec_ = importlib.util.spec_from_file_location(
    "proto", ROOT / "tools" / "build_newapp_prototype.py")
proto = importlib.util.module_from_spec(spec_)
spec_.loader.exec_module(proto)
SCREENS = [(sid, label, comps) for sid, label, comps, _ in proto.SCREENS]

wb = Workbook()
wb.remove(wb.active)

# ---- 00 Cover -------------------------------------------------------------
ws = wb.create_sheet("00_Cover")
ws.sheet_view.showGridLines = False
ws["A1"] = "Project Management APP (PM_APP)"
ws["A1"].font = Font(name=FONT, size=20, bold=True, color=NAVY)
ws["A2"] = "UI Component List"
ws["A2"].font = Font(name=FONT, size=14, color=NAVY)

cover = [
    ("Document ID", "PRAP-NAPP-UI-001"),
    ("Document type", "UI component list (Step N3 deliverable)"),
    ("Version", f"v{DOC_VERSION}"),
    ("Status", DOC_STATUS),
    ("Issue date", DOC_DATE),
    ("Author", "Claude Code"),
    ("Reviewer", "Requester - one round; five points settled, one added, approved 2026-08-13"),
    ("Governed by", f"{SPEC} (Gate N2 closed) and {PLAN} (Gate N1 closed)"),
    ("Review it against", f"{PROTO} - open it in a browser. It is THE REAL APPLICATION, "
                          f"pre-loaded with the 62-project dummy dataset: every tab, table, "
                          f"chart and figure is computed by the engine that will ship, and "
                          f"every person-month in it has been checked against the Python "
                          f"reference implementation - 1,225 of them, exact. The desktop "
                          f"chrome is around it, and the eight new screens open over it from "
                          f"the button top right. Only those eight are mocked."),
    ("Also", "app/PM_APP_Prototype_v0.1.html - the same eight screens on their own, if you "
             "want to read the wording without the application behind it"),
    ("What is being designed", "Only what the browser has no counterpart for. Every tab, "
                               "table, chart, filter and editing behaviour is the web "
                               "application's, unchanged - decision N-11"),
    ("Repository", "Dan5050-now/project1"),
]
r = 4
for k, v in cover:
    ws.cell(row=r, column=1, value=k).font = BOLD_F
    c = ws.cell(row=r, column=2, value=v)
    c.font = BODY_F
    c.alignment = WRAP
    r += 1
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 112
for rr in range(4, r):
    ws.row_dimensions[rr].height = 30

r += 1
r = section(ws, r, "How to review this")
r = lines(ws, r, [
    "Sheet 02 is the component list: 20 components, each with a verdict column - OK, Change or Discuss.",
    "Sheet 03 is the divergence register, which the plan specifically requires (NR-PAR-04): every place the",
    "desktop interface differs from the web one, with the reason. A divergence nobody wrote down is a",
    "divergence nobody agreed to.",
    "Sheet 04 gives every string a user can read, in one place, because wording is most of what these screens",
    "ARE - and a message that says 'error' where it could have named a colleague is a defect no screenshot",
    "shows.",
    "",
    "The useful question is not whether a dialog is pretty. It is whether this is the behaviour you want when",
    "it happens to you at 16:40 on a Friday.",
])

# ---- 01 Screens -----------------------------------------------------------
ws, r = sheet(wb, "01_Screens", "The eight screens",
              f"All eight are in {PROTO}. Nothing else about the interface changes.")
rows = [[str(i), label.split(" · ")[1], comps, f"#{sid}"]
        for i, (sid, label, comps) in enumerate(SCREENS, start=1)]
r = table(ws, r, ["#", "Screen", "Components", "In the prototype"],
          rows, [5, 40, 26, 22], wrap_cols=(2,))
r = note(ws, r, "Eight screens is the whole of the new interface. Everything else the desktop application "
                "shows - four tabs, every table, every chart, the provisional-edit model, the filters, the "
                "change log - is the web application, running in a window instead of a browser. The "
                "prototype demonstrates that literally: what is behind the eight screens is not a mock-up "
                "of the application, it IS the application, with 62 projects and 20 people in it.")
r += 1
r = section(ws, r, "What the prototype proves, and what it does not")
r = lines(ws, r, [
    "PROVES   Every figure, table and chart you see is computed. tools/test_demo.py compares all 1,225",
    "         person-months against the Python reference implementation on every run - worst difference",
    "         0.00e+00. If a number looks wrong in the prototype, it is wrong in the product.",
    "",
    "DOES NOT PROVE   Anything about the desktop shell. No file is read, no plan is opened, no claim is",
    "         taken, nothing is saved. The eight screens are pictures with the specified words on them.",
    "         Those become real at Step N4, and the tests that prove them are listed on sheet 08 of the",
    "         specification.",
])
r += 1
r = section(ws, r, "Why so few")
r = lines(ws, r, [
    "Because the difference between the two applications is not what they show. It is where the bytes live,",
    "and who is allowed to change them. Almost all of that is invisible: a save that is atomic looks exactly",
    "like a save that is not, until the power goes off.",
    "",
    "These eight are the places where it becomes visible - and they are, without exception, the places where a",
    "user finds out that something is not as they assumed. Somebody else has the plan. The figures moved. The",
    "application closed with your work in it. That is why the wording on sheet 04 matters more than the layout.",
])

# ---- 02 Components --------------------------------------------------------
ws, r = sheet(wb, "02_Components", "Component list",
              "Mark each OK, Change or Discuss. Add a note in the last column.")

comps = [
    ["NC-01", "Identity dialog", "1", "Asks name and department at launch, before any plan opens. First run pre-fills the name from the Windows account; afterwards both are pre-filled and the primary button reads 'Continue as <name>'.", "NR-USR-01..04", ""],
    ["NC-02", "'Not me' switch", "1", "Clears the remembered identity and asks again. Present on every launch, not buried in a menu - a shared PC needs it in the place where the wrong name is showing.", "NR-USR-02", ""],
    ["NC-03", "Open-a-plan screen", "2", "Recent list on the left, four ways to start on the right. Each recent entry shows where it lives, when it was saved and by whom.", "NR-APP-03, NR-APP-06, NR-USR-07", ""],
    ["NC-04", "'Held by' marker in the recent list", "2", "A plan somebody else is editing says so BEFORE it is opened, not after. Costs one read of the claim file per listed plan - agreed as worth it.", "NR-STO-12", "U-N01"],
    ["NC-05", "Blocked dialog - holder active", "3", "Names the holder and department, says when they started and that they are active now. Buttons: copy name and department, open read-only.", "NR-STO-11, NR-STO-12, NR-USR-05", ""],
    ["NC-06", "Blocked dialog - holder silent", "3", "The same dialog in amber, saying since when they stopped responding and the exact time the plan becomes free.", "NR-STO-14, NR-STO-12", ""],
    ["NC-07", "Superseded strip", "4", "A strip above the content, not a dialog: says who saved and when, and what the figures on screen are from. Offers Reload or Keep looking.", "NR-STO-16", ""],
    ["NC-08", "Recovery dialog", "5", "Names the plan, the number of unsaved changes and when they were made, and lists EVERY ONE of them in a scrollable box. Keep or discard.", "NR-STO-07", "U-N02"],
    ["NC-09", "Update-first question", "6", "'This plan already contains data. Update it from <file>?' asked BEFORE any difference is computed or shown.", "NR-IMP-02", ""],
    ["NC-10", "Difference report", "6", "Per sheet: add, change, only-here, and a tick. 'Only here' is highlighted because those are the hand-entered rows. Expands on request to the individual rows, showing current and incoming values side by side - for looking, not for choosing; the tick stays at sheet level.", "NR-IMP-02", "U-N05"],
    ["NC-11", "Window title", "7", "<plan> — Project Management APP, with a bullet for unsaved changes and [read-only] when somebody else holds it.", "NR-APP-07, NR-APP-08", ""],
    ["NC-12", "Menu bar", "7", "File, Edit, View, Plan, Help. Save is disabled without the claim, with the reason on hover.", "NR-APP-05", ""],
    ["NC-13", "Status strip", "7", "Who you are, whether you hold the plan, the unsaved count, and the three buttons the web application already has. 'You are editing this plan' appears when a VALUE CHANGES, not when you click into a cell.", "NR-USR-01, NR-STO-10", "U-N03"],
    ["NC-14", "About dialog", "8", "Version, schema, application folder, DATA FOLDER, and which of the four rules chose it.", "NR-APP-07, NR-DEP-10", ""],
    ["NC-15", "Save / Save As / Close plan", "7", "Native dialogs, not browser downloads. Save As works even when blocked, so being unable to edit never means being unable to work.", "NR-APP-05, NR-STO-17", ""],
    ["NC-16", "Recent-workspaces menu", "7", "At least ten, most recent first, per user, with relative paths where possible.", "NR-APP-06, NR-DEP-08", ""],
    ["NC-17", "Restore previous version", "7", "Plan menu. Loads the retained version AS A PENDING EDIT rather than overwriting - so a mistaken restore costs nothing.", "NR-STO-06", ""],
    ["NC-18", "'Look at a source file'", "2", "Opens a workbook read-only with no workspace created. Save disabled throughout; Export and Save As are not.", "NR-IMP-05", ""],
    ["NC-19", "Protected-file message", "-", "A file that cannot be read because it is protected says so and says what to do, rather than reporting corruption.", "NR-IMP-06", ""],
    ["NC-21", "Export for editing outside the application", "7", "File > Export to Excel and Export JSON, offered as a way to CHANGE a plan as well as to read one: export, edit the file elsewhere, hand it to the manager, import it back through the difference report. The round trip loses nothing, and derived values are recomputed rather than trusted.", "NR-IMP-08", "New at round 1"],
    ["NC-20", "Read-only mode", "3", "Not a refusal: every tab, figure, chart, filter and export works. Only writing to that plan is withheld, and the window says why.", "NR-STO-11", ""],
]
r_start = r
r = table(ws, r, ["ID", "Component", "Screen", "Behaviour", "Requirements", "Your verdict"],
          [c + [""] for c in comps], [8, 32, 8, 72, 26, 16, 30],
          wrap_cols=(4, 5, 7))
verdict_column(ws, r_start + 1, len(comps), 6)
ws.cell(row=r_start, column=7, value="Note").font = HDR_F
ws.cell(row=r_start, column=7).fill = HDR_FILL
ws.cell(row=r_start, column=7).border = BOX
ws.cell(row=r_start, column=7).alignment = WRAP_C

r = note(ws, r, "Twenty components. For comparison the web application's component list ran to fifty-two - "
                "which is the point: this is a shell around a finished application, not a second application.")

# ---- 03 Divergences -------------------------------------------------------
ws, r = sheet(wb, "03_Divergences", "Where the desktop interface differs, and why   [N3.2]",
              "NR-PAR-04 requires this list. Anything not on it is identical to the web application.")

div = [
    ["D-N01", "There is a menu bar.", "A desktop application without one is the odd thing, and File > Open is where people look for it. The web application has no menu because a browser page with a fake menu bar is worse than none.", "Convention", ""],
    ["D-N02", "Load workbook becomes File > Open plan, and Export becomes File > Save / Save As.", "The web application's verbs describe what a browser can do: pick a file, download a file. The desktop verbs describe what is actually happening - a plan is opened and saved in place.", "Persistence", ""],
    ["D-N03", "There is an identity dialog at launch.", "Nothing in the browser needed to know who you are. Sharing does.", "NR-USR-01", ""],
    ["D-N04", "There is an open-a-plan screen.", "The web application starts from 'upload or start blank'. The desktop one usually starts from 'the plan you had yesterday', so the empty state is a recent list rather than two buttons.", "NR-APP-03", ""],
    ["D-N05", "The status strip gains who you are and whether you hold the plan.", "Two facts that do not exist in a single-user browser tab, and that a user needs before the first keystroke rather than after it.", "NR-STO-10", ""],
    ["D-N06", "The title bar carries the plan name and an unsaved marker.", "A browser tab shows a page title; a window shows a document. The bullet is the standard Windows convention for unsaved work.", "Convention", ""],
    ["D-N07", "Export still exists, unchanged, alongside Save.", "They are different acts: Save keeps the plan, Export hands somebody an Excel file. Collapsing them would lose the archive that R-N19 depends on.", "Deliberate", ""],
    ["D-N08", "No file association, so a .prap file cannot be double-clicked open.", "The non-installed rule forbids the registry entry. Mitigated by the recent list and drag-and-drop onto the window.", "Q-N13", ""],
    ["D-N09", "The manual light/dark toggle is REMOVED. The window follows the Windows setting and offers no control of its own.", "Changed at review round 1. A desktop application that ignores the system theme looks broken; one that offers a second place to set it invites the two to disagree. The web application keeps its toggle, because a browser page cannot always read the system setting reliably.", "U-N04", "Changed"],
    ["D-N10", "Nothing else.", "Every tab, table, chart, filter, pop-up, value list, scroll behaviour, validation message and editing rule is the web application's, unchanged (N-11). Twenty-five rounds of review are not being re-opened.", "N-11", ""],
]
r_start = r
r = table(ws, r, ["ID", "Divergence", "Why", "Basis", "Your verdict"],
          div, [8, 54, 74, 14, 16], wrap_cols=(2, 3))
verdict_column(ws, r_start + 1, len(div), 5)
r = note(ws, r, "Nine divergences, and the tenth row is the important one. The temptation in a project like this "
                "is to improve the interface while rebuilding the shell around it; every such improvement would "
                "be a change you did not ask for, in a product you have already reviewed.")

# ---- 04 Messages ----------------------------------------------------------
ws, r = sheet(wb, "04_Messages", "Every string a user can read",
              "Wording is most of what these screens are. Mark up the text directly.")

msgs = [
    ["M-01", "Identity", "This is how colleagues will see you when you are editing a plan. It is not a login and it is not checked.", "NR-USR-08", ""],
    ["M-02", "Blocked, active", "Kim Min-jun (Data Management) is editing this plan. Started 09:14, active now.", "NR-STO-12", ""],
    ["M-03", "Blocked, silent", "Kim Min-jun (Data Management) has been editing this plan since 09:14, but their session has not responded since 09:22. The plan becomes free at 09:52.", "NR-STO-14", ""],
    ["M-04", "Blocked, expired", "Kim Min-jun's session stopped responding at 09:22 and no longer holds this plan. You may take over.", "NR-STO-14", ""],
    ["M-05", "Your own stalled claim", "This plan is held by an earlier session of your own, on this machine. Take it back?", "NR-STO-19", ""],
    ["M-06", "It has just freed", "This plan is free now. Start editing?", "NR-STO-15", ""],
    ["M-07", "Superseded", "Kim Min-jun saved this plan at 11:07. The figures on screen are from 09:00.", "NR-STO-16", ""],
    ["M-08", "Claim lost mid-save", "Your hold on this plan was taken over while you were working. Nothing has been saved. Save As a copy to keep your changes.", "NR-STO-10", ""],
    ["M-09", "Recovery", "Q3 resourcing.prap has 7 unsaved changes from 16:41 yesterday. The application closed before they were saved.", "NR-STO-07", ""],
    ["M-10", "Import, first question", "This plan already contains data. Update it from PRAP_SourceData_2026Q3.xlsx?", "NR-IMP-02", ""],
    ["M-11", "Import, only-here note", "'Only here' counts rows you added by hand that this file does not mention. Accepting a sheet never deletes them - an import adds and changes.", "NR-IMP-02", ""],
    ["M-12", "Protected file", "<name> could not be opened. It looks like it is protected by file security - open it in Excel first to unlock it, then import it again.", "NR-IMP-06", ""],
    ["M-13", "Not a plan file", "<name> is not a plan file. It may be a source workbook - use Import instead.", "NR-IMP-06", ""],
    ["M-14", "Newer version", "This plan was saved by version 1.2. This is version 1.0. Install the newer version to open it.", "NR-DEP-04", ""],
    ["M-15", "Read-only folder", "<path> cannot be written to. Save As somewhere else, or ask for write access.", "NR-DEP-09", ""],
    ["M-16", "No space", "There is not enough room to save. Nothing has been changed - your previous save is intact.", "NR-STO-05", ""],
    ["M-17", "Read-only mode", "You can look at everything - figures, charts, filters and exports all work. Only saving into this plan is held while somebody else has it.", "NR-STO-11", ""],
    ["M-18", "Last saved by", "Last saved by Kim Min-jun (Data Management), 11:07 today.", "NR-USR-07", ""],
]
r_start = r
r = table(ws, r, ["ID", "Where", "Text", "Requirement", "Your wording"],
          msgs, [7, 24, 88, 16, 34], wrap_cols=(3, 5))
verdict_column(ws, r_start + 1, len(msgs), 5)
for rr in range(r_start + 1, r_start + 1 + len(msgs)):
    ws.cell(row=rr, column=5).fill = INPUT_FILL
r = note(ws, r, "The last column takes replacement text, not a verdict - if a sentence is wrong, the quickest "
                "correction is the right sentence. M-08 and M-12 are the two worth the most attention: M-08 is "
                "the only message that appears when somebody may be about to lose work, and M-12 is the "
                "difference between a thirty-second fix and an afternoon hunting for a backup.")

# ---- 05 Open points -------------------------------------------------------
ws, r = sheet(wb, "05_Open_Points", "Open points for the Gate N3 review")
pts = [
    ["U-N01", "NC-04", "The recent list reads each listed plan's claim file, so it can say 'held by' before you open it.", "Cost per entry on a slow share, against finding out after you open.", "AGREED - the cost is acceptable. Built as proposed."],
    ["U-N02", "NC-08", "The recovery dialog lists the first three changes and then 'and 4 more'.", "Enough to recognise the work without a wall of text.", "CHANGED - list ALL of them, scrollable. Nobody can decide from three lines out of forty."],
    ["U-N03", "NC-01", "When is a session recognised as editing?", "The claim has to attach to something, and 'the first edit' was open to a loose reading.", "CLARIFIED - it attaches when an unsaved DATA CHANGE has happened, not when a user clicks something. Specification sheet 07 now says so in those words."],
    ["U-N04", "D-N09", "The window follows the Windows light/dark setting, with the manual toggle kept.", "Same rule as the web application, in desktop form.", "CHANGED - remove the manual toggle. The window follows Windows and offers no setting of its own."],
    ["U-N05", "NC-10", "The difference report is expandable to the individual rows, but collapsed by default.", "Sheet counts are what the decision is taken on.", "AGREED - and stated as a requirement rather than an option: the expansion is provided, collapsed by default."],
    ["U-N06", "NC-21", "NEW, raised at review: export the existing data so it can be updated OUTSIDE the application, in either format, and brought back.", "Bulk update or input, archiving, rollback, or simply reference. One manager receives the edited files and applies them - not every user importing their own.", "ACCEPTED as change C-N01. NR-IMP-08 added to the plan; assumption A-N13 records the one-manager scoping; specification sheet 09 carries the round trip it has to survive."],
]
r_start = r
r = table(ws, r, ["ID", "Component", "Point", "Reasoning", "Your answer"],
          pts, [8, 12, 52, 62, 34], wrap_cols=(3, 4, 5))
for rr in range(r_start + 1, r_start + 1 + len(pts)):
    ws.cell(row=rr, column=5).fill = NEW_FILL
r = note(ws, r, "All six settled. U-N03 and U-N06 reach beyond this document: U-N03 clarifies the write claim "
                "in the specification, and U-N06 adds a requirement to the plan - a change against an "
                "approved baseline, so it carries its own approval on sheet 12 of each.")

r = section(ws, r, "Approval - Gate N3")
appr = [["PRAP NewApp Component List v1.0", "Dan", "2026-08-13",
         "APPROVED - BASELINE. Gate N3 is closed and Step N4 authorised. 21 components, 10 divergences and "
         "18 user-visible strings are the contract for what the desktop application looks like and says. "
         "The approval also covers change C-N01, carried by plan v1.4 and specification v1.2: exporting is "
         "a supported way to edit a plan outside the application, and the round trip must lose nothing."]]
r_start2 = r
r = table(ws, r, ["Document", "Approver", "Date", "Decision"], appr, [34, 14, 14, 84], wrap_cols=(4,))
for cc in (1, 2, 3, 4):
    ws.cell(row=r_start2 + 1, column=cc).fill = NEW_FILL
r = note(ws, r, "STEP N3 IS COMPLETE. What was designed here now has to be built, and the tests that will "
                "prove it are already written down - sheet 08 of the specification lists them, including the "
                "three that can only be run on your own machine.")

wb.save(OUT)
print(f"Written: {OUT}")
print(f"  {len(comps)} components, {len(div)} divergences, {len(msgs)} messages, "
      f"{len(SCREENS)} screens")
