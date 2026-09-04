"""Generate the PRAP programming specification workbook (Step 2 deliverable).

Traceability is read from the approved plan rather than re-typed, so a requirement
cannot be dropped silently between the two documents.

    python tools/build_prog_spec.py

Output: docs/PRAP_Programming_Specification_v0.1.xlsx
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DOC_VERSION = "1.18"
DOC_STATUS = "APPROVED - Dan, 2026-08-02. Step 2 gate closed; this governs Step 4."
DOC_DATE = "2026-08-01"
# The APPROVED BASELINE is v2.0, and the traceability sheet used to read from it.
# It has to read from the CURRENT issue instead: a requirement added after the
# baseline - REQ-CAL-14 is the first - would otherwise be invisible here while
# check_consistency.py reported it as untraced, which is the drift both documents
# exist to prevent.
PLAN = "PRAP_Development_Plan_v2.44.xlsx"
PLAN_BASELINE = "PRAP_Development_Plan_v2.0.xlsx"    # approved, and unamended
ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / f"PRAP_Programming_Specification_v{DOC_VERSION}.xlsx"

FONT = "Arial"
NAVY = "1F3864"
TITLE_F = Font(name=FONT, size=16, bold=True, color=NAVY)
H1_F = Font(name=FONT, size=12, bold=True, color=NAVY)
HDR_F = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_F = Font(name=FONT, size=10)
BOLD_F = Font(name=FONT, size=10, bold=True)
NOTE_F = Font(name=FONT, size=9, italic=True, color="808080")
MONO_F = Font(name="Consolas", size=9)

HDR_FILL = PatternFill("solid", fgColor="2F5597")
BAND_FILL = PatternFill("solid", fgColor="F2F5FB")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
CODE_FILL = PatternFill("solid", fgColor="F7F7F7")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
WRAP_C = Alignment(vertical="top", wrap_text=True, horizontal="center")


def sheet(wb, name, title, subtitle=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_F
    row = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = NOTE_F
        row = 3
    ws.freeze_panes = f"A{row + 2}"
    return ws, row + 1


def table(ws, start_row, headers, rows, widths, wrap_cols=()):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font, c.fill, c.border, c.alignment = HDR_F, HDR_FILL, BOX, WRAP_C
    ws.row_dimensions[start_row].height = 28
    for r, data in enumerate(rows, start=start_row + 1):
        for i, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font, c.border = BODY_F, BOX
            c.alignment = WRAP if i in wrap_cols else Alignment(vertical="top")
            if (r - start_row) % 2 == 0:
                c.fill = BAND_FILL
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return start_row + len(rows) + 2


def section(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = H1_F
    return row + 1


def note(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = NOTE_F
    return row + 1


def lines(ws, row, texts):
    for t in texts:
        ws.cell(row=row, column=1, value=t).font = BODY_F
        row += 1
    return row


def code(ws, row, texts):
    for t in texts:
        c = ws.cell(row=row, column=1, value=t)
        c.font, c.fill = MONO_F, CODE_FILL
        row += 1
    return row + 1


wb = Workbook()
wb.remove(wb.active)

# ---- 00 Cover -------------------------------------------------------------
ws = wb.create_sheet("00_Cover")
ws.sheet_view.showGridLines = False
ws["A1"] = "Project Resource Assignment Program (PRAP)"
ws["A1"].font = Font(name=FONT, size=20, bold=True, color=NAVY)
ws["A2"] = "Programming Specification"
ws["A2"].font = Font(name=FONT, size=14, color=NAVY)

cover = [
    ("Document ID", "PRAP-SPEC-001"),
    ("Document type", "Programming specification (Step 2 deliverable)"),
    ("Version", f"v{DOC_VERSION}"),
    ("Status", DOC_STATUS),
    ("Issue date", DOC_DATE),
    ("Author", "Claude Code"),
    ("Governing document", f"{PLAN} - APPROVED BASELINE, Dan 2026-08-02"),
    ("Schema version specified", "11"),
    ("Repository", "Dan5050-now/project1"),
    ("Branch", "claude/project-resource-assignment-app-1vjdzh"),
]
r = 4
for k, v in cover:
    ws.cell(row=r, column=1, value=k).font = BOLD_F
    c = ws.cell(row=r, column=2, value=v)
    c.font, c.alignment = BODY_F, WRAP
    r += 1
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 84

r += 1
r = section(ws, r, "What this document is")
r = lines(ws, r, [
    "The plan says WHAT the application must do. This says HOW, in enough detail that the code can be written",
    "from it and checked against it. Every section cites the REQ-IDs it implements, and sheet 09 shows the",
    "reverse: each of the 70 requirements mapped to the section that satisfies it.",
    "",
    "Two things make this specification unusual, and both are deliberate:",
    "",
    "  - The data schema is not described in prose. It already exists as a working file,",
    "    templates/PRAP_SourceData_Template_v1.6.xlsx, and sheet 03 documents the parse contract against it.",
    "  - The calculation and validation logic already has a reference implementation in",
    "    tools/verify_source_workbook.py, which runs against the dummy data. Sheet 05 gives the pseudocode;",
    "    that script is the executable check that the pseudocode is right.",
    "",
    "So the specification describes behaviour that has been exercised, not behaviour that is only imagined.",
])

r += 1
r = section(ws, r, "Contents")
guide = [
    ("00_Cover", "Document control and reading guide."),
    ("01_Version_History", "Change log."),
    ("02_Scope", "What is specified here, what is deferred to Step 3, and the source documents."),
    ("03_Data_Schema", "The parse contract: every sheet, column, type and coercion rule."),
    ("04_Validation", "V-01..V-24 with trigger, severity and the exact message shown."),
    ("05_Calculation", "Period derivation, the load formula, aggregation and the two thresholds."),
    ("06_UI_Spec", "The four tabs, their components, filters and states."),
    ("07_Editing_IO", "Edit buffer, dirty state, cascading identifier edits, import and export."),
    ("08_Versioning", "Schema compatibility check and version display."),
    ("09_Traceability", "Every requirement in the plan, mapped to the section that implements it."),
    ("10_Open_Points", "The six points raised at the v0.3 review, with the answers given and what changed. None open."),
]
r = table(ws, r, ["Sheet", "Contents"], guide, [24, 86], wrap_cols=(2,))

# ---- 01 Version history ---------------------------------------------------
ws, r = sheet(wb, "01_Version_History", "Version history")
rows = [["1.0", "2026-08-02", "Claude Code", "Dan",
         "APPROVED. Content identical to v0.9; the version number changes because 1.0 marks an approved "
         "issue rather than a draft. Approved together with development plan v2.0 and UI component list "
         "v1.0. This document now governs Step 4: the code is written from it and checked against it.",
         "APPROVED 2026-08-02 by Dan"],
        ["0.9", DOC_DATE, "Claude Code", "Dan",
         "Plan change R-13, from the component-list v0.7 review. Sheet 06 gains two entries: the interim "
         "and final DB lock milestones are emphasised on the timeline in red and at a larger size, and "
         "the project source-data tab gains a utilisation graph with RELATIVE reference lines "
         "(REQ-DSH-12). The relative framing is the point - a project has no absolute ceiling or floor, "
         "so the person tab's threshold model does not transfer. Written against plan v1.10.", "Draft"],
        ["0.8", DOC_DATE, "Claude Code", "Dan",
         "Plan change R-12, from a reviewer question: why does the PersonPeriodWeight key include "
         "period_start when assignment_id is unique in Assignment? The key is unchanged and sheet 03 "
         "now says why - PersonPeriodWeight is a CHILD of Assignment and one assignment may carry "
         "several non-overlapping windows, so assignment_id alone would cap overrides at one per "
         "assignment. Answering it exposed two rules specified but never implemented: V-06's "
         "assignment-window overlap half, and referential integrity on PersonPeriodWeight.assignment_id. "
         "Both are now in the reference implementation, the second as new rule V-24. The dummy fixture "
         "gains an assignment with two windows. No schema change.", "Draft"],
        ["1.18", "2026-09-04", "Claude Code", "Dan",
         "R-36. Sheet 10 gains the CHANGE LOG and the COLUMN FILTERS; no schema change "
         "and no figure moves. The log records each saved change with its UTC time, the "
         "author, the record\'s natural key, the column and both values, plus the errors "
         "and warnings standing at each save and which were knowingly kept. It lives in "
         "memory for the session; the desktop shell APPENDS it to a monthly CSV in an "
         "audit folder at every save, and the browser - which has no folder to append to "
         "- hands it over from Export narrowed by a date range. Author comes from the "
         "Windows account in the desktop shell and is asked for once in the edit bar in "
         "the browser. The column filters are on the six tables named in the request and "
         "narrow the table only, never the figures. Written against plan v2.44.",
         "Issued"],
        ["1.17", "2026-09-04", "Claude Code", "Dan",
         "R-35, the date picker. Sheet 10 records two corrections. The month arrows "
         "stepped once and no further, because step() re-read the cell and snapped the "
         "grid back before stepping; the panel now distinguishes the cell steering the "
         "grid (opening, typing) from the user steering it (the arrows), and the cell "
         "keeps its text while navigating. \'Today\' wrote the UTC date rather than the "
         "date on the user\'s clock - a day out either way for anyone not near "
         "Greenwich - and now reads the local calendar date. No calculation changes. "
         "Written against plan v2.43.", "Issued"],
        ["1.16", "2026-09-04", "Claude Code", "Dan",
         "The 'Monthly demand by project' pop-up now states THE MONTH'S TOTAL across "
         "every project in view, under a rule, and the band's share of it - the two "
         "things 'Monthly demand by person' already carried and this one did not. A band "
         "read on its own cannot say whether the month came to five FTE or to fifty. "
         "Sheet 10 Graph 1 updated. No change to any figure: the total is the sum the "
         "chart was already drawing, and it is the same figure Graph 2 states for that "
         "month summed along the other axis - test_charts.py now holds the two to "
         "agreeing in WORDS as well as in pixels, month for month. Written against plan "
         "v2.42.", "Issued"],
        ["1.15", "2026-09-04", "Claude Code", "Dan",
         "R-33, SCHEMA 11, and no change to any figure - the 62-project fixture totals "
         "4,333.46 FTE-months before and after. Sheet 03 renames the sheet "
         "PeriodWeightStandard to PeriodFTEStandard: schema 10 already renamed its column "
         "to standard_fte on the grounds that it holds a magnitude in FTE, and leaving the "
         "SHEET called a weight kept the misreading alive that R-32 had to correct. A "
         "schema 10 workbook still opens - the old sheet name is translated on read and "
         "reported once, as a renamed column is. Sheet 10 renames the panel to 'Standard "
         "period FTE for project types'. Sheet 05 corrects the monthly-estimation notes, "
         "which still described the pre-R-32 product of four factors; they now state the "
         "project month and the assignment month separately, because under R-32 the two "
         "are arrived at differently. Sheet 09 records that the Python shell's sticky "
         "chrome and the page's sticky band no longer overlap. Written against plan "
         "v2.41.", "Issued"],
        ["1.14", "2026-09-03", "Claude Code", "Dan",
         "R-32: REQ-CAL-19 amended. A project-month IS its standard x period weight x "
         "the month it ran, and the people on it DIVIDE that - person_weight, coverage "
         "and the role factor all make a CLAIM, and the share is the claim over the sum "
         "of the claims. Sheet 05 replaces the formula with demand x share and records "
         "that the shares add to one by construction. Sheet 08 records month_run and "
         "demand_fte on the results export, and that every row is now two numbers rather "
         "than five. Written against plan v2.40.", "Issued"],
        ["1.13", "2026-09-02", "Claude Code", "Dan",
         "R-31, REQ-CAL-19, SCHEMA 10: the standard monthly FTE reaches the calculation. "
         "Sheet 05 replaces the formula: standard_fte x period weight x role_share x "
         "person weight x month coverage, with role_share normalised over the roles "
         "ACTUALLY STAFFED so the shares add to one. Sheet 03 renames "
         "PeriodWeightStandard.weight to standard_fte and records that it is a magnitude "
         "rather than a multiplier, and that ProjectPeriod.weight is now an adjustment "
         "around 1.00. Sheet 08 records the four columns the results export gains so a "
         "reader can see where a figure got its size. Written against plan v2.39.",
         "Issued"],
        ["1.12", "2026-08-30", "Claude Code", "Dan",
         "R-30, REQ-CAL-18, SCHEMA 9: a monthly figure may be STATED instead of "
         "calculated. Sheet 03 gains the MonthlyEstimate sheet and its key - scope + "
         "ref_id + month, the only key in the workbook that carries the name of its own "
         "parent table, because `scope` decides whether ref_id points at a project or an "
         "assignment - and estimation_type on Project and Assignment. Sheet 04 gains "
         "V-31 and V-32, both raised from the CALCULATION rather than from a sheet, so "
         "neither can refuse an edit. Sheet 05 records the order the two levels are "
         "applied in and why a project figure SCALES its people rather than replacing "
         "them: the results export guarantees every total is exactly the sum of its "
         "detail rows, and a project month that did not equal the sum of its people "
         "would break it. Sheet 08 records the four columns the results export gains so "
         "a stated row can be told from a calculated one. Written against plan v2.38.",
         "Issued"],
        ["1.11", "2026-08-28", "Claude Code", "Dan",
         "R-29, REQ-OUT-06: the calculated-results export. Sheet 08 gains its seven "
         "sheets column by column, the two properties that make it checkable - monthly "
         "figures summed FROM the detail rather than from the engine, and every detail "
         "row reconciling to its own four numbers - and why it deliberately does not "
         "round-trip. Sheet 05 records that calculate() now returns `lines`, one per "
         "assignment-month, so the explanation and the figure come from the same pass. "
         "Written against plan v2.37.", "Issued"],
        ["1.10", "2026-08-28", "Claude Code", "Dan",
         "R-28, REQ-IMP-14. Sheet 07 gains the config check on import: what is compared, "
         "when, and why a Config row is not treated like any other row in the file - "
         "every other sheet describes the plan, Config describes how the plan is read. "
         "The comparison itself is in core/06a_diff.js beside the import difference "
         "engine, because both shells import. Written against plan v2.36.", "Issued"],
        ["1.9", "2026-08-27", "Claude Code", "Dan",
         "R-26: the parse contract's row for capacity_unit named 'percent', which the "
         "application has never implemented - anything that is not 'hours' displays as "
         "FTE. The row now names the two real values and says what the unit means: FTE "
         "is a WEIGHT, 1.00 being one person for a full month, so ordinary values run "
         "about 0.1 to 1.0. R-27: sheet 04 gains V-30, and sheet 06 records that the "
         "Configuration table offers neither 'Delete' nor '+ row' - its nine settings are "
         "read BY NAME, so a new row is read by nothing and a deleted one hands its figure "
         "to a built-in default without saying so. What a user changes there is a VALUE, "
         "and the value cell stays as editable as any other. Written against plan v2.35.",
         "Issued"],
        ["1.8", "2026-08-27", "Claude Code", "Dan",
         "R-21 to R-25, all in the application rather than the arithmetic - no figure "
         "moves. Sheet 04 gains the RULE CLASS: must / conditional / incomplete, what "
         "each one does at a cell edit and at Save, and the class of every rule that is "
         "not must. Sheet 06 gains the calendar on a date column (an offer beside the "
         "cell, never a gate in front of it), the drawn scrollbars and why the browser's "
         "own could not be used, the single opaque sticky band, and the rule that a "
         "generator is only offered where it can run. Written against plan v2.34.", "Issued"],
        ["1.7", "2026-08-27", "Claude Code", "Dan",
         "R-19 and R-20. Sheet 04: V-03 and V-23 are marked REPORTED, NEVER REFUSING - "
         "the retirement of V-28 at v1.6 fixed one third of the problem, because all "
         "three rules asked the same question and all three refused the row. V-23 also "
         "moves out of the validation pass and into the CALCULATION: keyed on the whole "
         "composition the lookup uses and counted in the person-months actually "
         "calculated at 1.00, which is both the right grouping and the reason it can no "
         "longer refuse anything. Sheet 05 gains REQ-CAL-17: a project's window is the "
         "span of its PERIODS, not of its milestones, so no month outside the plan draws "
         "resource at weight 1.00. Written against plan v2.33.", "Issued"],
        ["1.6", "2026-08-26", "Claude Code", "Dan",
         "R-18: V-28 IS RETIRED. Sheet 04 records the retirement rather than dropping "
         "the row. The rule was right about the data and wrong about the moment: an "
         "error refuses the edit that raised it, and V-28 was the first assumptions "
         "error that did not need the project to have periods, so it fired on a project "
         "still being built and stopped a user entering their own assignments until the "
         "standing assumptions caught up. V-03 and V-23 still cover the case at the "
         "point where it can be acted on. V-27, REQ-CAL-16 and the absorption arithmetic "
         "are unchanged; no figure moves. Written against plan v2.32.", "Issued"],
        ["1.5", "2026-08-26", "Claude Code", "Dan",
         "R-17, REQ-CAL-16 and three new rules. Sheet 05 gains the ABSORPTION rule - an "
         "unstaffed role's factor lands on the role named to cover for it, per month, "
         "one hop - and says why the mapping is a column on RoleFactor rather than two "
         "role names in the code. Sheet 04 gains V-27 and V-28, which ask on the PROJECT "
         "what V-19 and V-23 ask on its periods, and V-29, which reports the case "
         "absorption cannot fix. Schema 7 to 8. Written against plan v2.31.", "Issued"],
        ["1.4", "2026-08-25", "Claude Code", "Dan",
         "R-16, SCHEMA 7. outsourcing_type becomes outsourcing_scope_det and becomes FREE "
         "TEXT - a note for a person, never read by the calculation. V-25 is RETIRED: it "
         "existed only to police two controlled vocabularies on the same axis, and there "
         "is now only one. Sheet 04 records the retirement rather than dropping the row, "
         "because a rule id that vanishes leaves a reader wondering whether they have an "
         "old document. Sheet 03 gains the RENAMED-COLUMN rule: a workbook carrying the "
         "old name is read, its values moved across, and the reader told once - a rename "
         "is the one schema change that otherwise loses a filled-in column in silence. "
         "Written against plan v2.30.", "Issued"],
        ["1.3", "2026-08-25", "Claude Code", "Dan",
         "R-15, REQ-CAL-15: both assignment dates are optional and a blank one means the "
         "project's own, so an assignment with neither runs for the whole project. Sheet "
         "05 gains the window rule and says why it is one function rather than two - the "
         "months a person is counted IN must not differ from the months they are counted "
         "AMONG, or a shared role stops adding up to one. Written against plan v2.29.",
         "Issued"],
        ["1.2", "2026-08-22", "Claude Code", "Dan",
         "TWO CHANGES. (1) R-13: the role factor is DIVIDED between the people holding "
         "that role on that project in that month. Sheet 05 carries the formula, the "
         "counting rule and the worked example; the count is per month, by distinct "
         "people, and each person's own weight applies to their share afterwards. New "
         "setting split_shared_role_fte, default 1, restores the previous arithmetic when "
         "set to 0. (2) R-14: PeriodWeightStandard and RoleFactor are DELIVERED FILLED IN "
         "- 84 and 429 rows of default assumptions - and the application's blank start "
         "seeds the identical figures, read from the template rather than restated. Sheet "
         "03 records where they come from and what a company is expected to do with them. "
         "Written against plan v2.28.", "Issued"],
        ["1.1", "2026-08-20", "Claude Code", "Dan",
         "SCHEMA 6. Two changes to the data model, requested 2026-08-20. (1) work_scope_type joins "
         "PeriodWeightStandard and RoleFactor beside clinical_phase, so the keys become "
         "(project_type, clinical_phase, work_scope_type, period_name) and that plus role_name. "
         "Project gains the same column, because the key selects a row FOR A PROJECT and nothing "
         "could choose between the scope rows without it. A ROW WHOSE work_scope_type IS EMPTY "
         "APPLIES TO EVERY SCOPE, and a project falls back to it - which is what keeps RoleFactor "
         "at 429 rows instead of 1,269, and what lets a schema 5 file go on calculating. (2) "
         "'Biosimilar CT' becomes 'Biosimilar CT (Healthy)' and 'Biosimilar CT (Patient)'; a type "
         "is recognised as a clinical trial by the START of its name. New rules V-25 (the two "
         "scope columns must not contradict) and V-26 (the retired type, reported with the two it "
         "became). Sheets 03, 04, 05 and 06 updated. Written against plan v2.27.", "Issued"],
        ["0.7", DOC_DATE, "Claude Code", "Dan",
         "Plan change R-11 applied. The conduct phase is split by NAME: 'Conduct (interim)' where the "
         "project has an interim DB lock and the stretch runs before it, 'Conduct (final)' after it or "
         "where there is no interim lock. Period names are therefore unique within a project, and "
         "ProjectPeriod is keyed on (project_id, period_name) - sheets 03 and 05 updated, V-18 becomes a "
         "plain uniqueness check, and period_seq carries order only. The clinical period set grows to "
         "seven names, so PeriodWeightStandard is 56 rows and RoleFactor 289. Schema 4 to 5. Weights "
         "unchanged: the dummy dataset returns identical load figures, which is the evidence that this "
         "renamed without reweighting. Written against plan v1.8.", "Draft"],
        ["0.6", DOC_DATE, "Claude Code", "Dan",
         "Plan change R-10 applied: the role factor is keyed on project type, clinical phase, period "
         "and role rather than type and role alone. Sheet 03 documents the two new RoleFactor columns "
         "and the four-part key; sheet 05 selects the factor by the period the month falls in; V-03 "
         "reworded and V-23 added for a factor missing on a period an assignment actually spans. "
         "Schema version 3 to 4. Sheet 05 also records the double-counting risk this creates, since "
         "RoleFactor and PeriodWeightStandard now vary over the same three dimensions. Sheet 06 gains "
         "the scroll-containment rule for every table on the source-data and assumptions tabs. "
         "Written against plan v1.7.", "Draft"],
        ["0.5", DOC_DATE, "Claude Code", "Dan",
         "Component-list v0.3 review applied (plan change R-09). Sheet 06 gains a fourth tab, "
         "'General assumptions', carrying PeriodWeightStandard, RoleFactor, Config and Lists "
         "(REQ-DSH-11) - and the FTE/hours unit moves there, since it is a setting rather than a "
         "filter. The filter set gains clinical phase. Row insertion is specified: a new row lands "
         "immediately below the row acted on (REQ-IMP-11). The load stamp carries its time zone. The "
         "timeline is specified as coloured by period name with weight as a lightness step, which "
         "supersedes design decision D-06 on the reviewer's own instruction. Tooltip content specified "
         "for the demand chart and the timeline. Written against plan v1.6.", "Draft"],
        ["0.4", DOC_DATE, "Claude Code", "Dan",
         "Answers to the six open points applied. S2-01 and S2-05: both thresholds stay ABSOLUTE, with the "
         "under-allocation floor moved 0.80 -> 0.60; the relative-threshold proposal is withdrawn and V-22 added "
         "to catch a capacity below the floor. S2-04: repeated period names are numbered on screen, 'Conduct (1)' "
         "and 'Conduct (2)' (REQ-DSH-10). S2-06: target volume re-baselined to 100 projects and 1,000 people, "
         "which makes table virtualisation and chart aggregation requirements rather than optimisations "
         "(REQ-DSH-09). S2-02 and S2-03 confirmed as drafted. ProjectPeriod is keyed on project_id + "
         "period_name + period_start. Written against plan v1.5.", "Draft"],
        ["0.3", DOC_DATE, "Claude Code", "-",
         "Change R-05: project_type split into 'NewDrug CT' and 'Biosimilar CT'. Parse contract, keys and "
         "schema version updated; RoleFactor and PeriodWeightStandard are now keyed on the type. Sheet 06 "
         "gains the UI changes requested at the Step 3 review: a global filter bar with Reset, per-project "
         "stacking, horizontal scroll regions, and row expansion on both Overall tables.", "Draft"],
        ["0.2", "2026-08-01", "Claude Code", "-",
         "Parse contract updated for change R-04: a free-text note column on every sheet, and source schema "
         "version 2. Column counts on the sheet-index table adjusted. No change to validation, calculation, "
         "UI or IO behaviour.", "Draft"],
        ["0.1", "2026-08-01", "Claude Code", "-",
         "First draft. Written against development plan v1.2 (approved baseline). Data schema documented from "
         "the delivered template; calculation and validation cross-checked against the reference implementation "
         "in tools/verify_source_workbook.py and the dummy dataset.", "Draft"]]
r = table(ws, r, ["Version", "Date", "Author", "Reviewer", "Summary of change", "Status"],
          rows, [10, 12, 15, 14, 92, 14], wrap_cols=(5,))
r = note(ws, r, "Version alignment across plan, specification and application is recorded on sheet 09 of the plan.")

# ---- 02 Scope -------------------------------------------------------------
ws, r = sheet(wb, "02_Scope", "Scope and source documents")

r = section(ws, r, "Source documents")
src = [
    [PLAN, "Development plan, v1.3 baseline approved by Dan 2026-08-01 plus changes APPROVED BASELINE 2026-08-02. 70 requirements, 24 validation rules, 11 decisions, source schema version 5.", "Governs this document"],
    ["templates/PRAP_SourceData_Template_v1.6.xlsx", "The blank source workbook as delivered.", "The schema on sheet 03 documents this file"],
    ["templates/PRAP_SourceData_Dummy_v1.8.xlsx", "34 NewDrug CT + 16 Biosimilar CT + 12 'Others', 20 people, 289 assignments over 73 months.", "The acceptance data for sheet 05"],
    ["tools/verify_source_workbook.py", "Reference implementation of parsing, validation and the monthly engine.", "Executable check on sheets 04 and 05"],
    ["docs/STEP2_OPEN_POINTS.md", "Points raised while building the template.", "Carried into sheet 10"],
]
r = table(ws, r, ["Document", "What it is", "Relationship to this specification"], src, [46, 62, 46], wrap_cols=(2, 3))

r = section(ws, r, "In scope for this specification")
ins = [
    ["The parse contract for the source workbook: sheets, columns, types, coercion and defaults."],
    ["All 21 validation rules with their exact severity and message."],
    ["Period derivation for both project types, including every degenerate timeline."],
    ["The monthly load calculation, aggregation, and the over- and under-allocation rules."],
    ["The behaviour of each dashboard tab: components, filters, interactions and empty states."],
    ["On-screen editing: the edit buffer, dirty state, cascading identifier changes, and export round-tripping."],
    ["Schema version checking and version display."],
]
r = table(ws, r, ["In scope"], ins, [132], wrap_cols=(1,))

r = section(ws, r, "Deferred to Step 3")
defer = [
    ["Exact visual layout, spacing, colour and typography.", "Step 3 fixes the UI design; sheet 06 fixes behaviour and content, not pixels."],
    ["Which chart library, if any, is embedded.", "Sheet 06 specifies what each graph must show. The rendering choice is made at Step 3 against the prototype."],
    ["Keyboard shortcuts and accessibility details.", "Reviewed with the prototype, where they can be tried rather than described."],
]
r = table(ws, r, ["Deferred", "Why"], defer, [56, 76], wrap_cols=(1, 2))

# ---- 03 Data schema -------------------------------------------------------
ws, r = sheet(wb, "03_Data_Schema", "Data schema - the parse contract",
              "Documents templates/PRAP_SourceData_Template_v1.6.xlsx. Sheet and column names are matched "
              "exactly and case-sensitively.")

r = section(ws, r, "Reading the workbook")
r = code(ws, r, [
    "  for each required sheet:",
    "      if absent                     -> fatal, abort the load and report (V-00)",
    "      read row 1 as the header",
    "      for each column in the contract below:",
    "          locate it by exact name; if absent -> fatal for a required column,",
    "                                                ignore for an optional one",
    "      read rows 2..n; a row whose every cell is empty is skipped, not an error",
    "      coerce each cell per its type; a failed coercion is an error on that row",
])

r = section(ws, r, "Type coercion")
types = [
    ["Text", "Trim leading and trailing whitespace. An empty string becomes null.", "Never coerce a number to text silently - a project_id of 001 read as 1 breaks every join."],
    ["Date", "Accept an Excel serial date or an ISO yyyy-mm-dd string. Normalise to a date with no time part.", "Any other format is an error, never a guess (REQ-NFR-05). A dd/mm vs mm/dd guess is silently wrong half the time."],
    ["Decimal", "Accept a number, or a string parseable as one with '.' as the decimal separator.", "A percentage-formatted cell reads as its underlying fraction; 40% is 0.4, which is what person_weight expects."],
    ["Integer", "As Decimal, then require a whole number.", "A fractional period_seq is an error."],
    ["List", "Text, then checked against the Lists sheet by V-11.", "The check is a warning, not a coercion - an unrecognised value is kept and reported."],
]
r = table(ws, r, ["Type", "Coercion", "Why it matters"], types, [12, 62, 66], wrap_cols=(2, 3))

r = section(ws, r, "Sheets and keys")
sheets = [
    ["Project", "project_id", "-", "23", "Master."],
    ["Milestone", "project_id + milestone_name + milestone_date", "Project", "6", "milestone_name is NOT unique alone - 'Inspection' repeats (REQ-PRJ-13)."],
    ["ProjectPeriod", "project_id + period_name", "Project", "7", "Since R-11 no period name repeats in a project, so the name alone identifies the row. period_seq carries order, not identity (V-18)."],
    ["PeriodFTEStandard", "project_type + clinical_phase + work_scope_type + period_name", "-", "6", "84 baseline rows at schema 6: three trial types keyed separately (R-05, R-12), four phases, seven periods (R-11), with work_scope_type EMPTY. A row per scope as well would be 252. 'Others' take manual weights (Q-28)."],
    ["RoleFactor", "project_type + clinical_phase + work_scope_type + period_name + role_name", "-", "7", "429 baseline rows at schema 6 - the largest sheet in the workbook, and 1,269 if every scope were spelled out. Keyed on all five so a role's burden can vary across the life of a project (R-10) and with how much of the work is kept (R-12). clinical_phase is EMPTY on the nine 'Others' rows - the lookup must match null to null, not fall through."],
    ["Person", "person_id", "-", "12", "Master."],
    ["Assignment", "assignment_id", "Person, Project, RoleFactor", "11", "One row per person + project + role."],
    ["PersonPeriodWeight", "assignment_id + period_start", "Assignment", "5", "Optional. Overrides person_weight for its window. period_start IS part of the key: one assignment may carry several non-overlapping windows, so assignment_id alone does not identify a row. See the note below."],
    ["MonthlyEstimate", "scope + ref_id + month", "Project or Assignment", "6", "Schema 9. Monthly FTE STATED rather than calculated (REQ-CAL-18). The only sheet whose parent depends on a value IN the row: scope says whether ref_id names a project or an assignment, which is why it is part of the key. Read only where the owning row carries estimation_type = 'manual'."],
    ["Lists", "list_name + value", "-", "3", "Value lists, long format. Each list occupies a contiguous block."],
    ["Config", "parameter", "-", "3", "Thresholds and settings."],
]
r = table(ws, r, ["Sheet", "Key", "References", "Cols", "Note"], sheets, [22, 40, 26, 7, 45], wrap_cols=(2, 5))

r = section(ws, r, "Why PersonPeriodWeight is keyed on TWO columns   [R-12]")
r = lines(ws, r, [
    "assignment_id is unique in Assignment - it is that sheet's primary key. It is NOT unique in",
    "PersonPeriodWeight, and it is not meant to be: this is a child table of windows, and one assignment",
    "may carry several.",
])
r = code(ws, r, [
    "  ASG-902  2026-07-01 .. 2026-09-30   0.20   part-time, parental leave",
    "  ASG-902  2027-04-01 .. 2027-06-30   0.75   covering the start-up peak",
    "           everything outside those windows  ->  the assignment's own person_weight",
])
r = lines(ws, r, [
    "One person, one project, one role, one assignment - three different weights over its life. That is",
    "what the table is for, and REQ-PSN-05, V-06 and the data model all assume it.",
])
r += 1
r = table(ws, r, ["If the key were assignment_id alone", "Consequence"],
          [["An assignment could carry at most ONE override window.",
            "The case above becomes inexpressible. A second spell of changed capacity would need a second Assignment row."],
           ["Splitting into ASG-902a / ASG-902b / ASG-902c would be the workaround.",
            "It breaks Assignment's own meaning - one row per person + project + role - and fragments the person's history across rows that are not really different assignments."],
           ["V-06's assignment-window half becomes dead.",
            "Two windows could never coexist, so they could never overlap. A rule that can never fire is worse than no rule: it reads as protection."]],
          [46, 92], wrap_cols=(1, 2))
r = note(ws, r, "So the key stays (assignment_id, period_start). period_start is the right second column "
                "because windows are ordered in time and cannot overlap - which V-06 enforces - so the "
                "start date identifies one window unambiguously.")
r += 1
r = note(ws, r, "Two gaps surfaced while answering this, both now closed. V-06's assignment-window half was "
                "specified from plan v1.0 but never implemented, so an overlapping pair passed silently and "
                "the weight that applied in the shared months depended on the order rows happened to sit in "
                "the file. And nothing checked that a PersonPeriodWeight row pointed at a real assignment, "
                "so an orphan override was accepted and then ignored without a word - the user's typed "
                "weight simply never applied. V-24 covers the second. The dummy fixture only ever had one "
                "window per assignment, which is why neither surfaced: a fixture that does not exercise a "
                "path cannot test it.")
r += 1

r = section(ws, r, "Derived columns - read, do not trust")
der = [
    ["Project.total_period_months", "Recompute from start_date and end_date.", "The template holds a formula, but an exported or hand-edited file may not. Recomputing costs nothing and cannot disagree with the dates."],
    ["Milestone.project_name", "Recompute by lookup on project_id.", "Denormalised for display only. V-13 reports a mismatch; the master value always wins."],
    ["Assignment.person_name", "Recompute by lookup on person_id.", "As above."],
]
r = table(ws, r, ["Column", "On import", "Reason"], der, [30, 44, 66], wrap_cols=(2, 3))

r = section(ws, r, "Free-text note columns   [R-04]")
r = lines(ws, r, [
    "Every sheet carries at least one. They are parsed as Text, carried through export unchanged, and never",
    "read by the calculation - so a note can hold anything without affecting a single figure.",
])
r += 1
notes_tbl = [
    ["Project", "note_1 .. note_5"], ["Milestone", "note_1"],
    ["ProjectPeriod", "note_1"], ["PeriodFTEStandard", "note_1"],
    ["RoleFactor", "role_note"], ["Person", "note_1 .. note_5"],
    ["Assignment", "note_1 .. note_3"], ["PersonPeriodWeight", "reason"],
    ["MonthlyEstimate", "note_1"],
    ["Lists", "note_1"], ["Config", "note"],
]
r = table(ws, r, ["Sheet", "Note column(s)"], notes_tbl, [26, 34])
r = note(ws, r, "A note column that is absent is not an error - it is optional on every sheet. An import must not "
                "fail because someone deleted a column they were not using.")
r += 1

r = section(ws, r, "Renamed columns   [R-16]")
r = lines(ws, r, [
    "A rename is the one schema change that can lose data in SILENCE. The old column is",
    "read, assigned to a key nothing looks at, and the new column comes back empty: the file",
    "still opens, every rule still passes, and the value is simply gone.",
    "",
    "So both readers translate the header on the way in, and the reader is told once:",
    "",
    "    schema 6 -> 7    Project.outsourcing_type  ->  Project.outsourcing_scope_det",
    "",
    "This is not leniency about the schema - the workbook is still expected to be the current",
    "one, and V-09 still reports the version mismatch. It is refusing to be the reason",
    "somebody loses a column they filled in last week.",
])
r += 1

r = section(ws, r, "The delivered default assumptions   [R-14]")
r = lines(ws, r, [
    "PeriodFTEStandard and RoleFactor are DELIVERED FILLED IN - 84 and 429 rows. They used",
    "to arrive empty, and a plan started inside the application filled both grids with a",
    "placeholder 1.00.",
    "",
    "1.00 everywhere is not a cautious starting point; it is the absence of one. At 1.00 the",
    "period weight and the role factor cancel out of the multiplication, so every project",
    "reduced to person_weight x coverage and the application produced figures that looked",
    "like an answer and were not one. Nothing said so, because nothing was wrong.",
])
r += 1
dflt = [
    ["Defined once", "tools/build_source_workbook.py - DEFAULT_PHASE_PROFILE and the tables beside it.", "One definition"],
    ["Read by the template", "PeriodFTEStandard and RoleFactor are written from it at build time.", "R-14"],
    ["Read by the blank start", "tools/build_app_seed.py lifts the same rows OUT OF THE TEMPLATE into SEED_PWS and SEED_RF. Not restated - lifted, so 'the workbook and the application hold the same assumptions' is a fact rather than a promise.", "R-14"],
    ["Read by the dummy data", "The same baseline, plus illustrative scope-specific rows on top.", "R-12"],
    ["Checked", "check_consistency.py compares the embedded seed with the template row for row, and fails if either grid has reverted to one flat value.", "R-14"],
    ["Still built from the lists", "The blank start builds the GRID from the value lists, so a company that adds a role gets that combination immediately. It arrives at 1.00, marked as a placeholder, because nobody has supplied a figure for a role nobody had yesterday.", "-"],
    ["Baseline only", "Every default row has work_scope_type EMPTY, which covers every scope. Shipping scope-specific defaults would mean asserting how much cheaper an outsourced trial is - the judgment a company has to make for itself.", "R-12"],
    ["Marked as defaults", "Every row carries 'Default assumption - replace with your own figure' in its note column, so a figure nobody has reviewed can be told from one somebody chose.", "R-14"],
]
r = table(ws, r, ["Point", "Detail", "Basis"], dflt, [30, 100, 14], wrap_cols=(2,))
r += 1

r = section(ws, r, "Config parameters")
cfg = [
    ["schema_version", "Integer", "11", "Compared with the version this application expects (sheet 08)."],
    ["absorb_unstaffed_role_factor", "Integer", "1", "1 = where nobody holds a role on a project, its factor is added to the role named in RoleFactor.absorbed_by (sheet 05). 0 = an unstaffed role costs nothing, the arithmetic of every version before this one."],
    ["split_shared_role_fte", "Integer", "1", "1 = the role factor is divided between the people sharing a role in a month (sheet 05). 0 = each carries the whole factor, the arithmetic of every version before this one. A switch, not a threshold - so the Config reader must distinguish a value of 0 from an absent value, which is the defect this setting exposed."],
    ["fte_hours_per_month", "Decimal", "160", "Converts FTE to hours for display."],
    ["over_allocation_fte", "Decimal", "1.50", "Absolute, not scaled by capacity_fte (S2-01). See sheet 05."],
    ["under_allocation_fte", "Decimal", "0.60", "Absolute, not scaled by capacity_fte. Moved from 0.80 at S2-05. See sheet 05."],
    ["under_allocation_min_months", "Integer", "3", "Consecutive months before a run is flagged."],
    ["default_horizon_months", "Integer", "24", "Months shown on opening."],
    ["capacity_unit", "List", "FTE", "'FTE' or 'hours'. FTE is a WEIGHT - 1.00 is one person for a full month, so ordinary values run about 0.1 to 1.0; 'hours' multiplies it by fte_hours_per_month. Anything else displays as FTE. 'percent' was named in the contract until v1.9 and was never implemented."],
]
r = table(ws, r, ["parameter", "Type", "Default", "Use"], cfg, [30, 12, 12, 86], wrap_cols=(4,))
r = note(ws, r, "schema_version stepped from 3 to 4 at R-10 (RoleFactor gained two columns), from 4 to 5 at "
                "R-11, and from 5 to 6 at R-12. R-11 changed no columns at all - but it retired the period "
                "name 'Conduct', so every ProjectPeriod row of a v4 file would fail V-15. A value set is part "
                "of the contract, not decoration. R-12 does both: three new columns AND a retired "
                "project_type. V-09 reports the mismatch, and a schema 5 file still CALCULATES - its "
                "standards rows carry no scope, which is exactly the every-scope row - so what it needs is "
                "the type renamed, which V-26 names.")
r += 1
r = note(ws, r, "A missing Config row falls back to the default above and raises a warning. A Config value that fails "
                "coercion is an error - a threshold read as text would silently disable a flag.")

# ---- 04 Validation --------------------------------------------------------
ws, r = sheet(wb, "04_Validation", "Validation rules",
              "Every rule runs on import and again on any on-screen edit (REQ-IMP-09). The load never stops at "
              "the first problem: findings are collected and presented as one report (REQ-IMP-02).")

r = section(ws, r, "Severities")
sev = [
    ["Fatal", "The workbook cannot be used at all.", "Abort the load. Show what is missing and the template download link.", "Missing sheet; missing required column."],
    ["Error", "This row cannot be trusted.", "Reject the row, keep loading the rest, list it with its sheet and row number.", "Unknown project_id on an assignment; end before start."],
    ["Warning", "Usable, but someone should look.", "Keep the row, list the finding.", "A value not in its list; an assignment outside its project's dates."],
    ["Information", "Explains a decision the reader might otherwise question.", "List only.", "V-21: an early inspection did not open period 7."],
]
r = table(ws, r, ["Severity", "Meaning", "Behaviour", "Example"], sev, [14, 40, 56, 46], wrap_cols=(2, 3, 4))

r = section(ws, r, "The rules")
rules = [
    ["V-00", "Fatal", "A required sheet or column is absent.", "Sheet 'Assignment' not found. The workbook must contain all 10 sheets - download the template to compare."],
    ["V-01", "Error", "Assignment.project_id not found in Project.", "Assignment ASG-014 refers to project PRJ-099, which does not exist."],
    ["V-02", "Error", "Assignment.person_id not found in Person.", "Assignment ASG-014 refers to person PSN-099, which does not exist."],
    ["V-03", "Error", "Assignment.role_name appears nowhere in RoleFactor for that project's type.", "Assignment ASG-014: role 'Main staff' is not valid for a project of type 'NewDrug CT'. Valid roles for this type: ..."],
    ["V-04", "Warning", "project_category empty on either clinical trial type.", "Project PRJ-003 is a clinical trial with no product category."],
    ["V-05", "Error", "An end date precedes its start date.", "Project PRJ-003: end_date 2026-01-01 is before start_date 2026-06-01."],
    ["V-06", "Error", "Two periods of one project, or two override windows of one assignment, overlap.", "ASG-902: override windows overlap - 2026-07-01..2026-09-30 and 2026-08-15..2026-11-30. Which weight applies in the shared months would depend on row order."],
    ["V-07", "Warning", "Assignment dates fall outside the project's own dates.", "Assignment ASG-014 runs to 2029-06-30, after project PRJ-005 ends on 2029-03-31."],
    ["V-08", "Error", "A duplicate identifier.", "person_id PSN-004 appears on rows 5 and 12."],
    ["V-09", "Warning", "Config.schema_version differs from the expected version.", "This file is schema version 2; this application expects version 1. Some columns may be ignored."],
    ["V-10", "Warning", "A clinical trial missing clinical_phase or any *_setup value.", "Project PRJ-004 has no RBQM_setup recorded."],
    ["V-11", "Warning", "A list-typed value not found in the Lists sheet.", "Project PRJ-002: outsourcing_type 'Partial in house' is not a known value. Did you mean 'Full In-house'?"],
    ["V-12", "Warning", "A project's periods leave a gap or do not cover its full span.", "Project PRJ-006: 2026-07-01 to 2026-07-31 belongs to no period. Those months are calculated at weight 1.00."],
    ["V-13", "Warning", "A denormalised name does not match its master row.", "Assignment ASG-014 records person_name 'Kim' but PSN-001 is 'Kim S.'. The master value is used."],
    ["V-14", "Error / Warning", "A boundary milestone is out of chronological order, or a milestone falls outside the project window.", "Project PRJ-003: 'First SIV' 2025-12-01 precedes 'CTA submission' 2026-01-15, so periods cannot be derived."],
    ["V-15", "Error", "A period_name not in the set for that project's type.", "Project PRJ-006 is type 'Others' but has a period named 'Conduct (final)'. Valid: Planning, Develop, Close."],
    ["V-16", "Error", "A clinical trial lacking CTA submission or any DB lock.", "Project PRJ-007 has no DB lock milestone, so its periods cannot be derived. Enter them manually or add the milestone."],
    ["V-17", "Error", "An edit would orphan a reference (see sheet 07).", "PSN-001 cannot be deleted: 3 assignments still refer to it."],
    ["V-18", "Error", "Within a project, period_name is not unique, or period_seq is duplicated.", "Project PRJ-003: period_name 'Conduct (final)' appears 2 times; (project_id, period_name) must be unique. period_seq must also be unique - it fixes their order."],
    ["V-19", "Error", "A clinical trial with no clinical_phase, or no PeriodFTEStandard row for its phase AND scope - counting the every-scope row.", "Project PRJ-005: no standard weight for NewDrug CT / Phase 3 / fully in-housed / Start-up. Add a row for that scope, or one with work_scope_type empty to cover every scope."],
    ["V-20", "Warning", "A milestone other than 'Inspection' recorded more than once.", "Project PRJ-002 records 'CTA submission' twice. Only 'Inspection' is expected to repeat."],
    ["V-21", "Information", "An 'Inspection' dated on or before the final DB lock.", "Project PRJ-002: 1 inspection on or before the final DB lock is treated as a marker and does not open the final period."],
    ["V-24", "Error", "PersonPeriodWeight.assignment_id not found in Assignment, or two windows of one assignment share a period_start.", "ASG-999: PersonPeriodWeight refers to an assignment that does not exist; its override is silently ignored."],
    ["V-23", "Error", "No RoleFactor row for a (project_type, clinical_phase, work_scope_type, period_name, role_name) that an assignment actually spans - counting the every-scope row.", "No role factor for NewDrug CT / Phase 3 / fully outsourced / Conduct (final) / Data Analyst - assignments covering that period would be calculated at factor 1.00."],
    ["V-25", "Warning", "A project's work_scope_type contradicts its outsourcing_type at one of the two unambiguous ends.", "Project PRJ-012: outsourcing_type says 'Full In-house' but work_scope_type says 'fully outsourced'. The weights follow work_scope_type; check which is right."],
    ["V-26", "Error", "A project carries a project_type that schema 6 retired.", "Project PRJ-003: project_type 'Biosimilar CT' was split in schema 6. Change it to Biosimilar CT (Healthy) or Biosimilar CT (Patient)."],
    ["V-22", "Warning", "Person.capacity_fte is below config.under_allocation_fte.", "PSN-018: capacity 0.50 FTE is below the under-allocation floor of 0.60, so this person can never clear it however fully they are booked. Lower the floor or raise the capacity."],
    ["V-31", "Error", "A project or assignment set to estimation_type = 'manual' has months it covers with no MonthlyEstimate row.", "Project PRJ-019 is set to MANUAL but MonthlyEstimate has no figure for 3 of its month(s): 2028-01, 2028-02, 2028-03. Those months are counted as 0.00."],
    ["V-32", "Error", "A manual PROJECT has a figure for a month in which nobody is assigned to it, so there is nobody to share it out to.", "Project PRJ-019 has a manual figure for 2 month(s) in which nobody is assigned to it. It has NOT been applied - the project would otherwise show a total that none of its people account for."],
]
r = table(ws, r, ["ID", "Severity", "Trigger", "Message shown to the user"],
          rules, [8, 15, 56, 76], wrap_cols=(3, 4))
r = note(ws, r, "Messages name the row and the offending value, and where there is an obvious fix they suggest it. "
                "A validation report that says only 'invalid data' forces the user to hunt, which is how findings "
                "get ignored.")

r = section(ws, r, "The findings report")
rep = [
    ["Grouped by severity, then by sheet.", "Fatals first - if any exist, nothing else matters yet."],
    ["Each finding shows sheet, row number, rule ID and message.", "The row number is the Excel row, so the user can go straight to it."],
    ["A one-line summary banner persists until dismissed.", "e.g. '3 errors, 11 warnings - 2 rows were not loaded'."],
    ["Exportable to .xlsx.", "So a data maintainer can work through it away from the application."],
]
r = table(ws, r, ["Behaviour", "Reason"], rep, [62, 70], wrap_cols=(1, 2))

# ---- 05 Calculation -------------------------------------------------------
ws, r = sheet(wb, "05_Calculation", "Calculation",
              "Pure functions: no DOM, no file access (plan sheet 07, layer 5). "
              "tools/verify_source_workbook.py is the reference implementation.")

r = section(ws, r, "Period derivation - clinical trial, either type   [REQ-CAL-09, REQ-CAL-12, REQ-CAL-13]")
r = code(ws, r, [
    "  protocol = milestone 'Protocol (v1)'          cta  = milestone 'CTA submission'",
    "  siv      = 'First SIV' else 'FPI'             idbl = 'interim DB lock'",
    "  fdbl     = 'final DB lock' else idbl          insp = all 'Inspection' dates",
    "",
    "  if cta is null or fdbl is null:  raise V-16 and derive nothing",
    "",
    "  su_start = max( protocol + 1 day  if protocol else  cta - 1 month , project.start )",
    "  su_end   = siv  if siv and siv >= su_start  else  su_start + 4 months - 1 day",
    "",
    "  later    = [ d in insp where d > fdbl ]            # V-21: earlier ones stay markers",
    "  p7_start = min(later) if later else null",
    "  p7_end   = max( max(later), project.end ) if later else null",
    "",
    "  cof_start = fdbl - 3 months",
    "  cof_end   = p7_start - 1 day  if p7_start  else  max( fdbl, project.end )",
    "",
    "  segments = []",
    "  if su_start > project.start:  segments += ('Before-Start-up', project.start, su_start - 1 day)",
    "  segments += ('Start-up', su_start, su_end)",
    "  if idbl and idbl < fdbl:                            # R-11: the split case",
    "      coi_start = idbl - 3 months",
    "      segments += ('Conduct (interim)',    su_end + 1 day,  coi_start - 1 day)",
    "      segments += ('Close-out (interim)',  coi_start,       idbl)",
    "      cof_start = max( cof_start, idbl + 1 day )",
    "      segments += ('Conduct (final)',      idbl + 1 day,    cof_start - 1 day)",
    "  else:                                               # no interim lock to be interim to",
    "      segments += ('Conduct (final)',      su_end + 1 day,  cof_start - 1 day)",
    "  segments += ('Close-out (final)', cof_start, cof_end)",
    "  if p7_start:  segments += ('After Close-out (final)', p7_start, p7_end)",
    "",
    "  drop every segment whose end < start          # REQ-CAL-12, decision C-11",
    "  number the survivors 1..n as period_seq",
])
r = note(ws, r, "'Conduct (interim)' is emitted ONLY in the split branch, so the name is never a lie: it exists "
                "only where there is an interim DB lock to be interim to. Every other conduct stretch - "
                "including the single one of a project that never has an interim lock - is 'Conduct (final)'. "
                "That is what makes period_name unique within a project and gives ProjectPeriod its natural "
                "key (R-11).")
r += 1
r = note(ws, r, "A project may legitimately end up with 'Conduct (interim)' and no 'Conduct (final)': where the "
                "two locks are less than three months apart, the final stretch is squeezed to nothing and "
                "dropped by the rule above. The names describe position, not a required pair.")
r = note(ws, r, "'month' arithmetic clamps to the end of a short month: 31 Mar minus 1 month is 28 Feb (29 in a leap "
                "year). Test that explicitly - it is the classic off-by-one in this kind of code.")
r += 1
r = note(ws, r, "For 'Others' the derivation does not run at all: periods are read from ProjectPeriod as entered (Q-25, Q-28).")
r += 1

r = section(ws, r, "Monthly load   [REQ-CAL-02, REQ-CAL-05, REQ-CAL-08, REQ-CAL-14]")
r = code(ws, r, [
    "  load(assignment, month) = project_period_weight( project, month )",
    "",
    "                            role_factor( project.type, project.phase, project.scope,",
    "                                         period_name( project, month ), assignment.role )",
    "                          x -------------------------------------------------------------",
    "                            sharers( project, assignment.role, month )",
    "",
    "                          x person_weight( assignment, month )",
    "                          x coverage( assignment, month )",
    "",
    "  project_period_weight : weight of the period containing the FIRST DAY of the month  (decision C-08)",
    "                          no period covers it -> 1.00, and raise V-12",
    "  role_factor           : selected by the SAME period as the weight above, so the two always",
    "                          agree about which period the month is in  (R-10)",
    "                          project.phase is NULL for 'Others'; the lookup matches null to null",
    "                          project.scope falls back to the empty-scope row  (R-12)",
    "                          no row matches -> 1.00, and raise V-23",
    "  sharers               : how many DISTINCT PEOPLE hold that role on that project in that",
    "                          month, counting anyone whose assignment touches it  (R-13)",
    "                          never 0; when nobody else holds the role it is 1",
    "                          config.split_shared_role_fte = 0 -> always 1, the old arithmetic",
    "  person_weight         : PersonPeriodWeight.weight_override if a window covers the",
    "                          first day of the month, else assignment.person_weight",
    "  coverage              : calendar days of the month inside the ASSIGNMENT WINDOW,",
    "                          divided by the days in that month",
    "",
    "  THE ASSIGNMENT WINDOW  [REQ-CAL-15]",
    "      start = assignment.assign_start_date  or  project.start_date",
    "      end   = assignment.assign_end_date    or  project.end_date",
    "",
    "  Both dates are OPTIONAL. Most people are on a project for the whole of it, and asking",
    "  for two dates that repeat the project's means copying the same pair onto every row and",
    "  keeping them in step when the project moves. Fill them in for a PARTIAL involvement.",
    "",
    "  The end date always behaved this way. The start did not, and a blank one made the",
    "  assignment contribute NOTHING - the row on screen, the person apparently unassigned,",
    "  and no finding to say why. That was a defect rather than a rule.",
    "",
    "  ONE FUNCTION, used by the calculation AND by the sharing pre-pass above. The months a",
    "  person is counted IN must not differ from the months they are counted AMONG; two",
    "  copies of this rule would eventually disagree, and the symptom would be a shared role",
    "  that does not add up to one.",
    "",
    "  result is FTE.  hours = FTE x config.fte_hours_per_month",
    "",
    "  WHEN NOBODY HOLDS A ROLE  [REQ-CAL-16]",
    "  role_factor answers 'what does this ROLE cost the project'. If nobody is holding it,",
    "  that cost does not disappear - the work is done by whoever is there. So the factor is",
    "  added to the role named in RoleFactor.absorbed_by, and the project stops looking",
    "  cheaper than it is simply because a post was never filled.",
    "",
    "      effective(role) = factor(role)",
    "                      + SUM factor(x) for every x with absorbed_by = role",
    "                                     that NOBODY holds in this month",
    "",
    "  PER MONTH, so cover ends by itself the month somebody arrives.",
    "  ONE HOP: if the absorbing role is unstaffed too, the work is not passed further along.",
    "  There is nobody to pass it to, and a chain would pile three absent roles onto whoever",
    "  happened to be left. V-29 reports that case rather than inventing an answer.",
    "  THE MAPPING IS DATA. Which role covers for which is a judgment about how a team works,",
    "  and REQ-CAL-06 has said since the first plan that factors live in the workbook.",
    "",
  "  WHY THE FACTOR IS DIVIDED  [REQ-CAL-14]",
    "  role_factor answers 'what does this ROLE cost the project in this period', not 'what",
    "  does each person holding it cost'. Charge every holder the whole factor and a trial run",
    "  by two data managers costs twice a trial run by one - the same work, priced by how many",
    "  people happen to be named against it. The property to preserve is that the PROJECT'S",
    "  total does not move when the same work is shared out differently, and dividing by the",
    "  headcount is what preserves it.",
    "",
    "  COUNTED PER MONTH, not once per assignment. When one of two sharers leaves in June, the",
    "  other must be back to a full share in July - by itself, with nobody editing anything.",
    "  A per-assignment divisor would halve them for the whole project.",
    "",
    "  COUNTED BY PEOPLE, not by rows. Two rows for one person on the same project and role are",
    "  one person doing one job, and must not halve their own load.",
    "",
    "  PERSON WEIGHT APPLIES AFTERWARDS, to the share. Two people at 1.00 and 0.50 on a factor",
    "  of 2.00 carry 1.00 and 0.50, not 1.33 and 0.67: the division is by headcount, which is",
    "  what was asked for, and a person's own weight then says how much of their share they",
    "  actually give it.",
])

r = section(ws, r, "The two weight tables now overlap   [R-10, R-12]")
r = lines(ws, r, [
    "PeriodFTEStandard is keyed on (project_type, clinical_phase, work_scope_type, period_name).",
    "RoleFactor adds role_name to the same key.",
    "The calculation multiplies them, so the pair is mathematically collapsible into one table.",
    "",
    "THE LOOKUP, IN BOTH TABLES, IS TWO STEPS AND NOT ONE:",
    "",
    "    1. the row for this project's own work_scope_type   ->   use it",
    "    2. otherwise the row whose work_scope_type is EMPTY  ->   use that; it means any scope",
    "    3. otherwise nothing                                ->   V-19 or V-23 says so",
    "",
    "An empty scope is not a missing value. It is a row that deliberately declines to distinguish,",
    "and it is what keeps the two tables a size a person can fill by hand. Any program reading these",
    "sheets must do the same, or it will report a missing weight where the application finds one.",
])
r = table(ws, r, ["Table", "Answers", "Change it when"],
          [["PeriodFTEStandard", "How busy is the PROJECT in this period?",
            "The shape of a project changes - e.g. Phase 3 close-out is heavier than assumed. One edit covers every role."],
           ["RoleFactor", "How much of that falls on THIS role?",
            "The split between roles changes - e.g. the analyst starts earlier than assumed. Five edits, one per role."]],
          [26, 52, 76], wrap_cols=(2, 3))
r = note(ws, r, "The separation is a maintenance convention, not something the arithmetic enforces. Raising a "
                "project's Conduct load by editing all five RoleFactor rows produces the right answer today "
                "and double-counts the next time PeriodFTEStandard moves. The application cannot detect "
                "this - both tables are legitimate inputs - so it belongs in the workbook README and in "
                "whatever process maintains the file.")
r += 1
r = note(ws, r, "If the distinction ever stops being maintained in practice, the honest fix is to collapse the "
                "two into one table keyed on all four columns, not to keep a separation that exists only on "
                "paper. That would be a schema change, so it is flagged here rather than done.")
r += 1

r = section(ws, r, "Aggregation   [REQ-CAL-03, REQ-CAL-04, REQ-CAL-07]")
agg = [
    ["project_month[p][m]", "sum of load over assignments on project p", "Overall tab, table A and the stacked graph"],
    ["person_month[s][m]", "sum of load over all assignments held by person s", "Overall tab, table B and the person graph"],
    ["cell[s][p][m]", "the individual load", "Drill-down under a person or project row"],
    ["headcount[p][m]", "count of assignments with load > 0", "Compared with planned_member_count"],
    ["over[s][m]", "true where person_month > config.over_allocation_fte - the same figure for everyone", "Red cell, counted in the summary tiles"],
    ["under_runs[s]", "maximal runs of >= min_months consecutive months where 0 < person_month < config.under_allocation_fte", "Amber across the run, counted once per run"],
]
r = table(ws, r, ["Output", "Definition", "Where it surfaces"], agg, [24, 76, 40], wrap_cols=(2, 3))

r = section(ws, r, "Thresholds   [settled at S2-01 and S2-05]")
r = lines(ws, r, [
    "Both thresholds are ABSOLUTE. They are not scaled by capacity_fte.",
])
r = code(ws, r, [
    "  over  when  person_month > config.over_allocation_fte    (1.50)",
    "  under when  person_month < config.under_allocation_fte   (0.60)",
])
r = lines(ws, r, [
    "The v0.3 draft proposed scaling both by capacity, because a part-timer at 0.60 could never reach an",
    "absolute 0.80 floor and so was flagged permanently. The review kept the thresholds absolute but moved",
    "the floor to 0.60, which resolves the same defect a different way - every capacity in the data can now",
    "clear it:",
])
r = table(ws, r, ["capacity_fte", "utilisation needed to clear the 0.60 floor"],
          [["1.00", "60%"], ["0.80", "75%"], ["0.60", "100%"]], [16, 46])
r = note(ws, r, "The risk returns only for a capacity BELOW the floor, which V-22 warns about on import. "
                "Nothing else in the calculation depends on capacity_fte, so it is now used for display "
                "and for that one check rather than in the arithmetic.")
r += 1

r = section(ws, r, "Under-allocation runs")
r = code(ws, r, [
    "  walk the person's months in order across the horizon",
    "  a month counts toward a run when  0 < person_month < under_threshold",
    "  a month at exactly 0 BREAKS the run          # no assignments at all is not under-allocation",
    "  a run of >= config.under_allocation_min_months is reported once, at its first month,",
    "      carrying its length",
])
r = note(ws, r, "Reporting the run once rather than flagging each month is what makes the signal readable: three "
                "separate amber cells look like three problems.")
r += 1

r = section(ws, r, "Acceptance test")
ex = [
    ["Assignment", "PSN-001 on PRJ-001 (Clinical Trial), role 'Lead data manager'", ""],
    ["person_weight", "0.40", ""],
    ["Window", "2026-03-10 to 2026-12-31", "Joins part-way through March"],
    ["Periods", "Start-up to 2026-06-30 at weight 1.50; Conduct (final) from 2026-07-01 at 1.00", "From the trial's clinical phase"],
    ["role_factor", "1.20", "Illustrative; the real value is data"],
    ["March 2026", "1.50 x 1.20 x 0.40 x (22/31) = 0.511 FTE", "81.8 hours"],
    ["April 2026", "1.50 x 1.20 x 0.40 x 1.0000 = 0.720 FTE", "115.2 hours"],
    ["July 2026", "1.00 x 1.20 x 0.40 x 1.0000 = 0.480 FTE", "76.8 hours"],
]
r = table(ws, r, ["Element", "Value", "Note"], ex, [22, 62, 44], wrap_cols=(2, 3))
r = note(ws, r, "Plus the whole dummy dataset: running tools/verify_source_workbook.py against "
                "PRAP_SourceData_Dummy_v1.8.xlsx must give no errors and no warnings, across 62 projects, "
                "20 people, 289 assignments and 308 periods spanning 73 months. Every period set must be "
                "contiguous, all 50 trials must carry 'Conduct (final)', 30 must also carry "
                "'Conduct (interim)', and 12 must carry the final inspection period. No project may carry "
                "a repeated period name. Those figures are the regression baseline for Step 4.")

# ---- 06 UI ----------------------------------------------------------------
ws, r = sheet(wb, "06_UI_Spec", "User interface",
              "Behaviour and content. Layout and typography are fixed at Step 3. Colour is fixed there too, "
              "with one exception now settled here: on the timeline, colour carries the period name, so it "
              "is meaning rather than styling.")

r = section(ws, r, "Global")
glob = [
    ["Header", "Application name, application version, expected schema version, loaded file name, and the load time WITH ITS TIME ZONE - offset and abbreviation both, e.g. '2026-08-01 09:14 (GMT+9, KST)'. A bare timestamp is ambiguous to a reader in another country.", "REQ-VC-02, REQ-IMP-05"],
    ["Tab bar", "Overall | Source data (project) | Source data (person) | General assumptions. Tab state survives a re-import.", "REQ-DSH-01..04, REQ-DSH-11"],
    ["Findings banner", "Appears when the last load or edit produced findings; opens the full report on click.", "REQ-IMP-02"],
    ["Unsaved-edit counter", "Always visible once any edit exists, and states the validation standing of those edits - 'n unsaved edits, all n pass validation'. Warns before unload or a new import. An edit that fails a rule is rejected at entry and never enters the buffer, so everything counted here is exportable.", "REQ-IMP-08, REQ-IMP-09"],
    ["Empty state", "With nothing loaded, every tab shows the same load panel plus a template download link.", "REQ-IMP-03"],
]
r = table(ws, r, ["Component", "Behaviour", "REQ-ID"], glob, [24, 90, 20], wrap_cols=(2,))

r = section(ws, r, "Tab 1 - Overall")
ov = [
    ["Horizon control", "From/to month. Defaults to 24 months from the current month. One control expands it to span every project's dates.", "REQ-CAL-01, REQ-DSH-07"],
    ["Filters", "project type (NewDrug CT / Biosimilar CT / Others), CLINICAL PHASE, project, person, role, department. Multi-select, combined with AND. GLOBAL: one setting drives every tab, not just the Overall tab. Clinical phase sits immediately right of project type, so the two type-ish controls read as a pair.", "REQ-DSH-05"],
    ["Unit toggle", "FTE or hours, seeded from config.capacity_unit. NOT in the filter bar: it changes how every figure is written, not which figures are shown, so it is a setting and lives on the assumptions tab.", "REQ-CAL-08, REQ-DSH-11"],
    ["Table A - by project", "Rows projects, columns months, cells FTE. Row and column totals. A project row expands to its people.", "REQ-DSH-01"],
    ["Table B - by person", "Rows people, columns months, cells FTE summed across projects. Over-allocated cells red, under-allocation runs amber. A person row expands to their projects.", "REQ-DSH-01, REQ-DSH-08"],
    ["Graph 1", "Stacked bar: total monthly demand, ONE BAND PER PROJECT, ordered by total resource with the largest on the baseline. 'Others' projects are grey; trials take the extended colour set. NO LEGEND - a list of 62 entries cannot be matched against the chart. Identity comes from the hover pop-up, which carries project name and type, that month's FTE and its hour equivalent, its share of the month, the headcount, every person on the project that month with their role, and - under a rule - THE MONTH'S TOTAL ACROSS EVERY PROJECT IN VIEW. The total is what a band on its own cannot give: 4.02 FTE means nothing until you know whether the month came to five or to fifty. It is the same figure Graph 2 states for that month, summed along the other axis, so the two pop-ups are held to agreeing in words as well as in pixels.", "REQ-DSH-02"],
    ["Graph 2", "Monthly FTE per person, with reference lines at the two thresholds - one pair of lines, since both are absolute. Above the bar budget it shows a ranked subset with the rest rolled into one 'others' band, and says which it is showing.", "REQ-DSH-02, REQ-DSH-08, REQ-DSH-09"],
    ["Graph 3", "Timeline per project - the FIRST panel on the tab, above the summary tiles. Each row carries the project name with its start, end and length beneath. Bands are coloured BY PERIOD NAME (see the colour rule below), with the period weight as a lightness step inside each hue. Milestones are inverted triangles in a lane above the bands; 'Inspection' takes the same marker as every other milestone. The hover pop-up gives the period, its dates, its weight and the FTE per month the project draws across it.", "REQ-DSH-02, REQ-PRJ-05, REQ-DSH-10"],
    ["Summary tiles", "Active projects; people assigned; total FTE in the horizon; over-allocated person-months; under-allocation runs.", "REQ-DSH-08"],
    ["Reset filters", "Clears every filter and restores the default 24-month horizon in one action.", "REQ-DSH-05"],
    ["Scroll regions", "Every chart and table sits in its own scroll region - horizontal for wide content, and a bounded height with vertical scroll for tall content. Wide or long content scrolls INSIDE its panel; the page body never scrolls sideways, and a long sub-table never pushes the panels below it down the page. A scrolled table keeps its header row visible.", "REQ-NFR-02"],
    ["Row expansion - project", "Clicking a project name reveals one row per person and role on it, each with its own monthly figures. Clicking again collapses.", "REQ-DSH-01"],
    ["Row expansion - person", "Clicking a person name reveals one row per project and role they hold, ordered NewDrug CT, Biosimilar CT, Others, then earliest project first.", "REQ-DSH-01"],
    ["Type and phase pills", "Project type and clinical phase shown as labelled pills. The text carries the meaning; the colour only speeds recognition.", "REQ-PRJ-01, REQ-PRJ-09"],
    ["Repeated period labels", "Retained as a guard only. Since R-11 no period name repeats in a project, so the numbering never fires and a name alone identifies a period on screen. V-18 rejects a repeat on import; this keeps the renderer honest if one ever reaches it another way.", "REQ-DSH-10"],
    ["Row virtualisation", "Both tables render only the rows inside the viewport plus a small overscan, with row height fixed so the scrollbar stays truthful. Sorting, filtering and totals run over the whole model, never over the rendered slice.", "REQ-DSH-09, REQ-NFR-03"],
]
r = table(ws, r, ["Component", "Behaviour", "REQ-ID"], ov, [24, 90, 20], wrap_cols=(2,))
r = note(ws, r, "Both tables are the same numbers aggregated differently, so they must always reconcile: the grand "
                "total of table A equals that of table B. Worth asserting in code, not just hoping.")

r = section(ws, r, "Timeline colour rule   [O-10, supersedes design decision D-06]")
r = lines(ws, r, [
    "Hue carries the PERIOD NAME. Weight carries a lightness step inside that hue, over a deliberately",
    "narrow range so it never competes with the hue for the reader's attention.",
])
colour = [
    ["Before-Start-up", "grey, light", "Not started. The two greys bracket the active work."],
    ["Start-up", "red, shifted toward orange", "Requested as red. Shifted because red beside green is the pair red-green colour blindness collapses, and REQ-DSH-08's principle - never colour alone - applies here too."],
    ["Conduct (interim)", "green, light", "The stretch before an interim DB lock. Exists only where there is one."],
    ["Conduct (final)", "green, deep", "The stretch that runs to the final lock - one green family, two steps, because they are the same phase of work at different points."],
    ["Close-out (interim)", "orange, light", "Split from the final close-out: a trial with an interim DB lock shows both in one row, and a single shared orange read as one interrupted band."],
    ["Close-out (final)", "orange, deep", "As above."],
    ["After Close-out (final)", "grey, dark", "Both greys are chosen to stay legible on the dark surface as well as the light one."],
    ["Planning / Develop / Close", "grey / green / orange", "The 'Others' period set, mapped to the same semantics."],
]
r = table(ws, r, ["Period", "Hue", "Why"], colour, [26, 26, 82], wrap_cols=(3,))
r = lines(ws, r, [
    "Colour is never the only carrier. A band wide enough to hold its name shows it as text, the pop-up",
    "names the period outright, and a repeated name carries its number (REQ-DSH-10).",
])
r = note(ws, r, "This overturns D-06, which shaded bands by weight on the grounds that weight is what drives "
                "the simulation. The reviewer accepted D-06 and then asked for the opposite in O-10. O-10 is "
                "the more specific instruction, so it governs - and D-06's point is preserved by keeping "
                "weight as the lightness step and exact in the pop-up.")
r += 1

r = section(ws, r, "Rendering at the target volume   [S2-06, REQ-DSH-09, REQ-NFR-03]")
r = lines(ws, r, [
    "REQ-NFR-03 now reads 100 projects and 1,000 people over a 60-month horizon. That is a different rendering",
    "problem from the 62 projects the prototype was drawn against, and it changes two components from",
    "'could be optimised later' to 'cannot be built the obvious way':",
])
scale = [
    ["Person table", "1,000 rows x 60 months = 60,000 cells, plus the same again on expansion.",
     "Virtualise: render the visible window plus overscan. A 60,000-node table is seconds of layout on every filter change."],
    ["Person chart", "1,000 bars across a 1,200px panel is 1.2px each - narrower than the gap between them.",
     "Aggregate. Show a ranked subset (default: the 20 most loaded) with the remainder as one band, and name what is shown."],
    ["Project chart", "100 stacked bands per month, against 8 hues that can be told apart.",
     "Unchanged from decision D-11: identity comes from legend order and tooltip, not hue. The band order is already by total resource."],
    ["Calculation", "100 x 1,000 x 60 is bounded by assignments, order 8,000, so about 480,000 month-rows worst case.",
     "Compute once per import or edit into typed arrays keyed by month index; never recompute inside a render."],
]
r = table(ws, r, ["Component", "What the volume implies", "What the specification requires"],
          scale, [20, 54, 60], wrap_cols=(2, 3))
r = note(ws, r, "The working figure in REQ-NFR-03 is an upper bound to design against, not the expected daily "
                "dataset. The point is that nothing in the design may assume the small case - a table that is "
                "fast at 20 rows and unusable at 1,000 fails the requirement.")

r = section(ws, r, "Tab 2 - Source data (project)")
t2 = [
    ["Project table", "All 23 Project columns, sortable and filterable, total_period_months recomputed. Editable.", "REQ-DSH-03, REQ-IMP-07"],
    ["Milestone sub-table", "Milestones of the selected project in date order. 'Inspection' may appear several times.", "REQ-PRJ-05, REQ-PRJ-13"],
    ["Period sub-table", "Derived periods with seq, dates and weight, in seq order. Names are unique within a project since R-11, so project_id + period_name identifies a row. Shows whether each date was derived or hand-set.", "REQ-PRJ-06, REQ-CAL-09, REQ-DSH-10"],
    ["Recompute periods", "Re-derives from current milestones, warning that hand-set dates will be replaced.", "decision C-10"],
    ["Utilisation graph", "The selected project's monthly resource across the horizon, as bars, with THREE reference lines: 2x and 0.5x the average an ACTIVE project-month draws across the portfolio, and the project's own average over its full life. Sits directly under the project table, mirroring the person tab's strip. Months where the project draws nothing are excluded from the portfolio average - averaging them in would drag the norm toward zero and make every running project look heavy.", "REQ-DSH-12"],
    ["Export", "Visible table to .xlsx.", "REQ-DSH-06"],
]
r = table(ws, r, ["Component", "Behaviour", "REQ-ID"], t2, [24, 90, 20], wrap_cols=(2,))
r = note(ws, r, "The reference lines are RELATIVE, and that is the whole design. A person has a capacity, "
                "so an absolute ceiling and floor mean something. A project does not: 3 FTE a month is "
                "heavy for a Phase 1 and light for a Phase 3 close-out. Comparing a project to the "
                "portfolio and to its own history is the only framing that carries meaning across "
                "62 projects of different types and phases. The lines are context, not pass or fail, and "
                "the caption must say so - a dashed line above a bar reads as a limit unless it is "
                "labelled otherwise.")

r = section(ws, r, "Tab 3 - Source data (person)")
t3 = [
    ["Person table", "All 12 Person columns, sortable and filterable. Editable.", "REQ-DSH-04, REQ-IMP-07"],
    ["Assignment sub-table", "The selected person's assignments: project, role, dates, person_weight.", "REQ-PSN-02, REQ-PSN-03"],
    ["Override sub-table", "PersonPeriodWeight windows for the selected assignment.", "REQ-PSN-05"],
    ["Utilisation strip", "The person's monthly FTE across the horizon with both absolute thresholds marked, and capacity_fte shown alongside for context.", "REQ-DSH-08"],
    ["Export", "Visible table to .xlsx.", "REQ-DSH-06"],
]
r = table(ws, r, ["Component", "Behaviour", "REQ-ID"], t3, [24, 90, 20], wrap_cols=(2,))

r = section(ws, r, "Tab 4 - General assumptions   [REQ-DSH-11, added at the component-list review]")
r = lines(ws, r, [
    "Every figure on the Overall tab starts from the standard monthly FTE for the project's type, phase and",
    "work scope, adjusted by the project's own period weight, and is then shared out among the people on it",
    "by the role factors and their own weights. Those standards lived only in the workbook, so a reader who",
    "wanted to know why a number was what it is had to leave the application. This tab is the fix.",
])
t4 = [
    ["Role factors", "RoleFactor as TWO matrices - clinical trials (type + phase + role down the side, the six periods across) and 'Others' (role down the side, the three periods across). 249 rows flat is unreadable, and reading a role ACROSS the periods is the whole point of keying it that way. Both sit in a bounded scroll region.", "REQ-DSH-11"],
    ["Standard period FTE for project types", "PeriodFTEStandard as a MATRIX - project type and clinical phase down the side, period name across, cells shaded on the same ramp as the Overall tables. It is a standard, and a standard is read across; 48 flat rows read as a list. 'Others' projects are absent by design - their weights are hand-entered per project (Q-28).", "REQ-DSH-11"],
    ["Configuration", "Config in full, with each parameter's note. The display-unit control sits here, with a line saying why it is not in the filter bar.", "REQ-DSH-11, REQ-CAL-08"],
    ["Value lists", "Lists, one row per list, showing every accepted value and the count. Answers 'what may I type here' without opening the workbook. A value outside its list is kept and reported by V-11, never dropped. READ-ONLY and with no insert control: a value added here with nothing referring to it is noise.", "REQ-DSH-11"],
    ["Editing", "The role-factor and config tables are editable and insertable like any other table, and an edit here recalculates everything - these are the multipliers.", "REQ-IMP-07, REQ-IMP-11"],
]
r = table(ws, r, ["Component", "Behaviour", "REQ-ID"], t4, [24, 90, 20], wrap_cols=(2,))
r = note(ws, r, "This tab is read-mostly but not read-only. A role factor is exactly the kind of figure a "
                "planner wants to try a different value for, and locking it would send them back to the "
                "workbook - which is what the tab exists to avoid.")

# ---- 07 Editing and IO ----------------------------------------------------
ws, r = sheet(wb, "07_Editing_IO", "Editing, import and export")

r = section(ws, r, "Import   [REQ-IMP-01, REQ-IMP-02]")
r = code(ws, r, [
    "  user picks a file, or drops one on the page",
    "  parse -> validate -> build the model -> derive periods -> calculate -> render",
    "  a fatal finding stops after validate and shows the report",
    "  a successful load replaces the model entirely; unsaved edits are warned about first",
])

r = section(ws, r, "The edit buffer   [REQ-IMP-07, REQ-IMP-09]")
edit = [
    ["Every field is editable, identifiers included.", "Q-20. Consistency is protected by rule, not by locking fields."],
    ["An edit is validated with the same rules as an import, at the moment of entry.", "REQ-IMP-09. A rejected edit never reaches the model, so the model is always valid."],
    ["An edited cell is marked; the count of unsaved edits is always visible.", "REQ-IMP-08."],
    ["Editing a milestone date re-derives that project's periods, unless its period dates were hand-set.", "Decision C-10. This is the frequent-timeline-change case the tool exists for."],
    ["Editing a weight, date or assignment recalculates immediately.", "The dashboard must never show numbers that disagree with the data on screen."],
]
r = table(ws, r, ["Rule", "Basis"], edit, [76, 56], wrap_cols=(1, 2))

r = section(ws, r, "Inserting a row   [REQ-IMP-11, added at the component-list review]")
r = code(ws, r, [
    "  every editable table shows an insert control on EVERY row, not once per table",
    "  the control LEADS the row: at the end of a 23-column table it scrolls out of view",
    "  on press:",
    "      create a blank row of that sheet's shape",
    "      position it IMMEDIATELY BELOW the row the control was pressed on",
    "      pre-fill the parent key where the table is a child of a selection",
    "          (a period inserted under project PRJ-003 gets project_id = PRJ-003)",
    "      leave every other field empty and focus the first editable cell",
    "      mark the row unsaved; it joins the edit count like any other change",
    "  validation runs on entry, exactly as for an edit (REQ-IMP-09):",
    "      a blank row is INCOMPLETE, not invalid - it is not reported until a field is filled",
    "      a row still missing its identifier when export is pressed blocks the export, naming it",
])
r = lines(ws, r, [
    "Position matters and is worth stating. Appending to the bottom of a 62-row table means the user has to",
    "find the new row and then move it; inserting where they are looking means they can just type. The stored",
    "order follows the display order, so the export carries the row where the user put it.",
])
r = note(ws, r, "Deleting stays asymmetric with inserting, as it already is with editing: a row can be created "
                "freely because an empty row references nothing, while a row that is referenced cannot be "
                "deleted at all (V-17). Creation is cheap to undo; deletion is not.")
r += 1

r = section(ws, r, "Cascading identifier edits   [REQ-IMP-10, V-17]")
r = code(ws, r, [
    "  on changing an identifier:",
    "      count the rows referencing the old value",
    "      if 0 : apply",
    "      else : confirm ('PRJ-003 is referenced by 14 rows. Update them all?')",
    "             on confirm, rewrite every reference in one atomic step",
    "",
    "  on deleting a row:",
    "      if anything still references it -> refuse, naming what does (V-17)",
    "      never cascade a delete: losing 14 assignments to one keystroke is unrecoverable",
])
r = note(ws, r, "Cascading an edit and refusing a delete are deliberately asymmetric. A cascaded rename is visible and "
                "reversible; a cascaded delete is neither.")

r = section(ws, r, "Export   [REQ-IMP-04, REQ-IMP-07]")
exp = [
    ["Writes all 11 sheets in template order, with the template's headers and formats.", "So the export re-imports without edits."],
    ["Includes every edit made on screen.", "REQ-IMP-07 - this is the point of the feature."],
    ["Derived columns are written as values, not formulas.", "A formula referencing a row that moved would be wrong; the importer recomputes them anyway."],
    ["Filename defaults to the loaded name with a date suffix.", "Never silently overwrite the source of record."],
    ["Clears the unsaved-edit state on success.", "REQ-IMP-08."],
]
r = table(ws, r, ["Behaviour", "Reason"], exp, [76, 56], wrap_cols=(1, 2))

# ---- 08 Versioning --------------------------------------------------------
ws, r = sheet(wb, "08_Versioning", "Versioning and compatibility")
ver = [
    ["Application version", "Constant in the HTML, shown in the header and footer.", "REQ-VC-02"],
    ["Expected schema version", "Constant in the HTML. Currently 10.", "REQ-VC-02"],
    ["Check on load", "Compare Config.schema_version with the expected value.", "REQ-VC-03"],
    ["Equal", "Proceed silently.", "REQ-VC-03"],
    ["File older", "Proceed; warn that columns added since may be missing.", "REQ-VC-03"],
    ["File newer", "Proceed; warn that columns may be ignored, and suggest updating the application.", "REQ-VC-03"],
    ["Missing", "Treat as version 1 and warn.", "REQ-VC-03"],
]
r = table(ws, r, ["Item", "Behaviour", "REQ-ID"], ver, [26, 88, 18], wrap_cols=(2,))
r = note(ws, r, "The check never blocks a load. A planner with a slightly stale file still needs an answer today; "
                "a warning gives them one with the caveat attached.")

# ---- 09 Traceability ------------------------------------------------------
ws, r = sheet(wb, "09_Traceability", "Requirement traceability",
              f"Read directly from {PLAN} sheet 03, so a requirement cannot be lost between the two documents.")

plan_wb = load_workbook(ROOT / "docs" / PLAN)
plan_ws = plan_wb["03_Requirements"]
SECTION_BY_PREFIX = {
    "REQ-OUT": "02_Scope, 07_Editing_IO",
    "REQ-PRJ": "03_Data_Schema",
    "REQ-PSN": "03_Data_Schema",
    "REQ-CAL": "05_Calculation",
    "REQ-DSH": "06_UI_Spec",
    "REQ-IMP": "07_Editing_IO",
    "REQ-VC": "08_Versioning",
    "REQ-NFR": "02_Scope, 07_Editing_IO",
}
trace, counts = [], {}
for rr in range(5, plan_ws.max_row + 1):
    rid = plan_ws.cell(rr, 1).value
    if not rid or not str(rid).startswith("REQ"):
        continue
    rid = str(rid).strip()
    prefix = "-".join(rid.split("-")[:2])
    sec = SECTION_BY_PREFIX.get(prefix, "-")
    counts[sec] = counts.get(sec, 0) + 1
    trace.append([rid, str(plan_ws.cell(rr, 3).value)[:150], plan_ws.cell(rr, 4).value, sec])
r = table(ws, r, ["REQ-ID", "Requirement (from the plan)", "Priority", "Specified in"],
          trace, [15, 96, 10, 30], wrap_cols=(2,))
r = note(ws, r, f"{len(trace)} requirements, all mapped. Regenerated from the plan on every build, so a requirement "
                f"added there appears here automatically rather than being remembered.")

# ---- 10 Open points -------------------------------------------------------
ws, r = sheet(wb, "10_Open_Points", "Open points - all answered at the v0.3 review",
              "The six points raised in v0.3, the answer given on 2026-08-01, and what changed in this "
              "document as a result. No point remains open.")
op = [
    ["S2-01", "Calculation",
     "The under-allocation threshold is an absolute 0.80 FTE while Person.capacity_fte allows a part-timer at 0.60, who can therefore never clear it and is flagged permanently. Should both thresholds be relative to capacity_fte?",
     "No, both thresholds (alerting over work-load, and under work-load) not to be relative to capacity_fte.",
     "Answered - draft overturned. The relative-threshold proposal is withdrawn; sheet 05 now specifies both thresholds as absolute, and capacity_fte no longer appears in the arithmetic. The part-timer defect this point raised is instead resolved by S2-05, and V-22 (new) warns on import about the one case that remains - a capacity below the floor. Plan change R-07.",
     "Closed"],
    ["S2-02", "Data model",
     "period_name and role_name cannot carry an Excel dropdown, because the valid list depends on the project's type.",
     "No need to carry out the dropdown list to period_name and role_name.",
     "Answered - draft confirmed. No dependent dropdowns are built. V-15 and V-03 remain the enforcement, at import and at edit. No change to this document.",
     "Closed"],
    ["S2-03", "Data model",
     "Adding a value to a list means inserting a row inside that list's block on the Lists sheet, since dropdowns bind to a contiguous range.",
     "Keep current setting.",
     "Answered - draft confirmed. The Lists sheet keeps its long format with contiguous blocks, documented in the template README. No change to this document.",
     "Closed"],
    ["S2-04", "UI",
     "Should the two 'Conduct' stretches of one project be distinguishable on screen, or shown as one Conduct total?",
     "The two 'Conduct' stretches needs to be distinguishable on screen.",
     "Answered - draft extended. They were already separate bands on the timeline but summed in the tables, which is not 'distinguishable'. Sheet 06 now numbers any repeated period name wherever it is named - 'Conduct (1)', 'Conduct (2)' - in bands, tooltips, the period sub-table and exports. New REQ-DSH-10; plan change R-08.",
     "Closed"],
    ["S2-05", "Calculation",
     "A month with zero load breaks an under-allocation run rather than continuing it - somebody with no assignments at all is not 'under-allocated', they are unassigned.",
     "The person who has under 0.6 FTE should be considered an under-allocated person.",
     "Answered - the zero-breaks-the-run rule stands as drafted, and the floor moves from 0.80 to 0.60 FTE in Config and in the template and dummy workbooks. At 0.60 every capacity in the data can clear the floor (1.00 needs 60% utilisation, 0.80 needs 75%, 0.60 needs 100%), which is what settles S2-01. Plan change R-07.",
     "Closed"],
    ["S2-06", "Non-functional",
     "The dummy dataset now holds 62 projects and 289 assignments, above REQ-NFR-03's headroom figure of 50 projects and 500 assignments. The requirement's working-volume figure (20 projects, 30 people) is well below what the dataset represents.",
     "Change the requirement: 100 projects, 1,000 people.",
     "Answered - requirement re-baselined, and raised from Should to Must. REQ-NFR-03 now reads 100 projects and 1,000 people over 60 months (order 8,000 assignments). At that volume the person table is 60,000 cells and the person chart 1,000 bars, so virtualisation and chart aggregation stop being optimisations - see the new section on sheet 06 and REQ-DSH-09. Plan change R-06.",
     "Closed"],
]
r = table(ws, r, ["ID", "Topic", "Question as raised in v0.3", "Answer (Dan, 2026-08-01)", "What changed in v0.4", "State"],
          op, [8, 14, 56, 42, 62, 9], wrap_cols=(3, 4, 5))
r = note(ws, r, "Two of the six were confirmed as drafted, three changed a rule the application reads as data, and "
                "one (S2-06) changed what the design has to survive. Only the last has architectural weight, which "
                "is why it produced a requirement rather than a configuration value.")
r += 1
r = note(ws, r, "Nothing in this document is now waiting on an answer. What it is waiting on is signature of plan "
                "v1.5, which carries the three changes these answers produced.")

wb.save(OUT)
print(f"Written: {OUT}  ({len(trace)} requirements traced)")
