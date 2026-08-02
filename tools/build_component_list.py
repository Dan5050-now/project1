"""Generate the Step 3 UI component list - the thing being reviewed alongside the prototype.

From v0.4 this is a disposition document as well as a review form: it carries the
reviewer's own words against each component, what was done about them, and what is
left. The review trail lives in the deliverable rather than in a chat log.

    python tools/build_component_list.py

Output: docs/PRAP_UI_Component_List_v0.8.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

VERSION = "1.0"
DATE = "2026-08-01"
PROTOTYPE = "app/PRAP_Prototype_v0.8.html"
OUT = Path(__file__).resolve().parents[1] / "docs" / f"PRAP_UI_Component_List_v{VERSION}.xlsx"

FONT = "Arial"
NAVY = "1F3864"
TITLE_F = Font(name=FONT, size=16, bold=True, color=NAVY)
H1_F = Font(name=FONT, size=12, bold=True, color=NAVY)
HDR_F = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_F = Font(name=FONT, size=10)
BOLD_F = Font(name=FONT, size=10, bold=True)
NOTE_F = Font(name=FONT, size=9, italic=True, color="808080")
HDR_FILL = PatternFill("solid", fgColor="2F5597")
BAND = PatternFill("solid", fgColor="F2F5FB")
INPUT = PatternFill("solid", fgColor="FFFF00")
NEWF = PatternFill("solid", fgColor="E2F0D9")      # added by this review
CHGF = PatternFill("solid", fgColor="FFF2CC")      # changed by this review
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
WRAPC = Alignment(vertical="top", wrap_text=True, horizontal="center")


def table(ws, r0, headers, rows, widths, wrap_cols=(), yellow_col=None, mark_col=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(r0, i, h)
        c.font, c.fill, c.border, c.alignment = HDR_F, HDR_FILL, BOX, WRAPC
    ws.row_dimensions[r0].height = 28
    for r, data in enumerate(rows, r0 + 1):
        mark = ""
        if mark_col is not None:
            v = str(data[mark_col - 1] or "")
            for tag, fill in (("[NEW]", NEWF), ("[CHANGED]", CHGF)):
                if v.startswith(tag):
                    mark, data = fill, list(data)
                    data[mark_col - 1] = v.replace(tag, "", 1).lstrip()
                    break
        for i, v in enumerate(data, 1):
            c = ws.cell(r, i, v)
            c.font, c.border = BODY_F, BOX
            c.alignment = WRAP if i in wrap_cols else Alignment(vertical="top")
            if yellow_col == i:
                c.fill = INPUT
            elif mark:
                c.fill = mark
            elif (r - r0) % 2 == 0:
                c.fill = BAND
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return r0 + len(rows) + 2


wb = Workbook()
wb.remove(wb.active)

# ---- cover ---------------------------------------------------------------
ws = wb.create_sheet("00_Cover")
ws.sheet_view.showGridLines = False
ws["A1"] = "PRAP — UI component list"
ws["A1"].font = TITLE_F
ws["A2"] = "Step 3, task 3.1 — APPROVED by Dan, 2026-08-02. Step 3 gate closed."
ws["A2"].font = NOTE_F
meta = [("Version", f"v{VERSION}"), ("Date", DATE), ("Author", "Claude Code"),
        ("Reviewer", "Dan — v0.3, v0.4 and v0.5 reviewed 2026-08-01 to 08-02"),
        ("Prototype", PROTOTYPE),
        ("Governing plan", "PRAP_Development_Plan_v2.0.xlsx (APPROVED BASELINE)"),
        ("Specification", "PRAP_Programming_Specification_v1.0.xlsx (APPROVED)")]
r = 4
for k, v in meta:
    ws.cell(r, 1, k).font = BOLD_F
    ws.cell(r, 2, v).font = BODY_F
    r += 1
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 92

r += 1
ws.cell(r, 1, "What happened to your v0.3 review").font = H1_F
r += 1
for line in [
    "All 15 design decisions accepted. Of the 38 components, 29 marked Keep and 9 marked Change.",
    "Every one of the 9 is applied in the prototype named above - none is deferred, none is partial.",
    "",
    "Sheet 01 carries your decision and your comment verbatim against each component, with a column",
    "saying what was actually done. Six new components appeared as a result of your changes; they are",
    "tinted green. The nine you changed are tinted amber.",
    "",
    "Sheet 02 records the decisions. One needs your attention: you accepted D-06 (the timeline shades",
    "bands by weight) and separately asked in O-10 for the timeline to be coloured by period name.",
    "Those cannot both hold. O-10 is the more specific instruction and the later intent, so it wins -",
    "D-06 is marked SUPERSEDED and weight now rides as a lightness step inside each period's hue,",
    "so nothing that drove the simulation stopped being visible. Overturn that on sheet 02 if I have",
    "read you backwards.",
    "",
    "Sheet 04 lists what your review changed outside this document: two new requirements, two",
    "reworded, and the plan and specification versions that carry them.",
    "",
    "ROUND 3 (this issue) applied three further items you raised, listed in full on sheet 05:",
    "the role factor is now keyed on role AND project type AND clinical phase AND period; every table",
    "on the three data tabs scrolls inside its own panel; and the insert buttons missing from the",
    "assumptions tab are fixed. You were right about the last one, and it was worse than reported -",
    "the role-factor table had an 'insert' HEADER with no cells beneath it, so that table was a column",
    "out of alignment, and the value-lists table had buttons it should never have had.",
    "",
    "ROUND 4 (this issue) is on sheet 06. You named the two conduct stretches apart -",
    "'Conduct (interim)' and 'Conduct (final)' - which makes period_name unique in a project and",
    "gives ProjectPeriod a natural key again. That is the alternative D-15 offered and was not taken",
    "at the time, so D-15 is now SUPERSEDED. It also retires the need for the display numbering that",
    "REQ-DSH-10 introduced: the requirement is satisfied by the data model instead.",
    "",
    "ROUND 6 (this issue) is on sheet 08. You confirmed all 13 changed components and all 7 changed",
    "decisions as OK - nothing was marked Rework - and added two requests, both applied: the DB lock",
    "milestones are emphasised on the timeline, and the project tab gains a utilisation graph with",
    "relative reference lines. The second needed a new requirement, REQ-DSH-12, for the reason you",
    "gave: a project has no static threshold, so the person tab's model does not transfer.",
    "",
    "Rows already confirmed carry 'OK (v0.7)' in the Confirm column, so a blank cell means an item",
    "that is new this round rather than one nobody has looked at.",
    "",
    "APPROVED by Dan on 2026-08-02, together with development plan v2.0 and programming",
    "specification v1.0. Step 3 is closed and Step 4 - code generation - is authorised.",
    "",
    "The two judgement calls flagged on sheet 08 were approved as made: the '...cut-off' milestones",
    "stay ordinary, and the portfolio average covers active project-months only. Both are recorded",
    "here so a later reader knows they were decided rather than defaulted.",
]:
    ws.cell(r, 1, line).font = BODY_F
    r += 1

# ---- components ----------------------------------------------------------
ws = wb.create_sheet("01_Components")
ws.sheet_view.showGridLines = False
ws["A1"] = "Component list — your decisions and what was done"
ws["A1"].font = TITLE_F
SUBTITLE_CELL = ws["A2"]          # filled once the component list below is defined
ws["A2"].font = NOTE_F
ws.freeze_panes = "A5"

# Confirmed OK at the v0.7 review. Recorded here so the trail lives in the document:
# a blank column would read as "not yet looked at" a round later.
CONFIRMED = {"G-02", "G-06", "G-07", "O-03", "O-04", "O-06", "O-10", "P-01", "S-01",
             "X-01", "X-03", "A-02", "A-04",
             "D-06", "D-11", "D-15", "D-16", "D-17", "D-18", "D-19"}

K = "Keep"
C = [
    # id, area, component, what it does, REQ-IDs, your decision, your comment, what was done
    ("G-01", "Global", "Header", "Application name, version, expected schema version.",
     "REQ-VC-02", K, "", "Unchanged."),
    ("G-02", "Global", "[CHANGED] Loaded-file line",
     "Which workbook is loaded, and when it was read - now with the time zone.",
     "REQ-IMP-05", "Change",
     "Add GMT timezone for loaded date & time information. E.g. loaded 2026-08-01 09:14 (GMT+G6, KST)",
     "Done. The stamp now reads 'loaded 2026-08-01 09:14 (GMT+9, KST)'. The offset and the zone "
     "abbreviation both show, because either alone is ambiguous to a reader in another country. Taken "
     "as GMT+9 / KST - your note says 'GMT+G6', which I have read as a typing slip for GMT+9 given "
     "KST beside it. Correct me if not."),
    ("G-03", "Global", "Load workbook", "File picker / drag-and-drop. Warns first if edits are unsaved.",
     "REQ-IMP-01, REQ-IMP-08", K, "", "Unchanged."),
    ("G-04", "Global", "Export", "Writes all ten sheets back in template layout, edits included.",
     "REQ-IMP-04, REQ-IMP-07", K, "", "Unchanged."),
    ("G-05", "Global", "Findings banner", "Summary of the last import; opens the full report.",
     "REQ-IMP-02", K, "", "Unchanged."),
    ("G-06", "Global", "[CHANGED] Unsaved-edit counter",
     "Always visible once anything is edited, and now states the validation standing of those edits.",
     "REQ-IMP-08, REQ-IMP-09", "Change",
     "When anything is changed, the update to be checked with validation rules and if no doubt on the "
     "updated all the changes to be included into the exported file when a use runs exporting.",
     "Done, and it was already the specified behaviour - what was missing was any sign of it on screen. "
     "The counter now reads '3 unsaved edits · all 3 pass validation', and says outright that a failing "
     "edit is rejected at entry so nothing invalid can reach the export. The rule has not changed; it "
     "has become visible."),
    ("G-07", "Global", "[CHANGED] Tab bar",
     "Overall | Source data (project) | Source data (person) | General assumptions.",
     "REQ-DSH-01..04, REQ-DSH-11", "Change",
     "Add 'General assumtions' tab having other general information (e.g. 'PeriodWeightStandard' "
     "sheet, 'RoleFactor' sheet)",
     "Done - a fourth tab, and the one gap in the design worth closing. Every figure on the Overall "
     "tab is the product of these two tables, yet they were reachable only by opening the workbook. "
     "See A-01 to A-04 for what the tab holds. This one needed a new requirement, REQ-DSH-11."),
    ("G-08", "Global", "Empty state", "With nothing loaded: one load panel and a template download link.",
     "REQ-IMP-03", K, "", "Unchanged."),

    ("O-01", "Overall", "Horizon control", "From/to month, defaulting to 24 months.",
     "REQ-CAL-01", K, "", "Unchanged."),
    ("O-02", "Overall", "Expand to all projects", "One click widens the horizon to the latest project end date.",
     "REQ-DSH-07", K, "", "Unchanged."),
    ("O-03", "Overall", "[CHANGED] Filters (GLOBAL)",
     "Project type, clinical phase, project, person, role, department. One setting drives every tab.",
     "REQ-DSH-05", "Change",
     "Filter for 'Unit' isn't needed because of no meaning. Instead, add 'Clinical trial phase' right "
     "side of 'Project type'.",
     "Done both ways. 'Unit' is gone from the bar and moved to the new tab as configuration (see O-04) "
     "- you are right that it was never a filter: it changed how figures are written, not which are "
     "shown. 'Clinical phase' now sits immediately right of 'Project type', where the two type-ish "
     "controls read as a pair."),
    ("O-03b", "Overall", "Reset filters", "Clears every filter and restores the 24-month default in one action.",
     "REQ-DSH-05", K, "", "Unchanged."),
    ("O-04", "Overall", "[CHANGED] Unit toggle",
     "FTE or hours. Now a setting on the General assumptions tab, not a filter.",
     "REQ-CAL-08", "Change",
     "Include configuration information into the new sheet (i.e. General assumptions).",
     "Done. The whole Config sheet is on the new tab, and the display-unit control sits with it, "
     "under a line explaining why it is not in the filter bar."),
    ("O-05", "Overall", "Summary tiles",
     "Projects, people, total demand, over-allocated months, under-allocation runs.",
     "REQ-DSH-08", K, "", "Unchanged."),
    ("O-06", "Overall", "[CHANGED] Demand chart",
     "Stacked monthly demand, one band per project, largest on the baseline. No legend; hover pop-up instead.",
     "REQ-DSH-02", "Change",
     "In 'Monthly demand by project' section, remove legend information on the bottom. Instead, while "
     "hovering over each stack in the graph, provide pop-up information (e.g. project name, project's "
     "month FTE, acting people in the months).",
     "Done. The legend is gone and a real hover pop-up replaces it, carrying project name and type, "
     "that month's FTE and its hour equivalent, the headcount, and every person on the project that "
     "month with their role. Note the knock-on: D-11 argued that identity came from the legend order "
     "because 62 hues cannot be told apart. The tooltip now carries that alone - which is stronger, "
     "since it names one band rather than asking you to match a colour against a list of 62."),
    ("O-07", "Overall", "Resource by project (table)",
     "Project x month heatmap. Sorted NewDrug CT, Biosimilar CT, Others, then earliest first.",
     "REQ-DSH-01", K, "", "Unchanged."),
    ("O-07b", "Overall", "Project row expansion",
     "Clicking a project name reveals a row per person and role, each with its own monthly figures.",
     "REQ-DSH-01", K, "", "Unchanged."),
    ("O-08", "Overall", "Mean load per person (chart)",
     "One bar per person against both absolute thresholds; above the bar budget, a ranked subset.",
     "REQ-DSH-02, REQ-DSH-08, REQ-DSH-09", K, "", "Unchanged."),
    ("O-09", "Overall", "Resource by person (table)",
     "Person x month, summed across projects, over/under flagged.",
     "REQ-DSH-01, REQ-DSH-08", K, "", "Unchanged."),
    ("O-09b", "Overall", "Person row expansion",
     "Clicking a person name reveals a row per project and role, in the same type-then-date order.",
     "REQ-DSH-01", K, "", "Unchanged."),
    ("O-10", "Overall", "[CHANGED] Project timeline (Gantt)",
     "First panel on the tab. Duration under each name, bands coloured by period, milestone markers.",
     "REQ-DSH-02, REQ-PRJ-05, REQ-DSH-10", "Change",
     "1. Add project duration (start date, end date, total month) under individual project name. "
     "2. Change color more intuitive. e.g. before start-up (Grey), Start-up (Red), Conduct (Green), "
     "Close-out (Orange), After close-out (Dark Grey). Information given when hovering over, should "
     "include monthly FTE for each period. "
     "3. No need to seperate 'Inspection' because of that being one of milestones. Change the icon for "
     "milestone from circle to inverted triangle highlighted. "
     "4. Location: as first section right after 'Horizon and filters' section.",
     "All four done. (1) 'start -> end · n months' sits under each project name. (2) Your mapping is "
     "followed, with one departure: red beside green is the pair red-green colour blindness collapses, "
     "and you accepted D-04 making 'never colour alone' a floor - so the red is shifted slightly toward "
     "orange, every band wide enough carries its period name as text, and the tooltip names it outright. "
     "The two close-outs take a light and a deep orange so they are not one indistinguishable block. "
     "The tooltip carries dates, weight and the FTE per month the project draws across that period. "
     "(3) 'Inspection' now takes the same marker as every other milestone, and the marker is an inverted "
     "triangle in its own lane above the bands, so it never lands on a band label. (4) The timeline is "
     "now the first panel, above the summary tiles. This change supersedes D-06 - see sheet 02."),

    ("P-01", "Project tab", "[CHANGED] Project table",
     "All 23 columns, sortable, filterable, editable, with an insert control on every row.",
     "REQ-DSH-03, REQ-IMP-07, REQ-IMP-11", "Change",
     "Add 'Insert new row' button for all rows and all sections. When clicking the button, new row to "
     "be added right after the row where a user clicks that button.",
     "Done, and generalised - see X-06. Every row of every editable table now carries '+ row', and the "
     "new row lands directly below the row you pressed, not at the bottom of the table. This needed a "
     "new requirement, REQ-IMP-11: nothing in the plan said a row could be created at all."),
    ("P-02", "Project tab", "Milestone sub-table",
     "Milestones of the selected project in date order; Inspection may repeat.",
     "REQ-PRJ-05, REQ-PRJ-13", K, "", "Unchanged except for gaining the insert control (X-06)."),
    ("P-03", "Project tab", "Period sub-table",
     "Derived periods with seq, dates, weight and note; repeated names numbered.",
     "REQ-PRJ-06, REQ-CAL-09, REQ-DSH-10", K, "", "Unchanged except for gaining the insert control (X-06)."),
    ("P-04", "Project tab", "Recompute periods",
     "Re-derives from current milestones; warns before replacing hand-set dates.",
     "decision C-10", K, "", "Unchanged."),
    ("P-05", "Project tab", "Export visible table", "Current table to .xlsx.",
     "REQ-DSH-06", K, "", "Unchanged."),
    ("P-06", "Project tab", "[NEW] Project utilisation graph",
     "The selected project's monthly resource across the horizon, against three relative reference "
     "lines. Sits directly under the project table.",
     "REQ-DSH-12", "Add (round 6)",
     "In 'Source Data (project)', add 'Utilisation' graph for projects specific monthly resource with "
     "bars (upper: 2 times x average of all projects FTE, lower: 0.5 times x average of all projects, "
     "additional: the project's average FTE during entire period). The graph location is next to the "
     "'Project' table.",
     "Added, with your three lines exactly as specified. Your reasoning is what made it a new "
     "requirement rather than a copy of the person strip: a project has no static threshold, so the "
     "person tab's absolute ceiling and floor do not transfer. Hence REQ-DSH-12. One judgement call to "
     "flag - the portfolio average is taken over ACTIVE project-months only. Including months where a "
     "project draws nothing would pull the average toward zero and make every running project look "
     "heavy against it. Say if you meant the simple mean across all months instead."),

    ("S-01", "Person tab", "[CHANGED] Person table",
     "All 12 columns, sortable, filterable, editable, with an insert control on every row.",
     "REQ-DSH-04, REQ-IMP-07, REQ-IMP-11", "Change",
     "Add 'Insert new row' button for all rows and all sections. When clicking the button, new row to "
     "be added right after the row where a user clicks that button.",
     "Done - same treatment as P-01, and the same generalisation to every section (X-06)."),
    ("S-02", "Person tab", "Utilisation strip",
     "Selected person's monthly load with both absolute thresholds drawn.",
     "REQ-DSH-08", K, "", "Unchanged."),
    ("S-03", "Person tab", "Assignment sub-table", "Project, role, dates and person weight.",
     "REQ-PSN-02, REQ-PSN-03", K, "", "Unchanged except for gaining the insert control (X-06)."),
    ("S-04", "Person tab", "Override sub-table", "PersonPeriodWeight windows for the selected assignment.",
     "REQ-PSN-05", K, "", "Unchanged except for gaining the insert control (X-06)."),
    ("S-05", "Person tab", "Export visible table", "Current table to .xlsx.",
     "REQ-DSH-06", K, "", "Unchanged."),

    ("E-01", "Editing", "Inline cell edit", "Every field editable, validated at the point of entry.",
     "REQ-IMP-09", K, "", "Unchanged."),
    ("E-02", "Editing", "Cascade confirm",
     "Changing an identifier states how many rows will follow, then rewrites them.",
     "REQ-IMP-10, V-17", K, "", "Unchanged."),
    ("E-03", "Editing", "Delete guard",
     "Refuses to delete a row that is still referenced, naming what points at it.",
     "V-17", K, "", "Unchanged."),

    ("X-01", "Layout", "[CHANGED] Scroll regions",
     "Every chart and table scrolls inside its own panel - horizontally when wide, and within a bounded "
     "height when tall, with the header row staying visible.",
     "REQ-NFR-02", "Change (round 3)",
     "Apply scroll bar on the bottom of all sections of 'Source data (project)'/'Source data (person)'/"
     "'General assumptions'. Some sections have problems that table size over the fixed section site. "
     "(e.g. Periods of 'Source data (project)', Assignments and Weight overrides of 'Source data (person)')",
     "Done. Six tables were rendering outside their panels: the milestone and period sub-tables, "
     "assignments, weight overrides, and both assumptions tables. Each now sits in a scroll region that "
     "is bounded in BOTH directions - the earlier rule only covered width, which is why a long "
     "sub-table still grew the page and pushed the panels below it down. Verified by measuring the "
     "document's scroll width against the viewport: they match, so nothing overflows sideways."),
    ("X-02", "Layout", "Type and phase pills",
     "Project type and clinical phase as labelled pills; text carries the meaning, colour only speeds recognition.",
     "REQ-PRJ-01, REQ-PRJ-09", K, "", "Unchanged."),
    ("X-03", "Layout", "[CHANGED] Numbered repeated periods — now a guard only",
     "Retained, but it never fires: since R-11 no period name repeats, so a name alone identifies a "
     "period on screen.",
     "REQ-DSH-10", "Change (round 4)",
     "Change period name for conduct as 'Conduct (interim)' and 'Conduct (final)' in order to "
     "distinguish multiple periods.",
     "Done - see sheet 06. This component is what REQ-DSH-10 asked for, and your change satisfies that "
     "requirement structurally instead. The numbering code is kept as a guard: V-18 now rejects a "
     "repeated name on import, but a hand-built model could still reach the renderer, and a silent "
     "collision there would be worse than a redundant ten lines."),
    ("X-04", "Layout", "Row virtualisation",
     "Both Overall tables render only the rows in the viewport plus overscan.",
     "REQ-DSH-09, REQ-NFR-03", K, "", "Unchanged."),

    ("A-01", "Assumptions tab", "[NEW] Standard period weights",
     "PeriodWeightStandard as a matrix - phase down the side, period across, shaded by magnitude.",
     "REQ-DSH-11", "", "",
     "Added for G-07. Shown as a matrix, not 48 flat rows: it is a standard, and a standard is read "
     "across. 'Others' projects are absent by design - their weights are hand-entered per project."),
    ("A-02", "Assumptions tab", "[CHANGED] Role factors — clinical trials",
     "RoleFactor as a matrix: type + phase + role down the side, the six periods across.",
     "REQ-DSH-11", "Change (round 3)",
     "Weight of Role factors should be given by Role & Project type & Clinical phase & Periods.",
     "Done - see sheet 05. The key is now all four columns and the sheet grows from 13 rows to 249, so "
     "it is shown as a matrix: 40 rows of six periods each. Reading a role ACROSS the periods is the "
     "point of the change - the database programmer peaks at start-up, the analyst at lock. Also fixed "
     "here: this table had an 'insert' header with no cells beneath it, so every row was a column out "
     "of alignment. It is a reference matrix now, and reference matrices carry no insert control."),
    ("A-02b", "Assumptions tab", "[NEW] Role factors — Others",
     "The same matrix for non-trial projects: role down the side, the three 'Others' periods across.",
     "REQ-DSH-11", "", "",
     "Split out because 'Others' projects carry no clinical phase and run a different period set, so "
     "they cannot share a matrix with the trials without a column of blanks."),
    ("A-03", "Assumptions tab", "[NEW] Configuration",
     "Config: thresholds and settings, plus the display-unit control moved here from the filter bar.",
     "REQ-DSH-11, REQ-CAL-08", "", "",
     "Added for G-07 and O-04. The thresholds that colour the tables now sit next to a note saying "
     "what they mean."),
    ("A-04", "Assumptions tab", "[CHANGED] Value lists",
     "Lists: what each list-typed column will accept, and how many values each list holds. Read-only.",
     "REQ-DSH-11", "Change (round 3)",
     "Insert row button ... Check all tables of the 'General assumptions'.",
     "Fixed. This table had insert BUTTONS it should never have had - the header carried no insert "
     "column, so it too was out of alignment, the opposite way round from A-02. It is now explicitly "
     "read-only: a value added here with nothing referring to it is noise, and the note says so."),
    ("X-05", "Layout", "[NEW] Hover pop-up layer",
     "A real tooltip - follows the cursor, flips at the screen edge, carries formatted multi-line content.",
     "REQ-DSH-02", "", "",
     "Added for O-06 and O-10. The native SVG tooltip cannot show a list of people and waits half a "
     "second to appear, so it could not do what you asked for. This moves off the deferred list, where "
     "v0.3 had it."),
    ("X-06", "Layout", "[NEW] Insert-row control",
     "'+ row' leading every row of every editable table; the new row lands directly below that row.",
     "REQ-IMP-11", "", "",
     "Added for P-01 and S-01, generalised to 'all sections' as you asked: both source-data tables, "
     "all four sub-tables, and the role-factor and config tables on the new tab. The control leads the "
     "row - see D-18 for why."),
]
# Derived, not typed: the counts have gone stale twice already.
SUBTITLE_CELL.value = (
    f"{len(C)} components. {sum(1 for c in C if c[5] == K)} Keep, "
    f"{sum(1 for c in C if str(c[5]).startswith('Change'))} Change (amber), "
    f"{sum(1 for c in C if '[NEW]' in c[2])} added by those changes (green). "
    f"Your comments are quoted exactly as written.")
rows = [list(c) + ["OK (v0.7)" if c[0] in CONFIRMED else ""] for c in C]
r = table(ws, 4,
          ["ID", "Area", "Component", "What it does", "REQ-IDs",
           "Your decision", "Your comment (verbatim)", "What was done in v0.4", "Confirm"],
          rows, [8, 15, 30, 56, 26, 12, 62, 86, 14],
          wrap_cols=(3, 4, 7, 8, 9), yellow_col=9, mark_col=3)
dv = DataValidation(type="list", formula1='"OK,Rework"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"I5:I{4 + len(rows)}")
r = table(ws, r, ["Count", "Keep", "Changed", "Added"],
          [[f"=COUNTA(A5:A{4 + len(rows)})",
            sum(1 for c in C if c[5] == K),
            sum(1 for c in C if c[5] == "Change"),
            sum(1 for c in C if "[NEW]" in c[2])]], [14, 10, 10, 10])
ws.cell(r, 1, "Green = added by this review. Amber = changed by this review.").font = NOTE_F

# ---- design decisions ----------------------------------------------------
ws = wb.create_sheet("02_Design_Decisions")
ws.sheet_view.showGridLines = False
ws["A1"] = "Design decisions — all 15 accepted, one now superseded"
ws["A1"].font = TITLE_F
ws["A2"] = ("You accepted every decision. D-06 is nonetheless overturned, because O-10 asks for the "
            "opposite; that conflict is the one thing on this sheet needing your eye.")
ws["A2"].font = NOTE_F
ws.freeze_panes = "A5"

D = [
    ("D-01", "SUPERSEDED at the v0.2 review - the demand chart stacks by individual project.",
     "You asked for per-project bands, so that is what the chart does: ordered by total resource with "
     "the largest on the baseline, 'Others' grey.", "Accept", "Stands. No change."),
    ("D-06", "[CHANGED] SUPERSEDED by your own O-10 - the timeline now colours bands by PERIOD NAME.",
     "D-06 said the timeline shades by weight, because weight is what drives the simulation and naming "
     "periods by colour would spend the palette on labels already in the tooltip. You accepted that, and "
     "then asked in O-10 for exactly the opposite - grey, red, green, orange, dark grey by period.",
     "Accept (conflicts with O-10)",
     "O-10 wins: it is the more specific instruction and the clearer statement of what you want to see. "
     "But D-06's point was real, so weight was not thrown away - it rides as a lightness step inside each "
     "period's hue, deliberately over a narrow range so it never competes with the hue, and it stays exact "
     "in the tooltip. If you did mean to keep weight as the colour, mark this Rework and O-10.2 goes back."),
    ("D-11", "[CHANGED] Beyond seven projects, colour comes from an EXTENDED palette, not the validated one.",
     "The validated set caps at eight hues; per-project colour needs ~50, so the seven validated hues are "
     "stepped in lightness. Hue alone no longer identifies a band.", "Accept",
     "Stands, but its safety net changed. D-11 said identity came from 'the legend order, the tooltip and "
     "the table below'. O-06 removed the legend, so the tooltip now carries it alone - which is the stronger "
     "half of that pair anyway, since it names one band rather than asking you to match a hue against 62."),
    ("D-12", "The project table samples the head of each type rather than listing a flat top-14.",
     "Your sort puts all 34 NewDrug CT projects first, so a flat top-14 would never reach the other types "
     "and the ordering could not be seen working.", "Accept", "Stands. No change."),
    ("D-13", "The global filter bar is not sticky.",
     "A sticky bar 110px tall covered panel headings as soon as the page scrolled.", "Accept",
     "Stands. No change."),
    ("D-02", "The project table lists the ten busiest projects plus an aggregate row.",
     "62 rows x 12 months does not fit on screen; the real table lists all 62 with sort and filter.",
     "Accept", "Stands. No change."),
    ("D-03", "Heat shading in the tables is one blue ramp, light to dark.",
     "Magnitude is a sequential quantity, so it takes one hue. The aggregate row is deliberately not shaded.",
     "Accept", "Stands, and now also carries the period-weight matrix on the new tab, so the same ramp "
     "means the same thing everywhere."),
    ("D-04", "Over- and under-allocation carry an icon and a value, never colour alone.",
     "Red-green colour blindness affects around 8% of men; a cell that says only 'red' is unreadable to "
     "them and unprintable in mono.", "Accept",
     "Stands, and it is what shaped the O-10 palette: your red/green pair is exactly the collapse this "
     "decision guards against, so the bands carry text labels and the red sits slightly toward orange."),
    ("D-05", "Both thresholds are drawn on the person chart and the utilisation strip.",
     "A ceiling without a floor makes under-use invisible, and under-use is half of what the tool is for.",
     "Accept", "Stands. No change."),
    ("D-07", "Editing is inline in the tables, not in a separate form or dialog.",
     "A dialog per row would make bulk correction - the common case after a timeline slips - slow enough "
     "that people would go back to editing the workbook by hand.", "Accept",
     "Stands, and extends naturally to X-06: a row is inserted in place, where the user is looking."),
    ("D-08", "The findings banner persists until dismissed rather than fading.",
     "A toast that disappears is a finding nobody read.", "Accept", "Stands. No change."),
    ("D-09", "Dark mode is supported, with its own selected colour steps.",
     "Not a flip of the light palette: the ramp, the categorical hues and the status inks are each "
     "re-stepped for the dark surface.", "Accept",
     "Stands, and it caught a real fault in the new period palette: your 'dark grey' for After Close-out "
     "vanished against the dark surface. Both greys were re-picked to a pair that separates on either "
     "background."),
    ("D-10", "The page is one scrolling column per tab, not a fixed dashboard grid.",
     "It prints, it works on a laptop screen, and it needs no layout engine.", "Accept",
     "Stands, and the fourth tab follows the same pattern."),
    ("D-14", "At the target volume the person chart shows the 20 most loaded people, not all 1,000.",
     "1,000 bars across a 1,200px panel is 1.2px each - narrower than the gap between them.",
     "Accept", "Stands. No change."),
    ("D-15", "[CHANGED] SUPERSEDED — the conduct stretches are RENAMED, not numbered.",
     "D-15 argued for numbering because renaming would break the PeriodWeightStandard lookup, which is "
     "keyed on the period name. Its stated alternative was to label them by what separates them - "
     "'Conduct (pre-interim)' and 'Conduct (post-interim)'. You took that alternative in round 4.",
     "Accept, then superseded",
     "The objection was answered rather than ignored: the lookup does not break because the new names "
     "were added to the standard period set, so PeriodWeightStandard and RoleFactor are keyed on them "
     "like any other period. The cost is the one D-15 named - the set grows from six names to seven, and "
     "both weight tables grow with it. The gain is larger: period_name is unique within a project, so "
     "ProjectPeriod has a natural key and the display numbering is no longer needed at all."),
    ("D-16", "[NEW] The two 'Close-out' periods take a light and a deep orange, not one shared orange.",
     "Your mapping gave 'Close-out' a single colour, but a trial with an interim DB lock shows both in "
     "the same row, adjacent to each other. One shared hue made them read as a single interrupted band.",
     "", "Light amber for interim, deep orange for final. Both still read as 'the close-out family', "
         "which is what your mapping intended."),
    ("D-20", "[NEW] The two DB locks are drawn in red and larger; the '...cut-off' milestones are not.",
     "You asked for the interim and final DB locks to stand out because they matter more than the other "
     "milestones. They do, and specifically: they are what the whole period derivation hangs on - move a "
     "lock and every period after it moves with it. Nothing else on the timeline has that reach.",
     "", "Applied. Red plus a larger marker, so size carries the emphasis as well as hue - D-04, which "
         "you accepted, makes 'never colour alone' a floor for this UI. The markers sit in their own "
         "lane above the bands (D-17), so red reads cleanly there and does not collide with the red "
         "Start-up band. Judgement call to flag: 'interim DB lock cut-off' and 'final DB lock cut-off' "
         "are LEFT ordinary. The cut-off is preparation; the lock is the event that moves the timeline. "
         "Say if you want the cut-offs emphasised too."),
    ("D-21", "[NEW] The project utilisation graph uses RELATIVE reference lines, not thresholds.",
     "Your own reasoning: a project has no static threshold for over-burden or under-resource. So the "
     "lines are 2x and 0.5x the portfolio average, plus the project's own lifetime average - three "
     "reference points rather than two limits.",
     "", "Applied as specified, and it is why this needed a new requirement (REQ-DSH-12) rather than "
         "reusing the person strip. One consequence worth stating: because the lines move as the "
         "portfolio changes, a bar that is 'above the line' this month can fall below it next month "
         "with no change to the project itself. That is correct behaviour for a relative measure, but "
         "it means the graph answers 'heavy compared to what we usually run' rather than 'over "
         "budget'. The caption says so, because a dashed line above a bar reads as a limit unless it "
         "is labelled otherwise."),
    ("D-19", "[NEW] PeriodWeightStandard and RoleFactor are kept as two tables, not collapsed into one.",
     "Keying the role factor on phase and period means both tables now vary over (type, phase, period), "
     "and the calculation multiplies them - so they are mathematically collapsible into a single table "
     "keyed on all four columns. They are kept apart because they answer different questions: one is "
     "how busy the PROJECT is in a period, the other how much of that falls on a ROLE.",
     "", "Kept separate. But the separation is a maintenance convention, not something the arithmetic "
         "enforces: raising a project's Conduct load by editing all five role rows gives the right answer "
         "today and double-counts the next time the period weight moves. If that distinction is not kept "
         "in practice, the honest fix is to collapse them - a schema change, so flagged rather than done."),
    ("D-18", "[NEW] The insert control leads each row rather than trailing it.",
     "Placed at the end of the row, '+ row' sat off-screen on the 23-column project table and needed a "
     "horizontal scroll every time. Leading, it is always in view and reads as a row-action gutter.",
     "", "Applies to every editable table, so the control is in the same place everywhere. Alternative: "
         "trailing but pinned to the viewport edge, which costs a sticky column."),
    ("D-17", "[NEW] Milestone markers sit in their own lane above the bands, not on them.",
     "Inverted triangles drawn on the bands landed on top of the period labels, and on a busy row the "
     "two fought for the same pixels.",
     "", "The markers occupy a thin lane above each row's band. Alternative: draw them on the bands and "
         "drop the in-band labels, which trades one legibility problem for another."),
]
rows = [list(x) + ["OK (v0.7)" if x[0] in CONFIRMED else ""] for x in D]
r = table(ws, 4, ["ID", "Decision", "Why", "Your decision", "State after your review", "Confirm"],
          rows, [8, 60, 66, 18, 76, 14], wrap_cols=(2, 3, 5, 6), yellow_col=6, mark_col=2)
dv2 = DataValidation(type="list", formula1='"OK,Rework"', allow_blank=True)
ws.add_data_validation(dv2)
dv2.add(f"F5:F{4 + len(rows)}")

# ---- deferred ------------------------------------------------------------
ws = wb.create_sheet("03_Deferred")
ws.sheet_view.showGridLines = False
ws["A1"] = "Deliberately not in the prototype"
ws["A1"].font = TITLE_F
ws["A2"] = "So their absence is not read as an oversight."
ws["A2"].font = NOTE_F
rows = [
    ["Any loading, calculation, filtering or export behaviour", "Task 3.1 is design only. The figures shown "
     "are a fixed snapshot computed in Python and baked into the markup. The controls are real controls; "
     "nothing is wired behind them."],
    ["Exact spacing, type scale and final colour values", "Fixed at task 3.3, once the components are agreed. "
     "The period hues are the exception - they now carry meaning, so they were settled here."],
    ["Keyboard shortcuts and full accessibility pass", "Reviewed against the working prototype at task 3.3, "
     "where they can be tried rather than described."],
    ["Row virtualisation at the target volume", "Specified (X-04, REQ-DSH-09) but not built: the prototype "
     "shows 9 rows of 62, where virtualisation would be invisible. It is a Step 4 concern."],
    ["The validation findings report itself", "Specified on sheet 04 of the specification; the banner that "
     "opens it is in the prototype."],
    ["NO LONGER DEFERRED - hover tooltips", "v0.3 deferred the tooltip layer to code generation. O-06 and "
     "O-10 both depend on it, so it is built and reviewable now (X-05)."],
]
r = table(ws, 4, ["Not shown", "Why"], rows, [56, 100], wrap_cols=(1, 2))

# ---- what the review changed elsewhere -----------------------------------
ws = wb.create_sheet("04_Change_Log")
ws.sheet_view.showGridLines = False
ws["A1"] = "What your review changed outside this document"
ws["A1"].font = TITLE_F
ws["A2"] = ("Two components could not be satisfied by design alone - they asked for behaviour the plan "
            "did not require. Those became requirements.")
ws["A2"].font = NOTE_F
rows = [
    ["REQ-DSH-11", "NEW", "G-07",
     "The application presents the standing assumptions - standard period weights, role factors, "
     "configuration and value lists - on their own tab, without the user opening the workbook.",
     "Must", "Plan v1.6, change R-09"],
    ["REQ-IMP-11", "NEW", "P-01, S-01",
     "A new row can be inserted into any editable table, immediately below the row the user acts on, "
     "and is validated on entry like any other edit.",
     "Must", "Plan v1.6, change R-09"],
    ["REQ-DSH-05", "REWORDED", "O-03",
     "The filter set now names clinical phase, and no longer implies the display unit is a filter.",
     "Must", "Plan v1.6, change R-09"],
    ["REQ-IMP-05", "REWORDED", "G-02",
     "The loaded-file stamp carries its time zone.",
     "Should", "Plan v1.6, change R-09"],
    ["Specification sheet 06", "UPDATED", "all nine",
     "Tab 4 specified; filter set, load stamp, insert-row behaviour, timeline colour rule, tooltip "
     "content and edit-counter wording all restated.",
     "-", "Specification v0.5"],
    ["Design decision D-06", "SUPERSEDED", "O-10",
     "Timeline colour moves from weight to period name; weight becomes a lightness step within the hue.",
     "-", "Sheet 02 of this document"],
]
r = table(ws, 4, ["Item", "Change", "Raised by", "What it says now", "Priority", "Carried in"],
          rows, [22, 14, 16, 88, 10, 26], wrap_cols=(4,))
r = ws.max_row + 2
ws.cell(r, 1, "Nothing else in the plan or specification moved. Everything else you asked for was a "
              "design change the existing requirements already permitted.").font = NOTE_F

# ---- round 3 ------------------------------------------------------------
ws = wb.create_sheet("05_Round3")
ws.sheet_view.showGridLines = False
ws["A1"] = "Review round 3 — three items"
ws["A1"].font = TITLE_F
ws["A2"] = "Raised after prototype v0.4. All three applied in v0.5."
ws["A2"].font = NOTE_F
rows = [
    ["1", "Data model",
     "Weight of Role factors should be given by Role & Project type & Clinical phase & Periods. Update all "
     "relevant outputs (plan, specification, sourcedata template, dummy test sourcedata file, prototype, "
     "UI_Component_List).",
     "Applied as plan change R-10, and it is a genuine gain: a role's burden is not flat across a project, "
     "and the dummy data now shows the database programmer peaking at start-up and the analyst at lock. "
     "RoleFactor gains clinical_phase and period_name; the key becomes all four columns; the sheet grows "
     "from 13 rows to 249; schema version steps 3 to 4. Every output listed was regenerated. "
     "TWO COSTS, both stated rather than hidden. (a) 249 rows is a large table to maintain by hand. "
     "(b) RoleFactor now varies over the same three dimensions as PeriodWeightStandard and the two "
     "multiply, so editing both for the same reason double-counts - see D-19 and specification sheet 05. "
     "A new rule, V-23, catches a factor missing for a period an assignment actually spans, which would "
     "otherwise drop that stretch silently to 1.00.",
     "Plan v1.7 (R-10), spec v0.6, template v1.5, dummy v1.6, prototype v0.5"],
    ["2", "Layout",
     "Apply scroll bar on the bottom of all sections of 'Source data (project)' / 'Source data (person)' / "
     "'General assumptions'. Some sections have problems that table size over the fixed section site. "
     "(e.g. Periods of 'Source data (project)', Assignments and Weight overrides of 'Source data (person)')",
     "Applied. Six tables were rendering outside their panels. The v0.4 rule bounded WIDTH only, which is "
     "why the ones you named still overflowed - Periods is too wide for a half-width panel, Assignments "
     "and Weight overrides are too tall. Each table now sits in a region bounded in both directions, and "
     "keeps its header row visible while scrolled. Checked by measuring the document scroll width against "
     "the viewport width: equal, so nothing overflows sideways.",
     "Prototype v0.5, component X-01, spec v0.6 sheet 06"],
    ["3", "Layout",
     "Insert row button should be added, however only column is added now but no button to add new row is "
     "added under 'Insert' column. It looks like wrong information. Check all tables of the "
     "'General assumptions'.",
     "Fixed, and it was worse than you saw. The role-factor table had an 'insert' HEADER with no cells "
     "beneath it, so every row of that table was a column out of alignment - the note you read as wrong "
     "information was wrong. The value-lists table had the opposite fault: insert buttons with no header. "
     "Both came from one careless edit in v0.4 that moved the control to the front of the row and missed "
     "these two tables. Now: role factors are reference matrices with no insert control at all, value "
     "lists are explicitly read-only, and Config keeps its per-row control.",
     "Prototype v0.5, components A-02, A-02b, A-04"],
]
r = table(ws, 4, ["#", "Area", "What you raised (verbatim)", "What was done", "Carried in"],
          rows, [5, 14, 66, 104, 34], wrap_cols=(3, 4, 5))
r = ws.max_row + 2
ws.cell(r, 1, "Item 3 is the kind of fault a rendered screenshot catches and a code diff does not. The "
              "prototype is now rendered and inspected on every build for exactly that reason.").font = NOTE_F

# ---- round 4 ------------------------------------------------------------
ws = wb.create_sheet("06_Round4")
ws.sheet_view.showGridLines = False
ws["A1"] = "Review round 4 — two items"
ws["A1"].font = TITLE_F
ws["A2"] = "Raised after prototype v0.5. Both applied in v0.6. The second follows from the first."
ws["A2"].font = NOTE_F
rows = [
    ["1", "Data model",
     "Change period name for conduct as 'Conduct (interim)' and 'Conduct (final)' in order to distinguish "
     "multiple periods. Conduct (interim): if the milestone has interim DB lock and is before interim DB "
     "lock then use this period name. Conduct (final): if the milestone is after interim DB lock or the "
     "project has only final DB lock without interim DB lock, use this period name.",
     "Applied as plan change R-11. Your rule maps exactly onto the derivation already in the "
     "specification: the split branch runs only when an interim DB lock exists, so 'Conduct (interim)' is "
     "emitted only there, and every other conduct stretch - including the single one of a project with no "
     "interim lock - is 'Conduct (final)'. The clinical period set grows from six names to seven. "
     "PeriodWeightStandard grows 48 to 56 rows and RoleFactor 249 to 289; both new entries carry the same "
     "weights the single 'Conduct' did, so this RENAMES without reweighting - the dummy dataset returns "
     "189 over-allocated person-months before and after, which is the evidence. Schema 4 to 5: no column "
     "changed, but the period_name value set did, so a v4 file's 'Conduct' rows would now fail V-15.",
     "Plan v1.8 (R-11), spec v0.7, template v1.6, dummy v1.7, prototype v0.6"],
    ["2", "Data model",
     "Change key of ProjectPeriod to 'project_id + period_name' because now the period table has a unique "
     "period name list. Use the changed key to map relative information.",
     "Applied, and this is the payoff of item 1. period_name is now unique within a project, so "
     "(project_id, period_name) is a natural key and period_seq goes back to carrying ORDER rather than "
     "identity. V-18 becomes a plain uniqueness check on the name - proved by renaming one row in a copy "
     "of the dummy file and confirming the error. Three things simplify as a result: the ProjectPeriod key "
     "loses period_start, which was only ever there to tell two identically-named rows apart; decision "
     "D-15 is superseded; and REQ-DSH-10's display numbering becomes unnecessary, since a name alone now "
     "identifies a period on screen. The numbering code is kept as a guard, not as the mechanism.",
     "Plan v1.8 (R-11), spec v0.7 sheets 03 and 04, component X-03, decision D-15"],
]
r = table(ws, 4, ["#", "Area", "What you raised (verbatim)", "What was done", "Carried in"],
          rows, [5, 14, 66, 104, 34], wrap_cols=(3, 4, 5))
r = ws.max_row + 2
ws.cell(r, 1, "A note on what this change is worth: the previous design carried the repeated name through "
              "four places - a composite key, a validation rule, a display numbering rule and a "
              "requirement. Naming the two stretches apart removes the need for all four. It is the "
              "cheaper design and it was available from the start; D-15 argued against it on a lookup "
              "objection that turned out to be answerable.").font = NOTE_F

# ---- round 5 ------------------------------------------------------------
ws = wb.create_sheet("07_Round5")
ws.sheet_view.showGridLines = False
ws["A1"] = "Review round 5 — a question, not a change"
ws["A1"].font = TITLE_F
ws["A2"] = "The key was right. Checking it found two rules that were not implemented."
ws["A2"].font = NOTE_F
rows = [
    ["1", "Data model",
     "In the programming specification, key setting of 'PersonPeriodWeight' is 'assignment_id + "
     "period_start'. My opinion is why period_start is needed as key. 'assignment_id' is unique in the "
     "'PersonPeriodWeight' and 'Assignment' sheets. Explain me that point. If my opinion isn't hurting "
     "other data consistency then change the key setting of 'PersonPeriodWeight' as 'assignment_id' only.",
     "NOT CHANGED, and here is why. assignment_id is unique in Assignment - it is that sheet's primary "
     "key. It was also unique in PersonPeriodWeight, but only because the dummy file happened to carry "
     "one window per assignment; that is a property of the sample, not of the schema. "
     "PersonPeriodWeight is a CHILD of Assignment and one assignment may carry SEVERAL non-overlapping "
     "windows - a spell of leave, back to normal, then a peak - which is what REQ-PSN-05's 'overrides' "
     "(plural), V-06's assignment-window clause and the data model's 'periods within one assignment must "
     "not overlap' all assume. Keying on assignment_id alone would cap it at one window, and the second "
     "spell would need a second Assignment row, fragmenting one person-project-role across rows that are "
     "not really different assignments. Your condition - 'if it isn't hurting other data consistency' - "
     "is therefore not met, so the key stays as it is. Say the word if you want the single-window model "
     "anyway; it is a small change and the cost is exactly the one described above.",
     "Specification v0.8 sheet 03 now carries this reasoning"],
    ["2", "Validation",
     "(not raised - found while answering item 1)",
     "Two rules were specified but never implemented, and the question is what exposed them. (a) V-06 "
     "says periods within one project AND within one assignment must not overlap. Only the project half "
     "ran, so two overlapping override windows passed silently and the weight that applied in the shared "
     "months depended on the order the rows happened to sit in the file. (b) Nothing checked that a "
     "PersonPeriodWeight row pointed at a real assignment, so an orphan override was accepted and then "
     "ignored without a word - the typed weight simply never applied. Both are now in the reference "
     "implementation, the second as new rule V-24, and both were proved to fire against deliberately "
     "broken copies of the dummy file. The fixture itself is the root cause: it only ever had one window "
     "per assignment, so the multi-window path was never exercised. It now carries an assignment with "
     "two.",
     "Plan v1.9 (R-12), spec v0.8, dummy v1.8"],
]
r = table(ws, 4, ["#", "Area", "What you raised (verbatim)", "What was done", "Carried in"],
          rows, [5, 14, 66, 110, 30], wrap_cols=(3, 4, 5))
r = ws.max_row + 2
ws.cell(r, 1, "Worth noting for its own sake: a question about a key found two silent-wrong-answer bugs. "
              "Neither was reachable by reading the code, because both were absences - a rule that was "
              "written down and never built.").font = NOTE_F

# ---- round 6 ------------------------------------------------------------
ws = wb.create_sheet("08_Round6")
ws.sheet_view.showGridLines = False
ws["A1"] = "Review round 6 — everything confirmed, two additions"
ws["A1"].font = TITLE_F
ws["A2"] = ("All 13 changed components and all 7 changed decisions marked OK; none marked Rework. "
            "Two new requests, both applied in prototype v0.8.")
ws["A2"].font = NOTE_F
rows = [
    ["1", "UI",
     "Milestone markers color highlights 'Interim DB Lock' and 'Final DB Lock' as Red. Those milestones "
     "(interim/final DB Lock) are more important and it should be recognizable than others.",
     "Applied. Red, and a larger marker with it - size as well as hue, because D-04 makes 'never colour "
     "alone' a floor for this UI and a red triangle alone would be invisible to a reader with red-green "
     "colour blindness. The markers already sit in their own lane above the bands (D-17), so red reads "
     "cleanly and does not collide with the red Start-up band. The legend gains an entry naming what "
     "they are. JUDGEMENT CALL: 'interim DB lock cut-off' and 'final DB lock cut-off' are left "
     "ordinary - the cut-off is preparation, the lock is the event that moves the timeline. Say if you "
     "want the cut-offs emphasised too.",
     "Prototype v0.8, decision D-20, spec v0.9 sheet 06"],
    ["2", "UI",
     "In 'Source Data (project)', add 'Utilisation' graph for projects specific monthly resource with "
     "bars (upper: 2 times x average of all projects FTE, lower: 0.5 times x average of all projects, "
     "additional: the project's average FTE during entire period). The graph location is next to the "
     "'Project' table. Project also can provide chronological resource trend like 'Source data (person)' "
     "tab. In addition, a project doesn't have static threshold to measure over-burden or "
     "under-resource. It suggests some bars to give more informative stuffs.",
     "Applied with your three lines exactly as specified, directly under the project table where the "
     "person tab puts its strip. Your second sentence is the important one and it is why this became a "
     "new requirement rather than a copy: a person has a capacity, so absolute thresholds mean "
     "something; a project does not, so the references had to be relative. That is REQ-DSH-12. "
     "JUDGEMENT CALL: the portfolio average is taken over ACTIVE project-months only. Including months "
     "where a project draws nothing would pull the average toward zero and make every running project "
     "look heavy against it. Say if you meant the simple mean across all months. "
     "One property to be aware of: because the lines move with the portfolio, a bar can cross a line "
     "with no change to the project itself. Correct for a relative measure, but it means the graph "
     "answers 'heavy compared to what we usually run', not 'over budget' - and the caption says so.",
     "Prototype v0.8, component P-06, decision D-21, plan v1.10 REQ-DSH-12, spec v0.9"],
]
r = table(ws, 4, ["#", "Area", "What you raised (verbatim)", "What was done", "Carried in"],
          rows, [5, 10, 66, 112, 32], wrap_cols=(3, 4, 5))
r = ws.max_row + 2
ws.cell(r, 1, "Both items carry a judgement call I made rather than guessed at silently. Neither blocks "
              "anything: each is one line to change if I have read you wrong.").font = NOTE_F

wb.save(OUT)
print(f"Written: {OUT}  ({len(C)} components, {len(D)} decisions)")
