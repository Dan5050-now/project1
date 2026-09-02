"""Validate a PRAP source workbook against the rules baselined in the plan.

Implements enough of the calculation engine to prove the data is coherent and that
the dummy file demonstrates what its README claims. Doubles as a reference
implementation for the Step 4 calculation layer.

    python tools/verify_source_workbook.py templates/PRAP_SourceData_Dummy_v1.3.xlsx
"""

import calendar
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

CLINICAL_PERIODS = ["Before-Start-up", "Start-up", "Conduct (interim)",
                    "Close-out (interim)", "Conduct (final)",
                    "Close-out (final)", "After Close-out (final)"]
OTHER_PERIODS = ["Planning", "Develop", "Close"]
class _ClinicalTypes:
    """Anything whose name begins 'NewDrug CT' or 'Biosimilar CT' is a clinical trial.
    Schema 6 split the biosimilar type in two; asking the name rather than holding a
    fixed set means the next subdivision is a value-list change, not a code change."""

    def __contains__(self, t):
        return str(t or "").startswith(("NewDrug CT", "Biosimilar CT"))


CLINICAL_TYPES = _ClinicalTypes()

ANY_SCOPE = ""                                    # schema 6: 'this row fits any scope'


def scope_of(row):
    return str((row or {}).get("work_scope_type") or "")


def project_window(periods, proj):
    """How long the project runs, for the purpose of working out a number.

    THE PERIODS ARE THE PROJECT (REQ-CAL-17). Milestones are reference dates that the
    period derivation reads; several of them mark moments inside the run rather than
    its edges. A project with no periods keeps its own typed dates.
    """
    lo = hi = None
    for s in (periods.get(proj["project_id"]) or []):
        ps, pe = d(s.get("period_start")), d(s.get("period_end"))
        if ps is not None and (lo is None or ps < lo):
            lo = ps
        if pe is not None and (hi is None or pe > hi):
            hi = pe
    return (lo or d(proj["start_date"]), hi or d(proj["end_date"]))


def assignment_window(periods, proj, a):
    """Both assignment dates are optional; a blank one means the project's own
    (REQ-CAL-15). A blank START used to mean the row contributed nothing at all."""
    ps, pe = project_window(periods, proj)
    return (d(a["assign_start_date"]) or ps,
            d(a["assign_end_date"]) or pe)


def lookup(table, key, scope, tail):
    """Schema 6's two-step: the project's own scope first, then the any-scope row.

    An empty work_scope_type is a row that deliberately declines to distinguish, not a
    row somebody forgot to fill in - which is why the fallback is part of the contract
    rather than leniency. Any program reading these sheets must do the same, or it will
    report a missing weight where the application finds one.
    """
    v = table.get((*key, scope, *tail))
    return table.get((*key, ANY_SCOPE, *tail)) if v is None else v


def rows(ws):
    hdr = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        yield dict(zip(hdr, r))


def d(v):
    return v.date() if hasattr(v, "date") else v


def months_between(a, b):
    y, m = a.year, a.month
    while (y, m) <= (b.year, b.month):
        yield y, m
        m += 1
        if m == 13:
            y, m = y + 1, 1


def coverage(y, m, s, e):
    days = calendar.monthrange(y, m)[1]
    m0, m1 = date(y, m, 1), date(y, m, days)
    lo, hi = max(m0, s), min(m1, e)
    return 0.0 if hi < lo else ((hi - lo).days + 1) / days


def main(path):
    wb = load_workbook(path, data_only=False)
    P = {r["project_id"]: r for r in rows(wb["Project"])}
    # A milestone name may repeat within a project ('Inspection'), so this maps
    # name -> list of dates rather than name -> date (REQ-PRJ-13).
    MS = defaultdict(lambda: defaultdict(list))
    for r in rows(wb["Milestone"]):
        MS[r["project_id"]][r["milestone_name"]].append(d(r["milestone_date"]))
    PER = defaultdict(list)
    for r in rows(wb["ProjectPeriod"]):
        PER[r["project_id"]].append(r)
    # R-10: the factor is keyed on type + phase + period + role, so a role's burden can
    # move across the life of a project rather than being one number for the whole run.
    RF = {(r["project_type"], r["clinical_phase"], scope_of(r), r["period_name"],
           r["role_name"]): r["role_factor"] for r in rows(wb["RoleFactor"])}
    RF_ROLES = defaultdict(set)                 # project_type -> {role_name}
    for k in RF:
        RF_ROLES[k[0]].add(k[4])
    # Indexed by the ABSORBING role (REQ-CAL-16): which absent roles land on this one.
    RF_ABSORB = defaultdict(list)
    for r in rows(wb["RoleFactor"]):
        if r.get("absorbed_by"):
            RF_ABSORB[(r["project_type"], r["clinical_phase"], scope_of(r),
                       r["period_name"], r["absorbed_by"])].append(r["role_name"])
    PWS = {(r["project_type"], r["clinical_phase"], scope_of(r), r["period_name"]):
           r["standard_fte"] for r in rows(wb["PeriodWeightStandard"])}
    PSN = {r["person_id"]: r for r in rows(wb["Person"])}
    ASG = list(rows(wb["Assignment"]))
    PPW = defaultdict(list)
    for r in rows(wb["PersonPeriodWeight"]):
        PPW[r["assignment_id"]].append(r)
    CFG = {r["parameter"]: r["value"] for r in rows(wb["Config"])}
    LISTS = defaultdict(set)
    for r in rows(wb["Lists"]):
        LISTS[r["list_name"]].add(r["value"])

    errors, warnings = [], []

    # ---- referential integrity (V-01, V-02, V-03, V-08) ----
    for a in ASG:
        if a["project_id"] not in P:
            errors.append(f"V-01 {a['assignment_id']}: unknown project {a['project_id']}")
            continue
        if a["person_id"] not in PSN:
            errors.append(f"V-02 {a['assignment_id']}: unknown person {a['person_id']}")
        ptype = P[a["project_id"]]["project_type"]
        if a["role_name"] not in RF_ROLES[ptype]:
            errors.append(f"V-03 {a['assignment_id']}: role '{a['role_name']}' not valid for {ptype}")
    ids = [a["assignment_id"] for a in ASG]
    if len(ids) != len(set(ids)):
        errors.append("V-08: duplicate assignment_id")

    # ---- override windows: V-24 referential, V-06 overlap ----
    # PersonPeriodWeight is a child of Assignment, and (assignment_id, period_start) is
    # its key: ONE assignment may carry several non-overlapping windows. Both halves of
    # that sentence need a rule, and neither had one.
    known = set(ids)
    for aid in sorted(PPW):
        if aid not in known:
            errors.append(f"V-24 {aid}: PersonPeriodWeight refers to an assignment that "
                          f"does not exist; its override is silently ignored")
            continue
        wins = sorted(PPW[aid], key=lambda w: d(w["period_start"]))
        starts = [d(w["period_start"]) for w in wins]
        if len(starts) != len(set(starts)):
            errors.append(f"V-24 {aid}: two override windows share a period_start; "
                          f"(assignment_id, period_start) must be unique")
        for a_, b_ in zip(wins, wins[1:]):
            if d(b_["period_start"]) <= d(a_["period_end"]):
                errors.append(f"V-06 {aid}: override windows overlap - "
                              f"{d(a_['period_start'])}..{d(a_['period_end'])} and "
                              f"{d(b_['period_start'])}..{d(b_['period_end'])}. Which "
                              f"weight applies in the shared months would depend on row order")
        for w in wins:
            if d(w["period_end"]) < d(w["period_start"]):
                errors.append(f"V-05 {aid}: override window ends before it starts")

    # ---- period coverage, ordering, naming (V-06, V-12, V-15, V-18) ----
    for pid, proj in P.items():
        segs = sorted(PER.get(pid, []), key=lambda r: r["period_seq"])
        if not segs:
            errors.append(f"V-12 {pid}: no periods")
            continue
        allowed = (CLINICAL_PERIODS if proj["project_type"] in CLINICAL_TYPES
                   else OTHER_PERIODS)
        # V-18: (project_id, period_name) is the key since R-11, so a repeated name is
        # now a duplicate key rather than a legitimate second stretch. period_seq must
        # still be unique - it is what orders them.
        names = [s["period_name"] for s in segs]
        dupes = sorted({n for n in names if names.count(n) > 1})
        for n in dupes:
            errors.append(f"V-18 {pid}: period_name '{n}' appears {names.count(n)} times; "
                          f"(project_id, period_name) must be unique")
        seqs = [s["period_seq"] for s in segs]
        if len(seqs) != len(set(seqs)):
            errors.append(f"V-18 {pid}: duplicate period_seq")
        prev_end = None
        for s in segs:
            if s["period_name"] not in allowed:
                errors.append(f"V-15 {pid}: '{s['period_name']}' not in the {proj['project_type']} set")
            ps, pe = d(s["period_start"]), d(s["period_end"])
            if pe < ps:
                errors.append(f"V-05 {pid} seq {s['period_seq']}: end before start")
            if prev_end is not None:
                gap = (ps - prev_end).days
                if gap > 1:
                    errors.append(f"V-12 {pid}: {gap - 1}-day GAP before seq {s['period_seq']}")
                elif gap < 1:
                    errors.append(f"V-06 {pid}: OVERLAP at seq {s['period_seq']}")
            prev_end = pe
        pstart, pend = d(proj["start_date"]), d(proj["end_date"])
        if d(segs[0]["period_start"]) > pstart:
            warnings.append(f"V-12 {pid}: periods start after the project does")
        if d(segs[-1]["period_end"]) < pend:
            warnings.append(f"V-12 {pid}: periods end before the project does")

    # ---- weights present (V-19) ----
    for pid, proj in P.items():
        if proj["project_type"] not in CLINICAL_TYPES:
            continue
        ph = proj["clinical_phase"]
        if not ph:
            errors.append(f"V-19 {pid}: clinical trial with no clinical_phase")
            continue
        for s in PER.get(pid, []):
            if lookup(PWS, (proj["project_type"], ph), scope_of(proj),
                      (s["period_name"],)) is None:
                errors.append(f"V-19 {pid}: no standard weight for "
                              f"{proj['project_type']} / {ph} / {s['period_name']}")

    # ---- role factor coverage (V-23) ----
    # Raised from the CALCULATION, further down, not from the sheets. Asked here it
    # walked every period of every project an assignment belonged to and demanded a row
    # for each - including periods no assignment ever reached. Asked from the arithmetic
    # it names exactly the compositions a person-month had to guess a factor for.

    # ---- list membership (V-11) ----
    for pid, proj in P.items():
        for col, lname in [("project_type", "project_type"),
                           ("work_scope_type", "work_scope_type"),
                           ("status", "project_status")]:
            v = proj.get(col)
            if v and v not in LISTS[lname]:
                warnings.append(f"V-11 {pid}: '{v}' not in list {lname}")
    for pid, mm in MS.items():
        for nm, dates in mm.items():
            if nm not in LISTS["milestone_name"]:
                warnings.append(f"V-11 {pid}: milestone '{nm}' not in the standard list")
            # V-20: only 'Inspection' is expected to repeat
            if nm != "Inspection" and len(dates) > 1:
                warnings.append(f"V-20 {pid}: milestone '{nm}' recorded {len(dates)} times")
        # V-21: an inspection on or before the final DB lock stays a marker
        fdbl = (mm.get("final DB lock") or mm.get("interim DB lock") or [None])[0]
        if fdbl:
            early = [x for x in mm.get("Inspection", []) if x <= fdbl]
            if early:
                warnings.append(f"V-21 {pid}: {len(early)} Inspection date(s) on or before the "
                                f"final DB lock - treated as markers, not opening period 7")

    # ---- monthly simulation ----
    def period_weight(pid, y, m):
        for s in PER.get(pid, []):
            if d(s["period_start"]) <= date(y, m, 1) <= d(s["period_end"]):
                return s["weight"]
        return 1.00

    def period_of(pid, y, m):
        for s in PER.get(pid, []):
            if d(s["period_start"]) <= date(y, m, 1) <= d(s["period_end"]):
                return s["period_name"]
        return None

    def role_factor(proj, a, y, m):
        ph = proj["clinical_phase"] if proj["project_type"] in CLINICAL_TYPES else None
        pn = period_of(a["project_id"], y, m)
        v = lookup(RF, (proj["project_type"], ph), scope_of(proj), (pn, a["role_name"]))
        return 1.00 if v is None else v

    def person_weight(a, y, m):
        for w in PPW.get(a["assignment_id"], []):
            if d(w["period_start"]) <= date(y, m, 1) <= d(w["period_end"]):
                return w["weight_override"]
        return a["person_weight"]

    # Who shares a role, and when. The role factor is what the ROLE costs the project,
    # not what each person holding it costs, so it is divided between them - counted
    # per month, by distinct people. Off when split_shared_role_fte is 0.
    split = str(CFG.get("split_shared_role_fte", 1)) not in ("0", "0.0", "False")
    absorb = str(CFG.get("absorb_unstaffed_role_factor", 1)) not in ("0", "0.0", "False")
    # Built ALWAYS: the count is the divisor, the presence decides absorption.
    sharers = defaultdict(set)
    for a in ASG:
        if a["project_id"] not in P or a["person_id"] not in PSN:
            continue
        s, e = assignment_window(PER, P[a["project_id"]], a)
        if not s or not e:
            continue
        for y, m in months_between(s, e):
            if coverage(y, m, s, e) > 0:
                sharers[(a["project_id"], a["role_name"], y, m)].add(a["person_id"])

    roles_on = defaultdict(set)
    for (pid_, role_, y_, m_) in sharers:
        roles_on[(pid_, y_, m_)].add(role_)

    def std_monthly(proj, pn):
        """REQ-CAL-19: the month's demand in FTE, from PeriodWeightStandard."""
        if pn is None:
            return 1.0
        v = lookup(PWS, (proj["project_type"], proj["clinical_phase"]), scope_of(proj), (pn,))
        return 1.0 if v is None else float(v)

    def effective_factor(proj, a, y, m):
        """The role's factor plus the factor of any role that names it as cover and
        that nobody holds this month (REQ-CAL-16). One hop, deliberately."""
        rf = role_factor(proj, a, y, m)
        if not absorb:
            return rf
        ph = proj["clinical_phase"] if proj["project_type"] in CLINICAL_TYPES else None
        pn = period_of(a["project_id"], y, m)
        for role in (lookup(RF_ABSORB, (proj["project_type"], ph), scope_of(proj),
                            (pn, a["role_name"])) or []):
            if sharers.get((a["project_id"], role, y, m)):
                continue
            v = lookup(RF, (proj["project_type"], ph), scope_of(proj), (pn, role))
            rf += 0.0 if v is None else v
        return rf

    load = defaultdict(float)          # (person, y, m) -> FTE
    horizon = set()
    gaps = {}                          # V-23: what the arithmetic had to guess at
    lines = []                         # REQ-CAL-18: manual figures replace these
    for a in ASG:
        if a["project_id"] not in P or a["person_id"] not in PSN:
            continue
        proj = P[a["project_id"]]
        s, e = assignment_window(PER, proj, a)
        if not s or not e:
            continue
        ph = proj["clinical_phase"] if proj["project_type"] in CLINICAL_TYPES else None
        for y, m in months_between(s, e):
            cov = coverage(y, m, s, e)
            if cov <= 0:
                continue
            pn = period_of(a["project_id"], y, m)
            # A month in no period at all is V-12's finding: there is no period name for
            # a factor to be missing FOR.
            if pn is not None and lookup(RF, (proj["project_type"], ph), scope_of(proj),
                                         (pn, a["role_name"])) is None:
                g = gaps.setdefault((proj["project_type"], ph, scope_of(proj), pn,
                                     a["role_name"]), [set(), 0])
                g[0].add(a["project_id"])
                g[1] += 1
            share = (len(sharers.get((a["project_id"], a["role_name"], y, m), ())) or 1) \
                if split else 1
            # REQ-CAL-19: the standard is the demand, the project's period weight
            # adjusts it, and the role factors divide it between the roles STAFFED.
            denom = 0.0
            for r_ in roles_on.get((a["project_id"], y, m), ()):
                denom += effective_factor(proj, {**a, "role_name": r_}, y, m)
            frac = (effective_factor(proj, a, y, m) / share) / denom if denom > 0 else 0.0
            v = (std_monthly(proj, pn) * period_weight(a["project_id"], y, m)
                 * frac * person_weight(a, y, m) * cov)
            lines.append({"aid": a["assignment_id"], "pid": a["project_id"],
                          "sid": a["person_id"], "y": y, "m": m, "fte": v})
            horizon.add((y, m))

    # ---- REQ-CAL-18: manual figures, assignment first, then the project scaling ----
    EST = {}
    for r in rows(wb["MonthlyEstimate"]) if "MonthlyEstimate" in wb.sheetnames else []:
        if r.get("scope") and r.get("ref_id") and r.get("month"):
            EST[(str(r["scope"]), str(r["ref_id"]), str(r["month"]))] = r.get("fte")
    manual_p = {pid for pid, pr in P.items()
                if str(pr.get("estimation_type") or "").strip().lower() == "manual"}
    manual_a = {a["assignment_id"] for a in ASG
                if str(a.get("estimation_type") or "").strip().lower() == "manual"}

    def mkey(y, m):
        return f"{y}-{m:02d}"

    for L in lines:
        if L["aid"] in manual_a:
            v = EST.get(("assignment", L["aid"], mkey(L["y"], L["m"])))
            L["fte"] = 0.0 if v is None else float(v)
    grp = defaultdict(list)
    for L in lines:
        if L["pid"] in manual_p:
            grp[(L["pid"], L["y"], L["m"])].append(L)
    for (pid, y, m), g in grp.items():
        want = EST.get(("project", pid, mkey(y, m)))
        if want is None:
            for L in g:
                L["fte"] = 0.0
            continue
        have = sum(L["fte"] for L in g)
        if abs(have) < 1e-9:
            continue
        for L in g:
            L["fte"] *= float(want) / have
    for L in lines:
        load[(L["sid"], L["y"], L["m"])] += L["fte"]

    for k in sorted(gaps, key=lambda x: tuple(str(v) for v in x)):
        pl, n = sorted(gaps[k][0]), gaps[k][1]
        errors.append(f"V-23: no role factor for {k[0]} / {k[1] or '-'} / "
                      f"{k[2] or 'any scope'} / {k[3]} / {k[4]} - {n} person-month(s) on "
                      f"{len(pl)} project(s) ({', '.join(pl[:3])}"
                      f"{', ...' if len(pl) > 3 else ''}) were calculated at factor 1.00")

    # V-22: an absolute floor is only meaningful if everyone can reach it
    _floor = float(CFG["under_allocation_fte"])
    for sid, per in PSN.items():
        cap = per.get("capacity_fte")
        if cap is not None and float(cap) < _floor:
            warnings.append(f"V-22 {sid}: capacity {float(cap):.2f} FTE is below the "
                            f"under-allocation floor of {_floor:.2f}, so this person can never "
                            f"clear it however fully they are booked")

    over_fte = float(CFG["over_allocation_fte"])
    under_fte = float(CFG["under_allocation_fte"])
    min_months = int(CFG["under_allocation_min_months"])

    over = sorted([(p, y, m, v) for (p, y, m), v in load.items() if v > over_fte],
                  key=lambda t: (t[0], t[1], t[2]))

    runs = []
    for pid_ in sorted(PSN):
        seq = sorted([(y, m) for (p, y, m) in load if p == pid_])
        if not seq:
            continue
        cur = []
        for y, m in months_between(date(*seq[0], 1), date(*seq[-1], 1)):
            v = load.get((pid_, y, m), 0.0)
            if 0 < v < under_fte:
                cur.append((y, m, v))
            else:
                if len(cur) >= min_months:
                    runs.append((pid_, cur[0][:2], len(cur)))
                cur = []
        if len(cur) >= min_months:
            runs.append((pid_, cur[0][:2], len(cur)))

    # ---- report ----
    print(f"File: {Path(path).name}")
    print(f"  projects {len(P)} | people {len(PSN)} | assignments {len(ASG)} | "
          f"periods {sum(len(v) for v in PER.values())} | months spanned {len(horizon)}")
    print()
    if errors:
        print(f"ERRORS ({len(errors)}):")
        for e_ in errors:
            print("   ", e_)
    else:
        print("ERRORS: none - referential integrity, period coverage and weights all check out.")
    if warnings:
        print(f"\nWARNINGS ({len(warnings)}):")
        for w in warnings:
            print("   ", w)
    else:
        print("WARNINGS: none.")

    print(f"\nOVER-ALLOCATION (> {over_fte} FTE): {len(over)} person-months")
    for p, y, m, v in over[:8]:
        print(f"    {p} {y}-{m:02d}  {v:.2f} FTE")
    if len(over) > 8:
        print(f"    ... and {len(over) - 8} more")

    print(f"\nUNDER-ALLOCATION (< {under_fte} FTE for >= {min_months} months): {len(runs)} runs")
    for p, (y, m), n in runs:
        print(f"    {p} from {y}-{m:02d}, {n} months")

    peak = max(load.items(), key=lambda kv: kv[1]) if load else None
    if peak:
        (p, y, m), v = peak
        print(f"\nPeak person-month: {p} {y}-{m:02d} at {v:.2f} FTE "
              f"({v * float(CFG['fte_hours_per_month']):.0f} hours)")

    return 1 if errors else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else
                  "templates/PRAP_SourceData_Dummy_v1.8.xlsx"))
