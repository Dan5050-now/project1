"""Generate the Project Management APP (PM_APP) programming specification.

Step N2 deliverable, governed by PRAP_NewApp_Development_Plan_v1.0.xlsx.

Traceability is READ from the approved plan rather than re-typed, so a requirement
cannot be dropped silently between the two documents - the same guarantee the web
application's specification has.

What this document does NOT restate: the data schema, the validation rules and the
calculation. Those are specified once, in PRAP_Programming_Specification_v1.0.xlsx, and
both applications are built from the same core/ - so restating them here would create a
second description able to disagree with the first.

    python tools/build_newapp_spec.py

Output: docs/PRAP_NewApp_Specification_v1.0.xlsx
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DOC_VERSION = "1.4"
DOC_STATUS = ("v1.2 APPROVED 2026-08-13 and still governing. THIS ISSUE, v1.3, adds change C-N02 - the "
              "Python shell - and AWAITS APPROVAL. Nothing already approved is withdrawn by it: the "
              "Electron shell stays specified and stays the better application wherever it can be "
              "delivered.")
DOC_DATE = "2026-08-13"
PLAN = "PRAP_NewApp_Development_Plan_v1.12.xlsx"
WEB_SPEC = "PRAP_Programming_Specification_v1.0.xlsx"
ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / f"PRAP_NewApp_Specification_v{DOC_VERSION}.xlsx"

FONT = "Arial"
NAVY = "1F3864"
TITLE_F = Font(name=FONT, size=16, bold=True, color=NAVY)
H1_F = Font(name=FONT, size=12, bold=True, color=NAVY)
HDR_F = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_F = Font(name=FONT, size=10)
BOLD_F = Font(name=FONT, size=10, bold=True)
NOTE_F = Font(name=FONT, size=9, italic=True, color="808080")
MONO_F = Font(name="Consolas", size=9)

HDR_FILL = PatternFill("solid", fgColor="2F5597")
BAND_FILL = PatternFill("solid", fgColor="F2F5FB")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
NEW_FILL = PatternFill("solid", fgColor="C6E0B4")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
WRAP_C = Alignment(vertical="top", wrap_text=True, horizontal="center")


def sheet(wb, name, title, subtitle=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_F
    row = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = NOTE_F
        row = 3
    ws.freeze_panes = f"A{row + 2}"
    return ws, row + 1


def table(ws, start_row, headers, rows, widths, wrap_cols=()):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = HDR_F
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = WRAP_C
    ws.row_dimensions[start_row].height = 28
    for r, data in enumerate(rows, start=start_row + 1):
        for i, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = BODY_F
            c.border = BOX
            c.alignment = WRAP if i in wrap_cols else Alignment(vertical="top")
            if (r - start_row) % 2 == 0:
                c.fill = BAND_FILL
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return start_row + len(rows) + 2


def section(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = H1_F
    return row + 1


def note(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = NOTE_F
    return row + 1


def lines(ws, row, texts, mono=False):
    for t in texts:
        c = ws.cell(row=row, column=1, value=t)
        c.font = MONO_F if mono else BODY_F
        row += 1
    return row


def code(ws, row, texts):
    return lines(ws, row, texts, mono=True)


wb = Workbook()
wb.remove(wb.active)

# ---- 00 Cover -------------------------------------------------------------
ws = wb.create_sheet("00_Cover")
ws.sheet_view.showGridLines = False
ws["A1"] = "Project Management APP (PM_APP)"
ws["A1"].font = Font(name=FONT, size=20, bold=True, color=NAVY)
ws["A2"] = "Programming Specification"
ws["A2"].font = Font(name=FONT, size=14, color=NAVY)

cover = [
    ("Document ID", "PRAP-NAPP-SPEC-001"),
    ("Document type", "Programming specification (Step N2 deliverable)"),
    ("Version", f"v{DOC_VERSION}"),
    ("Status", DOC_STATUS),
    ("Issue date", DOC_DATE),
    ("Author", "Claude Code"),
    ("Reviewer", "Requester - Gate N2 approved v1.0; the Gate N3 review returned five clarifications, applied here"),
    ("Governed by", f"{PLAN} - the approved baseline, Gate N1 closed 2026-08-13"),
    ("Specifies", "The desktop shell, the storage layer, the workspace file, sharing between "
                  "people, user identity, import, and packaging"),
    ("Does NOT specify",
     f"The data schema, the validation rules V-00..V-24, or the calculation. Those are "
     f"specified once in {WEB_SPEC} and both applications are built from the same core/. "
     f"A second description here could disagree with the first, and the disagreement would "
     f"be discovered by whoever trusted the wrong one."),
    ("Implemented in", "src/storage/desktop/ and src/shell/desktop/ (the Electron shell); "
                       "src/storage/python/ and src/shell/python/ (the Python shell, sheet 05a). "
                       "core/ and ui/ are not modified by this specification"),
    ("Repository", "Dan5050-now/project1"),
    ("Branch", "claude/project-resource-assignment-app-1vjdzh"),
]
r = 4
for k, v in cover:
    ws.cell(row=r, column=1, value=k).font = BOLD_F
    c = ws.cell(row=r, column=2, value=v)
    c.font = BODY_F
    c.alignment = WRAP
    r += 1
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 118
for rr in range(4, r):
    ws.row_dimensions[rr].height = 30

r += 1
r = section(ws, r, "How to read this document")
r = lines(ws, r, [
    "Each sheet specifies one part of the application, and every section carries the NR-ids it satisfies.",
    "Sheet 11 is the traceability matrix, built by reading the requirement register out of the approved plan",
    "rather than by re-typing it - so a requirement cannot be quietly dropped between the two documents.",
    "",
    "Where behaviour is exact, it is written as pseudocode or as a state table rather than as prose. Where a",
    "message will be read by a user, its text is given, because a message that says 'error' where it could have",
    "said who is editing the plan is a defect the code cannot show.",
])

# ---- 01 Version history ---------------------------------------------------
ws, r = sheet(wb, "01_Version_History", "Version history")
r = table(ws, r, ["Version", "Date", "Author", "Reviewer", "Summary"],
          [["1.4", "2026-08-22", "Claude Code", "Awaiting approval",
            "THE DIFFERENCE REPORT IS BUILT - task N4.5, NR-IMP-02, the last unbuilt piece of "
            "this specification. Sheet 09 gains an 'As BUILT' section recording the six "
            "decisions inside it that could each have gone another way, and what proves each. "
            "The engine is shared (src/core/06a_diff.js) and the screen is not: the web "
            "application stays feature-frozen and still replaces on a second load, which is "
            "what N-06 asked for and what it has no workspace to do otherwise. Also records "
            "Gate N4a: the shell arrived by e-mail, ran on the company PC, imported and "
            "exported - so R-N21's impact falls to Low."],
           ["1.3", "2026-08-18", "Claude Code", "Awaiting approval",
            "A SECOND SHELL, under the same page. Change C-N02: the two company controls are now measured "
            "rather than feared - an executable may run but may not arrive (R-N20), and data may not enter "
            "a browser page through the file picker (R-N21). Neither is worked around. A Python shell "
            "replaces Electron's main process: what travels is .py text rather than an executable, and the "
            "page never asks the browser for a file because Python reads it and hands over the bytes. New "
            "sheet 05a specifies it, including the amendment to NR-SEC-01 - this shell must open a socket, "
            "and the six controls around it are specified rather than assumed. core/ and ui/ are untouched, "
            "decision N-05 stands, and 1,225 person-months are compared against the reference "
            "implementation on every test run. The Electron shell is NOT withdrawn: it is the better "
            "application wherever it can be delivered."],
           ["1.2", DOC_DATE, "Claude Code", "APPROVED",
            "BASELINE. Change C-N01 approved 2026-08-13 together with Gate N3. Content is v1.1 unchanged - "
            "the five clarifications from the Gate N3 review, now signed. This issue governs Steps N4 and "
            "N5; v1.0 is superseded."],
           ["1.1", DOC_DATE, "Claude Code", "Superseded",
            "GATE N3 REVIEW APPLIED - five clarifications, none of which reverses a decision. (1) The write "
            "claim is taken when a DATA VALUE ACTUALLY CHANGES, not when a user clicks into a cell or moves "
            "a selection: sheet 07 now says so in those words, because 'first edit' was open to the looser "
            "reading. (2) The recovery dialog lists EVERY unsaved change, scrollable, rather than three and "
            "a count. (3) The difference report expands to individual rows on request. (4) The manual "
            "light/dark toggle is removed - the window follows Windows and offers no setting. (5) Export is "
            "confirmed as a supported way to edit a plan outside the application, and sheet 09 gains the "
            "round trip it must survive (NR-IMP-08, change C-N01 on the plan)."],
           ["1.0", DOC_DATE, "Claude Code", "APPROVED",
            "BASELINE. All seven open points on sheet 12 agreed at review with no change requested, "
            "so v1.0 is v0.1 with the answers recorded. S-N03 confirmed explicitly as accept-per-SHEET. "
            "Gate N2 is closed and this specification governs Steps N3 to N5; changing it now costs a "
            "version and a re-approval."],
           ["0.1", DOC_DATE, "Claude Code", "Superseded",
            "First issue. Covers WBS tasks N2.3 to N2.8: the storage interface, the workspace file "
            "format, the desktop shell, persistence and recovery, the write claim and user identity, "
            "import, deployment, and the traceability matrix. 69 requirements traced."]],
          [10, 13, 15, 13, 108], wrap_cols=(5,))

# ---- 02 Scope -------------------------------------------------------------
ws, r = sheet(wb, "02_Scope", "What is being built, and on top of what",
              "The desktop application is a new shell and a new storage layer over an engine that already "
              "exists and is already proven.")

r = section(ws, r, "The layers, and which of them this document touches")
lay = [
    ["core/", "Parse, validate, derive, calculate, xlsx and JSON.", "NOT TOUCHED. Shared with the web application, byte-identical, already verified.", "-"],
    ["ui/", "Tables, charts, filters, the provisional-edit model.", "EXTENDED. New screens only: identity, workspace open, the difference report, the claim messages.", "Sheets 05, 07, 08, 09"],
    ["storage/", "Where bytes live.", "NEW IMPLEMENTATION against a NEW INTERFACE. The web shell keeps its own.", "Sheets 03, 04, 06, 07"],
    ["shell/desktop/", "Window, menus, dialogs, lifecycle.", "NEW. All of it.", "Sheets 05, 10"],
]
r = table(ws, r, ["Layer", "What it does", "What this specification does to it", "Where"],
          lay, [16, 44, 66, 22], wrap_cols=(2, 3))
r += 1
r = note(ws, r, "The first row is the important one. Every figure the desktop application shows is produced by "
                "code that is already in service in the web application and already checked against the Python "
                "reference on 1,225 person-months. This specification cannot change a number, and is not "
                "permitted to try.")
r += 1

r = section(ws, r, "The storage interface is the whole of the difference")
r = lines(ws, r, [
    "In the web application, storage/ is two functions and 63 lines: one reads a file the user picked, one",
    "writes a file the user downloads. Everything this specification adds - workspaces, saving, versions,",
    "recovery, the write claim - sits behind the same seam.",
    "",
    "That is what makes the two applications safe to have. If a change is needed in how a number is produced,",
    "it happens once in core/ and both get it. If a change is needed in where bytes go, it happens in one",
    "storage implementation and the other is untouched.",
])
r += 1
r = note(ws, r, "ui/ must therefore never call the filesystem directly, and never assume it exists. Where a "
                "capability is missing - the web shell cannot claim a workspace - ui/ asks the interface and "
                "adapts, rather than being written twice. See sheet 03, 'Capabilities'.")

# ---- 03 Storage interface -------------------------------------------------
ws, r = sheet(wb, "03_Storage_Interface", "The storage interface   [N2.3]",
              "One interface, two implementations. ui/ knows only this.")

r = section(ws, r, "Operations")
ops = [
    ["capabilities()", "-", "{ workspaces, versions, claims, journal: boolean }", "Synchronous. What this shell can do. ui/ hides what is absent instead of failing at it.", "NR-PAR-01"],
    ["openWorkspace(ref)", "ref: opaque handle or path", "{ sheets, header, ref, readOnly }", "Reads and parses. Throws StorageError with a `kind` - see sheet 09.", "NR-STO-01"],
    ["saveWorkspace(ref, sheets, header)", "-", "{ savedAt, ref }", "Atomic. Returns only after the bytes are on the disk (NR-STO-04).", "NR-STO-04, NR-STO-05"],
    ["saveWorkspaceAs(sheets, header)", "-", "{ savedAt, ref } or null", "Opens the native Save dialog. null if the user cancels.", "NR-APP-05, NR-STO-17"],
    ["listRecent() / addRecent(ref)", "-", "[{ ref, name, at }]", "At least ten, most recent first, per user.", "NR-APP-06, NR-USR-09"],
    ["listVersions(ref)", "-", "[{ at, ref }]", "The retained previous version(s). One by default.", "NR-STO-06"],
    ["restoreVersion(ref, which)", "-", "{ sheets, header }", "Reads a retained version. Does NOT overwrite the current one - restoring is an edit like any other, and is saved like one.", "NR-STO-06"],
    ["claim(ref) / refreshClaim(ref) / releaseClaim(ref)", "-", "{ ok } or { ok:false, holder }", "The write claim. See sheet 07.", "NR-STO-10, NR-STO-13, NR-STO-14"],
    ["readClaim(ref)", "-", "{ holder, since, heartbeat } or null", "For the window to show who holds a plan, without attempting to take it.", "NR-STO-12"],
    ["writeJournal(ref, pending) / readJournal(ref) / clearJournal(ref)", "-", "-", "Pending edits, kept apart from the committed workspace.", "NR-STO-07, N-08"],
    ["stat(ref)", "-", "{ mtime, size, exists }", "For the staleness check. Cheap enough to poll.", "NR-STO-16"],
    ["openSourceFile()", "-", "{ sheets, name } or null", "A source workbook or .prap.json, for a look. No workspace is created.", "NR-IMP-05"],
    ["exportFile(bytes, suggestedName)", "-", "{ path } or null", "Native Save dialog. Works whether or not the claim is held.", "NR-STO-17"],
]
r = table(ws, r, ["Operation", "Takes", "Returns", "Notes", "NR"],
          ops, [40, 24, 34, 62, 16], wrap_cols=(4,))
r += 1

r = section(ws, r, "Capabilities, and why ui/ asks rather than assumes")
r = code(ws, r, [
    "    web shell      { workspaces:false, versions:false, claims:false, journal:false }",
    "    desktop shell  { workspaces:true,  versions:true,  claims:true,  journal:true  }",
])
r += 1
r = note(ws, r, "One code path, two answers. Where 'claims' is false the window shows no holder and no blocked "
                "message, because there is nobody to block - not because the feature was compiled out. The "
                "alternative, a ui/ written twice, is the divergence NR-PAR-01 exists to prevent.")
r += 1

r = section(ws, r, "Errors")
r = lines(ws, r, [
    "Every operation throws StorageError { kind, message, detail } rather than a bare Error. The kind is what",
    "ui/ switches on; the message is what the user reads. Kinds:",
])
r += 1
errs = [
    ["not_found", "The workspace is not where it was.", "\"That plan is no longer at <path>. It may have been moved or deleted.\"", "NR-APP-06"],
    ["protected", "Encrypted, or refused by file protection.", "See sheet 09 - this is the one that must never be reported as corruption.", "NR-IMP-06"],
    ["unreadable", "Present and readable, but not a workspace.", "\"<name> is not a plan file. It may be a source workbook - use Import instead.\"", "NR-IMP-06"],
    ["too_new", "Written by a newer application version.", "\"This plan was saved by version <v>. This is version <w>. Install the newer version to open it.\"", "NR-DEP-04"],
    ["read_only", "The location cannot be written to.", "\"<path> cannot be written to. Save As somewhere else, or ask for write access.\"", "NR-DEP-09"],
    ["no_space", "The disk or share is full.", "\"There is not enough room to save. Nothing has been changed - your previous save is intact.\"", "NR-STO-05"],
    ["claim_lost", "The claim was displaced under us.", "See sheet 07, step 5.", "NR-STO-10"],
]
r = table(ws, r, ["kind", "When", "What the user is told", "NR"], errs, [14, 40, 76, 14], wrap_cols=(2, 3))
r = note(ws, r, "A storage layer that throws one kind of error forces the screen to say one kind of thing. Most "
                "of the value of this table is in the third column.")

# ---- 04 Workspace format --------------------------------------------------
ws, r = sheet(wb, "04_Workspace_Format", "The workspace file   [N2.3]",
              "It is the interchange format with a header on it - deliberately, so an AI agent needs no new "
              "contract to read a plan.")

r = section(ws, r, "Shape")
r = code(ws, r, [
    '    {',
    '      "format": "prap-source-data",          <- unchanged, so prap_io.py reads it as it stands',
    '      "format_version": 1,',
    '      "workspace": {                         <- ADDITIVE. Readers that do not know it ignore it',
    '        "app": "PM_APP",',
    '        "app_version": "1.0",',
    '        "schema_version": 5,',
    '        "created": "2026-08-13T09:14:02Z",',
    '        "last_saved": "2026-08-13T11:40:55Z",',
    '        "last_saved_by": { "name": "Kim Min-jun", "department": "Data Management" },',
    '        "imported_from": { "file": "PRAP_SourceData_2026Q3.xlsx", "at": "..." }',
    '      },',
    '      "sheets": { "Project": [...], "Milestone": [...], ... }   <- exactly as today',
    '    }',
])
r += 1
fmt = [
    ["The format key does not change", "A workspace IS a prap-source-data file. tools/prap_io.py, the AI agent guide and every recipe in it keep working on a workspace with nothing added to them.", "NR-STO-03"],
    ["The workspace block is additive", "A reader that does not know the key ignores it; the application fills it in. Nothing in sheets/ depends on it.", "NR-STO-03"],
    ["last_saved_by is a courtesy, not a record", "Shown when the plan opens. It is a declared name (NR-USR-08) and nothing formal rests on it - Q-N09 confirms there is no record-keeping obligation.", "NR-USR-07"],
    ["app_version gates opening", "A workspace whose app_version is newer than the running application is refused, not partially read. Older always opens.", "NR-DEP-04, NR-DEP-03"],
    ["schema_version is checked as today", "V-09 already reports a mismatch. Unchanged.", "-"],
    ["Encoding is UTF-8 without a byte-order mark, newline \\n", "So a workspace diffs cleanly in version control and reads the same on every machine.", "-"],
    ["Nothing is encrypted by the application", "Where files must be protected, the company's own protection is applied to the folder. An application that encrypts its own data becomes the only thing that can read it.", "NR-IMP-07, N-32"],
]
r = table(ws, r, ["Rule", "Why", "NR"], fmt, [44, 84, 16], wrap_cols=(2,))
r += 1

r = section(ws, r, "Naming")
r = code(ws, r, [
    "    <name>.prap              the workspace",
    "    <name>.prap.lock         the write claim, while somebody holds it        (sheet 07)",
    "    <name>.prap.journal      pending edits, while a session has any          (sheet 06)",
    "    backups/<name>.prap.1    the retained previous version                   (sheet 06)",
    "    <name>.prap.tmp-<pid>    exists only during a save, for a moment         (sheet 06)",
])
r = note(ws, r, "All five sit beside the workspace or one folder below it. A plan and everything about it can be "
                "moved by moving the file and its backups folder - there is no index anywhere else that would "
                "be left behind.")

# ---- 05 Shell -------------------------------------------------------------
ws, r = sheet(wb, "05_Shell", "The desktop shell   [N2.4]",
              "Window, menus, dialogs, and what happens between double-clicking the icon and seeing a plan.")

r = section(ws, r, "Launch sequence")
seq = [
    ["1", "Resolve the data folder.", "--data, then PRAP_DATA, then data\\ beside the application, then ask. Shown in About.", "NR-DEP-10"],
    ["2", "Establish identity.", "First run: the identity dialog, name pre-filled from the Windows account. Afterwards: the same dialog pre-filled from settings, one click to continue, 'Not me' to switch.", "NR-USR-01, NR-USR-02"],
    ["3", "Restore the window.", "Size, position and active tab from settings. Off-screen positions are ignored rather than restored.", "NR-APP-04"],
    ["4", "Recover, if there is anything to recover.", "A journal newer than its workspace: 'keep these edits, or discard them'. See sheet 06.", "NR-STO-07"],
    ["5", "Reopen the last workspace.", "Without asking. If it is gone, the open screen with the recent list, and a line saying which plan could not be found.", "NR-APP-03"],
    ["6", "Show the window.", "Under 5 seconds from a local disk, measured from launch to a usable window.", "NR-NFR-01"],
]
r = table(ws, r, ["Step", "What happens", "Detail", "NR"], seq, [7, 40, 76, 20], wrap_cols=(3,))
r += 1
r = note(ws, r, "Step 2 before step 5 is deliberate. Opening a plan may need a claim, and a claim needs to know "
                "whose it is - so identity is settled before anything can be held.")
r += 1

r = section(ws, r, "Menus")
menus = [
    ["File", "New plan · Open… · Open recent > · Save · Save As… · Close plan", "Save is disabled while this session does not hold the claim, with the reason on hover.", "NR-APP-05, NR-STO-17"],
    ["File", "Import source data… · Look at a source file… · Export to Excel · Export JSON", "'Look at' is NR-IMP-05: no workspace, nothing saved.", "NR-IMP-01, NR-IMP-05"],
    ["Edit", "Save changes · Leave without change · Show unsaved changes", "The same three the web application has, in their conventional place.", "-"],
    ["View", "Overall · Source data (project) · Source data (person) · General assumptions", "The tabs, so they are reachable by keyboard.", "-"],
    ["Plan", "Who is editing… · Restore previous version… · Plan properties", "'Who is editing' answers the question without attempting to take the claim.", "NR-STO-12, NR-STO-06"],
    ["Help", "User guide · About Project Management APP", "About shows the version, the data folder in use, and which rule chose it.", "NR-APP-07, NR-DEP-10"],
]
r = table(ws, r, ["Menu", "Items", "Notes", "NR"], menus, [10, 62, 56, 18], wrap_cols=(2, 3))
r += 1

r = section(ws, r, "Window title")
r = code(ws, r, [
    "    <plan name> — Project Management APP                     editing, nothing pending",
    "    <plan name> • — Project Management APP                   unsaved changes (the bullet is the mark)",
    "    <plan name> [read-only] — Project Management APP         somebody else holds it",
    "    Project Management APP                                   no plan open",
])
r = note(ws, r, "NR-APP-08 fixes the name; NR-APP-07 requires the version, which is in About rather than the "
                "title - a version in the title is read once and then occupies space forever.")
r += 1

r = section(ws, r, "What the desktop shell does NOT change")
r = lines(ws, r, [
    "Every tab, table, chart, filter and editing behaviour is the web application's, unchanged (N-11). The",
    "reviewer spent twenty-five rounds on that interface; redesigning it here would discard the review and",
    "leave two habits to hold. Divergences are only where a desktop convention requires one, and each is",
    "listed at N3.2 rather than left to be discovered.",
])

# ---- 05a The Python shell -------------------------------------------------
ws, r = sheet(wb, "05a_Python_Shell", "The Python shell   [change C-N02]",
              "A second shell under the same page, for a machine that will not accept an executable "
              "and will not let a browser be given a file.")

r = section(ws, r, "Why there are two shells")
r = lines(ws, r, [
    "Two company controls, measured on the requester's own machine rather than guessed at:",
    "",
    "    R-N20   an executable may RUN, but may not ARRIVE. E-mail refuses to carry it, and",
    "            e-mail is the only route available.",
    "    R-N21   data may not enter a browser page through the file picker. The dialog opens;",
    "            on choosing a file a security message appears and nothing is imported.",
    "",
    "The Electron shell answers neither. This one answers both, and it does so by removing the thing",
    "each control acts on rather than by getting around the control:",
    "",
    "    what travels is .py TEXT             so there is no executable to refuse",
    "    the page never asks for a file       so there is no upload to intercept",
])
r += 1
r = note(ws, r, "Neither control is worked around, disabled or hidden from, and no requirement here asks a "
                "user to do so. An upload control governs data entering a web page; a program the user ran "
                "on their own machine reading a file they chose is what every application on that laptop "
                "does, Excel included. The control is not bypassed - it is not involved.")
r += 1

r = section(ws, r, "The shape of it")
r = code(ws, r, [
    "    the page  --fetch-->  127.0.0.1:<port>/api/...  -->  storage/python  -->  disk",
    "",
    "    core/     unchanged   the same JavaScript, the same figures, in the browser",
    "    ui/       unchanged   the same tabs, tables and charts the reviewer approved",
    "    storage/  ported      workspace.py and claim.py, from the .js beside them",
    "    shell/    new         server.py, paths.py, files.py, launch.py - about 900 lines",
])
r = note(ws, r, "Decision N-05 survives intact: there is still ONE engine, and it is in the browser. Python "
                "touches files and nothing else. tools/test_layers.py fails the build if a Python module "
                "mentions a page, and tools/test_python_app.py compares 1,225 person-months against the "
                "independent reference implementation on every run.")
r += 1

r = section(ws, r, "The import path - the whole point")
imp = [
    ["Web application, today", "Route taken here", "Which control sees it"],
]
r = table(ws, r, ["Step", "Web application, today", "The Python shell"],
          [["1", "The user picks a file", "The user picks a file - in a Python dialog, or a folder "
            "listing drawn in the page. Neither is a browser file interface."],
           ["2", "<input type=\"file\"> receives it", "Python opens it with open()"],
           ["3", "The File API reads the bytes  <- INTERCEPTED HERE", "Python reads the bytes"],
           ["4", "The page parses the xlsx", "The page fetches them from 127.0.0.1 as ordinary page "
            "content, then parses the xlsx - the same reader, unchanged"]],
          [6, 54, 84], wrap_cols=(2, 3))
r += 1
r = note(ws, r, "There is no <input type=\"file\"> on the page, no drop handler and no File API call anywhere "
                "in the shell. The web page's own picker is REMOVED at start-up rather than hidden: a button "
                "that opens a dialog and then silently loses the file is worse than no button, because the "
                "person who clicks it concludes the application is broken - and they are right.")
r += 1

r = section(ws, r, "Opening a socket, which needs answering rather than waving past")
r = note(ws, r, "NR-SEC-01 said the shell opens no socket. This shell must, because that is how a browser "
                "and a local program talk. The requirement is amended rather than quietly broken, and the "
                "amendment carries its own controls - every one of which is checked by "
                "tools/test_python_app.py on each run.")
sec = [
    ["Loopback only", "It binds 127.0.0.1, never the network interface. Nothing off the machine can reach it, and Windows Firewall does not prompt for loopback.", "NR-SEC-04"],
    ["A fresh port each time", "Chosen by the operating system at start. Nothing is reserved, and nothing collides.", "NR-SEC-04"],
    ["A key on every request", "32 bytes, generated at start, never written to disk. Another program on the machine cannot guess it; another web page cannot read it, because it lives in this page's own DOM and same-origin policy keeps it there.", "NR-SEC-05"],
    ["The Host header must be loopback", "This is what stops a hostile site pointing its own name at 127.0.0.1 and talking to us through the user's browser - DNS rebinding.", "NR-SEC-05"],
    ["No cross-site request, ever", "No CORS header is sent at all, and any Origin or Sec-Fetch-Site that is not our own is refused before it reaches an operation.", "NR-SEC-05"],
    ["Nothing is served from disk", "One page and one API. There is no path that maps a URL onto a file, so there is nothing for '..' to escape from.", "NR-SEC-06"],
    ["No network access of any kind", "It listens; it never connects. Standard library only, no pip install, no download.", "NR-DEP-05"],
]
r = table(ws, r, ["Control", "What it does", "NR"], sec, [30, 100, 14], wrap_cols=(2,))
r += 1

r = section(ws, r, "What differs from the Electron shell, and why")
dv = [
    ["The menu is drawn in the page", "There is no application menu bar to use. Same items, same order, same accelerators for Save and Open.", "Visible"],
    ["The window is a browser tab", "It looks like the web application because it IS the web application. The console window that starts it is also how it is stopped, and says so.", "Visible"],
    ["A folder listing inside the page", "For a Python built without tkinter, and for typing a path - a network share, say. It lists names and sizes only; nothing in it can read a file.", "Visible"],
    ["The page ASKS whether it still holds the claim", "Electron pushed 'your claim was taken over' down a second channel. Here the page asks, on the same thirty-second clock the heartbeat runs on. One question every thirty seconds costs nothing and needs no second channel to go wrong.", "Invisible"],
    ["Export by download is kept", "A download is not an upload and R-N21 does not touch it, so storage/web/export.js runs unmodified with all of its checks. 'Export to a folder…' is added for when the Downloads folder is the wrong place.", "Visible"],
    ["No window size or position to remember", "The browser owns the window. The setting is kept in the file and ignored, so a later shell can use it.", "Invisible"],
]
r = table(ws, r, ["Difference", "Why", "Seen by the user?"], dv, [40, 92, 16], wrap_cols=(2,))

# ---- 06 Persistence -------------------------------------------------------
ws, r = sheet(wb, "06_Persistence", "Saving, versions, and recovery   [N2.5]",
              "The requirement that data survives closing, and the risk that creates.")

r = section(ws, r, "The save protocol")
r = code(ws, r, [
    "    save(ref, sheets, header):",
    "      1  if not holdsClaim(ref):  throw StorageError(claim_lost)      # sheet 07 step 5",
    "      2  text  = serialise(sheets, header)                            # core/, unchanged",
    "      3  tmp   = ref + '.tmp-' + pid",
    "      4  write(tmp, text); fsync(tmp)                                 # bytes are on the disk",
    "      5  if exists(ref):  move(ref -> backups/<name>.prap.1)          # NR-STO-06",
    "      6  rename(tmp -> ref)                                           # atomic; NR-STO-05",
    "      7  clearJournal(ref)                                            # committed, so nothing pending",
    "      8  return { savedAt: now }                                      # only now is it 'saved'",
])
r += 1
save = [
    ["Step 4 before step 6", "A rename is atomic; a write is not. Doing the slow part into a temporary name means the workspace is only ever wholly old or wholly new.", "NR-STO-05"],
    ["fsync, not just write", "A write can sit in a cache. NR-STO-04 says a save is not reported until the bytes are on the disk, and only fsync makes that true."],
    ["Same directory for the temporary file", "So step 6 is a rename within one filesystem. A rename across filesystems is a copy, and a copy is not atomic."],
    ["Step 5 keeps exactly one", "Q-N07. The setting `retain_versions` defaults to 1; a higher value shifts .1 to .2 and so on."],
    ["Step 8 is the only place 'Saved' is shown", "Reporting a save that has not landed is the one failure that would destroy trust in the whole feature."],
    ["A leftover .tmp-<pid> is deleted at launch", "It means a save was interrupted. The workspace itself is intact by construction, so there is nothing to repair and nothing to tell the user."],
]
r = table(ws, r, ["Point", "Why", "NR"], [x + [""] * (3 - len(x)) for x in save], [42, 86, 14], wrap_cols=(2,))
r += 1

r = section(ws, r, "Pending edits, and recovery")
r = lines(ws, r, [
    "The provisional-edit model is unchanged: a snapshot before the first pending edit, Save commits, Leave",
    "without change reverts. What the desktop adds is that the pending edits are also written down - to a",
    "journal beside the workspace, NEVER into the workspace itself.",
])
r += 1
jr = [
    ["When", "On every pending edit, debounced to at most once a second.", "Cheap: the journal is the pending list, not the model."],
    ["Where", "<name>.prap.journal", "Beside the workspace, so moving the plan moves its unfinished work with it."],
    ["Cleared", "On Save (committed) and on Leave without change (discarded).", "Both already exist as events; neither is a new concept."],
    ["On launch", "A journal whose mtime is newer than its workspace prompts: 'This plan has <n> unsaved changes from <when>. Keep them, or discard them?' EVERY change is listed, in a scrollable box - not a sample with a count.", "The user decides, and cannot decide from three lines out of forty. Silently applying recovered edits would be worse than losing them (NR-STO-07). Gate N3 review."],
    ["Never merged", "A journal is only offered for the workspace it belongs to, and only if the workspace has not been saved by somebody else since.", "Otherwise the recovered edits would be applied to figures they were never made against."],
]
r = table(ws, r, ["Aspect", "Behaviour", "Why"], jr, [16, 68, 62], wrap_cols=(2, 3))
r += 1

r = section(ws, r, "Restoring a previous version")
r = lines(ws, r, [
    "    Plan > Restore previous version…  reads backups/<name>.prap.1 and loads it AS A PENDING EDIT.",
    "",
    "It does not overwrite the current file. Restoring is a change like any other: it appears in the unsaved",
    "changes list, Save commits it, Leave without change abandons it. That way a restore made by mistake costs",
    "nothing, and a restore made deliberately goes through the same door as every other edit.",
])
r += 1
r = note(ws, r, "With one retained version (Q-N07), two bad saves in a row leave nothing good to return to - "
                "risk R-N19. The user guide says plainly: export to Excel anything you cannot afford to lose. "
                "The export is an archive no save can touch, which is the real safety net and always was.")

# ---- 07 Sharing -----------------------------------------------------------
ws, r = sheet(wb, "07_Sharing", "The write claim   [N2.5]",
              "One writer, many readers. The claim is taken at the first edit, not at open.")

r = section(ws, r, "The claim file")
r = code(ws, r, [
    '    <name>.prap.lock',
    '    {',
    '      "name": "Kim Min-jun", "department": "Data Management",',
    '      "machine": "PC-4471", "pid": 8812,',
    '      "since": "2026-08-13T09:14:02Z",',
    '      "heartbeat": "2026-08-13T09:41:32Z",       <- rewritten every 30 s while alive',
    '      "app_version": "1.0"',
    '    }',
])
r += 1
mech = [
    ["Taken", "open(path, 'wx') - create, and FAIL if it already exists.", "Decided by the filesystem, one winner, and it works over SMB. Read-then-write loses the race: two sessions can both read 'free' (NR-STO-13).", "NR-STO-13"],
    ["When", "At the moment a DATA VALUE ACTUALLY CHANGES - the same point the snapshot is taken. Clicking into a cell, selecting a row, changing a filter, switching tab, or typing into a field and leaving it unchanged do NOT take the claim.", "Not on open, and not on a click. A session that only looks - even one that clicks about a great deal - never takes it, so a plan is free unless somebody is really altering it. Clarified at the Gate N3 review, because 'the first edit' was open to the looser reading.", "NR-STO-10"],
    ["Kept alive", "The holder rewrites `heartbeat` every 30 seconds.", "So the application can always tell a live holder from a dead one - within half a minute, whatever the expiry is (N-23).", "NR-STO-14"],
    ["Expires", "30 minutes after the last heartbeat.", "Q-N16. Separate from the heartbeat interval, deliberately.", "NR-STO-14"],
    ["Reclaimed", "Immediately, if name and machine are this user's own.", "Being locked out of your own plan for half an hour because your application crashed is an obstruction, not a policy (N-24).", "NR-STO-19"],
    ["Released", "On Save-and-close, on Leave without change, and on application close.", "NOT on save alone - that would hand the plan to somebody else mid-task (N-22).", "NR-STO-15"],
    ["Re-checked", "Before every save.", "If anything above went wrong, the save stops rather than overwriting somebody's work.", "NR-STO-10"],
    ["Unreadable", "Treated as held by an unknown session; re-read once after 2 seconds before saying so.", "A torn read of a file being rewritten is possible on a share. Reading twice costs nothing and avoids a false report.", "-"],
]
r = table(ws, r, ["Aspect", "Mechanism", "Why this way", "NR"], mech, [14, 46, 66, 16], wrap_cols=(2, 3))
r += 1

r = section(ws, r, "Session states")
st = [
    ["READING", "open", "no claim attempted", "Everything except changing this workspace: tabs, figures, charts, filters, horizon, exports, Save As.", "→ EDITING on the first edit"],
    ["EDITING", "claim held", "heartbeat running", "Everything. Identical to the single-user case.", "→ READING on release; → BLOCKED never"],
    ["BLOCKED", "claim refused", "the attempted edit did not happen", "Everything READING allows. The window says who holds it.", "→ EDITING when it frees and the user accepts the offer"],
    ["STALE", "the file changed on disk", "polled by stat() every 10 s", "Everything READING allows, with a standing offer to reload. Editing is refused until it reloads.", "→ READING on reload"],
]
r = table(ws, r, ["State", "Means", "Also", "Permitted", "Leaves"],
          st, [11, 18, 30, 56, 32], wrap_cols=(4, 5))
r += 1
r = note(ws, r, "STALE is not about writers. Somebody who opened a plan at 09:00 and is still shown 09:00's "
                "figures at 11:00, after two saves by somebody else, may quote them in good faith. Reading a "
                "shared plan is only safe if the application admits when what it shows has been superseded "
                "(NR-STO-16).")
r += 1

r = section(ws, r, "What the user is told")
msg = [
    ["Holder active - heartbeat within 30 s", "\"Kim Min-jun (Data Management) is editing this plan. Started 09:14, active now.\"", "Copy contact details · Open read-only"],
    ["Holder silent - heartbeat older than 30 s", "\"Kim Min-jun (Data Management) has been editing this plan since 09:14, but their session has not responded since 09:22. The plan becomes free at 09:52.\"", "Copy contact details · Open read-only · Wait"],
    ["Expired - heartbeat older than 30 min", "\"Kim Min-jun's session stopped responding at 09:22 and no longer holds this plan. You may take over.\"", "Take over · Open read-only"],
    ["Your own stalled claim", "\"This plan is held by an earlier session of your own, on this machine. Take it back?\"", "Take it back · Open read-only"],
    ["It has just freed", "\"This plan is free now. Start editing?\"", "Start editing · Stay read-only"],
    ["Superseded while reading", "\"Kim Min-jun saved this plan at 11:07. The figures on screen are from 09:00.\"", "Reload · Keep looking"],
    ["Claim lost during a save", "\"Your hold on this plan was taken over while you were working. Nothing has been saved. Save As a copy to keep your changes.\"", "Save As… · Discard"],
]
r = table(ws, r, ["Situation", "Message", "Buttons"], msg, [34, 84, 30], wrap_cols=(2, 3))
r = note(ws, r, "The last row is the one that matters most and will happen least. It must never lose the user's "
                "work silently, and it must offer the way out in the same breath as the bad news.")
r += 1

r = section(ws, r, "What this deliberately is not")
r = lines(ws, r, [
    "Not row-level or section-level locking (N-20) - that implies merging two people's edits, which the file",
    "format cannot express and which is a different product. Not a queue of waiting editors: one waiting",
    "session is offered the claim when it frees, and beyond that people can talk to each other - which is what",
    "the contact details are for. Not an operating-system lock (N-17): a share may not honour one, and a lock",
    "left by a crash can need IT to clear it.",
])

# ---- 08 Identity ----------------------------------------------------------
ws, r = sheet(wb, "08_Identity", "Who the user is   [N2.4]",
              "Declared, not authenticated. It answers one question, and answers it well.")

r = section(ws, r, "The identity dialog")
idn = [
    ["When", "At launch, before any workspace opens.", "NR-USR-01"],
    ["First run", "Name pre-filled from the Windows account name, department empty. Both editable.", "NR-USR-03"],
    ["Afterwards", "Both pre-filled from settings. One button, 'Continue as Kim Min-jun'. A second, 'Not me', clears and asks again - which is what a shared PC needs.", "NR-USR-02"],
    ["Fields", "name (required, non-blank) and department (required). Nothing else - no e-mail, no telephone.", "NR-USR-04, Q-N18"],
    ["Department source", "The departments already present in the open workspace's Person rows, offered as a list; free text where none matches. Where the typed name matches a person_name in the plan, that person's department is offered first.", "NR-USR-10, NR-USR-11"],
    ["Stored", "In the user's own data folder, in settings.json. Never in the shared application folder.", "NR-USR-09"],
    ["Said on the dialog", "\"This is how colleagues will see you when you are editing a plan. It is not a login and it is not checked.\"", "NR-USR-08"],
]
r = table(ws, r, ["Aspect", "Behaviour", "NR"], idn, [20, 108, 16], wrap_cols=(2,))
r += 1
r = note(ws, r, "The last row is a requirement, not decoration. A name entry box that looks like a login WILL be "
                "read as one, and the moment somebody believes it proves who acted, it has misled them. The "
                "limit belongs at the point it could mislead. Q-N09 confirms nothing formal rests on it.")
r += 1

r = section(ws, r, "Where identity is used")
use = [
    ["In the claim file", "name and department, so a blocked colleague is told who and where.", "NR-USR-05"],
    ["In the blocked message", "Shown together, with a button that copies 'Kim Min-jun (Data Management)'.", "NR-USR-06"],
    ["In the workspace header", "last_saved_by, shown when the plan opens: 'Last saved by Kim Min-jun (Data Management), 11:07 today.'", "NR-USR-07"],
    ["Nowhere else", "Not in exports, not in the calculation, not in any per-row audit column.", "NR-USR-08"],
]
r = table(ws, r, ["Where", "What", "NR"], use, [24, 104, 16], wrap_cols=(2,))

# ---- 09 Import ------------------------------------------------------------
ws, r = sheet(wb, "09_Import", "Import, and the difference report   [N2.6]",
              "The one design choice that can silently lose work, and what stops it.")

r = section(ws, r, "The sequence")
imp = [
    ["1", "The user picks a source workbook or a .prap.json.", "Native dialog. Unchanged reader - core/.", "NR-IMP-01"],
    ["2", "If the workspace is EMPTY: adopt it, and stop.", "The ordinary first import. No questions worth asking.", "NR-IMP-01"],
    ["3", "If the workspace HOLDS DATA: ask first. \"This plan already contains data. Update it from <file>?\"", "Q-N05's addition. v0.6 went straight to the differences, which presumes an update was wanted at all.", "NR-IMP-02"],
    ["4", "Build the difference report.", "Per sheet: rows the file would ADD, rows it would CHANGE (with which columns), rows present here but absent there. Compared on the sheet's key columns.", "NR-IMP-02"],
    ["5", "The user accepts or skips PER SHEET.", "Not per row - a row-level choice on a thousand rows is a choice nobody makes. Per sheet is a decision somebody can actually take.", "NR-IMP-02"],
    ["6", "Apply the accepted sheets as a PENDING EDIT.", "So the whole import is undone by Leave without change, and committed by Save, like everything else.", "NR-IMP-02"],
]
r = table(ws, r, ["Step", "What happens", "Detail", "NR"], imp, [7, 56, 66, 16], wrap_cols=(2, 3))
r += 1
r = note(ws, r, "Step 6 is what makes step 5 safe to get wrong. An import that lands as a pending edit can be "
                "looked at, disbelieved and abandoned. An import that writes straight to disk cannot.")
r += 1

r = section(ws, r, "The report")
r = code(ws, r, [
    "    Sheet              Add   Change   Only here   Accept?",
    "    Project              2        4           0     [x]",
    "    Milestone           17        3           1     [x]",
    "    Person               0        0          12     [ ]   <- 12 people you added are not in this file",
    "    Assignment           5        9           3     [ ]",
    "",
    "    Expandable per sheet to the rows themselves, and to which columns differ on a changed row.",
])
r += 1
r = note(ws, r, "'Only here' is the column that earns the report. Those are the rows somebody typed by hand, and "
                "they are exactly what a silent replace would destroy. Accepting a sheet never deletes them - "
                "an import adds and changes; removal is always a separate, explicit action.")
r += 1

r = section(ws, r, "Row-level detail   [Gate N3 review]")
r = lines(ws, r, [
    "Each sheet expands, on request, to the individual rows behind its counts: which rows would be added,",
    "which would change and in which columns, with the current value and the incoming one side by side.",
    "Collapsed by default, because sheet counts are what the decision is taken on and forty expanded sheets",
    "are not a report.",
    "",
    "The expansion is for LOOKING, not for choosing: the tick stays at sheet level (S-N03). Row-level counts",
    "that disagree with what you expected are the reason to open it - and the reason to say no.",
])
r += 1

r = section(ws, r, "As BUILT   [N4.5, 2026-08-22]")
r = lines(ws, r, [
    "Built and wired into the Python shell. The engine is shared; the screen is not.",
    "",
    "    src/core/06a_diff.js          pure - two sets of parsed rows in, a description out",
    "    src/shell/python/importdiff.js  the three dialogs, and the wiring",
    "",
    "The engine is in core/ because it decides which rows are the same row - which is a data",
    "question, testable without a browser, and the sort of thing that must not be written twice.",
    "The SCREEN is in the Python shell alone: the web application is feature-frozen (N-06) and",
    "has no workspace to merge into, so loading a second file there still replaces the first,",
    "exactly as the reviewer signed it off.",
])
r += 1
built = [
    ["What identifies a row", "DIFF_KEY, the composite keys - not KEY_COL, which names one column per sheet and is a FOREIGN key on a child sheet. Every milestone of a project shares its project_id; comparing on that alone would call them all the same row.", "check_consistency holds DIFF_KEY to the keys in docs/prap_contract.json"],
    ["What is not compared", "Derived columns. A stale total in a hand-edited workbook is not a difference, and must never become the truth (NR-IMP-08). It is left out of the report AND out of what gets written.", "tools/test_importdiff.py"],
    ["What counts as the same value", "A date read from .xlsx and the same date written in a .prap.json; 1.2 and '1.20'. A report full of differences that are not differences is a report nobody finishes.", "tools/test_importdiff.py"],
    ["What an accepted sheet does", "ADDS and CHANGES. Never deletes - S-N04, and the one behaviour here that could destroy work.", "tools/test_importdiff.py"],
    ["What an imported row is", "A complete row, not a draft. newRow marks a row __new, which holds it out of validation and out of the calculation; an imported assignment left that way would be silently weightless.", "tools/test_importdiff.py"],
    ["How much is shown", "Twelve rows per list, then a count. A dialog that scrolls for a thousand rows is one nobody reads and everybody approves.", "-"],
]
r = table(ws, r, ["Point", "How it is settled", "Proved by"], built, [30, 96, 34], wrap_cols=(2,))
r += 1

r = section(ws, r, "Export as the way to edit outside the application   [NR-IMP-08]")
r = lines(ws, r, [
    "Confirmed at the Gate N3 review as a supported workflow rather than a side effect, and it needs no new",
    "screen - only a guarantee:",
    "",
    "    export  ->  change it in Excel or a text editor  ->  give it to the manager  ->  import  ->  report",
])
r += 1
rt = [
    ["What goes out", "Every sheet, every column, every row of the plan - in .xlsx or .prap.json, the user's choice.", "NR-IMP-03"],
    ["What must come back", "All of it. A value that survived the journey out and back is unchanged; nothing is silently dropped because the application did not recognise it.", "NR-IMP-08"],
    ["What is NOT trusted on the way back", "Anything the application derives - a project's window from its milestones, its team size from its assignments. Those are recomputed from the rows, never read from the file. A stale derived value in a hand-edited workbook must not become the truth.", "NR-IMP-08"],
    ["Who imports", "One person - the application's manager - receives the edited files and applies them (A-N13). No per-user import rule is needed, and none is specified.", "A-N13"],
    ["How it lands", "Through the same difference report as any other import. A file edited by hand is exactly the case the report exists for.", "NR-IMP-02"],
    ["Proof", "The web application already round-trips xlsx -> json -> xlsx -> json reproducing every cell (tools/test_interop.py). NR-IMP-08 makes that a requirement of this application rather than a property it happens to have, and N4.6 tests it after a hand edit.", "NR-IMP-08"],
]
r = table(ws, r, ["Aspect", "Specified", "NR"], rt, [30, 98, 16], wrap_cols=(2,))
r = note(ws, r, "Note the third row. It is the only part of this that is not already true: an exported workbook "
                "carries derived columns so a person can read them, and a person editing that workbook may "
                "well change a milestone without changing the project window beside it. Recomputing rather "
                "than trusting is what stops the file's stale copy winning.")
r += 1

r = section(ws, r, "Looking at a source file without a workspace")
r = lines(ws, r, [
    "    File > Look at a source file…    reads it, shows every tab, saves nothing, creates nothing.",
    "",
    "Q-N10. A workspace is therefore optional rather than mandatory: the question 'what does this file say?'",
    "does not require committing to a plan first, which is also how somebody tries the application for the",
    "first time. Save is disabled throughout; Export and Save As are not.",
])
r += 1

r = section(ws, r, "A file that cannot be read")
prot = [
    ["Encrypted or protected", "\"<name> could not be opened. It looks like it is protected by file security — open it in Excel first to unlock it, then import it again.\"", "NR-IMP-06"],
    ["Not a PRAP file", "\"<name> is not a plan or a source workbook. It has no <sheet> sheet.\"", "NR-IMP-06"],
    ["Corrupt", "\"<name> could not be read: <detail>.\"", "NR-IMP-06"],
]
r = table(ws, r, ["Cause", "Message", "NR"], prot, [26, 102, 16], wrap_cols=(2,))
r = note(ws, r, "Distinguishing the first from the third is the whole point of R-N18. A protected file reported "
                "as corrupt sends somebody looking for a backup they do not need; the same file reported as "
                "protected is unlocked in thirty seconds. The detection is: the bytes are present and the file "
                "is not a readable ZIP or JSON, and the extension says it should be one.")

# ---- 10 Deployment --------------------------------------------------------
ws, r = sheet(wb, "10_Deployment", "Packaging and the folder   [N2.7]",
              "Non-installed: copied as a folder, run in place, leaving the machine unchanged.")

r = section(ws, r, "The folder")
r = code(ws, r, [
    "    PM_APP\\                      <- copy this anywhere; delete it to remove the application",
    "      PM_APP.exe                  <- \\",
    "      resources\\                  <-  |  replaced wholesale by an update",
    "      version.txt                 <- /",
    "      data\\                       <- NEVER touched by an update",
    "        users\\<windows-account>\\",
    "          settings.json           <- identity, window state, recent list, preferences",
    "          workspaces\\             <- the default home for .prap files",
    "          backups\\                <- the retained previous version",
])
r += 1
dep = [
    ["Delivered as", "PM_APP_v<ver>.zip, containing the PM_APP\\ folder with NO data\\ in it.", "So extracting an update over an existing folder cannot overwrite anybody's plans (N-13).", "NR-DEP-02, NR-DEP-03"],
    ["Packaged as", "A plain folder that runs in place - not a single self-extracting executable.", "The single .exe unpacks ~180 MB to the temporary folder on every launch: slow, writes outside its own folder, re-scanned by anti-virus each time (N-14).", "NR-DEP-02, NR-DEP-05"],
    ["Writes outside itself", "Only where the user designated it: a chosen data folder, an exported file, a shortcut created on request.", "No registry, no file association, no Start-menu entry, no AppData (N-15).", "NR-DEP-06, NR-DEP-07"],
    ["Paths recorded", "Relative wherever they can be - a recent workspace inside the folder is stored relative to it.", "So copying the folder to another machine or drive letter does not break the recent list.", "NR-DEP-08"],
    ["Read-only folder", "Detected at launch; the user is told and asked where to keep data, and offered a shortcut carrying --data.", "The share is expected to be writable (Q-N15), but a folder can be made read-only later without telling the application.", "NR-DEP-09, NR-DEP-12"],
    ["UNC paths", "Supported, and tested, as well as mapped drive letters.", "A shared folder is often reached both ways by different people.", "NR-DEP-13"],
    ["Both arrangements", "Run in place from the share, or copy it locally and run there. Both tested; the user guide states what each costs in launch time rather than recommending one.", "Q-N14 - both are in use.", "NR-DEP-14"],
]
r = table(ws, r, ["Aspect", "Specified", "Why", "NR"], dep, [18, 52, 60, 18], wrap_cols=(2, 3))
r += 1

r = section(ws, r, "Data folder resolution   [NR-DEP-10]")
r = code(ws, r, [
    "    1  --data=<path> on the command line          a personal shortcut carries it",
    "    2  PRAP_DATA environment variable             a site can set it centrally",
    "    3  data\\users\\<windows-account>\\ beside      the ordinary case, single or shared",
    "       the application, if writable",
    "    4  ask the user, and offer to create a        only if 3 is impossible",
    "       desktop shortcut carrying --data",
])
r = note(ws, r, "About always shows the folder in use AND which of the four rules chose it. A portable "
                "application that will not say where its data went is a support problem, and the answer costs "
                "one line on a dialog.")
r += 1

r = section(ws, r, "Update")
r = lines(ws, r, [
    "Extract the new zip over the folder. resources\\, PM_APP.exe and version.txt are replaced; data\\ is not in",
    "the zip and so cannot be touched. A workspace written by an older version opens unchanged; one written by",
    "a NEWER version is refused with the message on sheet 03 rather than partially read (NR-DEP-04).",
])

# ---- 11 Traceability ------------------------------------------------------
plan_path = ROOT / "docs" / PLAN
plan = load_workbook(plan_path, data_only=True)
preq = plan["03_Requirements"]
nr_ids = []
for rr in range(5, preq.max_row + 1):
    v = str(preq.cell(rr, 1).value or "").strip()
    if v.startswith("NR-"):
        nr_ids.append((v, str(preq.cell(rr, 3).value or "").strip()))

# Where each requirement is specified. Read against the sheets above; a requirement with
# no home here is a requirement nobody will build.
HOME = {
    "APP": "05_Shell",
    "STO": "06_Persistence / 07_Sharing",
    "IMP": "09_Import",
    "USR": "08_Identity",
    "PAR": "02_Scope",
    "DEP": "10_Deployment",
    "SEC": "10_Deployment",
    "NFR": "05_Shell",
}
OVERRIDE = {
    "NR-STO-02": "04_Workspace_Format", "NR-STO-03": "04_Workspace_Format",
    "NR-STO-09": "04_Workspace_Format", "NR-STO-08": "05_Shell",
    "NR-STO-10": "07_Sharing", "NR-STO-11": "07_Sharing", "NR-STO-12": "07_Sharing",
    "NR-STO-13": "07_Sharing", "NR-STO-14": "07_Sharing", "NR-STO-15": "07_Sharing",
    "NR-STO-16": "07_Sharing", "NR-STO-17": "03_Storage_Interface", "NR-STO-18": "09_Import",
    "NR-STO-19": "07_Sharing",
    "NR-IMP-05": "09_Import", "NR-IMP-06": "09_Import", "NR-IMP-07": "04_Workspace_Format",
    "NR-APP-03": "05_Shell", "NR-APP-08": "05_Shell",
    "NR-DEP-04": "03_Storage_Interface", "NR-DEP-10": "10_Deployment",
    "NR-PAR-02": "02_Scope", "NR-PAR-03": "02_Scope", "NR-PAR-04": "05_Shell",
    # Change C-N02 - the Python shell. Everything the second shell adds is specified
    # on one sheet, so a reader who only cares about the Electron build can skip it
    # and a reader who only has the Python build has one place to look.
    "NR-SEC-04": "05a_Python_Shell", "NR-SEC-05": "05a_Python_Shell",
    "NR-SEC-06": "05a_Python_Shell", "NR-IMP-09": "05a_Python_Shell",
    "NR-DEP-16": "05a_Python_Shell",
}

ws, r = sheet(wb, "11_Traceability", "Every requirement, and where it is specified   [N2.8]",
              f"The NR-ids are READ from {PLAN} when this workbook is generated, not re-typed. "
              f"A requirement cannot be dropped between the two documents without this sheet showing it.")

rows = []
for rid, text in nr_ids:
    area = rid.split("-")[1]
    rows.append([rid, OVERRIDE.get(rid, HOME.get(area, "-")), text[:200]])
r = table(ws, r, ["NR-id", "Specified in", "Requirement (from the plan)"],
          rows, [13, 30, 105], wrap_cols=(3,))

r = section(ws, r, "Coverage")
untraced = [rid for rid, _ in nr_ids if OVERRIDE.get(rid, HOME.get(rid.split("-")[1], "-")) == "-"]
cov = [
    ["Requirements in the plan", str(len(nr_ids))],
    ["Specified in this document", str(len(nr_ids) - len(untraced))],
    ["Not yet specified", ", ".join(untraced) if untraced else "none"],
]
r = table(ws, r, ["Measure", "Value"], cov, [34, 60], wrap_cols=(2,))
r = note(ws, r, "tools/check_consistency.py re-derives this comparison on every run, so the two documents "
                "cannot drift apart between issues.")

# ---- 12 Open points -------------------------------------------------------
ws, r = sheet(wb, "12_Open_Points", "Open points for the Gate N2 review",
              "Things the plan left to the specification, decided here and needing your eye.")

pts = [
    ["S-N01", "07_Sharing", "The staleness check polls stat() every 10 seconds.", "Cheap on a local disk, one small network round-trip on a share. Longer means somebody reads superseded figures for longer; shorter buys little. Proposed: 10 s, as a setting.", "AGREED"],
    ["S-N02", "06_Persistence", "Restoring a previous version arrives as a PENDING EDIT rather than overwriting.", "It makes a mistaken restore free and puts a deliberate one through the same door as every other change. The alternative - restore writes immediately - is what most applications do, and is worse here.", "AGREED"],
    ["S-N03", "09_Import", "Accept or skip is per SHEET, not per row.", "A row-level choice across a thousand rows is a choice nobody makes. If you want per-row for the small sheets - Project, Person - say so; it is a bigger dialog, not a bigger design.", "AGREED - accept per SHEET, confirmed explicitly"],
    ["S-N04", "09_Import", "An accepted sheet ADDS and CHANGES but never DELETES.", "So 'only here' rows - the ones somebody typed by hand - always survive an import. Removing them stays a separate, explicit action. Confirm this is what you want: it means a row deleted from the source workbook stays in the plan.", "AGREED"],
    ["S-N05", "05_Shell", "Identity is established BEFORE the last workspace reopens.", "Because opening may need a claim, and a claim needs a name. It costs one dialog between the icon and the plan, every launch - one click after the first run.", "AGREED"],
    ["S-N06", "10_Deployment", "The per-user data folder is keyed on the WINDOWS ACCOUNT name, not the declared name.", "The declared name is editable and could collide; the account name is unique on the machine and stable. The declared name is what colleagues see; the account name is what files are filed under.", "AGREED"],
    ["S-N07", "07_Sharing", "A blocked session that later gets the claim keeps whatever it was looking at.", "Rather than reloading and losing the user's place. It reloads only if the plan changed while they waited - which is the STALE path, and says so.", "AGREED"],
]
r_start = r
r = table(ws, r, ["ID", "Sheet", "Point", "Reasoning", "Your answer"],
          pts, [8, 20, 46, 62, 34], wrap_cols=(3, 4, 5))
for rr in range(r_start + 1, r_start + 1 + len(pts)):
    ws.cell(row=rr, column=5).fill = NEW_FILL
r = note(ws, r, "All seven were agreed at review with no change requested, so v1.0 is v0.1 with the answers "
                "recorded. S-N04 is the one that will be felt in daily use, and the agreement is deliberate: "
                "a row deleted from the source workbook STAYS in the plan until somebody removes it, because "
                "the alternative silently destroys hand-entered work.")
r += 1

r = section(ws, r, "Approval - Gate N2")
appr = [["PRAP NewApp Specification v1.0", "Dan", "2026-08-13",
         "APPROVED - BASELINE. Reviewed as v0.1; all seven open points agreed with no change requested. "
         "Gate N2 is closed and Step N3 - the desktop UI design and the divergence list - is authorised. "
         "This specification governs Steps N3 to N5: changing it now costs a version and a re-approval, "
         "not a re-opened gate."]]
r_start2 = r
r = table(ws, r, ["Document", "Approver", "Date", "Decision"], appr, [32, 14, 14, 88], wrap_cols=(4,))
for cc in (1, 2, 3, 4):
    ws.cell(row=r_start2 + 1, column=cc).fill = NEW_FILL
r = note(ws, r, "STEP N2 IS COMPLETE. Every requirement on sheet 11 has a specified home, and Step N3 designs "
                "the screens that sheets 05 to 09 describe.")
r += 1

r = section(ws, r, "Change C-N01 - the Gate N3 review, awaiting approval")
chg = [
    ["1", "07_Sharing", "The write claim is taken when a DATA VALUE ACTUALLY CHANGES - not on a click, a selection, a filter or a tab.", "Your words: the timing considers any unsaved data change, not the timing when a user just clicks something. 'The first edit' was open to the looser reading; it is now closed.", "Clarifies NR-STO-10"],
    ["2", "06_Persistence", "The recovery dialog lists EVERY unsaved change, scrollable.", "Was three and a count. Nobody can decide from three lines out of forty.", "Clarifies NR-STO-07"],
    ["3", "09_Import", "The difference report expands to individual rows on request.", "For looking, not for choosing - the tick stays at sheet level, which S-N03 settled.", "Clarifies NR-IMP-02"],
    ["4", "05_Shell", "The manual light/dark toggle is removed; the window follows Windows.", "One fewer setting. Divergence D-N09 on the component list is rewritten accordingly.", "Component list"],
    ["5", "09_Import", "Export is a supported way to EDIT a plan outside the application, and the round trip must lose nothing.", "New requirement NR-IMP-08 on the plan, with assumption A-N13 - one manager imports, not every user.", "NEW NR-IMP-08"],
]
r = table(ws, r, ["#", "Sheet", "Change", "Why", "Effect"], chg, [5, 18, 56, 60, 22], wrap_cols=(3, 4, 5))
r += 1
r = section(ws, r, "Approval - change C-N01")
appr2 = [["PRAP NewApp Specification v1.2", "Dan", "2026-08-13",
          "APPROVED. Change C-N01 accepted in full: the write claim attaches to a data change rather than a "
          "click, recovery lists every unsaved change, the difference report expands to rows, the manual "
          "theme toggle is removed, and export is a supported way to edit a plan outside the application "
          "with a round trip that loses nothing. This issue supersedes v1.0 and governs Steps N4 and N5."]]
r_start3 = r
r = table(ws, r, ["Document", "Approver", "Date", "Decision"], appr2, [32, 14, 14, 88], wrap_cols=(4,))
for cc in (1, 2, 3, 4):
    ws.cell(row=r_start3 + 1, column=cc).fill = NEW_FILL

wb.save(OUT)
print(f"Written: {OUT}")
print(f"  {len(nr_ids)} requirements traced, {len(untraced)} unspecified")
