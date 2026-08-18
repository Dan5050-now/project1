"""Generate the PRAP desktop application development plan workbook.

This is a SEPARATE plan from PRAP_Development_Plan_v*.xlsx. That one governs the
single-file web application (app/PRAP.html), which is finished and stays finished.
This one governs a second, independently packaged application built on the same
calculation engine.

The two plans are siblings, not successors: neither supersedes the other.

    python tools/build_newapp_plan.py

Output: docs/PRAP_NewApp_Development_Plan_v0.1.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.styles.borders import Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DOC_VERSION = "1.7"
DOC_STATUS = ("Baseline v1.0 + change C-N01. Gates N1-N3 closed; Step N4 in progress. A-N09 IS NOW "
              "EVIDENCE rather than an assumption: an unsigned executable ran on the requester's own "
              "machine on 2026-08-15.")
DOC_DATE = "2026-08-13"
OUT = Path(__file__).resolve().parents[1] / "docs" / f"PRAP_NewApp_Development_Plan_v{DOC_VERSION}.xlsx"

WEB_PLAN = "PRAP_Development_Plan_v2.26.xlsx"
WEB_APP = "app/PRAP.html (application v1.24)"

FONT = "Arial"
NAVY = "1F3864"
BLUE_HDR = "2F5597"
BAND = "F2F5FB"
YELLOW = "FFFF00"
GREEN = "C6E0B4"
ORANGE = "FCE4D6"
GREY = "808080"

TITLE_F = Font(name=FONT, size=16, bold=True, color=NAVY)
H1_F = Font(name=FONT, size=12, bold=True, color=NAVY)
HDR_F = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_F = Font(name=FONT, size=10)
BOLD_F = Font(name=FONT, size=10, bold=True)
NOTE_F = Font(name=FONT, size=9, italic=True, color=GREY)
MONO_F = Font(name="Consolas", size=10)

HDR_FILL = PatternFill("solid", fgColor=BLUE_HDR)
BAND_FILL = PatternFill("solid", fgColor=BAND)
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
NEW_FILL = PatternFill("solid", fgColor=GREEN)
CHG_FILL = PatternFill("solid", fgColor=ORANGE)

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(vertical="top", wrap_text=True)
WRAP_C = Alignment(vertical="top", wrap_text=True, horizontal="center")

MARK_NEW = "[NEW]"
MARK_CHG = "[CHANGED]"


def sheet(wb, name, title, subtitle=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_F
    row = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = NOTE_F
        row = 3
    ws.freeze_panes = f"A{row + 2}"
    return ws, row + 1


def table(ws, start_row, headers, rows, widths, wrap_cols=(), mark_col=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = HDR_F
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = WRAP_C
    ws.row_dimensions[start_row].height = 28

    for r, data in enumerate(rows, start=start_row + 1):
        data = list(data)
        fill = None
        if mark_col is not None and isinstance(data[mark_col - 1], str):
            v = data[mark_col - 1]
            if v.startswith(MARK_NEW):
                fill = NEW_FILL
                data[mark_col - 1] = v[len(MARK_NEW):].strip()
            elif v.startswith(MARK_CHG):
                fill = CHG_FILL
                data[mark_col - 1] = v[len(MARK_CHG):].strip()
        for i, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = BODY_F
            c.border = BOX
            c.alignment = WRAP if i in wrap_cols else Alignment(vertical="top")
            if fill is not None:
                c.fill = fill
            elif (r - start_row) % 2 == 0:
                c.fill = BAND_FILL

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return start_row + len(rows) + 2


def section(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = H1_F
    return row + 1


def note(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = NOTE_F
    return row + 1


def lines(ws, row, texts, mono=False):
    for t in texts:
        c = ws.cell(row=row, column=1, value=t)
        c.font = MONO_F if (mono and t.startswith(" ")) else BODY_F
        row += 1
    return row


def legend(ws, row):
    ws.cell(row=row, column=1, value="Legend").font = BOLD_F
    row += 1
    for fill, txt in (
        (NEW_FILL, "Green = new to the desktop application; nothing in the web application corresponds to it."),
        (CHG_FILL, "Orange = inherited from the web application but deliberately changed here."),
        (INPUT_FILL, "Yellow = needs your decision before the work can proceed."),
    ):
        c = ws.cell(row=row, column=1, value=txt)
        c.font = BODY_F
        c.fill = fill
        row += 1
    return row + 1


wb = Workbook()
wb.remove(wb.active)

# ---- 00 Cover -------------------------------------------------------------
ws = wb.create_sheet("00_Cover")
ws.sheet_view.showGridLines = False
ws["A1"] = "Project Resource Assignment Program (PRAP)"
ws["A1"].font = Font(name=FONT, size=20, bold=True, color=NAVY)
ws["A2"] = "Desktop Application - Development Plan"
ws["A2"].font = Font(name=FONT, size=14, color=NAVY)

cover = [
    ("Document ID", "PRAP-NAPP-PLAN-001"),
    ("Document type", "Development plan (Step 1 deliverable) for a SECOND application"),
    ("Version", f"v{DOC_VERSION}"),
    ("Status", DOC_STATUS),
    ("Issue date", DOC_DATE),
    ("Author", "Claude Code"),
    ("Reviewer", "Requester - six review rounds, all 18 questions answered"),
    ("Baseline", "v1.0, approved by direction 2026-08-13. Gate N1 closed; unchanged by this issue"),
    ("Application name",
     "'Project Management APP' on screen; PM_APP as the folder and executable name (Q-N11). The documents "
     "keep the PRAP_NewApp_ prefix so this line stays findable beside the web application's - say if you "
     "would rather they were renamed too."),
    ("Sharing model",
     "SINGLE WRITER, MANY READERS. Instructed 2026-08-13: while one person is editing a plan, no other "
     "session may update it. Everyone may open and read it at any time. The claim is made at the first "
     "EDIT, not at open - see sheet 05a, which is given over to this."),
    ("Fixed constraints",
     "WINDOWS ONLY, and NON-INSTALLED. Instructed 2026-08-13. The application is copied as a folder and "
     "run from it - no installer, no administrator rights, no registry, nothing written outside its own "
     "folder. See sheet 04 for what this changes and sheet 10 for what it does not fix."),
    ("Under consideration",
     "The application folder MAY sit on a SHARED NETWORK FOLDER (Q-N12, round 2). Planned for as though it "
     "will: several people running one copy is designed for on sheet 05, and it is what makes workspace "
     "locking (NR-STO-10) necessary rather than optional."),
    ("Relationship to the web application",
     "PARALLEL, NOT SUCCESSOR. The web application and its documents are complete and stay in "
     "service. This plan does not supersede, amend or retire any of them."),
    ("Governs", "A packaged desktop application built on the same calculation engine"),
    ("Does NOT govern", f"{WEB_APP}, which remains under {WEB_PLAN}"),
    ("Repository", "Dan5050-now/project1"),
    ("Branch", "claude/project-resource-assignment-app-1vjdzh"),
    ("Supersedes", "Nothing. v1.0 remains the baseline; this issue records progress against it"),
]
r = 4
for k, v in cover:
    ws.cell(row=r, column=1, value=k).font = BOLD_F
    c = ws.cell(row=r, column=2, value=v)
    c.font = BODY_F
    c.alignment = WRAP
    r += 1
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 110
for rr in range(4, r):
    ws.row_dimensions[rr].height = 30

r += 1
r = section(ws, r, "Why a second plan rather than a new version of the first")
r = lines(ws, r, [
    "The web application is finished, verified and about to be used. Folding a desktop rewrite into its plan would",
    "put a completed, approved document back into draft, and would make it impossible to tell which requirements",
    "the finished tool actually satisfies. Two plans keep both answers clean:",
    "",
    f"    {WEB_PLAN}          the web application - CLOSED at Gate 4, going to Gate 5",
    f"    PRAP_NewApp_Development_Plan_v{DOC_VERSION}.xlsx      the desktop application - OPEN at Step 1",
    "",
    "They share one calculation engine and one data contract, and sheet 02 states exactly how that sharing is",
    "kept honest. Everything else about them is independent: their own versions, their own gates, their own",
    "approvals, their own release cycles.",
], mono=True)
r += 1
r = note(ws, r, "File naming follows the repository convention already in use - no spaces, version suffix, one "
                "generator script per document. The generator for this plan is tools/build_newapp_plan.py.")

# ---- 01 Version history ---------------------------------------------------
ws, r = sheet(wb, "01_Version_History", "Version history",
              "This document's own line. It does not continue the web application plan's numbering.")

hist = [
    [f"{MARK_NEW}1.7", "2026-08-15", "Claude Code", "-",
     "THE BLOCKER MOVED. The probe proved the machine will RUN an unsigned executable; the attempt to send "
     "one proved the company will not let it ARRIVE - e-mail refused the .exe on a general security check. "
     "Risk R-N20 records it, and it is a different risk from R-N01: execution policy is satisfied, "
     "transport is not. It also puts decision N-01 back in question, since an Electron package is an "
     ".exe however it is delivered. Three routes are set out for the requester on sheet 10a, and the "
     "recommendation is the one that removes the executable from the problem entirely: keep the engine "
     "and the interface exactly as they are, and replace the Electron main process with a small Python "
     "one. No new engine, no second set of figures - the storage interface at specification sheet 03 is "
     "the seam that makes it a swap rather than a rewrite."],
    ["1.6", "2026-08-15", "Claude Code", "-",
     "THE LAUNCH PROBE RAN. Assumption A-N09 - that the requester can extract a zip into a folder of "
     "their own and run an unsigned executable from it - was the one thing in this plan that could have "
     "invalidated the whole approach, and it is now evidence: PM_APP_Probe.exe ran on BOOK-R8USOPHQ11 at "
     "09:14 on 2026-08-15, from C:\\Users\\del09\\..., and could write beside itself. R-N01 falls from "
     "Medium/HIGH to Low/HIGH; R-N10 falls to Low, measured rather than guessed - the real folder leaves "
     "153 characters of headroom under the Windows limit. What is NOT yet known is how anti-virus treats "
     "the full 142 MB package, so R-N09 stands unchanged and A-N09 is confirmed only for the half that "
     "was tested. Evidence archived at docs/review/PM_APP_probe_result_2026-08-15.txt."],
    ["1.5", DOC_DATE, "Claude Code", "-",
     "STEP N4 IN PROGRESS - a progress record, not a change. Tasks N4.1 to N4.4c are built: the desktop "
     "shell (window, menu, native dialogs, window state), the filesystem storage adapter with its atomic "
     "write protocol and retained version, the workspace lifecycle, journalling, and the write claim with "
     "its heartbeat, expiry and same-user reclaim. Two suites prove them - tools/test_storage.mjs (33 "
     "checks in plain Node, including twelve kills mid-write and an eight-process race for one claim) and "
     "tools/desktop_smoke.js (23 checks inside a running Electron application). Both found real defects "
     "before delivery; sheet 12 records them. N4.4d, N4.4e and N4.5 - the four session states on screen, "
     "the identity dialog and the difference report - are next."],
    ["1.4", DOC_DATE, "Claude Code", "APPROVED",
     "GATE N3 CLOSED and CHANGE C-N01 APPROVED, both on 2026-08-13. Content is v1.3 unchanged; this issue "
     "records the approvals and opens Step N4. The requirement register now stands at 70 and is the "
     "contract for the build; the component list v1.0 and specification v1.2 are its companions. Nothing "
     "further is asked of the reviewer until Gate N4, when the application is functionally complete."],
    ["1.3", DOC_DATE, "Claude Code", "Superseded",
     "GATE N3 REVIEW APPLIED, and it carries ONE CHANGE against the approved baseline rather than progress "
     "alone. Change C-N01: exporting is confirmed as a supported way to EDIT a plan - out to Excel or JSON, "
     "changed outside the application, and back in - so NR-IMP-08 is added, requiring the round trip to lose "
     "nothing. Assumption A-N13 records the scoping the reviewer supplied with it: bulk editing goes to one "
     "manager who imports, not to every user importing their own file. A change against an approved "
     "baseline needs its own approval; sheet 12 carries it. The other four answers change the "
     "specification and the component list, not this document."],
    ["1.2", DOC_DATE, "Claude Code", "-",
     "STEP N3 ISSUED FOR REVIEW - a progress record, not a change. N3.1 designed the eight screens the "
     "browser has no counterpart for and built them as a clickable prototype rather than describing them, "
     "because a table describing a dialog is a poor way to review a dialog. N3.2 produced the divergence "
     "register the plan requires at NR-PAR-04: nine differences from the web interface, each with its "
     "reason, and a tenth row saying that nothing else differs. PRAP_NewApp_Component_List_v0.1.xlsx "
     "carries 20 components, 18 user-visible strings and 5 open points, and awaits Gate N3."],
    ["1.1", DOC_DATE, "Claude Code", "-",
     "STEP N2 COMPLETE - a progress record, not a change. v1.0 remains the approved baseline and no "
     "requirement, decision, risk or assumption is altered. N2.1 split the source into core / ui / storage / "
     "shell with tools/build_app.py reassembling it; N2.2 proved the split inert - the rebuilt app/PRAP.html "
     "is BYTE-IDENTICAL to the file the thirteen suites verified, and all thirteen pass against it "
     "unmodified. N2.3 to N2.8 produced PRAP_NewApp_Specification_v1.0.xlsx, approved at Gate N2 with all "
     "seven of its open points agreed. Sheet 07 carries the statuses; sheet 04 and sheet 08 record what the "
     "refactor turned out to cost, which was less than this plan assumed: storage/ in the web shell is two "
     "functions and 63 lines."],
    ["1.0", DOC_DATE, "Claude Code", "APPROVED",
     "BASELINE. Gate N1 closed by direction on 2026-08-13 - 'go to next step' - with no change to the "
     "content of v0.7. Decisions N-01..N-32 are confirmed by the approval itself rather than by a seventh "
     "review round, exactly as decisions C-06..C-11 of the web application plan were at its Gate 1. Two "
     "items were outstanding at v0.7 and are resolved the same way: the reading taken of the Q-N04 answer "
     "is ADOPTED, and the documents KEEP the PRAP_NewApp_ prefix so this line stays beside the web "
     "application's. Both remain correctable - say so and they change under a v1.1. The 69 requirements on "
     "sheet 03 are now the contract for Steps N2-N5."],
    ["0.7", DOC_DATE, "Claude Code", "Superseded",
     "REVIEW ROUND 6 APPLIED, from PRAP_NewApp_Development_Plan_v0.6_reviewed.xlsx. Every one of the 18 "
     "questions is now answered. Q-N02: the approach is ASSUMED workable rather than confirmed with IT - "
     "recorded honestly at A-N09 and R-N01, which stay open and are proven only at N5.7. Q-N03: not an "
     "answer but a request for a plainer question; sheet 10 gains a plain-language explanation of what code "
     "signing is, what it would cost and who would buy it, and the question is re-asked in those terms. "
     "Q-N05 confirms the difference report. Q-N07 cuts retained versions from ten to ONE. Q-N09 confirms no "
     "validation obligation, which settles R-N16. Q-N10 adds a read-only look at a source workbook. Q-N11 "
     "names the application. Q-N14 and Q-N15: both distribution modes, on a WRITABLE share - so per-user "
     "data folders can live on the share itself. Q-N04's answer is about file encryption rather than "
     "location; the reading taken is recorded at NR-IMP-06 and R-N18 and needs your confirmation. "
     "NR-APP-08, NR-IMP-05..07, NR-DEP-14 added; NR-STO-06 changed. Decisions N-29..N-32. Risks R-N18 and "
     "R-N19 added; R-N01, R-N09 and R-N16 re-rated."],
    ["0.6", DOC_DATE, "Claude Code", "Superseded",
     "REVIEW ROUND 5 APPLIED. Q-N18 closed - the identity is NAME and DEPARTMENT, and nothing else. "
     "E-mail and telephone are dropped rather than left as optional fields nobody fills in. This turns out "
     "to fit the data model already in place: Person.department exists in the source schema, so an identity "
     "and a person record now use one vocabulary rather than two. NR-USR-04 rewritten; NR-USR-06 rewritten "
     "so the message shows name and department and copies them for a directory lookup; NR-USR-10 and "
     "NR-USR-11 added - the department is offered from the departments already in the open workspace, and "
     "where the declared name matches a person in the plan their department is offered first. Decision "
     "N-28. No new questions."],
    ["0.5", DOC_DATE, "Claude Code", "Superseded",
     "REVIEW ROUND 4 APPLIED. Q-N16 answered - a silent session holds a plan for 30 MINUTES, not the 90 "
     "seconds proposed. Q-N17 answered, and it is what makes 30 minutes coherent: users identify themselves "
     "at launch, and a blocked colleague is shown who is editing and how to reach them, so the ordinary "
     "remedy is a telephone call rather than taking the plan away. New requirement area NR-USR-01..08 and a "
     "new section on sheet 05a. The expiry is separated from the heartbeat, which stays short, so the "
     "application always knows whether a holder is alive even while it waits out the 30 minutes - and a user "
     "may reclaim THEIR OWN crashed session at once, which removes the one genuinely infuriating case. "
     "Decisions N-23..N-27. Risks R-N16 (a typed name is not proof of identity) and R-N17 (a stranded plan). "
     "One question, Q-N18."],
    ["0.4", DOC_DATE, "Claude Code", "Superseded",
     "REVIEW ROUND 3 APPLIED. The sharing model is settled as SINGLE WRITER, MANY READERS on your "
     "instruction: while someone is editing, no other session may update. One change of substance against "
     "v0.3 - the claim is made at the FIRST EDIT rather than at open, which lets everybody read a plan "
     "somebody else is working on, and which lands exactly on the snapshot point the edit model already has. "
     "New sheet 05a carries the whole model: the four states a session can be in, what is permitted in each, "
     "how the claim is made atomically on a network share, how a dead session's claim is released, and the "
     "four ways a claim ends. NR-STO-10 rewritten; NR-STO-12..18 added; decisions N-19..N-22; risks R-N14 "
     "(SMB caching can defeat both the claim and the staleness check) and R-N15 (a blocked user forking the "
     "plan through Save As). Two questions, Q-N16 and Q-N17."],
    ["0.3", DOC_DATE, "Claude Code", "Superseded",
     "REVIEW ROUND 2 APPLIED. Q-N13 closed - no file association is wanted, so NR-IMP-04 stands as reduced "
     "and decision N-15 is unchallenged. Q-N12 answered 'a shared network folder could be considered', which "
     "is the larger change: one application folder that several people run means their settings would collide, "
     "and two of them could edit one plan at once and lose each other's work. Planned for rather than deferred. "
     "NR-DEP-06 amended to permit the places the USER designates; NR-DEP-09 raised from Should to Must; "
     "NR-DEP-10..13 and NR-STO-10 added, covering where data goes, per-user separation, the personal shortcut "
     "that carries it, and workspace locking. Decisions N-16, N-17 and N-18. Risks R-N12 (running from a UNC "
     "path) and R-N13 (a stale lock) added. Q-N06 is superseded - concurrent use is now designed for whether or "
     "not it happens. Two new questions, Q-N14 and Q-N15, both about how the share would actually be used."],
    ["0.2", DOC_DATE, "Claude Code", "Superseded",
     "REVIEW ROUND 1 APPLIED. Two instructions, both narrowing the product: it must run on Windows, and it "
     "must NOT be installed. Q-N01 is closed (Windows 10/11 only; macOS and Linux out of scope). Q-N02 is "
     "narrowed - no installer means no administrator rights are needed at all, so what remains to check with "
     "IT is execution policy, SmartScreen and anti-virus, not installation rights. NR-DEP-01..04 rewritten "
     "and NR-DEP-05..08 added; NR-IMP-04 reduced, because a file association would need a registry write the "
     "portable rule forbids. Decisions N-13, N-14 and N-15 added, and N-14 rejects the single self-extracting "
     "executable in favour of a plain folder. Sheet 05 gains the folder layout, which had to be settled here "
     "because putting workspaces inside the application folder would make an update delete them. Risks R-N09 "
     "(SmartScreen and anti-virus), R-N10 (Windows path length) and R-N11 (read-only network share) added; "
     "R-N01 re-scoped. Two new questions, Q-N12 and Q-N13."],
    ["0.1", DOC_DATE, "Claude Code", "Superseded",
     "FIRST ISSUE. Raised on your instruction of 2026-08-13: build the requested application as a separate "
     "program type, keep imported data across sessions, and plan it apart from the web application under its "
     "own document name. Contains 33 requirements (sheet 03), the layer split that makes one engine serve two "
     "applications (sheet 04), the workspace persistence model (sheet 05), 12 decisions (sheet 06) and 11 open "
     "questions (sheet 11). Distribution size is recorded as unconstrained on your instruction, which is what "
     "makes decision N-01 affordable."],
]
r = table(ws, r, ["Version", "Date", "Author", "Reviewer", "Summary of change"],
          hist, [10, 13, 15, 13, 108], wrap_cols=(5,), mark_col=1)

r = section(ws, r, "Numbering rules for this document line")
rules = [
    ["v0.x", "Draft issues during Step 1 review rounds."],
    ["v1.0", "The approved baseline, set at Gate N1."],
    ["v1.x", "Approved changes against that baseline, each separately approved."],
    ["v2.0", "Reserved for a structural change, on the same rule the web plan used."],
]
r = table(ws, r, ["Version", "Meaning"], rules, [12, 110], wrap_cols=(2,))
r = note(ws, r, "Gate prefixes are N1..N5 throughout this document, so a reference such as 'G4' can never be "
                "confused between the two plans.")
r = legend(ws, r)

# ---- 02 Scope -------------------------------------------------------------
ws, r = sheet(wb, "02_Scope", "Scope - two applications, one engine",
              "What the desktop application is for, what it inherits, and what is explicitly out of scope.")

r = section(ws, r, "Objective")
r = lines(ws, r, [
    "Deliver the same resource simulation as a conventional installed program: an application with its own icon,",
    "its own window and its own files, which remembers the data it was given and opens showing the plan the user",
    "was last working on. The simulation itself - the load formula, the period derivation, the validation rules -",
    "does not change at all, and must not.",
])
r += 1

r = section(ws, r, "The two applications side by side")
twin = [
    [f"{MARK_CHG}Form", "One HTML file, opened in a browser", "Program with its own window and icon, run from a copied folder - NOT installed", "NR-APP-01, NR-DEP-05"],
    [f"{MARK_CHG}Platform", "Any browser on any OS", "Windows 10 / 11 only", "NR-DEP-01"],
    [f"{MARK_CHG}Data after closing", "Lost - the file must be re-imported each session", "KEPT - the workspace reopens as it was left", "NR-STO-01"],
    [f"{MARK_CHG}Source of truth", "The Excel workbook outside the application", "The workspace file the application owns; Excel becomes exchange", "NR-STO-02"],
    ["Calculation engine", "Shared - one implementation", "Shared - the same implementation, not a copy", "NR-PAR-01"],
    ["Data contract", "Schema version 5, ten sheets", "Identical, unchanged", "NR-PAR-02"],
    ["AI agent interoperability", "prap-source-data JSON, format_version 1", "Identical, and now also the on-disk format", "NR-STO-03"],
    [f"{MARK_CHG}Distribution", "Copy one file", "Copy one folder. No installer, no administrator rights, no registry. Size unconstrained", "NR-DEP-02, NR-DEP-05"],
    [f"{MARK_NEW}Removal", "Delete the file", "Delete the folder. Nothing is left behind anywhere on the machine", "NR-DEP-07"],
    ["Network use", "None", "None", "NR-SEC-01"],
    ["Status", "Complete, in service, frozen at Gate 5", "Not started - this plan is Step 1", "-"],
]
r = table(ws, r, ["Aspect", "Web application", "Desktop application", "Requirement"],
          twin, [26, 46, 56, 14], wrap_cols=(2, 3), mark_col=1)
r += 1

r = section(ws, r, "What the web application keeps - not touched by this plan")
keep = [
    [f"{WEB_APP} continues in service, unchanged, and is released as v1.0 under its own Gate 5."],
    [f"{WEB_PLAN}, the programming specification v1.0 and the UI component list v1.0 stay approved as they are."],
    ["The source workbook template, the dummy datasets and the AI agent guide serve BOTH applications and are "
     "not forked."],
    ["The 13 verification suites under tools/ keep running against the HTML build after every change, which is "
     "how a regression in the shared engine is caught before it can reach either application."],
    ["Any engine fix found while building the desktop application is made in the shared engine, re-verified "
     "against both, and released to the HTML build as a point version. Fixes flow to both; features do not "
     "flow backwards without their own approval."],
]
r = table(ws, r, ["Point"], keep, [146], wrap_cols=(1,))
r += 1

r = section(ws, r, "In scope")
ins = [
    ["A desktop application for Windows 10 / 11, launched from an icon, no browser involved."],
    ["Portable operation: the whole application is one folder that is copied where the user wants it and run "
     "from there, with everything it needs and everything it writes kept inside that folder."],
    ["Persistent storage: a workspace the application owns, written on every Save and reopened at launch."],
    ["Import from the Excel source workbook and from prap-source-data JSON, as today - but as an INGEST into "
     "the workspace rather than as the session's only data."],
    ["Export to Excel and JSON, as today, for sharing and archiving."],
    ["Desktop conventions the web version cannot offer: native Open/Save dialogs, a recent-workspaces list, "
     "a menu bar, window state remembered between sessions."],
    ["Crash and power-loss recovery, including edits that were pending when the application stopped."],
    ["The refactor that makes one engine serve both applications, with the HTML build proven unchanged by it."],
]
r = table(ws, r, ["Item"], ins, [146], wrap_cols=(1,))
r += 1

r = section(ws, r, "Out of scope for v1.0")
outs = [
    ["Any change to the load formula, the period derivation, the validation rules or the data schema.",
     "The engine is shared. Changing it here would change the finished web application too."],
    ["Multi-user or concurrent editing of one workspace.",
     "Assumption A-06 of the web plan still holds - one person maintains a plan at a time. Q-N06 asks whether "
     "that is still true."],
    ["A server, a database service, or any network feature.",
     "The tool is offline by requirement, and that is a property worth keeping rather than an accident."],
    ["macOS and Linux builds.",
     "CLOSED at review round 1: Windows only. The same source would build for them, but they are neither "
     "verified nor supported."],
    ["An installer, a Start-menu entry, a file association, or any registry entry.",
     "CLOSED at review round 1: the application is non-installed. Each of these would write outside its own "
     "folder, which NR-DEP-06 forbids. The cost is recorded honestly at NR-IMP-04 and decision N-15."],
    ["Automatic update over the network.",
     "It would be the only thing in the product that needs a network. Updates are distributed as a new package."],
    ["Migrating the web application onto the desktop shell, or retiring it.",
     "Explicitly excluded on your instruction: both applications stay."],
]
r = table(ws, r, ["Excluded", "Why"], outs, [66, 80], wrap_cols=(1, 2))
r = legend(ws, r)

# ---- 03 Requirements ------------------------------------------------------
ws, r = sheet(wb, "03_Requirements", "Requirement register - desktop application",
              "NR-ids belong to this plan. REQ-ids referenced in the last column belong to the web application "
              "plan and are inherited unchanged.")

r = section(ws, r, "Inheritance rule")
r = lines(ws, r, [
    "Every requirement of the web application plan that concerns DATA or CALCULATION is inherited by this",
    "application unchanged, and is not restated here. This register holds only what is new or deliberately",
    "different. Where a requirement here replaces an inherited one, the inherited ID is named so the",
    "contradiction is visible rather than buried.",
])
r += 1

reqs = [
    [f"{MARK_CHG}NR-APP-01", "Application form", "The application is a desktop program with its own window, icon and menu bar. No browser is visible to the user at any point.", "Must", "Your instruction", "N4"],
    [f"{MARK_NEW}NR-APP-02", "Application form", "It runs fully offline. It makes no network request of any kind, including on first launch.", "Must", "Inherited intent", "N4"],
    [f"{MARK_NEW}NR-APP-03", "Application form", "It launches to the plan the user last had open, without asking for a file.", "Must", "Your instruction", "N4"],
    [f"{MARK_NEW}NR-APP-04", "Application form", "Window size, position and the active tab are remembered between sessions.", "Should", "Desktop convention", "N4"],
    [f"{MARK_NEW}NR-APP-05", "Application form", "Native Open, Save, Save As, Import and Export dialogs are used, not browser download prompts.", "Must", "Desktop convention", "N4"],
    [f"{MARK_CHG}NR-APP-06", "Application form", "A recent-workspaces list of at least ten entries is available from the menu. It is stored in the application's own data folder, not in the registry or in the user profile.", "Should", "Desktop convention + NR-DEP-06", "N4"],
    [f"{MARK_NEW}NR-APP-07", "Application form", "The application version is visible in an About dialog and in the window title alongside the open workspace name.", "Must", "Version control", "N4"],
    [f"{MARK_NEW}NR-APP-08", "Application form", "The application is called 'Project Management APP' on screen. Its folder and executable are named PM_APP, so the name survives being copied, shortened or pinned.", "Must", "Q-N11", "N4"],

    [f"{MARK_NEW}NR-STO-01", "Storage", "Imported and hand-entered data survives closing the application. Reopening shows the same plan, with the same figures, without re-importing anything.", "Must", "Your instruction", "N4"],
    [f"{MARK_NEW}NR-STO-02", "Storage", "The application owns a workspace file on disk. That file, not the source Excel workbook, is the working store from the moment data is imported into it.", "Must", "Derived from NR-STO-01", "N4"],
    [f"{MARK_NEW}NR-STO-03", "Storage", "The workspace file is prap-source-data JSON - the same format the AI agent guide already documents - so a workspace can be read, written or generated by an AI agent with no new contract.", "Must", "Interoperability", "N4"],
    [f"{MARK_NEW}NR-STO-04", "Storage", "A committed Save is written to disk before the application reports it as saved. A save that did not reach the disk is never reported as done.", "Must", "Data safety", "N4"],
    [f"{MARK_NEW}NR-STO-05", "Storage", "Workspace writes are atomic: an interrupted write can never leave a workspace file that is neither the old contents nor the new.", "Must", "Data safety", "N4"],
    [f"{MARK_CHG}NR-STO-06", "Storage", "The previous committed version of a workspace is retained and can be restored from within the application. The number retained is a setting; its default is ONE, so a workspace directory holds the current file and the one before it.", "Should", "Q-N07", "N4"],
    [f"{MARK_NEW}NR-STO-07", "Storage", "Edits pending at the moment of a crash or power loss are recovered on the next launch, and the user is asked whether to keep or discard them.", "Should", "Data safety", "N4"],
    [f"{MARK_NEW}NR-STO-08", "Storage", "Several workspaces may exist. The user chooses which to open, and may create a new empty one at any time.", "Must", "Scenario planning", "N4"],
    [f"{MARK_NEW}NR-STO-09", "Storage", "A workspace records which source file it was imported from, when, and by which application version.", "Should", "Traceability", "N4"],
    [f"{MARK_CHG}NR-STO-10", "Storage", "While one session is editing a workspace, no other session may update it. The claim is made when the FIRST EDIT is attempted - not when the workspace is opened - so that reading is never blocked.", "Must", "Your instruction, round 3", "N4"],
    [f"{MARK_CHG}NR-STO-11", "Storage", "Reading is a full working mode, not a refusal: every figure, table, graph, filter and export is available to a session that does not hold the claim. Only writing to that workspace is withheld.", "Must", "Consequence of NR-STO-10", "N4"],
    [f"{MARK_CHG}NR-STO-12", "Storage", "A session refused the claim is shown a message naming who holds it, how to reach them, since when, whether their session is still responding, and when the plan becomes free - never merely that the file is locked.", "Must", "Q-N17", "N4"],
    [f"{MARK_NEW}NR-STO-13", "Storage", "The claim is made by an operation that cannot be won by two sessions at once, even when they attempt it in the same instant and even across a network share.", "Must", "Correctness", "N4"],
    [f"{MARK_CHG}NR-STO-14", "Storage", "A holding session refreshes its claim every 30 seconds while it lives, so the application always knows whether the holder is alive. A claim that has stopped being refreshed expires after 30 MINUTES, after which anybody may take it over.", "Must", "Q-N16", "N4"],
    [f"{MARK_NEW}NR-STO-19", "Storage", "A user may reclaim a stalled claim held by their OWN name on their OWN machine immediately, without waiting out the expiry. Being locked out of your own plan because your application crashed is not a rule worth enforcing.", "Must", "Consequence of the 30-minute expiry", "N4"],
    [f"{MARK_NEW}NR-STO-15", "Storage", "The claim ends when the holder saves and closes, discards their edits, or closes the application - and a session waiting to edit is offered it without having to reopen the workspace.", "Must", "Usability", "N4"],
    [f"{MARK_NEW}NR-STO-16", "Storage", "A reading session notices when the workspace has changed on disk beneath it, says so, and offers to reload. It never presents figures it knows to be superseded as though they were current.", "Must", "Consequence of many readers", "N4"],
    [f"{MARK_NEW}NR-STO-17", "Storage", "A session that cannot claim the workspace may still Save As a copy of its own, and may still export. Being unable to edit a shared plan never means being unable to work.", "Must", "Usability", "N4"],
    [f"{MARK_NEW}NR-STO-18", "Storage", "Importing into a workspace is an edit and takes the claim like any other. It is the largest change the application can make, so it cannot be the one that bypasses the rule.", "Must", "Correctness", "N4"],

    [f"{MARK_CHG}NR-IMP-01", "Import / export", "Import reads the Excel source workbook and prap-source-data JSON exactly as the web application does, using the same reader and reporting the same findings.", "Must", "Replaces nothing - REQ-IMP-01 inherited", "N4"],
    [f"{MARK_CHG}NR-IMP-02", "Import / export", "Importing into a workspace that already holds data first asks whether the existing data is to be updated at all. If it is, a DIFFERENCE REPORT is presented - what the file would add, change and remove, sheet by sheet - and the user decides per sheet whether to accept it. Nothing is overwritten silently, and nothing is overwritten unasked.", "Must", "Q-N05", "N4"],
    [f"{MARK_CHG}NR-IMP-03", "Import / export", "Export to Excel and to JSON produces byte-for-byte the same content the web application would produce from the same data.", "Must", "Parity", "N4"],
    [f"{MARK_NEW}NR-IMP-05", "Import / export", "A source workbook can be opened for a look without creating a workspace: every table and graph, nothing saved, nothing left behind. The quick question - 'what does this file say?' - does not require committing to a plan first.", "Must", "Q-N10", "N4"],
    [f"{MARK_NEW}NR-IMP-06", "Import / export", "A file the application cannot read because it is encrypted or protected says so, names that as the likely cause, and says what to do about it. It never reports a protected file as corrupt, and never fails silently.", "Must", "Q-N04 - see R-N18", "N4"],
    [f"{MARK_NEW}NR-IMP-07", "Import / export", "Nothing the application writes is encrypted by the application itself. Where files must be protected, that is done by whatever the company already uses, on the folder - so the application never becomes the only thing that can open the user's own data.", "Must", "Q-N04 - see R-N18", "N2"],
    [f"{MARK_NEW}NR-IMP-08", "Import / export", "Export is a supported way to EDIT a plan, not only to read one. The exported file - in either format - can be changed outside the application and imported back with no loss: every row, column and value that went out comes back, and anything the application derives is recomputed rather than trusted.", "Must", "Gate N3 review - C-N01", "N4"],
    [f"{MARK_CHG}NR-IMP-04", "Import / export", "A workspace can be opened by dragging it onto the running application's window. Double-clicking a workspace file in Explorer does NOT open it, because a file association would require a registry entry that NR-DEP-06 forbids.", "Could", "Reduced by the non-installed rule", "N4"],

    [f"{MARK_NEW}NR-USR-01", "User identity", "The application asks who the user is when it starts, before any workspace is opened, and carries that identity for the session.", "Must", "Q-N17", "N4"],
    [f"{MARK_NEW}NR-USR-02", "User identity", "It asks in full only the first time. Afterwards it shows what it remembers, pre-filled, for the user to confirm with one click - and offers 'not me' for a shared PC.", "Must", "Q-N17, usability", "N4"],
    [f"{MARK_NEW}NR-USR-03", "User identity", "The name is pre-filled from the Windows account name the first time, so the common case is confirming rather than typing. It stays editable - a Windows account name is often not how colleagues know each other.", "Should", "Q-N17", "N4"],
    [f"{MARK_CHG}NR-USR-04", "User identity", "An identity is a name and a DEPARTMENT. Both are recorded; no e-mail, telephone or other field is collected. A name and a department are what a colleague needs to find somebody in the company directory.", "Must", "Q-N18", "N4"],
    [f"{MARK_NEW}NR-USR-05", "User identity", "A claim carries the holder's name and contact details, so the message a blocked colleague sees can be acted on rather than merely read. Those details are removed with the claim.", "Must", "Q-N17 - contact", "N4"],
    [f"{MARK_CHG}NR-USR-06", "User identity", "The blocked message shows the holder's name and department together, and offers to copy them, so a colleague can look them up or ask their department without writing anything down.", "Should", "Q-N18", "N4"],
    [f"{MARK_CHG}NR-USR-07", "User identity", "A workspace records who last saved it - name and department - and when, shown when it is opened. On a shared plan, knowing whose figures these are matters as much as knowing when they were made.", "Should", "Follows from having identities", "N4"],
    [f"{MARK_NEW}NR-USR-08", "User identity", "This identity is DECLARED, not verified. The application states as much where it is entered, and nothing in it is treated as proof of who acted.", "Must", "Honesty about what it is", "N2"],
    [f"{MARK_NEW}NR-USR-09", "User identity", "Identity is held per user in their own data folder, never in the shared application folder, so two people on one copy do not overwrite each other's details.", "Must", "NR-DEP-11", "N4"],
    [f"{MARK_NEW}NR-USR-10", "User identity", "The department is offered from the departments already present in the open workspace's Person records, and typed freely only where none matches. One vocabulary, not two - 'Data Management' and 'DM' should not both exist because one was typed at a login prompt.", "Should", "Q-N18, Person.department", "N4"],
    [f"{MARK_NEW}NR-USR-11", "User identity", "Where the declared name matches a person in the open workspace, that person's department is offered first. The user is usually one of the people in the plan, so their own record is the best answer available.", "Could", "Q-N18", "N4"],

    [f"{MARK_NEW}NR-PAR-01", "Parity", "The calculation engine, the validation rules and the period derivation are ONE implementation shared by both applications, not two copies kept in step by hand.", "Must", "Your instruction to keep both", "N2"],
    [f"{MARK_NEW}NR-PAR-02", "Parity", "Given identical input, both applications produce identical figures, identical findings and identical exports. This is proven by an automated test, not asserted.", "Must", "Derived from NR-PAR-01", "N4"],
    [f"{MARK_NEW}NR-PAR-03", "Parity", "The HTML application continues to be built from the shared engine and its behaviour is unchanged by the refactor, proven by the existing 13 suites passing unmodified.", "Must", "Protects the finished product", "N2"],
    [f"{MARK_NEW}NR-PAR-04", "Parity", "Screen layout, wording and interaction match the web application except where a desktop convention requires otherwise, and every such difference is listed.", "Should", "Two tools, one habit", "N3"],

    [f"{MARK_CHG}NR-DEP-01", "Deployment", "The application runs on Windows 10 and Windows 11, 64-bit. No other operating system is supported or tested.", "Must", "Your instruction", "N5"],
    [f"{MARK_CHG}NR-DEP-02", "Deployment", "There is NO INSTALLER. The application is a folder that is copied to wherever the user wants it and run from there. It requires no administrator rights, at any point, for any purpose.", "Must", "Your instruction", "N5"],
    [f"{MARK_CHG}NR-DEP-03", "Deployment", "Updating means replacing the application files inside that folder. No workspace, setting or recent-file entry is touched by an update, and a workspace written by an older version opens in a newer one.", "Must", "Data safety", "N5"],
    [f"{MARK_NEW}NR-DEP-04", "Deployment", "A workspace written by a NEWER version than the running one is refused with a clear message rather than partially read.", "Must", "Data safety", "N4"],
    [f"{MARK_NEW}NR-DEP-05", "Deployment", "The application folder is self-contained: it carries its own browser engine and every other dependency, and requires nothing to be present on the machine beyond Windows itself. Package size is not constrained.", "Must", "Your instruction", "N5"],
    [f"{MARK_CHG}NR-DEP-06", "Deployment", "The application writes nothing outside its own folder EXCEPT where the user has explicitly designated somewhere else - a chosen data folder, an exported file, or a shortcut created at their request. No registry key, no file association, no Start-menu entry, no AppData directory, and no hidden state anywhere.", "Must", "Non-installed rule, amended for the shared folder", "N5"],
    [f"{MARK_NEW}NR-DEP-07", "Deployment", "Deleting the folder removes the application completely, leaving no trace on the machine.", "Must", "Non-installed rule", "N5"],
    [f"{MARK_NEW}NR-DEP-08", "Deployment", "The folder can be copied to another Windows PC, or moved to a different path on the same PC, and works unchanged - including its recent-workspaces list where those workspaces travelled with it. No path is recorded absolutely where a relative one would do.", "Must", "Non-installed rule", "N5"],
    [f"{MARK_CHG}NR-DEP-09", "Deployment", "Where the application folder is read-only or shared - typically on a network folder - the application still runs, tells the user plainly that it cannot keep data beside itself, and keeps that data in a folder the user chooses instead.", "Must", "Raised from Should at round 2 - Q-N12", "N4"],
    [f"{MARK_NEW}NR-DEP-10", "Deployment", "Where the application keeps its data is resolved in one fixed, documented order, and the answer is shown in the About dialog so it is never a mystery: a --data command-line argument, then a PRAP_DATA environment variable, then a writable data folder beside the application, and failing all of those the application asks.", "Must", "Q-N12", "N4"],
    [f"{MARK_NEW}NR-DEP-11", "Deployment", "Where several people run one shared copy, each has their own settings, recent list, backups and default workspace folder. No user can see, overwrite or lock out another through the application.", "Must", "Q-N12", "N4"],
    [f"{MARK_NEW}NR-DEP-12", "Deployment", "The application can create a personal desktop shortcut, at the user's request, that carries their own --data path. This is how a shared copy is used day to day without being asked anything at launch.", "Should", "Q-N12", "N4"],
    [f"{MARK_CHG}NR-DEP-13", "Deployment", "The application runs correctly when launched from a UNC path (\\\\server\\share\\PM_APP) as well as from a mapped drive letter.", "Must", "Q-N12", "N5"],
    [f"{MARK_NEW}NR-DEP-14", "Deployment", "Both arrangements are supported and both are tested: run in place from the shared folder, and copied from it to a local folder and run there. Neither is a second-class case, and the user guide states what each costs in launch time.", "Must", "Q-N14", "N5"],
    [f"{MARK_NEW}NR-DEP-15", "Deployment", "On a writable shared folder, each person's data folder sits under the application's own data folder, one per user. Nobody has to choose a location, and nobody's settings meet anybody else's.", "Should", "Q-N15", "N4"],

    [f"{MARK_NEW}NR-SEC-01", "Security", "No telemetry, no crash reporting to any server, no automatic update check, no font or script loaded from a remote host. Verified by observing that the packaged application opens no socket.", "Must", "Offline by requirement", "N5"],
    [f"{MARK_NEW}NR-SEC-02", "Security", "The embedded browser engine runs with remote content disabled and Node integration off in the renderer, so a crafted workspace file cannot execute code.", "Must", "Standard hardening", "N4"],
    [f"{MARK_NEW}NR-SEC-03", "Security", "Workspace files contain business data only - no credentials, no connection strings, nothing that would be sensitive if the file were mailed to a colleague.", "Must", "Data hygiene", "N2"],

    [f"{MARK_CHG}NR-NFR-01", "Performance", "Launch to a usable window in under 5 seconds from a local disk, and reopening the last workspace adds no more than 2 seconds on the volumes in A-N01. Launch from a network folder is measured and stated rather than promised - see R-N12.", "Should", "Usability", "N4"],
    [f"{MARK_NEW}NR-NFR-02", "Performance", "A Save completes in under 1 second at those volumes, so autosave is never felt.", "Should", "Usability", "N4"],
]
r_start = r
r = table(ws, r, ["ID", "Area", "Requirement", "Priority", "Source", "Gate"],
          reqs, [13, 17, 92, 9, 30, 7], wrap_cols=(3, 5), mark_col=1)
last = r_start + len(reqs)

dv = DataValidation(type="list", formula1='"Must,Should,Could,Won\'t"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"D{r_start + 1}:D{last}")

r = section(ws, r, "Count")
cnt = [
    ["Total requirements", f"=COUNTA(A{r_start + 1}:A{last})"],
    ["Must", f'=COUNTIF(D{r_start + 1}:D{last},"Must")'],
    ["Should", f'=COUNTIF(D{r_start + 1}:D{last},"Should")'],
    ["Could", f'=COUNTIF(D{r_start + 1}:D{last},"Could")'],
]
r = table(ws, r, ["Measure", "Count"], cnt, [26, 12])
r = note(ws, r, "Please mark any requirement you disagree with in the Priority column, or strike it out. The "
                "register is the contract for Steps 2 to 5, exactly as sheet 03 of the web plan was.")
r = legend(ws, r)

# ---- 04 Architecture ------------------------------------------------------
ws, r = sheet(wb, "04_Architecture", "Architecture - one engine, two shells",
              "The single decision that makes two applications affordable to maintain.")

r = section(ws, r, "The problem this solves")
r = lines(ws, r, [
    "Two applications doing the same arithmetic will drift apart. The month one of them reports 3.4 FTE and the",
    "other 3.6 for the same person, both become untrustworthy - and nobody will know which to believe. The only",
    "durable answer is that there are not two implementations to drift.",
    "",
    f"{WEB_APP} is 4,865 lines, of which about 4,160 are script. Only 69 of those script lines touch the DOM",
    "at all. The engine is therefore already almost separable; this is a lift, not a rewrite.",
])
r += 1

r = section(ws, r, "Layers")
layers = [
    ["core/", "Parse, validate, derive periods, calculate load, read and write xlsx, read and write JSON. No DOM, no filesystem, no dialogs.", "Shared, identical", "The whole of the trust in both products sits here. Built at N2.1: 7 files, 1,210 lines, and tools/test_layers.py holds it to the boundary."],
    ["ui/", "Tabs, tables, charts, filters, the provisional-edit model, the change log.", "Shared, identical", "Renders into a DOM; indifferent to which shell provides it."],
    ["storage/", "An interface: open, save, listWorkspaces, restoreVersion, recoverPending.", "Interface shared, implementation per shell", "The whole of the difference between the two products sits here. In the web shell it is two functions and 63 lines - one reads a file, one writes one - which is how small the seam turned out to be."],
    ["shell/web/", "Storage over the browser's own storage; the build step that emits the single HTML file.", "Web only", "Keeps the finished product buildable and unchanged."],
    [f"{MARK_NEW}shell/desktop/", "Application window, menu bar, native dialogs, storage over the filesystem, recent files, window state.", "Desktop only", "New code, and the only genuinely new code in the product."],
    ["tools/", "The 13 verification suites, the document generators, the Python reference implementation.", "Shared", "Runs against both shells after the refactor."],
]
r = table(ws, r, ["Layer", "Responsibility", "Shared?", "Why the boundary is there"],
          layers, [18, 62, 26, 42], wrap_cols=(2, 3, 4), mark_col=1)
r += 1
r = note(ws, r, "Read the table by its third column. Everything that decides a number is shared; everything that "
                "differs is about where bytes are kept and which dialog opens. That is the whole architecture.")
r += 1

r = section(ws, r, "Build targets from one source tree")
r = lines(ws, r, [
    "    core/ + ui/ + shell/web/       --build-->   app/PRAP.html          one file, unchanged, still supported",
    "    core/ + ui/ + shell/desktop/   --build-->   PM_APP package         new",
    "    core/                          --used by--> tools/ test suites     unchanged",
], mono=True)
r += 1
r = note(ws, r, "The HTML file stops being hand-written and becomes a build output. Its content is identical - "
                "task N2.1 requires the rebuilt file to pass all 13 suites before anything else starts.")
r += 1

r = section(ws, r, "Shell technology - the choice, and why")
tech = [
    [f"{MARK_NEW}Electron", "RECOMMENDED", "Reuses the engine and the UI unchanged. Filesystem access is direct. Testable in this repository today - Playwright drives an Electron application natively, so all 13 suites plus new desktop suites run before you ever see a build.", "Package around 180 MB. Chromium security updates mean rebuilding periodically."],
    ["Tauri", "Later option", "Package around 10 MB, single executable. Same engine reuse.", "Needs a Rust toolchain and a Windows build host; cannot be verified from the current environment. Automated UI testing needs a WebDriver setup that is not available here."],
    ["Native rewrite (C# / .NET, Java)", "Rejected", "Smallest package, best desktop integration.", "Discards a verified engine and a verified UI, and creates the second implementation that NR-PAR-01 exists to prevent. Months of work to reach where the product already is."],
    ["Progressive web app", "Rejected", "No packaging at all.", "Does not meet NR-APP-01 - it is still the browser, with the browser's storage rules."],
]
r = table(ws, r, ["Option", "Status", "For", "Against"], tech, [26, 14, 58, 50], wrap_cols=(3, 4), mark_col=1)
r = note(ws, r, "Size was the strongest argument for Tauri, and your instruction removes it. What remains is "
                "verifiability: an Electron build can be tested here, before delivery, by the same means every "
                "other claim in this project has been tested. Recorded as decision N-01, and it is reversible - "
                "the shell layer is the only code that would change.")
r += 1

r = section(ws, r, "Portable packaging   [round 1: non-installed]")
r = lines(ws, r, [
    "'Non-installed' is a stronger constraint than 'no administrator rights', and it rules out two shapes of",
    "package that are otherwise normal on Windows. The two that remain differ in a way the user feels:",
])
r += 1
pack = [
    ["Installer (MSI / NSIS / MSIX)", "REJECTED", "Writes to Program Files and the registry, creates Start-menu entries and an uninstall record. This is precisely what 'non-installed' excludes."],
    ["Single self-extracting .exe", "REJECTED", "Looks the most portable, and behaves the least. Each launch unpacks around 180 MB into the Windows temporary folder before the window appears - slow to start (against NR-NFR-01), it writes outside its own folder (against NR-DEP-06), and anti-virus scans the freshly written files every time."],
    [f"{MARK_NEW}Plain folder, run in place", "SELECTED - N-14", "PRAP.exe sits in a folder with its engine and resources beside it. Copy the folder, double-click the exe, it runs. Nothing is unpacked, nothing is registered, launch is as fast as the disk. The folder is large and visibly full of files, which is the only real cost."],
    ["ZIP of that folder", "How it is delivered", "The delivered artifact is a .zip of the folder above. The user extracts it once, wherever they like, and that extracted folder IS the application."],
]
r = table(ws, r, ["Shape", "Status", "Assessment"], pack, [30, 20, 96], wrap_cols=(3,), mark_col=1)
r += 1
r = note(ws, r, "Worth being explicit, because the single .exe is what most people picture when they hear "
                "'portable': it is the worse option here on three of this plan's own requirements. A folder is "
                "less elegant and behaves better.")
r += 1

r = section(ws, r, "What the non-installed rule costs")
cost = [
    ["Double-clicking a .prap file in Explorer will not open it.", "A file association is a registry entry. Mitigated by drag-and-drop onto the window, and by the recent-workspaces list (NR-IMP-04, NR-APP-06)."],
    ["No Start-menu or taskbar entry appears by itself.", "The user can pin the exe by hand - which is a user action, not something the application does. The user guide will say how."],
    ["Windows will not treat the application as known software.", "SmartScreen and anti-virus judge an unsigned, unregistered executable more harshly, not less. See risks R-N01 and R-N09 - this is the one place where 'non-installed' makes a problem harder rather than easier."],
    ["The folder must not be nested too deeply.", "Windows path-length limits. See R-N10."],
]
r = table(ws, r, ["Cost", "What is done about it"], cost, [56, 90], wrap_cols=(1, 2))
r = note(ws, r, "None of these is a reason to reconsider - they are the price of the constraint, listed so that "
                "nobody is surprised by them at Gate N5.")
r = legend(ws, r)

# ---- 05 Data flow ---------------------------------------------------------
ws, r = sheet(wb, "05_Data_Flow", "Data flow and persistence",
              "The second of the two requirements: data that survives closing the application.")

r = section(ws, r, "Today, and what changes")
r = lines(ws, r, [
    "  WEB APPLICATION      Excel  --import-->  memory  --export-->  Excel",
    "                                             |",
    "                                          closing the tab loses everything",
    "",
    "  DESKTOP APPLICATION  Excel / JSON  --import-->  WORKSPACE (on disk)  --export-->  Excel / JSON",
    "                                                       ^   |",
    "                                                       |   +--> reopened automatically at launch",
    "                                                       +------- written on every committed Save",
], mono=True)
r += 1
r = note(ws, r, "The consequential change is not the file - it is that the workspace becomes the source of truth. "
                "After the first import, the Excel workbook is an exchange format: something you export FOR "
                "somebody, or import FROM somebody, not the thing the application reads every morning.")
r += 1

r = section(ws, r, "The workspace file")
wsf = [
    ["Format", "prap-source-data JSON, format_version 1 - the format the AI agent guide already documents.", "NR-STO-03"],
    ["Why not SQLite", "The data is small - tens of projects, hundreds of people, low thousands of rows. SQLite buys concurrency and query power that nothing here needs, and costs the property that an AI agent can read a workspace with no new contract. Revisit only if Q-N06 turns out to need multi-user access.", "N-02"],
    ["Why not the Excel workbook itself", "Writing xlsx on every Save would be slow, would lose the provisional-edit state, and would make an interrupted write corrupt the user's own archive file.", "N-02"],
    ["Extension", ".prap - but NOT registered with Windows, because that would need a registry entry the non-installed rule forbids. Workspaces are opened from within the application, or dragged onto its window.", "NR-IMP-04"],
    ["Contents", "The ten sheets exactly as imported, plus a small header: application version, schema version, source file name and import timestamp, and the retained version history.", "NR-STO-09"],
    ["Readable", "Yes - it is JSON. A workspace can be inspected, diffed, version-controlled, or handed to an AI agent as-is.", "NR-STO-03"],
]
r = table(ws, r, ["Aspect", "Decision", "Ref"], wsf, [22, 108, 12], wrap_cols=(2,))
r += 1

r = section(ws, r, "Where things live   [round 1: non-installed]")
r = lines(ws, r, [
    "The portable rule says everything lives in one folder. Data safety says an update must not be able to",
    "delete a plan. Those two pull against each other, and the layout below is what resolves them:",
    "",
    "    PM_APP\\                      <- copy THIS folder anywhere; delete it to remove the application",
    "      PM_APP.exe                  <- \\",
    "      resources\\                  <-  |  replaced wholesale by an update",
    "      version.txt                 <- /",
    "      data\\                       <- NEVER touched by an update",
    "        users\\<name>\\             <- one per person, where the folder is shared (NR-DEP-15)",
    "          settings.json           <- window state, recent workspaces, identity, preferences",
    "          workspaces\\             <- the default home for .prap files",
    "          backups\\                <- the ONE retained previous version (NR-STO-06)",
], mono=True)
r += 1
lay = [
    ["An update replaces the application files and leaves data\\ alone.", "Stated as a rule in the user guide AND enforced by the updater instructions: the delivered zip contains no data\\ folder, so extracting it over an existing installation cannot overwrite one."],
    ["Workspaces default into data\\workspaces\\ but are not confined there.", "The user may save a workspace anywhere - a project folder, a network drive. The default simply means the application always has somewhere sensible to put one."],
    ["Paths in settings.json are relative wherever they can be.", "So copying the folder to another PC, or to a different drive letter, does not break the recent list (NR-DEP-08)."],
    ["If the application folder is read-only, data\\ cannot be created.", "The application says so plainly at first launch and asks the user where to keep data instead, remembering the answer next to nothing else - see R-N11 and NR-DEP-09."],
]
r = table(ws, r, ["Rule", "How it is upheld"], lay, [56, 90], wrap_cols=(1, 2))
r = note(ws, r, "The one arrangement that must NOT be used is workspaces inside resources\\, or anywhere else "
                "among the application files: an update would silently delete the user's plans. It is an easy "
                "mistake to make and an unrecoverable one to ship, which is why it is written down here at "
                "Step 1 rather than discovered at Step 4.")
r += 1

r = section(ws, r, "When the folder is shared   [round 2: Q-N12]")
r = lines(ws, r, [
    "One copy on a network folder, several people running it, is a sensible way to distribute a portable",
    "application - and it breaks the simple layout above in two places. Both are cheap to solve now and",
    "expensive to solve after somebody has lost an afternoon's work.",
    "",
    "    1. SETTINGS COLLIDE.  One data\\settings.json cannot hold two people's window sizes and recent lists.",
    "    2. TWO WRITERS.       Two people open the same .prap file; the second Save silently discards the first.",
])
r += 1
r = note(ws, r, "Note that problem 2 exists as soon as workspaces are reachable by two people - which a shared "
                "folder makes likely even if the application itself is copied locally. So it is solved "
                "unconditionally, not only in the shared case.")
r += 1

r = section(ws, r, "Where the data folder is - resolved in one fixed order   [NR-DEP-10]")
res = [
    ["1", "--data=<path> on the command line", "How a personal shortcut carries it. Explicit, visible, and belongs to the user rather than to the application.", "Always wins"],
    ["2", "The PRAP_DATA environment variable", "For a site that wants to set it centrally, by login script or group policy.", "If no argument"],
    ["3", "data\\ beside the application - and data\\users\\<name>\\ where the folder is shared", "The ordinary single-user case, and the reason a copied folder simply works. Q-N15 says the share would be WRITABLE, so on a shared copy this rule still applies: each person gets their own folder under it and nobody has to choose anything (NR-DEP-15).", "If writable"],
    ["4", "Ask the user, once", "Only where the folder is read-only after all. The answer is not written to a hidden file - the application offers to create a desktop shortcut carrying --data, so the choice lives somewhere the user can see, move and delete.", "Last resort"],
]
r = table(ws, r, ["Order", "Source", "Why it is there", "Applies"], res, [7, 38, 76, 22], wrap_cols=(3,))
r = note(ws, r, "There is a chicken-and-egg problem hiding here: if the application cannot write beside itself, "
                "where does it remember the answer? The usual solution - a file in the user profile - is exactly "
                "the hidden state the non-installed rule forbids. A shortcut the user owns solves it honestly: "
                "it is visible, it is theirs, and deleting it undoes the decision. The About dialog always shows "
                "which of the four rules applied (NR-DEP-10).")
r += 1

r = section(ws, r, "Two people, one workspace")
r = note(ws, r, "Settled at review round 3 as SINGLE WRITER, MANY READERS, and given its own sheet: see 05a.")
r += 1

r = section(ws, r, "Writing safely")
safe = [
    ["1", "Serialise the committed model to JSON in memory.", "Nothing has touched the disk yet."],
    ["2", "Write it to a temporary file in the same directory, and flush it to the physical disk.", "Same directory, so the rename in step 4 cannot cross a filesystem boundary."],
    ["3", "Roll the current workspace file into the version history.", "This is what NR-STO-06 restores from."],
    ["4", "Rename the temporary file over the workspace file.", "A rename is atomic. The file is either wholly the old contents or wholly the new - never half of each (NR-STO-05)."],
    ["5", "Only now report the save as done.", "NR-STO-04. Reporting a save that has not reached the disk is the one failure that would destroy trust in the whole feature."],
]
r = table(ws, r, ["Step", "Action", "Why"], safe, [7, 76, 62], wrap_cols=(2, 3))
r += 1
r = note(ws, r, "Power loss at any point leaves a readable workspace: before step 4 the old one, after it the new "
                "one. A stray temporary file is detected and removed at the next launch.")
r += 1

r = section(ws, r, "How this meets the existing edit model")
edit = [
    ["A pending edit", "Held exactly as today - snapshot taken before the first change.", "Nothing written to the workspace."],
    ["Save", "Commits as today, then writes the workspace.", "The existing Save is already the commit point, so persistence attaches to it with no new state."],
    ["Leave without change", "Reverts as today.", "Nothing is written, so nothing needs undoing on disk."],
    [f"{MARK_NEW}Autosave of PENDING edits", "Pending edits are journalled separately, not into the workspace.", "So an unexpected close can offer them back (NR-STO-07) without a half-finished edit ever becoming committed data."],
    [f"{MARK_NEW}Recovery", "At launch, a journal newer than the workspace prompts: keep these edits, or discard them.", "The user decides. Silently applying recovered edits would be worse than losing them."],
]
r = table(ws, r, ["Event", "Behaviour", "Note"], edit, [30, 60, 56], wrap_cols=(2, 3), mark_col=1)
r += 1
r = note(ws, r, "The provisional-edit model built for the web application - snapshot, commit on Save, revert on "
                "Leave without change - turns out to be exactly the seam persistence needs. No part of it changes.")
r += 1

r = section(ws, r, "Re-importing over a workspace that already holds data")
r = lines(ws, r, [
    "This is the one genuinely new decision the change forces, and it is the one most likely to lose somebody's",
    "work if it is decided carelessly. Somebody else maintains the source workbook; you will have hand-entered",
    "assignments in the workspace. A silent replace destroys them, and a silent merge hides that it happened.",
])
r += 1
imp = [
    ["Replace everything", "Simple.", "Destroys hand-entered data with no warning.", "Rejected"],
    ["Merge by key, source wins", "Keeps rows the import does not mention.", "Silently changes rows you edited deliberately.", "Rejected"],
    [f"{MARK_NEW}Show the differences, decide per sheet", "Nothing is lost without being seen. The user keeps authority over their own data.", "One more dialog to build and to pass through.", "PROPOSED - NR-IMP-02, decision N-07"],
    ["Refuse; require a new workspace", "Impossible to lose anything.", "Makes the ordinary case - refreshed project dates - painful.", "Rejected"],
]
r = table(ws, r, ["Option", "For", "Against", "Status"], imp, [40, 44, 44, 22], wrap_cols=(2, 3), mark_col=1)
r = note(ws, r, "Q-N05 asks you to confirm this. It is worth a moment's thought: it is the one place in the "
                "design where the wrong choice quietly loses work.")
r = legend(ws, r)

# ---- 05a Sharing ----------------------------------------------------------
ws, r = sheet(wb, "05a_Sharing", "Sharing a plan between people",
              "Settled at review round 3: while someone is editing, no other session may update. "
              "Everyone may read, always.")

r = section(ws, r, "The rule")
r = lines(ws, r, [
    "    One writer at a time.  Any number of readers.  Nobody is ever blocked from LOOKING at a plan.",
    "",
    "The claim to write is made at the FIRST EDIT, not when the workspace is opened. That single choice is",
    "what separates this from a file lock, and it is worth being explicit about why:",
])
r += 1
why = [
    ["Claim on OPEN", "Simple to build.", "Somebody who opened a plan to glance at a figure, and then went to lunch, blocks everybody. Most sessions never edit anything, so most claims would be pointless."],
    [f"{MARK_NEW}Claim on FIRST EDIT", "Only a session that actually intends to change something takes the claim, so a plan is free unless somebody is really working on it. It also falls exactly on the point the application already has: the snapshot taken before the first pending edit. No new state, no new concept.", "The user learns they cannot edit one moment later than they would have. Mitigated by showing who holds the claim in the window from the moment the workspace opens, so it is never a surprise at the keystroke."],
]
r = table(ws, r, ["When the claim is made", "For", "Against"], why, [26, 62, 58], wrap_cols=(2, 3), mark_col=1)
r = note(ws, r, "The provisional-edit model built for the web application turns out to fit this exactly: "
                "'snapshot before the first pending edit' becomes 'claim before the first pending edit', and "
                "Save and Leave-without-change - which already commit and revert - become the two ways a claim "
                "is released. Nothing about the edit model changes.")
r += 1

r = section(ws, r, "The four states a session can be in")
states = [
    ["READING", "The workspace is open and nobody in this session has tried to change anything.", "Everything except changing the workspace: all tabs, figures, graphs, filters, horizon, exports, Save As.", "Entered on open. Leaves to EDITING on the first edit attempt."],
    ["EDITING", "This session holds the claim.", "Everything. This is the ordinary working state, identical to the single-user case.", "Entered by a successful claim. Leaves on Save, on Leave-without-change, or on close."],
    ["BLOCKED", "This session tried to edit and somebody else holds the claim.", "Everything READING allows. The attempted edit did not happen, and the window says who holds the plan and since when.", "Entered by a refused claim. Leaves to READING, and offers EDITING the moment the holder releases."],
    ["STALE", "Somebody else has saved this workspace since this session read it.", "Everything READING allows, with a standing offer to reload. Editing from a stale view is refused until it reloads.", "Entered when the file changes on disk. Leaves on reload."],
]
r = table(ws, r, ["State", "Meaning", "What the user can do", "How it is entered and left"],
          states, [12, 46, 50, 40], wrap_cols=(2, 3, 4))
r += 1
r = note(ws, r, "STALE deserves a word. Without it, somebody who opened a plan at 09:00 is still looking at "
                "09:00's figures at 11:00, after two saves by somebody else, with nothing on screen to say so - "
                "and may then quote them. Reading a shared plan is only safe if the application admits when what "
                "it is showing has been superseded.")
r += 1

r = section(ws, r, "How the claim is made, on a network share")
mech = [
    ["1", "Create a claim file beside the workspace, using an operation that FAILS if the file already exists.", "Not read-then-write: two sessions can both read 'no claim' and both then write one. Create-if-absent is decided by the filesystem, one winner, no race - and it works over SMB (NR-STO-13)."],
    ["2", "The file records user name, machine name, process, the application version, and the time.", "So a refusal can name a person rather than saying 'locked' (NR-STO-12). 'Kim is editing this, since 09:14, on PC-4471' is actionable; 'file in use' is not."],
    ["3", "While the session holds the claim it rewrites the time inside it, every 30 seconds.", "A heartbeat. It is how a live holder is told apart from a dead one, without asking the user to guess (NR-STO-14)."],
    ["4", "A claim whose heartbeat has not moved for 30 MINUTES expires and may be taken over.", "Your answer at Q-N16. Note that the heartbeat interval and the expiry are two different numbers - see the section below, which is where the value of keeping them apart shows."],
    ["5", "Saving re-reads the claim first, and refuses if it is no longer this session's.", "The last defence. If anything above went wrong - a clock, a takeover, a share that lied - the save stops instead of overwriting somebody's work."],
]
r = table(ws, r, ["Step", "Mechanism", "Why this way"], mech, [7, 62, 76], wrap_cols=(2, 3))
r += 1

r = section(ws, r, "Thirty minutes, and a thirty-second heartbeat   [round 4: Q-N16]")
r = lines(ws, r, [
    "The heartbeat interval and the expiry answer two different questions, and keeping them apart is what",
    "makes a long expiry comfortable rather than opaque:",
    "",
    "    HEARTBEAT, 30 seconds    'Is the holder still alive?'      Known within half a minute.",
    "    EXPIRY, 30 minutes       'May somebody else take it now?'  Your answer at Q-N16.",
    "",
    "So the application always knows the holder has gone quiet, long before anyone may act on it - and can",
    "say so, instead of showing the same message for a colleague who is mid-sentence and one whose laptop",
    "died twenty minutes ago:",
], mono=True)
r += 1
hb = [
    ["Holder active - heartbeat within the last 30 seconds", "\"Kim Min-jun is editing this plan (active now).\"", "Wait, or telephone them. The plan is genuinely in use."],
    ["Holder silent - heartbeat older than 30 seconds", "\"Kim Min-jun's session has not responded since 09:14. The plan becomes free at 09:44.\"", "The colleague can decide: wait 20 minutes, or call and ask. Both are informed choices."],
    ["Expired - heartbeat older than 30 minutes", "\"Kim Min-jun's session stopped responding at 09:14. You may take over.\"", "Take it over. The displacing session records whose claim it took."],
    [f"{MARK_NEW}Your own stalled claim, same name and machine", "\"This plan is held by your own earlier session. Take it back?\"", "Immediately, without waiting (NR-STO-19)."],
]
r = table(ws, r, ["Situation", "What the colleague is told", "What they can do"], hb, [42, 62, 42], wrap_cols=(2, 3), mark_col=1)
r = note(ws, r, "Thirty minutes is a long time to be locked out by a crash, and the last row is why it does not "
                "bite in the case that would annoy most: your own application falling over and then refusing to "
                "let you back into your own plan. Recovering your own session needs no waiting period at all.")
r += 1

r = section(ws, r, "Who is editing   [round 4: Q-N17]")
r = lines(ws, r, [
    "A thirty-minute expiry is only reasonable if the normal remedy is not waiting but ASKING. That is what",
    "your Q-N17 answer supplies: the application knows who people are, so a refusal can name someone and say",
    "how to reach them. The two answers work as a pair - neither is as good alone.",
])
r += 1
ident = [
    ["At launch, before any plan opens", "The application asks who you are. The first time it pre-fills your Windows account name and asks for the rest; afterwards it shows what it remembers and you confirm with one click. 'Not me' switches user, which matters on a shared PC.", "NR-USR-01..03"],
    ["What is recorded", "A name and a DEPARTMENT. Nothing else - no e-mail, no telephone. A name and a department are what a colleague needs to find somebody in the company directory, and they are two fields people will actually fill in.", "NR-USR-04"],
    ["Where the department comes from", "The departments already present in the open workspace's Person records, typed freely only where none matches. Person.department is in the source schema already, so an identity and a person record speak one vocabulary. Where the declared name matches a person in the plan, their own department is offered first.", "NR-USR-10, NR-USR-11"],
    ["Where it is kept", "In the user's own data folder - never in the shared application folder, where two people would overwrite each other.", "NR-USR-09"],
    ["What the claim carries", "The holder's name and contact details, so the message can be acted on. Removed when the claim ends.", "NR-USR-05"],
    ["What a blocked colleague sees", "A message naming the holder and their department, when they started, whether their session is still responding, and when the plan becomes free - with a button to copy the name and department. For example: \"Kim Min-jun (Data Management) has been editing this plan since 09:14, active now.\"", "NR-STO-12, NR-USR-06"],
    ["What else it improves", "A workspace records who last saved it and when, shown on opening. On a shared plan, whose figures these are matters as much as when they were made.", "NR-USR-07"],
]
r = table(ws, r, ["Aspect", "Behaviour", "Requirement"], ident, [34, 96, 16], wrap_cols=(2,))
r += 1
r = note(ws, r, "One thing must be said plainly, in the plan and on the screen where the name is typed: this "
                "identity is DECLARED, not verified. Anyone may type any name. It exists to answer 'who is "
                "editing this, and how do I reach them' - a question between colleagues - and it answers that "
                "well. It is not a login, it controls no access, and nothing in it should ever be relied on as "
                "evidence of who did what (NR-USR-08, R-N16). If an audit trail is ever needed, this is not one, "
                "and Q-N09 is where that would be settled.")
r += 1

r = section(ws, r, "How a claim ends   [NR-STO-15]")
ends = [
    ["The holder saves and closes the workspace.", "Released immediately. A waiting session is offered it without reopening anything."],
    ["The holder discards their edits (Leave without change).", "Released immediately - there is nothing to protect any more."],
    ["The holder closes the application.", "Released on the way out, including on an unexpected close where the application still gets to run."],
    ["The holder's session dies without warning.", "The heartbeat stops; after three missed intervals anybody may take it over."],
]
r = table(ws, r, ["Event", "What happens"], ends, [56, 90], wrap_cols=(1, 2))
r = note(ws, r, "The holder is NOT released merely by saving and continuing to work - that would hand the plan to "
                "somebody else mid-task. Saving keeps the claim; closing gives it up.")
r += 1

r = section(ws, r, "What this deliberately is not")
notdo = [
    ["Row-level or section-level locking.", "Two people editing different parts of one plan at the same time is a different product, and it needs a merge model the storage format does not have. The claim covers the whole workspace (N-20)."],
    ["Merging two people's changes.", "Out of scope, and it would change the file format. What is in scope is that no change is lost silently."],
    ["A queue of waiting editors.", "One waiting session is offered the claim when it frees. Beyond that, people can talk to each other."],
    ["An operating-system lock.", "A share may not honour one, and a lock left by a crash can need IT to clear it. The claim file plus a heartbeat fails softer and recovers by itself (N-21)."],
]
r = table(ws, r, ["Not this", "Why not"], notdo, [46, 100], wrap_cols=(1, 2))
r = note(ws, r, "The requirement you gave - block other sessions from updating while somebody is editing - is met "
                "in full by the whole-workspace claim. Everything above that line would be building a "
                "collaborative editor, which is a different tool.")
r = legend(ws, r)

# ---- 06 Decisions ---------------------------------------------------------
ws, r = sheet(wb, "06_Decisions", "Engineering decisions",
              "N-ids belong to this plan. C-01..C-11 of the web plan govern the calculation and are inherited "
              "unchanged.")

dec = [
    ["N-01", "The desktop shell is Electron, packaged without an installer requirement.", "Reuses a verified engine and a verified UI whole, and can be tested in this repository before delivery. Size was the argument against it, and your instruction removes that argument. Reversible: only shell/desktop/ would change.", "CONFIRMED at Gate N1"],
    ["N-02", "The workspace is prap-source-data JSON, not SQLite and not xlsx.", "Small data, an existing round-trip-tested format, and it keeps a workspace directly readable by an AI agent. See sheet 05.", "CONFIRMED at Gate N1"],
    ["N-03", "Workspace writes are atomic, with the previous versions retained.", "The cost of persistence is that a bad write now destroys data that used to live safely in Excel. Atomic rename plus history is the standard answer.", "CONFIRMED at Gate N1"],
    ["N-04", "Many workspaces, chosen by the user, with a recent list.", "You plan scenarios. A single implicit workspace would force overwriting one plan to explore another.", "CONFIRMED at Gate N1"],
    ["N-05", "One shared engine, two shells - never two implementations.", "Two implementations of the same arithmetic will diverge, and the divergence will be discovered by someone who trusted the wrong number.", "CONFIRMED at Gate N1"],
    ["N-06", "The HTML application is frozen in feature terms at its v1.0. Shared-engine fixes reach it; desktop-only features do not.", "It is finished and in service. Keeping it a build target protects it; keeping it feature-frozen keeps it from becoming a second thing to design.", "CONFIRMED at Gate N1"],
    ["N-07", "Re-import shows a difference and lets the user decide per sheet.", "The only option that cannot silently lose hand-entered work. See sheet 05.", "CONFIRMED at Gate N1"],
    ["N-08", "Pending edits are journalled separately from the committed workspace.", "So recovery can offer them back without a half-typed row ever becoming committed data.", "CONFIRMED at Gate N1"],
    ["N-09", "No network access of any kind, including update checks and crash reporting.", "The tool is offline by requirement. Anything that opens a socket makes it a different kind of product in the eyes of corporate IT.", "CONFIRMED at Gate N1"],
    ["N-10", "Engine parity between the applications is mandatory and tested; visual parity is best-effort and documented.", "The figures must agree or both tools are worthless. Whether a dialog looks native matters much less.", "CONFIRMED at Gate N1"],
    ["N-11", "The desktop application reuses the web UI rather than being redesigned.", "You have reviewed that UI over 25 rounds. Redesigning it would discard that and give you two habits to hold.", "CONFIRMED at Gate N1"],
    ["N-12", "A workspace from a newer application version is refused, not partially read.", "Reading unknown fields and writing them back out is how data quietly disappears.", "CONFIRMED at Gate N1"],
    ["N-13", "Application files and user data are separate folders under one parent, and an update replaces only the application files.", "Portability wants one folder; data safety forbids an update that can delete a plan. This layout gives both. Putting workspaces among the application files would eventually destroy somebody's work.", "CONFIRMED at Gate N1"],
    ["N-14", "The package is a plain folder that runs in place, delivered as a zip - not a single self-extracting executable.", "The single .exe unpacks about 180 MB to the temporary folder on every launch: slow to start, writes outside its own folder, and re-scanned by anti-virus each time. It only looks more portable. See sheet 04.", "CONFIRMED at Gate N1"],
    ["N-15", "No registry entry, no file association, no Start-menu entry, no AppData folder - ever.", "'Non-installed' means the machine is unchanged by the application's presence. Each of these would break that, and each is the kind of convenience that gets added without thinking. Stated as a decision so it has to be argued to be undone. CONFIRMED in substance at round 2: Q-N13 says the file association is not wanted.", "CONFIRMED at Gate N1"],
    ["N-16", "Where data lives is resolved in one fixed order, and the chosen location is shown in the About dialog.", "A portable application that guesses where its data went is a support problem. Four rules, always in the same order, always visible. The last of them creates a shortcut the user owns rather than hidden state in their profile.", "CONFIRMED at Gate N1"],
    ["N-17", "Workspace locking is an advisory marker file, not an operating-system lock.", "A network share may not support real locking, and an OS lock left by a crash can need IT to clear it. A marker anyone can read, and the holder's own machine can clear, fails softer - and the failure that matters is silent overwriting, which a marker prevents.", "CONFIRMED at Gate N1"],
    ["N-18", "Concurrent use is designed for unconditionally, not only if the shared folder is adopted.", "Q-N12 says 'could be considered'. The cost of building the marker anyway is one small file; the cost of not building it, and then adopting the share, is somebody's lost afternoon. This also supersedes Q-N06 - it no longer matters what the answer is.", "CONFIRMED at Gate N1"],
    ["N-19", "The write claim is made at the first EDIT, not when the workspace is opened.", "Claiming on open would let somebody who glanced at a plan and walked away block everyone else, and most sessions never edit anything. It also falls on a point the application already has - the snapshot before the first pending edit - so it costs no new concept. See sheet 05a.", "CONFIRMED at Gate N1"],
    ["N-20", "The claim covers the whole workspace, not a row or a section.", "Part-locking implies merging two people's edits, which the storage format cannot express and which is a different product. The requirement given - block other sessions from updating while somebody edits - is met in full by a whole-workspace claim.", "CONFIRMED at Gate N1"],
    ["N-21", "The claim is a file created by an operation that fails if it already exists, kept alive by a heartbeat.", "Read-then-write loses races; two sessions can both see 'free'. Create-if-absent has one winner, decided by the filesystem, and works over SMB. The heartbeat is what tells a live holder from a crashed one without asking a user to guess.", "CONFIRMED at Gate N1"],
    ["N-22", "Saving keeps the claim; only closing or discarding releases it.", "Releasing on save would hand the plan to somebody else in the middle of a working session, which is exactly when it must not move.", "CONFIRMED at Gate N1"],
    ["N-23", "The heartbeat interval (30 seconds) and the expiry (30 minutes) are separate numbers.", "They answer different questions - is the holder alive, and may somebody else take over. Keeping them apart lets the application distinguish a colleague who is mid-sentence from one whose laptop died, and say which, instead of showing the same message for both.", "CONFIRMED at Gate N1"],
    ["N-24", "A user may reclaim their own stalled claim immediately, without waiting out the expiry.", "Thirty minutes locked out of somebody else's plan is a policy; thirty minutes locked out of your OWN plan because your application crashed is just an obstruction. Same name and same machine is enough evidence for this purpose.", "CONFIRMED at Gate N1"],
    ["N-25", "Identity is declared by the user, not authenticated.", "There is no directory to check against in an offline, non-installed application, and building one would be a different product. A typed name answers 'who is editing this' completely, which is the question actually being asked. It answers no other question, and the plan says so rather than letting it be assumed (NR-USR-08).", "CONFIRMED at Gate N1"],
    ["N-26", "The name is pre-filled from the Windows account but stays editable.", "Free accuracy for the common case without a login system, and editable because a Windows account name is often not how colleagues know each other.", "CONFIRMED at Gate N1"],
    ["N-27", "The holder's name and department travel inside the claim, and are removed with it.", "A message that names somebody but not where to find them leaves the colleague exactly as stuck. Nothing personal outlives the claim it was written for.", "CONFIRMED at Gate N1"],
    ["N-28", "An identity is a name and a department, and nothing else.", "Your answer at Q-N18. E-mail and telephone are dropped rather than kept as optional fields: a field nobody fills in makes the message look incomplete when it is merely unused, and in a company a name and a department are enough to find anybody. It also matches Person.department, which the source schema already has, so identities and person records share one vocabulary instead of inventing a second.", "CONFIRMED at Gate N1"],
    ["N-29", "The application is 'Project Management APP' on screen, PM_APP as folder and executable.", "Your answer at Q-N11. The documents keep the PRAP_NewApp_ prefix so this plan stays findable beside the web application's; say if you would rather they were renamed to match.", "CONFIRMED at Gate N1"],
    ["N-30", "One previous version of a workspace is retained, and it is a setting rather than a constant.", "Your answer at Q-N07. Recorded with its consequence rather than silently: two bad saves in a row leave nothing good to return to (R-N19). Keeping it a setting means raising it later costs nothing, and the Excel export remains an archive no save can touch.", "CONFIRMED at Gate N1"],
    ["N-31", "Both distribution arrangements are first-class: run from the share, or copy from it and run locally.", "Your answer at Q-N14. Neither is treated as the fallback, both are tested, and the user guide states what each costs in launch time rather than recommending one and leaving the other undocumented.", "CONFIRMED at Gate N1"],
    ["N-32", "The application never encrypts anything itself; where files must be protected, the company's own protection is applied to the folder.", "Following from Q-N04. An application that encrypts its own data becomes the only thing that can read it - and when it will not start, the data is gone. What the application owes instead is a clear message when it meets a file it cannot read (NR-IMP-06).", "CONFIRMED at Gate N1"],
]
r_start = r
r = table(ws, r, ["ID", "Decision", "Rationale", "Status"], dec, [8, 62, 74, 14], wrap_cols=(2, 3))
last = r_start + len(dec)
for rr in range(r_start + 1, last + 1):
    ws.cell(row=rr, column=4).fill = NEW_FILL

dv = DataValidation(type="list", formula1='"CONFIRM,CONFIRMED,REJECTED,DEFERRED"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"D{r_start + 1}:D{last}")

r = note(ws, r, "All thirty-two are confirmed by the Gate N1 approval of 2026-08-13 rather than by a further "
                "review round - the same route decisions C-06..C-11 of the web application plan took at its "
                "own Gate 1. Any of them can still be changed, but from here it costs a version and a "
                "re-approval, and N-01 and N-02 become expensive to reverse once Gate N2 closes.")
r += 1
r = section(ws, r, "Inherited and unchanged")
r = note(ws, r, "C-01 weights multiply. C-02 partial months pro-rated by calendar days. C-03 a missing override "
                "means person_weight applies. C-04 1.00 FTE = 160 h/month. C-05 FTE is the display unit. "
                "C-06 over/under-allocation judged on the person total. C-07 weights seeded then editable. "
                "C-08 a month belongs to the period containing its first day. C-09 identifier edits cascade. "
                "C-10 boundaries recomputed on milestone change unless hand-edited. C-11 ordered clipping, "
                "empty periods omitted. None of these is reopened by this plan.")

# ---- 07 WBS ---------------------------------------------------------------
ws, r = sheet(wb, "07_WBS_Schedule", "Work breakdown - five steps, five gates",
              "The same method as the web application: each step ends at a review gate, and the next starts only "
              "after that gate is passed.")

wbs = [
    ["N1", "N1.1", "Draft this plan.", f"PRAP_NewApp_Development_Plan_v{DOC_VERSION}.xlsx", "Complete - issued for review"],
    ["N1", "N1.2", "Requester reviews across six rounds; all 18 questions answered.", "v0.6_reviewed mark-up", "Complete"],
    ["N1", "N1.3", "Apply the answers and re-issue, round by round.", "Plans v0.2 to v0.7", "Complete"],
    ["N1", "GN1", "GATE N1 CLOSED - approved by direction 2026-08-13. Decisions N-01..N-32 confirmed. Step N2 authorised.", "Plan v1.0", "Complete"],

    ["N2", "N2.1", "Split the source into core / ui / storage / shell-web, with a build step that re-emits app/PRAP.html. NO behaviour change.", "src/ + tools/build_app.py", "Complete"],
    ["N2", "N2.2", "Prove the refactor changed nothing: the rebuilt file is byte-identical, and all 13 existing suites pass against it unmodified.", "tools/test_layers.py + suite run", "Complete - byte-identical; 13/13 suites pass"],
    ["N2", "N2.3", "Specify the storage interface and the workspace file format.", "Specification sheets 03, 04", "Complete - issued for review"],
    ["N2", "N2.4", "Specify the desktop shell: window, menus, dialogs, recent files, drag-and-drop, About; and user identity.", "Specification sheets 05, 08", "Complete - issued for review"],
    ["N2", "N2.5", "Specify save safety, version retention, journalling and recovery; and the write claim.", "Specification sheets 06, 07", "Complete - issued for review"],
    ["N2", "N2.6", "Specify the re-import difference report, the look-without-a-workspace path, and protected files.", "Specification sheet 09", "Complete - issued for review"],
    ["N2", "N2.7", "Specify packaging, deployment, the folder layout, data resolution and update.", "Specification sheet 10", "Complete - issued for review"],
    ["N2", "N2.8", "Traceability matrix: every NR-id to a specification section, read from the plan rather than re-typed.", "Specification sheet 11", "Complete - 69/69 specified"],
    ["N2", "GN2", "GATE N2 CLOSED - specification approved 2026-08-13; all seven open points agreed. Step N3 authorised.", "PRAP_NewApp_Specification_v1.0.xlsx", "Complete"],

    ["N3", "N3.1", "Desktop-specific UI design: eight screens, built as a clickable prototype; then the real application wrapped in the desktop chrome and loaded with the 62-project dummy dataset, so the figures and charts can be reviewed as they will be.", "app/PM_APP_Prototype_v0.2.html", "Complete - issued for review"],
    ["N3", "N3.2", "List every deliberate divergence from the web UI, with its reason (NR-PAR-04).", "Component list sheet 03 - 9 divergences", "Complete - issued for review"],
    ["N3", "N3.3", "Requester reviews the prototype and the component list.", "v0.2_reviewed mark-up", "Complete"],
    ["N3", "GN3", "GATE N3 CLOSED - component list v1.0 approved 2026-08-13, with change C-N01. Step N4 authorised.", "PRAP_NewApp_Component_List_v1.0.xlsx", "Complete"],

    ["N4", "N4.1", "Desktop shell: application window, menu bar, native dialogs, window state.", "src/shell/desktop/main.js, preload.js, paths.js", "Complete - 23 checks in a running application"],
    ["N4", "N4.2", "Filesystem storage adapter: open, save, atomic write, version history.", "src/storage/desktop/workspace.js", "Complete - killed mid-write 12 times, always wholly old or wholly new"],
    ["N4", "N4.3", "Workspace lifecycle: new, open, save, save as, recent list, reopen last at launch.", "src/shell/desktop/main.js", "Complete"],
    ["N4", "N4.4", "Journalling and crash recovery.", "src/storage/desktop/workspace.js", "Complete - the storage half; the prompt is N4.4d"],
    ["N4", "N4.4b", "Data-location resolution, the About dialog that shows it, and the personal shortcut.", "src/shell/desktop/paths.js", "Complete - the shortcut is outstanding"],
    ["N4", "N4.4c", "The write claim: create-if-absent, the heartbeat, expiry and takeover, and the re-check before every save.", "src/storage/desktop/claim.js", "Complete - one winner from eight racing processes"],
    ["N4", "N4.4d", "The four session states on screen: who holds the plan, the refusal that names them, the offer when it frees, and the staleness notice with its reload.", "ui/", "Not started"],
    ["N4", "N4.4e", "User identity: the launch prompt, remembering and confirming it, switching user, contact details, and the blocked message that carries them.", "shell/desktop/ + ui/", "Not started"],
    ["N4", "N4.5", "Import into an occupied workspace: the difference view and the per-sheet decision.", "ui/", "Not started"],
    ["N4", "N4.6", "Parity suite: identical input through both applications, identical figures, findings and exports.", "tools/test_parity.py", "Not started"],
    ["N4", "N4.7", "Persistence suite: save, close, relaunch, verify; kill mid-write, verify; recover pending edits.", "tools/test_persist.py", "Not started"],
    ["N4", "N4.8", "Requester reviews against real data; refinements folded in.", "Updated code", "Not started"],
    ["N4", "GN4", "GATE N4 - application functionally complete.", "Project Management APP v0.9", "Not started"],

    ["N5", "N5.1", "Package as a plain Windows folder delivered in a zip; verify it launches with no installation and no administrator rights.", "PM_APP_v<ver>.zip", "Not started"],
    ["N5", "N5.2", "Verify no socket is opened, offline, for a full session (NR-SEC-01).", "Test evidence", "Not started"],
    ["N5", "N5.3", "Portability suite: run the folder, then copy it to another path and another PC and run it again; verify the recent list and settings survive (NR-DEP-08).", "tools/test_portable.py", "Not started"],
    ["N5", "N5.4", "Cleanliness check: record every file and registry key touched during a full session; verify nothing outside the application folder is written, and that deleting the folder leaves nothing behind (NR-DEP-06, NR-DEP-07).", "Test evidence", "Not started"],
    ["N5", "N5.5", "Update check: extract a newer zip over an existing folder; verify workspaces, backups and settings are untouched (NR-DEP-03).", "Test evidence", "Not started"],
    ["N5", "N5.6", "Awkward-location tests: a deeply nested path (R-N10), a read-only network folder (R-N11), and a UNC path (NR-DEP-13).", "Test evidence", "Not started"],
    ["N5", "N5.6b", "Shared-copy tests: separate settings per user; the claim race run many times with two sessions starting together, verifying exactly one wins; a killed holder's claim expiring and being taken over; a reader seeing a save and being offered the reload; a save refused after its claim was displaced; Save As while blocked. Repeated on a real share, because share behaviour belongs to the server (R-N14).", "tools/test_shared.py", "Not started"],
    ["N5", "N5.6c", "Measure launch time from a local disk and from a network folder, and state both in the user guide rather than promising one.", "Test evidence", "Not started"],
    ["N5", "N5.7", "Run on a real company PC - the only place execution policy, SmartScreen and anti-virus can be judged.", "Test evidence", "Pending you"],
    ["N5", "N5.8", "User guide: where to extract it, first launch and the SmartScreen prompt, workspaces, import and export, backup, updating, and what to do if it will not start.", "User guide", "Not started"],
    ["N5", "N5.9", "Full pass over the traceability matrix; version alignment across both product lines.", "Traceability matrix", "Not started"],
    ["N5", "GN5", "GATE N5 - release.", "Project Management APP v1.0 (PM_APP)", "Not started"],
]
r_start = r
r = table(ws, r, ["Step", "Task", "Activity", "Deliverable", "Status"],
          wbs, [7, 9, 94, 40, 26], wrap_cols=(3, 4))
last = r_start + len(wbs)

dv2 = DataValidation(type="list",
                     formula1='"Not started,In progress,Pending you,Complete - issued for review,Complete,Blocked"',
                     allow_blank=True)
ws.add_data_validation(dv2)
dv2.add(f"E{r_start + 1}:E{last}")

r = section(ws, r, "Progress")
prog = [
    ["Tasks total", f"=COUNTA(B{r_start + 1}:B{last})"],
    ["Complete", f'=COUNTIF(E{r_start + 1}:E{last},"Complete")+COUNTIF(E{r_start + 1}:E{last},"Complete - issued for review")'],
    ["Awaiting the requester", f'=COUNTIF(E{r_start + 1}:E{last},"Pending you")'],
    ["Not started", f'=COUNTIF(E{r_start + 1}:E{last},"Not started")'],
]
r = table(ws, r, ["Measure", "Count"], prog, [26, 12])
r += 1
r = note(ws, r, "Task N2.1 is deliberately first and deliberately dull. Until the engine is shared and the HTML "
                "file is proven unchanged by the sharing, every later task risks the finished product. No "
                "durations: the pace is set by review turnaround, as before.")

# ---- 08 Verification ------------------------------------------------------
ws, r = sheet(wb, "08_Verification", "How each claim will be proven",
              "Same standard as the web application: a claim is not made until something demonstrates it.")

ver = [
    ["The refactor changed nothing", "The rebuilt file is BYTE-IDENTICAL to the one that was verified, and the 13 existing suites pass against it unmodified. Identity is a stronger claim than equivalence and needs no argument.", "NR-PAR-03", "DONE at N2.2 - 13/13 pass, sha256 unchanged"],
    ["The two applications agree", "tools/test_parity.py - one dataset through both, comparing every person-month figure, every validation finding and both export files.", "NR-PAR-02", "Automated, in this repository"],
    ["Data survives closing", "tools/test_persist.py - import, save, close the application, relaunch, compare the model to what was saved.", "NR-STO-01", "Automated, in this repository"],
    ["A save is never half-written", "Kill the process during a write, relaunch, verify the workspace is wholly old or wholly new. Repeated across many timings.", "NR-STO-05", "Automated, in this repository"],
    ["Pending edits are recovered", "Kill the process with edits pending, relaunch, verify the prompt appears and both answers behave.", "NR-STO-07", "Automated, in this repository"],
    ["Re-import loses nothing", "Import a changed workbook over hand-entered data; verify every difference is shown and nothing changes without a decision.", "NR-IMP-02", "Automated, in this repository"],
    ["An old workspace still opens", "Open a workspace written by an earlier version; verify it loads and reports the upgrade.", "NR-DEP-03", "Automated, in this repository"],
    [f"{MARK_NEW}No network access", "Run packaged, offline, with socket activity observed for a full session.", "NR-SEC-01", "Automated on the build machine; repeated on the company PC"],
    [f"{MARK_CHG}It needs no installation", "Extract the zip, run the exe, use it, delete the folder - with no administrator rights at any point.", "NR-DEP-02", "Automated on the build machine; repeated by you at N5.7"],
    [f"{MARK_NEW}It leaves the machine unchanged", "Record every file and registry key touched during a full session; verify nothing outside the application folder is written and nothing survives deleting it.", "NR-DEP-06, NR-DEP-07", "Automated, at N5.4"],
    [f"{MARK_NEW}It survives being moved", "Copy the folder to another path and another PC; verify it runs and its settings and recent list still work.", "NR-DEP-08", "Automated, at N5.3"],
    [f"{MARK_NEW}An update cannot delete a plan", "Extract a newer zip over an existing folder; verify workspaces, backups and settings are untouched.", "NR-DEP-03", "Automated, at N5.5"],
    [f"{MARK_NEW}One shared copy serves several people", "Two users run one folder with different --data paths; verify their settings, recent lists and backups never meet.", "NR-DEP-11", "Automated, at N5.6b"],
    [f"{MARK_CHG}Only one session can write", "Two sessions attempt the first edit together, many times over; verify exactly one is granted it and the other is told who holds it. Then verify a save is refused if the claim was displaced beneath it.", "NR-STO-10, NR-STO-13", "Automated, at N5.6b"],
    [f"{MARK_NEW}Reading is never blocked", "With one session editing, verify another can open, filter, chart and export the same workspace.", "NR-STO-11, NR-STO-17", "Automated, at N5.6b"],
    [f"{MARK_CHG}A crash does not strand a plan", "Kill the holding session; verify the heartbeat stops within 30 seconds and is reported as silent, that the claim expires at 30 minutes and can then be taken over naming whose it was, and that the SAME user on the SAME machine can reclaim it at once.", "NR-STO-14, NR-STO-19, R-N13", "Automated, at N5.6b"],
    [f"{MARK_CHG}A blocked colleague can act", "Block a session and verify the message names the holder and their department, says whether they are still responding and when the plan frees, and copies name and department on request.", "NR-STO-12, NR-USR-05, NR-USR-06", "Automated, at N5.6b"],
    [f"{MARK_NEW}A look without a workspace", "Open a source workbook directly; verify every table and graph works, that nothing can be saved into it, and that no workspace or setting is created.", "NR-IMP-05", "Automated, at N4"],
    [f"{MARK_NEW}Re-import asks before it changes anything", "Import a changed workbook over hand-entered data; verify the update question comes first, that the difference report lists what would be added, changed and removed per sheet, and that declining leaves the workspace byte-identical.", "NR-IMP-02", "Automated, at N4.5"],
    [f"{MARK_NEW}A protected file says so", "Present an encrypted or unreadable workbook; verify the message names protection as the likely cause rather than reporting corruption.", "NR-IMP-06, R-N18", "Automated, at N4"],
    [f"{MARK_NEW}One version back, and no further", "Save twice; verify exactly one previous version is retained, that it can be restored, and that the older one is gone rather than silently accumulating.", "NR-STO-06", "Automated, at N5.5"],
    [f"{MARK_NEW}One vocabulary for departments", "Declare an identity against a workspace whose people carry departments; verify those departments are offered rather than typed, and that a name matching a person offers that person's department first.", "NR-USR-10, NR-USR-11", "Automated, at N5.6b"],
    [f"{MARK_NEW}Identities do not collide", "Two users on one shared copy; verify each keeps their own name and contact details and neither sees the other's.", "NR-USR-09", "Automated, at N5.6b"],
    [f"{MARK_NEW}Nobody reads superseded figures unknowingly", "Save from one session; verify the other notices, says so, and offers the reload rather than continuing to show the old numbers.", "NR-STO-16", "Automated, at N5.6b"],
    [f"{MARK_NEW}The share itself behaves", "Repeat the claim, heartbeat and staleness tests against the real network share, not a local folder.", "R-N14", "ONLY provable on your share, at N5.6b"],
    [f"{MARK_NEW}It is allowed to run at all", "Launch it on a company-managed PC and see whether execution policy, SmartScreen or anti-virus stops it.", "R-N01, R-N09", "ONLY provable by you - see the risk sheet"],
    ["The figures are right", "Inherited: the engine is already proven against the Python reference on all 1,225 person-months of the large dataset and 433 of the 10x10 dataset.", "Inherited", "Already done, and re-run by the parity suite"],
]
r = table(ws, r, ["Claim", "How it is proven", "Requirement", "Where"],
          ver, [34, 66, 14, 34], wrap_cols=(2, 4), mark_col=1)
r += 1
r = note(ws, r, "One row says ONLY PROVABLE BY YOU, and round 1 is the reason there is now only one: with no "
                "installer, the administrator-rights question answers itself and can be demonstrated here. "
                "Whether your machine will EXECUTE an unsigned folder-based application cannot be, from any "
                "environment but yours. That single row is why risk R-N01 is rated as it is.")
r = legend(ws, r)

# ---- 09 Version control ---------------------------------------------------
ws, r = sheet(wb, "09_Version_Control", "Version control across two product lines")

r = section(ws, r, "Artifacts")
vc = [
    ["Desktop plan", "PRAP_NewApp_Development_Plan_v<ver>.xlsx", "docs/", "Sheet 01 of this document"],
    ["Desktop specification", "PRAP_NewApp_Specification_v<ver>.xlsx", "docs/", "Its own version-history sheet"],
    ["Desktop component list", "PRAP_NewApp_Component_List_v<ver>.xlsx", "docs/", "Its own version-history sheet"],
    ["Desktop application", "PM_APP_v<ver>.zip, extracting to a PM_APP\\ folder", "dist/", "version.txt in the folder, plus the About dialog and window title"],
    ["Workspace file", "<user's own name>.prap", "data\\workspaces\\ by default, or wherever the user chooses", "Header block inside the file"],
    ["User settings", "settings.json", "data\\", "Written by the application; never carried in the zip"],
    ["Web plan and its documents", "PRAP_Development_Plan_v<ver>.xlsx and siblings", "docs/", "Unchanged - their own sheets"],
    ["Web application", "app/PRAP.html", "app/", "APP_VERSION constant, shown in the footer"],
    ["Shared source tree", "core/, ui/, storage/, shell/", "repository root", "Git history"],
    ["Shared data contract", "schema_version in the Config sheet; format_version in the JSON", "templates/", "Unchanged at 5 and 1"],
]
r = table(ws, r, ["Artifact", "Naming", "Folder", "Where its version is recorded"],
          vc, [26, 52, 24, 42], wrap_cols=(2, 4))
r += 1

r = section(ws, r, "Rules")
vr = [
    ["The two product lines version independently.", "Desktop v1.0 and web v1.0 are unrelated numbers. Nothing forces them to move together."],
    ["The DATA contract versions once, for both.", "Schema version 5 and JSON format_version 1 are shared. A change to either is a change to both products and needs approval on both plans."],
    ["An engine fix is released to both.", "Fix in core/, re-verify against both, re-issue the HTML file as a point version and the desktop package as a point version. The fix is recorded in both version histories."],
    ["A desktop feature does not reach the web application.", "Decision N-06. The web application is feature-frozen; adding to it would reopen a closed gate."],
    ["Neither plan supersedes the other.", "Both remain live documents. The web plan's status stays exactly what it is today."],
]
r = table(ws, r, ["Rule", "Consequence"], vr, [56, 90], wrap_cols=(1, 2))

# ---- 10 Risks -------------------------------------------------------------
ws, r = sheet(wb, "10_Risks", "Risks and assumptions")

r = section(ws, r, "Risks")
risks = [
    [f"{MARK_CHG}R-N01", "Corporate policy blocks the application from running at all - typically application allow-listing, or a rule against executables outside Program Files.", "Low", "HIGH",
     "LARGELY SETTLED 2026-08-15. It was the largest risk in this plan and it is now largely evidence: an unsigned executable, extracted from a zip into C:\\Users\\del09\\..., started and ran. There is no application allow-list standing in the way on that machine. Likelihood drops from Medium to Low; the impact stays HIGH because what remains would still be fatal - a policy that tolerates a 257 KB probe may yet quarantine a 142 MB application carrying a browser engine (R-N09). Proven for one machine and one user; a policy applied to a group could differ."],
    [f"{MARK_NEW}R-N02", "The refactor to share the engine breaks the finished web application.", "Medium", "High",
     "Task N2.2 gates everything else on the 13 existing suites passing unmodified against the rebuilt HTML file. If they do not pass, the refactor is wrong and nothing proceeds. This is why N2.1 is the first task and carries no other change."],
    [f"{MARK_NEW}R-N03", "A workspace file is corrupted and a plan is lost - data that used to be safe in Excel because the application never wrote to it.", "Low", "HIGH",
     "This risk is CREATED by the persistence requirement. Atomic writes (N-03), retained versions (NR-STO-06) and the fact that Excel export still exists as an independent archive. The user guide will say plainly: export to Excel for anything you cannot afford to lose."],
    [f"{MARK_NEW}R-N04", "The two applications drift apart and disagree about the same data.", "Low", "HIGH",
     "Prevented structurally by N-05 - there is one engine - and detected by the parity suite at N4.6, which compares figures rather than trusting the structure."],
    [f"{MARK_NEW}R-N05", "Chromium security advisories oblige periodic rebuilds of an application that is otherwise finished.", "High", "Low",
     "Accepted, and inherent to the option chosen at N-01. The application is offline and opens no remote content (NR-SEC-02), so the realistic exposure is a crafted workspace file. Rebuild when a package update matters; the source does not change."],
    [f"{MARK_NEW}R-N06", "The Windows build cannot be verified in the development environment, which is Linux.", "High", "Medium",
     "Structural, not incidental. Mitigation: automated tests run against the Linux build of the same source, and N5.3 - your test on a real company PC - is a named, gating task rather than an afterthought. Any claim about Windows behaviour will be labelled as unverified until you verify it."],
    [f"{MARK_NEW}R-N07", "Two products double the maintenance from here on.", "Medium", "Medium",
     "Reduced by the architecture - most maintenance lands in the shared core and serves both - and by N-06 freezing the web application's feature set. It is not eliminated. It is the price of the bilingual arrangement, and worth stating plainly."],
    [f"{MARK_NEW}R-N08", "Users are unsure which of the two applications to use, or work in both and diverge.", "Medium", "Medium",
     "A one-page note in the user guide: the desktop application for planning work you keep, the web application for a quick look on a machine where nothing may be installed. A workspace exports to Excel, which imports into either, so no work is trapped."],
    [f"{MARK_NEW}R-N09", "Windows SmartScreen warns about the executable, or anti-virus quarantines it, because it is unsigned and unknown.", "HIGH", "Medium",
     "Likely rather than possible - an unsigned executable arriving in a zip carries the mark-of-the-web and is treated as untrusted. Usually a one-time 'More info - Run anyway', which the user guide will show with a screenshot. Being non-installed does not help here and may hurt: unregistered software has no reputation. If anti-virus quarantines it outright, only IT can allow it - Q-N02 covers this. A code-signing certificate would remove the warning, and remains an IT purchase rather than a coding task (Q-N03)."],
    [f"{MARK_CHG}R-N10", "The application fails in a deeply nested folder because Windows path limits are exceeded.", "Low", "Medium",
     "MEASURED 2026-08-15 rather than guessed. The requester's own folder is C:\\Users\\del09\\DL-ClaudeCW\\project management app develop\\ - 64 characters to the application root, which leaves the deepest file inside the package at 107 of the 260 Windows allows: 153 characters of headroom, and the path contains spaces without trouble. The risk is real for somebody who extracts into a much deeper tree, so the user guide still states a recommended location and N5.6 still tests a deliberately deep path."],
    [f"{MARK_CHG}R-N11", "The application folder is put on a read-only or shared network folder, so it cannot keep its data beside itself.", "HIGH", "Medium",
     "Raised from Medium at round 2 - Q-N12 says this is under consideration, so it is now the expected case rather than an edge one. NR-DEP-09 (now Must) and the resolution order at NR-DEP-10 handle it: detected at launch, stated plainly, and the user chooses. Without that it would fail at the first Save, which is the worst possible moment to discover it."],
    [f"{MARK_NEW}R-N12", "Launched from a network folder, the application starts slowly, or Windows treats an executable on a UNC path as untrusted.", "Medium", "Medium",
     "A packaged browser engine is a lot of bytes to pull across a network at every launch, and Windows applies stricter zone rules to executables on network paths than on local ones. Measured rather than promised (NR-NFR-01), and the user guide will recommend the arrangement that is fastest: use the share to DISTRIBUTE the folder, copy it locally once, run it from there. Q-N14 asks which you intend."],
    [f"{MARK_CHG}R-N13", "A crash leaves a workspace claimed, and nobody can edit it.", "Medium", "Low",
     "The direct cost of NR-STO-10, and largely removed at round 3: the heartbeat in NR-STO-14 means a dead session's claim expires by itself after three missed intervals, rather than needing anybody to delete a file. What remains is a wait of a minute or two, and the displacing session says whose claim it took."],
    [f"{MARK_NEW}R-N14", "A network share's caching defeats the claim or the staleness check, so two sessions both believe they hold the plan.", "Low", "HIGH",
     "The one failure that would break the guarantee rather than inconvenience somebody. SMB client caching can delay both the visibility of a new file and a change to modification time. Three defences: the claim uses create-if-absent, which the server decides rather than the client; the heartbeat is re-read rather than remembered; and every save re-checks the claim before writing (step 5 on sheet 05a), so a lost race still stops short of overwriting. Must be tested on your actual share - N5.6b - because share behaviour is a property of the server, not of the application."],
    [f"{MARK_CHG}R-N16", "A declared name is mistaken for a verified one, and somebody relies on the tool to say who did what.", "Low", "Low",
     "Much reduced at round 6: Q-N09 confirms there is no internal validation or record-keeping obligation, so nothing formal rests on this identity. It remains stated at NR-USR-08 and on the screen where the name is entered, because the limit should be visible at the point it could mislead - but it is now a courtesy rather than a control."],
    [f"{MARK_NEW}R-N20", "The company will not let an executable ARRIVE, whatever its policy on running one.", "HIGH", "HIGH",
     "PROVEN 2026-08-15, and it is now the binding constraint. The probe ran happily once it was on the machine, but e-mail refused to carry a .exe at all on a general security check. Execution policy is satisfied; transport is not. Two things follow. First, no amount of work on the package fixes this - an Electron build is an .exe however it is zipped, and defeating a mail filter is not a solution anybody should want. Second, decision N-01 is genuinely reopened: the choice of shell is now governed by what can be delivered rather than by what is pleasant to build. See sheet 10a."],
    [f"{MARK_NEW}R-N18", "Company file protection encrypts a workbook or a workspace, and the application cannot read it.", "Medium", "Medium",
     "Raised from your Q-N04 answer, which is about decrypting a file before it can be imported. Where a document-security product encrypts files on a share, an ordinary application sees bytes it cannot parse. Two requirements follow: NR-IMP-06, so a protected file is reported as protected rather than as corrupt - the failure that would otherwise waste an afternoon - and NR-IMP-07, so the application never adds encryption of its own and never becomes the only thing that can open your data. Protection stays the company's, applied to the folder."],
    [f"{MARK_NEW}R-N19", "Two bad saves in a row leave nothing good to go back to.", "Low", "Medium",
     "The direct cost of the Q-N07 answer - one retained version rather than ten. Save a mistake, notice it, save again while fixing it, and the good version has been pushed out of history. Kept as a setting so it can be raised without a code change, and the Excel export remains an independent archive that no save can touch. The user guide will say plainly: export to Excel anything you cannot afford to lose."],
    [f"{MARK_NEW}R-N17", "A plan is stranded for up to 30 minutes after a crash, and somebody needs it now.", "Medium", "Low",
     "The cost of the Q-N16 answer, and much reduced by two things: the message states exactly when the plan becomes free, so nobody is left guessing, and it names the holder so the ordinary remedy is a telephone call. Your own crashed session is reclaimable at once (NR-STO-19), which removes the case that would otherwise be infuriating. The expiry is a setting, so it can be shortened later without a code change."],
    [f"{MARK_NEW}R-N15", "A blocked user takes a Save As copy, works in it, and the plan quietly forks in two.", "Medium", "Medium",
     "The cost of NR-STO-17, and worth accepting: the alternative is telling somebody they cannot work at all. Mitigated by the copy carrying its own name and its own import history (NR-STO-09), and by the user guide saying plainly what a copy is for - a scenario of your own, not a second master. A fork somebody chose is a better outcome than an edit somebody lost."],
]
r = table(ws, r, ["ID", "Risk", "Likelihood", "Impact", "Mitigation"],
          risks, [9, 56, 12, 10, 66], wrap_cols=(2, 5), mark_col=1)

r = section(ws, r, "Code signing, in plain terms   [asked at Q-N03]")
r = lines(ws, r, [
    "You asked what the purchase actually is. Here it is without the jargon.",
])
r += 1
sign = [
    ["What it is", "A digital seal applied to the .exe that says which ORGANISATION published it. Windows checks the seal when the file is run."],
    ["What happens without it", "Windows does not know who made the file, so SmartScreen shows 'Windows protected your PC - unknown publisher' the first time it is run. The user clicks 'More info' then 'Run anyway'. Anti-virus also treats an unknown publisher with more suspicion. This is R-N09."],
    ["What happens with it", "The warning names your company instead, and usually disappears entirely once the file has been seen a few times. Anti-virus is far less likely to quarantine it."],
    ["What is bought", "A CODE-SIGNING CERTIFICATE, from a certificate authority - DigiCert, Sectigo, GlobalSign and similar. Roughly USD 200-600 per year for the standard kind; more for the 'extended validation' kind, which removes the warning immediately rather than gradually. Treat these as indicative figures to check, not quotations."],
    ["Why the COMPANY has to buy it", "The authority verifies the ORGANISATION exists before issuing - company registration documents, a verifiable telephone number, that sort of thing. An individual cannot buy a certificate in the company's name, which is the whole point of it."],
    ["What else it needs", "Industry rules now require the signing key to live on a hardware token or in a cloud signing service, so somebody has to hold that token and run the signing step when a new version is built."],
    ["Who would own it", "Whoever owns software distribution or endpoint security - IT or InfoSec. It is a procurement request, not a development task."],
    ["Is it needed?", "NO - not to make the application work. Your Q-N02 answer is that a zip in your own folder can be extracted and run, so the application is usable unsigned. Signing removes friction and reduces the chance of anti-virus interference; it is a comfort measure, not a prerequisite. It is worth raising only if the SmartScreen prompt or a quarantine turns out to be a real nuisance at N5.7."],
]
r = table(ws, r, ["Question", "Answer"], sign, [30, 116], wrap_cols=(2,))
r = note(ws, r, "So: nothing needs buying now. If it later does, the request to IT is 'a code-signing certificate "
                "for internally distributed software, and somewhere to keep the token'. Q-N03 stays open only "
                "in the sense that you may want to ask; nothing in this plan waits for it.")
r += 1

r = section(ws, r, "Assumptions")
assum = [
    ["A-N01", "Data volume stays as assumed by the web plan - up to about 100 projects and 1,000 people.", "Inherited, standing"],
    [f"{MARK_CHG}A-N02", "Windows 10 or 11, 64-bit, is the only target. macOS and Linux are out of scope.", "CONFIRMED at review round 1"],
    [f"{MARK_NEW}A-N08", "The application is never installed: it is copied as a folder and run in place, and the machine is unchanged by its presence.", "CONFIRMED at review round 1"],
    [f"{MARK_CHG}A-N09", "The user can extract a zip and run an executable from their own folder, and nothing will later block or remove it.", "CONFIRMED IN PART 2026-08-15 - an unsigned .exe extracted from a zip ran on BOOK-R8USOPHQ11 and wrote beside itself. The clause about a LARGE package and anti-virus is still untested"],
    [f"{MARK_NEW}A-N11", "No internal validation, qualification or record-keeping obligation applies to this tool.", "CONFIRMED at Q-N09"],
    [f"{MARK_NEW}A-N12", "The shared folder, if used, is writable by the people using it.", "CONFIRMED at Q-N15"],
    [f"{MARK_NEW}A-N10", "If a shared network folder is used, it is reachable whenever the application is needed. A plan kept only on an unreachable share cannot be opened.", "Standing - the user guide will recommend keeping working copies locally"],
    [f"{MARK_NEW}A-N13", "Bulk editing outside the application is done by exporting, changing the file, and giving it to ONE person - the application's manager - who imports it. Every user does not import their own files.", "CONFIRMED at the Gate N3 review"],
    [f"{MARK_CHG}A-N03", "One person works on a given workspace at a time - but the application no longer RELIES on it, because a shared folder makes it unsafe to assume.", "WITHDRAWN as an assumption at round 2; enforced by NR-STO-10 instead"],
    ["A-N04", "Package size is unconstrained.", "CONFIRMED by your instruction, 2026-08-13"],
    ["A-N05", "Both applications remain in service indefinitely; neither is a migration path away from the other.", "CONFIRMED by your instruction, 2026-08-13"],
    ["A-N06", "The source Excel workbook continues to be maintained by hand, by someone other than the tool.", "Inherited A-04, standing"],
    ["A-N07", "The user can copy a package into their own profile or a network share and run it from there.", "To confirm - Q-N02"],
]
r = table(ws, r, ["ID", "Assumption", "Status"], assum, [9, 108, 26], wrap_cols=(2,), mark_col=1)
r = legend(ws, r)

# ---- 10a Delivery routes ---------------------------------------------------
ws, r = sheet(wb, "10a_Delivery", "Getting it onto the laptop   [R-N20]",
              "The probe ran, but e-mail refused to carry it. Execution is settled; delivery is not.")

r = section(ws, r, "What the two tests together showed")
r = lines(ws, r, [
    "  RUNNING an unsigned executable from a user folder    ALLOWED   (probe, 2026-08-15)",
    "  SENDING one through company e-mail                   REFUSED   (general security check)",
    "",
    "Those are different controls, and only the second is now in the way. It is also the one that no",
    "amount of work on the package can fix: an Electron build is an .exe however it is zipped, renamed",
    "or split.",
], mono=True)
r += 1
r = note(ws, r, "Working around the mail filter is not on this sheet and will not be. It is a control the "
                "company put there deliberately; going around it is the behaviour it exists to catch, and it "
                "would put the requester rather than the tool in the wrong.")
r += 1

r = section(ws, r, "The three routes")
routes = [
    ["A", "Ask IT for the sanctioned channel", "The software-distribution share, the internal artifact store, or an allow-list entry. Nothing is rebuilt - the package is finished and tested.", "Best fidelity, no work. Depends entirely on somebody else, and the answer may be no or slow.", "Ask first - it costs one conversation"],
    ["B", "Replace the Electron shell with a PYTHON one", "The interface and the engine do not change at all. Python does what Electron's main process does - files, folders, the claim - and serves the existing page to the browser. What travels is .py text, not an executable.", "Removes the executable from the problem. Needs Python on the laptop. Roughly 700 lines of new Python, no new engine.", "RECOMMENDED if A is slow or refused"],
    ["C", "Persistence in the web application instead", "Give the existing HTML file real storage, through the browser's own file access. It already arrives - it is a document.", "Fastest, and it satisfies NR-STO-01. But browser storage is per profile, so the shared-folder and claim model largely goes; and it is Edge/Chrome only.", "Partial - keeps the data, loses the sharing"],
]
r = table(ws, r, ["", "Route", "What it is", "What it costs", "Verdict"],
          routes, [4, 34, 62, 58, 30], wrap_cols=(3, 4, 5))
r += 1

r = section(ws, r, "Why route B is a swap rather than a rewrite")
r = lines(ws, r, [
    "Because the seam is already there. Specification sheet 03 defines thirteen storage operations, and",
    "preload.js already exposes exactly those to the page. A Python shell implements the same thirteen",
    "over HTTP instead of IPC; the page calls fetch() instead of ipcRenderer.invoke(). Nothing above that",
    "line moves.",
    "",
    "    core/    unchanged   the same JavaScript, the same figures, in the browser",
    "    ui/      unchanged   the same tabs, tables and charts the reviewer approved",
    "    storage/ NEW         workspace.py and claim.py, ported from the .js beside them",
    "    shell/   NEW         http.server, about 200 lines, serving the page and the thirteen calls",
])
r += 1
why = [
    ["Decision N-05 survives", "There is still ONE engine. The numbers stay in core/, in the browser. Python touches files and nothing else - if a calculation appears in the Python, something has gone wrong.", "N-05"],
    ["No pip install", "http.server, json, pathlib, zipfile and xml are all standard library. The xlsx reading and writing stays in the browser, where it is already dependency-free - so Python never needs openpyxl.", "NR-DEP-05"],
    ["The storage design carries over intact", "Atomic save, retained version, journal, the claim with its heartbeat and expiry - all of it is filesystem work, and all of it ports directly. The 33 checks in test_storage.mjs become 33 checks in Python.", "NR-STO-04..19"],
    ["The shared folder still works", "Which route C cannot offer. Real files, real claim files, several people on one plan - the whole of sheet 07 keeps its meaning.", "NR-STO-10"],
    ["What is lost", "A window of its own. The application would open in the browser, pointed at a local address, launched from a shortcut. It looks like the web application because it IS the web application - which is a fair trade for arriving at all.", "D-N01..D-N09"],
]
r = table(ws, r, ["Point", "Detail", "Ref"], why, [34, 96, 14], wrap_cols=(2,))
r = note(ws, r, "The Python engine that already exists - tools/prap_io.py, 876 lines, proven identical to the "
                "browser on 1,225 person-months - is NOT what route B uses. It stays what it is: the "
                "independent reference that checks the browser's arithmetic. Using it as the engine would "
                "create the second implementation N-05 exists to prevent.")
r += 1

r = section(ws, r, "What has to be known before route B is chosen")
q = [
    ["Is Python installed on the laptop, and which version?", "`python --version` at a command prompt. 3.9 or newer is enough. If it is absent, route B needs IT to install it - which is the same conversation as route A, but for a runtime that is far easier to approve than an unknown application."],
    ["What internal route exists for files that are not e-mail?", "SharePoint, Teams, an internal GitLab, a software share. Route B still has to arrive; .py is text rather than an executable, which most filters treat differently, but the sanctioned route is better than the tolerated one."],
    ["Would IT rather approve a runtime than an application?", "Usually yes, and it is worth asking in those terms. 'May I have Python' is a smaller request than 'may I run this 142 MB thing a contractor sent me'."],
]
r = table(ws, r, ["Question", "Why it decides the route"], q, [56, 90], wrap_cols=(1, 2))

# ---- 11 Open questions ----------------------------------------------------
ws, r = sheet(wb, "11_Open_Questions", "Questions for review round 1",
              "Please answer in the Answer column. Q-N02 is the one worth asking someone else before you answer.")

qs = [
    [f"{MARK_CHG}Q-N01", "Deployment", "Which operating systems must this run on?", "ANSWERED 2026-08-13: Windows only. Closed - A-N02, NR-DEP-01. macOS and Linux are out of scope.", "Windows. CLOSED"],
    [f"{MARK_CHG}Q-N02", "Deployment", "Can you extract a zip into your own folder and run the .exe inside it?", "ANSWERED 2026-08-13: yes - a zip saved in your own folder can be extracted and executed. On the rest - whether anything might later block or remove it - checking with IT is impractical at the moment, and you have chosen to proceed assuming no issue. That is recorded as ASSUMED rather than confirmed: assumption A-N09 and risk R-N01 stay OPEN, and N5.7 on a real company PC is the first moment anybody will actually know. The plan proceeds on your decision; it simply does not pretend to more certainty than exists.", "Yes for running it; the rest ASSUMED, not verified. Proceeding"],
    [f"{MARK_CHG}Q-N03", "Deployment", "Is a company code-signing certificate obtainable, and by whom?", "ROUND 6 asked for this in plainer terms - what is actually bought, and what is a code-signing certificate. Answered in full on sheet 10, 'Code signing, in plain terms'. The short version: it is a digital seal saying which company published the .exe; without it Windows shows an 'unknown publisher' warning once; it costs roughly USD 200-600 a year; only the company can buy it because the authority verifies the organisation; and NOTHING here needs it - the application works unsigned. Raise it only if the warning or an anti-virus quarantine turns out to be a nuisance in practice.", "Explained on sheet 10. Nothing waits on it - raise with IT only if the warning becomes a nuisance"],
    [f"{MARK_CHG}Q-N04", "Storage", "Where should workspaces live by default?", "ANSWERED 2026-08-13, and the answer is about ENCRYPTION rather than location: 'even in a shared network, there is a certain way to decrypt a file to be imported into the system'. READING TAKEN - a shared network is acceptable, and files there may carry company protection that has to be lifted before the application can read them. Applied as NR-IMP-06 (a protected file is reported as protected, not as corrupt), NR-IMP-07 and decision N-32 (the application adds no encryption of its own). Location itself falls out of Q-N15: the share is writable, so data\\users\\<name>\\ on the share is the default and nobody has to choose. PLEASE CONFIRM this reading - it is the one place in v0.7 where I have interpreted rather than transcribed.", "READING ADOPTED at Gate N1 - correctable under a v1.1 if it was not your meaning"],
    [f"{MARK_CHG}Q-N05", "Storage", "When you re-import a source workbook over a workspace that already holds your hand-entered data, what should happen?", "ANSWERED 2026-08-13: ask first whether existing data is to be updated at all, then give a DIFFERENCE REPORT so the user decides what the import may override. Confirms the proposal and adds the explicit first question. NR-IMP-02 rewritten accordingly.", "Ask first, then a difference report. CLOSED"],
    [f"{MARK_CHG}Q-N06", "Storage", "Will more than one person ever have the same workspace open at once?", "SUPERSEDED at round 2. Once a shared folder is on the table the safe answer is 'assume yes', and the cost of doing so is one small marker file. Handled by NR-STO-10 and decision N-18 whichever way it turns out. No answer needed.", "Superseded - no longer needs an answer"],
    [f"{MARK_CHG}Q-N07", "Storage", "How many previous versions of a workspace should be retained?", "ANSWERED 2026-08-13: one - the version immediately before the current one. Applied at NR-STO-06 and decision N-30, kept as a setting so it can be raised without a code change. One consequence recorded rather than left implicit: two bad saves in a row leave nothing good to return to (R-N19), so the Excel export remains the archive for anything that matters.", "One previous version. CLOSED"],
    [f"{MARK_CHG}Q-N08", "Deployment", "How will updates reach you?", "ANSWERED 2026-08-13: both - a copied file and a shared folder. Neither needs anything special of the package beyond what NR-DEP-03 already requires, since an update is an extraction over the application files and never touches data\\.", "Copied file and shared folder. CLOSED"],
    [f"{MARK_CHG}Q-N09", "Governance", "Does this tool fall under any internal validation or record-keeping obligation, now that it HOLDS data rather than only displaying it?", "ANSWERED 2026-08-13: no. Recorded as assumption A-N11. It settles R-N16 - with nothing formal resting on it, a declared user name is a courtesy rather than a control - and it keeps Step N5 to functional verification rather than qualification evidence.", "No obligation. CLOSED"],
    [f"{MARK_CHG}Q-N10", "Scope", "Should the application be able to open a source workbook for a quick look, without creating a workspace?", "ANSWERED 2026-08-13: yes. Added as NR-IMP-05. A workspace is therefore optional rather than mandatory - the question 'what does this file say?' can be answered without committing to a plan, which is also how somebody tries the application for the first time.", "Yes. CLOSED"],
    [f"{MARK_CHG}Q-N11", "Naming", "What should the application be called on screen?", "ANSWERED 2026-08-13: 'Project Management APP', or 'PM_APP'. Applied as NR-APP-08 and decision N-29 - the long form on screen and in the window title, PM_APP as the folder and executable name. The DOCUMENTS keep the PRAP_NewApp_ prefix so this plan stays findable beside the web application's; say if you would rather they were renamed to match.", "Project Management APP / PM_APP. CLOSED"],
    [f"{MARK_CHG}Q-N12", "Deployment", "Where will you keep the application folder?", "ANSWERED 2026-08-13: a shared network folder could be considered. Planned for as though it will be - NR-DEP-09..13, NR-STO-10, decisions N-16..N-18, risks R-N11..R-N13. Sheet 05 sets out how data and locking work in that arrangement.", "Shared network folder possible. CLOSED, with Q-N14 and Q-N15 following from it"],
    [f"{MARK_CHG}Q-N13", "Deployment", "Is it acceptable that double-clicking a .prap file in Explorer will NOT open the application?", "ANSWERED 2026-08-13: not needed, drag-and-drop is fine. Closed. NR-IMP-04 stands as reduced, decision N-15 is unchallenged, and no registry entry will be written for any reason.", "Not needed. CLOSED"],
    [f"{MARK_CHG}Q-N14", "Deployment", "Is the share where people RUN it from, or where they copy it from once?", "ANSWERED 2026-08-13: both are possible. Applied as NR-DEP-14 - both are supported, both are tested, and neither is second-class. What the user guide gains instead of a recommendation is a measurement: what each costs in launch time (N5.6c), so the choice can be made on figures rather than on advice.", "Both. CLOSED"],
    [f"{MARK_CHG}Q-N15", "Deployment", "Would the share be writable, or read-only?", "ANSWERED 2026-08-13: writable. This is the simpler outcome - each person gets data\\users\\<name>\\ under the application folder automatically (NR-DEP-15), nobody is asked where to put anything, and the shortcut at NR-DEP-12 becomes a convenience rather than the normal way in. The read-only path (NR-DEP-09) stays built, because a folder can be made read-only later without anybody telling the application.", "Writable. CLOSED"],
    [f"{MARK_CHG}Q-N16", "Sharing", "How long should a silent session hold a plan before others may take it over?", "ANSWERED 2026-08-13: 30 minutes. Applied at NR-STO-14, kept separate from the 30-second heartbeat so the application can still tell a live holder from a dead one throughout (N-23), and softened by NR-STO-19 - your own crashed session is reclaimable at once. It remains a setting rather than a constant.", "30 minutes. CLOSED"],
    [f"{MARK_CHG}Q-N17", "Sharing", "Should the application be able to say who is editing, and let a blocked colleague contact them?", "ANSWERED 2026-08-13: yes - ask for the user's name at launch, show who is editing in the message a blocked user sees, and give enough information to contact them. Applied as a new requirement area NR-USR-01..09, decisions N-25..N-27, and a section on sheet 05a. This is what makes the 30-minute expiry reasonable: the normal remedy becomes asking rather than waiting.", "Yes, with contact details. CLOSED"],
    [f"{MARK_CHG}Q-N18", "User identity", "Which details should the application record beside the name?", "ANSWERED 2026-08-13: the user's department. E-mail and telephone dropped rather than kept as optional. Applied at NR-USR-04, NR-USR-06, NR-USR-07 and decision N-28; NR-USR-10 and NR-USR-11 added so the department comes from the same vocabulary as Person.department rather than being typed afresh.", "Department. CLOSED"],
]
r_start = r
r = table(ws, r, ["ID", "Area", "Question", "Why it matters", "Answer"],
          qs, [9, 15, 62, 56, 40], wrap_cols=(3, 4, 5), mark_col=1)
last = r_start + len(qs)
for rr in range(r_start + 1, last + 1):
    ws.cell(row=rr, column=5).fill = INPUT_FILL
r = note(ws, r, "Nothing here blocks drafting the specification except Q-N02, Q-N05 and Q-N06 - those three change "
                "the design rather than fill it in. The rest can be answered while Step 2 proceeds.")
r = legend(ws, r)

# ---- 12 Review log --------------------------------------------------------
ws, r = sheet(wb, "12_Review_Log", "Review log and approval",
              "This document's own log. The web application's review log is unaffected and stays where it is.")

log = [
    ["1", "Instruction 2026-08-13", "Change the program from a web application to a separate program type, like other open-source applications.", "Accepted as the purpose of this plan.", "Sheet 02 scope; NR-APP-01..07; decision N-01 selects the shell technology; sheet 04 records why the engine is not rewritten.", "Closed"],
    ["2", "Instruction 2026-08-13", "Change the data flow to keep imported data even after the application closes.", "Accepted.", "NR-STO-01..09; sheet 05 sets out the workspace model, the write protocol and the recovery behaviour; decisions N-02, N-03, N-08; risk R-N03 records the new risk this creates.", "Closed"],
    ["3", "Instruction 2026-08-13", "Use both ways - the web application goes final and stays usable; keep the HTML and its documents.", "Accepted, and it shapes the architecture.", "Sheet 02 states what is kept and untouched. Decisions N-05, N-06 and N-10 make the arrangement maintainable: one engine, the web application feature-frozen, parity tested rather than assumed. Risk R-N07 records the cost.", "Closed"],
    ["4", "Instruction 2026-08-13", "Develop the new application under a separate plan with a different name.", "Accepted.", "This document, PRAP-NAPP-PLAN-001, in its own version line from v0.1 with its own gates N1..N5. The web application plan is not amended.", "Closed"],
    ["5", "Instruction 2026-08-13", "Distribution size would be allowed.", "Accepted - and it decides an open question.", "Recorded as assumption A-N04. It removes the only strong argument against Electron, so decision N-01 selects it on the grounds of verifiability instead. Sheet 04 keeps Tauri as a documented later option.", "Closed"],
    ["6", "-", "Not requester input - raised while drafting.", "Two claims in this plan cannot be verified from the development environment.", "Whether the package is allowed to run on a company PC (R-N01), and whether it installs without administrator rights (NR-DEP-02). Both are marked on sheet 08 as provable only by you, and Q-N02 asks you to check the first before Gate N1 rather than after Gate N5. Round 1 closed the second of the two - see item 8.", "Closed"],

    ["7", "Instruction 2026-08-13 (round 1)", "The new application should work well under Windows OS.", "Accepted - Q-N01 closed.", "Windows 10 / 11, 64-bit, is the only supported target: NR-DEP-01, assumption A-N02 confirmed, macOS and Linux moved to out of scope on sheet 02. Risk R-N06 stands - the development environment is Linux, so Windows behaviour is verified by you at N5.7 and every claim about it is labelled unverified until then.", "Closed"],
    ["8", "Instruction 2026-08-13 (round 1)", "It should be a non-installed application.", "Accepted, and it reaches further than one requirement.", "NR-DEP-02 rewritten from 'no administrator rights' to 'no installer at all'; NR-DEP-05..09 added; NR-APP-06 and NR-IMP-04 amended. Decisions N-13 (data folder separate from application files), N-14 (a plain folder, not a self-extracting exe) and N-15 (no registry, ever). Sheet 04 gains the packaging comparison, sheet 05 the folder layout. Three risks added: R-N09 SmartScreen and anti-virus, R-N10 path length, R-N11 read-only share. Six verification tasks added at Step N5.", "Closed"],
    ["9", "-", "Not requester input - found while applying item 8.", "Two consequences worth your explicit agreement.", "First, an update that extracted over the application folder would delete the user's workspaces if they lived among the application files - resolved by decision N-13 before it could ever be built. Second, no file association is possible without a registry entry, so double-clicking a .prap file will not open the application; raised as Q-N13 rather than decided silently.", "Closed"],
    ["10", "-", "Not requester input - found while applying item 8.", "Non-installed makes one thing HARDER, not easier.", "An unsigned, unregistered executable has no reputation with SmartScreen or anti-virus, so it is judged more harshly than installed software would be. Recorded as R-N09 with likelihood HIGH, and folded into Q-N02 so you ask IT the right question - about execution, not about installation rights.", "Closed"],

    ["11", "Q-N13 answer (round 2)", "File association not needed; drag-and-drop is fine.", "Accepted - closes the question with no change.", "NR-IMP-04 stands as reduced at v0.2, and decision N-15 is now unchallenged: no registry entry will be written for any purpose. The 'optional associate files action' floated at v0.2 is dropped rather than left hanging.", "Closed"],
    ["12", "Q-N12 answer (round 2)", "A shared network folder could be considered.", "Accepted, and planned for as though it will happen.", "'Could be considered' is enough: the cost of building for it now is one marker file and one resolution rule, and the cost of not building for it and then adopting it is somebody's lost work. NR-DEP-09 raised to Must; NR-DEP-10..13 and NR-STO-10..11 added; decisions N-16..N-18; risks R-N11 raised to HIGH likelihood, R-N12 and R-N13 added; sheet 05 gains two new sections; five verification tasks added at Steps N4 and N5.", "Closed"],
    ["13", "-", "Not requester input - found while applying item 12.", "A shared folder breaks the v0.2 layout in two places, and one of them silently destroys work.", "First, one settings file cannot serve two people - solved by the data-location resolution order at NR-DEP-10. Second, two people editing one workspace means the second Save discards the first, with nothing said - solved by the advisory marker at NR-STO-10. The second is the serious one: it is the only failure in this plan that loses work without anybody being told, which is why it is a Must and why it is built whether or not the share is adopted (N-18).", "Closed"],
    ["14", "-", "Not requester input - found while applying item 12.", "Where does a portable application remember its data location, if it cannot write beside itself?", "The usual answer - a file in the user profile - is the hidden state the non-installed rule exists to prevent. Resolved at N-16 by having the application offer to create a desktop shortcut carrying --data: visible, owned by the user, and undone by deleting it. The About dialog always shows which rule applied.", "Closed"],
    ["15", "Q-N06", "Not requester input - a consequence of item 12.", "The question no longer needs answering.", "Q-N06 asked whether two people would ever open one workspace at once. With concurrent use designed for unconditionally (N-18), the answer changes nothing. Marked superseded rather than left open.", "Closed"],

    ["16", "Instruction 2026-08-13 (round 3)", "Block other sessions from updating while someone starts editing.", "Accepted, and it settles the sharing model: single writer, many readers.", "One change of substance against v0.3 - the claim is made at the FIRST EDIT rather than at open, so reading is never blocked and a plan is only held while somebody is genuinely working on it. New sheet 05a carries the model in full. NR-STO-10 rewritten; NR-STO-12..18 added; decisions N-19 (claim on first edit), N-20 (whole workspace, not part of it), N-21 (create-if-absent plus heartbeat) and N-22 (saving keeps the claim, closing releases it).", "Closed"],
    ["17", "-", "Not requester input - noticed while applying item 16.", "The instruction fits the application's existing edit model exactly.", "'Snapshot before the first pending edit' becomes 'claim before the first pending edit', and Save and Leave-without-change - which already commit and revert - become the two ways a claim is released. Nothing in the edit model changes, which is the strongest sign the rule is the right shape.", "Closed"],
    ["18", "-", "Not requester input - a gap in v0.3 that item 16 makes plain.", "Blocking writers is not enough on its own: a READER can be misled just as badly.", "Somebody who opened a plan at 09:00 was still shown 09:00's figures after two saves by somebody else, with nothing to say so, and could quote them in good faith. NR-STO-16 added: a reading session notices the file changed beneath it and offers the reload rather than presenting figures it knows to be superseded.", "Closed"],
    ["19", "-", "Not requester input - raised while applying item 16.", "One failure mode belongs to the network, not to the application.", "SMB client caching can delay both the appearance of a new file and a change to a modification time, which could in principle let two sessions both believe they hold a plan. Three defences are built in - a claim the server decides rather than the client, a heartbeat re-read rather than remembered, and a re-check before every save - but the behaviour belongs to your file server. Recorded as R-N14 and marked on sheet 08 as provable only on your own share.", "Closed"],

    ["20", "Q-N16 answer (round 4)", "A silent session may hold a plan for 30 minutes.", "Accepted - twenty times the 90 seconds proposed, and coherent with the Q-N17 answer.", "NR-STO-14 set to 30 minutes. The heartbeat stays at 30 seconds and becomes a separate number (N-23), so the application still knows within half a minute whether a holder is alive, and can say 'active now' or 'silent since 09:14, free at 09:44' rather than one message for both. Kept as a setting, not a constant.", "Closed"],
    ["21", "Q-N17 answer (round 4)", "Ask for the user's name at log-in; show who is editing in the pop-up; let others contact that person from what the pop-up shows.", "Accepted, and it is what makes item 20 work.", "New requirement area NR-USR-01..09; decisions N-25 (declared, not authenticated), N-26 (pre-filled from the Windows account, editable) and N-27 (contact details travel in the claim and are removed with it); a section on sheet 05a; three verification rows. NR-STO-12 rewritten so the message names the holder, shows how to reach them, and says when the plan frees.", "Closed"],
    ["22", "-", "Not requester input - noticed while applying items 20 and 21.", "The two answers are stronger together than separately.", "A 30-minute expiry alone would be a long wait with no recourse. Naming the holder turns the normal remedy from waiting into asking, which is what makes the long expiry reasonable rather than obstructive. Recorded on sheet 05a so the reasoning is not lost if either is revisited on its own.", "Closed"],
    ["23", "-", "Not requester input - found while applying item 20.", "Thirty minutes has one genuinely infuriating case.", "Your own application crashes, you restart it, and you cannot edit your own plan for half an hour. NR-STO-19 and decision N-24 let a user reclaim a stalled claim held by their own name on their own machine at once. Everybody else waits.", "Closed"],
    ["24", "-", "Not requester input - raised while applying item 21.", "What a typed name is, and is not.", "It answers 'who is editing this, and how do I reach them' completely. It is not a login, it verifies nothing, and anyone can type any name. Stated at NR-USR-08, on the screen where the name is entered, and as risk R-N16 - because the danger is not the feature but somebody later relying on it as a record of who did what. If that is ever needed, Q-N09 is where it belongs.", "Closed"],

    ["25", "Q-N18 answer (round 5)", "Record the user's department beside their name.", "Accepted, and e-mail and telephone are dropped rather than kept as optional.", "NR-USR-04 rewritten to name and department only; NR-USR-06 and NR-USR-07 follow; decision N-28. Dropping the optional fields is the point of the answer, not an omission from it - a field nobody fills in makes the message look incomplete when it is merely unused, and in a company a name and a department find anybody.", "Closed"],
    ["26", "-", "Not requester input - noticed while applying item 25.", "The answer fits a column the data model already has.", "Person.department is in the source schema. So the identity's department is offered from the departments already in the open workspace rather than typed afresh (NR-USR-10), and where the declared name matches a person in the plan, their own department is offered first (NR-USR-11). Without this, 'Data Management' and 'DM' would both come into existence because one of them was typed at a login prompt.", "Closed"],

    ["27", "v0.6_reviewed (round 6)", "Q-N02: a zip can be saved in your own folder and executed. Confirming the rest with IT is difficult at the moment - assume no issue.", "Accepted as a decision to proceed, recorded as an assumption rather than a fact.", "A-N09 reworded from 'to confirm' to 'ASSUMED, not verified'; R-N01 stays OPEN with the same wording. The plan proceeds on your call - it simply does not claim more certainty than exists, and N5.7 remains the first moment anybody will know. Everything else in this plan has been demonstrated before it was claimed; this one cannot be, from here.", "Closed"],
    ["28", "v0.6_reviewed - Q-N03", "Not an answer: 'elaborate the question more easily. What IT purchase is needed? What is the company code-signing certificate?'", "Fair - the question assumed knowledge it should have supplied.", "Sheet 10 gains 'Code signing, in plain terms': what the seal is, what happens with and without it, what is bought and for roughly how much, why only the company can buy it, what else it needs, who would own it, and - the part that matters - that NOTHING here needs it. The application works unsigned; signing removes friction. Q-N03 rewritten in those terms.", "Closed"],
    ["29", "v0.6_reviewed - Q-N04", "'Even in a shared network, there is a certain way to decrypt a file to be imported into the system.'", "Read as being about file PROTECTION rather than about location - flagged for your confirmation.", "The reading taken: a shared network is acceptable, and files on it may carry company protection that must be lifted before an ordinary application can read them. NR-IMP-06 makes a protected file report itself as protected rather than as corrupt, which is the failure that would otherwise waste an afternoon; NR-IMP-07 and N-32 keep the application from adding encryption of its own, so it never becomes the only thing that can open your data. Risk R-N18. The location question itself is answered by Q-N15 instead. THIS IS THE ONE PLACE IN v0.7 WHERE I HAVE INTERPRETED RATHER THAN TRANSCRIBED - please confirm or correct it.", "Closed"],
    ["30", "v0.6_reviewed - Q-N05", "Ask whether to update existing data, and give a difference report so the user decides what may be overridden.", "Accepted - confirms the v0.3 proposal and adds an explicit first question.", "NR-IMP-02 rewritten: the import asks whether existing data is to be updated at all, then presents a difference report sheet by sheet. The 'ask first' step is the addition - v0.6 went straight to the differences, which presumes the user wanted an update at all.", "Closed"],
    ["31", "v0.6_reviewed - Q-N07", "Retain only one previous version.", "Accepted, with its consequence recorded rather than left implicit.", "NR-STO-06 changed from ten to one, decision N-30, kept as a setting. R-N19 added: save a mistake, notice it, save again while fixing it, and the good version has been pushed out of history. The Excel export remains an archive no save can touch, and the user guide will say so plainly.", "Closed"],
    ["32", "v0.6_reviewed - Q-N09", "No internal validation or record-keeping obligation.", "Accepted - and it settles an open risk.", "Assumption A-N11. R-N16 falls from Medium/Medium to Low/Low: with nothing formal resting on it, a declared user name is a courtesy rather than a control. It also keeps Step N5 to functional verification rather than qualification evidence.", "Closed"],
    ["33", "v0.6_reviewed - Q-N10", "Yes - open a source workbook for a look without creating a workspace.", "Accepted.", "NR-IMP-05. A workspace becomes optional rather than mandatory, which also gives somebody a way to try the application without committing to anything.", "Closed"],
    ["34", "v0.6_reviewed - Q-N11", "'Project Management APP' or 'PM_APP'.", "Accepted, with one thing to confirm.", "NR-APP-08 and N-29: the long form on screen and in the window title, PM_APP as folder and executable. Sheet 05's layout and NR-DEP-13 updated to match. The DOCUMENTS keep the PRAP_NewApp_ prefix so this plan stays findable beside the web application's - say if you would rather they were renamed too.", "Closed"],
    ["35", "v0.6_reviewed - Q-N14 and Q-N15", "Both ways of using the share; and the share would be writable.", "Accepted - and the writable answer is the simpler outcome.", "NR-DEP-14: both arrangements first-class, both tested, and the user guide states what each costs in launch time rather than recommending one. NR-DEP-15: on a writable share each person gets data\\users\\<name>\\ automatically, so nobody is asked where to put anything. The read-only path stays built - a folder can be made read-only later without telling the application.", "Closed"],
    ["36", "-", "Not requester input - the state of the document.", "Every question is answered.", "Eighteen questions across six review rounds, all answered. Two items were outstanding at v0.7 - the reading at item 29 and the naming at item 34 - and both are resolved by the Gate N1 approval below: the reading is adopted, and the documents keep their prefix. Both remain correctable under a v1.1.", "Closed"],
    ["37", "Gate N1", "'Go to next step', 2026-08-13.", "Recorded as approval of the plan as a baseline.", "v1.0 issued with no content change against v0.7. Decisions N-01..N-32 are confirmed by the approval itself rather than by a seventh review round - the same route the web application plan took at its own Gate 1, where C-06..C-11 were resolved at approval. Step N2 opens with the refactor, which is gated on the 13 existing suites passing unmodified against the rebuilt HTML file.", "Closed"],
]
r_start = r
r = table(ws, r, ["No.", "Source", "Input", "Response", "Action taken in v0.1", "Status"],
          log, [6, 22, 56, 40, 76, 12], wrap_cols=(3, 4, 5))
last = r_start + len(log)

dv3 = DataValidation(type="list", formula1='"Open,Accepted,Rejected,Deferred,Closed"', allow_blank=True)
ws.add_data_validation(dv3)
dv3.add(f"F{r_start + 1}:F{last}")

r = note(ws, r, "Every item is closed. Five of them - 9, 10, 19, 24 and 29 - were carried open through the "
                "review rounds and are closed at Gate N1 because their ACTIONS are complete: each was "
                "recorded as a requirement, a decision or a risk. What remains live from them lives on "
                "sheet 10, which is where a risk belongs; a review log records what was done, not what is "
                "still feared.")
r += 1

r = section(ws, r, "Approval - Gate N1")
appr = [["PRAP NewApp Development Plan v1.0", "Dan", "2026-08-13",
         "APPROVED - BASELINE, by direction ('go to next step'). Gate N1 is closed and Step N2 is "
         "authorised. The approval confirms decisions N-01 to N-32 on sheet 06, adopts the reading taken "
         "of the Q-N04 answer (review-log item 29), and leaves the documents named PRAP_NewApp_* rather "
         "than renaming them to match the application. The 69 requirements on sheet 03 are the contract "
         "for Steps N2 to N5. Any of these can still be changed - it costs a version and a re-approval, "
         "not a re-opened gate."]]
r_start2 = r
r = table(ws, r, ["Document", "Approver", "Date", "Decision"], appr, [34, 16, 14, 84], wrap_cols=(4,))
for cc in (1, 2, 3, 4):
    ws.cell(row=r_start2 + 1, column=cc).fill = NEW_FILL

r = note(ws, r, "STEP N1 IS COMPLETE. Work now proceeds under Step N2, whose first task is the refactor that "
                "gives both applications one engine - gated on the 13 existing suites passing unmodified "
                "against the rebuilt HTML file, so the finished web application cannot be damaged by it.")

wb.save(OUT)
print(f"Written: {OUT}")
