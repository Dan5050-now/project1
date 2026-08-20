"""A row with no identifier must not become a record called "null".

This is the bug it was written for. A project saved with every field filled in EXCEPT
project_id was indexed as M.projects[null] - which JavaScript turns into the string key
"null". The application then believed in a project called "null": it appeared in the
filter list, it became the selection, and the row filter `keep.has(r.project_id)`
compared the set {"null"} against the row's actual null and missed. So the table the user
had just typed into reported "No rows. Use + row to add one." Its milestones and periods
went looking for a parent named "null" and found none, and any row added under them was
stamped "null" too. A person with no person_id did the same, which is why the Assignments
and Weight overrides tables beneath them offered nothing to fill in.

What is checked, for a project and for a person:

  1. the row is still on screen after the save attempt - it is the row being repaired
  2. Save is REFUSED, and the refusal names the column to fill in
  3. no phantom record is created: the model stays empty, and nothing called "null"
     reaches the model or the filter
  4. the child sections stay locked rather than offering entry that would go nowhere
  5. supplying the identifier recovers completely - the record appears, the child
     sections unlock, and a child row entered there carries the real parent key
  6. a WORKBOOK that already contains a keyless row loads, reports it, and shows it,
     rather than hiding it where it can never be repaired

    python tools/test_nokey.py
"""

import pathlib
import sys
import tempfile

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

ROOT = pathlib.Path(__file__).resolve().parents[1]
APP = (ROOT / "app" / "PRAP.html").as_uri()
DUMMY = ROOT / "templates" / "PRAP_SourceData_Dummy_10x10_v1.2.xlsx"
CHROME = "/opt/pw-browsers/chromium-1194/chrome-linux/chrome"
TMP = pathlib.Path(tempfile.mkdtemp(prefix="prap_nokey_"))

fails = []


def check(ok, label, detail=""):
    print(f"  {'ok  ' if ok else 'FAIL'} {label}{'   ' + detail if detail else ''}")
    if not ok:
        fails.append(label)


def add_row(pg, loc, values):
    pg.locator(f"{loc} button[data-ins]").last.click()
    pg.wait_for_timeout(800)
    row = pg.locator(f"{loc} tbody tr td[data-col]").evaluate_all(
        "es => Math.max(...es.map(e => +e.dataset.row))")
    for col, v in values.items():
        cell(pg, loc, row, col, v)
    return row


def cell(pg, loc, row, col, v):
    """Type v into a cell. An empty v CLEARS it, which is how a keyless row is reached
    now that + row allocates an identifier for you - a user who types over the suggested
    id and leaves it blank, or a workbook that arrived with the column empty."""
    td = pg.locator(f"{loc} td[data-row='{row}'][data-col='{col}']")
    if td.count() == 0:
        return
    td.first.click()
    pg.wait_for_timeout(140)
    pg.keyboard.press("Control+A")
    if str(v) == "":
        pg.keyboard.press("Delete")
    else:
        pg.keyboard.type(str(v))
    pg.keyboard.press("Enter")
    pg.wait_for_timeout(320)


def rows_on(pg, loc):
    return pg.locator(f"{loc} tbody tr td[data-col]").evaluate_all(
        "es => new Set(es.map(e => e.dataset.row)).size")


def panel_of(pg, sheet, pane):
    """The panel a table lives in: how many rows it shows, and whether it offers entry."""
    return pg.evaluate("""a => {const t = document.querySelector(
        a.pane + " .data-t[data-sheet='" + a.sheet + "']");
        if (!t) return null;
        const p = t.closest('.panel');
        return {rows: new Set([...t.querySelectorAll('td[data-col]')].map(e => e.dataset.row)).size,
                ins: p.querySelectorAll('button[data-ins]').length,
                why: (p.querySelector('td.muted[colspan]') || {}).textContent || ''};}""",
                       {"sheet": sheet, "pane": pane})


with sync_playwright() as pw:
    browser = pw.chromium.launch(executable_path=CHROME)
    pg = browser.new_page(viewport={"width": 1500, "height": 950})
    errors = []
    pg.on("pageerror", lambda e: errors.append(str(e)))
    pg.on("dialog", lambda d: d.accept())

    print("app/PRAP.html — a row with no identifier")
    pg.goto(APP)
    pg.wait_for_timeout(300)
    pg.click("#startBtn")
    pg.wait_for_timeout(2200)

    # ---- the project ---------------------------------------------------------
    pg.click("text=Source data (project)")
    pg.wait_for_timeout(900)
    P = "#t-proj .data-t[data-sheet='Project']"
    row = add_row(pg, P, {"project_id": "", "project_name": "Trial One",
                          "project_type": "NewDrug CT", "clinical_phase": "Phase 2",
                          "start_date": "2027-01-01", "end_date": "2029-06-30"})
    pg.click("#saveBtn")
    pg.wait_for_timeout(1600)
    banner = pg.inner_text("#banner")
    model = pg.evaluate("() => ({projects: Object.keys(S.model.projects), sel: S.selProj, "
                        "raw: S.model.raw.Project.length})")

    check(rows_on(pg, P) == 1,
          "the project row is still on screen after the save attempt",
          f"{rows_on(pg, P)} row(s), {model['raw']} in the data")
    check("refused" in banner.lower() and "project_id" in banner,
          "Save is refused, and the refusal names the column to fill in",
          banner.strip()[:96])
    check(model["projects"] == [] and model["sel"] is None,
          "no phantom record is created — nothing called 'null' reaches the model",
          f"projects={model['projects']}, selected={model['sel']}")
    ms = panel_of(pg, "Milestone", "#t-proj")
    check(ms["ins"] == 0 and "project_id" in ms["why"],
          "the child sections stay locked rather than offering entry that goes nowhere",
          ms["why"][:70])

    cell(pg, P, row, "project_id", "PRJ-001")
    pg.click("#saveBtn")
    pg.wait_for_timeout(1700)
    model = pg.evaluate("() => ({projects: Object.keys(S.model.projects), sel: S.selProj})")
    ms = panel_of(pg, "Milestone", "#t-proj")
    check(model["projects"] == ["PRJ-001"] and model["sel"] == "PRJ-001" and ms["ins"] == 1,
          "supplying project_id recovers completely — the record appears and the "
          "child sections unlock",
          f"projects={model['projects']}, Milestones offers + row={ms['ins'] == 1}")

    add_row(pg, "#t-proj .data-t[data-sheet='Milestone']",
            {"milestone_name": "CTA submission", "milestone_date": "2027-03-01"})
    parents = pg.evaluate("S.model.raw.Milestone.map(r => r.project_id)")
    check(parents == ["PRJ-001"],
          "a child row entered afterwards carries the real parent key, not 'null'",
          f"{parents}")

    # ---- the person ----------------------------------------------------------
    pg.click("text=Source data (person)")
    pg.wait_for_timeout(900)
    PER = "#t-pers .data-t[data-sheet='Person']"
    prow = add_row(pg, PER, {"person_id": "", "person_name": "Alex R.",
                             "department": "Data Management",
                             "primary_role": "Lead data manager", "capacity_fte": "1.00"})
    pg.click("#saveBtn")
    pg.wait_for_timeout(1600)
    banner = pg.inner_text("#banner")
    people = pg.evaluate("Object.keys(S.model.people)")
    asg = panel_of(pg, "Assignment", "#t-pers")
    check(rows_on(pg, PER) == 1 and "refused" in banner.lower() and "person_id" in banner
          and people == [],
          "a person with no person_id behaves the same way",
          f"{rows_on(pg, PER)} row(s) on screen, people={people}")
    check(asg["ins"] == 0 and "person_id" in asg["why"],
          "Assignments stays locked while the person has no identifier", asg["why"][:70])

    cell(pg, PER, prow, "person_id", "PSN-001")
    pg.click("#saveBtn")
    pg.wait_for_timeout(1700)
    people = pg.evaluate("Object.keys(S.model.people)")
    asg = panel_of(pg, "Assignment", "#t-pers")
    ppw = panel_of(pg, "PersonPeriodWeight", "#t-pers")
    check(people == ["PSN-001"] and asg["ins"] == 1 and ppw is not None,
          "supplying person_id recovers completely — Assignments and Weight overrides "
          "both become enterable",
          f"people={people}, Assignments + row={asg['ins'] == 1}")

    # ---- 6. a workbook that already carries a keyless row ---------------------
    broken = TMP / "keyless.xlsx"
    wb = load_workbook(DUMMY)
    wb["Project"].cell(row=2, column=1).value = None          # blank the first project_id
    wb.save(broken)
    pg.goto(APP)
    pg.wait_for_timeout(300)
    pg.set_input_files("#picker", str(broken))
    pg.wait_for_timeout(4500)
    found = pg.evaluate("""() => {
      const f = S.model.findings.filter(x => x.msg.includes('no project_id'));
      return {n: f.length, sev: (f[0] || {}).sev,
              keys: Object.keys(S.model.projects).filter(k => k === 'null' || k === 'undefined')};}""")
    pg.click("text=Source data (project)")
    pg.wait_for_timeout(1100)
    onscreen = pg.evaluate("""() => {const t = document.querySelector(
        "#t-proj .data-t[data-sheet='Project']");
        return [...t.querySelectorAll("td[data-col='project_name']")]
          .some(td => td.textContent.trim() === 'ONV-101 Phase 1');}""")
    check(found["n"] == 1 and found["sev"] == "error" and not found["keys"] and onscreen,
          "a workbook that already carries a keyless row reports it and shows it",
          f"{found['n']} finding(s) at {found['sev']}, phantom keys {found['keys']}, "
          f"row on screen={onscreen}")

    check(not errors, "no uncaught errors in the page", "; ".join(errors[:2]))
    browser.close()

print()
print("FAILURES: " + (", ".join(fails) if fails else "none"))
sys.exit(1 if fails else 0)
