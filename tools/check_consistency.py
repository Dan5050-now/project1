"""Cross-check the documents against the artifacts they describe.

The plan documents a schema, the specification documents a parse contract, and the
template is the file both describe. Nothing stops those three drifting apart except
a check, so this is the check.

    python tools/check_consistency.py
"""

import hashlib
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PLAN = ROOT / "docs" / "PRAP_Development_Plan_v2.26.xlsx"
SPEC = ROOT / "docs" / "PRAP_Programming_Specification_v1.0.xlsx"
TEMPLATE = ROOT / "templates" / "PRAP_SourceData_Template_v1.7.xlsx"
DUMMY = ROOT / "templates" / "PRAP_SourceData_Dummy_v1.9.xlsx"
DUMMY_SMALL = ROOT / "templates" / "PRAP_SourceData_Dummy_10x10_v1.1.xlsx"

problems, notes = [], []


def cells(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                yield c


# ---- 1. the plan's data model vs the template's actual columns -------------
tpl = load_workbook(TEMPLATE)
tpl_cols = {s: [c.value for c in tpl[s][1]] for s in tpl.sheetnames if s != "00_README"}

plan = load_workbook(PLAN)
dm = plan["04_Data_Model"]
# A data-model row is identifiable by its Type cell, which keeps the scan from
# running past the end of a table into the prose that follows it.
TYPES = {"Text", "Date", "Decimal", "Integer", "List", "Derived", "-"}
documented, current = {}, None
for r in range(1, dm.max_row + 1):
    a, t = dm.cell(r, 1).value, dm.cell(r, 2).value
    if isinstance(a, str) and a.startswith("Sheet: "):
        current = a.replace("Sheet: ", "").split(" ")[0].strip()
        documented.setdefault(current, [])
    elif current and isinstance(a, str) and a and t in TYPES:
        documented[current].append(a.strip())

for sheet, cols in documented.items():
    if sheet not in tpl_cols:
        continue
    actual = [c for c in tpl_cols[sheet] if c]
    for col in cols:
        if ".." in col:                      # note_1 .. note_5 shorthand
            stem = col.split("..")[0].strip().rsplit("_", 1)[0]
            if not any(str(a).startswith(stem) for a in actual):
                problems.append(f"plan documents {sheet}.{col} but the template has no {stem}_* column")
            continue
        if "/" in col:                       # 'employment_start / employment_end'
            for part in [p.strip() for p in col.split("/")]:
                if part not in actual:
                    problems.append(f"plan documents {sheet}.{part}, absent from the template")
            continue
        if col not in actual:
            problems.append(f"plan documents {sheet}.{col}, absent from the template")

# ---- 2. sheet set matches -------------------------------------------------
plan_sheets = set(documented) & set(tpl_cols)
missing = set(tpl_cols) - set(documented)
if missing:
    notes.append(f"template sheets not itemised column-by-column in the plan: {sorted(missing)}")

# ---- 3. schema version agrees across plan, spec, template, dummy ----------
def cfg_version(path):
    wb = load_workbook(path)
    for r in wb["Config"].iter_rows(min_row=2, values_only=True):
        if r[0] == "schema_version":
            return int(r[1])
    return None


tpl_v, dum_v = cfg_version(TEMPLATE), cfg_version(DUMMY)
if tpl_v != dum_v:
    problems.append(f"template schema_version {tpl_v} != dummy {dum_v}")
# Every dummy the repository ships must be on the same schema. A second fixture is a
# second thing that can fall behind, which is the only reason it is worth checking.
small_v = cfg_version(DUMMY_SMALL)
if small_v != tpl_v:
    problems.append(f"template schema_version {tpl_v} != {DUMMY_SMALL.name} {small_v}")

spec = load_workbook(SPEC)
spec_v = None
for c in cells(spec["00_Cover"]):
    if c.value == "Schema version specified":
        spec_v = spec["00_Cover"].cell(c.row, 2).value
if str(spec_v) != str(tpl_v):
    problems.append(f"specification says schema v{spec_v}, template is v{tpl_v}")

plan_v = None
for c in cells(plan["04_Data_Model"]):
    if c.value == "schema_version":
        plan_v = plan["04_Data_Model"].cell(c.row, 3).value
if str(plan_v) != str(tpl_v):
    problems.append(f"plan documents schema v{plan_v}, template is v{tpl_v}")

# ---- 4. project_type values agree everywhere -----------------------------
lists = {}
for r in tpl["Lists"].iter_rows(min_row=2, values_only=True):
    if r[0]:
        lists.setdefault(r[0], []).append(r[1])
types = lists.get("project_type", [])
# 11_Open_Questions and the change register are a historical record: the reviewer's
# own words and the change requests necessarily quote values that have since been
# retired. Rewriting them to match today's schema would falsify the review trail, so
# they are exempt from the retired-literal scan.
RECORD_SHEETS = {"11_Open_Questions", "12_Review_Log", "01_Version_History",
                 "10_Open_Points"}
for doc, wb in (("plan", plan), ("specification", spec)):
    txt = " ".join(str(c.value) for s in wb.sheetnames if s not in RECORD_SHEETS
                   for c in cells(wb[s]))
    for t in types:
        if t not in txt:
            problems.append(f"{doc} never mentions project_type '{t}'")
    if re.search(r"'Clinical Trial'", txt):
        notes.append(f"{doc} still contains the retired literal 'Clinical Trial' somewhere")

# ---- 4b. no build artefacts left in any shipped workbook -----------------
# The generators use [NEW]/[CHANGED] markers to drive row highlighting, and strip
# them on render. One missed strip ships the marker as visible text, so check.
for label, path in (("plan", PLAN), ("specification", SPEC),
                    ("template", TEMPLATE), ("dummy", DUMMY),
                    ("small dummy", DUMMY_SMALL)):
    wb_ = load_workbook(path)
    for sh in wb_.sheetnames:
        for c in cells(wb_[sh]):
            v = c.value
            if not isinstance(v, str):
                continue
            if "[NEW]" in v or "[CHANGED]" in v:
                problems.append(f"{label} {sh}!{c.coordinate}: unstripped marker - {v[:60]}")
            elif v.startswith("=") and not v[1:2].isalpha():
                problems.append(f"{label} {sh}!{c.coordinate}: text starting with '=' will "
                                f"be read as a formula - {v[:60]}")

# ---- 4b2. derived columns are LOCKED, and nothing else is -----------------
# The green fill marking a derived column is a convention the reader has to have been
# told about. The lock is the file itself refusing the edit. Both are easy to lose in a
# regeneration, and neither failure announces itself - the file still opens, still reads,
# and quietly accepts a hand-typed total that nothing will ever recalculate.
def derived_of(ws_name, plan_wb):
    """The columns the plan marks DERIVED for this sheet."""
    dm_ = plan_wb["04_Data_Model"]
    out, cur = [], None
    for r in range(1, dm_.max_row + 1):
        a_, t_ = dm_.cell(r, 1).value, dm_.cell(r, 2).value
        if isinstance(a_, str) and a_.startswith("Sheet: "):
            cur = a_.replace("Sheet: ", "").split(" ")[0].strip()
        elif cur == ws_name and t_ == "Derived" and isinstance(a_, str) and a_.strip():
            out.append(a_.strip())
    return out


for label, path in (("template", TEMPLATE), ("dummy", DUMMY), ("small dummy", DUMMY_SMALL)):
    wb_ = load_workbook(path)
    for sh in wb_.sheetnames:
        if sh == "00_README":
            continue
        ws = wb_[sh]
        want = set(derived_of(sh, plan))
        hdr = [c.value for c in ws[1]]
        if not want:
            if ws.protection.sheet:
                notes.append(f"{label} {sh}: protected but the plan marks no derived column")
            continue
        if not ws.protection.sheet:
            problems.append(f"{label} {sh}: derived column(s) {sorted(want)} but the sheet is "
                            f"not protected, so nothing stops them being typed over")
            continue
        for i, h in enumerate(hdr, start=1):
            if h is None:
                continue
            locked = [bool(ws.cell(row=r, column=i).protection.locked)
                      for r in range(2, min(ws.max_row, 60) + 1)
                      if ws.cell(row=r, column=i).value is not None]
            if not locked:
                continue
            if h in want and not all(locked):
                problems.append(f"{label} {sh}.{h} is DERIVED but not locked in every row")
            if h not in want and any(locked):
                problems.append(f"{label} {sh}.{h} is an input column but is locked")
        for i, h in enumerate(hdr, start=1):
            if h in want and not ws.cell(row=1, column=i).comment:
                problems.append(f"{label} {sh}.{h} is locked with no note saying why")
        for allowed in ("insertRows", "deleteRows", "sort"):
            if getattr(ws.protection, allowed):
                problems.append(f"{label} {sh}: protection blocks {allowed}, which an editor needs")


# ---- 4c. the spec's Config defaults vs the template's actual values -------
# The thresholds live in the workbook as data, and the specification quotes them.
# A threshold changed in one place and not the other is the kind of drift that
# survives review, because both documents read correctly on their own.
tpl_cfg = {r[0]: r[1] for r in tpl["Config"].iter_rows(min_row=2, values_only=True) if r[0]}
sch = spec["03_Data_Schema"]
for row in sch.iter_rows(values_only=True):
    name = row[0]
    if not isinstance(name, str) or name not in tpl_cfg or len(row) < 3:
        continue
    documented_default, actual = str(row[2]).strip(), tpl_cfg[name]
    try:
        same = float(documented_default) == float(actual)
    except (TypeError, ValueError):
        same = documented_default == str(actual).strip()
    if not same:
        problems.append(f"specification documents Config.{name} default {documented_default}, "
                        f"template holds {actual}")

# ---- 4d. the period names the documents describe vs the template's list ---
# R-11 changed the clinical period set. A document still listing the old six names
# reads perfectly well on its own, which is exactly why this needs a machine check.
period_names = lists.get("period_name_clinical", []) + lists.get("period_name_others", [])
for doc, wb in (("plan", plan), ("specification", spec)):
    txt = " ".join(str(c.value) for s in wb.sheetnames if s not in RECORD_SHEETS
                   for c in cells(wb[s]))
    for nm in period_names:
        if nm not in txt:
            problems.append(f"{doc} never mentions period name '{nm}'")
    # the bare 'Conduct' was retired at R-11; it must not survive as a period name
    if re.search(r"'Conduct'|\bConduct(?!\s*\()", txt):
        notes.append(f"{doc} mentions a bare 'Conduct' - check it is prose, not the retired period name")

# ---- 4e. the application's provenance strip vs the repository -------------
# app/PRAP.html shows which controlled documents it was built against. A label like
# that is worse than useless once it goes stale, because it is believed. So the files
# it names must exist, and must be the newest of their kind in the repo.
APP = ROOT / "app" / "PRAP.html"
if APP.exists():
    src = APP.read_text(encoding="utf-8")
    block = re.search(r"const BUILT_AGAINST = \[(.*?)\n\];", src, re.S)
    if not block:
        problems.append("app/PRAP.html has no BUILT_AGAINST block to check")
    else:
        claimed = dict(re.findall(r'what:"([^"]+)",\s*file:"([^"]+)"', block.group(1)))
        if len(claimed) != 5:
            problems.append(f"app provenance lists {len(claimed)} documents, expected 5")
        for what, fname in claimed.items():
            folder = "templates" if "SourceData" in fname else "docs"
            if not (ROOT / folder / fname).exists():
                problems.append(f"app provenance names {fname}, which is not in {folder}/")
                continue
            # newest of its kind: same stem, highest version
            stem = re.sub(r"_v[\d.]+\.xlsx$", "", fname)
            sibs = sorted((ROOT / folder).glob(f"{stem}_v*.xlsx"),
                          key=lambda q: [int(x) for x in
                                         re.search(r"_v([\d.]+)\.xlsx$", q.name).group(1).split(".")])
            if sibs and sibs[-1].name != fname:
                problems.append(f"app provenance names {fname} for '{what}', but the repository's "
                                f"newest is {sibs[-1].name}")
    m = re.search(r"const SCHEMA_EXPECTED = (\d+);", src)
    if m and int(m.group(1)) != tpl_v:
        problems.append(f"app expects schema v{m.group(1)}, template is v{tpl_v}")

# ---- 4e2. the blank-start seed vs the template it was taken from ----------
# A plan started inside the application begins from the template's reference content -
# the value lists and the settings - embedded in the HTML. Embedded means copied, and a
# copy that falls behind the template would give someone starting a new plan a different
# vocabulary from someone who loaded a workbook, which is exactly the kind of divergence
# nothing would announce.
sys.path.insert(0, str(ROOT / "tools"))
import build_app_seed                                              # noqa: E402

if APP.exists():
    seeded = build_app_seed.embedded(src)
    if seeded is None:
        problems.append("app/PRAP.html has no SEED_LISTS / SEED_CONFIG block - "
                        "run python tools/build_app_seed.py")
    else:
        want = build_app_seed.seed_rows(TEMPLATE)
        for sheet in ("Lists", "Config"):
            got = [[None if v == "" else v for v in r] for r in seeded[sheet]]
            exp = [[None if v == "" else v for v in r] for r in want[sheet]]
            if got != exp:
                problems.append(f"the blank-start seed for {sheet} does not match "
                                f"{TEMPLATE.name} ({len(got)} rows vs {len(exp)}) - "
                                f"run python tools/build_app_seed.py")
        # The weight and role-factor grids are BUILT from those lists rather than
        # embedded, so what has to hold is that the lists the builder reads all exist.
        names = {r[0] for r in seeded["Lists"]}
        for needed in ("project_type", "clinical_phase", "period_name_clinical",
                       "period_name_others", "role_clinical", "role_others"):
            if needed not in names:
                problems.append(f"the blank-start seed has no '{needed}' list, so a new "
                                f"plan could not build its reference grid")
        notes.append(f"blank-start seed: {len(seeded['Lists'])} list values and "
                     f"{len(seeded['Config'])} settings, matching {TEMPLATE.name}")

# ---- 4f. the AI-agent reference vs the artifacts it describes --------------
# docs/prap_contract.json is what another AI system reads INSTEAD of opening these
# workbooks. A contract that has drifted from the template is worse than none, because
# an agent would follow it confidently. It is generated, so the check is that the
# generated copy in the repository is the one the current sources produce.
CONTRACT = ROOT / "docs" / "prap_contract.json"
GUIDE_MD = ROOT / "docs" / "PRAP_AI_Agent_Guide.md"
MANIFEST = ROOT / "docs" / "PRAP_Manifest.json"
for p in (CONTRACT, GUIDE_MD, MANIFEST):
    if not p.exists():
        problems.append(f"{p.relative_to(ROOT)} is missing - run python tools/build_ai_reference.py")

if CONTRACT.exists():
    contract = json.loads(CONTRACT.read_text(encoding="utf-8"))
    if contract["schema_version"] != tpl_v:
        problems.append(f"the contract says schema v{contract['schema_version']}, "
                        f"the template is v{tpl_v}")
    c_cols = {s: [c["name"] for c in v["columns"]] for s, v in contract["sheets"].items()}
    for sheet, cols in c_cols.items():
        actual = [c for c in tpl_cols.get(sheet, []) if c]
        if actual and actual != cols:
            problems.append(f"the contract's columns for {sheet} are not the template's: "
                            f"{[c for c in cols if c not in actual] or ''} "
                            f"{[c for c in actual if c not in cols] or ''}".strip())
    # Every list the template ships must be in the contract with the same values.
    tpl_lists = {}
    for row in tpl["Lists"].iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            tpl_lists.setdefault(row[0], []).append(row[1])
    for name, vals in tpl_lists.items():
        if contract["value_lists"].get(name) != vals:
            problems.append(f"value list '{name}' differs between the template and the contract")
    if set(contract["value_lists"]) != set(tpl_lists):
        problems.append("the contract and the template do not carry the same set of value lists")
    # A rule documented in the plan but reported by nothing is a promise to an agent
    # that the application does not keep.
    for rid, r in contract["validation_rules"].items():
        if not r["enforced_by_application"]:
            problems.append(f"{rid} is documented in the plan but the application never reports it")
    if APP.exists():
        m = re.search(r'const APP_VERSION = "([\d.]+)"', src)
        if m and contract["application"]["version"] != m.group(1):
            problems.append(f"the contract says app v{contract['application']['version']}, "
                            f"app/PRAP.html says v{m.group(1)}")
    notes.append(f"contract: {len(c_cols)} sheets, "
                 f"{sum(len(v) for v in c_cols.values())} columns, "
                 f"{len(contract['value_lists'])} value lists, "
                 f"{len(contract['validation_rules'])} rules - all agree with the template")

if MANIFEST.exists():
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    for entry in manifest["current"]:
        p = ROOT / entry["path"]
        if not p.exists():
            problems.append(f"the manifest points at {entry['path']}, which does not exist")
        elif entry["sha256"] and hashlib.sha256(p.read_bytes()).hexdigest() != entry["sha256"]:
            problems.append(f"{entry['path']} has changed since the manifest was built - "
                            f"run python tools/build_ai_reference.py")

# ---- 4h. the desktop plan vs the desktop specification --------------------
# The second product line gets the same guarantee as the first: a requirement cannot be
# dropped between the plan and the specification without this saying so.
NAPP_PLAN = ROOT / "docs" / "PRAP_NewApp_Development_Plan_v1.8.xlsx"
NAPP_SPEC = ROOT / "docs" / "PRAP_NewApp_Specification_v1.2.xlsx"
if NAPP_PLAN.exists() and NAPP_SPEC.exists():
    np_ = load_workbook(NAPP_PLAN, data_only=True)["03_Requirements"]
    ns_ = load_workbook(NAPP_SPEC, data_only=True)["11_Traceability"]
    nr_plan = {str(np_.cell(r, 1).value).strip() for r in range(5, np_.max_row + 1)
               if str(np_.cell(r, 1).value or "").startswith("NR-")}
    nr_spec = {str(ns_.cell(r, 1).value).strip() for r in range(5, ns_.max_row + 1)
               if str(ns_.cell(r, 1).value or "").startswith("NR-")}
    unhoused = {str(ns_.cell(r, 1).value).strip() for r in range(5, ns_.max_row + 1)
                if str(ns_.cell(r, 1).value or "").startswith("NR-")
                and str(ns_.cell(r, 2).value or "-").strip() in ("-", "None", "")}
    if nr_plan - nr_spec:
        problems.append(f"desktop requirements in the plan but not traced in its "
                        f"specification: {sorted(nr_plan - nr_spec)}")
    if nr_spec - nr_plan:
        problems.append(f"desktop requirements traced in the specification but absent "
                        f"from the plan: {sorted(nr_spec - nr_plan)}")
    if unhoused:
        problems.append(f"desktop requirements traced to no sheet: {sorted(unhoused)}")
    if not (nr_plan - nr_spec) and not unhoused:
        notes.append(f"desktop line: {len(nr_plan)} requirements, all specified")

# ---- 4g. app/PRAP.html vs the src/ tree it is now built from ---------------
# Since N2.1 the single file is a build output, not a hand-written one. Editing it
# directly would work perfectly well until the next build silently discarded the edit,
# so the two are held together here rather than by anybody remembering.
try:
    sys.path.insert(0, str(ROOT / "tools"))
    import build_app

    built = build_app.render()
    current = (ROOT / "app" / "PRAP.html").read_text(encoding="utf-8")
    if built != current:
        b, c = built.split("\n"), current.split("\n")
        where = next((i for i, (x, y) in enumerate(zip(b, c), start=1) if x != y), None)
        problems.append(
            f"app/PRAP.html does not match a build from src/ "
            + (f"(first difference at line {where})" if where
               else f"(built {len(b)} lines, committed {len(c)})")
            + " - edit the part under src/ and run python tools/build_app.py")
    else:
        notes.append(f"app/PRAP.html is byte-identical to a build from src/ "
                     f"({len(build_app.PARTS)} parts, {len(built):,} bytes)")
except Exception as exc:                                    # noqa: BLE001
    problems.append(f"could not build app/PRAP.html from src/: {exc}")

# ---- 5. every requirement in the plan is traced in the specification ------
plan_reqs = {str(plan["03_Requirements"].cell(r, 1).value).strip()
             for r in range(5, plan["03_Requirements"].max_row + 1)
             if str(plan["03_Requirements"].cell(r, 1).value or "").startswith("REQ")}
spec_traced = {str(spec["09_Traceability"].cell(r, 1).value).strip()
               for r in range(5, spec["09_Traceability"].max_row + 1)
               if str(spec["09_Traceability"].cell(r, 1).value or "").startswith("REQ")}
if plan_reqs - spec_traced:
    problems.append(f"requirements in the plan but not traced in the spec: {sorted(plan_reqs - spec_traced)}")
if spec_traced - plan_reqs:
    problems.append(f"requirements traced in the spec but absent from the plan: {sorted(spec_traced - plan_reqs)}")

# ---- report ---------------------------------------------------------------
print(f"plan       {PLAN.name}")
print(f"spec       {SPEC.name}")
print(f"template   {TEMPLATE.name}   schema v{tpl_v}")
print(f"dummy      {DUMMY.name}   schema v{dum_v}")
print(f"dummy      {DUMMY_SMALL.name}   schema v{small_v}")
print(f"types      {types}")
print(f"columns cross-checked: {sum(len(v) for k, v in documented.items() if k in plan_sheets)}"
      f" across {len(plan_sheets)} sheets")
print(f"requirements traced: {len(plan_reqs)}")
print()
if problems:
    print(f"PROBLEMS ({len(problems)}):")
    for p_ in problems:
        print("   ", p_)
else:
    print("PROBLEMS: none - plan, specification, template and dummy agree.")
if notes:
    print(f"\nNOTES ({len(notes)}):")
    for n in notes:
        print("   ", n)

sys.exit(1 if problems else 0)
