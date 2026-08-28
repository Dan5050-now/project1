"""The second export: the calculated figures, not the plan.

Two exports that must never be confused with each other. The SOURCE export is the plan
and its whole point is to come back; this one is the ANSWER, for a spreadsheet or a
report, and it deliberately does not round-trip.

What is checked, and why each one matters:

  * BOTH are on the export menu, described so the choice can be made without trying it.
  * The results file carries seven sheets, and the first one says it cannot be imported -
    because somebody will try, and finding out by losing a plan is the wrong way to learn.
  * IT ADDS UP. Every project-month and every person-month is EXACTLY the sum of its
    Detail rows, and the Summary total is exactly the sum of both. A workbook whose
    purpose is to be checked must survive being checked: add the column, get the total.
  * EVERY DETAIL ROW IS ITS OWN MULTIPLICATION - period weight x (factor / sharers) x
    person weight x coverage - so a reader who disagrees with a figure can see which of
    the four numbers they disagree with.
  * IT FOLLOWS THE SCREEN. Filter to one person and the file holds that person, with the
    filter named on the ReadMe, and the project totals fall to match. A file that
    silently held more than the screen would be the worse failure of the two.
  * The figures equal the independent Python reference, to rounding.
  * The source export still round-trips, and is untouched by any of this.

    python tools/test_results.py
"""

import pathlib
import sys
import tempfile
from collections import defaultdict

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

ROOT = pathlib.Path(__file__).resolve().parents[1]
APP = (ROOT / "app" / "PRAP.html").as_uri()
CHROME = "/opt/pw-browsers/chromium-1194/chrome-linux/chrome"
FIX = ROOT / "templates" / "PRAP_SourceData_Dummy_10x10_v1.5.xlsx"
TMP = pathlib.Path(tempfile.mkdtemp(prefix="prap_results_"))

sys.path.insert(0, str(ROOT / "tools"))
import prap_io                                                       # noqa: E402

fails = []


def check(ok, label, detail=""):
    print(f"  {'ok  ' if ok else 'FAIL'} {label}{'   ' + detail if detail else ''}")
    if not ok:
        fails.append(label)


def sheet(wb, name):
    it = wb[name].iter_rows(values_only=True)
    hdr = next(it)
    return [dict(zip(hdr, r)) for r in it]


def export(pg, item, name):
    pg.click("#exportBtn")
    pg.wait_for_timeout(300)
    with pg.expect_download() as dl:
        pg.click(item)
    out = TMP / name
    dl.value.save_as(out)
    return out


with sync_playwright() as pw:
    browser = pw.chromium.launch(executable_path=CHROME, downloads_path=str(TMP))
    ctx = browser.new_context(accept_downloads=True)
    pg = ctx.new_page()
    pg.set_viewport_size({"width": 1500, "height": 950})
    pg.set_default_timeout(25000)
    errors = []
    pg.on("pageerror", lambda e: errors.append(str(e)))
    pg.on("dialog", lambda d: d.accept())

    print("app/PRAP.html — exporting the figures rather than the plan")
    pg.goto(APP)
    pg.wait_for_timeout(400)
    check(pg.eval_on_selector("#expMenu", "e => e.classList.contains('off')"),
          "with nothing loaded there is nothing to export, and the menu says so")

    pg.set_input_files("#picker", str(FIX))
    pg.wait_for_timeout(3500)
    pg.click("#exportBtn")
    pg.wait_for_timeout(400)
    items = pg.eval_on_selector_all(".expitem", "es => es.map(e => e.innerText)")
    check(len(items) == 3
          and any("Calculated monthly FTE" in t for t in items)
          and any("Cannot be imported back" in t for t in items),
          "THE MENU OFFERS BOTH, and says which one comes back and which does not",
          " | ".join(t.split("\n")[0] for t in items))
    pg.keyboard.press("Escape")
    pg.wait_for_timeout(200)

    res = export(pg, "#exportCalcBtn", "calc.xlsx")
    check(res.name.endswith(".xlsx") and res.stat().st_size > 20000,
          "the calculated export downloads", f"{res.stat().st_size:,} bytes")

    wb = load_workbook(res)
    want = ["00_ReadMe", "Summary", "ProjectMonth", "PersonMonth", "Detail", "Flags",
            "Assumptions"]
    check(wb.sheetnames == want, "seven sheets, ReadMe first",
          ", ".join(wb.sheetnames))

    readme = " ".join(str(c.value or "") for r in wb["00_ReadMe"].iter_rows() for c in r)
    check("CANNOT be imported" in readme and "FTE  =  period weight" in readme
          and "V-12" in readme and "V-23" in readme,
          "the ReadMe says it cannot be imported, gives the formula, and names where a "
          "figure can be short of an assumption")

    detail = sheet(wb, "Detail")
    pmn = sheet(wb, "ProjectMonth")
    smn = sheet(wb, "PersonMonth")
    summ = sheet(wb, "Summary")

    # ---- it adds up -------------------------------------------------------------
    byproj, bypers = defaultdict(float), defaultdict(float)
    for r in detail:
        byproj[(r["month_iso"], r["project_id"])] += r["fte"]
        bypers[(r["month_iso"], r["person_id"])] += r["fte"]
    bad = [k for k, v in byproj.items()
           if abs(v - next((x["fte"] for x in pmn
                            if x["month_iso"] == k[0] and x["project_id"] == k[1]), 0)) > 1e-9]
    check(not bad and len(byproj) == len(pmn),
          "EVERY PROJECT-MONTH IS EXACTLY THE SUM OF ITS DETAIL ROWS",
          f"{len(pmn)} project-months" + (f"; {len(bad)} differ" if bad else ""))
    bad = [k for k, v in bypers.items()
           if abs(v - next((x["fte"] for x in smn
                            if x["month_iso"] == k[0] and x["person_id"] == k[1]), 0)) > 1e-9]
    check(not bad and len(bypers) == len(smn),
          "and every person-month too",
          f"{len(smn)} person-months" + (f"; {len(bad)} differ" if bad else ""))

    td = sum(r["fte"] for r in detail)
    tp = sum(r["fte"] for r in pmn)
    ts = sum(r["fte"] for r in smn)
    tot = next(r["value"] for r in summ
               if r["measure"] == "Total demand" and r["unit"] == "FTE-months")
    check(max(abs(x - td) for x in (tp, ts, tot)) < 1e-9,
          "and all four totals in the file are the same number",
          f"detail {td:.4f}, project {tp:.4f}, person {ts:.4f}, summary {tot:.4f}")

    # ---- each row is its own multiplication ---------------------------------------
    bad = [r for r in detail
           if abs(r["period_weight"] * (r["role_factor_effective"] / r["sharers"])
                  * r["person_weight"] * r["month_coverage"] - r["fte"]) > 5e-4]
    check(not bad,
          "EVERY DETAIL ROW IS ITS OWN FOUR NUMBERS — a reader can check any figure "
          "without the application",
          f"{len(detail)} rows" + (f"; {len(bad)} do not reconcile" if bad else ""))

    shared = [r for r in detail if r["sharers"] > 1]
    absorbed = [r for r in detail if r["absorbed_from"]]
    check(shared and all(r["role_factor_effective"] / r["sharers"] < r["role_factor_effective"]
                         for r in shared),
          "the working shows a shared role being divided",
          f"{len(shared)} row(s) with more than one holder")
    check(all(r["role_factor_effective"] > (r["role_factor"] or 0) for r in absorbed)
          if absorbed else True,
          "and an absorbed factor as larger than the role's own",
          f"{len(absorbed)} row(s) carrying an absorbed factor")

    # ---- against the reference ------------------------------------------------------
    M = prap_io.Model(prap_io.read_xlsx(FIX))
    C = prap_io.calculate(M)
    ref = defaultdict(float)
    for (sid, k), v in C["pers_month"].items():
        ref[(f"{k // 12}-{(k % 12) + 1:02d}", sid)] += v
    worst = max(abs(v - ref.get(k, 0)) for k, v in bypers.items())
    check(worst < 1e-3,
          "and the figures are the Python reference implementation's, to rounding",
          f"{len(bypers)} person-months, worst difference {worst:.2e}")

    # ---- it follows the screen ------------------------------------------------------
    who = smn[0]["person_id"]
    pg.evaluate("""(sid) => { S.f.pers = new Set([sid]); fillFilters(); renderAll();
                              showTab(S.tab); }""", who)
    pg.wait_for_timeout(900)
    res2 = export(pg, "#exportCalcBtn", "calc_filtered.xlsx")
    wb2 = load_workbook(res2)
    d2, s2 = sheet(wb2, "Detail"), sheet(wb2, "PersonMonth")
    rm2 = " ".join(str(c.value or "") for r in wb2["00_ReadMe"].iter_rows() for c in r)
    check(d2 and {r["person_id"] for r in d2} == {who}
          and {r["person_id"] for r in s2} == {who},
          "FILTER TO ONE PERSON AND THE FILE HOLDS THAT PERSON",
          f"{len(d2)} detail row(s), {len({r['person_id'] for r in d2})} person")
    check(f"Person: {who}" in rm2,
          "with the filter named on the ReadMe, so the file explains its own scope",
          next((str(c.value) for r in wb2["00_ReadMe"].iter_rows() for c in r
                if c.value and "Person:" in str(c.value)), "(not named)"))
    byproj2 = defaultdict(float)
    for r in d2:
        byproj2[(r["month_iso"], r["project_id"])] += r["fte"]
    p2 = sheet(wb2, "ProjectMonth")
    bad = [k for k, v in byproj2.items()
           if abs(v - next((x["fte"] for x in p2
                            if x["month_iso"] == k[0] and x["project_id"] == k[1]), 0)) > 1e-9]
    check(not bad,
          "and the project totals fall to match — they are not the unfiltered ones",
          f"{len(p2)} project-month(s), all equal to the rows beneath them")
    pg.evaluate("() => { S.f.pers = new Set(); fillFilters(); renderAll(); showTab(S.tab); }")
    pg.wait_for_timeout(700)

    # ---- the source export is untouched ---------------------------------------------
    src = export(pg, "#exportBtn2", "source.xlsx")
    back = prap_io.read_xlsx(src)
    check(len(back["Project"]) == 10 and len(back["Assignment"]) == 48
          and set(back) == set(prap_io.SHEET_ORDER),
          "THE SOURCE EXPORT STILL ROUND-TRIPS, with every sheet the reader expects",
          f"{sum(len(v) for v in back.values() if isinstance(v, list))} rows across "
          f"{len(back)} sheets")

    check(not errors, "no uncaught errors in the page", "; ".join(errors[:2]))
    browser.close()

print(f"\nFAILURES: {'none' if not fails else len(fails)}")
for f in fails:
    print(f"  FAILED  {f}")
sys.exit(1 if fails else 0)
