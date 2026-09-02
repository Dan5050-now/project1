"""Manual monthly estimation (REQ-CAL-18), driven through the screen.

The engine half is checked by test_app.py, which compares four independent
implementations on every person-month of a fixture that carries a manual project and a
manual assignment. This is the other half: the part a user touches, where the failure
modes are different in kind.

What is checked, and why each one earns its place:

  * BOTH LEVELS ARE ON SCREEN, and say which they are - a project panel on the project
    tab, an assignment panel beside the selected assignment, each badged MANUAL or
    AUTOMATIC. A figure that has been stated rather than calculated and does not say so
    is the whole hazard this feature introduces.
  * EVERY CHANGE OF CALCULATION WAY ASKS FIRST. In both directions, and from the cell as
    well as the button. This is a change to a whole run of figures, and the only two
    outcomes of doing it by accident are silent zeroes or deleted work.
  * SWITCHING TO MANUAL DOES NOT MOVE ANYTHING. Every month is copied across as it stood,
    so the totals immediately after a switch equal the totals immediately before it. A
    switch that jumped would make the feature unusable on a plan anybody had reviewed.
  * A STATED FIGURE IS THEN USED. Type a month, and the project total, the person total
    and the chart all move to it - not to the calculation.
  * SWITCHING BACK DISCARDS THE LOT, and returns the figures to exactly what they were
    before the manual run started.
  * IT IS AN EDIT LIKE ANY OTHER: it appears in the change log, and 'Leave without
    change' puts everything back.
  * IT SURVIVES THE ROUND TRIP. Export the plan, read it back, and the stated figures and
    both estimation_type flags are still there and still in force.
  * THE CALCULATED-FTE EXPORT SAYS WHICH ROWS WERE STATED, and at which level.

    python tools/test_manual.py
"""

import pathlib
import sys
import tempfile

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

ROOT = pathlib.Path(__file__).resolve().parents[1]
APP = (ROOT / "app" / "PRAP.html").as_uri()
CHROME = "/opt/pw-browsers/chromium-1194/chrome-linux/chrome"
FIX = ROOT / "templates" / "PRAP_SourceData_Dummy_10x10_v1.7.xlsx"
TMP = pathlib.Path(tempfile.mkdtemp(prefix="prap_manual_"))

sys.path.insert(0, str(ROOT / "tools"))
import prap_io                                                       # noqa: E402

fails = []


def check(ok, label, detail=""):
    print(f"  {'ok  ' if ok else 'FAIL'} {label}{'   ' + detail if detail else ''}")
    if not ok:
        fails.append(label)


def sheet(wb, name):
    it = wb[name].iter_rows(values_only=True)
    hdr = next(it)
    return [dict(zip(hdr, r)) for r in it]


with sync_playwright() as pw:
    browser = pw.chromium.launch(executable_path=CHROME, downloads_path=str(TMP))
    ctx = browser.new_context(accept_downloads=True)
    pg = ctx.new_page()
    pg.set_viewport_size({"width": 1600, "height": 1000})
    pg.set_default_timeout(25000)
    errors = []
    pg.on("pageerror", lambda e: errors.append(str(e)))
    pg.on("dialog", lambda d: d.accept())

    print("app/PRAP.html — stating a figure instead of calculating it")
    pg.goto(APP)
    pg.set_input_files("#picker", str(FIX))
    pg.wait_for_timeout(3500)

    # ---- the fixture's own manual rows are in force -------------------------------
    man = pg.evaluate("""() => ({
        proj: S.model.raw.Project.filter(p => p.estimation_type === 'manual')
                .map(p => p.project_id),
        asg:  S.model.raw.Assignment.filter(a => a.estimation_type === 'manual')
                .map(a => a.assignment_id),
        est:  S.model.raw.MonthlyEstimate.length,
        lines_p: S.calc.lines.filter(L => L.manual_project).length,
        lines_a: S.calc.lines.filter(L => L.manual_assignment).length})""")
    check(len(man["proj"]) == 1 and len(man["asg"]) == 1 and man["est"] > 0
          and man["lines_p"] > 0 and man["lines_a"] > 0,
          "THE FILE'S OWN MANUAL ROWS ARE IN FORCE — both levels, straight from the workbook",
          f"project {man['proj']}, assignment {man['asg']}, {man['est']} stated month(s); "
          f"{man['lines_p']} line(s) carry a project figure, {man['lines_a']} an assignment one")
    MP, MA = man["proj"][0], man["asg"][0]

    # ---- the panel is there, and says which it is ---------------------------------
    pg.click("text=Source data (project)")
    pg.wait_for_timeout(1200)
    pg.evaluate("(pid) => { S.selProj = pid; renderAll(); showTab('t-proj'); }", MP)
    pg.wait_for_timeout(1200)
    badge = pg.eval_on_selector_all("#t-proj .est", "es => es.map(e => e.textContent.trim())")
    check(badge == ["MANUAL"],
          "the manual project's panel is badged MANUAL, not left to be inferred",
          " / ".join(badge) or "(no badge)")
    txt = pg.inner_text("#t-proj")
    check("Monthly estimation" in txt and "automatic_fte" in txt,
          "with the calculated figure beside every stated one, so the departure is readable")

    auto_pid = pg.evaluate("""() => Object.keys(S.model.projects)
        .find(p => (S.model.raw.Project.find(r => r.project_id === p) || {}).estimation_type
                   !== 'manual')""")
    pg.evaluate("(pid) => { S.selProj = pid; renderAll(); showTab('t-proj'); }", auto_pid)
    pg.wait_for_timeout(1000)
    badge = pg.eval_on_selector_all("#t-proj .est", "es => es.map(e => e.textContent.trim())")
    check(badge == ["AUTOMATIC"], "and an ordinary project is badged AUTOMATIC",
          f"{auto_pid}: " + (" / ".join(badge) or "(no badge)"))

    # ---- nothing changes without being asked --------------------------------------
    before = pg.evaluate("""() => { const o = {}; for (const [k, v] of S.calc.projMonth) o[k] = v;
                                    return o; }""")
    total_before = sum(before.values())
    pg.eval_on_selector("#t-proj .est ~ .btn", "b => b.click()")
    pg.wait_for_timeout(600)
    asked = pg.eval_on_selector("#estchg", "d => d.open")
    title = pg.inner_text("#estTitle")
    body = pg.inner_text("#estBody")
    check(asked and "manual" in title.lower(),
          "SWITCHING TO MANUAL ASKS FIRST, in its own dialog", title)
    check("responsibility" in body.lower() or "your job" in body.lower(),
          "and says what the user is taking on, not just what the button does")
    pg.click("#estNo")
    pg.wait_for_timeout(700)
    still = pg.evaluate("""() => { let t = 0; for (const v of S.calc.projMonth.values()) t += v;
                                   return t; }""")
    check(pg.evaluate("() => S.pending.length") == 0 and abs(still - total_before) < 1e-9,
          "and 'Go back' changes nothing at all — no pending edit, no moved figure",
          f"{still:.4f} FTE-months, unchanged")

    # ---- switching copies the months across, and nothing jumps --------------------
    n_months = pg.evaluate("""(pid) => new Set(S.calc.lines.filter(L => L.project_id === pid)
                                .map(L => L.month)).size""", auto_pid)
    pg.eval_on_selector("#t-proj .est ~ .btn", "b => b.click()")
    pg.wait_for_timeout(500)
    pg.click("#estYes")
    pg.wait_for_timeout(1500)
    after = pg.evaluate("""() => { const o = {}; for (const [k, v] of S.calc.projMonth) o[k] = v;
                                   return o; }""")
    worst = max((abs(before.get(k, 0) - after.get(k, 0)) for k in set(before) | set(after)),
                default=0)
    # Within the rounding of the figure that was seeded, and no further. The seed is
    # written to TWO decimal places, because a stated figure lands in a cell a person then
    # reads and edits and 0.8814374999 is not a number anybody states - and because two
    # places is the finest edit that means anything in the unit people think in: at 160
    # hours to the FTE, 0.01 is 1.6 hours. So a month can move by up to half a unit in the
    # last place, 0.005 FTE, which is 48 minutes. The claim being tested is that nothing
    # MEANINGFUL moves; asserting bit-equality here would be asserting the wrong thing,
    # and would be satisfied only by making the panel worse.
    check(worst <= 5e-3,
          "SWITCHING TO MANUAL MOVES NOTHING — every month was copied across as it stood, "
          "to the precision it is stated in",
          f"{len(after)} project-months, worst difference {worst:.2e} FTE "
          f"({worst * 160:.2f} hours), against the 0.005 FTE the rounding allows")
    seeded = pg.evaluate("""(pid) => S.model.raw.MonthlyEstimate
        .filter(r => r.scope === 'project' && r.ref_id === pid).length""", auto_pid)
    check(seeded == n_months,
          "and EVERY month it covers got a figure — manual is all or nothing",
          f"{seeded} stated for {n_months} calculated month(s)")
    flag = pg.evaluate("""(pid) => (S.model.raw.Project.find(r => r.project_id === pid)
                                    || {}).estimation_type""", auto_pid)
    check(flag == "manual", "the project itself is now manual", str(flag))

    logged = pg.evaluate("() => S.pending.length")
    check(logged == seeded + 1,
          "IT IS AN EDIT LIKE ANY OTHER — the flag and every seeded month are in the log",
          f"{logged} pending change(s)")

    # ---- a stated figure is then USED ---------------------------------------------
    mk = pg.evaluate("""(pid) => { const ms = S.model.raw.MonthlyEstimate
        .filter(r => r.scope === 'project' && r.ref_id === pid)
        .sort((a, b) => String(a.month).localeCompare(String(b.month)));
        return ms[Math.floor(ms.length / 2)].month; }""", auto_pid)
    row = pg.evaluate("""([pid, mm]) => S.model.raw.MonthlyEstimate
        .find(r => r.scope === 'project' && r.ref_id === pid && r.month === mm).__row""",
        [auto_pid, mk])
    cell = f'#t-proj td[data-sheet="MonthlyEstimate"][data-row="{row}"][data-col="fte"]'
    pg.eval_on_selector(cell, "e => e.scrollIntoView({block:'center'})")
    pg.click(cell)
    pg.keyboard.press("Control+a")
    pg.type(cell, "9.5")
    pg.keyboard.press("Tab")
    pg.wait_for_timeout(1200)
    key = f"{auto_pid}|{int(mk.split('-')[0]) * 12 + int(mk.split('-')[1]) - 1}"
    got = pg.evaluate("(k) => S.calc.projMonth.get(k)", key)
    check(got is not None and abs(got - 9.5) < 1e-6,
          "A STATED FIGURE IS THE ONE USED — the project's month is what was typed",
          f"{auto_pid} {mk}: {got if got is None else round(got, 4)} (was "
          f"{round(before.get(key, 0), 4)})")
    people = pg.evaluate("""(k) => { const m = S.calc.projPers.get(k); let t = 0;
        if (m) for (const v of m.values()) t += v; return t; }""", key)
    check(people is not None and abs(people - 9.5) < 1e-6,
          "and the people on it are scaled so they still add up to it",
          f"sum over the people that month: {round(people, 4) if people else people}")

    # ---- switching back discards it, exactly ---------------------------------------
    pg.eval_on_selector("#t-proj .est ~ .btn", "b => b.click()")
    pg.wait_for_timeout(500)
    t2 = pg.inner_text("#estTitle")
    b2 = pg.inner_text("#estBody")
    check(pg.eval_on_selector("#estchg", "d => d.open") and "automatic" in t2.lower(),
          "SWITCHING BACK ASKS TOO, in the other direction", t2)
    check("delet" in b2.lower(), "and says the stated figures are deleted, not set aside")
    pg.click("#estYes")
    pg.wait_for_timeout(1500)
    back = pg.evaluate("""() => { const o = {}; for (const [k, v] of S.calc.projMonth) o[k] = v;
                                  return o; }""")
    worst = max((abs(before.get(k, 0) - back.get(k, 0)) for k in set(before) | set(back)),
                default=0)
    left = pg.evaluate("""(pid) => S.model.raw.MonthlyEstimate
        .filter(r => r.scope === 'project' && r.ref_id === pid).length""", auto_pid)
    check(worst < 1e-9 and left == 0,
          "SWITCHING BACK RESTORES THE CALCULATION EXACTLY, and leaves no stated month behind",
          f"{left} stated month(s) left, worst difference {worst:.2e}")

    # ---- and the whole lot can be abandoned ----------------------------------------
    pg.click("#discardBtn")
    pg.wait_for_timeout(1200)
    check(pg.evaluate("() => S.pending.length") == 0
          and pg.evaluate("""(pid) => (S.model.raw.Project.find(r => r.project_id === pid)
                                       || {}).estimation_type""", auto_pid) != "manual",
          "'Leave without change' undoes a switch like any other edit")

    # ---- the assignment level, on the other tab ------------------------------------
    sid = pg.evaluate("""(aid) => (S.model.raw.Assignment.find(a => a.assignment_id === aid)
                                   || {}).person_id""", MA)
    pg.click("text=Source data (person)")
    pg.wait_for_timeout(1000)
    pg.evaluate("([s, a]) => { S.selPers = s; S.selAsg = a; renderAll(); showTab('t-pers'); }",
                [sid, MA])
    pg.wait_for_timeout(1400)
    badge = pg.eval_on_selector_all("#t-pers .est", "es => es.map(e => e.textContent.trim())")
    ptxt = pg.inner_text("#t-pers")
    check(badge == ["MANUAL"] and "Monthly estimation" in ptxt,
          "THE ASSIGNMENT LEVEL HAS ITS OWN PANEL, badged for the selected assignment",
          f"{MA}: " + (" / ".join(badge) or "(no badge)"))
    check(MA in ptxt,
          "WHICH ASSIGNMENT IS STATED IS NAMED, not left as a property of the person")

    # ---- the project panel names the assignments that are manual --------------------
    apid = pg.evaluate("""(aid) => (S.model.raw.Assignment.find(a => a.assignment_id === aid)
                                    || {}).project_id""", MA)
    pg.click("text=Source data (project)")
    pg.wait_for_timeout(800)
    pg.evaluate("(pid) => { S.selProj = pid; renderAll(); showTab('t-proj'); }", apid)
    pg.wait_for_timeout(1200)
    txt = pg.inner_text("#t-proj")
    check(MA in txt,
          "AND THE PROJECT PANEL NAMES THE ASSIGNMENT — so a project total that is manual "
          "says at WHICH level it is manual",
          f"{apid} names {MA}")

    # ---- typing into the cell is refused, and offers the switch ---------------------
    prow = pg.evaluate("""(pid) => (S.model.raw.Project.find(r => r.project_id === pid)
                                    || {}).__row""", apid)
    pcell = f'#t-proj td[data-sheet="Project"][data-row="{prow}"][data-col="estimation_type"]'
    pg.eval_on_selector(pcell, "e => e.scrollIntoView({block:'center'})")
    pg.click(pcell)
    pg.keyboard.press("Control+a")
    pg.type(pcell, "manual")
    pg.keyboard.press("Tab")
    pg.wait_for_timeout(800)
    check(pg.eval_on_selector("#estchg", "d => d.open"),
          "TYPING 'manual' INTO THE CELL DOES NOT SET IT SILENTLY — it opens the same "
          "confirmation, because the flag without the figures would zero every month")
    pg.click("#estNo")
    pg.wait_for_timeout(600)
    check(pg.evaluate("""(pid) => (S.model.raw.Project.find(r => r.project_id === pid)
                                   || {}).estimation_type""", apid) != "manual"
          and pg.evaluate("() => S.pending.length") == 0,
          "and declining leaves the cell exactly as it was")

    # ---- a stated figure keeps its own row alive --------------------------------------
    # V-17 already refuses to delete a row anything still points at, and MonthlyEstimate
    # points at both a project and an assignment. Worth pinning rather than assuming:
    # deleting a manual assignment and leaving its stated months behind would put figures
    # in the file belonging to something that no longer exists.
    pg.click("text=Source data (person)")
    pg.wait_for_timeout(800)
    pg.evaluate("([s, a]) => { S.selPers = s; S.selAsg = a; renderAll(); showTab('t-pers'); }",
                [sid, MA])
    pg.wait_for_timeout(1200)
    arow = pg.evaluate("""(aid) => (S.model.raw.Assignment.find(a => a.assignment_id === aid)
                                    || {}).__row""", MA)
    n_before = pg.evaluate("() => S.model.raw.Assignment.length")
    pg.eval_on_selector(f'#t-pers button[data-del="Assignment"][data-row="{arow}"]',
                        "b => { b.scrollIntoView({block:'center'}); b.click(); }")
    pg.wait_for_timeout(900)
    ban = pg.inner_text("#banner") if pg.locator("#banner").count() else ""
    check(pg.evaluate("() => S.model.raw.Assignment.length") == n_before
          and "MonthlyEstimate" in ban,
          "DELETING A MANUAL ASSIGNMENT IS REFUSED, naming the stated months that would "
          "otherwise be left pointing at nothing",
          " ".join(ban.split())[:110])

    # ---- it survives the round trip --------------------------------------------------
    with pg.expect_download() as dl:
        pg.click("#exportBtn")
        pg.click("#exportBtn2")
    src = TMP / "source.xlsx"
    dl.value.save_as(src)
    a, b = prap_io.read_xlsx(FIX), prap_io.read_xlsx(src)
    same = (len(a["MonthlyEstimate"]) == len(b["MonthlyEstimate"])
            and [r["project_id"] for r in b["Project"] if r.get("estimation_type") == "manual"]
                == [MP]
            and [r["assignment_id"] for r in b["Assignment"]
                 if r.get("estimation_type") == "manual"] == [MA])
    CA, CB = prap_io.calculate(prap_io.Model(a)), prap_io.calculate(prap_io.Model(b))
    keys = set(CA["pers_month"]) | set(CB["pers_month"])
    worst = max(abs(CA["pers_month"].get(k, 0) - CB["pers_month"].get(k, 0)) for k in keys)
    check(same and worst < 1e-9,
          "IT SURVIVES THE ROUND TRIP — the stated months and both flags come back, and "
          "still produce the same figures",
          f"{len(b['MonthlyEstimate'])} stated month(s), {len(keys)} person-months, "
          f"worst difference {worst:.2e}")

    # ---- and the calculated export says which rows were stated ------------------------
    with pg.expect_download() as dl:
        pg.click("#exportBtn")
        pg.click("#exportCalcBtn")
    res = TMP / "calc.xlsx"
    dl.value.save_as(res)
    wb = load_workbook(res)
    det = sheet(wb, "Detail")
    cols = set(det[0]) if det else set()
    stated = [r for r in det if str(r.get("estimation") or "").startswith("manual")]
    lvl = {str(r.get("estimation")) for r in stated}
    check({"estimation", "automatic_fte"} <= cols and stated,
          "THE CALCULATED EXPORT MARKS THE STATED ROWS, and keeps the calculated figure "
          "beside them",
          f"{len(stated)} of {len(det)} rows stated; " + " | ".join(sorted(lvl)[:3]))
    check(any("assignment" in x for x in lvl) and any("project" in x for x in lvl),
          "naming WHICH LEVEL each one came from",
          ", ".join(sorted(lvl)))

    check(not errors, "no uncaught errors in the page", "; ".join(errors[:2]))
    browser.close()

print(f"\nFAILURES: {'none' if not fails else len(fails)}")
for f in fails:
    print(f"  FAILED  {f}")
sys.exit(1 if fails else 0)
